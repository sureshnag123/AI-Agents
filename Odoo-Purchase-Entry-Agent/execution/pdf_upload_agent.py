#!/usr/bin/env python3
"""
PDF Purchase Upload Agent

Reads an Excel file, downloads each PDF from Google Drive,
extracts invoice data via Claude Vision, creates vendor bills in Odoo,
and writes results (Odoo Bill ID + status) back to the same Excel file.

Excel format (row 1 = header):
    Column A — Expense Ledger   (required) : exact Odoo account name
    Column B — Notes            (optional) : any notes for narration
    Column C — PDF Drive Link   (required) : Google Drive public share URL
    Column D — Odoo Bill ID     (written back by agent)
    Column E — Status           (written back by agent)

Usage:
    python pdf_upload_agent.py bills.xlsx
    python pdf_upload_agent.py bills.xlsx --dry-run
    python pdf_upload_agent.py bills.xlsx --start-row 3
    python pdf_upload_agent.py bills.xlsx --row 5        (single row)
"""

import sys
import json
import argparse
import traceback
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

SCRIPT_DIR  = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR     = PROJECT_ROOT / ".tmp"

load_dotenv(PROJECT_ROOT / ".env")

sys.path.insert(0, str(SCRIPT_DIR))
from pdf_ocr_extractor import extract_invoice_data
from odoo_bill_creator  import create_vendor_bill


# ── Column indices (1-based, for openpyxl) ───────────────────────────────────
COL_LEDGER  = 1   # A
COL_NOTES   = 2   # B
COL_DRIVE   = 3   # C
COL_BILL_ID = 4   # D  ← written back
COL_STATUS  = 5   # E  ← written back


def _cell(ws, row_num: int, col: int):
    """Return cell value as stripped string (empty string if None)."""
    val = ws.cell(row=row_num, column=col).value
    return str(val).strip() if val is not None else ""


def _write_back(ws, row_num: int, bill_name: str, status: str, excel_path: Path):
    """Write result columns D and E, then save immediately."""
    ws.cell(row=row_num, column=COL_BILL_ID).value = bill_name
    ws.cell(row=row_num, column=COL_STATUS).value  = status
    ws.parent.save(excel_path)


def process_row(ws, row_num: int, excel_path: Path, dry_run: bool = False) -> str:
    """
    Process a single Excel row.
    Returns one of: "skipped", "dry_run", "created", "duplicate", "error"
    """
    ledger    = _cell(ws, row_num, COL_LEDGER)
    notes     = _cell(ws, row_num, COL_NOTES)
    drive_url = _cell(ws, row_num, COL_DRIVE)
    existing  = _cell(ws, row_num, COL_BILL_ID)

    # Skip blank rows
    if not ledger and not drive_url:
        return "skipped"

    # Skip already-processed rows (D column already has a bill name)
    if existing and existing.upper().startswith("BILL"):
        print(f"Row {row_num}: Already uploaded ({existing}) — skipping")
        return "skipped"

    print(f"\n{'-'*60}")
    print(f"Row {row_num} | Ledger: {ledger}")

    # Validate required fields
    if not ledger:
        msg = "ERROR: Expense Ledger (column A) is empty"
        print(f"  {msg}")
        _write_back(ws, row_num, "", msg, excel_path)
        return "error"

    if not drive_url:
        msg = "ERROR: PDF Drive Link (column C) is empty"
        print(f"  {msg}")
        _write_back(ws, row_num, "", msg, excel_path)
        return "error"

    print(f"  Drive: {drive_url[:70]}{'...' if len(drive_url) > 70 else ''}")

    try:
        # Step 1: Download PDF + OCR via Claude Vision (always runs, even in dry-run)
        extracted = extract_invoice_data(drive_url, notes=notes)

        print(f"  Vendor:  {extracted.get('vendor_name')}")
        print(f"  Inv No:  {extracted.get('invoice_number')}")
        print(f"  Date:    {extracted.get('invoice_date')}")
        print(f"  Total:   {extracted.get('total')}  Tax: {extracted.get('tax_rate')}%")
        if extracted.get("line_items"):
            print(f"  Lines:   {len(extracted['line_items'])} item(s)")

        # Save debug JSON to .tmp (excludes raw PDF bytes)
        TMP_DIR.mkdir(parents=True, exist_ok=True)
        debug_file = TMP_DIR / f"row_{row_num}_extracted.json"
        debug_copy = {k: v for k, v in extracted.items() if not k.startswith("_")}
        with open(debug_file, "w", encoding="utf-8") as f:
            json.dump(debug_copy, f, indent=2, default=str)
        print(f"  Debug:   {debug_file}")

        if dry_run:
            print(f"  [DRY RUN] OCR complete — Odoo bill creation skipped")
            _write_back(ws, row_num, "", "DRY RUN — OCR OK", excel_path)
            return "dry_run"

        # Step 2: Create vendor bill in Odoo
        result = create_vendor_bill(extracted, expense_ledger=ledger)

        # Step 3: Write result back to Excel
        bill_name = result["bill_name"]
        if result["status"] == "duplicate":
            status = f"DUPLICATE — already exists as {bill_name}"
            _write_back(ws, row_num, bill_name, status, excel_path)
            return "duplicate"
        else:
            _write_back(ws, row_num, bill_name, "SUCCESS", excel_path)
            return "created"

    except Exception as e:
        err = str(e)
        short_err = err[:150]
        print(f"  ERROR: {short_err}")
        traceback.print_exc()
        _write_back(ws, row_num, "", f"ERROR: {short_err}", excel_path)
        return "error"


def process_excel(
    excel_path: Path,
    dry_run: bool = False,
    start_row: int = 2,
    single_row: int = None,
):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Ensure header labels in row 1
    if not ws.cell(1, COL_BILL_ID).value:
        ws.cell(1, COL_BILL_ID).value = "Odoo Bill ID"
    if not ws.cell(1, COL_STATUS).value:
        ws.cell(1, COL_STATUS).value = "Status"
    wb.save(excel_path)

    counts = {"created": 0, "duplicate": 0, "error": 0, "skipped": 0, "dry_run": 0}

    if single_row is not None:
        rows_to_process = [single_row]
    else:
        rows_to_process = range(start_row, ws.max_row + 1)

    for row_num in rows_to_process:
        outcome = process_row(ws, row_num, excel_path, dry_run=dry_run)
        counts[outcome] = counts.get(outcome, 0) + 1

    print(f"\n{'='*60}")
    print(f"Done.")
    print(f"  Created:    {counts['created']}")
    print(f"  Duplicates: {counts['duplicate']}")
    print(f"  Errors:     {counts['error']}")
    print(f"  Skipped:    {counts['skipped']}")
    if dry_run:
        print(f"  Dry run:    {counts['dry_run']}")
    print(f"\nExcel updated: {excel_path}")
    print(f"Debug JSON files: {TMP_DIR}/row_*_extracted.json")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Upload purchase bills to Odoo from Excel + Google Drive PDF links"
    )
    parser.add_argument("excel",
                        help="Path to Excel file (.xlsx)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Download + OCR each PDF but do NOT create bills in Odoo")
    parser.add_argument("--start-row", type=int, default=2,
                        help="First data row number (default: 2, skips header row 1)")
    parser.add_argument("--row", type=int, default=None,
                        help="Process only this single row number")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    if not excel_path.suffix.lower() == ".xlsx":
        print(f"ERROR: File must be .xlsx format: {excel_path}")
        sys.exit(1)

    print(f"PDF Purchase Upload Agent")
    print(f"Excel: {excel_path}")
    if args.dry_run:
        print("Mode:  DRY RUN (no Odoo changes)")
    else:
        print("Mode:  LIVE (will create bills in Odoo as Draft)")

    process_excel(
        excel_path,
        dry_run=args.dry_run,
        start_row=args.start_row,
        single_row=args.row,
    )


if __name__ == "__main__":
    main()
