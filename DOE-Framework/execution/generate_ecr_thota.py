"""
generate_ecr_thota.py
---------------------
Generates EPFO ECR 2.0 text file for THOTA HOSPITALITY LLP
from the payroll Excel (sheet: Salary Stmt-MAR26).

ECR 2.0 format (pipe-separated, 11 fields per line):
  UAN#~#MEMBER_NAME#~#GROSS_WAGES#~#EPF_WAGES#~#EPS_WAGES#~#EDLI_WAGES#~#
  EPF_CONT_EE#~#EPS_CONT_ER#~#EPF_DIFF_ER#~#NCP_DAYS#~#REFUND_OF_ADVANCES

Column mapping for Thota MAR26 sheet (0-indexed):
  0: Sl No   1: Position   2: Emp Name   3: UAN NO   4: ESI NO
  6: Calendar  7: Present  8: Holiday  10: LOP/Absent
  12: Paid Days  13: LOP Days
  14: Gross Salary  15: LOP  16: Eff.Gross  17: Basic  18: HRA
  19: Convnc  20: Medical  21: Spl.Allow  22: Incentive  23: OT
  24: Total Eff. Gross
  25: Employee PF   26: Employee ESI   27: Employer PF   28: Employer ESI
  29: PT  30: TDS  31: Sal.Adv  32: Total Ded  33: Arrears  34: Bonus
  35: Net Sal  36: Mode  37: CTC

Usage:
    python generate_ecr_thota.py --file <path> --sheet "Salary Stmt-MAR26"
                                  --wage-month 03 --wage-year 2026
                                  --establishment-id <EPFO_ID>
"""

import argparse
import json
import math
import os
import re
import sys
from pathlib import Path

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SEP = "#~#"

# Column indices (0-based) for Thota format
COL = {
    "sl_no":          0,
    "position":       1,
    "emp_name":       2,
    "uan":            3,
    "esi_no":         4,
    "doj":            5,
    "calendar":       6,
    "present":        7,
    "holiday":        8,
    "sl_el_cl":       9,
    "lop_absent":     10,
    "total_days":     11,
    "paid_days":      12,
    "lop_days":       13,
    "gross":          14,
    "lop_amount":     15,
    "eff_gross":      16,
    "basic":          17,
    "hra":            18,
    "convnc":         19,
    "medical":        20,
    "spl_allow":      21,
    "incentive":      22,
    "overtime":       23,
    "total_eff_gross":24,
    "emp_pf":         25,
    "emp_esi":        26,
    "er_pf":          27,
    "er_esi":         28,
    "pt":             29,
    "tds":            30,
    "sal_adv":        31,
    "total_ded":      32,
    "arrears":        33,
    "bonus":          34,
    "net_sal":        35,
    "mode":           36,
    "ctc":            37,
}


def round_epfo(val: float) -> int:
    """Round half-up to integer (EPFO ECR requires integer rupee values)."""
    return math.floor(float(val) + 0.5)


def sanitize_name(name: str) -> str:
    name = name.upper()
    name = re.sub(r"[^A-Z0-9 ]", " ", name)
    name = re.sub(r" +", " ", name).strip()
    return name[:50]


def clean_uan(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().lstrip("'")
    # Remove decimal from float representation (e.g., '100666890455.0' → '100666890455')
    if '.' in s:
        s = s.split('.')[0]
    if re.fullmatch(r"\d{12}", s):
        return s
    return ""


def clean_val(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_sheet(ws) -> list:
    """Parse salary sheet, return list of employee dicts."""
    employees = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue  # skip header rows

        # Stop at summary / consultant rows (col 0 is not a positive integer, or col 2 is None)
        sl = row[COL["sl_no"]]
        if sl is None or not isinstance(sl, (int, float)) or sl <= 0:
            continue

        emp_name = str(row[COL["emp_name"]] or "").strip()
        if not emp_name:
            continue

        emp = {
            "sl_no":          int(sl),
            "position":       str(row[COL["position"]] or "").strip(),
            "emp_name":       emp_name,
            "uan":            clean_uan(row[COL["uan"]]),
            "esi_no":         str(row[COL["esi_no"]] or "").strip(),
            "calendar":       clean_val(row[COL["calendar"]]),
            "present":        clean_val(row[COL["present"]]),
            "lop_days":       clean_val(row[COL["lop_days"]]),
            "gross":          clean_val(row[COL["gross"]]),
            "lop_amount":     clean_val(row[COL["lop_amount"]]),
            "eff_gross":      clean_val(row[COL["eff_gross"]]),
            "total_eff_gross":clean_val(row[COL["total_eff_gross"]]),
            "emp_pf":         clean_val(row[COL["emp_pf"]]),
            "emp_esi":        clean_val(row[COL["emp_esi"]]),
            "er_pf":          clean_val(row[COL["er_pf"]]),
            "er_esi":         clean_val(row[COL["er_esi"]]),
            "pt":             clean_val(row[COL["pt"]]),
            "net_sal":        clean_val(row[COL["net_sal"]]),
        }
        employees.append(emp)
    return employees


def build_ecr_line(emp: dict) -> str | None:
    """Build one ECR 2.0 line. Returns None if employee is not PF-eligible."""
    emp_pf = emp.get("emp_pf", 0)
    er_pf  = emp.get("er_pf", 0)
    uan    = emp.get("uan", "")

    if emp_pf <= 0 or not uan:
        return None

    epf_wages   = round_epfo(emp_pf / 0.12)
    eps_wages   = min(epf_wages, 15000)
    edli_wages  = min(epf_wages, 15000)
    eps_cont_er = round_epfo(eps_wages * 8.33 / 100)
    epf_diff_er = round_epfo(er_pf) - eps_cont_er
    epf_cont_ee = round_epfo(emp_pf)
    gross_wages = round_epfo(emp.get("total_eff_gross", emp.get("eff_gross", 0)))
    ncp_days    = int(emp.get("lop_days", 0))
    refund_adv  = 0
    name        = sanitize_name(emp.get("emp_name", ""))

    fields = [
        uan, name, gross_wages, epf_wages, eps_wages, edli_wages,
        epf_cont_ee, eps_cont_er, epf_diff_er, ncp_days, refund_adv,
    ]
    return SEP.join(str(f) for f in fields)


def load_thota_config() -> dict:
    """Load directives/thota_config.json if it exists."""
    cfg_path = Path(__file__).parent.parent / "directives" / "thota_config.json"
    if cfg_path.exists():
        with open(cfg_path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def main():
    cfg = load_thota_config()
    default_estab = cfg.get("epfo_establishment_id", "ESTAB_CODE")
    default_file  = cfg.get("payroll_file", "")

    parser = argparse.ArgumentParser(description="Generate EPFO ECR 2.0 for Thota Hospitality LLP")
    parser.add_argument("--file",             default=default_file, help="Path to payroll .xlsx")
    parser.add_argument("--sheet",            default="Salary Stmt-MAR26", help="Sheet name")
    parser.add_argument("--wage-month",       required=True,  help="MM  e.g. 03")
    parser.add_argument("--wage-year",        required=True,  help="YYYY e.g. 2026")
    parser.add_argument("--establishment-id", default=default_estab,
                        help=f"EPFO establishment code (default from thota_config: {default_estab})")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"ERROR: sheet '{args.sheet}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[args.sheet]
    employees = parse_sheet(ws)

    ecr_lines  = []
    included   = []
    skipped_no_uan = []
    skipped_no_pf  = []

    for emp in employees:
        line = build_ecr_line(emp)
        if line:
            ecr_lines.append(line)
            included.append(emp)
        elif emp.get("emp_pf", 0) > 0:
            skipped_no_uan.append(emp["emp_name"])

    out_dir = Path("outputs/hr_compliance")
    out_dir.mkdir(parents=True, exist_ok=True)
    ecr_filename = f"ECR_{args.wage_year}{args.wage_month}_{args.establishment_id}.txt"
    ecr_path = out_dir / ecr_filename

    with open(ecr_path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(ecr_lines))

    print(f"\n[OK]  ECR file written : {ecr_path}")
    print(f"      Employees included : {len(included)}")
    for emp in included:
        epf_wages   = round_epfo(emp["emp_pf"] / 0.12)
        eps_wages   = min(epf_wages, 15000)
        eps_cont_er = round_epfo(eps_wages * 8.33 / 100)
        epf_diff_er = round_epfo(emp["er_pf"]) - eps_cont_er
        print(f"    {emp['emp_name']:<22} UAN:{emp['uan']}  "
              f"EPFWages:{epf_wages:>6}  EE:{round_epfo(emp['emp_pf']):>5}  "
              f"ER:{round_epfo(emp['er_pf']):>5}  EPS:{eps_cont_er:>5}  Diff:{epf_diff_er:>4}  LOP:{int(emp['lop_days'])}")

    if skipped_no_uan:
        print(f"\n[WARN] PF deducted but NO UAN — excluded from ECR (must be resolved):")
        for name in skipped_no_uan:
            print(f"       ✗ {name}")

    pf_emps = included
    total_ee_pf = sum(round_epfo(e["emp_pf"]) for e in pf_emps)
    total_er_pf = sum(round_epfo(e["er_pf"])  for e in pf_emps)
    total_eps   = sum(round_epfo(min(round_epfo(e["emp_pf"]/0.12), 15000) * 8.33 / 100) for e in pf_emps)
    total_diff  = total_er_pf - total_eps

    next_month = int(args.wage_month) % 12 + 1
    next_year  = int(args.wage_year) + (1 if int(args.wage_month) == 12 else 0)

    print(f"\n  Establishment ID   : {args.establishment_id}")
    print(f"  Wage Month         : {args.wage_month}/{args.wage_year}")
    print(f"  Total Employee PF  : ₹{total_ee_pf:,}")
    print(f"  Total Employer PF  : ₹{total_er_pf:,}")
    print(f"    └─ EPS (8.33%)   : ₹{total_eps:,}")
    print(f"    └─ EPF Diff(3.67%): ₹{total_diff:,}")
    print(f"  ──────────────────────────────────")
    print(f"  Total PF Remittance: ₹{total_ee_pf + total_er_pf:,}")
    print(f"\n  Statutory Deadline : 15th {next_year}-{next_month:02d}")
    print(f"  >> TARGET (1 day early): 14th {next_year}-{next_month:02d}")

    # ── PF Summary Excel ──────────────────────────────────────────────────────
    try:
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "PF Summary"

        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF")
        headers = [
            "Sl No", "Employee Name", "UAN", "Position",
            "Total Eff. Gross", "EPF Wages", "EPS Wages", "EDLI Wages",
            "EE PF (12%)", "ER EPS (8.33%)", "ER EPF Diff (3.67%)", "Total ER PF",
            "NCP Days",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        row_num = 2
        for emp in pf_emps:
            epf_wages   = round_epfo(emp["emp_pf"] / 0.12)
            eps_wages   = min(epf_wages, 15000)
            edli_wages  = eps_wages
            eps_cont_er = round_epfo(eps_wages * 8.33 / 100)
            epf_diff_er = round_epfo(emp["er_pf"]) - eps_cont_er
            ws2.append([
                emp["sl_no"],
                emp["emp_name"],
                emp["uan"],
                emp["position"],
                round_epfo(emp.get("total_eff_gross", emp.get("eff_gross", 0))),
                epf_wages,
                eps_wages,
                edli_wages,
                round_epfo(emp["emp_pf"]),
                eps_cont_er,
                epf_diff_er,
                round_epfo(emp["er_pf"]),
                int(emp.get("lop_days", 0)),
            ])
            row_num += 1

        # Totals row
        ws2.append([
            "", "TOTAL", "", "",
            f"=SUM(E2:E{row_num-1})", f"=SUM(F2:F{row_num-1})", "",  "",
            f"=SUM(I2:I{row_num-1})", f"=SUM(J2:J{row_num-1})",
            f"=SUM(K2:K{row_num-1})", f"=SUM(L2:L{row_num-1})", "",
        ])
        for cell in ws2[row_num]:
            cell.font = Font(bold=True)

        for col in ws2.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 32)

        xls_path = out_dir / f"PF_Summary_Thota_{args.wage_year}{args.wage_month}.xlsx"
        wb2.save(xls_path)
        print(f"\n[OK]  PF Summary Excel : {xls_path}")
    except Exception as e:
        print(f"  (Excel summary skipped: {e})", file=sys.stderr)


if __name__ == "__main__":
    main()
