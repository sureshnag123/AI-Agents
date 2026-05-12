"""
generate_ecr.py
---------------
Generates EPFO ECR 2.0 (Electronic Challan cum Return) text file
from parsed payroll JSON.

ECR 2.0 format (pipe-separated, one employee per line):
  #~#
  [UAN]~[MEMBER_NAME]~[GROSS_WAGES]~[EPF_WAGES]~[EPS_WAGES]~[EDLI_WAGES]~
  [EPF_CONT_EE]~[EPS_CONT_ER]~[EPF_DIFF_ER]~[EPS_DIFF]~[NCP_DAYS]~[REFUND_ADV]

Field derivation from Excel values:
  EPF_WAGES     = round(emp_pf / 0.12)  — reverse from Excel's computed PF
  EPS_WAGES     = min(EPF_WAGES, 15000)
  EPS_CONT_ER   = round(EPS_WAGES * 8.33 / 100)
  EPF_DIFF_ER   = er_pf - EPS_CONT_ER
  GROSS_WAGES   = total_eff_gross (as per EPFO requirement)
  NCP_DAYS      = lop_days

Usage:
    python generate_ecr.py --json .tmp/payroll_Salary_Sheet_Mar.json
                           --wage-month 03
                           --wage-year 2026
                           --establishment-id KARBN001234   (your EPFO estab code)

Output: outputs/hr_compliance/ECR_YYYYMM_<estab>.txt
"""

import sys, io
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import argparse
import json
import math
import os
import re
import sys
from pathlib import Path


SEPARATOR = "#~#"   # EPFO ECR 2.0 field separator


def round_epfo(val: float) -> int:
    """EPFO ECR requires integer rupee values (round half-up)."""
    return math.floor(val + 0.5)


def sanitize_name(name: str) -> str:
    """ECR member name: uppercase, only alphanumeric + spaces, max 50 chars."""
    name = name.upper()
    name = re.sub(r"[^A-Z0-9 ]", " ", name)
    name = re.sub(r" +", " ", name).strip()
    return name[:50]


def build_ecr_line(emp: dict) -> str | None:
    """
    Build one ECR line for an employee.
    Returns None if employee is not PF-eligible.
    """
    emp_pf = emp.get("emp_pf", 0)
    er_pf  = emp.get("er_pf", 0)
    uan    = emp.get("uan", "")

    # Skip if no PF contribution
    if emp_pf <= 0 or not uan or len(uan) != 12:
        return None

    # Derive EPF wage from Excel-computed PF amount
    epf_wages = round_epfo(emp_pf / 0.12)

    # EPS wage capped at ₹15,000
    # Non-EPS members (flagged in directives/non_eps_members.json): EPS=0
    non_eps = emp.get("_non_eps", False)
    eps_wages   = 0 if non_eps else min(epf_wages, 15000)

    # EDLI wage = EPF wage, capped at ₹15,000 (required field)
    edli_wages  = min(epf_wages, 15000)

    # Employer contributions
    eps_cont_er = 0 if non_eps else round_epfo(eps_wages * 8.33 / 100)
    epf_diff_er = round_epfo(er_pf) - eps_cont_er  # EPF_ER - EPS = 3.67% portion

    # Employee PF contribution (from Excel)
    epf_cont_ee = round_epfo(emp_pf)

    # Gross wages for EPFO = total effective gross (incl. OT, incentives)
    gross_wages = round_epfo(emp.get("total_eff_gross", 0))

    ncp_days    = int(emp.get("lop_days", 0))
    refund_adv  = 0

    name = sanitize_name(emp.get("emp_name", ""))

    # 11 fields — EPFO ECR 2.0 correct format (EPS_DIFF excluded, EDLI included)
    fields = [
        uan,          # 1.  UAN (12 digits)
        name,         # 2.  MEMBER_NAME
        gross_wages,  # 3.  GROSS_WAGES
        epf_wages,    # 4.  EPF_WAGES
        eps_wages,    # 5.  EPS_WAGES  (0 if non-EPS member)
        edli_wages,   # 6.  EDLI_WAGES (= EPF_WAGES capped at 15000)
        epf_cont_ee,  # 7.  EPF_CONT_EE (12%)
        eps_cont_er,  # 8.  EPS_CONT_ER (8.33%, 0 if non-EPS)
        epf_diff_er,  # 9.  EPF_DIFF_ER (3.67%)
        ncp_days,     # 10. NCP_DAYS
        refund_adv,   # 11. REFUND_OF_ADVANCES
    ]
    return SEPARATOR.join(str(f) for f in fields)


def load_non_eps_members() -> set:
    """Load UANs of non-EPS members from directives/non_eps_members.json."""
    path = Path(__file__).parent.parent / "directives" / "non_eps_members.json"
    if not path.exists():
        return set()
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    return {v["uan"] for k, v in raw.items() if not k.startswith("_") and "uan" in v}


def generate_ecr(data: dict, wage_month: str, wage_year: str, estab_id: str) -> str:
    """Returns full ECR file content as string."""
    employees = data.get("employees", [])
    non_eps_uans = load_non_eps_members()

    lines = []
    skipped = []
    included = []
    non_eps_flagged = []

    for emp in employees:
        if emp.get("uan") in non_eps_uans:
            emp["_non_eps"] = True
            non_eps_flagged.append(emp["emp_name"])
        line = build_ecr_line(emp)
        if line:
            lines.append(line)
            included.append(emp["emp_name"])
        elif emp.get("emp_pf", 0) > 0:
            skipped.append(emp["emp_name"])

    return "\n".join(lines), included, skipped, non_eps_flagged


def load_company_config() -> dict:
    """Load directives/company_config.json if it exists."""
    cfg_path = Path(__file__).parent.parent / "directives" / "company_config.json"
    if cfg_path.exists():
        with open(cfg_path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def main():
    cfg = load_company_config()
    default_estab = cfg.get("epfo_establishment_id", "ESTAB_CODE")

    parser = argparse.ArgumentParser(description="Generate EPFO ECR 2.0 file")
    parser.add_argument("--json", required=True, help="Path to parsed payroll JSON (.tmp/payroll_*.json)")
    parser.add_argument("--wage-month", required=True, help="Wage month (MM), e.g. 03")
    parser.add_argument("--wage-year",  required=True, help="Wage year (YYYY), e.g. 2026")
    parser.add_argument("--establishment-id", default=default_estab,
                        help=f"EPFO establishment code (default from company_config: {default_estab})")
    args = parser.parse_args()

    with open(args.json, encoding="utf-8") as f:
        data = json.load(f)

    content, included, skipped, non_eps_flagged = generate_ecr(
        data, args.wage_month, args.wage_year, args.establishment_id
    )

    out_dir = Path("outputs/hr_compliance")
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = f"ECR_{args.wage_year}{args.wage_month}_{args.establishment_id}.txt"
    out_path = out_dir / filename

    with open(out_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(content)

    print(f"[OK] ECR file written: {out_path}")
    print(f"  Employees included : {len(included)}")
    for name in included:
        print(f"    + {name}")
    if non_eps_flagged:
        print(f"  [INFO] Non-EPS members (EPS_WAGES=0): {', '.join(non_eps_flagged)}")
    if skipped:
        print(f"  [WARN] Skipped (bad UAN):", file=sys.stderr)
        for name in skipped:
            print(f"    - {name}", file=sys.stderr)

    # Print summary totals
    employees = data.get("employees", [])
    pf_emps = [e for e in employees if e.get("emp_pf", 0) > 0 and len(e.get("uan", "")) == 12]
    total_emp_pf = sum(e["emp_pf"] for e in pf_emps)
    total_er_pf  = sum(e["er_pf"]  for e in pf_emps)
    print(f"\n  Establishment ID   : {args.establishment_id}")
    print(f"  Total Employee PF  : ₹{total_emp_pf:,.2f}")
    print(f"  Total Employer PF  : ₹{total_er_pf:,.2f}")
    print(f"  Total PF Remit     : ₹{total_emp_pf + total_er_pf:,.2f}")

    next_month = int(args.wage_month) % 12 + 1
    next_year  = int(args.wage_year) + (1 if int(args.wage_month) == 12 else 0)
    print(f"\n  Deadline  : 15th {next_year}-{next_month:02d} (actual)")
    print(f"  >> Target : 14th {next_year}-{next_month:02d} (1 day early)")

    # Also write a human-readable summary Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PF Summary"

        headers = [
            "Sl No", "Employee Name", "UAN", "Dept",
            "Gross Wages", "EPF Wages", "EPS Wages",
            "EE PF (12%)", "ER EPS (8.33%)", "ER EPF Diff (3.67%)",
            "Total ER PF", "NCP Days"
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        row_num = 2
        for emp in pf_emps:
            epf_wages   = round_epfo(emp["emp_pf"] / 0.12)
            eps_wages   = min(epf_wages, 15000)
            eps_cont_er = round_epfo(eps_wages * 8.33 / 100)
            epf_diff_er = round_epfo(emp["er_pf"]) - eps_cont_er

            ws.append([
                int(emp["sl_no"]),
                emp["emp_name"],
                emp["uan"],
                emp["dept"],
                round_epfo(emp.get("total_eff_gross", 0)),
                epf_wages,
                eps_wages,
                round_epfo(emp["emp_pf"]),
                eps_cont_er,
                epf_diff_er,
                round_epfo(emp["er_pf"]),
                int(emp.get("lop_days", 0)),
            ])
            row_num += 1

        # Totals row
        ws.append([
            "", "TOTAL", "", "",
            f"=SUM(E2:E{row_num-1})",
            f"=SUM(F2:F{row_num-1})",
            "",
            f"=SUM(H2:H{row_num-1})",
            f"=SUM(I2:I{row_num-1})",
            f"=SUM(J2:J{row_num-1})",
            f"=SUM(K2:K{row_num-1})",
            "",
        ])
        total_cell_font = Font(bold=True)
        for cell in ws[row_num]:
            cell.font = total_cell_font

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

        xls_path = out_dir / f"PF_Summary_{args.wage_year}{args.wage_month}.xlsx"
        wb.save(xls_path)
        print(f"\n[OK] PF Summary Excel: {xls_path}")
    except Exception as e:
        print(f"  (Excel summary skipped: {e})", file=sys.stderr)


if __name__ == "__main__":
    main()
