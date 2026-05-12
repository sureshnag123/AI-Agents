"""
parse_payroll.py
----------------
Reads Fracktal Works payroll Excel and returns structured employee records
split by section: regular, directors, interns, professional_service.

Usage:
    python parse_payroll.py --file <path_to_excel> --sheet "Salary Sheet_Mar"

Output (stdout): JSON array of employee dicts
Also writes: .tmp/payroll_<sheet>.json
"""

import sys, io
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import argparse
import json
import os
import re
import sys
from datetime import datetime

import openpyxl


# ── Column indices (0-based) in the payroll sheet ──────────────────────────
COL = {
    "sl_no":        0,
    "dept":         1,
    "emp_name":     2,
    "designation":  3,
    "uan":          4,
    "pan":          5,
    "esi_no":       6,
    "bank_account": 7,
    "doj":          8,
    "calendar":     9,
    "present":      10,
    "holiday":      11,
    "leaves":       12,
    "lop_days_att": 13,
    "total":        14,
    "paid_days":    15,
    "lop_days":     16,
    "ctc":          17,
    "lop_amount":   18,
    "eff_gross":    19,
    "basic":        20,
    "hra":          21,
    "conv_allow":   22,
    "med_allow":    23,
    "spl_allow":    24,
    "incentive":    25,
    "overtime":     26,
    "total_eff_gross": 27,
    "emp_pf":       28,
    "emp_esi":      29,
    "er_pf":        30,
    "er_esi":       31,
    "pt":           32,
    "tds":          33,
    "sal_adv":      34,
    "total_ded":    35,
    "arrears":      36,
    "attn_bonus":   37,
    "net_sal":      38,
    "remarks":      39,
}

# Section labels as they appear in col C of the sheet
SECTION_LABELS = {
    "TOTAL",
    "DIRECTOR REMUNERATION",
    "INTERNSHIPS/ Under Contract",
    "PROFESSIONAL SERVICE",
    "GRAND TOTAL",
}


def clean_uan(raw) -> str:
    """Normalise UAN to a 12-digit string, or '' if invalid."""
    if raw is None:
        return ""
    s = str(raw).strip().replace(" ", "")
    # Remove leading apostrophe if any
    s = s.lstrip("'")
    if s == "0" or s == "":
        return ""
    # Must be numeric and exactly 12 digits
    if re.fullmatch(r"\d{12}", s):
        return s
    return s  # return as-is so caller can flag it


def clean_value(v):
    """Return numeric value or 0."""
    if v is None:
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def row_to_employee(row, section: str) -> dict:
    emp = {
        "section":      section,
        "sl_no":        clean_value(row[COL["sl_no"]]),
        "dept":         str(row[COL["dept"]] or "").strip(),
        "emp_name":     str(row[COL["emp_name"]] or "").strip(),
        "designation":  str(row[COL["designation"]] or "").strip(),
        "uan":          clean_uan(row[COL["uan"]]),
        "pan":          str(row[COL["pan"]] or "").strip(),
        "esi_no":       str(row[COL["esi_no"]] or "").strip(),
        "bank_account": str(row[COL["bank_account"]] or "").strip(),
        "doj":          str(row[COL["doj"]].date() if isinstance(row[COL["doj"]], datetime) else row[COL["doj"]] or ""),
        "calendar_days": clean_value(row[COL["calendar"]]),
        "present_days": clean_value(row[COL["present"]]),
        "paid_days":    clean_value(row[COL["paid_days"]]),
        "lop_days":     clean_value(row[COL["lop_days"]]),
        "ctc":          clean_value(row[COL["ctc"]]),
        "lop_amount":   clean_value(row[COL["lop_amount"]]),
        "eff_gross":    clean_value(row[COL["eff_gross"]]),
        "basic":        clean_value(row[COL["basic"]]),
        "hra":          clean_value(row[COL["hra"]]),
        "conv_allow":   clean_value(row[COL["conv_allow"]]),
        "med_allow":    clean_value(row[COL["med_allow"]]),
        "spl_allow":    clean_value(row[COL["spl_allow"]]),
        "incentive":    clean_value(row[COL["incentive"]]),
        "overtime":     clean_value(row[COL["overtime"]]),
        "total_eff_gross": clean_value(row[COL["total_eff_gross"]]),
        "emp_pf":       round(clean_value(row[COL["emp_pf"]]), 2),
        "emp_esi":      round(clean_value(row[COL["emp_esi"]]), 2),
        "er_pf":        round(clean_value(row[COL["er_pf"]]), 2),
        "er_esi":       round(clean_value(row[COL["er_esi"]]), 2),
        "pt":           round(clean_value(row[COL["pt"]]), 2),
        "tds":          round(clean_value(row[COL["tds"]]), 2),
        "sal_adv":      round(clean_value(row[COL["sal_adv"]]), 2),
        "total_ded":    round(clean_value(row[COL["total_ded"]]), 2),
        "arrears":      round(clean_value(row[COL["arrears"]]), 2),
        "attn_bonus":   round(clean_value(row[COL["attn_bonus"]]), 2),
        "net_sal":      round(clean_value(row[COL["net_sal"]]), 2),
        "remarks":      str(row[COL["remarks"]] or "").strip(),
    }
    return emp


def parse_sheet(ws) -> dict:
    """
    Parse one salary sheet.
    Returns dict with keys: month_label, employees (list), sections_summary.
    """
    employees = []
    current_section = "regular"
    month_label = ""

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Row 1: company name (skip)
        # Row 2: month label
        if i == 1:
            month_label = str(row[0] or "").strip()
            continue
        # Rows 3-4: headers (skip)
        if i < 4:
            continue

        # Check col C (index 2) for section labels
        col_c = str(row[COL["emp_name"]] or "").strip().upper()
        for label in SECTION_LABELS:
            if label.upper() in col_c:
                if "DIRECTOR" in col_c:
                    current_section = "director"
                elif "INTERN" in col_c or "CONTRACT" in col_c:
                    current_section = "intern"
                elif "PROFESSIONAL" in col_c:
                    current_section = "professional_service"
                break

        # Skip non-employee rows (TOTAL, labels, empty)
        sl = row[COL["sl_no"]]
        if sl is None or not isinstance(sl, (int, float)) or sl == 0:
            continue

        emp = row_to_employee(row, current_section)
        employees.append(emp)

    return {
        "month_label": month_label,
        "employees": employees,
    }


def validate(employees: list) -> list:
    """Returns list of warning strings."""
    warnings = []
    pf_emps = [e for e in employees if e["emp_pf"] > 0]
    for e in pf_emps:
        if not e["uan"] or len(e["uan"]) != 12:
            warnings.append(f"[UAN] {e['emp_name']}: UAN '{e['uan']}' is not 12 digits")
        if not e["pan"]:
            warnings.append(f"[PAN] {e['emp_name']}: PAN is missing")

    esi_emps = [e for e in employees if e["emp_esi"] > 0]
    for e in esi_emps:
        if not e["esi_no"]:
            warnings.append(f"[ESI] {e['emp_name']}: ESI insurance number missing — required for ESIC portal")

    return warnings


def main():
    parser = argparse.ArgumentParser(description="Parse Fracktal payroll Excel")
    parser.add_argument("--file", required=True, help="Path to payroll .xlsx")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: latest Salary Sheet)")
    parser.add_argument("--list-sheets", action="store_true", help="List available salary sheets and exit")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)

    salary_sheets = [s for s in wb.sheetnames if s.startswith("Salary Sheet")]

    if args.list_sheets:
        print("Available salary sheets:")
        for s in salary_sheets:
            print(f"  {s}")
        return

    sheet_name = args.sheet or salary_sheets[0]
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {salary_sheets}", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]
    result = parse_sheet(ws)

    # Apply PAN overrides from directives/pan_numbers.json
    pan_overrides_path = os.path.join(os.path.dirname(__file__), "..", "directives", "pan_numbers.json")
    if os.path.exists(pan_overrides_path):
        with open(pan_overrides_path, encoding="utf-8") as f:
            raw = json.load(f)
        pan_overrides = {k.upper(): v for k, v in raw.items() if not k.startswith("_")}
        patched = 0
        for emp in result["employees"]:
            if not emp.get("pan"):
                match = pan_overrides.get(emp["emp_name"].upper())
                if match:
                    emp["pan"] = match
                    patched += 1
        if patched:
            print(f"[OK] Applied PAN overrides to {patched} employee(s)", file=sys.stderr)

    warnings = validate(result["employees"])
    if warnings:
        print("\n[WARN] VALIDATION WARNINGS:", file=sys.stderr)
        for w in warnings:
            print(f"  {w}", file=sys.stderr)
        print("", file=sys.stderr)

    # Write to .tmp/
    os.makedirs(".tmp", exist_ok=True)
    safe_name = sheet_name.replace(" ", "_").replace("/", "-")
    out_path = f".tmp/payroll_{safe_name}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(json.dumps(result, indent=2, ensure_ascii=False))
    print(f"\n[OK] Written to {out_path}", file=sys.stderr)
    print(f"  Total employees: {len(result['employees'])}", file=sys.stderr)

    section_counts = {}
    for e in result["employees"]:
        section_counts[e["section"]] = section_counts.get(e["section"], 0) + 1
    for sec, cnt in section_counts.items():
        print(f"  {sec}: {cnt}", file=sys.stderr)


if __name__ == "__main__":
    main()
