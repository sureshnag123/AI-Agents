"""
generate_pt_challan.py
----------------------
Generates Karnataka Professional Tax deduction register and challan summary.

Karnataka PT Slabs (FY 2025-26, per the Karnataka Tax on Professions,
Trades, Callings and Employments Act, 1976):

  Gross Salary / Month     |  PT / Month
  -------------------------|-------------
  Up to ₹15,000            |  Nil
  ₹15,001 to ₹24,999       |  ₹150
  ₹25,000 and above        |  ₹200

  NOTE: For Directors with gross ≥ ₹25,000 the same ₹200 applies.
        The ₹300 seen in some payrolls is employer-side PT registration fee —
        the script uses the Excel's PT value as authoritative.

Challan payment schedule (Karnataka):
  - Employers with > 20 employees: pay monthly by 20th of following month
  - Employers with ≤ 20 employees: pay half-yearly (April–Sep by Oct 20,
    Oct–Mar by Apr 20)

Usage:
    python generate_pt_challan.py --json .tmp/payroll_Salary_Sheet_Mar.json
                                  --wage-month 03
                                  --wage-year 2026

Output:
    outputs/hr_compliance/PT_Challan_YYYYMM.xlsx
"""

import sys, io
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import argparse
import json
import math
import os
import sys
from pathlib import Path


# Karnataka PT slabs
KA_PT_SLABS = [
    (0,      15000,  0),
    (15001,  24999,  150),
    (25000,  float("inf"), 200),
]


def compute_pt_slab(gross: float) -> int:
    """Compute PT per Karnataka slab from gross monthly wages."""
    for low, high, pt in KA_PT_SLABS:
        if low <= gross <= high:
            return pt
    return 0


def load_company_config() -> dict:
    cfg_path = Path(__file__).parent.parent / "directives" / "company_config.json"
    if cfg_path.exists():
        with open(cfg_path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def main():
    cfg = load_company_config()
    default_rcn = cfg.get("pt_registration_no", "")

    parser = argparse.ArgumentParser(description="Generate Karnataka PT challan")
    parser.add_argument("--json", required=True, help="Path to parsed payroll JSON")
    parser.add_argument("--wage-month", required=True, help="Month MM")
    parser.add_argument("--wage-year",  required=True, help="Year YYYY")
    parser.add_argument("--verify-slabs", action="store_true",
                        help="Flag employees where Excel PT doesn't match Karnataka slab")
    parser.add_argument("--rcn", default=default_rcn,
                        help=f"PT Registration Certificate No (default: {default_rcn})")
    args = parser.parse_args()

    with open(args.json, encoding="utf-8") as f:
        data = json.load(f)

    employees = data.get("employees", [])
    month_label = data.get("month_label", "")

    # Only employees with PT deducted
    pt_employees = [e for e in employees if e.get("pt", 0) > 0]
    zero_pt      = [e for e in employees if e.get("pt", 0) == 0
                    and (e.get("total_eff_gross") or 0) > 15000
                    and e.get("section") not in ("intern", "professional_service")]

    out_dir = Path("outputs/hr_compliance")
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: PT Deduction Register ────────────────────────────────
        ws1 = wb.active
        ws1.title = "PT_Deduction_Register"

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="7030A0")
        hdr_font = Font(bold=True, color="FFFFFF")

        # Title
        ws1.merge_cells("A1:H1")
        title_cell = ws1["A1"]
        title_cell.value = f"FRACKTAL WORKS PRIVATE LIMITED — PT Deduction Register — {month_label or args.wage_month + '/' + args.wage_year}"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center")

        headers = [
            "Sl No", "Employee Name", "Dept", "Designation",
            "Gross Salary (₹)", "PT Slab (Karnataka)", "PT Deducted (₹)", "Remarks"
        ]
        for col, h in enumerate(headers, 1):
            c = ws1.cell(row=2, column=col, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")
            c.border = border

        row = 3
        total_pt = 0
        discrepancies = []

        for i, emp in enumerate(pt_employees, 1):
            gross     = emp.get("total_eff_gross") or emp.get("eff_gross") or 0
            pt_excel  = emp.get("pt", 0)
            pt_slab   = compute_pt_slab(gross)
            section   = emp.get("section", "")

            remark = ""
            if args.verify_slabs and pt_excel != pt_slab:
                remark = f"[WARN] Slab says ₹{pt_slab}, Excel has ₹{pt_excel}"
                discrepancies.append((emp["emp_name"], gross, pt_slab, pt_excel))

            # Slab label
            if gross <= 15000:
                slab_label = "Nil (≤₹15,000)"
            elif gross <= 24999:
                slab_label = "₹150 (₹15,001–₹24,999)"
            else:
                slab_label = "₹200 (₹25,000+)"

            ws1.append([
                i,
                emp.get("emp_name", ""),
                emp.get("dept", ""),
                emp.get("designation", ""),
                round(gross),
                slab_label,
                int(pt_excel),
                remark,
            ])
            for col in range(1, 9):
                ws1.cell(row=row, column=col).border = border

            total_pt += pt_excel
            row += 1

        # Totals row
        ws1.append(["", "TOTAL", "", "", "", "", int(total_pt), ""])
        total_row = ws1[row]
        for cell in total_row:
            cell.font = Font(bold=True)
            cell.border = border

        # Auto-width
        for col in ws1.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

        # ── Sheet 2: Challan Summary ───────────────────────────────────────
        ws2 = wb.create_sheet("Challan_Summary")

        ws2.merge_cells("A1:C1")
        ws2["A1"].value = "PROFESSIONAL TAX CHALLAN SUMMARY"
        ws2["A1"].font = Font(bold=True, size=13, color="7030A0")
        ws2["A1"].alignment = Alignment(horizontal="center")

        summary_data = [
            ("", ""),
            ("Employer Name",              "FRACKTAL WORKS PRIVATE LIMITED"),
            ("PT Registration No (RCN)",   args.rcn or "—"),
            ("State",                      "Karnataka"),
            ("Wage Month",                 f"{args.wage_month}/{args.wage_year}"),
            ("", ""),
            ("Employees with PT Deduction", len(pt_employees)),
            ("", ""),
        ]

        # Break down by slab
        pt_150 = [e for e in pt_employees if e.get("pt") == 150]
        pt_200 = [e for e in pt_employees if e.get("pt") == 200]
        pt_300 = [e for e in pt_employees if e.get("pt") == 300]

        summary_data += [
            ("PT @ ₹150 — Count", len(pt_150)),
            ("PT @ ₹150 — Amount", f"₹{len(pt_150)*150:,}"),
            ("PT @ ₹200 — Count", len(pt_200)),
            ("PT @ ₹200 — Amount", f"₹{len(pt_200)*200:,}"),
        ]
        if pt_300:
            summary_data += [
                ("PT @ ₹300 — Count", len(pt_300)),
                ("PT @ ₹300 — Amount", f"₹{len(pt_300)*300:,}"),
            ]

        summary_data += [
            ("", ""),
            ("TOTAL PT TO REMIT", f"₹{int(total_pt):,}"),
            ("", ""),
            ("Due Date (monthly filer)", f"20th of following month"),
            ("Portal", "https://ptax.karnataka.gov.in"),
        ]

        for key, val in summary_data:
            ws2.append([key, val])
            if key == "TOTAL PT TO REMIT":
                r = ws2.max_row
                ws2.cell(row=r, column=1).font = Font(bold=True, size=11, color="C00000")
                ws2.cell(row=r, column=2).font = Font(bold=True, size=11, color="C00000")

        for col in [1, 2]:
            max_len = max((len(str(c.value)) for c in ws2[get_column_letter(col)] if c.value), default=10)
            ws2.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 45)

        # ── Sheet 3: Zero PT flags (if any) ────────────────────────────────
        if zero_pt:
            ws3 = wb.create_sheet("Zero_PT_Review")
            ws3["A1"].value = "[WARN] Employees earning >₹15,000 with zero PT — review manually"
            ws3["A1"].font = Font(bold=True, color="C00000")
            ws3.append(["Name", "Section", "Gross Salary", "Slab PT", "PT in Excel", "Note"])
            for emp in zero_pt:
                gross = emp.get("total_eff_gross") or emp.get("eff_gross") or 0
                ws3.append([
                    emp["emp_name"],
                    emp["section"],
                    round(gross),
                    compute_pt_slab(gross),
                    emp.get("pt", 0),
                    "Possibly director/exempt or data to verify",
                ])

        out_path = out_dir / f"PT_Challan_{args.wage_year}{args.wage_month}.xlsx"
        wb.save(out_path)

        print(f"[OK] PT Challan file written: {out_path}")
        print(f"  Employees with PT : {len(pt_employees)}")
        print(f"  Total PT to remit : ₹{int(total_pt):,}")
        if pt_150:
            print(f"    @ ₹150 × {len(pt_150)} = ₹{len(pt_150)*150:,}")
        if pt_200:
            print(f"    @ ₹200 × {len(pt_200)} = ₹{len(pt_200)*200:,}")
        if pt_300:
            print(f"    @ ₹300 × {len(pt_300)} = ₹{len(pt_300)*300:,}")
        if discrepancies:
            print(f"\n  [WARN] Slab mismatches detected:")
            for name, gross, slab, excel in discrepancies:
                print(f"    {name}: gross ₹{gross:,.0f} → slab ₹{slab}, Excel ₹{excel}")
        if zero_pt:
            print(f"\n  [INFO] {len(zero_pt)} employee(s) earning >₹15K with ₹0 PT — see 'Zero_PT_Review' sheet")

        next_month = int(args.wage_month) % 12 + 1
        next_year  = int(args.wage_year) + (1 if int(args.wage_month) == 12 else 0)
        print(f"\n  Deadline  : 20th {next_year}-{next_month:02d} (actual)")
        print(f"  >> Target : 19th {next_year}-{next_month:02d} (1 day early)")

    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
