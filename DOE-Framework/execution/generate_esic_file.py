"""
generate_esic_file.py
---------------------
Generates ESIC monthly contribution file for upload to ESIC Employer Portal.

ESIC eligibility: employees with gross salary ≤ ₹21,000/month.
  Employee contribution : 0.75% of gross wages
  Employer contribution : 3.25% of gross wages

The ESIC portal accepts a CSV/Excel upload with columns:
  Insurance No | Employee Name | Days Worked | Total Wages | EE Contribution | ER Contribution

NOTE: ESI Insurance Numbers (col G in payroll) must be filled in the
      Excel before this script will include the employee. If missing,
      the employee is listed in a "pending ESI No" section.

Usage:
    python generate_esic_file.py --json .tmp/payroll_Salary_Sheet_Mar.json
                                 --wage-month 03
                                 --wage-year 2026
                                 --contribution-period "March 2026"

Output: outputs/hr_compliance/ESIC_YYYYMM.xlsx
"""

import sys, io
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import argparse
import json
import math
import os
import sys
from pathlib import Path


ESI_SALARY_LIMIT = 21000     # ₹/month gross wage ceiling
EE_RATE = 0.0075             # 0.75%
ER_RATE = 0.0325             # 3.25%


def round_esic(v: float) -> float:
    """Round to nearest rupee (half-up)."""
    return math.floor(v + 0.5)


def get_esic_employees(employees: list) -> tuple[list, list]:
    """
    Split employees into:
      - eligible: have ESI contributions in payroll (emp_esi > 0)
      - no_esi_no: eligible by wage but ESI number missing
    """
    eligible    = []
    no_esi_no   = []
    zero_esi    = []  # above threshold, no ESI — informational

    for emp in employees:
        gross = emp.get("total_eff_gross") or emp.get("eff_gross") or 0
        emp_esi = emp.get("emp_esi", 0)
        er_esi  = emp.get("er_esi",  0)

        if emp_esi > 0 or er_esi > 0:
            # Has ESI contribution in payroll — needs insurance number
            if emp.get("esi_no"):
                eligible.append(emp)
            else:
                no_esi_no.append(emp)
        # Note: employees with gross <= 21000 but zero ESI in Excel are
        # assumed intentionally excluded by the employer (e.g. LOP, FNF,
        # or employer opted not to cover). Respect the Excel values.

    return eligible, no_esi_no


def load_company_config() -> dict:
    cfg_path = Path(__file__).parent.parent / "directives" / "company_config.json"
    if cfg_path.exists():
        with open(cfg_path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def main():
    cfg = load_company_config()
    default_esic = cfg.get("esic_employer_code", "EMPLOYER_CODE")

    parser = argparse.ArgumentParser(description="Generate ESIC contribution file")
    parser.add_argument("--json", required=True, help="Path to parsed payroll JSON")
    parser.add_argument("--wage-month", required=True, help="Month MM")
    parser.add_argument("--wage-year",  required=True, help="Year YYYY")
    parser.add_argument("--contribution-period", default=None,
                        help="Human label, e.g. 'March 2026'")
    parser.add_argument("--employer-code", default=default_esic,
                        help=f"ESIC employer code (default from config: {default_esic})")
    args = parser.parse_args()

    with open(args.json, encoding="utf-8") as f:
        data = json.load(f)

    employees = data.get("employees", [])
    period = args.contribution_period or f"{args.wage_month}/{args.wage_year}"

    # Load ESI number overrides (directives/esi_numbers.json)
    esi_overrides_path = Path("directives/esi_numbers.json")
    esi_overrides = {}
    if esi_overrides_path.exists():
        with open(esi_overrides_path, encoding="utf-8") as f:
            raw = json.load(f)
        esi_overrides = {k.upper(): v for k, v in raw.items() if not k.startswith("_")}

    # Apply overrides to employees missing ESI number
    patched = 0
    for emp in employees:
        if not emp.get("esi_no"):
            name_upper = emp.get("emp_name", "").upper()
            # Try full name, then first word
            match = esi_overrides.get(name_upper) or esi_overrides.get(name_upper.split()[0])
            if match:
                emp["esi_no"] = match
                patched += 1
    if patched:
        print(f"[OK] Applied ESI number overrides to {patched} employee(s) from {esi_overrides_path}")

    eligible, no_esi_no = get_esic_employees(employees)

    out_dir = Path("outputs/hr_compliance")
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: ESIC Upload File (portal format) ──────────────────────
        ws1 = wb.active
        ws1.title = "ESIC_Upload"

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="1F7A4D")
        hdr_font = Font(bold=True, color="FFFFFF")

        headers = [
            "Insurance No", "Employee Name", "Days Worked",
            "Total Wages (₹)", "EE Contribution (0.75%)", "ER Contribution (3.25%)",
            "Total Contribution"
        ]
        for col, h in enumerate(headers, 1):
            c = ws1.cell(row=1, column=col, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")
            c.border = border

        row = 2
        total_wages = total_ee = total_er = 0

        for emp in eligible:
            gross       = round_esic(emp.get("total_eff_gross") or emp.get("eff_gross") or 0)
            ee_contrib  = round_esic(emp.get("emp_esi", gross * EE_RATE))
            er_contrib  = round_esic(emp.get("er_esi",  gross * ER_RATE))
            days_worked = int(emp.get("paid_days") or emp.get("present_days") or 0)

            ws1.append([
                emp.get("esi_no", ""),
                emp.get("emp_name", ""),
                days_worked,
                gross,
                ee_contrib,
                er_contrib,
                ee_contrib + er_contrib,
            ])
            for col in range(1, 8):
                ws1.cell(row=row, column=col).border = border

            total_wages += gross
            total_ee    += ee_contrib
            total_er    += er_contrib
            row += 1

        # Totals row
        totals_font = Font(bold=True)
        ws1.append(["", "TOTAL", "", total_wages, total_ee, total_er, total_ee + total_er])
        for col in range(1, 8):
            c = ws1.cell(row=row, column=col)
            c.font = totals_font
            c.border = border

        # Auto-width
        for col in ws1.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 35)

        # ── Sheet 2: Pending ESI Numbers ───────────────────────────────────
        if no_esi_no:
            ws2 = wb.create_sheet("Pending_ESI_No")
            ws2.append(["[WARN] Employees missing ESI Insurance Number — fill these and re-run"])
            ws2["A1"].font = Font(bold=True, color="C00000")

            ws2.append(["Sl No", "Name", "Dept", "Gross Salary", "Emp ESI (calc)", "Action Required"])
            for col in range(1, 7):
                c = ws2.cell(row=2, column=col)
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="FFE699")

            for i, emp in enumerate(no_esi_no, 1):
                gross = emp.get("total_eff_gross") or emp.get("eff_gross") or 0
                ws2.append([
                    i,
                    emp.get("emp_name", ""),
                    emp.get("dept", ""),
                    round_esic(gross),
                    round_esic(gross * EE_RATE),
                    "Enter ESI Insurance No in payroll Excel col G",
                ])

        out_path = out_dir / f"ESIC_{args.wage_year}{args.wage_month}.xlsx"
        wb.save(out_path)
        print(f"[OK] ESIC file written: {out_path}")

        if eligible:
            print(f"  Employer Code      : {args.employer_code}")
            print(f"  Employees with ESI : {len(eligible)}")
            for emp in eligible:
                gross = emp.get("total_eff_gross") or emp.get("eff_gross") or 0
                print(f"    {emp['emp_name']}: wages ₹{gross:,.0f}, EE ₹{emp['emp_esi']:,.2f}, ER ₹{emp['er_esi']:,.2f}")
            print(f"  Total EE ESI : ₹{total_ee:,.2f}")
            print(f"  Total ER ESI : ₹{total_er:,.2f}")
            print(f"  Total Remit  : ₹{total_ee + total_er:,.2f}")
        else:
            print("  [INFO] No employees with ESI contributions found in this month's payroll.")

        next_month = int(args.wage_month) % 12 + 1
        next_year  = int(args.wage_year) + (1 if int(args.wage_month) == 12 else 0)
        print(f"\n  Deadline  : 15th {next_year}-{next_month:02d} (actual)")
        print(f"  >> Target : 14th {next_year}-{next_month:02d} (1 day early)")

        if no_esi_no:
            print(f"\n  [WARN] {len(no_esi_no)} employee(s) need ESI Insurance Numbers — see 'Pending_ESI_No' sheet")

    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
