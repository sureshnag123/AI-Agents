"""
payroll_reader.py — Reads Thota Hospitality payroll Excel and returns structured employee data.

Usage:
    from payroll_reader import read_payroll_sheet, list_salary_sheets

Handles column variations across different monthly sheets automatically.
"""

import re
import sys
from pathlib import Path
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# Column-name normalisation rules
# ---------------------------------------------------------------------------
# The Excel sheets use slightly different header names across months.
# We map them all to a canonical set so downstream code never has to care.

_CANONICAL_MAP = {
    # identity mappings (already canonical)
    "sl no": "sl_no",
    "position": "position",
    "emp name": "emp_name",
    "designation": "designation",
    "department": "department",
    "uan no": "uan_no",
    "pf no": "pf_no",
    "esi no": "esi_no",
    "branch": "branch",
    "bank account no": "bank_account_no",
    "doj": "doj",
    "calender": "calendar_days",
    "calendar": "calendar_days",
    "present": "present_days",
    "holiday": "holidays",
    "sl/el/cl": "leaves",
    "lop / absent": "lop_absent",
    "lop/absent": "lop_absent",
    "total": "total_days",
    "paid days": "paid_days",
    "lop days": "lop_days",
    # Earnings
    "gross": "gross_salary",
    "gross salary": "gross_salary",
    "ctc": "ctc",
    "lop": "lop_amount",
    "eff.gs =gross -lop": "eff_gross",
    "eff.gross =gross -lop": "eff_gross",
    "eff.gs =ctc -lop": "eff_gross",
    "basic": "basic",
    "hra": "hra",
    "convnc allowance": "conveyance",
    "conveyance allowance": "conveyance",
    "spl. allow": "special_allowance",
    "special allowance": "special_allowance",
    "medical allowance": "medical_allowance",
    "incentive": "incentive",
    "incentive (variable)": "incentive",
    "over time": "overtime",
    "total eff.gross": "total_eff_gross",
    # Deductions  (PF/ESI handled separately because of merged cells)
    "pt": "pt",
    "tds": "tds",
    "sal. adv.": "salary_advance",
    "salary advance": "salary_advance",
    "total ded.": "total_deductions",
    "total deduction": "total_deductions",
    # Net
    "arrears": "arrears",
    "attn bonus": "bonus",
    "bonus": "bonus",
    "net sal.": "net_salary",
    "net salary": "net_salary",
    "mode of pay": "mode_of_pay",
}


def _normalise_header(raw: str) -> str:
    """Return canonical key for a raw header string."""
    if raw is None:
        return None
    cleaned = str(raw).strip().lower()
    # Remove extra whitespace
    cleaned = re.sub(r"\s+", " ", cleaned)
    return _CANONICAL_MAP.get(cleaned, cleaned)


def _safe_float(val, default=0.0) -> float:
    """Convert a cell value to float, treating None / '-' / '' as *default*."""
    if val is None or val == "" or val == "-" or val == "–":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_workbook(excel_path: str):
    """Load the workbook once and return the openpyxl Workbook object."""
    return openpyxl.load_workbook(excel_path, data_only=True)


def list_salary_sheets(excel_path: str = None, *, wb=None) -> list[str]:
    """Return names of sheets that look like monthly salary statements."""
    close_after = False
    if wb is None:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        close_after = True
    sheets = [
        name for name in wb.sheetnames
        if re.search(r"salary\s*stmt|salary\s*statement", name, re.IGNORECASE)
    ]
    if close_after:
        wb.close()
    return sheets


def _extract_month_label(ws) -> str:
    """Pull the month label from the title row (row 2)."""
    title = ws.cell(row=2, column=1).value or ""
    # e.g. "SALARY STATEMENT FOR THE MONTH OF JAN-2025"
    m = re.search(r"MONTH\s+OF\s+(.+)", str(title), re.IGNORECASE)
    if m:
        return m.group(1).strip().replace(" ", "").upper()
    # fallback: use sheet name
    return ws.title.replace("Salary Stmt", "").replace("Salary stmt", "").strip(" -")


def _normalize_name(name: str) -> str:
    """Normalize a name for fuzzy matching: lowercase, strip, collapse spaces."""
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name).strip().lower())


def read_pf_esi_data(excel_path: str = None, *, wb=None) -> dict:
    """
    Read the 'PF UAN & ESI IP' sheet and return lookup dicts.

    Returns
    -------
    dict with keys:
        uan_map : dict[normalized_name -> str(UAN number)]
        esi_map : dict[normalized_name -> str(ESI IP number)]
    """
    close_after = False
    if wb is None:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        close_after = True

    uan_map = {}
    esi_map = {}

    sheet_name = None
    for sn in wb.sheetnames:
        if "PF UAN" in sn.upper() or "ESI IP" in sn.upper():
            sheet_name = sn
            break

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section = None  # 'uan' or 'esi'
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            cell0 = str(row[0]).strip().upper() if row[0] else ""
            cell1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            # Detect section headers
            if cell0 == "UAN":
                section = "uan"
                continue
            if cell0 in ("EMPE_NAME", "EMPLOYEE NAME", "EMPE NAME"):
                section = "esi"
                continue
            if not row[0] or not cell1:
                continue

            if section == "uan":
                # row[0] = UAN number, row[1] = name
                uan_map[_normalize_name(cell1)] = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0]).strip()
            elif section == "esi":
                # row[0] = name, row[1] = ESI IP number
                esi_map[_normalize_name(str(row[0]))] = str(int(row[1])) if isinstance(row[1], (int, float)) else str(row[1]).strip()

    if close_after:
        wb.close()

    return {"uan_map": uan_map, "esi_map": esi_map}


def _find_in_map(emp_name: str, lookup: dict) -> str | None:
    """Find an employee name in a lookup dict using fuzzy matching."""
    norm = _normalize_name(emp_name)
    if norm in lookup:
        return lookup[norm]
    # Try partial match: check if any key contains the name or vice versa
    for key, val in lookup.items():
        # Split into parts for matching
        name_parts = set(norm.split())
        key_parts = set(key.split())
        # Match if at least first name + one other part match
        common = name_parts & key_parts
        if len(common) >= 1 and (len(name_parts) <= 1 or len(common) >= min(2, len(name_parts))):
            return val
    return None


def enrich_employees_with_pf_esi(
    employees: list[dict],
    pf_esi_data: dict,
    default_doj: str = "01-Jul-2025",
) -> list[dict]:
    """
    Enrich employee dicts with UAN numbers, ESI IP numbers, and default DOJ.

    Parameters
    ----------
    employees    : list of employee dicts from read_payroll_sheet()
    pf_esi_data  : dict from read_pf_esi_data()
    default_doj  : default date of joining string
    """
    uan_map = pf_esi_data.get("uan_map", {})
    esi_map = pf_esi_data.get("esi_map", {})

    for emp in employees:
        name = emp.get("emp_name", "")

        # UAN number
        uan = _find_in_map(name, uan_map)
        emp["uan_no"] = uan if uan else "N.A."

        # ESI IP number
        esi = _find_in_map(name, esi_map)
        emp["esi_no"] = esi if esi else "N.A."

        # Date of Joining
        emp["doj"] = default_doj

    return employees


def read_payroll_sheet(excel_path: str = None, sheet_name: str = "", *, wb=None) -> dict:
    """
    Read a single salary-statement sheet and return structured data.

    Parameters
    ----------
    excel_path : path to Excel (used only if *wb* is None)
    sheet_name : name of the sheet to read
    wb         : pre-loaded openpyxl Workbook (avoids re-loading)

    Returns
    -------
    dict with keys:
        company   : str
        month     : str  (e.g. "JAN-2025")
        employees : list[dict]   — one dict per employee row
    """
    close_after = False
    if wb is None:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        close_after = True

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    company = str(ws.cell(row=1, column=1).value or "THOTA HOSPITALITY LLP").strip()
    month_label = _extract_month_label(ws)

    # ---- Build column index from header row (row 3) + sub-header row (row 4) ----
    raw_headers = [cell.value for cell in list(ws.iter_rows(min_row=3, max_row=3))[0]]
    sub_headers = [cell.value for cell in list(ws.iter_rows(min_row=4, max_row=4))[0]]

    col_map: dict[str, int] = {}  # canonical_name -> 0-based column index

    # Handle merged "EMPLOYEE SHARE" / "EMPLOYER SHARE" with PF & ESI sub-columns
    i = 0
    while i < len(raw_headers):
        h = raw_headers[i]
        sh = sub_headers[i] if i < len(sub_headers) else None

        if h is not None:
            norm = _normalise_header(h)

            if norm == "employee share" or (h and "EMPLOYEE" in str(h).upper() and "SHARE" in str(h).upper()):
                # Next two columns are PF, ESI under employee share
                if sh and str(sh).strip().upper() == "PF":
                    col_map["employee_pf"] = i
                elif i + 1 < len(sub_headers):
                    col_map["employee_pf"] = i
                # Check next col
                if i + 1 < len(raw_headers):
                    sh_next = sub_headers[i + 1] if i + 1 < len(sub_headers) else None
                    if sh_next and str(sh_next).strip().upper() == "ESI":
                        col_map["employee_esi"] = i + 1
                    elif raw_headers[i + 1] is None:
                        col_map["employee_esi"] = i + 1
                i += 2
                continue

            elif norm == "employer share" or (h and "EMPLOYER" in str(h).upper() and "SHARE" in str(h).upper()):
                if sh and str(sh).strip().upper() == "PF":
                    col_map["employer_pf"] = i
                elif i + 1 < len(sub_headers):
                    col_map["employer_pf"] = i
                if i + 1 < len(raw_headers):
                    sh_next = sub_headers[i + 1] if i + 1 < len(sub_headers) else None
                    if sh_next and str(sh_next).strip().upper() == "ESI":
                        col_map["employer_esi"] = i + 1
                    elif raw_headers[i + 1] is None:
                        col_map["employer_esi"] = i + 1
                i += 2
                continue

            else:
                if norm:
                    col_map[norm] = i
        else:
            # None header — might be a sub-column of merged cell; check sub_header
            if sh:
                sh_upper = str(sh).strip().upper()
                if sh_upper == "PF":
                    # Determine if employee or employer based on prior context
                    if "employee_pf" not in col_map:
                        col_map["employee_pf"] = i
                    elif "employer_pf" not in col_map:
                        col_map["employer_pf"] = i
                elif sh_upper == "ESI":
                    if "employee_esi" not in col_map:
                        col_map["employee_esi"] = i
                    elif "employer_esi" not in col_map:
                        col_map["employer_esi"] = i
        i += 1

    # ---- Read employee rows starting from row 5 ----
    employees = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        sl = row[col_map.get("sl_no", 0)] if "sl_no" in col_map else row[0]
        if sl is None:
            continue  # skip empty rows
        try:
            float(sl)
        except (ValueError, TypeError):
            continue  # skip non-numeric (totals row, etc.)

        def _get(key, as_float=False):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return 0.0 if as_float else None
            val = row[idx]
            if as_float:
                return _safe_float(val)
            return val

        emp = {
            "sl_no": int(_safe_float(sl)),
            "position": str(_get("position") or "N/A").strip(),
            "emp_name": str(_get("emp_name") or "Unknown").strip(),
            "designation": str(_get("designation") or "N/A").strip(),
            "department": str(_get("department") or "N/A").strip(),
            "uan_no": str(_get("uan_no") or "N/A").strip(),
            "pf_no": str(_get("pf_no") or "N/A").strip(),
            "esi_no": str(_get("esi_no") or "N/A").strip(),
            "branch": str(_get("branch") or "N/A").strip(),
            "bank_account_no": str(_get("bank_account_no") or "N/A").strip(),
            "doj": _get("doj"),
            # Attendance
            "calendar_days": _get("calendar_days", as_float=True),
            "present_days": _get("present_days", as_float=True),
            "holidays": _get("holidays", as_float=True),
            "leaves": _get("leaves", as_float=True),
            "lop_absent": _get("lop_absent", as_float=True),
            "paid_days": _get("paid_days", as_float=True),
            "lop_days": _get("lop_days", as_float=True),
            # Earnings
            "gross_salary": _get("gross_salary", as_float=True),
            "lop_amount": _get("lop_amount", as_float=True),
            "eff_gross": _get("eff_gross", as_float=True),
            "basic": _get("basic", as_float=True),
            "hra": _get("hra", as_float=True),
            "conveyance": _get("conveyance", as_float=True),
            "special_allowance": _get("special_allowance", as_float=True),
            "medical_allowance": _get("medical_allowance", as_float=True),
            "incentive": _get("incentive", as_float=True),
            "overtime": _get("overtime", as_float=True),
            "total_eff_gross": _get("total_eff_gross", as_float=True),
            # Deductions
            "employee_pf": _get("employee_pf", as_float=True),
            "employee_esi": _get("employee_esi", as_float=True),
            "employer_pf": _get("employer_pf", as_float=True),
            "employer_esi": _get("employer_esi", as_float=True),
            "pt": _get("pt", as_float=True),
            "tds": _get("tds", as_float=True),
            "salary_advance": _get("salary_advance", as_float=True),
            "total_deductions": _get("total_deductions", as_float=True),
            # Net
            "arrears": _get("arrears", as_float=True),
            "bonus": _get("bonus", as_float=True),
            "net_salary": _get("net_salary", as_float=True),
            "mode_of_pay": str(_get("mode_of_pay") or "NEFT").strip(),
            "ctc": _get("ctc", as_float=True),
        }
        employees.append(emp)

    if close_after:
        wb.close()
    return {
        "company": company,
        "month": month_label,
        "employees": employees,
    }


# ---------------------------------------------------------------------------
# CLI quick-test
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python payroll_reader.py <excel_path> [sheet_name]")
        sys.exit(1)

    excel = sys.argv[1]
    sheets = list_salary_sheets(excel)
    print(f"Salary sheets found: {sheets}\n")

    sheet = sys.argv[2] if len(sys.argv) > 2 else sheets[0]
    data = read_payroll_sheet(excel, sheet)
    print(f"Company : {data['company']}")
    print(f"Month   : {data['month']}")
    print(f"Employees: {len(data['employees'])}")
    for emp in data["employees"][:3]:
        print(f"  #{emp['sl_no']} {emp['emp_name']:20s}  Net: ₹{emp['net_salary']:,.0f}")
