#!/usr/bin/env python3
"""
MIS Master Sheet Generator — Formula-Linked to Trial Balance

Creates a professional Excel workbook where:
- Sheet "TB" is the ONLY input (paste monthly cumulative Trial Balance values)
- All report sheets (P&L, Balance Sheet, Cash Flow, KPIs, Dashboard)
  use Excel FORMULAS referencing TB — zero hardcoded numbers
- Monthly, Quarterly, and Annual views auto-calculate
- Ready for investor/director presentations

Usage:
    python execution/mis_master_generator.py --source "path/to/TrialBal.xlsx"
    python execution/mis_master_generator.py --source "path/to/TrialBal.xlsx" --output "MIS_Master.xlsx"
"""

import os
import sys
import argparse
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(exist_ok=True)

# ── Style Constants ──────────────────────────────────────────────────────────
NAVY = "2F5496"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
RED = "C00000"
DARK_GREEN = "548235"

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color=NAVY)
SECTION_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
INPUT_FILL = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="4472C4")
PCT_FONT = Font(name="Calibri", italic=True, size=10, color="4472C4")
NORMAL_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
TOTAL_BORDER = Border(
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='double', color=NAVY),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)
INR = '#,##0'
INR2 = '#,##0.00'
PCT = '0.0%'
NUM2 = '0.00'

# Month columns on TB sheet: D=Apr ... O=Mar (col index 4..15)
# Month columns on report sheets: B=Apr ... M=Mar (col index 2..13)
# Quarterly on report: N=Q1, O=Q2, P=Q3, Q=Q4, R=FY (col index 14..18)
MONTH_NAMES = ['APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN', 'FEB (P)', 'MAR (P)']
TB_MONTH_START_COL = 4   # Column D on TB sheet
RPT_MONTH_START_COL = 2  # Column B on report sheets


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  ACCOUNT DEFINITIONS — Maps TB accounts to report line items           ║
# ╚══════════════════════════════════════════════════════════════════════════╝

# Each account: (id, group, display_name, nature, tb_value_jan)
# nature: "Dr" = debit balance positive, "Cr" = credit balance positive
# These define the TB input rows. Section headers are separate.

# --- TB Layout ---
# Row 5: Column headers
# Row 6+: sections and accounts

TB_SECTIONS = [
    # (row, title)
    (6, "EQUITY & CAPITAL"),
    (9, "LOANS (LIABILITY)"),
    (13, "CURRENT LIABILITIES"),
    (18, "FIXED ASSETS"),
    (22, "CURRENT ASSETS"),
    (27, "REVENUE (Cumulative — from P&L)"),
    (30, "COST OF GOODS SOLD (Cumulative)"),
    (34, "DIRECT EXPENSES (Cumulative)"),
    (37, "INDIRECT INCOME (Cumulative)"),
    (39, "INDIRECT EXPENSES / OPEX (Cumulative)"),
    (50, "DEPRECIATION"),
]

# Account definitions: (tb_row, id, group, name, nature, jan_cumulative_value)
TB_ACCOUNTS = [
    (7,  'cap_partner',    "Capital",           "Partner's Capital",              "Cr", 116001),
    (8,  'cap_pnl_prior',  "Capital",           "P&L A/c (Prior Year Loss)",      "Dr", 321920.38),
    (10, 'loan_akshatha',  "Loans",             "Partner's Loan — Akshatha",      "Cr", 1323443.49),
    (11, 'loan_tejas',     "Loans",             "Partner's Loan — Tejas",         "Cr", 217000),
    (12, 'loan_tvs',       "Loans",             "TVS Credit Services Ltd",        "Cr", 10600.40),
    (14, 'cl_duties',      "Current Liab.",     "Duties & Taxes (Net)",           "Cr", 142197.13),
    (15, 'cl_creditors',   "Current Liab.",     "Sundry Creditors (Net)",         "Cr", 24304.61),
    (16, 'cl_reimburse',   "Current Liab.",     "Reimbursements (Net)",           "Cr", -17931.67),
    (17, 'cl_salary',      "Current Liab.",     "Salary Payable (Net)",           "Cr", 282369.37),
    (19, 'fa_intangible',  "Fixed Assets",      "Intangible Assets",              "Dr", 69015.84),
    (20, 'fa_tangible',    "Fixed Assets",      "Tangible Assets",                "Dr", 1779109.98),
    (23, 'ca_advances',    "Current Assets",    "Loans & Advances",               "Dr", 100167),
    (24, 'ca_debtors',     "Current Assets",    "Sundry Debtors (Net)",           "Dr", 251631.72),
    (25, 'ca_cash',        "Current Assets",    "Cash-in-Hand",                   "Dr", 181394.89),
    (26, 'ca_bank',        "Current Assets",    "Bank Accounts",                  "Dr", 609924.76),
    (28, 'rev_hospitality',"Revenue",           "Hospitality Services @ 5%",      "Cr", 8472614.40),
    (29, 'rev_studio',     "Revenue",           "Studio Rental Services @ 18%",   "Cr", 165863.50),
    (31, 'cogs_kitchen',   "COGS",              "Thota Kitchen",                  "Dr", 1439306.54),
    (32, 'cogs_decor',     "COGS",              "Thota Decor",                    "Dr", 730979.78),
    (33, 'cogs_studio',    "COGS",              "Studio (Purchase)",              "Dr", 87739.79),
    (35, 'de_fuel',        "Direct Expenses",   "Fuel Expenses",                  "Dr", 249452),
    (36, 'de_transport',   "Direct Expenses",   "Transportation Expenses",        "Dr", 45233),
    (38, 'oi_discount',    "Other Income",      "Discount Received",              "Cr", 2699.40),
    (40, 'opex_admin',     "OPEX",              "Administrative Overheads",       "Dr", 543182.59),
    (41, 'opex_finance',   "OPEX",              "Finance Cost",                   "Dr", 7659.32),
    (42, 'opex_hr',        "OPEX",              "HR Expenses (Salary+Incentive)", "Dr", 3360542),
    (43, 'opex_marketing', "OPEX",              "Marketing & Ads",                "Dr", 62886.98),
    (44, 'opex_professional',"OPEX",            "Professional Charges",           "Dr", 271640),
    (45, 'opex_audit',     "OPEX",              "Audit / Legal Fee",              "Dr", 23600),
    (46, 'opex_pt',        "OPEX",              "Employer PT",                    "Dr", 2500),
    (47, 'opex_rates',     "OPEX",              "Rates & Taxes (Credit adj.)",    "Cr", 467.52),
    (48, 'opex_rm',        "OPEX",              "Repairs & Maintenance",          "Dr", 601754.83),
    (49, 'opex_roundoff',  "OPEX",              "Round Off (Credit adj.)",        "Cr", 12.25),
    (51, 'dep_charge',     "Depreciation",      "Depreciation Charge (Cumul.)",   "Dr", 0),
]

# Build lookup: account_id → tb_row
ACCT_ROW = {a[1]: a[0] for a in TB_ACCOUNTS}


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  HELPER FUNCTIONS                                                       ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def safe_float(val, default=0.0):
    if val is None: return default
    if isinstance(val, str):
        val = val.strip().replace(',','').replace('₹','').replace(' ','')
        if val in ('', '-', '#REF!', '\xa0'): return default
        try: return float(val)
        except ValueError: return default
    try: return float(val)
    except (ValueError, TypeError): return default


def tb_col(month_idx):
    """Get TB column letter for month index (0=Apr, 11=Mar)."""
    return get_column_letter(TB_MONTH_START_COL + month_idx)


def rpt_col(month_idx):
    """Get report column letter for month index (0=Apr, 11=Mar)."""
    return get_column_letter(RPT_MONTH_START_COL + month_idx)


def monthly_pnl_formula(tb_row, rpt_col_idx):
    """
    Generate Excel formula to extract MONTHLY P&L value from CUMULATIVE TB.
    rpt_col_idx: 2=B(Apr), 3=C(May), ..., 13=M(Mar)
    TB columns: 4=D(Apr), 5=E(May), ..., 15=O(Mar)
    
    April (first month): =TB!D{row}
    Other months: =TB!E{row}-TB!D{row}  (cumulative delta)
    """
    tb_col_letter = get_column_letter(rpt_col_idx + 2)
    if rpt_col_idx == 2:  # April = first month
        return f"=TB!{tb_col_letter}{tb_row}"
    prev_col_letter = get_column_letter(rpt_col_idx + 1)
    return f"=TB!{tb_col_letter}{tb_row}-TB!{prev_col_letter}{tb_row}"


def monthly_pnl_formula_negate(tb_row, rpt_col_idx):
    """Same as monthly_pnl_formula but NEGATED (for credit-nature expense adj)."""
    tb_col_letter = get_column_letter(rpt_col_idx + 2)
    if rpt_col_idx == 2:
        return f"=-TB!{tb_col_letter}{tb_row}"
    prev_col_letter = get_column_letter(rpt_col_idx + 1)
    return f"=-(TB!{tb_col_letter}{tb_row}-TB!{prev_col_letter}{tb_row})"


def bs_formula(tb_row, rpt_col_idx):
    """BS formula — direct closing balance from TB for that month."""
    tb_col_letter = get_column_letter(rpt_col_idx + 2)
    return f"=TB!{tb_col_letter}{tb_row}"


def quarterly_formula(row, q_num):
    """Q1=SUM(B:D), Q2=SUM(E:G), Q3=SUM(H:J), Q4=SUM(K:M)"""
    start_cols = {1: 'B', 2: 'E', 3: 'H', 4: 'K'}
    end_cols   = {1: 'D', 2: 'G', 3: 'J', 4: 'M'}
    return f"=SUM({start_cols[q_num]}{row}:{end_cols[q_num]}{row})"


def fy_formula(row):
    """FY Total = SUM(B:M)"""
    return f"=SUM(B{row}:M{row})"


def apply_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def apply_section(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER


def apply_total(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = TOTAL_BORDER


def apply_data_fmt(ws, row, start_col, end_col, fmt=INR):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.number_format = fmt
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')


def write_title(ws, row, col, title, subtitle=None):
    ws.cell(row=row, column=col, value=title).font = TITLE_FONT
    if subtitle:
        ws.cell(row=row+1, column=col, value=subtitle).font = SUBTITLE_FONT
        return row + 3
    return row + 2


def set_col_widths(ws, widths_dict):
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  READ SOURCE TRIAL BALANCE                                             ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def read_trial_balance(source_path):
    """Read TB file and extract account values. Returns dict of account_id → value."""
    wb = openpyxl.load_workbook(source_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Parse TB into account name → (debit, credit) mapping
    # Use FIRST occurrence for duplicates (Tally shows group subtotals first)
    tb_data = {}
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        debit = safe_float(ws.cell(row=row, column=2).value)
        credit = safe_float(ws.cell(row=row, column=3).value)
        if name and name.strip():
            key = name.strip()
            if key not in tb_data:  # First occurrence wins (group subtotal)
                tb_data[key] = (debit, credit)

    # Get the period from the TB
    period = ""
    for row in range(1, 10):
        val = ws.cell(row=row, column=1).value
        if val and "to" in str(val).lower():
            period = str(val)
            break

    wb.close()
    return tb_data, period


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 1: TB (Trial Balance Input)                                     ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_tb_sheet(wb, jan_values):
    """Create the TB input sheet with pre-filled Jan cumulative data."""
    ws = wb.create_sheet("TB", 0)

    # Title
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Trial Balance — Monthly Cumulative Input Sheet").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value=(
        "Instructions: Paste CUMULATIVE Trial Balance closing values for each month. "
        "Export from Tally: TB for 'Apr 1 to Apr 30', 'Apr 1 to May 31', etc. "
        "All report sheets auto-calculate via formulas."
    )).font = Font(name="Calibri", size=10, color="808080", italic=True)

    # Headers (Row 5)
    headers = ['Group', 'Account Name', 'Dr/Cr'] + MONTH_NAMES
    for ci, h in enumerate(headers, 1):
        ws.cell(row=5, column=ci, value=h)
    apply_header(ws, 5, len(headers))

    # Section headers
    for sec_row, sec_title in TB_SECTIONS:
        ws.cell(row=sec_row, column=1, value=sec_title)
        apply_section(ws, sec_row, len(headers))

    # Account rows
    for tb_row, acct_id, group, name, nature, jan_val in TB_ACCOUNTS:
        ws.cell(row=tb_row, column=1, value=group).font = NORMAL_FONT
        ws.cell(row=tb_row, column=2, value=name).font = NORMAL_FONT
        ws.cell(row=tb_row, column=3, value=nature).font = Font(name="Calibri", size=10, color="808080")

        # Yellow input cells for months
        for mi in range(12):
            col_idx = TB_MONTH_START_COL + mi
            cell = ws.cell(row=tb_row, column=col_idx)
            cell.fill = INPUT_FILL
            cell.number_format = INR2
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

        # Pre-fill January column (M = column 13 = month index 9)
        jan_col = TB_MONTH_START_COL + 9  # Column M (index 13)
        # Use provided jan_value or look up from TB data
        override = jan_values.get(acct_id)
        val = override if override is not None else jan_val
        ws.cell(row=tb_row, column=jan_col, value=val)

    # Verification row
    verify_row = 53
    ws.cell(row=verify_row, column=1, value="VERIFICATION").font = SECTION_FONT
    ws.cell(row=verify_row + 1, column=2, value="Total Debits").font = TOTAL_FONT
    ws.cell(row=verify_row + 2, column=2, value="Total Credits").font = TOTAL_FONT
    ws.cell(row=verify_row + 3, column=2, value="Difference (should be 0)").font = Font(name="Calibri", bold=True, color=RED)

    # Debit accounts: rows with Dr nature
    dr_rows = [a[0] for a in TB_ACCOUNTS if a[4] == "Dr"]
    cr_rows = [a[0] for a in TB_ACCOUNTS if a[4] == "Cr"]

    for mi in range(12):
        col_idx = TB_MONTH_START_COL + mi
        col_letter = get_column_letter(col_idx)
        # Total Debits
        dr_formula = "=" + "+".join([f"{col_letter}{r}" for r in dr_rows])
        ws.cell(row=verify_row + 1, column=col_idx, value=dr_formula).number_format = INR2
        # Total Credits
        cr_formula = "=" + "+".join([f"{col_letter}{r}" for r in cr_rows])
        ws.cell(row=verify_row + 2, column=col_idx, value=cr_formula).number_format = INR2
        # Difference
        ws.cell(row=verify_row + 3, column=col_idx,
                value=f"={col_letter}{verify_row+1}-{col_letter}{verify_row+2}").number_format = INR2

    # Column widths
    set_col_widths(ws, {'A': 16, 'B': 35, 'C': 6})
    for mi in range(12):
        ws.column_dimensions[get_column_letter(TB_MONTH_START_COL + mi)].width = 15

    # Freeze panes
    ws.freeze_panes = 'D6'

    return ws


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 2: P&L (Profit & Loss Statement)                                ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_pnl_sheet(wb):
    """Create P&L with all formulas referencing TB sheet."""
    ws = wb.create_sheet("PnL")

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP",
                      "Profit & Loss Statement — FY 2025-26 (Monthly | Quarterly | Annual)")
    MAX_COL = 18  # A through R

    # Headers (Row 4)
    r = 4
    headers = ['Particulars'] + MONTH_NAMES + ['Q1', 'Q2', 'Q3', 'Q4', 'FY TOTAL']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h)
    apply_header(ws, r, len(headers))

    # ── Helper: write a PnL data row with monthly formulas from TB ──
    def write_pnl_line(row, label, tb_rows, negate=False, indent=False):
        """Write a line pulling monthly deltas from one or more TB rows."""
        cell = ws.cell(row=row, column=1, value=("  " + label if indent else label))
        cell.font = NORMAL_FONT
        for mi in range(12):
            col_idx = RPT_MONTH_START_COL + mi
            if len(tb_rows) == 1:
                if negate:
                    f = monthly_pnl_formula_negate(tb_rows[0], col_idx)
                else:
                    f = monthly_pnl_formula(tb_rows[0], col_idx)
            else:
                # Sum multiple TB rows
                parts = []
                for tbr in tb_rows:
                    if negate:
                        parts.append(f"(-({_monthly_raw(tbr, col_idx)}))")
                    else:
                        parts.append(f"({_monthly_raw(tbr, col_idx)})")
                f = "=" + "+".join(parts)
            ws.cell(row=row, column=col_idx, value=f)
        # Quarterly & FY
        for q in range(1, 5):
            ws.cell(row=row, column=13 + q, value=quarterly_formula(row, q))
        ws.cell(row=row, column=18, value=fy_formula(row))
        apply_data_fmt(ws, row, 2, 18, INR)

    def _monthly_raw(tb_row, rpt_col_idx):
        """Raw expression (no leading =) for monthly delta from cumulative."""
        tc = get_column_letter(rpt_col_idx + 2)
        if rpt_col_idx == 2:
            return f"TB!{tc}{tb_row}"
        pc = get_column_letter(rpt_col_idx + 1)
        return f"TB!{tc}{tb_row}-TB!{pc}{tb_row}"

    def write_sum_row(row, label, sum_rows, is_total=False, subtract_rows=None):
        """Write a row that sums/subtracts other PnL rows."""
        ws.cell(row=row, column=1, value=label)
        for ci in range(2, 19):
            col_l = get_column_letter(ci)
            parts = [f"{col_l}{r}" for r in sum_rows]
            formula = "=" + "+".join(parts)
            if subtract_rows:
                for sr in subtract_rows:
                    formula += f"-{col_l}{sr}"
            ws.cell(row=row, column=ci, value=formula)
        apply_data_fmt(ws, row, 2, 18, INR)
        if is_total:
            apply_total(ws, row, 18)

    def write_pct_row(row, label, numerator_row, denominator_row):
        """Write a percentage row = numerator / denominator."""
        ws.cell(row=row, column=1, value=label).font = PCT_FONT
        for ci in range(2, 19):
            col_l = get_column_letter(ci)
            f = f'=IFERROR({col_l}{numerator_row}/{col_l}{denominator_row},0)'
            cell = ws.cell(row=row, column=ci, value=f)
            cell.number_format = PCT
            cell.font = PCT_FONT

    # ── P&L Line Items ──────────────────────────────────────────────────

    # Row 5: REVENUE section
    r = 5
    ws.cell(row=r, column=1, value="REVENUE FROM OPERATIONS")
    apply_section(ws, r, 18)

    # Row 6: Hospitality Services
    write_pnl_line(6, "Hospitality Services", [ACCT_ROW['rev_hospitality']], indent=True)
    # Row 7: Studio Rental
    write_pnl_line(7, "Studio Rental Services", [ACCT_ROW['rev_studio']], indent=True)
    # Row 8: Total Revenue
    write_sum_row(8, "TOTAL REVENUE", [6, 7], is_total=True)

    # Row 10: COGS section
    ws.cell(row=10, column=1, value="COST OF GOODS SOLD")
    apply_section(ws, 10, 18)
    write_pnl_line(11, "Thota Kitchen", [ACCT_ROW['cogs_kitchen']], indent=True)
    write_pnl_line(12, "Thota Decor", [ACCT_ROW['cogs_decor']], indent=True)
    write_pnl_line(13, "Studio (Purchase)", [ACCT_ROW['cogs_studio']], indent=True)
    write_sum_row(14, "TOTAL COGS", [11, 12, 13], is_total=True)

    # Row 16: Direct Expenses
    ws.cell(row=16, column=1, value="DIRECT EXPENSES")
    apply_section(ws, 16, 18)
    write_pnl_line(17, "Fuel Expenses", [ACCT_ROW['de_fuel']], indent=True)
    write_pnl_line(18, "Transportation", [ACCT_ROW['de_transport']], indent=True)
    write_sum_row(19, "TOTAL DIRECT EXPENSES", [17, 18], is_total=True)

    # Row 21: GROSS PROFIT = Revenue - COGS - Direct
    write_sum_row(21, "GROSS PROFIT", [8], subtract_rows=[14, 19], is_total=True)
    write_pct_row(22, "Gross Profit %", 21, 8)

    # Row 24: OTHER INCOME
    ws.cell(row=24, column=1, value="OTHER INCOME")
    apply_section(ws, 24, 18)
    write_pnl_line(25, "Discount Received", [ACCT_ROW['oi_discount']], indent=True)

    # Row 27: OPERATING EXPENSES
    ws.cell(row=27, column=1, value="OPERATING EXPENSES (OPEX)")
    apply_section(ws, 27, 18)
    write_pnl_line(28, "Administrative Overheads", [ACCT_ROW['opex_admin']], indent=True)
    write_pnl_line(29, "HR Expenses", [ACCT_ROW['opex_hr']], indent=True)
    write_pnl_line(30, "Marketing & Ads", [ACCT_ROW['opex_marketing']], indent=True)
    write_pnl_line(31, "Professional & Legal", [ACCT_ROW['opex_professional'], ACCT_ROW['opex_audit']], indent=True)
    write_pnl_line(32, "Employer PT", [ACCT_ROW['opex_pt']], indent=True)
    write_pnl_line(33, "Repairs & Maintenance", [ACCT_ROW['opex_rm']], indent=True)
    # Rates & Taxes (credit = reduces expense, so NEGATE)
    write_pnl_line(34, "Rates & Taxes (Adj.)", [ACCT_ROW['opex_rates']], negate=True, indent=True)
    # Round Off (credit = reduces expense)
    write_pnl_line(35, "Round Off (Adj.)", [ACCT_ROW['opex_roundoff']], negate=True, indent=True)
    write_sum_row(36, "TOTAL OPEX", [28, 29, 30, 31, 32, 33, 34, 35], is_total=True)

    # Row 38: EBITDA = Gross Profit + Other Income - OPEX
    ws.cell(row=38, column=1, value="EBITDA")
    for ci in range(2, 19):
        col_l = get_column_letter(ci)
        ws.cell(row=38, column=ci, value=f"={col_l}21+{col_l}25-{col_l}36")
    apply_data_fmt(ws, 38, 2, 18, INR)
    apply_total(ws, 38, 18)
    write_pct_row(39, "EBITDA %", 38, 8)

    # Row 41: Below EBITDA items
    write_pnl_line(41, "Finance Cost", [ACCT_ROW['opex_finance']])
    write_pnl_line(42, "Depreciation", [ACCT_ROW['dep_charge']])

    # Row 44: PBT = EBITDA - Finance - Depreciation
    ws.cell(row=44, column=1, value="PROFIT BEFORE TAX (PBT)")
    for ci in range(2, 19):
        col_l = get_column_letter(ci)
        ws.cell(row=44, column=ci, value=f"={col_l}38-{col_l}41-{col_l}42")
    apply_data_fmt(ws, 44, 2, 18, INR)
    apply_total(ws, 44, 18)

    # Row 45: Income Tax (input / formula)
    ws.cell(row=45, column=1, value="  Income Tax").font = NORMAL_FONT
    for ci in range(2, 14):
        cell = ws.cell(row=45, column=ci, value=0)
        cell.fill = INPUT_FILL
        cell.number_format = INR
        cell.border = THIN_BORDER
    for q in range(1, 5):
        ws.cell(row=45, column=13 + q, value=quarterly_formula(45, q)).number_format = INR
    ws.cell(row=45, column=18, value=fy_formula(45)).number_format = INR

    # Row 47: PAT = PBT - Tax
    write_sum_row(47, "PROFIT AFTER TAX (PAT)", [44], subtract_rows=[45], is_total=True)
    write_pct_row(48, "PAT %", 47, 8)

    # Conditional formatting: negative values in red
    red_font_rule = CellIsRule(operator='lessThan', formula=['0'],
                               font=Font(color=RED))
    ws.conditional_formatting.add(f'B6:R48', red_font_rule)

    # Column widths
    set_col_widths(ws, {'A': 32})
    for ci in range(2, 19):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = 'B5'
    ws.sheet_properties.tabColor = "4472C4"
    return ws


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 3: Balance Sheet                                                ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_bs_sheet(wb):
    """Create Balance Sheet with formulas referencing TB."""
    ws = wb.create_sheet("Balance Sheet")
    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP",
                      "Balance Sheet — Monthly Snapshots FY 2025-26")
    MAX_COL = 14  # A through N (12 months + Total col)

    # Headers
    r = 4
    headers = ['Particulars'] + MONTH_NAMES + ['Latest']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h)
    apply_header(ws, r, len(headers))

    def write_bs_line(row, label, tb_rows, sign=1, indent=False):
        """Write BS line referencing TB closing balances."""
        ws.cell(row=row, column=1, value=("  " + label if indent else label)).font = NORMAL_FONT
        for mi in range(12):
            col_idx = RPT_MONTH_START_COL + mi
            tb_col_l = get_column_letter(col_idx + 2)
            if len(tb_rows) == 1:
                if sign == -1:
                    f = f"=-TB!{tb_col_l}{tb_rows[0]}"
                else:
                    f = f"=TB!{tb_col_l}{tb_rows[0]}"
            else:
                parts = []
                for tbr in tb_rows:
                    if sign == -1:
                        parts.append(f"(-TB!{tb_col_l}{tbr})")
                    else:
                        parts.append(f"TB!{tb_col_l}{tbr}")
                f = "=" + "+".join(parts)
            ws.cell(row=row, column=col_idx, value=f)
        # Latest = last non-zero month (use M / column 13 for Jan)
        # Simple: reference column M (Jan)
        ws.cell(row=row, column=14, value=f"=M{row}")
        apply_data_fmt(ws, row, 2, 14, INR)

    def write_bs_sum(row, label, sum_rows, subtract_rows=None, is_total=False):
        ws.cell(row=row, column=1, value=label)
        for ci in range(2, 15):
            col_l = get_column_letter(ci)
            parts = [f"{col_l}{r}" for r in sum_rows]
            formula = "=" + "+".join(parts)
            if subtract_rows:
                for sr in subtract_rows:
                    formula += f"-{col_l}{sr}"
            ws.cell(row=row, column=ci, value=formula)
        apply_data_fmt(ws, row, 2, 14, INR)
        if is_total:
            apply_total(ws, row, 14)

    # ── LIABILITIES & EQUITY ──
    r = 5
    ws.cell(row=r, column=1, value="EQUITY & LIABILITIES")
    apply_section(ws, r, 14)

    write_bs_line(6, "Partner's Capital", [ACCT_ROW['cap_partner']], indent=True)
    write_bs_line(7, "Less: Prior Year Loss", [ACCT_ROW['cap_pnl_prior']], sign=-1, indent=True)

    # Current Year P&L — reference PnL cumulative PAT
    ws.cell(row=8, column=1, value="  Current Year Profit/(Loss)").font = NORMAL_FONT
    for mi in range(12):
        col_idx = RPT_MONTH_START_COL + mi
        # Cumulative PAT up to this month = sum of monthly PATs from PnL B47 to this col
        start_col_l = get_column_letter(RPT_MONTH_START_COL)
        end_col_l = get_column_letter(col_idx)
        ws.cell(row=8, column=col_idx,
                value=f"=SUM(PnL!{start_col_l}47:{end_col_l}47)")
    ws.cell(row=8, column=14, value=f"=M8")
    apply_data_fmt(ws, 8, 2, 14, INR)

    write_bs_sum(9, "Total Equity", [6, 7, 8], is_total=True)

    # Loans
    r = 11
    ws.cell(row=r, column=1, value="NON-CURRENT LIABILITIES")
    apply_section(ws, r, 14)
    write_bs_line(12, "Partner's Loan — Akshatha", [ACCT_ROW['loan_akshatha']], indent=True)
    write_bs_line(13, "Partner's Loan — Tejas", [ACCT_ROW['loan_tejas']], indent=True)
    write_bs_line(14, "TVS Credit Services Ltd", [ACCT_ROW['loan_tvs']], indent=True)
    write_bs_sum(15, "Total Non-Current Liabilities", [12, 13, 14], is_total=True)

    # Current Liabilities
    r = 17
    ws.cell(row=r, column=1, value="CURRENT LIABILITIES")
    apply_section(ws, r, 14)
    write_bs_line(18, "Duties & Taxes", [ACCT_ROW['cl_duties']], indent=True)
    write_bs_line(19, "Sundry Creditors", [ACCT_ROW['cl_creditors']], indent=True)
    write_bs_line(20, "Reimbursements", [ACCT_ROW['cl_reimburse']], indent=True)
    write_bs_line(21, "Salary Payable", [ACCT_ROW['cl_salary']], indent=True)
    write_bs_sum(22, "Total Current Liabilities", [18, 19, 20, 21], is_total=True)

    # TOTAL LIABILITIES
    write_bs_sum(24, "TOTAL LIABILITIES & EQUITY", [9, 15, 22], is_total=True)

    # ── ASSETS ──
    r = 26
    ws.cell(row=r, column=1, value="ASSETS")
    apply_section(ws, r, 14)

    ws.cell(row=27, column=1, value="NON-CURRENT ASSETS")
    apply_section(ws, 27, 14)
    write_bs_line(28, "Intangible Assets", [ACCT_ROW['fa_intangible']], indent=True)
    write_bs_line(29, "Tangible Assets", [ACCT_ROW['fa_tangible']], indent=True)

    # Depreciation — cumulative, reduces asset
    ws.cell(row=30, column=1, value="  Less: Depreciation").font = NORMAL_FONT
    for mi in range(12):
        col_idx = RPT_MONTH_START_COL + mi
        tb_col_l = get_column_letter(col_idx + 2)
        ws.cell(row=30, column=col_idx, value=f"=-TB!{tb_col_l}{ACCT_ROW['dep_charge']}")
    ws.cell(row=30, column=14, value=f"=M30")
    apply_data_fmt(ws, 30, 2, 14, INR)

    write_bs_sum(31, "Net Fixed Assets", [28, 29, 30], is_total=True)

    # Current Assets
    ws.cell(row=33, column=1, value="CURRENT ASSETS")
    apply_section(ws, 33, 14)
    write_bs_line(34, "Loans & Advances", [ACCT_ROW['ca_advances']], indent=True)
    write_bs_line(35, "Sundry Debtors", [ACCT_ROW['ca_debtors']], indent=True)
    write_bs_line(36, "Cash-in-Hand", [ACCT_ROW['ca_cash']], indent=True)
    write_bs_line(37, "Bank Accounts", [ACCT_ROW['ca_bank']], indent=True)
    write_bs_sum(38, "Total Current Assets", [34, 35, 36, 37], is_total=True)

    # TOTAL ASSETS
    write_bs_sum(40, "TOTAL ASSETS", [31, 38], is_total=True)

    # Balance check
    ws.cell(row=42, column=1, value="Balance Check (Assets - Liabilities)").font = Font(name="Calibri", bold=True, color=RED)
    for ci in range(2, 15):
        col_l = get_column_letter(ci)
        ws.cell(row=42, column=ci, value=f"={col_l}40-{col_l}24").number_format = INR

    set_col_widths(ws, {'A': 35})
    for ci in range(2, 15):
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws.freeze_panes = 'B5'
    ws.sheet_properties.tabColor = "548235"
    return ws


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 4: Cash Flow Statement (Indirect Method)                        ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_cashflow_sheet(wb):
    """Cash flow — derived from PnL and BS via formulas."""
    ws = wb.create_sheet("Cash Flow")
    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP",
                      "Cash Flow Statement (Indirect Method) — FY 2025-26")

    MAX_COL = 18
    r = 4
    headers = ['Particulars'] + MONTH_NAMES + ['Q1', 'Q2', 'Q3', 'Q4', 'FY TOTAL']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h)
    apply_header(ws, r, len(headers))

    def write_cf_formula(row, label, formula_template, is_total=False, indent=False):
        """Write a CF row with a formula that uses {c} as column placeholder."""
        ws.cell(row=row, column=1, value=("  " + label if indent else label)).font = NORMAL_FONT
        for ci in range(2, 14):
            col_l = get_column_letter(ci)
            f = formula_template.replace('{c}', col_l)
            ws.cell(row=row, column=ci, value=f)
        for q in range(1, 5):
            ws.cell(row=row, column=13 + q, value=quarterly_formula(row, q))
        ws.cell(row=row, column=18, value=fy_formula(row))
        apply_data_fmt(ws, row, 2, 18, INR)
        if is_total:
            apply_total(ws, row, 18)

    def write_wc_change(row, label, bs_row, sign=1):
        """Working capital change: (current month BS - previous month BS).
        For assets: increase = cash outflow = negative.
        For liabilities: increase = cash inflow = positive.
        sign=1 for liabilities, sign=-1 for assets.
        """
        ws.cell(row=row, column=1, value="  " + label).font = NORMAL_FONT
        # April: change from start = assume 0 start, so change = current value * sign
        for ci in range(2, 14):
            col_l = get_column_letter(ci)
            prev_col_l = get_column_letter(ci - 1) if ci > 2 else None
            bs_sheet = "'Balance Sheet'"
            if ci == 2:  # April - no previous month in our model
                f = f"=0"  # Can't determine change without prior year BS
            else:
                if sign == -1:
                    f = f"=-({bs_sheet}!{col_l}{bs_row}-{bs_sheet}!{prev_col_l}{bs_row})"
                else:
                    f = f"={bs_sheet}!{col_l}{bs_row}-{bs_sheet}!{prev_col_l}{bs_row}"
            ws.cell(row=row, column=ci, value=f)
        for q in range(1, 5):
            ws.cell(row=row, column=13 + q, value=quarterly_formula(row, q))
        ws.cell(row=row, column=18, value=fy_formula(row))
        apply_data_fmt(ws, row, 2, 18, INR)

    # A. Operating Activities
    r = 5
    ws.cell(row=r, column=1, value="A. CASH FROM OPERATING ACTIVITIES")
    apply_section(ws, r, 18)

    write_cf_formula(6, "Net Profit (PAT)", "=PnL!{c}47", indent=True)
    write_cf_formula(7, "Add: Depreciation", "=PnL!{c}42", indent=True)
    write_cf_formula(8, "Add: Finance Cost", "=PnL!{c}41", indent=True)

    ws.cell(row=9, column=1, value="  Working Capital Changes:").font = Font(name="Calibri", italic=True)
    # BS row references for WC items
    write_wc_change(10, "(Inc)/Dec in Sundry Debtors", 35, sign=-1)
    write_wc_change(11, "(Inc)/Dec in Loans & Advances", 34, sign=-1)
    write_wc_change(12, "Inc/(Dec) in Sundry Creditors", 19, sign=1)
    write_wc_change(13, "Inc/(Dec) in Salary Payable", 21, sign=1)
    write_wc_change(14, "Inc/(Dec) in Duties & Taxes", 18, sign=1)

    # Cash from operations
    ws.cell(row=16, column=1, value="Cash from Operations")
    for ci in range(2, 19):
        col_l = get_column_letter(ci)
        ws.cell(row=16, column=ci,
                value=f"={col_l}6+{col_l}7+{col_l}8+{col_l}10+{col_l}11+{col_l}12+{col_l}13+{col_l}14")
    apply_data_fmt(ws, 16, 2, 18, INR)
    apply_total(ws, 16, 18)

    # B. Investing Activities
    r = 18
    ws.cell(row=r, column=1, value="B. CASH FROM INVESTING ACTIVITIES")
    apply_section(ws, r, 18)

    # Change in fixed assets (increase = purchase = outflow)
    ws.cell(row=19, column=1, value="  Purchase of Fixed Assets").font = NORMAL_FONT
    for ci in range(2, 14):
        col_l = get_column_letter(ci)
        prev_col_l = get_column_letter(ci - 1) if ci > 2 else None
        bs = "'Balance Sheet'"
        if ci == 2:
            f = "=0"
        else:
            # Change in gross assets (row 28+29) = purchase
            f = f"=-({bs}!{col_l}28+{bs}!{col_l}29-{bs}!{prev_col_l}28-{bs}!{prev_col_l}29)"
        ws.cell(row=19, column=ci, value=f)
    for q in range(1, 5):
        ws.cell(row=19, column=13 + q, value=quarterly_formula(19, q))
    ws.cell(row=19, column=18, value=fy_formula(19))
    apply_data_fmt(ws, 19, 2, 18, INR)

    ws.cell(row=20, column=1, value="Cash from Investing")
    for ci in range(2, 19):
        col_l = get_column_letter(ci)
        ws.cell(row=20, column=ci, value=f"={col_l}19")
    apply_data_fmt(ws, 20, 2, 18, INR)
    apply_total(ws, 20, 18)

    # C. Financing Activities
    r = 22
    ws.cell(row=r, column=1, value="C. CASH FROM FINANCING ACTIVITIES")
    apply_section(ws, r, 18)

    # Change in loans
    ws.cell(row=23, column=1, value="  Proceeds / (Repayment) of Loans").font = NORMAL_FONT
    for ci in range(2, 14):
        col_l = get_column_letter(ci)
        prev_col_l = get_column_letter(ci - 1) if ci > 2 else None
        bs = "'Balance Sheet'"
        if ci == 2:
            f = "=0"
        else:
            f = f"={bs}!{col_l}15-{bs}!{prev_col_l}15"
        ws.cell(row=23, column=ci, value=f)
    for q in range(1, 5):
        ws.cell(row=23, column=13 + q, value=quarterly_formula(23, q))
    ws.cell(row=23, column=18, value=fy_formula(23))
    apply_data_fmt(ws, 23, 2, 18, INR)

    # Finance cost paid (outflow)
    write_cf_formula(24, "Finance Cost Paid", "=-PnL!{c}41", indent=True)

    ws.cell(row=25, column=1, value="Cash from Financing")
    for ci in range(2, 19):
        col_l = get_column_letter(ci)
        ws.cell(row=25, column=ci, value=f"={col_l}23+{col_l}24")
    apply_data_fmt(ws, 25, 2, 18, INR)
    apply_total(ws, 25, 18)

    # NET CASH FLOW
    ws.cell(row=27, column=1, value="NET CASH FLOW")
    for ci in range(2, 19):
        col_l = get_column_letter(ci)
        ws.cell(row=27, column=ci, value=f"={col_l}16+{col_l}20+{col_l}25")
    apply_data_fmt(ws, 27, 2, 18, INR)
    apply_total(ws, 27, 18)

    # Opening / Closing Cash
    ws.cell(row=29, column=1, value="Opening Cash & Bank").font = TOTAL_FONT
    # April opening = input cell (prior year closing)
    ws.cell(row=29, column=2, value=0)
    ws.cell(row=29, column=2).fill = INPUT_FILL
    ws.cell(row=29, column=2).number_format = INR
    for ci in range(3, 14):
        col_l = get_column_letter(ci)
        prev_col_l = get_column_letter(ci - 1)
        ws.cell(row=29, column=ci, value=f"={prev_col_l}30").number_format = INR
    for q in range(1, 5):
        ws.cell(row=29, column=13 + q, value="").number_format = INR
    ws.cell(row=29, column=18, value="=B29").number_format = INR

    ws.cell(row=30, column=1, value="Closing Cash & Bank").font = TOTAL_FONT
    for ci in range(2, 14):
        col_l = get_column_letter(ci)
        ws.cell(row=30, column=ci, value=f"={col_l}29+{col_l}27").number_format = INR
    ws.cell(row=30, column=18, value=f"=M30").number_format = INR
    apply_total(ws, 30, 18)

    red_rule = CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED))
    ws.conditional_formatting.add('B5:R30', red_rule)

    set_col_widths(ws, {'A': 35})
    for ci in range(2, 19):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = 'B5'
    ws.sheet_properties.tabColor = "C00000"
    return ws


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 5: KPIs & Financial Ratios                                      ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_kpi_sheet(wb):
    """KPIs with all formulas referencing PnL and BS sheets."""
    ws = wb.create_sheet("KPIs")
    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP",
                      "Key Performance Indicators & Financial Ratios — FY 2025-26")

    MAX_COL = 14  # A + 12 months + Latest
    r = 4
    headers = ['KPI / Ratio'] + MONTH_NAMES + ['FY Total']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h)
    apply_header(ws, r, len(headers))

    def write_kpi_formula(row, label, formula_tmpl, fmt=PCT, is_section=False):
        """formula_tmpl uses {c} for column letter placeholder."""
        ws.cell(row=row, column=1, value=label)
        if is_section:
            apply_section(ws, row, 14)
            return
        ws.cell(row=row, column=1).font = NORMAL_FONT
        for mi in range(12):
            ci = RPT_MONTH_START_COL + mi
            col_l = get_column_letter(ci)
            f = formula_tmpl.replace('{c}', col_l)
            cell = ws.cell(row=row, column=ci, value=f)
            cell.number_format = fmt
            cell.border = THIN_BORDER
        # FY column (col 14) — use R column from PnL (FY total, col 18)
        fy_f = formula_tmpl.replace('{c}', 'R').replace('PnL!', 'PnL!')
        # For FY, reference PnL column R (index 18)
        ws.cell(row=row, column=14, value=fy_f).number_format = fmt

    # Profitability
    write_kpi_formula(5, "PROFITABILITY RATIOS", "", is_section=True)
    write_kpi_formula(6, "Gross Profit Margin",
                      "=IFERROR(PnL!{c}21/PnL!{c}8,0)")
    write_kpi_formula(7, "EBITDA Margin",
                      "=IFERROR(PnL!{c}38/PnL!{c}8,0)")
    write_kpi_formula(8, "Net Profit Margin (PAT %)",
                      "=IFERROR(PnL!{c}47/PnL!{c}8,0)")
    write_kpi_formula(9, "COGS to Revenue",
                      "=IFERROR(PnL!{c}14/PnL!{c}8,0)")
    write_kpi_formula(10, "OPEX to Revenue",
                      "=IFERROR(PnL!{c}36/PnL!{c}8,0)")
    write_kpi_formula(11, "Direct Expense to Revenue",
                      "=IFERROR(PnL!{c}19/PnL!{c}8,0)")

    # Revenue KPIs
    write_kpi_formula(13, "REVENUE KPIs", "", is_section=True)
    write_kpi_formula(14, "Monthly Revenue (₹)",
                      "=PnL!{c}8", fmt=INR)
    write_kpi_formula(15, "Revenue Growth (MoM)",
                      "=IFERROR(PnL!{c}8/PnL!{prev}8-1,0)")
    # Fix MoM formula — needs previous column
    for mi in range(12):
        ci = RPT_MONTH_START_COL + mi
        col_l = get_column_letter(ci)
        if mi == 0:
            ws.cell(row=15, column=ci, value="=0")
        else:
            prev_l = get_column_letter(ci - 1)
            ws.cell(row=15, column=ci,
                    value=f"=IFERROR(PnL!{col_l}8/PnL!{prev_l}8-1,0)")
        ws.cell(row=15, column=ci).number_format = PCT
    ws.cell(row=15, column=14, value="").number_format = PCT

    write_kpi_formula(16, "Hospitality % of Revenue",
                      "=IFERROR(PnL!{c}6/PnL!{c}8,0)")
    write_kpi_formula(17, "Studio % of Revenue",
                      "=IFERROR(PnL!{c}7/PnL!{c}8,0)")

    # Liquidity (use latest BS values)
    write_kpi_formula(19, "LIQUIDITY RATIOS (Monthly BS)", "", is_section=True)
    # Current Ratio = Current Assets (BS row 38) / Current Liabilities (BS row 22)
    write_kpi_formula(20, "Current Ratio",
                      "=IFERROR('Balance Sheet'!{c}38/'Balance Sheet'!{c}22,0)", fmt=NUM2)
    write_kpi_formula(21, "Working Capital (₹)",
                      "='Balance Sheet'!{c}38-'Balance Sheet'!{c}22", fmt=INR)
    # Cash Ratio = (Cash + Bank) / CL
    write_kpi_formula(22, "Cash Ratio",
                      "=IFERROR(('Balance Sheet'!{c}36+'Balance Sheet'!{c}37)/'Balance Sheet'!{c}22,0)", fmt=NUM2)

    # Operational
    write_kpi_formula(24, "OPERATIONAL KPIs", "", is_section=True)
    # HR cost as % of revenue
    write_kpi_formula(25, "HR Cost % of Revenue",
                      "=IFERROR(PnL!{c}29/PnL!{c}8,0)")
    # Monthly OPEX
    write_kpi_formula(26, "Monthly OPEX (₹)",
                      "=PnL!{c}36", fmt=INR)

    # Fix column widths
    set_col_widths(ws, {'A': 32})
    for ci in range(2, 15):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = 'B5'
    ws.sheet_properties.tabColor = "ED7D31"
    return ws


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 6: Dashboard (Executive Summary with Charts)                    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_dashboard_sheet(wb):
    """Executive dashboard with summary metrics and charts."""
    ws = wb.create_sheet("Dashboard")
    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP",
                      "Executive Dashboard — FY 2025-26")

    # ── Key Metrics Summary (reference PnL FY column R=18) ──
    r = 4
    ws.cell(row=r, column=1, value="KEY METRICS (FY 2025-26)")
    apply_section(ws, r, 6)

    metrics = [
        ("Total Revenue", "=PnL!R8", INR),
        ("Gross Profit", "=PnL!R21", INR),
        ("Gross Profit %", "=IFERROR(PnL!R21/PnL!R8,0)", PCT),
        ("Total OPEX", "=PnL!R36", INR),
        ("EBITDA", "=PnL!R38", INR),
        ("EBITDA %", "=IFERROR(PnL!R38/PnL!R8,0)", PCT),
        ("PAT", "=PnL!R47", INR),
        ("PAT %", "=IFERROR(PnL!R47/PnL!R8,0)", PCT),
        ("Total Assets", "='Balance Sheet'!N40", INR),
        ("Total Equity", "='Balance Sheet'!N9", INR),
        ("Current Ratio", "=IFERROR('Balance Sheet'!N38/'Balance Sheet'!N22,0)", NUM2),
        ("Working Capital", "='Balance Sheet'!N38-'Balance Sheet'!N22", INR),
    ]

    ws.cell(row=5, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=5, column=1).fill = HEADER_FILL
    ws.cell(row=5, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=5, column=2).fill = HEADER_FILL

    for i, (label, formula, fmt) in enumerate(metrics):
        ws.cell(row=6 + i, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=6 + i, column=1).border = THIN_BORDER
        cell = ws.cell(row=6 + i, column=2, value=formula)
        cell.number_format = fmt
        cell.font = Font(name="Calibri", bold=True, size=12, color=NAVY)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')

    # ── Quarterly Summary Table ──
    qs_row = 4
    ws.cell(row=qs_row, column=4, value="QUARTERLY SUMMARY")
    apply_section(ws, qs_row, 9)
    # Adjust section to only cover cols 4-9
    for c in range(1, 4):
        ws.cell(row=qs_row, column=c).fill = PatternFill()

    q_headers = ['Metric', 'Q1', 'Q2', 'Q3', 'Q4', 'FY']
    for ci, h in enumerate(q_headers):
        cell = ws.cell(row=qs_row + 1, column=4 + ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    q_metrics = [
        ("Revenue", "N8", "O8", "P8", "Q8", "R8"),
        ("COGS", "N14", "O14", "P14", "Q14", "R14"),
        ("Gross Profit", "N21", "O21", "P21", "Q21", "R21"),
        ("OPEX", "N36", "O36", "P36", "Q36", "R36"),
        ("EBITDA", "N38", "O38", "P38", "Q38", "R38"),
        ("PAT", "N47", "O47", "P47", "Q47", "R47"),
    ]

    for i, (label, *refs) in enumerate(q_metrics):
        r = qs_row + 2 + i
        ws.cell(row=r, column=4, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER
        for j, ref in enumerate(refs):
            cell = ws.cell(row=r, column=5 + j, value=f"=PnL!{ref}")
            cell.number_format = INR
            cell.border = THIN_BORDER

    # ── Charts ──
    # Monthly Revenue Trend
    chart_data_start = 20
    ws.cell(row=chart_data_start, column=1, value="Chart Data (hidden)").font = Font(size=8, color="D9D9D9")
    for mi in range(12):
        ws.cell(row=chart_data_start + 1, column=1 + mi, value=MONTH_NAMES[mi])
    for mi in range(12):
        col_l = get_column_letter(RPT_MONTH_START_COL + mi)
        ws.cell(row=chart_data_start + 2, column=1 + mi, value=f"=PnL!{col_l}8")  # Revenue
        ws.cell(row=chart_data_start + 3, column=1 + mi, value=f"=PnL!{col_l}21")  # GP
        ws.cell(row=chart_data_start + 4, column=1 + mi, value=f"=PnL!{col_l}38")  # EBITDA
        ws.cell(row=chart_data_start + 5, column=1 + mi, value=f"=PnL!{col_l}47")  # PAT

    # Revenue & Profitability Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue & Profitability Trend"
    chart.y_axis.title = "Amount (₹)"
    chart.style = 10
    chart.width = 28
    chart.height = 14

    cats = Reference(ws, min_col=1, max_col=12, min_row=chart_data_start + 1)
    for offset, series_name in [(2, "Revenue"), (3, "Gross Profit"), (4, "EBITDA"), (5, "PAT")]:
        data = Reference(ws, min_col=1, max_col=12, min_row=chart_data_start + offset)
        chart.add_data(data, from_rows=True)
        chart.series[-1].title = openpyxl.chart.series.SeriesLabel(v=series_name)

    chart.set_categories(cats)
    ws.add_chart(chart, "A22")

    # Margin trend line chart
    # Write margin data
    for mi in range(12):
        col_l = get_column_letter(RPT_MONTH_START_COL + mi)
        ws.cell(row=chart_data_start + 7, column=1 + mi,
                value=f"=IFERROR(PnL!{col_l}21/PnL!{col_l}8,0)")  # GP%
        ws.cell(row=chart_data_start + 8, column=1 + mi,
                value=f"=IFERROR(PnL!{col_l}38/PnL!{col_l}8,0)")  # EBITDA%
        ws.cell(row=chart_data_start + 9, column=1 + mi,
                value=f"=IFERROR(PnL!{col_l}47/PnL!{col_l}8,0)")  # PAT%

    line_chart = LineChart()
    line_chart.title = "Margin Trends (%)"
    line_chart.y_axis.title = "Percentage"
    line_chart.style = 10
    line_chart.width = 28
    line_chart.height = 14

    cats2 = Reference(ws, min_col=1, max_col=12, min_row=chart_data_start + 1)
    for offset, series_name in [(7, "GP %"), (8, "EBITDA %"), (9, "PAT %")]:
        data = Reference(ws, min_col=1, max_col=12, min_row=chart_data_start + offset)
        line_chart.add_data(data, from_rows=True)
        line_chart.series[-1].title = openpyxl.chart.series.SeriesLabel(v=series_name)
    line_chart.set_categories(cats2)
    ws.add_chart(line_chart, "A40")

    set_col_widths(ws, {'A': 22, 'B': 18, 'C': 3, 'D': 16, 'E': 14, 'F': 14, 'G': 14, 'H': 14, 'I': 14})
    ws.sheet_properties.tabColor = "7030A0"
    return ws


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 7: Quarterly P&L (Compact Presentation View)                    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_quarterly_pnl_sheet(wb):
    """Compact quarterly P&L for board presentations."""
    ws = wb.create_sheet("Quarterly PnL")
    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP",
                      "Quarterly Performance Summary — FY 2025-26")

    r = 4
    headers = ['Particulars', 'Q1 (Apr-Jun)', 'Q2 (Jul-Sep)', 'Q3 (Oct-Dec)',
               'Q4 (Jan-Mar)', 'FY 2025-26', 'YoY Target', 'Variance']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h)
    apply_header(ws, r, len(headers))

    # Map PnL quarterly columns: N=Q1(14), O=Q2(15), P=Q3(16), Q=Q4(17), R=FY(18)
    pnl_q_cols = ['N', 'O', 'P', 'Q', 'R']

    def write_q_row(row, label, pnl_row, is_total=False, is_pct=False):
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        fmt = PCT if is_pct else INR
        for qi, qcol in enumerate(pnl_q_cols):
            cell = ws.cell(row=row, column=2 + qi, value=f"=PnL!{qcol}{pnl_row}")
            cell.number_format = fmt
            cell.border = THIN_BORDER
        if not is_pct:
            # YoY Target (30% growth — user adjustable)
            ws.cell(row=row, column=7, value=f"=F{row}*1.3").number_format = INR
            ws.cell(row=row, column=7).fill = INPUT_FILL
            # Variance
            ws.cell(row=row, column=8, value=f"=F{row}-G{row}").number_format = INR
        if is_total:
            apply_total(ws, row, 8)

    ws.cell(row=5, column=1, value="REVENUE")
    apply_section(ws, 5, 8)
    write_q_row(6, "  Hospitality Services", 6)
    write_q_row(7, "  Studio Rental Services", 7)
    write_q_row(8, "TOTAL REVENUE", 8, is_total=True)

    ws.cell(row=10, column=1, value="COST OF GOODS SOLD")
    apply_section(ws, 10, 8)
    write_q_row(11, "  Total COGS", 14)
    write_q_row(12, "  Direct Expenses", 19)

    write_q_row(14, "GROSS PROFIT", 21, is_total=True)
    write_q_row(15, "GP %", 22, is_pct=True)

    ws.cell(row=17, column=1, value="OPERATING EXPENSES")
    apply_section(ws, 17, 8)
    write_q_row(18, "  Total OPEX", 36)

    write_q_row(20, "EBITDA", 38, is_total=True)
    write_q_row(21, "EBITDA %", 39, is_pct=True)

    write_q_row(23, "  Finance Cost", 41)
    write_q_row(24, "  Depreciation", 42)

    write_q_row(26, "PROFIT BEFORE TAX", 44, is_total=True)
    write_q_row(27, "  Income Tax", 45)
    write_q_row(29, "PROFIT AFTER TAX", 47, is_total=True)
    write_q_row(30, "PAT %", 48, is_pct=True)

    # Conditional formatting
    red_rule = CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED))
    ws.conditional_formatting.add('B5:H30', red_rule)

    set_col_widths(ws, {'A': 28, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 18})
    ws.sheet_properties.tabColor = "FFC000"
    return ws


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 8: Budget vs Actuals                                            ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_budget_sheet(wb):
    """Budget vs Actuals with variance analysis."""
    ws = wb.create_sheet("Budget vs Actuals")
    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP",
                      "Budget vs Actuals — FY 2025-26")

    r = 4
    headers = ['Particulars', 'Actuals (Q1-Q3)', 'Budget (Q4)',
               'FY Actual+Proj', 'Annual Budget', 'Variance (₹)', 'Variance %', 'Status']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h)
    apply_header(ws, r, len(headers))

    # PnL quarterly: N=Q1, O=Q2, P=Q3, Q=Q4, R=FY
    def write_bva_row(row, label, pnl_row, is_total=False):
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        # Actuals Q1-Q3
        ws.cell(row=row, column=2, value=f"=PnL!N{pnl_row}+PnL!O{pnl_row}+PnL!P{pnl_row}").number_format = INR
        # Budget Q4
        ws.cell(row=row, column=3, value=f"=PnL!Q{pnl_row}").number_format = INR
        # FY Actual+Proj
        ws.cell(row=row, column=4, value=f"=PnL!R{pnl_row}").number_format = INR
        # Annual Budget (user input — yellow)
        cell = ws.cell(row=row, column=5, value=0)
        cell.number_format = INR
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        # Variance
        ws.cell(row=row, column=6, value=f"=D{row}-E{row}").number_format = INR
        # Variance %
        ws.cell(row=row, column=7, value=f"=IFERROR(F{row}/E{row},0)").number_format = PCT
        # Status
        ws.cell(row=row, column=8,
                value=f'=IF(E{row}=0,"",IF(F{row}>=0,"✓ On Track","✗ Below"))').font = NORMAL_FONT
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = THIN_BORDER
        if is_total:
            apply_total(ws, row, 8)

    write_bva_row(5, "Total Revenue", 8, is_total=True)
    write_bva_row(6, "Total COGS", 14)
    write_bva_row(7, "Direct Expenses", 19)
    write_bva_row(8, "Gross Profit", 21, is_total=True)
    write_bva_row(9, "Total OPEX", 36)
    write_bva_row(10, "EBITDA", 38, is_total=True)
    write_bva_row(11, "Finance Cost", 41)
    write_bva_row(12, "Depreciation", 42)
    write_bva_row(13, "PBT", 44, is_total=True)
    write_bva_row(14, "PAT", 47, is_total=True)

    # Instructions
    ws.cell(row=16, column=1, value="Enter Annual Budget figures in the yellow cells (Column E)").font = Font(
        name="Calibri", italic=True, color="808080")
    ws.cell(row=17, column=1, value="Variance = Actual - Budget; Positive = favorable").font = Font(
        name="Calibri", italic=True, color="808080")

    set_col_widths(ws, {'A': 22, 'B': 20, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 14, 'H': 14})
    ws.sheet_properties.tabColor = "00B050"
    return ws


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  MAIN GENERATOR                                                        ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def generate_mis_master(source_path, output_path, company="THOTA HOSPITALITY LLP"):
    """Generate the complete MIS Master sheet."""
    print(f"📊 Reading Trial Balance from: {source_path}")

    # Read TB values
    tb_data, period = read_trial_balance(source_path)
    print(f"   Period: {period}")
    print(f"   Accounts found: {len(tb_data)}")

    # Map TB data to our account IDs
    jan_values = {}

    # Direct mapping from TB account names to our IDs
    tb_mapping = {
        "Partner's Capital":              ('cap_partner', 'Cr'),
        "Profit & Loss A/c":              ('cap_pnl_prior', 'Dr'),
        "Partner's Loan From Akshatha":   ('loan_akshatha', 'Cr'),
        "Partner's Loan From Tejas":      ('loan_tejas', 'Cr'),
        "TVS Credit Services Limited":    ('loan_tvs', 'Cr'),
        "Intangible Assets":              ('fa_intangible', 'Dr'),
        "Tangible Assets":                ('fa_tangible', 'Dr'),
        "Loans & Advances (Asset)":       ('ca_advances', 'Dr'),
        "Cash-in-Hand":                   ('ca_cash', 'Dr'),
        "Bank Accounts":                  ('ca_bank', 'Dr'),
        "Hospitality Services @ 5%":      ('rev_hospitality', 'Cr'),
        "Studio Rental Services @ 18%":   ('rev_studio', 'Cr'),
        "Thota Kitchen":                  ('cogs_kitchen', 'Dr'),
        "Thota Decor":                    ('cogs_decor', 'Dr'),
        "Studio":                         ('cogs_studio', 'Dr'),
        "Fuel Expenses":                  ('de_fuel', 'Dr'),
        "Transportation Expenses":        ('de_transport', 'Dr'),
        "Discount Received":              ('oi_discount', 'Cr'),
        "Administrative Overheads":       ('opex_admin', 'Dr'),
        "Finance Cost":                   ('opex_finance', 'Dr'),
        "HR Expenses":                    ('opex_hr', 'Dr'),
        "Marketing & Ads":                ('opex_marketing', 'Dr'),
        "Professional Charges":           ('opex_professional', 'Dr'),
        "AUDIT / LEGAL PROFESSIONAL FEE": ('opex_audit', 'Dr'),
        "Employer PT":                    ('opex_pt', 'Dr'),
        "Rates and Taxes":                ('opex_rates', 'Cr'),
        "Repairs & Maintenance":          ('opex_rm', 'Dr'),
        "Round Off":                      ('opex_roundoff', 'Cr'),
    }

    for tb_name, (acct_id, nature) in tb_mapping.items():
        if tb_name in tb_data:
            debit, credit = tb_data[tb_name]
            if nature == 'Dr':
                jan_values[acct_id] = debit if debit else 0
            else:
                jan_values[acct_id] = credit if credit else 0

    # Special: Net accounts (have both debit and credit)
    net_mappings = {
        'cl_duties': 'Duties & Taxes',
        'cl_creditors': 'Sundry Creditors',
        'cl_reimburse': 'Reimburements',
        'cl_salary': 'Salary Payable',
        'ca_debtors': 'Sundry Debtors',
    }
    for acct_id, tb_name in net_mappings.items():
        if tb_name in tb_data:
            debit, credit = tb_data[tb_name]
            if acct_id.startswith('cl_'):  # Credit nature
                jan_values[acct_id] = credit - debit
            else:  # Debit nature (debtors)
                jan_values[acct_id] = debit - credit

    # Handle sub-accounts that may need aggregation
    # Reimbursements may be split across multiple employee names
    if 'cl_reimburse' not in jan_values or jan_values.get('cl_reimburse', 0) == 0:
        reimb_debit = 0
        reimb_credit = 0
        for name, (d, c) in tb_data.items():
            if 'reimbur' in name.lower() or 'Reimburements' in name:
                reimb_debit += d
                reimb_credit += c
        if reimb_credit > 0 or reimb_debit > 0:
            jan_values['cl_reimburse'] = reimb_credit - reimb_debit

    print(f"   Mapped {len(jan_values)} accounts to model")

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("📋 Creating TB Input sheet...")
    create_tb_sheet(wb, jan_values)

    print("📈 Creating P&L sheet (formula-linked)...")
    create_pnl_sheet(wb)

    print("📊 Creating Balance Sheet (formula-linked)...")
    create_bs_sheet(wb)

    print("💰 Creating Cash Flow sheet (formula-linked)...")
    create_cashflow_sheet(wb)

    print("📉 Creating KPIs sheet (formula-linked)...")
    create_kpi_sheet(wb)

    print("🎯 Creating Dashboard (charts + summary)...")
    create_dashboard_sheet(wb)

    print("📋 Creating Quarterly P&L (presentation view)...")
    create_quarterly_pnl_sheet(wb)

    print("📊 Creating Budget vs Actuals...")
    create_budget_sheet(wb)

    # Set Dashboard as the active sheet when opening
    wb.active = wb.sheetnames.index("Dashboard")

    # Save
    wb.save(output_path)
    print(f"\n✅ MIS Master Sheet generated: {output_path}")
    print(f"📌 IMPORTANT: This file has {len(jan_values)} accounts pre-filled for January.")
    print(f"   → Paste monthly TBs into columns Apr-Mar on the 'TB' sheet.")
    print(f"   → All report sheets auto-update via formulas. No manual changes needed.")
    return output_path


# ── CLI Entry Point ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate MIS Master Sheet from Trial Balance")
    parser.add_argument("--source", required=True, help="Path to Trial Balance Excel file")
    parser.add_argument("--output", default=None, help="Output file path")
    parser.add_argument("--company", default="THOTA HOSPITALITY LLP", help="Company name")
    args = parser.parse_args()

    source = Path(args.source)
    if not source.exists():
        print(f"❌ File not found: {source}")
        sys.exit(1)

    if args.output:
        output = args.output
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = str(TMP_DIR / f"MIS_Master_{ts}.xlsx")

    generate_mis_master(str(source), output, args.company)


if __name__ == "__main__":
    main()
