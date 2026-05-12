#!/usr/bin/env python3
"""
Formula-Linked Financial Model Generator — THOTA HOSPITALITY LLP

Creates an Excel workbook where:
- Trial_Balance sheet = ONLY INPUT (paste from Tally)
- All other sheets use EXCEL FORMULAS referencing the TB
- ZERO hardcoded numbers in any report sheet
- Monthly tracking via Monthly_TB input sheet

Output sheets:
1. Trial_Balance (INPUT)
2. Monthly_TB (MONTHLY INPUT - cumulative TB per month)
3. P&L_YTD (Annual/YTD P&L from TB)
4. P&L_Monthly (Monthly + Quarterly P&L from Monthly_TB)
5. Balance_Sheet (from TB)
6. Cash_Flow (derived from P&L and BS)
7. Revenue_Summary (segment-wise from TB)
8. OPEX_Schedule (detailed expense from TB)
9. KPIs_Ratios (all formula-calculated)
10. MIS_Dashboard (summary for management)
"""

import os
import sys
import shutil
import argparse
from pathlib import Path
from datetime import datetime
from copy import copy

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

# ── Configuration ────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(exist_ok=True)

# ── Style Constants ──────────────────────────────────────────────────────────

NAVY = "2F5496"
LIGHT_BLUE = "D6E4F0"
DARK_BLUE = "1F3864"
GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
RED = "C00000"
ORANGE = "ED7D31"
LIGHT_GRAY = "F2F2F2"
GRAY = "808080"
WHITE = "FFFFFF"

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11, color=NAVY)
SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="4472C4")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
FORMULA_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
GROUP_FONT = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
DETAIL_FONT = Font(name="Calibri", size=10)

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
BOTTOM_DOUBLE = Border(
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='double', color=NAVY),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)
INPUT_BORDER = Border(
    left=Side(style='thin', color=ORANGE),
    right=Side(style='thin', color=ORANGE),
    top=Side(style='thin', color=ORANGE),
    bottom=Side(style='thin', color=ORANGE)
)

INR_FMT = '₹#,##0'
INR_DEC = '₹#,##0.00'
NUM_FMT = '#,##0'
NUM_DEC = '#,##0.00'
PCT_FMT = '0.0%'
PCT_DEC = '0.00%'

# ── TB Account Row Mapping ──────────────────────────────────────────────────
# These are the ROW NUMBERS in the Trial Balance sheet where each account appears.
# When the user pastes a new TB, as long as the chart of accounts order is the same,
# these references remain valid. If Tally adds accounts, the user adjusts these.

# ROW numbers in the TB sheet (1-indexed, matching the source file layout)
# NOTE: These will be set dynamically after copying TB data

TB_ROWS = {}  # Will be populated when TB is loaded


def find_tb_rows(ws):
    """Scan the Trial Balance sheet and find row numbers for key accounts."""
    rows = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        val_str = str(val).strip()

        # Map account names to keys
        mappings = {
            'Capital Account': 'capital_account',
            "Partner's Capital": 'partners_capital',
            'Loans (Liability)': 'loans_liability',
            'Unsecured Loans': 'unsecured_loans',
            "Partner's Loan From Akshatha": 'loan_akshatha',
            "Partner's Loan From Tejas": 'loan_tejas',
            'TVS Credit Services Limited': 'tvs_credit',
            'Current Liabilities': 'current_liabilities',
            'Duties & Taxes': 'duties_taxes',
            'Sundry Creditors': 'sundry_creditors',
            'Employee Advances': 'employee_advances',
            'Reimburements': 'reimbursements',
            'Salary Payable': 'salary_payable',
            'Fixed Assets': 'fixed_assets',
            'Intangible Assets': 'intangible_assets',
            'Tangible Assets': 'tangible_assets',
            'Current Assets': 'current_assets',
            'Loans & Advances (Asset)': 'loans_advances_asset',
            'Sundry Debtors': 'sundry_debtors',
            'Cash-in-Hand': 'cash_in_hand',
            'Bank Accounts': 'bank_accounts',
            'Sales Accounts': 'sales_accounts',
            'Hospitality Services @ 5%': 'hospitality_revenue',
            'Studio Rental Services @ 18%': 'studio_revenue',
            'Purchase Accounts': 'purchase_accounts',
            'Studio': 'purchase_studio',
            'Thota Kitchen': 'purchase_kitchen',
            'Thota Decor': 'purchase_decor',
            'Direct Expenses': 'direct_expenses',
            'Fuel Expenses': 'fuel_expenses',
            'Transportation Expenses': 'transportation',
            'Indirect Incomes': 'indirect_incomes',
            'Discount Received': 'discount_received',
            'Indirect Expenses': 'indirect_expenses',
            'Administrative Overheads': 'admin_overheads',
            'Finance Cost': 'finance_cost',
            'Bank Charges': 'bank_charges',
            'Interest Expense': 'interest_expense',
            'HR Expenses': 'hr_expenses',
            'Employee Salaries': 'employee_salaries',
            'Incentives': 'incentives',
            'PF Admin Charges': 'pf_admin',
            'Stipend': 'stipend',
            'AUDIT / LEGAL PROFESSIONAL FEE': 'audit_legal',
            'Employer PT': 'employer_pt',
            'Marketing & Ads': 'marketing',
            'Professional Charges': 'professional_charges',
            'Rates and Taxes': 'rates_taxes',
            'Repairs & Maintenance': 'repairs_maintenance',
            'Round Off': 'round_off',
            'Profit & Loss A/c': 'pnl_account',
            'Grand Total': 'grand_total',
        }

        for account_name, key in mappings.items():
            if val_str == account_name and key not in rows:
                rows[key] = row
                break

    return rows


# ── Helper Functions ─────────────────────────────────────────────────────────

def apply_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def apply_subheader(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def apply_total_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = BOTTOM_DOUBLE


def write_title(ws, row, col, title, subtitle=None):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = TITLE_FONT
    if subtitle:
        cell2 = ws.cell(row=row + 1, column=col, value=subtitle)
        cell2.font = SUBTITLE_FONT
        return row + 3
    return row + 2


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def tb_ref(key, col='B', sheet='Trial_Balance'):
    """Generate a formula reference to a Trial Balance cell."""
    row = TB_ROWS.get(key)
    if row is None:
        return '0'
    return f"'{sheet}'!{col}{row}"


def tb_dr(key, sheet='Trial_Balance'):
    """Reference the Debit column (B) of a TB account."""
    return tb_ref(key, 'B', sheet)


def tb_cr(key, sheet='Trial_Balance'):
    """Reference the Credit column (C) of a TB account."""
    return tb_ref(key, 'C', sheet)


def mtb_dr(key, month_col_offset, sheet='Monthly_TB'):
    """Reference a monthly TB debit column."""
    row = MTB_ROWS.get(key)
    if row is None:
        return '0'
    col_letter = get_column_letter(3 + month_col_offset * 2)  # C, E, G, ...
    return f"'{sheet}'!{col_letter}{row}"


def mtb_cr(key, month_col_offset, sheet='Monthly_TB'):
    """Reference a monthly TB credit column."""
    row = MTB_ROWS.get(key)
    if row is None:
        return '0'
    col_letter = get_column_letter(4 + month_col_offset * 2)  # D, F, H, ...
    return f"'{sheet}'!{col_letter}{row}"


# Monthly TB row mapping (will be set when creating Monthly_TB)
MTB_ROWS = {}

# Month names
MONTHS = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
QUARTERS = {
    'Q1': [0, 1, 2],     # Apr, May, Jun
    'Q2': [3, 4, 5],     # Jul, Aug, Sep
    'Q3': [6, 7, 8],     # Oct, Nov, Dec
    'Q4': [9, 10, 11],   # Jan, Feb, Mar
}


# ── Sheet 1: Trial Balance (Input) ──────────────────────────────────────────

def create_trial_balance_sheet(wb, source_path):
    """Copy the Trial Balance data as the input sheet."""
    ws = wb.create_sheet("Trial_Balance")
    src_wb = openpyxl.load_workbook(source_path, data_only=True)
    src_ws = src_wb['Trial Balance']

    # Copy all data from source
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=3):
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            # Copy basic formatting
            if cell.value is not None:
                new_cell.border = THIN_BORDER
                if isinstance(cell.value, (int, float)):
                    new_cell.number_format = NUM_FMT
                    new_cell.alignment = Alignment(horizontal='right')

    # Style the headers
    for r in range(1, 9):
        ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11, color=NAVY)

    # Column headers
    ws.cell(row=10, column=1).font = HEADER_FONT
    ws.cell(row=10, column=1).fill = HEADER_FILL
    ws.cell(row=11, column=2).font = HEADER_FONT
    ws.cell(row=11, column=2).fill = HEADER_FILL
    ws.cell(row=12, column=2).font = HEADER_FONT
    ws.cell(row=12, column=2).fill = HEADER_FILL
    ws.cell(row=12, column=3).font = HEADER_FONT
    ws.cell(row=12, column=3).fill = HEADER_FILL

    # Highlight group headers
    group_rows = []
    for r in range(13, src_ws.max_row + 1):
        val = src_ws.cell(row=r, column=1).value
        if val:
            val_str = str(val).strip()
            if val_str in [
                'Capital Account', 'Loans (Liability)', 'Current Liabilities',
                'Fixed Assets', 'Current Assets', 'Sales Accounts',
                'Purchase Accounts', 'Direct Expenses', 'Indirect Incomes',
                'Indirect Expenses', 'Profit & Loss A/c', 'Grand Total'
            ]:
                group_rows.append(r)
                ws.cell(row=r, column=1).font = GROUP_FONT
                ws.cell(row=r, column=1).fill = PatternFill(start_color="E8EDF4", end_color="E8EDF4", fill_type="solid")
                if ws.cell(row=r, column=2).value:
                    ws.cell(row=r, column=2).fill = PatternFill(start_color="E8EDF4", end_color="E8EDF4", fill_type="solid")
                if ws.cell(row=r, column=3).value:
                    ws.cell(row=r, column=3).fill = PatternFill(start_color="E8EDF4", end_color="E8EDF4", fill_type="solid")

    # Mark input cells (B and C, rows 13+) with light yellow
    for r in range(13, src_ws.max_row + 1):
        for c in [2, 3]:
            cell = ws.cell(row=r, column=c)
            if r not in group_rows:
                cell.fill = INPUT_FILL
                cell.border = INPUT_BORDER

    # Grand total row
    gt_row = src_ws.max_row
    for c in [1, 2, 3]:
        ws.cell(row=gt_row, column=c).font = TOTAL_FONT
        ws.cell(row=gt_row, column=c).fill = TOTAL_FILL
        ws.cell(row=gt_row, column=c).border = BOTTOM_DOUBLE

    # Add instruction note
    note_row = gt_row + 2
    ws.cell(row=note_row, column=1, value="📌 INSTRUCTIONS:").font = Font(name="Calibri", bold=True, size=11, color=RED)
    ws.cell(row=note_row + 1, column=1, value="• This is the INPUT sheet. Paste your Trial Balance from Tally here.").font = Font(name="Calibri", size=10, color=GRAY)
    ws.cell(row=note_row + 2, column=1, value="• Yellow cells = Input values. Update Debit (B) and Credit (C) columns.").font = Font(name="Calibri", size=10, color=GRAY)
    ws.cell(row=note_row + 3, column=1, value="• All other sheets auto-calculate via formulas. DO NOT edit report sheets.").font = Font(name="Calibri", size=10, color=GRAY)
    ws.cell(row=note_row + 4, column=1, value="• Keep account names and row order consistent with Tally export.").font = Font(name="Calibri", size=10, color=GRAY)

    set_widths(ws, {'A': 40, 'B': 18, 'C': 18})

    # Find row mapping
    global TB_ROWS
    TB_ROWS = find_tb_rows(ws)

    src_wb.close()
    return ws


# ── Sheet 2: Monthly TB (Monthly Input) ─────────────────────────────────────

# Key accounts for monthly tracking
MONTHLY_ACCOUNTS = [
    ('sales_accounts', 'Sales Accounts', 'Revenue'),
    ('hospitality_revenue', '  Hospitality Services', 'Revenue'),
    ('studio_revenue', '  Studio Rental Services', 'Revenue'),
    ('purchase_accounts', 'Purchase Accounts', 'COGS'),
    ('purchase_kitchen', '  Thota Kitchen', 'COGS'),
    ('purchase_decor', '  Thota Decor', 'COGS'),
    ('purchase_studio', '  Studio Purchases', 'COGS'),
    ('direct_expenses', 'Direct Expenses', 'Direct Cost'),
    ('fuel_expenses', '  Fuel Expenses', 'Direct Cost'),
    ('transportation', '  Transportation', 'Direct Cost'),
    ('indirect_incomes', 'Indirect Incomes', 'Other Income'),
    ('indirect_expenses', 'Indirect Expenses (Total)', 'OPEX'),
    ('admin_overheads', '  Admin Overheads', 'OPEX'),
    ('finance_cost', '  Finance Cost', 'OPEX'),
    ('hr_expenses', '  HR Expenses', 'OPEX'),
    ('marketing', '  Marketing & Ads', 'OPEX'),
    ('professional_charges', '  Professional Charges', 'OPEX'),
    ('repairs_maintenance', '  Repairs & Maintenance', 'OPEX'),
    ('audit_legal', '  Audit / Legal Fee', 'OPEX'),
    ('employer_pt', '  Employer PT', 'OPEX'),
    ('rates_taxes', '  Rates & Taxes', 'OPEX'),
    ('round_off', '  Round Off', 'OPEX'),
    # Balance Sheet items for monthly tracking
    ('capital_account', 'Capital Account', 'BS-Equity'),
    ('loans_liability', 'Loans (Liability)', 'BS-Liability'),
    ('current_liabilities', 'Current Liabilities', 'BS-CL'),
    ('duties_taxes', '  Duties & Taxes', 'BS-CL'),
    ('sundry_creditors', '  Sundry Creditors', 'BS-CL'),
    ('salary_payable', '  Salary Payable', 'BS-CL'),
    ('fixed_assets', 'Fixed Assets', 'BS-NCA'),
    ('current_assets', 'Current Assets', 'BS-CA'),
    ('sundry_debtors', '  Sundry Debtors', 'BS-CA'),
    ('cash_in_hand', '  Cash-in-Hand', 'BS-CA'),
    ('bank_accounts', '  Bank Accounts', 'BS-CA'),
    ('loans_advances_asset', '  Loans & Advances', 'BS-CA'),
    ('pnl_account', 'Profit & Loss A/c', 'BS-RE'),
]


def create_monthly_tb_sheet(wb):
    """Create the Monthly TB input sheet for month-by-month tracking."""
    ws = wb.create_sheet("Monthly_TB")
    global MTB_ROWS

    row = 1
    ws.cell(row=row, column=1, value="THOTA HOSPITALITY LLP").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Monthly Trial Balance — Cumulative Figures from Tally").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Enter CUMULATIVE TB figures for each month. Monthly movement is auto-calculated.").font = Font(name="Calibri", size=10, color=GRAY, italic=True)
    row += 2  # Row 5

    # Headers
    ws.cell(row=row, column=1, value="Account")
    ws.cell(row=row, column=2, value="Category")

    # Month headers - 2 columns per month (Dr/Cr)
    for i, month in enumerate(MONTHS):
        dr_col = 3 + i * 2
        cr_col = 4 + i * 2
        ws.cell(row=row, column=dr_col, value=f"{month}\nDr")
        ws.cell(row=row, column=cr_col, value=f"{month}\nCr")

    # YTD columns
    ytd_dr_col = 3 + 12 * 2  # Column 27
    ytd_cr_col = ytd_dr_col + 1  # Column 28
    ws.cell(row=row, column=ytd_dr_col, value="YTD\nDr")
    ws.cell(row=row, column=ytd_cr_col, value="YTD\nCr")

    max_col = ytd_cr_col
    apply_header(ws, row, max_col)
    ws.row_dimensions[row].height = 30
    row += 1  # Row 6

    # Account rows
    data_start_row = row
    for key, name, category in MONTHLY_ACCOUNTS:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=category)

        # Style
        if not name.startswith('  '):
            ws.cell(row=row, column=1).font = GROUP_FONT
            ws.cell(row=row, column=2).font = Font(name="Calibri", size=9, color=GRAY)
        else:
            ws.cell(row=row, column=1).font = DETAIL_FONT
            ws.cell(row=row, column=2).font = Font(name="Calibri", size=9, color=GRAY)

        # Monthly input cells (yellow background)
        for i in range(12):
            dr_col = 3 + i * 2
            cr_col = 4 + i * 2
            ws.cell(row=row, column=dr_col).fill = INPUT_FILL
            ws.cell(row=row, column=dr_col).border = INPUT_BORDER
            ws.cell(row=row, column=dr_col).number_format = NUM_FMT
            ws.cell(row=row, column=cr_col).fill = INPUT_FILL
            ws.cell(row=row, column=cr_col).border = INPUT_BORDER
            ws.cell(row=row, column=cr_col).number_format = NUM_FMT

        # YTD formulas - sum of all months (but only the LATEST non-empty month matters for cumulative)
        # Since values are cumulative, YTD = the latest filled month
        # We'll use: pick the rightmost non-empty value
        # Formula: use the last filled month's value
        # Simpler: just reference Trial_Balance for YTD
        ytd_dr_formula = f"=Trial_Balance!B{TB_ROWS.get(key, 1)}" if key in TB_ROWS else "=0"
        ytd_cr_formula = f"=Trial_Balance!C{TB_ROWS.get(key, 1)}" if key in TB_ROWS else "=0"
        ws.cell(row=row, column=ytd_dr_col, value=ytd_dr_formula).number_format = NUM_FMT
        ws.cell(row=row, column=ytd_cr_col, value=ytd_cr_formula).number_format = NUM_FMT
        ws.cell(row=row, column=ytd_dr_col).fill = FORMULA_FILL
        ws.cell(row=row, column=ytd_cr_col).fill = FORMULA_FILL

        MTB_ROWS[key] = row
        row += 1

    # Instructions
    row += 2
    ws.cell(row=row, column=1, value="📌 HOW TO USE:").font = Font(name="Calibri", bold=True, size=11, color=RED)
    ws.cell(row=row + 1, column=1, value="• Each month, export your cumulative Trial Balance from Tally").font = Font(name="Calibri", size=10, color=GRAY)
    ws.cell(row=row + 2, column=1, value="• Enter the cumulative Dr/Cr values for each account in that month's columns").font = Font(name="Calibri", size=10, color=GRAY)
    ws.cell(row=row + 3, column=1, value="• The P&L_Monthly sheet auto-calculates monthly movements (Current - Previous month)").font = Font(name="Calibri", size=10, color=GRAY)
    ws.cell(row=row + 4, column=1, value="• YTD columns link to the Trial_Balance sheet. Update TB sheet for latest YTD.").font = Font(name="Calibri", size=10, color=GRAY)
    ws.cell(row=row + 5, column=1, value="• Yellow cells = Input. Gray cells = Formulas (do not edit).").font = Font(name="Calibri", size=10, color=GRAY)

    # Column widths
    set_widths(ws, {'A': 30, 'B': 12})
    for i in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws.freeze_panes = 'C6'
    return ws


# ── Sheet 3: P&L (YTD from Trial Balance) ───────────────────────────────────

def create_pnl_ytd_sheet(wb):
    """Create YTD Profit & Loss linked to Trial_Balance."""
    ws = wb.create_sheet("P&L_YTD")
    S = 'Trial_Balance'

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP", "Profit & Loss Statement — Year to Date")

    # Period reference
    ws.cell(row=row, column=1, value="Period:").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=row, column=2, value=f"='{S}'!A8").font = Font(name="Calibri", size=10, color="4472C4")
    row += 2

    headers = ['Particulars', 'Amount (₹)', 'Notes / Tally Reference']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header(ws, row, 3)
    row += 1

    def pnl_row(ws, row, label, formula, note='', is_total=False, is_subtotal=False, indent=0, is_pct=False):
        """Write a P&L line with formula."""
        prefix = '  ' * indent
        ws.cell(row=row, column=1, value=f"{prefix}{label}")

        cell = ws.cell(row=row, column=2)
        cell.value = formula
        if is_pct:
            cell.number_format = PCT_FMT
        else:
            cell.number_format = INR_FMT

        ws.cell(row=row, column=3, value=note).font = Font(name="Calibri", size=9, color=GRAY)

        if is_total:
            apply_total_row(ws, row, 3)
        elif is_subtotal:
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=11)
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = Border(
                    top=Side(style='thin', color=NAVY),
                    bottom=Side(style='thin', color=NAVY),
                )
        else:
            ws.cell(row=row, column=1).font = Font(name="Calibri", size=10)
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = THIN_BORDER

        return row + 1

    # ── REVENUE ──
    ws.cell(row=row, column=1, value='REVENUE FROM OPERATIONS')
    apply_subheader(ws, row, 3)
    row += 1

    rev_total_row = row  # Will use this for GP calculation
    row = pnl_row(ws, row, 'Hospitality Services',
                  f"='{S}'!C{TB_ROWS.get('hospitality_revenue', 1)}",
                  f"TB Row {TB_ROWS.get('hospitality_revenue', '?')}", indent=1)
    row = pnl_row(ws, row, 'Studio Rental Services',
                  f"='{S}'!C{TB_ROWS.get('studio_revenue', 1)}",
                  f"TB Row {TB_ROWS.get('studio_revenue', '?')}", indent=1)

    sales_row = row
    row = pnl_row(ws, row, 'Total Revenue',
                  f"='{S}'!C{TB_ROWS.get('sales_accounts', 1)}",
                  f"TB Row {TB_ROWS.get('sales_accounts', '?')} (Credit)", is_subtotal=True)
    row += 1

    # ── COST OF GOODS SOLD ──
    ws.cell(row=row, column=1, value='COST OF GOODS SOLD')
    apply_subheader(ws, row, 3)
    row += 1

    row = pnl_row(ws, row, 'Thota Kitchen',
                  f"='{S}'!B{TB_ROWS.get('purchase_kitchen', 1)}",
                  'Food & Kitchen materials', indent=1)
    row = pnl_row(ws, row, 'Thota Decor',
                  f"='{S}'!B{TB_ROWS.get('purchase_decor', 1)}",
                  'Decoration materials', indent=1)
    row = pnl_row(ws, row, 'Studio Purchases',
                  f"='{S}'!B{TB_ROWS.get('purchase_studio', 1)}",
                  'Studio materials', indent=1)

    cogs_row = row
    row = pnl_row(ws, row, 'Total COGS',
                  f"='{S}'!B{TB_ROWS.get('purchase_accounts', 1)}",
                  f"TB Row {TB_ROWS.get('purchase_accounts', '?')} (Debit)", is_subtotal=True)
    row += 1

    # ── DIRECT EXPENSES ──
    row = pnl_row(ws, row, 'Fuel Expenses',
                  f"='{S}'!B{TB_ROWS.get('fuel_expenses', 1)}",
                  '', indent=1)
    row = pnl_row(ws, row, 'Transportation Expenses',
                  f"='{S}'!B{TB_ROWS.get('transportation', 1)}",
                  '', indent=1)
    direct_row = row
    row = pnl_row(ws, row, 'Total Direct Expenses',
                  f"='{S}'!B{TB_ROWS.get('direct_expenses', 1)}",
                  f"TB Row {TB_ROWS.get('direct_expenses', '?')}", is_subtotal=True)
    row += 1

    # ── GROSS PROFIT ──
    gp_row = row
    row = pnl_row(ws, row, 'GROSS PROFIT',
                  f"=B{sales_row}-B{cogs_row}-B{direct_row}",
                  'Revenue - COGS - Direct Expenses', is_total=True)
    gp_pct_row = row
    row = pnl_row(ws, row, 'Gross Profit %',
                  f"=IF(B{sales_row}>0,B{gp_row}/B{sales_row},0)",
                  '', is_pct=True)
    row += 1

    # ── INDIRECT INCOME ──
    ii_row = row
    row = pnl_row(ws, row, 'Indirect Income (Discount Received)',
                  f"='{S}'!C{TB_ROWS.get('indirect_incomes', 1)}",
                  f"TB Row {TB_ROWS.get('indirect_incomes', '?')}")
    row += 1

    # ── OPERATING EXPENSES ──
    ws.cell(row=row, column=1, value='OPERATING EXPENSES (OPEX)')
    apply_subheader(ws, row, 3)
    row += 1

    opex_items = [
        ('Administrative Overheads', 'admin_overheads'),
        ('Finance Cost', 'finance_cost'),
        ('HR Expenses', 'hr_expenses'),
        ('Audit / Legal Professional Fee', 'audit_legal'),
        ('Employer PT', 'employer_pt'),
        ('Marketing & Ads', 'marketing'),
        ('Professional Charges', 'professional_charges'),
        ('Rates & Taxes', 'rates_taxes'),
        ('Repairs & Maintenance', 'repairs_maintenance'),
        ('Round Off', 'round_off'),
    ]

    for label, key in opex_items:
        tb_row = TB_ROWS.get(key, 1)
        # Most are debit. Rates & Taxes and Round Off are credit.
        if key in ('rates_taxes', 'round_off'):
            formula = f"='{S}'!B{tb_row}-'{S}'!C{tb_row}"
        else:
            formula = f"='{S}'!B{tb_row}"
        row = pnl_row(ws, row, label, formula, f"TB Row {tb_row}", indent=1)

    opex_total_row = row
    # Total OPEX = TB Indirect Expenses Debit - Credit
    row = pnl_row(ws, row, 'Total OPEX',
                  f"='{S}'!B{TB_ROWS.get('indirect_expenses', 1)}-'{S}'!C{TB_ROWS.get('indirect_expenses', 1)}",
                  'Net Indirect Expenses', is_subtotal=True)
    row += 1

    # ── EBITDA ──
    ebitda_row = row
    row = pnl_row(ws, row, 'EBITDA',
                  f"=B{gp_row}+B{ii_row}-B{opex_total_row}",
                  'Gross Profit + Other Income - OPEX', is_total=True)
    ebitda_pct_row = row
    row = pnl_row(ws, row, 'EBITDA %',
                  f"=IF(B{sales_row}>0,B{ebitda_row}/B{sales_row},0)",
                  '', is_pct=True)
    row += 1

    # ── DEPRECIATION (estimated) ──
    dep_row = row
    row = pnl_row(ws, row, 'Depreciation (Estimated)',
                  f"='{S}'!B{TB_ROWS.get('fixed_assets', 1)}*0.15",
                  '15% of Fixed Assets (adjust as needed)')
    row += 1

    # ── PBT ──
    pbt_row = row
    row = pnl_row(ws, row, 'PROFIT BEFORE TAX (PBT)',
                  f"=B{ebitda_row}-B{dep_row}",
                  'EBITDA - Depreciation', is_total=True)
    row += 1

    # ── TAX ──
    tax_row = row
    row = pnl_row(ws, row, 'Income Tax (LLP - 30%)',
                  f"=IF(B{pbt_row}>0,B{pbt_row}*0.30,0)",
                  '30% on PBT (LLP tax rate - adjust as needed)')
    row += 1

    # ── PAT ──
    pat_row = row
    row = pnl_row(ws, row, 'PROFIT AFTER TAX (PAT)',
                  f"=B{pbt_row}-B{tax_row}",
                  'PBT - Tax', is_total=True)
    pat_pct_row = row
    row = pnl_row(ws, row, 'PAT %',
                  f"=IF(B{sales_row}>0,B{pat_row}/B{sales_row},0)",
                  '', is_pct=True)

    # Store key row references for other sheets
    ws._pnl_refs = {
        'sales': sales_row,
        'cogs': cogs_row,
        'direct': direct_row,
        'gp': gp_row,
        'indirect_income': ii_row,
        'opex': opex_total_row,
        'ebitda': ebitda_row,
        'dep': dep_row,
        'pbt': pbt_row,
        'tax': tax_row,
        'pat': pat_row,
    }

    set_widths(ws, {'A': 38, 'B': 22, 'C': 35})
    ws.sheet_properties.tabColor = "4472C4"
    return ws


# ── Sheet 4: P&L Monthly ────────────────────────────────────────────────────

def create_pnl_monthly_sheet(wb):
    """Create monthly P&L that derives movement from Monthly_TB cumulative figures."""
    ws = wb.create_sheet("P&L_Monthly")

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP", "Monthly & Quarterly P&L — FY 2025-26")

    # Explanation
    ws.cell(row=row, column=1, value="Monthly movement = Current month cumulative - Previous month cumulative (from Monthly_TB)").font = Font(name="Calibri", italic=True, size=9, color=GRAY)
    row += 2

    # Headers: Particulars | Apr | May | ... | Mar | Q1 | Q2 | Q3 | Q4 | FY
    headers = ['Particulars'] + MONTHS + ['Q1', 'Q2', 'Q3', 'Q4', 'FY Total']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header(ws, row, len(headers))
    header_row = row
    row += 1

    M = 'Monthly_TB'

    # P&L accounts (Revenue, COGS are P&L items - cumulative)
    pnl_accounts = [
        ('REVENUE', None, 'subheader'),
        ('Hospitality Services', 'hospitality_revenue', 'cr'),
        ('Studio Rental Services', 'studio_revenue', 'cr'),
        ('Total Revenue', 'sales_accounts', 'cr_total'),
        ('', None, 'blank'),
        ('COST OF GOODS SOLD', None, 'subheader'),
        ('Thota Kitchen', 'purchase_kitchen', 'dr'),
        ('Thota Decor', 'purchase_decor', 'dr'),
        ('Studio Purchases', 'purchase_studio', 'dr'),
        ('Total COGS', 'purchase_accounts', 'dr_total'),
        ('', None, 'blank'),
        ('DIRECT EXPENSES', None, 'subheader'),
        ('Fuel Expenses', 'fuel_expenses', 'dr'),
        ('Transportation', 'transportation', 'dr'),
        ('Total Direct Expenses', 'direct_expenses', 'dr_total'),
        ('', None, 'blank'),
        # Gross Profit will be calculated
        ('', None, 'gross_profit'),
        ('', None, 'blank'),
        ('OPERATING EXPENSES', None, 'subheader'),
        ('Admin Overheads', 'admin_overheads', 'dr'),
        ('Finance Cost', 'finance_cost', 'dr'),
        ('HR Expenses', 'hr_expenses', 'dr'),
        ('Marketing & Ads', 'marketing', 'dr'),
        ('Professional Charges', 'professional_charges', 'dr'),
        ('Repairs & Maintenance', 'repairs_maintenance', 'dr'),
        ('Other OPEX', None, 'opex_other'),
        ('Total OPEX', 'indirect_expenses', 'dr_net_total'),
        ('', None, 'blank'),
        ('', None, 'ebitda'),
    ]

    # Track key rows for GP, EBITDA calculations
    row_refs = {}

    for label, key, style in pnl_accounts:
        if style == 'blank':
            row += 1
            continue

        if style == 'subheader':
            ws.cell(row=row, column=1, value=label)
            apply_subheader(ws, row, len(headers))
            row += 1
            continue

        if style == 'gross_profit':
            ws.cell(row=row, column=1, value='GROSS PROFIT')
            rev_row = row_refs.get('sales_accounts')
            cogs_row = row_refs.get('purchase_accounts')
            direct_row = row_refs.get('direct_expenses')
            for c in range(2, len(headers) + 1):
                col_l = get_column_letter(c)
                formula = f"={col_l}{rev_row}-{col_l}{cogs_row}-{col_l}{direct_row}"
                ws.cell(row=row, column=c, value=formula).number_format = INR_FMT
            apply_total_row(ws, row, len(headers))
            row_refs['gross_profit'] = row
            row += 1
            # GP %
            ws.cell(row=row, column=1, value='Gross Profit %').font = Font(name="Calibri", italic=True, color="4472C4")
            gp_r = row_refs['gross_profit']
            for c in range(2, len(headers) + 1):
                col_l = get_column_letter(c)
                formula = f"=IF({col_l}{rev_row}>0,{col_l}{gp_r}/{col_l}{rev_row},0)"
                cell = ws.cell(row=row, column=c, value=formula)
                cell.number_format = PCT_FMT
                cell.font = Font(name="Calibri", italic=True, color="4472C4")
            row += 1
            continue

        if style == 'ebitda':
            ws.cell(row=row, column=1, value='EBITDA')
            gp_r = row_refs.get('gross_profit')
            opex_r = row_refs.get('indirect_expenses')
            for c in range(2, len(headers) + 1):
                col_l = get_column_letter(c)
                formula = f"={col_l}{gp_r}-{col_l}{opex_r}"
                ws.cell(row=row, column=c, value=formula).number_format = INR_FMT
            apply_total_row(ws, row, len(headers))
            row_refs['ebitda'] = row
            row += 1
            # EBITDA %
            ws.cell(row=row, column=1, value='EBITDA %').font = Font(name="Calibri", italic=True, color=GREEN)
            rev_r = row_refs['sales_accounts']
            ebitda_r = row_refs['ebitda']
            for c in range(2, len(headers) + 1):
                col_l = get_column_letter(c)
                formula = f"=IF({col_l}{rev_r}>0,{col_l}{ebitda_r}/{col_l}{rev_r},0)"
                cell = ws.cell(row=row, column=c, value=formula)
                cell.number_format = PCT_FMT
                cell.font = Font(name="Calibri", italic=True, color=GREEN)
            row += 1
            continue

        if style == 'opex_other':
            # Other OPEX = Total OPEX - named OPEX items
            ws.cell(row=row, column=1, value='  Other OPEX').font = DETAIL_FONT
            named_keys = ['admin_overheads', 'finance_cost', 'hr_expenses', 'marketing',
                          'professional_charges', 'repairs_maintenance']
            for c in range(2, len(headers) + 1):
                col_l = get_column_letter(c)
                named_refs = '+'.join([f"{col_l}{row_refs[k]}" for k in named_keys if k in row_refs])
                # Use a placeholder formula - will be overwritten after we know total OPEX row
                ws.cell(row=row, column=c, value=0).number_format = INR_FMT
                ws.cell(row=row, column=c).border = THIN_BORDER
            row_refs['opex_other'] = row
            row += 1
            continue

        # Standard account row
        ws.cell(row=row, column=1, value=label)
        mtb_row = MTB_ROWS.get(key, 1)

        is_credit = style.startswith('cr')
        is_total = style.endswith('total')
        is_net = 'net' in style

        # Monthly columns (movement = current cumulative - previous cumulative)
        for i in range(12):
            c = i + 2  # Column B=2 for Apr, C=3 for May, etc.
            dr_col = get_column_letter(3 + i * 2)  # Monthly_TB Dr column
            cr_col = get_column_letter(4 + i * 2)  # Monthly_TB Cr column

            if i == 0:
                # First month: movement = cumulative value itself
                if is_credit:
                    formula = f"='{M}'!{cr_col}{mtb_row}"
                elif is_net:
                    formula = f"='{M}'!{dr_col}{mtb_row}-'{M}'!{cr_col}{mtb_row}"
                else:
                    formula = f"='{M}'!{dr_col}{mtb_row}"
            else:
                # Subsequent months: movement = current cumulative - previous cumulative
                prev_dr = get_column_letter(3 + (i - 1) * 2)
                prev_cr = get_column_letter(4 + (i - 1) * 2)
                if is_credit:
                    formula = f"='{M}'!{cr_col}{mtb_row}-'{M}'!{prev_cr}{mtb_row}"
                elif is_net:
                    formula = f"=('{M}'!{dr_col}{mtb_row}-'{M}'!{cr_col}{mtb_row})-('{M}'!{prev_dr}{mtb_row}-'{M}'!{prev_cr}{mtb_row})"
                else:
                    formula = f"='{M}'!{dr_col}{mtb_row}-'{M}'!{prev_dr}{mtb_row}"

            cell = ws.cell(row=row, column=c, value=formula)
            cell.number_format = INR_FMT
            cell.border = THIN_BORDER

        # Quarterly columns: Q1 = Sum(Apr, May, Jun), etc.
        for qi, (qname, month_indices) in enumerate(QUARTERS.items()):
            q_col = 14 + qi  # Columns 14, 15, 16, 17
            month_cols = [get_column_letter(m + 2) for m in month_indices]
            formula = '=' + '+'.join([f"{mc}{row}" for mc in month_cols])
            ws.cell(row=row, column=q_col, value=formula).number_format = INR_FMT
            ws.cell(row=row, column=q_col).border = THIN_BORDER

        # FY Total = Sum of monthly
        fy_col = 18
        month_col_letters = [get_column_letter(m + 2) for m in range(12)]
        formula = '=' + '+'.join([f"{mc}{row}" for mc in month_col_letters])
        ws.cell(row=row, column=fy_col, value=formula).number_format = INR_FMT
        ws.cell(row=row, column=fy_col).border = THIN_BORDER

        if is_total:
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=row, column=fy_col).font = Font(name="Calibri", bold=True, size=11)
        else:
            ws.cell(row=row, column=1).font = DETAIL_FONT

        row_refs[key] = row
        row += 1

    # Now fix the "Other OPEX" row formula
    if 'opex_other' in row_refs and 'indirect_expenses' in row_refs:
        other_row = row_refs['opex_other']
        total_row = row_refs['indirect_expenses']
        named_keys = ['admin_overheads', 'finance_cost', 'hr_expenses', 'marketing',
                      'professional_charges', 'repairs_maintenance']
        for c in range(2, len(headers) + 1):
            col_l = get_column_letter(c)
            named_refs = '-'.join([f"{col_l}{row_refs[k]}" for k in named_keys if k in row_refs])
            formula = f"={col_l}{total_row}-{named_refs}"
            ws.cell(row=other_row, column=c, value=formula).number_format = INR_FMT

    set_widths(ws, {'A': 28})
    for i in range(2, 19):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws.freeze_panes = 'B6'
    ws.sheet_properties.tabColor = "4472C4"
    return ws


# ── Sheet 5: Balance Sheet ───────────────────────────────────────────────────

def create_balance_sheet_sheet(wb):
    """Create Balance Sheet linked to Trial_Balance."""
    ws = wb.create_sheet("Balance_Sheet")
    S = 'Trial_Balance'

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP", "Balance Sheet")
    ws.cell(row=row, column=1, value="Period:").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=row, column=2, value=f"='{S}'!A8").font = Font(name="Calibri", size=10, color="4472C4")
    row += 2

    headers = ['Particulars', 'Sub-Total (₹)', 'Total (₹)']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header(ws, row, 3)
    row += 1

    def bs_line(ws, row, label, formula_sub=None, formula_total=None, is_group=False, is_total=False, indent=0):
        prefix = '  ' * indent
        ws.cell(row=row, column=1, value=f"{prefix}{label}")

        if formula_sub:
            ws.cell(row=row, column=2, value=formula_sub).number_format = INR_FMT
        if formula_total:
            ws.cell(row=row, column=3, value=formula_total).number_format = INR_FMT

        if is_total:
            apply_total_row(ws, row, 3)
        elif is_group:
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color=NAVY)
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = PatternFill(start_color="E8EDF4", end_color="E8EDF4", fill_type="solid")
                ws.cell(row=row, column=c).border = THIN_BORDER
        else:
            ws.cell(row=row, column=1).font = DETAIL_FONT
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = THIN_BORDER

        return row + 1

    # ═══ LIABILITIES ═══
    ws.cell(row=row, column=1, value='LIABILITIES & EQUITY')
    apply_subheader(ws, row, 3)
    row += 1

    # Capital Account
    row = bs_line(ws, row, 'Capital Account', None,
                  f"='{S}'!C{TB_ROWS.get('capital_account', 1)}", is_group=True)
    row = bs_line(ws, row, "Partner's Capital",
                  f"='{S}'!C{TB_ROWS.get('partners_capital', 1)}", indent=1)
    row += 1

    # Reserves & Surplus (P&L Account)
    pnl_acct_row = TB_ROWS.get('pnl_account', 1)
    row = bs_line(ws, row, 'Reserves & Surplus (P&L)', None,
                  f"=-'{S}'!B{pnl_acct_row}", is_group=True)
    row += 1

    # Loans
    row = bs_line(ws, row, 'Loans (Liability)', None,
                  f"='{S}'!C{TB_ROWS.get('loans_liability', 1)}", is_group=True)
    row = bs_line(ws, row, "Partner's Loan - Akshatha",
                  f"='{S}'!C{TB_ROWS.get('loan_akshatha', 1)}", indent=1)
    row = bs_line(ws, row, "Partner's Loan - Tejas",
                  f"='{S}'!C{TB_ROWS.get('loan_tejas', 1)}", indent=1)
    row = bs_line(ws, row, "TVS Credit Services",
                  f"='{S}'!C{TB_ROWS.get('tvs_credit', 1)}", indent=1)
    row += 1

    # Current Liabilities
    cl_row = TB_ROWS.get('current_liabilities', 1)
    row = bs_line(ws, row, 'Current Liabilities', None,
                  f"='{S}'!C{cl_row}-'{S}'!B{cl_row}", is_group=True)
    cl_detail = [
        ('Duties & Taxes', 'duties_taxes'),
        ('Sundry Creditors', 'sundry_creditors'),
        ('Salary Payable', 'salary_payable'),
        ('Reimbursements', 'reimbursements'),
    ]
    for label, key in cl_detail:
        r = TB_ROWS.get(key, 1)
        row = bs_line(ws, row, label,
                      f"='{S}'!C{r}-'{S}'!B{r}", indent=1)
    row += 1

    # Retained Earnings from P&L_YTD
    pnl_ws = wb['P&L_YTD']
    pat_row = pnl_ws._pnl_refs.get('pat', 1)
    re_row = row
    row = bs_line(ws, row, 'Current Year Profit (from P&L)', None,
                  f"=P&L_YTD!B{pat_row}", is_group=True)
    row += 1

    # TOTAL LIABILITIES
    total_liab_row = row
    # Sum all liability groups
    cap_ref = f"C{row - (row - 6)}"  # We'll use formula summing specific cells
    row = bs_line(ws, row, 'TOTAL LIABILITIES & EQUITY', None,
                  f"='{S}'!C{TB_ROWS.get('capital_account', 1)}"
                  f"-'{S}'!B{pnl_acct_row}"
                  f"+'{S}'!C{TB_ROWS.get('loans_liability', 1)}"
                  f"+'{S}'!C{cl_row}-'{S}'!B{cl_row}"
                  f"+P&L_YTD!B{pat_row}",
                  is_total=True)
    row += 2

    # ═══ ASSETS ═══
    ws.cell(row=row, column=1, value='ASSETS')
    apply_subheader(ws, row, 3)
    row += 1

    # Fixed Assets
    row = bs_line(ws, row, 'Fixed Assets (Non-Current)', None,
                  f"='{S}'!B{TB_ROWS.get('fixed_assets', 1)}", is_group=True)
    row = bs_line(ws, row, 'Intangible Assets',
                  f"='{S}'!B{TB_ROWS.get('intangible_assets', 1)}", indent=1)
    row = bs_line(ws, row, 'Tangible Assets',
                  f"='{S}'!B{TB_ROWS.get('tangible_assets', 1)}", indent=1)
    row += 1

    # Current Assets
    ca_row = TB_ROWS.get('current_assets', 1)
    row = bs_line(ws, row, 'Current Assets', None,
                  f"='{S}'!B{ca_row}-'{S}'!C{ca_row}", is_group=True)

    ca_detail = [
        ('Loans & Advances', 'loans_advances_asset'),
        ('Sundry Debtors', 'sundry_debtors'),
        ('Cash-in-Hand', 'cash_in_hand'),
        ('Bank Accounts', 'bank_accounts'),
    ]
    for label, key in ca_detail:
        r = TB_ROWS.get(key, 1)
        row = bs_line(ws, row, label,
                      f"='{S}'!B{r}-'{S}'!C{r}", indent=1)
    row += 1

    # TOTAL ASSETS
    total_asset_row = row
    row = bs_line(ws, row, 'TOTAL ASSETS', None,
                  f"='{S}'!B{TB_ROWS.get('fixed_assets', 1)}"
                  f"+'{S}'!B{ca_row}-'{S}'!C{ca_row}",
                  is_total=True)
    row += 2

    # Variance check
    ws.cell(row=row, column=1, value="Balance Check (Assets - Liabilities):").font = Font(name="Calibri", bold=True, size=10, color=RED)
    ws.cell(row=row, column=3, value=f"=C{total_asset_row}-C{total_liab_row}").number_format = INR_FMT
    ws.cell(row=row, column=3).font = Font(name="Calibri", bold=True, size=11, color=RED)

    # Store refs
    ws._bs_refs = {
        'total_liabilities': total_liab_row,
        'total_assets': total_asset_row,
    }

    set_widths(ws, {'A': 38, 'B': 20, 'C': 20})
    ws.sheet_properties.tabColor = "548235"
    return ws


# ── Sheet 6: Cash Flow ──────────────────────────────────────────────────────

def create_cashflow_sheet(wb):
    """Create Cash Flow statement derived from P&L and Balance Sheet."""
    ws = wb.create_sheet("Cash_Flow")
    S = 'Trial_Balance'

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP", "Cash Flow Statement (Indirect Method)")
    ws.cell(row=row, column=1, value="Period:").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=row, column=2, value=f"='{S}'!A8").font = Font(name="Calibri", size=10, color="4472C4")
    row += 2

    headers = ['Particulars', 'Amount (₹)', 'Reference']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header(ws, row, 3)
    row += 1

    pnl_ws = wb['P&L_YTD']
    pnl_refs = pnl_ws._pnl_refs

    def cf_line(ws, row, label, formula, note='', is_total=False, is_sub=False, indent=0):
        prefix = '  ' * indent
        ws.cell(row=row, column=1, value=f"{prefix}{label}")
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = INR_FMT
        ws.cell(row=row, column=3, value=note).font = Font(name="Calibri", size=9, color=GRAY)

        if is_total:
            apply_total_row(ws, row, 3)
        elif is_sub:
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True)
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = Border(
                    top=Side(style='thin', color=NAVY),
                    bottom=Side(style='thin', color=NAVY))
        else:
            ws.cell(row=row, column=1).font = DETAIL_FONT
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = THIN_BORDER
        return row + 1

    # A. OPERATING ACTIVITIES
    ws.cell(row=row, column=1, value='A. CASH FROM OPERATING ACTIVITIES')
    apply_subheader(ws, row, 3)
    row += 1

    pat_ref = pnl_refs['pat']
    row = cf_line(ws, row, 'Net Profit (PAT)',
                  f"=P&L_YTD!B{pat_ref}", 'From P&L')

    # Add back non-cash items
    dep_ref = pnl_refs['dep']
    row = cf_line(ws, row, 'Add: Depreciation',
                  f"=P&L_YTD!B{dep_ref}", 'Non-cash expense', indent=1)

    fc_row = TB_ROWS.get('finance_cost', 1)
    row = cf_line(ws, row, 'Add: Finance Cost',
                  f"='{S}'!B{fc_row}", 'To be reclassified', indent=1)

    op_profit_row = row
    row = cf_line(ws, row, 'Operating Profit before WC changes',
                  f"=B{row-3}+B{row-2}+B{row-1}", '', is_sub=True)
    row += 1

    # Working Capital changes
    ws.cell(row=row, column=1, value='Working Capital Changes:').font = Font(name="Calibri", bold=True, italic=True, size=10, color=NAVY)
    row += 1

    debtors_row = TB_ROWS.get('sundry_debtors', 1)
    row = cf_line(ws, row, 'Decrease/(Increase) in Sundry Debtors',
                  f"=-('{S}'!B{debtors_row}-'{S}'!C{debtors_row})",
                  'Negative = more debtors = cash outflow', indent=1)

    creditors_row = TB_ROWS.get('sundry_creditors', 1)
    row = cf_line(ws, row, 'Increase/(Decrease) in Sundry Creditors',
                  f"='{S}'!C{creditors_row}-'{S}'!B{creditors_row}",
                  'Positive = more creditors = cash saved', indent=1)

    salary_row = TB_ROWS.get('salary_payable', 1)
    row = cf_line(ws, row, 'Increase/(Decrease) in Salary Payable',
                  f"='{S}'!C{salary_row}-'{S}'!B{salary_row}",
                  '', indent=1)

    duties_row = TB_ROWS.get('duties_taxes', 1)
    row = cf_line(ws, row, 'Increase/(Decrease) in Duties & Taxes',
                  f"='{S}'!C{duties_row}-'{S}'!B{duties_row}",
                  '', indent=1)
    row += 1

    cash_ops_row = row
    row = cf_line(ws, row, 'Net Cash from Operating Activities',
                  f"=SUM(B{op_profit_row}:B{row-2})",
                  '', is_total=True)
    row += 1

    # B. INVESTING ACTIVITIES
    ws.cell(row=row, column=1, value='B. CASH FROM INVESTING ACTIVITIES')
    apply_subheader(ws, row, 3)
    row += 1

    fa_row = TB_ROWS.get('fixed_assets', 1)
    row = cf_line(ws, row, 'Purchase of Fixed Assets',
                  f"=-'{S}'!B{fa_row}",
                  'Investment in FA (outflow)', indent=1)

    la_row = TB_ROWS.get('loans_advances_asset', 1)
    row = cf_line(ws, row, 'Loans & Advances Given',
                  f"=-'{S}'!B{la_row}",
                  '', indent=1)

    cash_inv_row = row
    row = cf_line(ws, row, 'Net Cash from Investing Activities',
                  f"=B{row-2}+B{row-1}",
                  '', is_total=True)
    row += 1

    # C. FINANCING ACTIVITIES
    ws.cell(row=row, column=1, value='C. CASH FROM FINANCING ACTIVITIES')
    apply_subheader(ws, row, 3)
    row += 1

    cap_row = TB_ROWS.get('capital_account', 1)
    row = cf_line(ws, row, "Partner's Capital",
                  f"='{S}'!C{cap_row}",
                  'Equity contribution', indent=1)

    loans_row = TB_ROWS.get('loans_liability', 1)
    row = cf_line(ws, row, 'Loans Received / (Repaid)',
                  f"='{S}'!C{loans_row}",
                  '', indent=1)

    row = cf_line(ws, row, 'Finance Cost Paid',
                  f"=-'{S}'!B{fc_row}",
                  'Outflow', indent=1)

    cash_fin_row = row
    row = cf_line(ws, row, 'Net Cash from Financing Activities',
                  f"=B{row-3}+B{row-2}+B{row-1}",
                  '', is_total=True)
    row += 2

    # NET CHANGE
    net_row = row
    row = cf_line(ws, row, 'NET INCREASE/(DECREASE) IN CASH',
                  f"=B{cash_ops_row}+B{cash_inv_row}+B{cash_fin_row}",
                  'Operating + Investing + Financing', is_total=True)
    row += 1

    # Cash Position
    cash_row = TB_ROWS.get('cash_in_hand', 1)
    bank_row = TB_ROWS.get('bank_accounts', 1)
    row = cf_line(ws, row, 'Cash & Bank Balance (as per TB)',
                  f"='{S}'!B{cash_row}+'{S}'!B{bank_row}",
                  'Cash + Bank Accounts from TB', is_sub=True)

    set_widths(ws, {'A': 42, 'B': 22, 'C': 40})
    ws.sheet_properties.tabColor = "ED7D31"
    return ws


# ── Sheet 7: Revenue Summary ────────────────────────────────────────────────

def create_revenue_summary_sheet(wb):
    """Create revenue analysis with segment-wise breakdown."""
    ws = wb.create_sheet("Revenue_Summary")
    S = 'Trial_Balance'

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP", "Revenue Summary — Segment-wise Analysis")
    row += 1

    headers = ['Revenue Segment', 'Amount (₹)', '% of Revenue', 'Avg Monthly (₹)']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header(ws, row, 4)
    row += 1

    # Period (# months)
    ws.cell(row=row, column=1, value="Number of months in period:").font = Font(name="Calibri", bold=True, size=10)
    months_cell = ws.cell(row=row, column=2, value=10)  # Apr-Jan = 10 months
    months_cell.fill = INPUT_FILL
    months_cell.border = INPUT_BORDER
    months_cell.font = Font(name="Calibri", bold=True, size=11, color=ORANGE)
    ws.cell(row=row, column=3, value="← Edit this when TB period changes").font = Font(name="Calibri", size=9, color=GRAY, italic=True)
    months_ref_row = row
    row += 2

    # Hospitality
    hosp_row = row
    ws.cell(row=row, column=1, value='Hospitality Services').font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=2, value=f"='{S}'!C{TB_ROWS.get('hospitality_revenue', 1)}").number_format = INR_FMT
    sales_row_ref = TB_ROWS.get('sales_accounts', 1)
    ws.cell(row=row, column=3, value=f"=IF('{S}'!C{sales_row_ref}>0,B{row}/'{S}'!C{sales_row_ref},0)").number_format = PCT_FMT
    ws.cell(row=row, column=4, value=f"=B{row}/B{months_ref_row}").number_format = INR_FMT
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

    # Studio
    studio_row = row
    ws.cell(row=row, column=1, value='Studio Services').font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=2, value=f"='{S}'!C{TB_ROWS.get('studio_revenue', 1)}").number_format = INR_FMT
    ws.cell(row=row, column=3, value=f"=IF('{S}'!C{sales_row_ref}>0,B{row}/'{S}'!C{sales_row_ref},0)").number_format = PCT_FMT
    ws.cell(row=row, column=4, value=f"=B{row}/B{months_ref_row}").number_format = INR_FMT
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

    # Total
    total_rev_row = row
    ws.cell(row=row, column=1, value='TOTAL REVENUE')
    ws.cell(row=row, column=2, value=f"='{S}'!C{sales_row_ref}").number_format = INR_FMT
    ws.cell(row=row, column=3, value=f"=IF(B{row}>0,1,0)").number_format = PCT_FMT
    ws.cell(row=row, column=4, value=f"=B{row}/B{months_ref_row}").number_format = INR_FMT
    apply_total_row(ws, row, 4)
    row += 2

    # Revenue Unit Economics
    ws.cell(row=row, column=1, value='UNIT ECONOMICS')
    apply_subheader(ws, row, 4)
    row += 1

    pnl_ws = wb['P&L_YTD']
    pnl_refs = pnl_ws._pnl_refs

    unit_items = [
        ('Revenue per Month', f"=B{total_rev_row}/B{months_ref_row}", INR_FMT, 'Total Revenue / Months'),
        ('COGS per Month', f"=P&L_YTD!B{pnl_refs['cogs']}/B{months_ref_row}", INR_FMT, 'Total COGS / Months'),
        ('Gross Profit per Month', f"=P&L_YTD!B{pnl_refs['gp']}/B{months_ref_row}", INR_FMT, 'GP / Months'),
        ('OPEX per Month', f"=P&L_YTD!B{pnl_refs['opex']}/B{months_ref_row}", INR_FMT, 'OPEX / Months'),
        ('EBITDA per Month', f"=P&L_YTD!B{pnl_refs['ebitda']}/B{months_ref_row}", INR_FMT, 'EBITDA / Months'),
        ('PAT per Month', f"=P&L_YTD!B{pnl_refs['pat']}/B{months_ref_row}", INR_FMT, 'PAT / Months'),
    ]

    for label, formula, fmt, note in unit_items:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=2, value=formula).number_format = fmt
        ws.cell(row=row, column=4, value=note).font = Font(name="Calibri", size=9, color=GRAY)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Revenue Pie Chart
    row += 2
    ws.cell(row=row, column=4, value="Segment").font = HEADER_FONT
    ws.cell(row=row, column=4).fill = HEADER_FILL
    ws.cell(row=row, column=5, value="Amount").font = HEADER_FONT
    ws.cell(row=row, column=5).fill = HEADER_FILL

    ws.cell(row=row + 1, column=4, value="Hospitality")
    ws.cell(row=row + 1, column=5, value=f"=B{hosp_row}").number_format = INR_FMT
    ws.cell(row=row + 2, column=4, value="Studio")
    ws.cell(row=row + 2, column=5, value=f"=B{studio_row}").number_format = INR_FMT

    pie = PieChart()
    pie.title = "Revenue Mix"
    pie.style = 10
    pie.width = 16
    pie.height = 12

    cats = Reference(ws, min_col=4, min_row=row + 1, max_row=row + 2)
    vals = Reference(ws, min_col=5, min_row=row, max_row=row + 2)
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, f"D{hosp_row}")

    set_widths(ws, {'A': 30, 'B': 22, 'C': 16, 'D': 18, 'E': 18})
    ws.sheet_properties.tabColor = "FFC000"
    return ws


# ── Sheet 8: OPEX Schedule ──────────────────────────────────────────────────

def create_opex_schedule_sheet(wb):
    """Create detailed OPEX breakdown from Trial Balance."""
    ws = wb.create_sheet("OPEX_Schedule")
    S = 'Trial_Balance'

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP", "Operating Expenses Schedule — Detailed")
    ws.cell(row=row, column=1, value="Period:").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=row, column=2, value=f"='{S}'!A8").font = Font(name="Calibri", size=10, color="4472C4")
    row += 2

    headers = ['Expense Category', 'Amount (₹)', '% of Revenue', '% of Total OPEX']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header(ws, row, 4)
    row += 1

    sales_ref = f"'{S}'!C{TB_ROWS.get('sales_accounts', 1)}"
    opex_ref_dr = f"'{S}'!B{TB_ROWS.get('indirect_expenses', 1)}"
    opex_ref_cr = f"'{S}'!C{TB_ROWS.get('indirect_expenses', 1)}"

    opex_categories = [
        ('Administrative Overheads', 'admin_overheads', [
            ('Local Conveyance', 100), ('Office Expenses', 101),
            ('Printing & Stationery', 102), ('Razorpay Charges', 103),
            ('Registration & Subscription', 104), ('Staff Welfare', 105),
            ('Telephone & Internet', 106),
        ]),
        ('Finance Cost', 'finance_cost', [
            ('Bank Charges', 108), ('Interest Expense', 109),
        ]),
        ('HR Expenses', 'hr_expenses', [
            ('Employee Salaries', 111), ('Incentives', 112),
            ('PF Admin Charges', 113), ('Stipend', 114),
        ]),
        ('Audit / Legal Fee', 'audit_legal', []),
        ('Employer PT', 'employer_pt', []),
        ('Marketing & Ads', 'marketing', []),
        ('Professional Charges', 'professional_charges', []),
        ('Rates & Taxes', 'rates_taxes', []),
        ('Repairs & Maintenance', 'repairs_maintenance', []),
        ('Round Off', 'round_off', []),
    ]

    for cat_name, cat_key, sub_items in opex_categories:
        # Category header
        cat_tb_row = TB_ROWS.get(cat_key, 1)
        if cat_key in ('rates_taxes', 'round_off'):
            formula = f"='{S}'!B{cat_tb_row}-'{S}'!C{cat_tb_row}"
        else:
            formula = f"='{S}'!B{cat_tb_row}"

        ws.cell(row=row, column=1, value=cat_name).font = GROUP_FONT
        ws.cell(row=row, column=2, value=formula).number_format = INR_FMT
        ws.cell(row=row, column=3, value=f"=IF({sales_ref}>0,B{row}/{sales_ref},0)").number_format = PCT_FMT
        ws.cell(row=row, column=4, value=f"=IF({opex_ref_dr}-{opex_ref_cr}>0,B{row}/({opex_ref_dr}-{opex_ref_cr}),0)").number_format = PCT_FMT
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = PatternFill(start_color="E8EDF4", end_color="E8EDF4", fill_type="solid")
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        # Sub-items
        for sub_name, sub_tb_row in sub_items:
            ws.cell(row=row, column=1, value=f"  {sub_name}").font = DETAIL_FONT
            ws.cell(row=row, column=2, value=f"='{S}'!B{sub_tb_row}").number_format = INR_FMT
            ws.cell(row=row, column=3, value=f"=IF({sales_ref}>0,B{row}/{sales_ref},0)").number_format = PCT_FMT
            ws.cell(row=row, column=4, value='').number_format = PCT_FMT
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    # Total OPEX
    row += 1
    ws.cell(row=row, column=1, value='TOTAL OPEX')
    ws.cell(row=row, column=2, value=f"={opex_ref_dr}-{opex_ref_cr}").number_format = INR_FMT
    ws.cell(row=row, column=3, value=f"=IF({sales_ref}>0,B{row}/{sales_ref},0)").number_format = PCT_FMT
    ws.cell(row=row, column=4, value=1).number_format = PCT_FMT
    apply_total_row(ws, row, 4)

    # Bar chart for OPEX categories
    row += 2

    # Write chart data
    chart_row = row
    ws.cell(row=row, column=5, value='Category')
    ws.cell(row=row, column=6, value='Amount')
    for i, (cat_name, cat_key, _) in enumerate(opex_categories):
        cat_tb_row = TB_ROWS.get(cat_key, 1)
        ws.cell(row=row + 1 + i, column=5, value=cat_name)
        if cat_key in ('rates_taxes', 'round_off'):
            ws.cell(row=row + 1 + i, column=6, value=f"='{S}'!B{cat_tb_row}-'{S}'!C{cat_tb_row}")
        else:
            ws.cell(row=row + 1 + i, column=6, value=f"='{S}'!B{cat_tb_row}")

    chart = BarChart()
    chart.type = "bar"
    chart.title = "OPEX Breakdown"
    chart.y_axis.title = "Amount (₹)"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(ws, min_col=5, min_row=chart_row + 1, max_row=chart_row + len(opex_categories))
    vals = Reference(ws, min_col=6, min_row=chart_row, max_row=chart_row + len(opex_categories))
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"A{chart_row}")

    set_widths(ws, {'A': 30, 'B': 20, 'C': 16, 'D': 18, 'E': 22, 'F': 16})
    ws.sheet_properties.tabColor = "BF8F00"
    return ws


# ── Sheet 9: KPIs & Ratios ──────────────────────────────────────────────────

def create_kpi_sheet(wb):
    """Create KPI and financial ratios sheet - all formulas."""
    ws = wb.create_sheet("KPIs_Ratios")
    S = 'Trial_Balance'

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP", "Financial Ratios & Key Performance Indicators")
    row += 1

    headers = ['KPI / Ratio', 'Value', 'Formula / Reference', 'Benchmark']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header(ws, row, 4)
    row += 1

    pnl_ws = wb['P&L_YTD']
    pnl = pnl_ws._pnl_refs
    bs_ws = wb['Balance_Sheet']

    sales_ref = f"P&L_YTD!B{pnl['sales']}"
    gp_ref = f"P&L_YTD!B{pnl['gp']}"
    ebitda_ref = f"P&L_YTD!B{pnl['ebitda']}"
    pat_ref = f"P&L_YTD!B{pnl['pat']}"
    opex_ref = f"P&L_YTD!B{pnl['opex']}"
    cogs_ref = f"P&L_YTD!B{pnl['cogs']}"

    def kpi_row(ws, row, category, label, formula, fmt, ref_note, benchmark=''):
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = fmt
        cell.font = Font(name="Calibri", bold=True, size=12, color=NAVY)
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=ref_note).font = Font(name="Calibri", size=9, color=GRAY)
        ws.cell(row=row, column=4, value=benchmark).font = Font(name="Calibri", size=10, color=GREEN)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        return row + 1

    # PROFITABILITY
    ws.cell(row=row, column=1, value='PROFITABILITY RATIOS')
    apply_subheader(ws, row, 4)
    row += 1

    row = kpi_row(ws, row, 'P', 'Gross Profit Margin',
                  f"=IF({sales_ref}>0,{gp_ref}/{sales_ref},0)",
                  PCT_FMT, 'Gross Profit / Revenue', '> 60%')
    row = kpi_row(ws, row, 'P', 'EBITDA Margin',
                  f"=IF({sales_ref}>0,{ebitda_ref}/{sales_ref},0)",
                  PCT_FMT, 'EBITDA / Revenue', '> 25%')
    row = kpi_row(ws, row, 'P', 'Net Profit Margin (PAT %)',
                  f"=IF({sales_ref}>0,{pat_ref}/{sales_ref},0)",
                  PCT_FMT, 'PAT / Revenue', '> 15%')
    row = kpi_row(ws, row, 'P', 'COGS to Revenue',
                  f"=IF({sales_ref}>0,({cogs_ref}+P&L_YTD!B{pnl['direct']})/{sales_ref},0)",
                  PCT_FMT, '(COGS + Direct) / Revenue', '< 40%')
    row = kpi_row(ws, row, 'P', 'OPEX to Revenue',
                  f"=IF({sales_ref}>0,{opex_ref}/{sales_ref},0)",
                  PCT_FMT, 'OPEX / Revenue', '< 40%')

    # ROA
    fa_ref = f"'{S}'!B{TB_ROWS.get('fixed_assets', 1)}"
    ca_ref = f"'{S}'!B{TB_ROWS.get('current_assets', 1)}-'{S}'!C{TB_ROWS.get('current_assets', 1)}"
    total_assets = f"({fa_ref}+{ca_ref})"
    row = kpi_row(ws, row, 'P', 'Return on Assets (ROA)',
                  f"=IF({total_assets}>0,{pat_ref}/{total_assets},0)",
                  PCT_FMT, 'PAT / Total Assets', '> 10%')

    # ROE
    equity_ref = f"'{S}'!C{TB_ROWS.get('capital_account', 1)}"
    row = kpi_row(ws, row, 'P', 'Return on Equity (ROE)',
                  f"=IF({equity_ref}>0,{pat_ref}/{equity_ref},0)",
                  PCT_FMT, 'PAT / Equity', '> 15%')
    row += 1

    # LIQUIDITY
    ws.cell(row=row, column=1, value='LIQUIDITY RATIOS')
    apply_subheader(ws, row, 4)
    row += 1

    cl_row = TB_ROWS.get('current_liabilities', 1)
    cl_ref = f"('{S}'!C{cl_row}-'{S}'!B{cl_row})"

    row = kpi_row(ws, row, 'L', 'Current Ratio',
                  f"=IF({cl_ref}>0,({ca_ref})/{cl_ref},0)",
                  NUM_DEC, 'Current Assets / Current Liabilities', '> 1.5')

    cash_ref = f"('{S}'!B{TB_ROWS.get('cash_in_hand', 1)}+'{S}'!B{TB_ROWS.get('bank_accounts', 1)})"
    row = kpi_row(ws, row, 'L', 'Cash Ratio',
                  f"=IF({cl_ref}>0,{cash_ref}/{cl_ref},0)",
                  NUM_DEC, 'Cash & Bank / Current Liabilities', '> 0.5')

    row = kpi_row(ws, row, 'L', 'Working Capital',
                  f"=({ca_ref})-{cl_ref}",
                  INR_FMT, 'Current Assets - Current Liabilities', 'Positive')
    row += 1

    # OPERATIONAL
    ws.cell(row=row, column=1, value='OPERATIONAL KPIs')
    apply_subheader(ws, row, 4)
    row += 1

    months_ref = f"Revenue_Summary!B{7}"  # Months cell in Revenue Summary

    row = kpi_row(ws, row, 'O', 'Monthly Revenue (Avg)',
                  f"={sales_ref}/{months_ref}",
                  INR_FMT, 'Revenue / Number of Months', '')
    row = kpi_row(ws, row, 'O', 'Monthly EBITDA (Avg)',
                  f"={ebitda_ref}/{months_ref}",
                  INR_FMT, 'EBITDA / Number of Months', '')
    row = kpi_row(ws, row, 'O', 'Monthly PAT (Avg)',
                  f"={pat_ref}/{months_ref}",
                  INR_FMT, 'PAT / Number of Months', '')

    # DSO
    debtors_ref = f"('{S}'!B{TB_ROWS.get('sundry_debtors', 1)}-'{S}'!C{TB_ROWS.get('sundry_debtors', 1)})"
    row = kpi_row(ws, row, 'O', 'Days Sales Outstanding (DSO)',
                  f"=IF({sales_ref}>0,{debtors_ref}/{sales_ref}*365,0)",
                  NUM_FMT, 'Debtors / Revenue × 365', '< 45 days')

    # Debt to Equity
    loans_ref = f"'{S}'!C{TB_ROWS.get('loans_liability', 1)}"
    row = kpi_row(ws, row, 'O', 'Debt to Equity',
                  f"=IF({equity_ref}>0,{loans_ref}/{equity_ref},0)",
                  NUM_DEC, 'Loans / Equity', '< 2.0')

    # Asset Turnover
    row = kpi_row(ws, row, 'O', 'Asset Turnover',
                  f"=IF({total_assets}>0,{sales_ref}/{total_assets},0)",
                  NUM_DEC, 'Revenue / Total Assets', '> 1.0')

    # Cash Position
    row += 1
    ws.cell(row=row, column=1, value='CASH POSITION')
    apply_subheader(ws, row, 4)
    row += 1

    row = kpi_row(ws, row, 'C', 'Total Cash & Bank',
                  f"={cash_ref}",
                  INR_FMT, 'Cash + All Bank Accounts', '')
    row = kpi_row(ws, row, 'C', 'Net Debtors (Receivable)',
                  f"={debtors_ref}",
                  INR_FMT, 'Sundry Debtors Net', '')
    row = kpi_row(ws, row, 'C', 'Net Creditors (Payable)',
                  f"={cl_ref}",
                  INR_FMT, 'Current Liabilities Net', '')

    set_widths(ws, {'A': 32, 'B': 20, 'C': 40, 'D': 14})
    ws.sheet_properties.tabColor = "7030A0"
    return ws


# ── Sheet 10: MIS Dashboard ─────────────────────────────────────────────────

def create_dashboard_sheet(wb):
    """Create a summary MIS dashboard for management."""
    ws = wb.create_sheet("MIS_Dashboard")
    S = 'Trial_Balance'

    # Move to first position
    wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 1))

    row = write_title(ws, 1, 1, "THOTA HOSPITALITY LLP", "Management Information System (MIS) — Dashboard")
    ws.cell(row=row, column=1, value="Period:").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=row, column=2, value=f"='{S}'!A8").font = Font(name="Calibri", bold=True, size=10, color="4472C4")
    row += 2

    pnl_ws = wb['P&L_YTD']
    pnl = pnl_ws._pnl_refs

    headers = ['Key Metric', 'YTD Amount (₹)', 'Monthly Avg (₹)', 'Margin %']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header(ws, row, 4)
    row += 1

    months_ref = "Revenue_Summary!B7"

    # P&L Summary
    ws.cell(row=row, column=1, value='P&L SUMMARY')
    apply_subheader(ws, row, 4)
    row += 1

    dashboard_items = [
        ('Total Revenue', f"=P&L_YTD!B{pnl['sales']}", f"=B{row}/{months_ref}", ''),
        ('Total COGS', f"=P&L_YTD!B{pnl['cogs']}+P&L_YTD!B{pnl['direct']}", f"=B{row+1}/{months_ref}",
         f"=IF(B{row}>0,(B{row+1})/B{row},0)"),
        ('Gross Profit', f"=P&L_YTD!B{pnl['gp']}", f"=B{row+2}/{months_ref}",
         f"=IF(B{row}>0,B{row+2}/B{row},0)"),
        ('Total OPEX', f"=P&L_YTD!B{pnl['opex']}", f"=B{row+3}/{months_ref}",
         f"=IF(B{row}>0,B{row+3}/B{row},0)"),
        ('EBITDA', f"=P&L_YTD!B{pnl['ebitda']}", f"=B{row+4}/{months_ref}",
         f"=IF(B{row}>0,B{row+4}/B{row},0)"),
        ('PAT', f"=P&L_YTD!B{pnl['pat']}", f"=B{row+5}/{months_ref}",
         f"=IF(B{row}>0,B{row+5}/B{row},0)"),
    ]

    rev_row = row  # Save for chart
    for i, (label, ytd_formula, monthly_formula, margin_formula) in enumerate(dashboard_items):
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        ws.cell(row=row, column=2, value=ytd_formula).number_format = INR_FMT
        ws.cell(row=row, column=3, value=monthly_formula).number_format = INR_FMT
        if margin_formula:
            ws.cell(row=row, column=4, value=margin_formula).number_format = PCT_FMT
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER

        if label in ('Gross Profit', 'EBITDA', 'PAT'):
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=11)
        row += 1

    row += 1

    # Balance Sheet Summary
    ws.cell(row=row, column=1, value='BALANCE SHEET SUMMARY')
    apply_subheader(ws, row, 4)
    row += 1

    bs_items = [
        ('Total Assets', f"='{S}'!B{TB_ROWS.get('fixed_assets', 1)}"
                         f"+'{S}'!B{TB_ROWS.get('current_assets', 1)}"
                         f"-'{S}'!C{TB_ROWS.get('current_assets', 1)}"),
        ('Fixed Assets', f"='{S}'!B{TB_ROWS.get('fixed_assets', 1)}"),
        ('Current Assets', f"='{S}'!B{TB_ROWS.get('current_assets', 1)}"
                           f"-'{S}'!C{TB_ROWS.get('current_assets', 1)}"),
        ('Cash & Bank', f"='{S}'!B{TB_ROWS.get('cash_in_hand', 1)}"
                        f"+'{S}'!B{TB_ROWS.get('bank_accounts', 1)}"),
        ('Sundry Debtors', f"='{S}'!B{TB_ROWS.get('sundry_debtors', 1)}"
                           f"-'{S}'!C{TB_ROWS.get('sundry_debtors', 1)}"),
        ('Current Liabilities', f"='{S}'!C{TB_ROWS.get('current_liabilities', 1)}"
                                f"-'{S}'!B{TB_ROWS.get('current_liabilities', 1)}"),
        ('Loans', f"='{S}'!C{TB_ROWS.get('loans_liability', 1)}"),
        ('Equity', f"='{S}'!C{TB_ROWS.get('capital_account', 1)}"),
    ]

    for label, formula in bs_items:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        ws.cell(row=row, column=2, value=formula).number_format = INR_FMT
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        if label == 'Total Assets':
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
        row += 1

    row += 1

    # Key Ratios
    ws.cell(row=row, column=1, value='KEY RATIOS')
    apply_subheader(ws, row, 4)
    row += 1

    fa_ref = f"'{S}'!B{TB_ROWS.get('fixed_assets', 1)}"
    ca_ref = f"'{S}'!B{TB_ROWS.get('current_assets', 1)}-'{S}'!C{TB_ROWS.get('current_assets', 1)}"
    cl_ref = f"('{S}'!C{TB_ROWS.get('current_liabilities', 1)}-'{S}'!B{TB_ROWS.get('current_liabilities', 1)})"
    sales_f = f"P&L_YTD!B{pnl['sales']}"
    debtors_f = f"('{S}'!B{TB_ROWS.get('sundry_debtors', 1)}-'{S}'!C{TB_ROWS.get('sundry_debtors', 1)})"

    ratios = [
        ('GP Margin', f"=IF({sales_f}>0,P&L_YTD!B{pnl['gp']}/{sales_f},0)", PCT_FMT),
        ('EBITDA Margin', f"=IF({sales_f}>0,P&L_YTD!B{pnl['ebitda']}/{sales_f},0)", PCT_FMT),
        ('PAT Margin', f"=IF({sales_f}>0,P&L_YTD!B{pnl['pat']}/{sales_f},0)", PCT_FMT),
        ('Current Ratio', f"=IF({cl_ref}>0,({ca_ref})/{cl_ref},0)", NUM_DEC),
        ('DSO (Days)', f"=IF({sales_f}>0,{debtors_f}/{sales_f}*365,0)", NUM_FMT),
    ]

    for label, formula, fmt in ratios:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = fmt
        cell.font = Font(name="Calibri", bold=True, size=12, color=NAVY)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Revenue and Profitability bar chart
    chart_start = row + 2
    chart_labels = ['Revenue', 'COGS+Direct', 'Gross Profit', 'OPEX', 'EBITDA', 'PAT']
    ws.cell(row=chart_start, column=6, value='Metric')
    ws.cell(row=chart_start, column=7, value='Amount')

    for i, label in enumerate(chart_labels):
        ws.cell(row=chart_start + 1 + i, column=6, value=label)
        ws.cell(row=chart_start + 1 + i, column=7, value=f"=B{rev_row + i}").number_format = INR_FMT

    chart = BarChart()
    chart.type = "col"
    chart.title = "P&L Waterfall"
    chart.y_axis.title = "Amount (₹)"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(ws, min_col=6, min_row=chart_start + 1, max_row=chart_start + len(chart_labels))
    vals = Reference(ws, min_col=7, min_row=chart_start, max_row=chart_start + len(chart_labels))
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"E{rev_row - 1}")

    set_widths(ws, {'A': 28, 'B': 22, 'C': 18, 'D': 14, 'E': 3, 'F': 18, 'G': 18})
    ws.sheet_properties.tabColor = "002060"
    return ws


# ── Main Builder ─────────────────────────────────────────────────────────────

def build_financial_model(source_path: str, output_path: str, company: str = "THOTA HOSPITALITY LLP"):
    """Build the complete formula-linked financial model."""
    print(f"📊 Reading Trial Balance from: {source_path}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. Trial Balance (INPUT)
    print("  → Creating Trial_Balance sheet (INPUT)...")
    create_trial_balance_sheet(wb, source_path)

    # 2. Monthly TB (MONTHLY INPUT)
    print("  → Creating Monthly_TB sheet (MONTHLY INPUT)...")
    create_monthly_tb_sheet(wb)

    # 3. P&L YTD
    print("  → Creating P&L_YTD sheet (formulas)...")
    create_pnl_ytd_sheet(wb)

    # 4. P&L Monthly
    print("  → Creating P&L_Monthly sheet (formulas)...")
    create_pnl_monthly_sheet(wb)

    # 5. Balance Sheet
    print("  → Creating Balance_Sheet sheet (formulas)...")
    create_balance_sheet_sheet(wb)

    # 6. Cash Flow
    print("  → Creating Cash_Flow sheet (formulas)...")
    create_cashflow_sheet(wb)

    # 7. Revenue Summary
    print("  → Creating Revenue_Summary sheet (formulas)...")
    create_revenue_summary_sheet(wb)

    # 8. OPEX Schedule
    print("  → Creating OPEX_Schedule sheet (formulas)...")
    create_opex_schedule_sheet(wb)

    # 9. KPIs & Ratios
    print("  → Creating KPIs_Ratios sheet (formulas)...")
    create_kpi_sheet(wb)

    # 10. MIS Dashboard (first sheet)
    print("  → Creating MIS_Dashboard (formulas)...")
    create_dashboard_sheet(wb)

    # Save
    wb.save(output_path)
    print(f"\n✅ Financial Model saved to: {output_path}")
    print(f"\n📌 HOW TO USE:")
    print(f"   1. Open the file in Excel")
    print(f"   2. Go to 'Trial_Balance' sheet → Paste your latest TB from Tally")
    print(f"   3. All report sheets auto-update (P&L, BS, Cash Flow, KPIs...)")
    print(f"   4. For monthly tracking → Fill 'Monthly_TB' sheet each month")
    print(f"   5. 'MIS_Dashboard' has the summary for management")
    print(f"\n🔒 Yellow cells = Input.  Gray cells = Formulas (do not edit).")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate Formula-Linked Financial Model from Trial Balance")
    parser.add_argument("--source", required=True, help="Path to Trial Balance Excel file")
    parser.add_argument("--output", default=None, help="Output file path")
    parser.add_argument("--company", default="THOTA HOSPITALITY LLP", help="Company name")
    args = parser.parse_args()

    source = Path(args.source)
    if not source.exists():
        print(f"❌ File not found: {source}")
        sys.exit(1)

    if args.output:
        output = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = str(TMP_DIR / f"Financial_Model_TB_{timestamp}.xlsx")

    build_financial_model(str(source), output, args.company)


if __name__ == "__main__":
    main()
