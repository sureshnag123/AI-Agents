#!/usr/bin/env python3
"""
Financial Model Generator for THOTA HOSPITALITY LLP

Reads source financial data from Excel and generates a comprehensive
financial model workbook with:
- Performance Summary (P&L)
- Detailed P&L (Monthly)
- Balance Sheet
- Cash Flow Statement
- Revenue Summary & Segment Analysis
- Financial Ratios & KPIs
- Budget vs Actuals

Output: Professional Excel workbook ready for partner presentation.
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from copy import copy

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel

# ── Configuration ────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(exist_ok=True)

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=13, color="4472C4")
CURRENCY_FORMAT = '#,##0'
CURRENCY_DEC_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
BOTTOM_BORDER = Border(bottom=Side(style='medium', color='2F5496'))

# Indian number format
INR_FORMAT = '₹#,##0'
INR_DEC_FORMAT = '₹#,##0.00'


def safe_float(val, default=0.0):
    """Convert value to float safely."""
    if val is None:
        return default
    if isinstance(val, str):
        val = val.strip().replace(',', '').replace('₹', '').replace(' ', '')
        if val in ('', '-', '#REF!', '\xa0'):
            return default
        try:
            return float(val)
        except ValueError:
            return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def apply_header_style(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def apply_subheader_style(ws, row, max_col):
    """Apply subheader styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def apply_total_style(ws, row, max_col):
    """Apply total row styling."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = Border(
            top=Side(style='medium', color='2F5496'),
            bottom=Side(style='double', color='2F5496'),
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9')
        )


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_title(ws, row, col, title, subtitle=None):
    """Write title and optional subtitle."""
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='left')
    if subtitle:
        cell2 = ws.cell(row=row + 1, column=col, value=subtitle)
        cell2.font = SUBTITLE_FONT
        cell2.alignment = Alignment(horizontal='left')
        return row + 3
    return row + 2


# ── Data Extraction ──────────────────────────────────────────────────────────

def load_source_data(source_path: str) -> dict:
    """Load and extract all financial data from the source Excel file."""
    wb = openpyxl.load_workbook(source_path, data_only=True)
    data = {}

    # 1. Performance Summary
    if 'Performance Summary' in wb.sheetnames:
        ws = wb['Performance Summary']
        data['performance'] = extract_performance_summary(ws)

    # 2. P&L Detailed
    if 'P&L Detailed' in wb.sheetnames:
        ws = wb['P&L Detailed']
        data['pnl_detailed'] = extract_pnl_detailed(ws)

    # 3. Schedule OPEX
    if 'Schedule_OPEX' in wb.sheetnames:
        ws = wb['Schedule_OPEX']
        data['opex'] = extract_opex(ws)

    # 4. Balance Sheet
    if 'BalanceSheet Summary' in wb.sheetnames:
        ws = wb['BalanceSheet Summary']
        data['balance_sheet'] = extract_balance_sheet(ws)

    # 5. Cashflow
    if 'Cashflow' in wb.sheetnames:
        ws = wb['Cashflow']
        data['cashflow'] = extract_cashflow(ws)

    # 6. Fundflow
    if 'Fundflow_Post Investment' in wb.sheetnames:
        ws = wb['Fundflow_Post Investment']
        data['fundflow'] = extract_fundflow(ws)

    # 7. Segment Revenue
    if 'Segment_wise_Revenue' in wb.sheetnames:
        ws = wb['Segment_wise_Revenue']
        data['segment_revenue'] = extract_segment_revenue(ws)

    # 8. Fixed Costs
    if 'FC' in wb.sheetnames:
        ws = wb['FC']
        data['fixed_costs'] = extract_fixed_costs(ws)

    # 9. Revenue Projection (Pipeline)
    if 'Revenue Projection' in wb.sheetnames:
        ws = wb['Revenue Projection']
        data['revenue_pipeline'] = extract_revenue_pipeline(ws)

    wb.close()
    return data


def extract_performance_summary(ws):
    """Extract quarterly performance summary."""
    result = {'quarters': ['Q1', 'Q2', 'Q3', 'Q4'], 'metrics': {}}
    # Row mapping based on data analysis
    row_map = {
        6: 'Hospitality Services',
        7: 'Studio Services',
        10: 'Total Sales',
        12: 'COGS',
        13: 'Direct Expenses',
        15: 'Gross Profit',
        16: 'Gross Profit %',
        19: 'Indirect Expenses',
        21: 'EBITDA',
        22: 'EBITDA %',
        24: 'Finance Cost',
        25: 'Depreciation',
        28: 'PAT',
        29: 'PAT %',
    }
    for row_num, label in row_map.items():
        values = []
        for col in range(3, 8):  # C to G (Q1-Q4 + Total)
            values.append(safe_float(ws.cell(row=row_num, column=col).value))
        result['metrics'][label] = values
    return result


def extract_pnl_detailed(ws):
    """Extract detailed monthly P&L."""
    result = {
        'months': ['APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN', 'FEB', 'MAR'],
        'line_items': {}
    }
    # Key rows from the P&L Detailed sheet
    row_map = {
        4: 'Revenue from Operations',
        5: 'Morning Glory',
        6: 'Al Fresco',
        7: 'Sunset Soiree',
        8: 'Studio',
        10: 'Purchase Cost (COGS Material)',
        15: 'Direct Expenses',
        20: 'Total COGS',
        22: 'Gross Profit',
        23: 'Gross Profit %',
        25: 'Indirect Income',
        30: 'Indirect Expenses',
    }
    for row_num, label in row_map.items():
        values = []
        for col in range(2, 15):  # B to N (Apr-Mar + Total)
            values.append(safe_float(ws.cell(row=row_num, column=col).value))
        result['line_items'][label] = values
    return result


def extract_opex(ws):
    """Extract operating expenses schedule."""
    result = {'months': [], 'categories': {}}
    # Read month headers
    for col in range(2, 14):
        val = ws.cell(row=2, column=col).value
        result['months'].append(str(val) if val else '')

    row_map = {
        3: 'Office & Admin Overheads',
        11: 'Finance Cost',
        15: 'HR Expenses',
        22: 'Marketing & Ads',
        25: 'Professional Services',
        27: 'Repairs & Maintenance',
        29: 'Grand Total OPEX',
    }
    for row_num, label in row_map.items():
        values = []
        for col in range(2, 15):  # B to N
            values.append(safe_float(ws.cell(row=row_num, column=col).value))
        result['categories'][label] = values
    return result


def extract_balance_sheet(ws):
    """Extract balance sheet data."""
    result = {'liabilities': {}, 'assets': {}, 'totals': {}}

    liability_rows = {
        6: ('Capital Account', 'E'),
        10: ('Loans (Liability)', 'E'),
        15: ('Current Liabilities', 'E'),
        22: ('Profit & Loss A/c', 'E'),
        25: ('Total Liabilities', 'E'),
    }
    for row_num, (label, col_letter) in liability_rows.items():
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        result['liabilities'][label] = safe_float(ws.cell(row=row_num, column=col_idx).value)

    # Detailed liabilities
    detail_rows = {
        7: ('Reserves & Surplus', 'D'),
        8: ('Partners Capital', 'D'),
        11: ("Partner's Loan-Akshatha", 'D'),
        12: ("Partner's Loan-Tejas", 'D'),
        13: ('TVS Credit Services Ltd', 'D'),
        16: ('Duties & Taxes', 'D'),
        18: ('Sundry Creditors', 'D'),
        19: ('Reimbursements Payable', 'D'),
        20: ('Salaries Payable', 'D'),
    }
    for row_num, (label, col_letter) in detail_rows.items():
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        result['liabilities'][f'  {label}'] = safe_float(ws.cell(row=row_num, column=col_idx).value)

    asset_rows = {
        27: ('Fixed Assets', 'E'),
        31: ('Current Assets', 'E'),
        41: ('Total Assets', 'E'),
    }
    for row_num, (label, col_letter) in asset_rows.items():
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        result['assets'][label] = safe_float(ws.cell(row=row_num, column=col_idx).value)

    asset_details = {
        28: ('Intangible Assets', 'D'),
        29: ('Tangible Assets', 'D'),
        34: ('Loans & Advances', 'D'),
        35: ('Sundry Debtors', 'D'),
        36: ('Cash-in-Hand', 'D'),
        37: ('Bank Accounts', 'D'),
    }
    for row_num, (label, col_letter) in asset_details.items():
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        result['assets'][f'  {label}'] = safe_float(ws.cell(row=row_num, column=col_idx).value)

    return result


def extract_cashflow(ws):
    """Extract cash flow statement data."""
    result = {'months': [], 'items': {}}
    # Read month headers (row 3 has dates)
    for col in range(2, 12):
        val = ws.cell(row=3, column=col).value
        if val and hasattr(val, 'strftime'):
            result['months'].append(val.strftime('%b-%y'))
        else:
            result['months'].append(str(val) if val else '')

    row_map = {
        4: 'Opening Balance',
        8: 'Current Assets Inflow',
        13: 'Total Inflow',
        17: 'Current Liabilities Outflow',
        23: 'Current Assets Outflow',
        30: 'Indirect Expenses Outflow',
        38: 'Total Outflow',
        39: 'Closing Balance',
        40: 'Net Inflow',
    }
    for row_num, label in row_map.items():
        values = []
        for col in range(2, 12):
            values.append(safe_float(ws.cell(row=row_num, column=col).value))
        result['items'][label] = values
    return result


def extract_fundflow(ws):
    """Extract fund flow statement."""
    result = {'periods': {}}

    # Previous Financial Period (Column D = total)
    prev_items = {
        6: 'Beginning Cash',
        17: 'Total Cash Receipts',
        24: 'COGS Sub-Total',
        36: 'Operating Expenses Sub-Total',
        48: 'Other Payments Sub-Total',
        50: 'Total Cash Payments',
        52: 'Net Cash Change',
    }
    result['periods']['Previous (FY2024-25)'] = {}
    for row_num, label in prev_items.items():
        result['periods']['Previous (FY2024-25)'][label] = safe_float(
            ws.cell(row=row_num, column=4).value  # Column D
        )

    # Current Financial Period (Column R = total)
    result['periods']['Current (H1 FY2025-26)'] = {}
    for row_num, label in prev_items.items():
        result['periods']['Current (H1 FY2025-26)'][label] = safe_float(
            ws.cell(row=row_num, column=18).value  # Column R
        )

    # Cash position details
    cash_detail_rows = {
        55: 'Bank Balance',
        56: 'Cash Balance',
        57: 'Bank OverDraft (ICICI)',
        58: 'Bank OverDraft (BAJAJ)',
        59: 'Fixed Deposit',
        60: 'Cash Liquidity',
    }
    result['cash_position'] = {}
    for row_num, label in cash_detail_rows.items():
        prev_val = safe_float(ws.cell(row=row_num, column=3).value)  # Col C
        curr_val = safe_float(ws.cell(row=row_num, column=17).value)  # Col Q
        result['cash_position'][label] = {'previous': prev_val, 'current': curr_val}

    return result


def extract_segment_revenue(ws):
    """Extract segment-wise revenue."""
    result = {
        'months': ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar'],
        'segments': {
            'Morning Glory': [],
            'Al Fresco': [],
            'Sunset Soiree': [],
            'Studio': [],
            'Other Income': [],
            'Total': [],
        }
    }
    col_map = {'Morning Glory': 2, 'Al Fresco': 3, 'Sunset Soiree': 4, 'Studio': 5, 'Other Income': 6, 'Total': 7}
    for seg, col_idx in col_map.items():
        for row in range(3, 15):
            result['segments'][seg].append(safe_float(ws.cell(row=row, column=col_idx).value))
    return result


def extract_fixed_costs(ws):
    """Extract fixed cost structure."""
    result = {'payroll': [], 'totals': {}}
    for row in range(4, 22):
        name = ws.cell(row=row, column=3).value  # Col C = Emp Name
        ctc = safe_float(ws.cell(row=row, column=5).value)  # Col E = CTC
        if name and ctc > 0:
            result['payroll'].append({'name': str(name), 'ctc': ctc})
    result['totals']['Total HR Cost'] = safe_float(ws.cell(row=37, column=5).value)
    result['totals']['Total Fixed Cost/Month'] = safe_float(ws.cell(row=43, column=5).value)
    result['totals']['Total Fixed Cost/Year'] = safe_float(ws.cell(row=45, column=5).value)
    return result


def extract_revenue_pipeline(ws):
    """Extract revenue pipeline/projections."""
    result = []
    for row in range(2, 46):
        quarter = ws.cell(row=row, column=1).value
        amount = safe_float(ws.cell(row=row, column=3).value)
        deal = ws.cell(row=row, column=4).value
        stage = ws.cell(row=row, column=7).value
        prob = safe_float(ws.cell(row=row, column=8).value)
        expected_rev = safe_float(ws.cell(row=row, column=14).value)
        if deal:
            result.append({
                'quarter': str(quarter) if quarter else '',
                'amount': amount,
                'deal': str(deal),
                'stage': str(stage) if stage else '',
                'probability': prob,
                'expected_revenue': expected_rev,
            })
    return result


# ── Report Generation ────────────────────────────────────────────────────────

def generate_financial_model(data: dict, output_path: str, company_name: str = "THOTA HOSPITALITY LLP"):
    """Generate the comprehensive financial model workbook."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Executive Summary / Performance Dashboard
    create_performance_summary_sheet(wb, data, company_name)

    # 2. Detailed P&L
    create_detailed_pnl_sheet(wb, data, company_name)

    # 3. Balance Sheet
    create_balance_sheet(wb, data, company_name)

    # 4. Cash Flow Statement
    create_cashflow_sheet(wb, data, company_name)

    # 5. Revenue Analysis
    create_revenue_analysis_sheet(wb, data, company_name)

    # 6. Financial Ratios & KPIs
    create_kpi_sheet(wb, data, company_name)

    # 7. Budget vs Actuals
    create_budget_vs_actuals_sheet(wb, data, company_name)

    # 8. Fund Flow Statement
    create_fundflow_sheet(wb, data, company_name)

    # Save
    wb.save(output_path)
    print(f"✓ Financial model saved to: {output_path}")
    return output_path


# ── Sheet 1: Performance Summary ─────────────────────────────────────────────

def create_performance_summary_sheet(wb, data, company_name):
    """Create executive performance summary with charts."""
    ws = wb.create_sheet("Performance Summary")
    perf = data.get('performance', {})
    metrics = perf.get('metrics', {})

    # Title
    row = write_title(ws, 1, 1, company_name, "Performance Summary — FY 2025-26")

    # Headers
    headers = ['Particulars', 'Q1 (Apr-Jun)', 'Q2 (Jul-Sep)', 'Q3 (Oct-Dec)', 'Q4 (Jan-Mar)', 'FY Total']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(headers))
    row += 1

    # Revenue section
    revenue_items = [
        ('Hospitality Services', 'Hospitality Services'),
        ('Studio Services', 'Studio Services'),
    ]
    for label, key in revenue_items:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        vals = metrics.get(key, [0]*5)
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i+2, value=v)
            cell.number_format = INR_FORMAT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')
        row += 1

    # Total Sales
    ws.cell(row=row, column=1, value='TOTAL SALES')
    vals = metrics.get('Total Sales', [0]*5)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(headers))
    row += 2

    # Cost items
    cost_items = [
        ('COGS', 'COGS'),
        ('Direct Expenses', 'Direct Expenses'),
    ]
    for label, key in cost_items:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        vals = metrics.get(key, [0]*5)
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i+2, value=v)
            cell.number_format = INR_FORMAT
            cell.border = THIN_BORDER
        row += 1

    # Gross Profit
    row += 1
    ws.cell(row=row, column=1, value='GROSS PROFIT')
    vals = metrics.get('Gross Profit', [0]*5)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(headers))
    row += 1

    # GP %
    ws.cell(row=row, column=1, value='Gross Profit %').font = Font(name="Calibri", size=11, italic=True, color="4472C4")
    vals = metrics.get('Gross Profit %', [0]*5)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = PERCENT_FORMAT
        cell.font = Font(name="Calibri", italic=True, color="4472C4")
    row += 2

    # OPEX
    ws.cell(row=row, column=1, value='Indirect Expenses (OPEX)').font = Font(name="Calibri", size=11)
    vals = metrics.get('Indirect Expenses', [0]*5)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
        cell.border = THIN_BORDER
    row += 2

    # EBITDA
    ws.cell(row=row, column=1, value='EBITDA')
    vals = metrics.get('EBITDA', [0]*5)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(headers))
    row += 1

    ws.cell(row=row, column=1, value='EBITDA %').font = Font(name="Calibri", size=11, italic=True, color="4472C4")
    vals = metrics.get('EBITDA %', [0]*5)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = PERCENT_FORMAT
        cell.font = Font(name="Calibri", italic=True, color="4472C4")
    row += 2

    # Finance Cost & Depreciation
    for label, key in [('Finance Cost', 'Finance Cost'), ('Depreciation', 'Depreciation')]:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        vals = metrics.get(key, [0]*5)
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i+2, value=v)
            cell.number_format = INR_FORMAT
            cell.border = THIN_BORDER
        row += 1

    # PAT
    row += 1
    ws.cell(row=row, column=1, value='PROFIT AFTER TAX (PAT)')
    vals = metrics.get('PAT', [0]*5)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(headers))
    row += 1

    ws.cell(row=row, column=1, value='PAT %').font = Font(name="Calibri", size=11, italic=True, color="548235")
    vals = metrics.get('PAT %', [0]*5)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = PERCENT_FORMAT
        cell.font = Font(name="Calibri", italic=True, color="548235")

    # Column widths
    set_col_widths(ws, {'A': 30, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 18})

    # Add charts
    _add_performance_charts(ws, data, row + 3)


def _add_performance_charts(ws, data, chart_start_row):
    """Add performance charts to the summary sheet."""
    perf = data.get('performance', {})
    metrics = perf.get('metrics', {})

    # Revenue Trend Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Quarterly Revenue & Profitability"
    chart.y_axis.title = "Amount (₹)"
    chart.x_axis.title = "Quarter"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    # Write chart data
    chart_data_row = chart_start_row
    quarters = ['Q1', 'Q2', 'Q3', 'Q4']
    ws.cell(row=chart_data_row, column=8, value='Quarter')
    ws.cell(row=chart_data_row, column=9, value='Revenue')
    ws.cell(row=chart_data_row, column=10, value='Gross Profit')
    ws.cell(row=chart_data_row, column=11, value='EBITDA')
    ws.cell(row=chart_data_row, column=12, value='PAT')

    sales = metrics.get('Total Sales', [0]*5)
    gp = metrics.get('Gross Profit', [0]*5)
    ebitda = metrics.get('EBITDA', [0]*5)
    pat = metrics.get('PAT', [0]*5)

    for i, q in enumerate(quarters):
        r = chart_data_row + 1 + i
        ws.cell(row=r, column=8, value=q)
        ws.cell(row=r, column=9, value=sales[i] if i < len(sales) else 0)
        ws.cell(row=r, column=10, value=gp[i] if i < len(gp) else 0)
        ws.cell(row=r, column=11, value=ebitda[i] if i < len(ebitda) else 0)
        ws.cell(row=r, column=12, value=pat[i] if i < len(pat) else 0)

    cats = Reference(ws, min_col=8, min_row=chart_data_row + 1, max_row=chart_data_row + 4)
    for col_offset, series_name in [(9, 'Revenue'), (10, 'Gross Profit'), (11, 'EBITDA'), (12, 'PAT')]:
        values = Reference(ws, min_col=col_offset, min_row=chart_data_row, max_row=chart_data_row + 4)
        chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, f"A{chart_start_row}")

    # Margin Trend Line Chart
    line_chart = LineChart()
    line_chart.title = "Margin Trends (%)"
    line_chart.y_axis.title = "Percentage"
    line_chart.style = 10
    line_chart.width = 20
    line_chart.height = 12

    margin_row = chart_data_row + 6
    ws.cell(row=margin_row, column=8, value='Quarter')
    ws.cell(row=margin_row, column=9, value='GP %')
    ws.cell(row=margin_row, column=10, value='EBITDA %')
    ws.cell(row=margin_row, column=11, value='PAT %')

    gp_pct = metrics.get('Gross Profit %', [0]*5)
    ebitda_pct = metrics.get('EBITDA %', [0]*5)
    pat_pct = metrics.get('PAT %', [0]*5)

    for i, q in enumerate(quarters):
        r = margin_row + 1 + i
        ws.cell(row=r, column=8, value=q)
        ws.cell(row=r, column=9, value=gp_pct[i] if i < len(gp_pct) else 0)
        ws.cell(row=r, column=10, value=ebitda_pct[i] if i < len(ebitda_pct) else 0)
        ws.cell(row=r, column=11, value=pat_pct[i] if i < len(pat_pct) else 0)

    cats2 = Reference(ws, min_col=8, min_row=margin_row + 1, max_row=margin_row + 4)
    for col_offset in [9, 10, 11]:
        values = Reference(ws, min_col=col_offset, min_row=margin_row, max_row=margin_row + 4)
        line_chart.add_data(values, titles_from_data=True)
    line_chart.set_categories(cats2)
    ws.add_chart(line_chart, f"A{chart_start_row + 16}")


# ── Sheet 2: Detailed P&L ────────────────────────────────────────────────────

def create_detailed_pnl_sheet(wb, data, company_name):
    """Create detailed monthly P&L statement."""
    ws = wb.create_sheet("P&L Statement")
    pnl = data.get('pnl_detailed', {})
    line_items = pnl.get('line_items', {})

    row = write_title(ws, 1, 1, company_name, "Profit & Loss Statement — FY 2025-26 (Monthly)")

    # Headers
    months = ['Particulars', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN', 'FEB (P)', 'MAR (P)', 'TOTAL']
    for col_idx, h in enumerate(months, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(months))
    row += 1

    # Revenue Section
    ws.cell(row=row, column=1, value='REVENUE FROM OPERATIONS')
    apply_subheader_style(ws, row, len(months))
    row += 1

    revenue_segments = [
        ('  Morning Glory', 'Morning Glory'),
        ('  Al Fresco', 'Al Fresco'),
        ('  Sunset Soiree', 'Sunset Soiree'),
        ('  Studio', 'Studio'),
    ]
    for label, key in revenue_segments:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        vals = line_items.get(key, [0]*13)
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i+2, value=v)
            cell.number_format = INR_FORMAT
            cell.border = THIN_BORDER
        row += 1

    # Total Revenue
    ws.cell(row=row, column=1, value='Total Revenue')
    vals = line_items.get('Revenue from Operations', [0]*13)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(months))
    row += 2

    # COGS Section
    ws.cell(row=row, column=1, value='COST OF GOODS SOLD')
    apply_subheader_style(ws, row, len(months))
    row += 1

    ws.cell(row=row, column=1, value='  Purchase Cost (Materials)').font = Font(name="Calibri", size=11)
    vals = line_items.get('Purchase Cost (COGS Material)', [0]*13)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
        cell.border = THIN_BORDER
    row += 1

    ws.cell(row=row, column=1, value='  Direct Expenses').font = Font(name="Calibri", size=11)
    vals = line_items.get('Direct Expenses', [0]*13)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
        cell.border = THIN_BORDER
    row += 1

    ws.cell(row=row, column=1, value='Total COGS')
    vals = line_items.get('Total COGS', [0]*13)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(months))
    row += 2

    # Gross Profit
    ws.cell(row=row, column=1, value='GROSS PROFIT')
    vals = line_items.get('Gross Profit', [0]*13)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(months))
    row += 1

    ws.cell(row=row, column=1, value='Gross Profit %').font = Font(name="Calibri", italic=True, color="4472C4")
    vals = line_items.get('Gross Profit %', [0]*13)
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = PERCENT_FORMAT
        cell.font = Font(name="Calibri", italic=True, color="4472C4")
    row += 2

    # Indirect Expenses (OPEX)
    ws.cell(row=row, column=1, value='OPERATING EXPENSES (OPEX)')
    apply_subheader_style(ws, row, len(months))
    row += 1

    opex = data.get('opex', {}).get('categories', {})
    opex_items = [
        ('  Office & Admin Overheads', 'Office & Admin Overheads'),
        ('  Finance Cost', 'Finance Cost'),
        ('  HR Expenses', 'HR Expenses'),
        ('  Marketing & Ads', 'Marketing & Ads'),
        ('  Professional Services', 'Professional Services'),
        ('  Repairs & Maintenance', 'Repairs & Maintenance'),
    ]
    for label, key in opex_items:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        vals = opex.get(key, [0]*13)
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i+2, value=v)
            cell.number_format = INR_FORMAT
            cell.border = THIN_BORDER
        row += 1

    ws.cell(row=row, column=1, value='Total OPEX')
    vals = opex.get('Grand Total OPEX', line_items.get('Indirect Expenses', [0]*13))
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(months))

    # Column widths
    set_col_widths(ws, {'A': 32})
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 14


# ── Sheet 3: Balance Sheet ───────────────────────────────────────────────────

def create_balance_sheet(wb, data, company_name):
    """Create the balance sheet."""
    ws = wb.create_sheet("Balance Sheet")
    bs = data.get('balance_sheet', {})
    liabilities = bs.get('liabilities', {})
    assets = bs.get('assets', {})

    row = write_title(ws, 1, 1, company_name, "Balance Sheet — As at Jan 31, 2026")

    # LIABILITIES
    headers = ['Particulars', 'Sub-Total (₹)', 'Total (₹)']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(headers))
    row += 1

    ws.cell(row=row, column=1, value='LIABILITIES')
    apply_subheader_style(ws, row, len(headers))
    row += 1

    liability_structure = [
        ('Capital Account', True),
        ('  Reserves & Surplus', False),
        ('  Partners Capital', False),
        ('', None),
        ('Loans (Liability)', True),
        ("  Partner's Loan-Akshatha", False),
        ("  Partner's Loan-Tejas", False),
        ('  TVS Credit Services Ltd', False),
        ('', None),
        ('Current Liabilities', True),
        ('  Duties & Taxes', False),
        ('  Sundry Creditors', False),
        ('  Reimbursements Payable', False),
        ('  Salaries Payable', False),
        ('', None),
        ('Profit & Loss A/c', True),
    ]

    for label, is_total in liability_structure:
        if label == '':
            row += 1
            continue
        ws.cell(row=row, column=1, value=label)
        val = liabilities.get(label, 0)
        if is_total:
            ws.cell(row=row, column=3, value=val).number_format = INR_FORMAT
            ws.cell(row=row, column=1).font = TOTAL_FONT
        else:
            ws.cell(row=row, column=2, value=val).number_format = INR_FORMAT
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Total Liabilities
    ws.cell(row=row, column=1, value='TOTAL LIABILITIES')
    ws.cell(row=row, column=3, value=liabilities.get('Total Liabilities', 0)).number_format = INR_FORMAT
    apply_total_style(ws, row, len(headers))
    row += 2

    # ASSETS
    ws.cell(row=row, column=1, value='ASSETS')
    apply_subheader_style(ws, row, len(headers))
    row += 1

    asset_structure = [
        ('Fixed Assets', True),
        ('  Intangible Assets', False),
        ('  Tangible Assets', False),
        ('', None),
        ('Current Assets', True),
        ('  Loans & Advances', False),
        ('  Sundry Debtors', False),
        ('  Cash-in-Hand', False),
        ('  Bank Accounts', False),
    ]

    for label, is_total in asset_structure:
        if label == '':
            row += 1
            continue
        ws.cell(row=row, column=1, value=label)
        val = assets.get(label, 0)
        if is_total:
            ws.cell(row=row, column=3, value=val).number_format = INR_FORMAT
            ws.cell(row=row, column=1).font = TOTAL_FONT
        else:
            ws.cell(row=row, column=2, value=val).number_format = INR_FORMAT
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Total Assets
    row += 1
    ws.cell(row=row, column=1, value='TOTAL ASSETS')
    ws.cell(row=row, column=3, value=assets.get('Total Assets', 0)).number_format = INR_FORMAT
    apply_total_style(ws, row, len(headers))

    set_col_widths(ws, {'A': 35, 'B': 20, 'C': 20})


# ── Sheet 4: Cash Flow Statement ─────────────────────────────────────────────

def create_cashflow_sheet(wb, data, company_name):
    """Create cash flow statement."""
    ws = wb.create_sheet("Cash Flow")
    cf = data.get('cashflow', {})
    items = cf.get('items', {})
    cf_months = cf.get('months', [])

    row = write_title(ws, 1, 1, company_name, "Cash Flow Statement — FY 2025-26")

    headers = ['Particulars'] + cf_months
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(headers))
    row += 1

    cf_structure = [
        ('Opening Balance', 'Opening Balance', True),
        ('', None, None),
        ('INFLOW', None, 'subheader'),
        ('  Current Assets (Receivables)', 'Current Assets Inflow', False),
        ('Total Inflow', 'Total Inflow', True),
        ('', None, None),
        ('OUTFLOW', None, 'subheader'),
        ('  Current Liabilities', 'Current Liabilities Outflow', False),
        ('  Current Assets (Payments)', 'Current Assets Outflow', False),
        ('  Indirect Expenses', 'Indirect Expenses Outflow', False),
        ('Total Outflow', 'Total Outflow', True),
        ('', None, None),
        ('Closing Balance', 'Closing Balance', True),
        ('Net Cash Inflow / (Outflow)', 'Net Inflow', True),
    ]

    for label, key, style in cf_structure:
        if label == '':
            row += 1
            continue
        if style == 'subheader':
            ws.cell(row=row, column=1, value=label)
            apply_subheader_style(ws, row, len(headers))
            row += 1
            continue

        ws.cell(row=row, column=1, value=label)
        vals = items.get(key, [0] * len(cf_months))
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i + 2, value=v)
            cell.number_format = INR_FORMAT
            cell.border = THIN_BORDER
            # Color negative values red
            if v < 0:
                cell.font = Font(name="Calibri", color="C00000")

        if style:
            apply_total_style(ws, row, len(headers))
        row += 1

    set_col_widths(ws, {'A': 32})
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16


# ── Sheet 5: Revenue Analysis ────────────────────────────────────────────────

def create_revenue_analysis_sheet(wb, data, company_name):
    """Create revenue analysis with segment breakdown and charts."""
    ws = wb.create_sheet("Revenue Analysis")
    seg = data.get('segment_revenue', {})
    segments = seg.get('segments', {})
    seg_months = seg.get('months', [])

    row = write_title(ws, 1, 1, company_name, "Revenue Analysis — Segment-wise Breakdown FY 2025-26")

    # Monthly Segment Revenue Table
    headers = ['Month'] + list(segments.keys())
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(headers))
    data_start_row = row + 1
    row += 1

    for i, month in enumerate(seg_months):
        ws.cell(row=row, column=1, value=month)
        for j, (seg_name, seg_vals) in enumerate(segments.items()):
            cell = ws.cell(row=row, column=j + 2, value=seg_vals[i] if i < len(seg_vals) else 0)
            cell.number_format = INR_FORMAT
            cell.border = THIN_BORDER
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value='TOTAL')
    for j, (seg_name, seg_vals) in enumerate(segments.items()):
        total = sum(v for v in seg_vals if v)
        cell = ws.cell(row=row, column=j + 2, value=total)
        cell.number_format = INR_FORMAT
    apply_total_style(ws, row, len(headers))
    row += 1

    # Percentage row
    total_revenue = sum(v for v in segments.get('Total', []) if v)
    ws.cell(row=row, column=1, value='% of Revenue')
    for j, (seg_name, seg_vals) in enumerate(segments.items()):
        seg_total = sum(v for v in seg_vals if v)
        pct = seg_total / total_revenue if total_revenue else 0
        cell = ws.cell(row=row, column=j + 2, value=pct)
        cell.number_format = PERCENT_FORMAT
        cell.font = Font(name="Calibri", italic=True, color="4472C4")
    row += 2

    # Revenue Pipeline Summary
    pipeline = data.get('revenue_pipeline', [])
    if pipeline:
        row = write_title(ws, row, 1, "Revenue Pipeline", "Sales Deals by Stage")
        pipe_headers = ['Quarter', 'Deal', 'Amount (₹)', 'Stage', 'Probability %', 'Expected Revenue (₹)']
        for col_idx, h in enumerate(pipe_headers, 1):
            ws.cell(row=row, column=col_idx, value=h)
        apply_header_style(ws, row, len(pipe_headers))
        row += 1

        for deal in pipeline[:25]:  # Top 25 deals
            ws.cell(row=row, column=1, value=deal['quarter'])
            ws.cell(row=row, column=2, value=deal['deal'])
            ws.cell(row=row, column=3, value=deal['amount']).number_format = INR_FORMAT
            ws.cell(row=row, column=4, value=deal['stage'])
            ws.cell(row=row, column=5, value=deal['probability'] / 100 if deal['probability'] > 1 else deal['probability'])
            ws.cell(row=row, column=5).number_format = PERCENT_FORMAT
            ws.cell(row=row, column=6, value=deal['expected_revenue']).number_format = INR_FORMAT
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    set_col_widths(ws, {'A': 14, 'B': 18, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 16})

    # Pie chart for segment mix
    _add_segment_pie_chart(ws, data, data_start_row, len(seg_months))


def _add_segment_pie_chart(ws, data, data_start_row, num_months):
    """Add a pie chart showing revenue segment mix."""
    seg = data.get('segment_revenue', {})
    segments = seg.get('segments', {})

    # Write totals for pie chart (exclude Total and Other Income if zero)
    chart_row = data_start_row + num_months + 8
    ws.cell(row=chart_row, column=9, value='Segment')
    ws.cell(row=chart_row, column=10, value='Revenue')

    seg_data = []
    for seg_name, seg_vals in segments.items():
        if seg_name in ('Total', 'Other Income'):
            continue
        total = sum(v for v in seg_vals if v)
        if total > 0:
            seg_data.append((seg_name, total))

    for i, (name, total) in enumerate(seg_data):
        ws.cell(row=chart_row + 1 + i, column=9, value=name)
        ws.cell(row=chart_row + 1 + i, column=10, value=total)

    if seg_data:
        pie = PieChart()
        pie.title = "Revenue Mix by Segment"
        pie.style = 10
        pie.width = 18
        pie.height = 14

        cats = Reference(ws, min_col=9, min_row=chart_row + 1, max_row=chart_row + len(seg_data))
        vals = Reference(ws, min_col=10, min_row=chart_row, max_row=chart_row + len(seg_data))
        pie.add_data(vals, titles_from_data=True)
        pie.set_categories(cats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = False
        ws.add_chart(pie, f"I{data_start_row}")


# ── Sheet 6: Financial Ratios & KPIs ─────────────────────────────────────────

def create_kpi_sheet(wb, data, company_name):
    """Create financial ratios and KPI dashboard."""
    ws = wb.create_sheet("KPIs & Ratios")

    row = write_title(ws, 1, 1, company_name, "Financial Ratios & Key Performance Indicators — FY 2025-26")

    perf = data.get('performance', {}).get('metrics', {})
    bs = data.get('balance_sheet', {})
    fc = data.get('fixed_costs', {}).get('totals', {})

    total_sales = perf.get('Total Sales', [0]*5)[-1]  # FY Total
    total_cogs = perf.get('COGS', [0]*5)[-1]
    total_direct = perf.get('Direct Expenses', [0]*5)[-1]
    gross_profit = perf.get('Gross Profit', [0]*5)[-1]
    ebitda = perf.get('EBITDA', [0]*5)[-1]
    pat = perf.get('PAT', [0]*5)[-1]
    indirect_exp = perf.get('Indirect Expenses', [0]*5)[-1]
    depreciation = perf.get('Depreciation', [0]*5)[-1]
    finance_cost = perf.get('Finance Cost', [0]*5)[-1]

    total_assets = bs.get('assets', {}).get('Total Assets', 0)
    total_liabilities = bs.get('liabilities', {}).get('Total Liabilities', 0)
    current_assets = bs.get('assets', {}).get('Current Assets', 0)
    current_liabilities = bs.get('liabilities', {}).get('Current Liabilities', 0)
    fixed_assets = bs.get('assets', {}).get('Fixed Assets', 0)
    capital = bs.get('liabilities', {}).get('Capital Account', 0)
    loans = bs.get('liabilities', {}).get('Loans (Liability)', 0)
    debtors = bs.get('assets', {}).get('  Sundry Debtors', 0)
    cash = bs.get('assets', {}).get('  Cash-in-Hand', 0) + bs.get('assets', {}).get('  Bank Accounts', 0)

    monthly_fc = fc.get('Total Fixed Cost/Month', 0)

    # ─── Profitability Ratios ────────────────────────────────────────────
    headers = ['KPI / Ratio', 'Value', 'Benchmark / Notes']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(headers))
    row += 1

    ws.cell(row=row, column=1, value='PROFITABILITY RATIOS')
    apply_subheader_style(ws, row, len(headers))
    row += 1

    profitability_kpis = [
        ('Gross Profit Margin', gross_profit / total_sales if total_sales else 0, PERCENT_FORMAT, 'Target: >60%'),
        ('EBITDA Margin', ebitda / total_sales if total_sales else 0, PERCENT_FORMAT, 'Target: >25%'),
        ('Net Profit Margin (PAT %)', pat / total_sales if total_sales else 0, PERCENT_FORMAT, 'Target: >15%'),
        ('Return on Assets (ROA)', pat / total_assets if total_assets else 0, PERCENT_FORMAT, 'PAT / Total Assets'),
        ('Return on Equity (ROE)', pat / capital if capital and capital != 0 else 0, PERCENT_FORMAT, 'PAT / Equity'),
        ('COGS to Revenue', (total_cogs + total_direct) / total_sales if total_sales else 0, PERCENT_FORMAT, 'Lower is better'),
        ('OPEX to Revenue', indirect_exp / total_sales if total_sales else 0, PERCENT_FORMAT, 'Target: <40%'),
    ]

    for label, value, fmt, note in profitability_kpis:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        cell.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=note).font = Font(name="Calibri", size=10, color="808080")
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1

    # ─── Liquidity Ratios ────────────────────────────────────────────────
    ws.cell(row=row, column=1, value='LIQUIDITY RATIOS')
    apply_subheader_style(ws, row, len(headers))
    row += 1

    current_ratio = current_assets / current_liabilities if current_liabilities else 0
    quick_ratio = (current_assets - 0) / current_liabilities if current_liabilities else 0  # No inventory listed
    cash_ratio = cash / current_liabilities if current_liabilities else 0

    liquidity_kpis = [
        ('Current Ratio', current_ratio, '#,##0.00', 'Target: >1.5 (Current Assets/Liabilities)'),
        ('Quick Ratio', quick_ratio, '#,##0.00', 'Target: >1.0 (CA - Inventory)/CL'),
        ('Cash Ratio', cash_ratio, '#,##0.00', 'Cash & Bank / Current Liabilities'),
        ('Working Capital (₹)', current_assets - current_liabilities, INR_FORMAT, 'Current Assets - Current Liabilities'),
    ]

    for label, value, fmt, note in liquidity_kpis:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        cell.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=note).font = Font(name="Calibri", size=10, color="808080")
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1

    # ─── Operational KPIs ─────────────────────────────────────────────────
    ws.cell(row=row, column=1, value='OPERATIONAL KPIs')
    apply_subheader_style(ws, row, len(headers))
    row += 1

    monthly_revenue = total_sales / 12 if total_sales else 0
    monthly_breakeven = monthly_fc / (gross_profit / total_sales) if total_sales and gross_profit else 0
    dso = (debtors / total_sales * 365) if total_sales else 0
    runway_months = cash / monthly_fc if monthly_fc else 0

    operational_kpis = [
        ('Monthly Revenue (Avg)', monthly_revenue, INR_FORMAT, 'Annual Revenue / 12'),
        ('Monthly Fixed Cost', monthly_fc, INR_FORMAT, 'From FC Sheet'),
        ('Annual Fixed Cost', fc.get('Total Fixed Cost/Year', 0), INR_FORMAT, 'Annual Commitment'),
        ('Monthly Breakeven Revenue', monthly_breakeven, INR_FORMAT, 'FC / GP Margin'),
        ('Days Sales Outstanding (DSO)', dso, '#,##0', 'Debtors / Revenue × 365'),
        ('Cash Runway (Months)', runway_months, '#,##0.0', 'Cash / Monthly Fixed Cost'),
        ('Debt to Equity', loans / capital if capital and capital != 0 else 0, '#,##0.00', 'Loans / Equity'),
        ('Asset Turnover', total_sales / total_assets if total_assets else 0, '#,##0.00', 'Revenue / Total Assets'),
    ]

    for label, value, fmt, note in operational_kpis:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        cell.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=note).font = Font(name="Calibri", size=10, color="808080")
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    row += 1

    # ─── Quarterly Trend Table ────────────────────────────────────────────
    ws.cell(row=row, column=1, value='QUARTERLY TREND')
    apply_subheader_style(ws, row, 6)
    row += 1

    trend_headers = ['Metric', 'Q1', 'Q2', 'Q3', 'Q4', 'FY Total']
    for col_idx, h in enumerate(trend_headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(trend_headers))
    row += 1

    trend_metrics = [
        ('Revenue', 'Total Sales', INR_FORMAT),
        ('Gross Profit', 'Gross Profit', INR_FORMAT),
        ('EBITDA', 'EBITDA', INR_FORMAT),
        ('PAT', 'PAT', INR_FORMAT),
        ('GP Margin', 'Gross Profit %', PERCENT_FORMAT),
        ('EBITDA Margin', 'EBITDA %', PERCENT_FORMAT),
        ('PAT Margin', 'PAT %', PERCENT_FORMAT),
    ]

    for label, key, fmt in trend_metrics:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
        vals = perf.get(key, [0]*5)
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i + 2, value=v)
            cell.number_format = fmt
            cell.border = THIN_BORDER
        row += 1

    set_col_widths(ws, {'A': 35, 'B': 20, 'C': 45, 'D': 16, 'E': 16, 'F': 16})


# ── Sheet 7: Budget vs Actuals ───────────────────────────────────────────────

def create_budget_vs_actuals_sheet(wb, data, company_name):
    """Create budget vs actuals comparison."""
    ws = wb.create_sheet("Budget vs Actuals")

    row = write_title(ws, 1, 1, company_name, "Budget vs Actuals — FY 2025-26")

    perf = data.get('performance', {}).get('metrics', {})
    pnl = data.get('pnl_detailed', {}).get('line_items', {})

    # Calculate actuals (Q1+Q2+Q3 = first 3 quarters actual) and Q4 as projection/budget
    actual_quarters = 3  # Q1, Q2, Q3 are actuals
    sales = perf.get('Total Sales', [0]*5)
    gp = perf.get('Gross Profit', [0]*5)
    ebitda = perf.get('EBITDA', [0]*5)
    pat = perf.get('PAT', [0]*5)
    cogs = perf.get('COGS', [0]*5)
    direct = perf.get('Direct Expenses', [0]*5)
    opex = perf.get('Indirect Expenses', [0]*5)

    headers = ['Particulars', 'Actuals (Q1-Q3)', 'Budget / Proj (Q4)', 'FY Total',
               'YoY Growth Target', 'Variance', 'Variance %']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(headers))
    row += 1

    def write_bva_row(ws, row, label, vals, is_total=False, is_pct=False):
        """Write a budget vs actuals row."""
        actual = sum(vals[:actual_quarters])
        budget = vals[actual_quarters] if len(vals) > actual_quarters else 0
        fy_total = vals[-1] if len(vals) > 4 else actual + budget

        ws.cell(row=row, column=1, value=label)
        fmt = PERCENT_FORMAT if is_pct else INR_FORMAT

        ws.cell(row=row, column=2, value=actual).number_format = fmt
        ws.cell(row=row, column=3, value=budget).number_format = fmt
        ws.cell(row=row, column=4, value=fy_total).number_format = fmt

        # YoY target (assume 30% growth target for illustration)
        if not is_pct:
            yoy_target = fy_total * 1.3
            ws.cell(row=row, column=5, value=yoy_target).number_format = INR_FORMAT
            variance = fy_total - yoy_target
            ws.cell(row=row, column=6, value=variance).number_format = INR_FORMAT
            variance_pct = variance / yoy_target if yoy_target else 0
            ws.cell(row=row, column=7, value=variance_pct).number_format = PERCENT_FORMAT
            if variance < 0:
                ws.cell(row=row, column=6).font = Font(name="Calibri", color="C00000")
                ws.cell(row=row, column=7).font = Font(name="Calibri", color="C00000")

        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN_BORDER

        if is_total:
            apply_total_style(ws, row, len(headers))
        return row + 1

    ws.cell(row=row, column=1, value='REVENUE')
    apply_subheader_style(ws, row, len(headers))
    row += 1

    row = write_bva_row(ws, row, 'Total Sales', sales, is_total=True)
    row += 1

    ws.cell(row=row, column=1, value='COST OF GOODS SOLD')
    apply_subheader_style(ws, row, len(headers))
    row += 1
    row = write_bva_row(ws, row, 'COGS (Materials)', cogs)
    row = write_bva_row(ws, row, 'Direct Expenses', direct)
    row += 1

    row = write_bva_row(ws, row, 'GROSS PROFIT', gp, is_total=True)
    row += 1

    ws.cell(row=row, column=1, value='OPERATING EXPENSES')
    apply_subheader_style(ws, row, len(headers))
    row += 1
    row = write_bva_row(ws, row, 'Total OPEX', opex)
    row += 1

    row = write_bva_row(ws, row, 'EBITDA', ebitda, is_total=True)
    row += 1
    row = write_bva_row(ws, row, 'PAT', pat, is_total=True)

    # Variance chart
    row += 2
    chart = BarChart()
    chart.type = "col"
    chart.title = "Actuals vs Target (FY 2025-26)"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    # Write chart data
    chart_items = [
        ('Revenue', sales[-1] if len(sales) > 4 else 0, (sales[-1] if len(sales) > 4 else 0) * 1.3),
        ('Gross Profit', gp[-1] if len(gp) > 4 else 0, (gp[-1] if len(gp) > 4 else 0) * 1.3),
        ('EBITDA', ebitda[-1] if len(ebitda) > 4 else 0, (ebitda[-1] if len(ebitda) > 4 else 0) * 1.3),
        ('PAT', pat[-1] if len(pat) > 4 else 0, (pat[-1] if len(pat) > 4 else 0) * 1.3),
    ]

    ws.cell(row=row, column=8, value='Metric')
    ws.cell(row=row, column=9, value='Actual/Projected')
    ws.cell(row=row, column=10, value='Target (30% Growth)')

    for i, (label, actual_val, target_val) in enumerate(chart_items):
        ws.cell(row=row + 1 + i, column=8, value=label)
        ws.cell(row=row + 1 + i, column=9, value=actual_val)
        ws.cell(row=row + 1 + i, column=10, value=target_val)

    cats = Reference(ws, min_col=8, min_row=row + 1, max_row=row + len(chart_items))
    for col_off in [9, 10]:
        vals = Reference(ws, min_col=col_off, min_row=row, max_row=row + len(chart_items))
        chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{row}")

    set_col_widths(ws, {'A': 28, 'B': 20, 'C': 22, 'D': 18, 'E': 22, 'F': 16, 'G': 14})


# ── Sheet 8: Fund Flow Statement ─────────────────────────────────────────────

def create_fundflow_sheet(wb, data, company_name):
    """Create fund flow statement."""
    ws = wb.create_sheet("Fund Flow")
    ff = data.get('fundflow', {})
    periods = ff.get('periods', {})
    cash_pos = ff.get('cash_position', {})

    row = write_title(ws, 1, 1, company_name, "Fund Flow Statement — Post Investment")

    # Headers
    period_names = list(periods.keys())
    headers = ['PARTICULARS'] + period_names + ['Change']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(headers))
    row += 1

    ff_items = [
        ('Beginning Cash', True),
        ('Total Cash Receipts', True),
        ('', None),
        ('LESS: CASH PAYMENTS', None),
        ('COGS Sub-Total', False),
        ('Operating Expenses Sub-Total', False),
        ('Other Payments Sub-Total', False),
        ('Total Cash Payments', True),
        ('', None),
        ('Net Cash Change', True),
    ]

    for label, is_total in ff_items:
        if label == '':
            row += 1
            continue
        if is_total is None:
            ws.cell(row=row, column=1, value=label)
            apply_subheader_style(ws, row, len(headers))
            row += 1
            continue

        ws.cell(row=row, column=1, value=label)
        vals = []
        for period in period_names:
            v = periods.get(period, {}).get(label, 0)
            vals.append(v)
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i + 2, value=v)
            cell.number_format = INR_FORMAT
            cell.border = THIN_BORDER
            if v < 0:
                cell.font = Font(name="Calibri", color="C00000")
        # Change column
        if len(vals) >= 2:
            change = vals[-1] - vals[0]
            cell = ws.cell(row=row, column=len(period_names) + 2, value=change)
            cell.number_format = INR_FORMAT
            if change < 0:
                cell.font = Font(name="Calibri", color="C00000")

        if is_total:
            apply_total_style(ws, row, len(headers))
        row += 1

    # Cash Position
    row += 2
    ws.cell(row=row, column=1, value='CASH POSITION')
    apply_subheader_style(ws, row, len(headers))
    row += 1

    cp_headers = ['Item', 'Previous Period', 'Current Period']
    for col_idx, h in enumerate(cp_headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    apply_header_style(ws, row, len(cp_headers))
    row += 1

    for item, vals in cash_pos.items():
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=vals.get('previous', 0)).number_format = INR_FORMAT
        ws.cell(row=row, column=3, value=vals.get('current', 0)).number_format = INR_FORMAT
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        if item == 'Cash Liquidity':
            apply_total_style(ws, row, 3)
        row += 1

    set_col_widths(ws, {'A': 35, 'B': 25, 'C': 25, 'D': 20})


# ── Main Entry Point ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Financial Model from source Excel")
    parser.add_argument("--source", required=True, help="Path to source Excel file")
    parser.add_argument("--output", default=None, help="Output file path (default: .tmp/Financial_Model_<date>.xlsx)")
    parser.add_argument("--company", default="THOTA HOSPITALITY LLP", help="Company name")
    args = parser.parse_args()

    source_path = Path(args.source)
    if not source_path.exists():
        print(f"❌ Source file not found: {source_path}")
        sys.exit(1)

    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(TMP_DIR / f"Financial_Model_{timestamp}.xlsx")

    print(f"📊 Loading source data from: {source_path}")
    data = load_source_data(str(source_path))

    # Save extracted data as JSON for debugging
    json_path = str(TMP_DIR / "extracted_data.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, default=str)
    print(f"📋 Extracted data saved to: {json_path}")

    print(f"📈 Generating financial model...")
    generate_financial_model(data, output_path, args.company)

    print(f"\n✅ Financial model generation complete!")
    print(f"📁 Output: {output_path}")


if __name__ == "__main__":
    main()
