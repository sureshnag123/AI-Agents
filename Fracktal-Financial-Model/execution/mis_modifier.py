#!/usr/bin/env python3
"""
MIS Workbook Modifier — COGS Fix + CRM Projections  (v2 — with formula fixer)

Key insight: openpyxl's insert_rows() does NOT adjust formulas.
This script manually fixes ALL formula references after row insertions.

Order of operations:
  1. Copy workbook
  2. Insert rows in TB (Opening Stock + Closing Stock)
  3. Insert rows in P&L (Opening Stock + Total Purchases + Closing Stock)
  4. Fix ALL formulas in ALL sheets using regex-based reference adjuster
  5. Write new content (Opening/Closing Stock cells, COGS formula, etc.)
  6. Create CRM Data, Sales Projection, Q1 Projection sheets
  7. Save
"""

import re
import shutil
from pathlib import Path
from datetime import datetime
from copy import copy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──
SOURCE_MIS = Path(r"D:\Suresh_AGENTS\Fracktal-Financial-Model\.tmp\Fracktal_MIS_Master_20260224_181404.xlsx")
CRM_FILE   = Path(r"C:\Users\Lenovo\Downloads\Deals_2026_03_10.xlsx")
OUTPUT_DIR = Path(r"D:\Suresh_AGENTS\Fracktal-Financial-Model\.tmp")
OUTPUT_DIR.mkdir(exist_ok=True)

ts = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_MIS = OUTPUT_DIR / f"Fracktal_MIS_Master_{ts}.xlsx"

# ── Style Constants ──
NAVY       = "1F3864"
DARK_BLUE  = "2F5496"
MID_BLUE   = "4472C4"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN= "E2EFDA"
LIGHT_YELLOW="FFF2CC"
WHITE      = "FFFFFF"
RED        = "C00000"
DARK_GREEN = "548235"
ORANGE     = "ED7D31"

TITLE_FONT       = Font(name="Calibri", bold=True, size=14, color=NAVY)
SUBTITLE_FONT    = Font(name="Calibri", bold=True, size=12, color=MID_BLUE)
HEADER_FONT      = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL      = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SECTION_FONT     = Font(name="Calibri", bold=True, size=11, color=NAVY)
SECTION_FILL     = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
TOTAL_FONT       = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL       = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
GRAND_TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
GRAND_TOTAL_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
INPUT_FILL       = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
NORMAL_FONT      = Font(name="Calibri", size=11)
SMALL_FONT       = Font(name="Calibri", size=10, color="808080", italic=True)
PCT_FONT         = Font(name="Calibri", italic=True, size=10, color=MID_BLUE)
NOTE_FONT        = Font(name="Calibri", size=10, color="808080")

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
TOTAL_BORDER = Border(
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='double', color=NAVY),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)

INR  = '#,##0'
INR2 = '#,##0.00'
PCT_FMT = '0.0%'
MONTHS = ['APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC','JAN','FEB','MAR']

def cl(c):
    return get_column_letter(c)


# ═══════════════════════════════════════════════════════════════════════════════
#  FORMULA FIXER — The core of v2
# ═══════════════════════════════════════════════════════════════════════════════
#
# After inserting rows in TB and P&L, ALL formulas in ALL sheets have stale
# row references. This function scans every cell and adjusts references.
#
# TB row shifts (2 inserts: 1 row at 13, then 1 row at 17 in new space):
#   Original row < 13  → no shift
#   Original row 13-15 → +1
#   Original row >= 16 → +2
#
# P&L row shifts (2 inserts: 1 row at 20, then 2 rows at 24 in new space):
#   Original row < 20  → no shift
#   Original row 20-22 → +1
#   Original row >= 23 → +3

def tb_shift(old_row):
    """Map original TB row to new TB row after insertions."""
    if old_row < 13:
        return old_row
    elif old_row <= 15:
        return old_row + 1
    else:
        return old_row + 2

def pnl_shift(old_row):
    """Map original P&L row to new P&L row after insertions."""
    if old_row < 20:
        return old_row
    elif old_row <= 22:
        return old_row + 1
    else:
        return old_row + 3


# Regex to match cell references in Excel formulas.
# Captures: (1) optional sheet prefix with !, (2) $ + column letters, (3) $ for row, (4) row digits
# Does NOT match function names (e.g., SUM, IFERROR) because they don't have digits after letters.
CELL_REF_RE = re.compile(
    r"((?:'[^']*'|[A-Za-z_]\w*)!)?"   # (1) Optional sheet prefix: TB! or 'P&L'!
    r"(\$?[A-Z]{1,3})"                # (2) Column with optional $
    r"(\$?)(\d+)"                      # (3)(4) Row with optional $
    r"(?!\w)"                          # Negative lookahead: not part of longer token
)


def fix_formula(formula, current_sheet_name):
    """
    Adjust all cell references in a formula string.
    References to TB rows get tb_shift; references to P&L rows get pnl_shift.
    Unqualified references use current_sheet_name to determine which shift to apply.
    """
    def _replacer(m):
        sheet_prefix = m.group(1) or ""           # "TB!" or "'P&L'!" or ""
        col_part     = m.group(2)                  # "$A" or "A"
        row_dollar   = m.group(3)                  # "$" or ""
        row_str      = m.group(4)                  # "13"
        old_row      = int(row_str)

        # Determine target sheet
        if sheet_prefix:
            sname = sheet_prefix.rstrip("!").strip("'")
        else:
            sname = current_sheet_name

        # Apply shift
        if sname == "TB":
            new_row = tb_shift(old_row)
        elif sname == "P&L":
            new_row = pnl_shift(old_row)
        else:
            new_row = old_row  # No shift for other sheets

        return f"{sheet_prefix}{col_part}{row_dollar}{new_row}"

    return CELL_REF_RE.sub(_replacer, formula)


def fix_all_formulas(wb):
    """Scan every cell in every sheet and fix formula references."""
    fixed_count = 0
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    new_val = fix_formula(val, ws_name)
                    if new_val != val:
                        cell.value = new_val
                        fixed_count += 1
    return fixed_count


# ═══════════════════════════════════════════════════════════════════════════════
#  STYLING HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_section_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER

def style_total_row(ws, row, max_col, grand=False):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = GRAND_TOTAL_FONT if grand else TOTAL_FONT
        cell.fill = GRAND_TOTAL_FILL if grand else TOTAL_FILL
        cell.border = TOTAL_BORDER

def fmt_number_cells(ws, row, scol, ecol, fmt=INR):
    for c in range(scol, ecol + 1):
        cell = ws.cell(row=row, column=c)
        cell.number_format = fmt
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')

def fmt_pct_cells(ws, row, scol, ecol):
    for c in range(scol, ecol + 1):
        cell = ws.cell(row=row, column=c)
        cell.number_format = PCT_FMT
        cell.font = PCT_FONT
        cell.border = THIN_BORDER


# ═══════════════════════════════════════════════════════════════════════════════
#  PHASE 2: INSERT ROWS (no formula writing yet — just structural inserts)
# ═══════════════════════════════════════════════════════════════════════════════

def insert_tb_rows(wb):
    """Insert 2 rows in TB for Opening Stock (row 13) and Closing Stock (row 17)."""
    ws = wb['TB']
    ws.insert_rows(13, 1)   # Opening Stock → row 13
    ws.insert_rows(17, 1)   # Closing Stock → row 17
    print("    TB: Inserted rows 13 (Opening Stock) and 17 (Closing Stock)")


def insert_pnl_rows(wb):
    """
    Insert 3 rows in P&L for Opening Stock, Total Purchases, and Closing Stock.

    Original P&L COGS area:
      19: Cost of Goods Sold (header)
      20: Purchase of RM          → after insert 1 at 20 → 21
      21: Import of RM            → 22
      22: Other Purchases         → 23
      23: Total COGS              → 24 (after 1st); then insert 2 at 24 → 26
      24: blank                   → 25 → 27
      25: Direct / Manufacturing  → 26 → 28

    After all inserts:
      19: header
      20: EMPTY → Opening Stock
      21: Purchase of RM
      22: Import of RM
      23: Other Purchases
      24: EMPTY → Total Purchases
      25: EMPTY → Closing Stock
      26: Total COGS (moved)
      27: blank (moved)
      28: Direct / Manufacturing (moved)
    """
    ws = wb['P&L']
    ws.insert_rows(20, 1)   # Opening Stock slot → row 20
    ws.insert_rows(24, 2)   # Total Purchases (24) + Closing Stock (25) slots
    print("    P&L: Inserted 1 row at 20 + 2 rows at 24")


# ═══════════════════════════════════════════════════════════════════════════════
#  PHASE 5: WRITE NEW CONTENT (after formula fixer has run)
# ═══════════════════════════════════════════════════════════════════════════════

def write_tb_new_content(wb):
    """Fill Opening Stock (row 13), Closing Stock (row 17), and rewrite COGS formula."""
    ws = wb['TB']
    last_col = 14  # N = FY Total

    # Opening Stock (row 13) — yellow manual input
    ws.cell(row=13, column=1, value="Opening Stock").font = NORMAL_FONT
    for c in range(2, 14):
        cell = ws.cell(row=13, column=c)
        cell.fill = INPUT_FILL
        cell.number_format = INR2
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
    ws.cell(row=13, column=14, value="=SUM(B13:M13)").number_format = INR2
    ws.cell(row=13, column=14).border = THIN_BORDER

    # Closing Stock (row 17) — yellow manual input
    ws.cell(row=17, column=1, value="Closing Stock").font = NORMAL_FONT
    for c in range(2, 14):
        cell = ws.cell(row=17, column=c)
        cell.fill = INPUT_FILL
        cell.number_format = INR2
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
    ws.cell(row=17, column=14, value="=SUM(B17:M17)").number_format = INR2
    ws.cell(row=17, column=14).border = THIN_BORDER

    # Find Total COGS row by label (it was shifted by insert_rows to a new position)
    # After inserts, the label is still "Total COGS" but at a new row.
    # After formula fixer, its formula has been adjusted.
    # We need to find it and OVERWRITE with our new COGS formula.
    total_cogs_row = None
    for r in range(70, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and "Total COGS" in str(val):
            total_cogs_row = r
            break

    if total_cogs_row:
        print(f"    TB: Total COGS found at row {total_cogs_row}")
        ws.cell(row=total_cogs_row, column=1,
                value="Total COGS (Opening + Purchases \u2212 Closing Stock)")
        for c in range(2, last_col + 1):
            col = cl(c)
            if c <= 13:
                # Opening Stock + 3 purchases - Closing Stock
                ws.cell(row=total_cogs_row, column=c,
                        value=f"={col}13+{col}14+{col}15+{col}16-{col}17")
            else:
                ws.cell(row=total_cogs_row, column=c,
                        value=f"=SUM(B{total_cogs_row}:M{total_cogs_row})")
            ws.cell(row=total_cogs_row, column=c).number_format = INR2

    # Build row_map for new TB layout (after inserts + formula fixes)
    row_map = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val:
            row_map[str(val).strip()] = r

    # Update COGS note
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and "COGS (Cost of Goods Sold)" in str(val) and "domestic purchases" in str(val).lower():
            ws.cell(row=r, column=1,
                    value="  \u2022  COGS (Cost of Goods Sold): Opening Stock + Purchases \u2212 Closing Stock. Enter stock values manually each month.")

    return row_map


def write_pnl_new_content(wb, tb_row_map):
    """
    Fill P&L COGS breakdown rows.
    After inserts + formula fixes:
      20: EMPTY → Opening Stock
      21: Purchase of RM (formula already fixed to =TB!B14)
      22: Import of RM (=TB!B15)
      23: Other Purchases (=TB!B16)
      24: EMPTY → Total Purchases
      25: EMPTY → Closing Stock
      26: Total COGS (formula already fixed, we'll overwrite)
    """
    ws = wb['P&L']
    last_col = 14  # N = FY Total

    # Row 20: Opening Stock
    ws.cell(row=20, column=1, value="Opening Stock").font = NORMAL_FONT
    for c in range(2, last_col + 1):
        ws.cell(row=20, column=c, value=f"=TB!{cl(c)}13").number_format = INR2
        ws.cell(row=20, column=c).border = THIN_BORDER

    # Row 24: Total Purchases
    ws.cell(row=24, column=1, value="Total Purchases").font = TOTAL_FONT
    for c in range(2, last_col + 1):
        col = cl(c)
        ws.cell(row=24, column=c, value=f"={col}21+{col}22+{col}23").number_format = INR2
    style_total_row(ws, 24, last_col)

    # Row 25: Less: Closing Stock
    ws.cell(row=25, column=1, value="Less: Closing Stock").font = NORMAL_FONT
    for c in range(2, last_col + 1):
        ws.cell(row=25, column=c, value=f"=TB!{cl(c)}17").number_format = INR2
        ws.cell(row=25, column=c).border = THIN_BORDER

    # Row 26: Total COGS — overwrite with explicit reference to TB Total COGS
    total_cogs_tb = tb_row_map.get("Total COGS (Opening + Purchases \u2212 Closing Stock)")
    if total_cogs_tb:
        ws.cell(row=26, column=1, value="Total COGS").font = TOTAL_FONT
        for c in range(2, last_col + 1):
            ws.cell(row=26, column=c, value=f"=TB!{cl(c)}{total_cogs_tb}").number_format = INR2
        style_total_row(ws, 26, last_col)


def modify_balance_sheet(wb):
    """Rename 'Opening Stock' to 'Inventories' in Balance Sheet."""
    ws = wb['Balance Sheet']
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and str(val).strip() == "Opening Stock":
            ws.cell(row=r, column=1, value="Inventories")
            print(f"    BS row {r}: 'Opening Stock' \u2192 'Inventories'")
            break


# ═══════════════════════════════════════════════════════════════════════════════
#  PHASE 6: CRM DATA + PROJECTION SHEETS
# ═══════════════════════════════════════════════════════════════════════════════

def read_crm_data():
    wb_crm = openpyxl.load_workbook(str(CRM_FILE), data_only=True)
    ws = wb_crm.active
    deals = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        pipeline = ws.cell(r, 3).value
        stage    = ws.cell(r, 4).value
        quarter  = str(ws.cell(r, 5).value) if ws.cell(r, 5).value else "Unknown"
        exp_rev  = ws.cell(r, 6).value or 0
        amount   = ws.cell(r, 7).value or 0
        prob_raw = ws.cell(r, 8).value or 0
        try:
            prob = float(prob_raw) / 100.0
        except (ValueError, TypeError):
            prob = 0.0
        deals.append({
            'name': name, 'pipeline': pipeline, 'stage': stage,
            'quarter': quarter, 'expected_revenue': float(exp_rev),
            'amount': float(amount), 'probability': prob,
            'weighted_revenue': float(amount) * prob,
        })
    wb_crm.close()
    return deals


def build_crm_data_sheet(wb, deals):
    ws = wb.create_sheet("CRM Data")
    ws.sheet_properties.tabColor = ORANGE

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="CRM Pipeline Data  |  Zoho CRM Export  |  As of March 2026").font = SUBTITLE_FONT

    headers = ['Deal Name', 'Stage', 'Expected Closure', 'Deal Amount (\u20b9)',
               'Expected Revenue (\u20b9)', 'Probability (%)', 'Weighted Revenue (\u20b9)']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, len(headers))

    stage_order = {
        'PO Received': 0, 'Awaiting PO': 1, 'Negotiation / Tendering': 2,
        'Proposal / Quote Sent': 3, 'Demo / Samples Done': 4,
        'Discovery Meeting Done': 5, 'Sales Qualified Lead': 6,
    }
    deals_sorted = sorted(deals, key=lambda d: (stage_order.get(d['stage'], 99), -d['weighted_revenue']))

    r = 5
    for d in deals_sorted:
        ws.cell(row=r, column=1, value=d['name']).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=d['stage']).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=d['quarter']).font = NORMAL_FONT
        ws.cell(row=r, column=4, value=d['amount']).number_format = INR
        ws.cell(row=r, column=5, value=d['expected_revenue']).number_format = INR
        ws.cell(row=r, column=6, value=d['probability']).number_format = PCT_FMT
        ws.cell(row=r, column=7, value=d['weighted_revenue']).number_format = INR
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    last_data = r - 1
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=4, value=f"=SUM(D5:D{last_data})").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(E5:E{last_data})").number_format = INR
    ws.cell(row=r, column=7, value=f"=SUM(G5:G{last_data})").number_format = INR
    style_total_row(ws, r, 7)
    r += 2

    # Summary by Stage
    ws.cell(row=r, column=1, value="PIPELINE SUMMARY BY STAGE").font = SECTION_FONT
    style_section_row(ws, r, 7)
    r += 1
    for ci, h in enumerate(['Stage', 'Count', 'Total Amount', 'Weighted Rev'], 1):
        ws.cell(row=r, column=ci, value=h)
    style_header_row(ws, r, 4)
    r += 1

    stage_agg = {}
    for d in deals:
        s = d['stage'] or "Unknown"
        if s not in stage_agg:
            stage_agg[s] = {'count': 0, 'amount': 0, 'weighted': 0}
        stage_agg[s]['count'] += 1
        stage_agg[s]['amount'] += d['amount']
        stage_agg[s]['weighted'] += d['weighted_revenue']

    for sn in ['PO Received', 'Awaiting PO', 'Negotiation / Tendering',
               'Proposal / Quote Sent', 'Demo / Samples Done',
               'Discovery Meeting Done', 'Sales Qualified Lead']:
        if sn in stage_agg:
            data = stage_agg[sn]
            ws.cell(row=r, column=1, value=sn).font = NORMAL_FONT
            ws.cell(row=r, column=2, value=data['count']).border = THIN_BORDER
            ws.cell(row=r, column=3, value=data['amount']).number_format = INR
            ws.cell(row=r, column=4, value=data['weighted']).number_format = INR
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = THIN_BORDER
            r += 1

    r += 1
    # Summary by Quarter
    ws.cell(row=r, column=1, value="PIPELINE SUMMARY BY EXPECTED CLOSURE").font = SECTION_FONT
    style_section_row(ws, r, 7)
    r += 1
    for ci, h in enumerate(['Quarter', 'Count', 'Total Amount', 'Weighted Rev', 'Expected Rev'], 1):
        ws.cell(row=r, column=ci, value=h)
    style_header_row(ws, r, 5)
    r += 1

    quarter_agg = {}
    for d in deals:
        q = d['quarter']
        if q not in quarter_agg:
            quarter_agg[q] = {'count': 0, 'amount': 0, 'weighted': 0, 'expected': 0}
        quarter_agg[q]['count'] += 1
        quarter_agg[q]['amount'] += d['amount']
        quarter_agg[q]['weighted'] += d['weighted_revenue']
        quarter_agg[q]['expected'] += d['expected_revenue']

    for qn in ['Q3', 'Q4', 'Q1', 'Next Financial Year']:
        if qn in quarter_agg:
            data = quarter_agg[qn]
            ws.cell(row=r, column=1, value=qn).font = NORMAL_FONT
            ws.cell(row=r, column=2, value=data['count']).border = THIN_BORDER
            ws.cell(row=r, column=3, value=data['amount']).number_format = INR
            ws.cell(row=r, column=4, value=data['weighted']).number_format = INR
            ws.cell(row=r, column=5, value=data['expected']).number_format = INR
            for c in range(1, 6):
                ws.cell(row=r, column=c).border = THIN_BORDER
            r += 1

    r += 2
    notes = [
        "\U0001f4cb  Understanding the CRM Pipeline",
        "  \u2022  Data exported from Zoho CRM as of March 2026.",
        "  \u2022  Deal Amount = full value if won. Weighted Revenue = Amount \u00d7 Probability.",
        "  \u2022  PO Received=100%, Awaiting PO=80%, Negotiation=50%, Proposal=20%, Demo=15%, SQL=5%, Discovery=10%.",
        "  \u2022  Q3=Jan-Mar FY25-26 (current). Q4=Apr-Jun FY26-27 (next). Pipeline feeds Sales Projection & Q1 Projection.",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=r + i, column=1, value=n).font = NOTE_FONT

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 28
    for c in 'CDEFG':
        ws.column_dimensions[c].width = 18
    ws.freeze_panes = 'A5'

    return quarter_agg, stage_agg


def build_sales_projection_sheet(wb, deals, quarter_agg):
    ws = wb.create_sheet("Sales Projection")
    ws.sheet_properties.tabColor = DARK_GREEN

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Sales Projection  |  Q1 FY 2026-27 (Apr\u2013Jun)  |  Based on CRM Pipeline").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value="Source: Zoho CRM Pipeline as of March 2026. Revenue distributed across Q1 months.").font = SMALL_FONT

    headers = ['Particulars', 'APR', 'MAY', 'JUN', 'Q1 Total']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=5, column=ci, value=h)
    style_header_row(ws, 5, 5)

    q4_deals = [d for d in deals if d['quarter'] == 'Q4']
    q3_po = [d for d in deals if d['quarter'] == 'Q3' and d['stage'] == 'PO Received']

    high_conf = [d for d in q4_deals if d['stage'] in ('PO Received', 'Awaiting PO')]
    med_conf  = [d for d in q4_deals if d['stage'] == 'Negotiation / Tendering']
    low_conf  = [d for d in q4_deals if d['stage'] not in ('PO Received','Awaiting PO','Negotiation / Tendering')]

    r = 6

    # A. HIGH CONFIDENCE
    ws.cell(row=r, column=1, value="A. CONFIRMED / HIGH PROBABILITY DEALS").font = SECTION_FONT
    style_section_row(ws, r, 5)
    r += 1
    high_start = r
    for d in sorted(high_conf, key=lambda x: -x['weighted_revenue']):
        ws.cell(row=r, column=1, value=f"  {d['name']} ({d['stage']})").font = NORMAL_FONT
        w = d['weighted_revenue']
        ws.cell(row=r, column=2, value=round(w * 0.30)).number_format = INR
        ws.cell(row=r, column=3, value=round(w * 0.40)).number_format = INR
        ws.cell(row=r, column=4, value=round(w * 0.30)).number_format = INR
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    if q3_po:
        ws.cell(row=r, column=1, value="  Q3 PO spill-over (delivery in Apr)").font = NORMAL_FONT
        total_q3 = sum(d['weighted_revenue'] for d in q3_po)
        ws.cell(row=r, column=2, value=round(total_q3 * 0.5)).number_format = INR
        ws.cell(row=r, column=3, value=round(total_q3 * 0.3)).number_format = INR
        ws.cell(row=r, column=4, value=round(total_q3 * 0.2)).number_format = INR
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    high_end = r - 1
    ws.cell(row=r, column=1, value="Sub-total: High Confidence").font = TOTAL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c, value=f"=SUM({cl(c)}{high_start}:{cl(c)}{high_end})").number_format = INR
    style_total_row(ws, r, 5)
    HIGH_TOTAL = r
    r += 2

    # B. MEDIUM CONFIDENCE
    ws.cell(row=r, column=1, value="B. MEDIUM PROBABILITY DEALS (Negotiation)").font = SECTION_FONT
    style_section_row(ws, r, 5)
    r += 1
    med_start = r
    for d in sorted(med_conf, key=lambda x: -x['weighted_revenue']):
        ws.cell(row=r, column=1, value=f"  {d['name']}").font = NORMAL_FONT
        w = d['weighted_revenue']
        ws.cell(row=r, column=2, value=round(w * 0.25)).number_format = INR
        ws.cell(row=r, column=3, value=round(w * 0.35)).number_format = INR
        ws.cell(row=r, column=4, value=round(w * 0.40)).number_format = INR
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    med_end = r - 1
    ws.cell(row=r, column=1, value="Sub-total: Medium Confidence").font = TOTAL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c, value=f"=SUM({cl(c)}{med_start}:{cl(c)}{med_end})").number_format = INR
    style_total_row(ws, r, 5)
    MED_TOTAL = r
    r += 2

    # C. LOW CONFIDENCE
    ws.cell(row=r, column=1, value="C. LOW PROBABILITY DEALS (Proposals / Demos / Leads)").font = SECTION_FONT
    style_section_row(ws, r, 5)
    r += 1
    low_start = r
    for d in sorted(low_conf, key=lambda x: -x['weighted_revenue']):
        ws.cell(row=r, column=1, value=f"  {d['name']} ({d['stage']})").font = NORMAL_FONT
        w = d['weighted_revenue']
        ws.cell(row=r, column=2, value=round(w * 0.20)).number_format = INR
        ws.cell(row=r, column=3, value=round(w * 0.35)).number_format = INR
        ws.cell(row=r, column=4, value=round(w * 0.45)).number_format = INR
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    low_end = r - 1
    ws.cell(row=r, column=1, value="Sub-total: Low Confidence").font = TOTAL_FONT
    for c in range(2, 6):
        if low_start <= low_end:
            ws.cell(row=r, column=c, value=f"=SUM({cl(c)}{low_start}:{cl(c)}{low_end})").number_format = INR
        else:
            ws.cell(row=r, column=c, value=0).number_format = INR
    style_total_row(ws, r, 5)
    LOW_TOTAL = r
    r += 2

    # GRAND TOTAL
    ws.cell(row=r, column=1, value="TOTAL PROJECTED REVENUE (Q1 FY 2026-27)").font = GRAND_TOTAL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c,
                value=f"={cl(c)}{HIGH_TOTAL}+{cl(c)}{MED_TOTAL}+{cl(c)}{LOW_TOTAL}").number_format = INR
    style_total_row(ws, r, 5, grand=True)
    PROJ_REV = r
    r += 2

    # Scenarios
    ws.cell(row=r, column=1, value="REVENUE SCENARIOS").font = SECTION_FONT
    style_section_row(ws, r, 5)
    r += 1

    ws.cell(row=r, column=1, value="Conservative (High Only)").font = NORMAL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c, value=f"={cl(c)}{HIGH_TOTAL}").number_format = INR
    fmt_number_cells(ws, r, 2, 5)
    r += 1

    ws.cell(row=r, column=1, value="Most Likely (High + Medium)").font = NORMAL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c, value=f"={cl(c)}{HIGH_TOTAL}+{cl(c)}{MED_TOTAL}").number_format = INR
    fmt_number_cells(ws, r, 2, 5)
    MOST_LIKELY = r
    r += 1

    ws.cell(row=r, column=1, value="Optimistic (All Pipeline)").font = NORMAL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c, value=f"={cl(c)}{PROJ_REV}").number_format = INR
    fmt_number_cells(ws, r, 2, 5)
    r += 2

    notes = [
        "\U0001f4ca  Understanding Sales Projections",
        "  \u2022  All projections from Zoho CRM pipeline weighted by deal probability.",
        "  \u2022  High: PO Received (100%) + Awaiting PO (80%). Near-certain revenue.",
        "  \u2022  Medium: Negotiation/Tendering (50%). Actively being discussed.",
        "  \u2022  Low: Proposals/Demos/Leads (<20%). Speculative.",
        "  \u2022  Revenue distributed across Apr-Jun by typical delivery patterns.",
        "  \u2022  Conservative = confirmed deals only (cash planning).",
        "  \u2022  Most Likely = confirmed + negotiation (best estimate).",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=r + i, column=1, value=n).font = NOTE_FONT

    ws.column_dimensions['A'].width = 55
    for c in range(2, 6):
        ws.column_dimensions[cl(c)].width = 18
    ws.freeze_panes = 'B6'

    return {'most_likely_row': MOST_LIKELY, 'high_total': HIGH_TOTAL}


def build_q1_projection_sheet(wb, proj_rows, tb_row_map):
    """Q1 FY2026-27 Projected P&L: Revenue from CRM, expenses from FY avg +10%."""
    ws = wb.create_sheet("Q1 FY26-27 Projection")
    ws.sheet_properties.tabColor = "7030A0"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Projected P&L  |  Q1 FY 2026-27 (Apr\u2013Jun)  |  CRM Pipeline + Historical Costs").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value="Revenue from CRM (Most Likely). Expenses at FY25-26 avg monthly +10%.").font = SMALL_FONT

    headers = ['Particulars', 'APR (P)', 'MAY (P)', 'JUN (P)', 'Q1 Total', 'Basis / Assumption']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=5, column=ci, value=h)
    style_header_row(ws, 5, 6)

    most_likely = proj_rows['most_likely_row']

    # TB row references (already in NEW row space after inserts)
    tb_total_rev   = tb_row_map.get("Total Revenue")
    tb_total_cogs  = tb_row_map.get("Total COGS (Opening + Purchases \u2212 Closing Stock)") or tb_row_map.get("Total COGS")
    tb_total_direct= tb_row_map.get("Total Direct Expenses")
    tb_admin  = tb_row_map.get("Admin Overheads Total")
    tb_finance= tb_row_map.get("Finance Cost Total")
    tb_hr     = tb_row_map.get("HR Expenses Total")
    tb_mktg   = tb_row_map.get("Marketing Total")
    tb_rnd    = tb_row_map.get("R&D Total")
    tb_other  = tb_row_map.get("Other OPEX (Professional, Round Off, Write Off)")
    tb_opex   = tb_row_map.get("Total Indirect Expenses (OPEX)")
    tb_gp     = tb_row_map.get("Gross Profit")
    tb_ebitda = tb_row_map.get("EBITDA")

    r = 6

    # I. REVENUE
    ws.cell(row=r, column=1, value="I. REVENUE").font = SECTION_FONT
    style_section_row(ws, r, 6)
    r += 1
    ws.cell(row=r, column=1, value="Projected Revenue (Most Likely)").font = NORMAL_FONT
    ws.cell(row=r, column=6, value="CRM: High + Medium confidence deals").font = SMALL_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=f"='Sales Projection'!{cl(c)}{most_likely}").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    fmt_number_cells(ws, r, 2, 5)
    REV = r
    r += 2

    # II. COGS
    ws.cell(row=r, column=1, value="II. COST OF GOODS SOLD (PROJECTED)").font = SECTION_FONT
    style_section_row(ws, r, 6)
    r += 1

    ws.cell(row=r, column=1, value="Opening Stock").font = NORMAL_FONT
    ws.cell(row=r, column=6, value="Manual input \u2014 physical stock at month start").font = SMALL_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=c).number_format = INR
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    OS = r
    r += 1

    ws.cell(row=r, column=1, value="Projected Purchases").font = NORMAL_FONT
    if tb_total_cogs and tb_total_rev:
        ws.cell(row=r, column=6, value="FY25-26 COGS/Revenue ratio \u00d7 projected revenue").font = SMALL_FONT
        for c in range(2, 5):
            ws.cell(row=r, column=c,
                    value=f"=IFERROR({cl(c)}{REV}*(TB!N{tb_total_cogs}/TB!N{tb_total_rev}),0)").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    fmt_number_cells(ws, r, 2, 5)
    PURCH = r
    r += 1

    ws.cell(row=r, column=1, value="Closing Stock").font = NORMAL_FONT
    ws.cell(row=r, column=6, value="Manual input \u2014 physical stock at month end").font = SMALL_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=c).number_format = INR
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    CS = r
    r += 1

    ws.cell(row=r, column=1, value="Total COGS (Opening + Purchases \u2212 Closing)").font = TOTAL_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=f"={cl(c)}{OS}+{cl(c)}{PURCH}-{cl(c)}{CS}").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    style_total_row(ws, r, 5)
    COGS = r
    r += 2

    # III. DIRECT
    ws.cell(row=r, column=1, value="III. DIRECT / MANUFACTURING EXPENSES").font = SECTION_FONT
    style_section_row(ws, r, 6)
    r += 1
    ws.cell(row=r, column=1, value="Projected Direct Expenses").font = NORMAL_FONT
    if tb_total_direct and tb_total_rev:
        ws.cell(row=r, column=6, value="FY25-26 Direct/Revenue ratio applied").font = SMALL_FONT
        for c in range(2, 5):
            ws.cell(row=r, column=c,
                    value=f"=IFERROR({cl(c)}{REV}*(TB!N{tb_total_direct}/TB!N{tb_total_rev}),0)").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    fmt_number_cells(ws, r, 2, 5)
    DIRECT = r
    r += 2

    # GROSS PROFIT
    ws.cell(row=r, column=1, value="GROSS PROFIT").font = TOTAL_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=f"={cl(c)}{REV}-{cl(c)}{COGS}-{cl(c)}{DIRECT}").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    style_total_row(ws, r, 5)
    GP = r
    r += 1
    ws.cell(row=r, column=1, value="Gross Profit %").font = PCT_FONT
    for c in range(2, 6):
        ref_r = f"{cl(c)}{REV}" if c <= 4 else f"E{REV}"
        ref_g = f"{cl(c)}{GP}" if c <= 4 else f"E{GP}"
        ws.cell(row=r, column=c, value=f"=IFERROR({ref_g}/{ref_r},0)").number_format = PCT_FMT
    fmt_pct_cells(ws, r, 2, 5)
    r += 2

    # IV. OPEX
    ws.cell(row=r, column=1, value="IV. INDIRECT EXPENSES (OPEX) \u2014 FY Avg +10%").font = SECTION_FONT
    style_section_row(ws, r, 6)
    r += 1

    opex_items = [
        ("Admin Overheads", tb_admin),
        ("Finance Cost", tb_finance),
        ("HR / Payroll Expenses", tb_hr),
        ("Marketing & Ads", tb_mktg),
        ("R&D", tb_rnd),
        ("Other OPEX", tb_other),
    ]
    opex_s = r
    for label, tb_r in opex_items:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=6, value="FY25-26 monthly avg \u00d7 1.10").font = SMALL_FONT
        if tb_r:
            for c in range(2, 5):
                ws.cell(row=r, column=c, value=f"=IFERROR(TB!N{tb_r}/12*1.1,0)").number_format = INR
        ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
        fmt_number_cells(ws, r, 2, 5)
        r += 1
    opex_e = r - 1

    ws.cell(row=r, column=1, value="Total OPEX").font = TOTAL_FONT
    for c in range(2, 6):
        ws.cell(row=r, column=c, value=f"=SUM({cl(c)}{opex_s}:{cl(c)}{opex_e})").number_format = INR
    style_total_row(ws, r, 5)
    OPEX = r
    r += 2

    # EBITDA
    ws.cell(row=r, column=1, value="EBITDA").font = GRAND_TOTAL_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=f"={cl(c)}{GP}-{cl(c)}{OPEX}").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    style_total_row(ws, r, 5, grand=True)
    EBITDA = r
    r += 1
    ws.cell(row=r, column=1, value="EBITDA %").font = PCT_FONT
    for c in range(2, 6):
        ref_r = f"{cl(c)}{REV}" if c <= 4 else f"E{REV}"
        ref_e = f"{cl(c)}{EBITDA}" if c <= 4 else f"E{EBITDA}"
        ws.cell(row=r, column=c, value=f"=IFERROR({ref_e}/{ref_r},0)").number_format = PCT_FMT
    fmt_pct_cells(ws, r, 2, 5)
    r += 2

    # Depreciation + PBT
    ws.cell(row=r, column=1, value="Depreciation").font = NORMAL_FONT
    ws.cell(row=r, column=6, value="Manual input").font = SMALL_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=c).number_format = INR
        ws.cell(row=r, column=c).border = THIN_BORDER
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    DEP = r
    r += 1

    ws.cell(row=r, column=1, value="Profit Before Tax (PBT)").font = TOTAL_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=f"={cl(c)}{EBITDA}-{cl(c)}{DEP}").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").number_format = INR
    style_total_row(ws, r, 5, grand=True)
    PBT = r
    r += 1

    ws.cell(row=r, column=1, value="PBT %").font = PCT_FONT
    for c in range(2, 6):
        ref_r = f"{cl(c)}{REV}" if c <= 4 else f"E{REV}"
        ref_p = f"{cl(c)}{PBT}" if c <= 4 else f"E{PBT}"
        ws.cell(row=r, column=c, value=f"=IFERROR({ref_p}/{ref_r},0)").number_format = PCT_FMT
    fmt_pct_cells(ws, r, 2, 5)
    r += 2

    # Comparison table
    ws.cell(row=r, column=1, value="COMPARISON: FY25-26 ACTUAL vs Q1 PROJECTION").font = SECTION_FONT
    style_section_row(ws, r, 6)
    r += 1
    for ci, h in enumerate(['Metric', 'FY25-26 Mo. Avg', 'Q1 Proj Mo. Avg', 'Change', 'Change %'], 1):
        ws.cell(row=r, column=ci, value=h)
    style_header_row(ws, r, 5)
    r += 1

    comp = [
        ("Revenue",       tb_total_rev,    REV),
        ("COGS",          tb_total_cogs,   COGS),
        ("Direct Exp.",   tb_total_direct, DIRECT),
        ("Gross Profit",  tb_gp,           GP),
        ("Total OPEX",    tb_opex,         OPEX),
        ("EBITDA",        tb_ebitda,       EBITDA),
    ]
    for label, tb_r, proj_r in comp:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        if tb_r:
            ws.cell(row=r, column=2, value=f"=IFERROR(TB!N{tb_r}/12,0)").number_format = INR
        ws.cell(row=r, column=3, value=f"=IFERROR(E{proj_r}/3,0)").number_format = INR
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = INR
        ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/B{r},0)").number_format = PCT_FMT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    r += 2
    notes = [
        "\U0001f680  Understanding the Q1 Projection",
        "  \u2022  Revenue from CRM Sales Projection (Most Likely = High + Medium confidence).",
        "  \u2022  COGS projected using FY25-26 COGS-to-Revenue ratio. Opening & Closing Stock are manual inputs.",
        "  \u2022  Direct Expenses projected using FY25-26 Direct/Revenue ratio.",
        "  \u2022  OPEX = FY25-26 average monthly spend \u00d7 1.10 (10% increase for growth/inflation).",
        "  \u2022  This is a projection, not a budget. Actuals will vary with deal closures and operations.",
        "  \u2022  Positive EBITDA trajectory indicates business on track for profitability.",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=r + i, column=1, value=n).font = NOTE_FONT

    ws.column_dimensions['A'].width = 50
    for c in range(2, 7):
        ws.column_dimensions[cl(c)].width = 22
    ws.freeze_panes = 'B6'


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("Fracktal MIS Modifier v2 — COGS Fix + CRM Projections")
    print(f"  Source: {SOURCE_MIS}")
    print(f"  CRM:    {CRM_FILE}")
    print(f"  Output: {OUTPUT_MIS}")

    # Phase 1: Copy
    print("\n  Phase 1: Copying source workbook...")
    shutil.copy2(str(SOURCE_MIS), str(OUTPUT_MIS))
    wb = openpyxl.load_workbook(str(OUTPUT_MIS))

    # Phase 2: Insert rows (structural changes only)
    print("\n  Phase 2: Inserting rows...")
    insert_tb_rows(wb)
    insert_pnl_rows(wb)

    # Phase 3: Fix ALL formulas in ALL sheets
    print("\n  Phase 3: Fixing formulas after row insertions...")
    n = fix_all_formulas(wb)
    print(f"    Fixed {n} formula(s) across all sheets")

    # Phase 4: Write new content
    print("\n  Phase 4: Writing new content...")
    tb_map = write_tb_new_content(wb)
    write_pnl_new_content(wb, tb_map)
    modify_balance_sheet(wb)

    # Phase 5: CRM + Projections
    print("\n  Phase 5: CRM Data & Projections...")
    deals = read_crm_data()
    print(f"    Loaded {len(deals)} CRM deals")
    q_agg, s_agg = build_crm_data_sheet(wb, deals)
    proj_rows = build_sales_projection_sheet(wb, deals, q_agg)
    build_q1_projection_sheet(wb, proj_rows, tb_map)

    # Phase 6: Reorder & save
    desired = ['TB', 'P&L', 'Cash Flow', 'OPEX Schedule', 'Balance Sheet',
               'Performance Summary', 'KPIs', 'Dashboard',
               'CRM Data', 'Sales Projection', 'Q1 FY26-27 Projection']
    for idx, name in enumerate(desired):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            wb.move_sheet(name, offset=idx - cur)

    print(f"\n  Phase 6: Saving to {OUTPUT_MIS}")
    wb.save(str(OUTPUT_MIS))
    print(f"  Sheets: {wb.sheetnames}")
    print("  Done! ✓")
    return str(OUTPUT_MIS)


if __name__ == "__main__":
    main()
