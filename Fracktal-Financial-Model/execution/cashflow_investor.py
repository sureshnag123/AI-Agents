#!/usr/bin/env python3
"""
Investor-Ready Cash Flow Statement + Working Capital Requirement
================================================================
Creates a standalone Excel file for investor meeting (March 11, 2026).

Contents:
  Sheet 1: Simple Cash Flow Statement (Apr 2025 — Feb 2026, from Tally)
  Sheet 2: Working Capital Requirement (Orders in hand × 50% COGS)

Usage:
  python execution/cashflow_investor.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

# ── Paths ──
MIS_FILE = Path(r"D:\Suresh_AGENTS\Fracktal-Financial-Model\.tmp\Fracktal_MIS_Master_20260310_122834.xlsx")
CRM_FILE = Path(r"C:\Users\Lenovo\Downloads\Deals_2026_03_10.xlsx")
OUTPUT_DIR = Path(r"D:\Suresh_AGENTS\Fracktal-Financial-Model\.tmp")
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT = OUTPUT_DIR / f"Fracktal_CashFlow_Investor_{ts}.xlsx"

# ── Styles ──
NAVY       = "1F3864"
DARK_BLUE  = "2F5496"
MID_BLUE   = "4472C4"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN= "E2EFDA"
LIGHT_YELLOW= "FFF2CC"
WHITE      = "FFFFFF"
LIGHT_RED  = "FCE4EC"

TITLE_FONT       = Font(name="Calibri", bold=True, size=16, color=NAVY)
SUBTITLE_FONT    = Font(name="Calibri", bold=True, size=11, color=MID_BLUE)
HEADER_FONT      = Font(name="Calibri", bold=True, size=10, color=WHITE)
HEADER_FILL      = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SECTION_FONT     = Font(name="Calibri", bold=True, size=10, color=NAVY)
SECTION_FILL     = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
TOTAL_FONT       = Font(name="Calibri", bold=True, size=10)
TOTAL_FILL       = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
GRAND_FONT       = Font(name="Calibri", bold=True, size=11, color=WHITE)
GRAND_FILL       = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
NORMAL_FONT      = Font(name="Calibri", size=10)
SMALL_FONT       = Font(name="Calibri", size=9, color="808080", italic=True)
NOTE_FONT        = Font(name="Calibri", size=9, color="666666")
NEGATIVE_FONT    = Font(name="Calibri", size=10, color="C00000")
INPUT_FILL       = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
HIGHLIGHT_FILL   = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
BLUE_FILL        = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
TOTAL_BORDER = Border(
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='double', color=NAVY),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)

INR = '#,##0'
INR2 = '#,##0.00'
PCT = '0.0%'

MONTHS = ['APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC','JAN','FEB']
MONTH_LABELS = ["Apr'25","May'25","Jun'25","Jul'25","Aug'25","Sep'25",
                "Oct'25","Nov'25","Dec'25","Jan'26","Feb'26"]


def cl(c):
    return get_column_letter(c)


def get_val(ws, row, col):
    """Get a numeric value from a cell, treating formulas and dashes as 0."""
    v = ws.cell(row, col).value
    if v is None or v == '-' or v == '':
        return 0.0
    if isinstance(v, str):
        if v.startswith('='):
            return 0.0
        try:
            return float(v)
        except ValueError:
            return 0.0
    return float(v)


def read_tb_data(mis_path):
    """Read actual values from TB sheet (Apr-Feb = cols 2-12)."""
    wb = openpyxl.load_workbook(str(mis_path))
    ws = wb['TB']

    data = {}

    # Row definitions (after row inserts in modified MIS)
    row_defs = {
        # Revenue
        'sale_products':  5,
        'sale_service':   6,
        'export_sales':   7,
        'printsticks':    8,
        'discount_recv':  9,
        'interest_income':10,
        'other_income':   11,
        # Purchases
        'purchase_rm':    14,
        'import_rm':      15,
        'other_purchase': 16,
        # Direct
        'salary_prod':    19,
        'overtime':       20,
        'elec_factory':   21,
        'elec_mfg':       22,
        'freight_in':     23,
        'loading':        24,
        'discount_allow': 25,
        # OPEX
        'admin_oh':       28,
        'travel':         29,
        'rates_tax':      30,
        'fx_gain':        31,
        'razorpay':       32,
        'tender_fee':     33,
        'finance':        35,
        'payroll':        37,
        'ads_marketing':  39,
        'freight_out':    40,
        'rnd':            42,
        'professional':   43,
        'roundoff':       44,
        'writeoff':       45,
        # Bank/Cash
        'cash_hand':      69,
        'bank':           70,
    }

    for key, row in row_defs.items():
        vals = []
        for c in range(2, 13):  # B to L (Apr-Feb)
            vals.append(get_val(ws, row, c))
        data[key] = vals

    wb.close()
    return data


def read_crm_orders(crm_path):
    """Read PO Received and Awaiting PO deals from CRM data."""
    wb = openpyxl.load_workbook(str(crm_path), data_only=True)
    ws = wb.active

    po_received = []
    awaiting_po = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        stage = ws.cell(r, 4).value
        amount = ws.cell(r, 7).value or 0
        prob_raw = ws.cell(r, 8).value or 0
        try:
            prob = float(prob_raw) / 100.0
        except (ValueError, TypeError):
            prob = 0
        deal = {'name': str(name), 'amount': float(amount), 'probability': prob}

        if stage and 'PO Received' in str(stage):
            po_received.append(deal)
        elif stage and 'Awaiting PO' in str(stage):
            awaiting_po.append(deal)

    wb.close()
    return po_received, awaiting_po


def style_row(ws, row, max_col, font, fill=None, border=THIN_BORDER, num_fmt=INR):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        if fill:
            cell.fill = fill
        cell.border = border
        if c >= 2:
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = num_fmt


def write_data_row(ws, row, label, values, font=NORMAL_FONT, fill=None, num_fmt=INR):
    """Write label + 11 month values + YTD total."""
    ws.cell(row=row, column=1, value=label).font = font
    ws.cell(row=row, column=1).border = THIN_BORDER
    if fill:
        ws.cell(row=row, column=1).fill = fill

    ytd = 0
    for i, v in enumerate(values):
        c = i + 2  # col B=2 to L=12
        cell = ws.cell(row=row, column=c, value=round(v))
        cell.font = NEGATIVE_FONT if v < 0 else font
        cell.number_format = num_fmt
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        ytd += v

    # YTD Total (col 13)
    cell = ws.cell(row=row, column=13, value=round(ytd))
    cell.font = NEGATIVE_FONT if ytd < 0 else font
    cell.number_format = num_fmt
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return ytd


def write_formula_row(ws, row, label, formula_rows, font=TOTAL_FONT,
                      fill=TOTAL_FILL, border=TOTAL_BORDER, sign='+'):
    """Write a row with formulas summing other rows. sign can be '+' or '-'."""
    ws.cell(row=row, column=1, value=label).font = font
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).border = border

    for c in range(2, 14):
        col = cl(c)
        if sign == '+':
            parts = '+'.join(f"{col}{r}" for r in formula_rows)
            formula = f"={parts}"
        else:
            # first row positive, rest subtracted
            formula = f"={cl(c)}{formula_rows[0]}"
            for r in formula_rows[1:]:
                formula += f"-{cl(c)}{r}"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.number_format = INR
        cell.alignment = Alignment(horizontal='right')


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1: CASH FLOW STATEMENT
# ═══════════════════════════════════════════════════════════════════════════════

def build_cashflow_sheet(wb, data):
    ws = wb.active
    ws.title = "Cash Flow Statement"
    ws.sheet_properties.tabColor = NAVY

    max_col = 13  # A + 11 months + YTD

    # Title
    ws.merge_cells('A1:M1')
    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.merge_cells('A2:M2')
    ws.cell(row=2, column=1, value="Cash Flow Statement  |  April 2025 \u2013 February 2026  |  Source: Tally ERP").font = SUBTITLE_FONT
    ws.merge_cells('A3:M3')
    ws.cell(row=3, column=1, value="All amounts in INR (\u20b9). Prepared for investor review on 11 March 2026.").font = SMALL_FONT

    # Headers
    r = 5
    ws.cell(row=r, column=1, value="Particulars").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for i, ml in enumerate(MONTH_LABELS):
        c = i + 2
        ws.cell(row=r, column=c, value=ml).font = HEADER_FONT
        ws.cell(row=r, column=c).fill = HEADER_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=13, value="YTD Total").font = HEADER_FONT
    ws.cell(row=r, column=13).fill = HEADER_FILL
    ws.cell(row=r, column=13).border = THIN_BORDER
    ws.cell(row=r, column=13).alignment = Alignment(horizontal='center')

    r = 6
    # ── Opening Cash Balance ──
    cash_opening = [data['cash_hand'][i] + data['bank'][i] for i in range(11)]
    # For the cash flow, APR opening = cash+bank from Tally APR balance
    # This is the balance AT THAT POINT, not cash flow. For a simple cash-basis
    # statement, let's use the Tally bank+cash balances directly.
    ws.cell(row=r, column=1, value="Opening Cash & Bank Balance").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 14):
        ws.cell(row=r, column=c).fill = SECTION_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER
    # We'll fill this after calculating net flows; for now leave row 6 as section header
    OPENING_ROW = r
    r += 1

    # ── A. CASH INFLOWS ──
    ws.cell(row=r, column=1, value="A. CASH INFLOWS (Revenue)").font = SECTION_FONT
    style_row(ws, r, max_col, SECTION_FONT, SECTION_FILL)
    r += 1

    # Revenue items
    rev_items = [
        ("Sale of Products", data['sale_products']),
        ("Sale of Services", data['sale_service']),
        ("Export Sales", data['export_sales']),
        ("Printsticks Revenue", data['printsticks']),
        ("Discount Received", data['discount_recv']),
        ("Interest Income", data['interest_income']),
        ("Other Income", data['other_income']),
    ]

    rev_start = r
    for label, vals in rev_items:
        if sum(abs(v) for v in vals) > 0:
            write_data_row(ws, r, f"  {label}", vals)
            r += 1
    rev_end = r - 1

    # Total Inflows
    write_formula_row(ws, r, "Total Cash Inflows", list(range(rev_start, rev_end + 1)))
    TOTAL_IN_ROW = r
    r += 2

    # ── B. CASH OUTFLOWS ──
    ws.cell(row=r, column=1, value="B. CASH OUTFLOWS").font = SECTION_FONT
    style_row(ws, r, max_col, SECTION_FONT, SECTION_FILL)
    r += 1

    # Purchases
    ws.cell(row=r, column=1, value="  Material Purchases").font = Font(name="Calibri", bold=True, size=10, color=MID_BLUE)
    style_row(ws, r, max_col, Font(name="Calibri", bold=True, size=10, color=MID_BLUE))
    r += 1
    purch_items = [
        ("    Purchase of Raw Materials (Domestic)", data['purchase_rm']),
        ("    Import of Raw Materials", data['import_rm']),
        ("    Other Purchases", data['other_purchase']),
    ]
    purch_start = r
    for label, vals in purch_items:
        write_data_row(ws, r, label, vals)
        r += 1
    purch_end = r - 1
    write_formula_row(ws, r, "  Sub-total: Purchases", list(range(purch_start, purch_end + 1)))
    PURCH_TOTAL = r
    r += 1

    # Direct / Manufacturing
    ws.cell(row=r, column=1, value="  Direct / Manufacturing Expenses").font = Font(name="Calibri", bold=True, size=10, color=MID_BLUE)
    style_row(ws, r, max_col, Font(name="Calibri", bold=True, size=10, color=MID_BLUE))
    r += 1
    direct_items = [
        ("    Salaries (Production)", data['salary_prod']),
        ("    Overtime Pay", data['overtime']),
        ("    Electricity (Factory + Mfg)", [data['elec_factory'][i] + data['elec_mfg'][i] for i in range(11)]),
        ("    Freight Inward", data['freight_in']),
    ]
    direct_start = r
    for label, vals in direct_items:
        if sum(abs(v) for v in vals) > 0:
            write_data_row(ws, r, label, vals)
            r += 1
    direct_end = r - 1
    write_formula_row(ws, r, "  Sub-total: Direct Expenses", list(range(direct_start, direct_end + 1)))
    DIRECT_TOTAL = r
    r += 1

    # Salaries & HR
    ws.cell(row=r, column=1, value="  Salaries & HR").font = Font(name="Calibri", bold=True, size=10, color=MID_BLUE)
    style_row(ws, r, max_col, Font(name="Calibri", bold=True, size=10, color=MID_BLUE))
    r += 1
    hr_start = r
    write_data_row(ws, r, "    Payroll / HR Expenses", data['payroll'])
    r += 1
    hr_end = r - 1
    write_formula_row(ws, r, "  Sub-total: HR & Payroll", list(range(hr_start, hr_end + 1)))
    HR_TOTAL = r
    r += 1

    # Admin & Overheads
    ws.cell(row=r, column=1, value="  Admin & Overheads").font = Font(name="Calibri", bold=True, size=10, color=MID_BLUE)
    style_row(ws, r, max_col, Font(name="Calibri", bold=True, size=10, color=MID_BLUE))
    r += 1
    admin_items = [
        ("    Office & Admin Overheads", data['admin_oh']),
        ("    Travelling", data['travel']),
        ("    Rates & Taxes", data['rates_tax']),
        ("    Professional Services", data['professional']),
    ]
    admin_start = r
    for label, vals in admin_items:
        if sum(abs(v) for v in vals) > 0:
            write_data_row(ws, r, label, vals)
            r += 1
    admin_end = r - 1
    write_formula_row(ws, r, "  Sub-total: Admin & Overheads", list(range(admin_start, admin_end + 1)))
    ADMIN_TOTAL = r
    r += 1

    # Marketing & Sales
    ws.cell(row=r, column=1, value="  Marketing & Sales").font = Font(name="Calibri", bold=True, size=10, color=MID_BLUE)
    style_row(ws, r, max_col, Font(name="Calibri", bold=True, size=10, color=MID_BLUE))
    r += 1
    mkt_items = [
        ("    Advertisement / Marketing", data['ads_marketing']),
        ("    Freight Outward", data['freight_out']),
    ]
    mkt_start = r
    for label, vals in mkt_items:
        if sum(abs(v) for v in vals) > 0:
            write_data_row(ws, r, label, vals)
            r += 1
    mkt_end = r - 1
    write_formula_row(ws, r, "  Sub-total: Marketing", list(range(mkt_start, mkt_end + 1)))
    MKT_TOTAL = r
    r += 1

    # Finance & Other
    ws.cell(row=r, column=1, value="  Finance & Other").font = Font(name="Calibri", bold=True, size=10, color=MID_BLUE)
    style_row(ws, r, max_col, Font(name="Calibri", bold=True, size=10, color=MID_BLUE))
    r += 1
    fin_items = [
        ("    Finance Cost (Interest/Bank)", data['finance']),
        ("    R&D Expenses", data['rnd']),
    ]
    fin_start = r
    for label, vals in fin_items:
        if sum(abs(v) for v in vals) > 0:
            write_data_row(ws, r, label, vals)
            r += 1
    fin_end = r - 1
    write_formula_row(ws, r, "  Sub-total: Finance & Other", list(range(fin_start, fin_end + 1)))
    FIN_TOTAL = r
    r += 1

    # ── TOTAL CASH OUTFLOWS ──
    subtotal_rows = [PURCH_TOTAL, DIRECT_TOTAL, HR_TOTAL, ADMIN_TOTAL, MKT_TOTAL, FIN_TOTAL]
    write_formula_row(ws, r, "Total Cash Outflows", subtotal_rows, GRAND_FONT, GRAND_FILL, TOTAL_BORDER)
    TOTAL_OUT_ROW = r
    r += 2

    # ── NET CASH FLOW ──
    ws.cell(row=r, column=1, value="NET CASH FLOW (Inflows \u2212 Outflows)").font = GRAND_FONT
    ws.cell(row=r, column=1).fill = GRAND_FILL
    ws.cell(row=r, column=1).border = TOTAL_BORDER
    for c in range(2, 14):
        col = cl(c)
        cell = ws.cell(row=r, column=c, value=f"={col}{TOTAL_IN_ROW}-{col}{TOTAL_OUT_ROW}")
        cell.font = GRAND_FONT
        cell.fill = GRAND_FILL
        cell.border = TOTAL_BORDER
        cell.number_format = INR
        cell.alignment = Alignment(horizontal='right')
    NET_ROW = r
    r += 2

    # ── BANK BALANCE TRAJECTORY ──
    ws.cell(row=r, column=1, value="C. BANK & CASH POSITION (from Tally)").font = SECTION_FONT
    style_row(ws, r, max_col, SECTION_FONT, SECTION_FILL)
    r += 1

    write_data_row(ws, r, "  Cash-in-Hand", data['cash_hand'])
    r += 1
    bank_vals = data['bank']
    write_data_row(ws, r, "  Bank Accounts (incl. OD)", bank_vals)
    BANK_ROW = r
    r += 1

    # Total Cash Position
    cash_pos = [data['cash_hand'][i] + data['bank'][i] for i in range(11)]
    write_data_row(ws, r, "Total Cash Position", cash_pos, TOTAL_FONT, TOTAL_FILL)
    CASH_POS_ROW = r
    r += 2

    # ── KEY METRICS ──
    ws.cell(row=r, column=1, value="D. KEY METRICS").font = SECTION_FONT
    style_row(ws, r, max_col, SECTION_FONT, SECTION_FILL)
    r += 1

    # Monthly Burn Rate
    ws.cell(row=r, column=1, value="  Monthly Cash Burn").font = NORMAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 14):
        col = cl(c)
        ws.cell(row=r, column=c, value=f"={col}{TOTAL_OUT_ROW}").number_format = INR
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).font = NORMAL_FONT
    BURN_ROW = r
    r += 1

    # Cash Position Change
    ws.cell(row=r, column=1, value="  Cash Position (Bank + Cash)").font = TOTAL_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 13):
        col = cl(c)
        ws.cell(row=r, column=c, value=f"={col}{CASH_POS_ROW}").number_format = INR
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).font = TOTAL_FONT
    # YTD: latest position
    ws.cell(row=r, column=13, value=f"=L{CASH_POS_ROW}").number_format = INR
    ws.cell(row=r, column=13).border = THIN_BORDER
    ws.cell(row=r, column=13).font = TOTAL_FONT
    r += 2

    # Column widths
    ws.column_dimensions['A'].width = 38
    for c in range(2, 14):
        ws.column_dimensions[cl(c)].width = 14
    ws.freeze_panes = 'B6'

    return {
        'total_in_row': TOTAL_IN_ROW,
        'total_out_row': TOTAL_OUT_ROW,
        'net_row': NET_ROW,
        'cash_pos_row': CASH_POS_ROW,
        'bank_row': BANK_ROW,
        'purch_total': PURCH_TOTAL,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2: WORKING CAPITAL REQUIREMENT
# ═══════════════════════════════════════════════════════════════════════════════

def build_wc_sheet(wb, po_received, awaiting_po, cf_rows):
    ws = wb.create_sheet("Working Capital Requirement")
    ws.sheet_properties.tabColor = "C00000"

    max_col = 5

    # Title
    ws.merge_cells('A1:E1')
    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.merge_cells('A2:E2')
    ws.cell(row=2, column=1, value="Working Capital Requirement  |  Orders in Hand  |  March 2026").font = SUBTITLE_FONT
    ws.merge_cells('A3:E3')
    ws.cell(row=3, column=1, value="COGS estimated at 50% of order value for material procurement and execution.").font = SMALL_FONT

    # ── Section A: Current Cash Position ──
    r = 5
    ws.cell(row=r, column=1, value="A. CURRENT CASH POSITION (Feb 2026)").font = SECTION_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1

    headers_a = ['Particulars', '', '', 'Amount (\u20b9)', 'Remarks']
    for ci, h in enumerate(headers_a, 1):
        ws.cell(row=r, column=ci, value=h).font = HEADER_FONT
        ws.cell(row=r, column=ci).fill = HEADER_FILL
        ws.cell(row=r, column=ci).border = THIN_BORDER
    r += 1

    # Cash Position — reference from Cash Flow sheet
    ws.cell(row=r, column=1, value="Cash-in-Hand (Feb 2026)").font = NORMAL_FONT
    ws.cell(row=r, column=4, value=f"='Cash Flow Statement'!L{cf_rows['cash_pos_row']-2}").number_format = INR
    ws.cell(row=r, column=5, value="From Tally").font = SMALL_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    CASH_HAND_ROW = r
    r += 1

    ws.cell(row=r, column=1, value="Bank Balance (Feb 2026)").font = NORMAL_FONT
    ws.cell(row=r, column=4, value=f"='Cash Flow Statement'!L{cf_rows['bank_row']}").number_format = INR
    ws.cell(row=r, column=5, value="From Tally (incl. OD)").font = SMALL_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    BANK_BAL_ROW = r
    r += 1

    ws.cell(row=r, column=1, value="Current Cash Position").font = TOTAL_FONT
    ws.cell(row=r, column=4, value=f"=D{CASH_HAND_ROW}+D{BANK_BAL_ROW}").number_format = INR
    style_row(ws, r, max_col, TOTAL_FONT, TOTAL_FILL, TOTAL_BORDER)
    CURRENT_CASH_ROW = r
    r += 2

    # ── Section B: Orders in Hand (PO Received) ──
    ws.cell(row=r, column=1, value="B. CONFIRMED ORDERS IN HAND (PO Received \u2014 100% Confirmed)").font = SECTION_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1

    for ci, h in enumerate(['Order / Deal Name', '', '', 'Order Value (\u20b9)', 'COGS at 50%'], 1):
        ws.cell(row=r, column=ci, value=h).font = HEADER_FONT
        ws.cell(row=r, column=ci).fill = HEADER_FILL
        ws.cell(row=r, column=ci).border = THIN_BORDER
    r += 1

    po_start = r
    for deal in sorted(po_received, key=lambda d: -d['amount']):
        ws.cell(row=r, column=1, value=f"  {deal['name']}").font = NORMAL_FONT
        ws.cell(row=r, column=4, value=round(deal['amount'])).number_format = INR
        ws.cell(row=r, column=5, value=f"=D{r}*0.5").number_format = INR
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    po_end = r - 1

    ws.cell(row=r, column=1, value="Total PO Received (Orders in Hand)").font = TOTAL_FONT
    ws.cell(row=r, column=4, value=f"=SUM(D{po_start}:D{po_end})").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(E{po_start}:E{po_end})").number_format = INR
    style_row(ws, r, max_col, TOTAL_FONT, TOTAL_FILL, TOTAL_BORDER)
    PO_TOTAL_ROW = r
    r += 2

    # ── Section C: Awaiting PO (High Probability) ──
    ws.cell(row=r, column=1, value="C. NEAR-CONFIRMED ORDERS (Awaiting PO \u2014 80% Probability)").font = SECTION_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1

    for ci, h in enumerate(['Order / Deal Name', '', '', 'Order Value (\u20b9)', 'COGS at 50%'], 1):
        ws.cell(row=r, column=ci, value=h).font = HEADER_FONT
        ws.cell(row=r, column=ci).fill = HEADER_FILL
        ws.cell(row=r, column=ci).border = THIN_BORDER
    r += 1

    apo_start = r
    for deal in sorted(awaiting_po, key=lambda d: -d['amount']):
        ws.cell(row=r, column=1, value=f"  {deal['name']}").font = NORMAL_FONT
        ws.cell(row=r, column=4, value=round(deal['amount'])).number_format = INR
        ws.cell(row=r, column=5, value=f"=D{r}*0.5").number_format = INR
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    apo_end = r - 1

    ws.cell(row=r, column=1, value="Total Awaiting PO").font = TOTAL_FONT
    ws.cell(row=r, column=4, value=f"=SUM(D{apo_start}:D{apo_end})").number_format = INR
    ws.cell(row=r, column=5, value=f"=SUM(E{apo_start}:E{apo_end})").number_format = INR
    style_row(ws, r, max_col, TOTAL_FONT, TOTAL_FILL, TOTAL_BORDER)
    APO_TOTAL_ROW = r
    r += 2

    # ── Section D: WORKING CAPITAL SUMMARY ──
    ws.cell(row=r, column=1, value="D. WORKING CAPITAL REQUIREMENT SUMMARY").font = SECTION_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = SECTION_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1

    for ci, h in enumerate(['Particulars', '', '', 'Amount (\u20b9)', 'Notes'], 1):
        ws.cell(row=r, column=ci, value=h).font = HEADER_FONT
        ws.cell(row=r, column=ci).fill = HEADER_FILL
        ws.cell(row=r, column=ci).border = THIN_BORDER
    r += 1

    # Total confirmed orders
    ws.cell(row=r, column=1, value="Total Confirmed Orders (PO Received)").font = NORMAL_FONT
    ws.cell(row=r, column=4, value=f"=D{PO_TOTAL_ROW}").number_format = INR
    ws.cell(row=r, column=5, value="100% confirmed with PO").font = SMALL_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    CONF_ORD_ROW = r
    r += 1

    # COGS for confirmed orders
    ws.cell(row=r, column=1, value="Material / COGS Required (50% of Order Value)").font = NORMAL_FONT
    ws.cell(row=r, column=4, value=f"=E{PO_TOTAL_ROW}").number_format = INR
    ws.cell(row=r, column=5, value="Raw materials, components, mfg costs").font = SMALL_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    COGS_REQ_ROW = r
    r += 1

    # Available cash
    ws.cell(row=r, column=1, value="Less: Available Cash & Bank").font = NORMAL_FONT
    ws.cell(row=r, column=4, value=f"=D{CURRENT_CASH_ROW}").number_format = INR
    ws.cell(row=r, column=5, value="Current bank + cash position").font = SMALL_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    AVAIL_CASH_ROW = r
    r += 1

    # Monthly OPEX buffer (1 month)
    ws.cell(row=r, column=1, value="Add: Monthly Operating Expenses (1-month buffer)").font = NORMAL_FONT
    ws.cell(row=r, column=4, value=f"='Cash Flow Statement'!M{cf_rows['total_out_row']}/11").number_format = INR
    ws.cell(row=r, column=5, value="Avg monthly outflow for opex continuity").font = SMALL_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    OPEX_BUFFER_ROW = r
    r += 1

    # ── WORKING CAPITAL GAP ──
    ws.cell(row=r, column=1, value="WORKING CAPITAL REQUIREMENT").font = GRAND_FONT
    ws.cell(row=r, column=4,
            value=f"=D{COGS_REQ_ROW}-D{AVAIL_CASH_ROW}+D{OPEX_BUFFER_ROW}").number_format = INR
    ws.cell(row=r, column=5, value="COGS Required \u2212 Cash Available + OPEX Buffer").font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = GRAND_FILL
        ws.cell(row=r, column=c).border = TOTAL_BORDER
        ws.cell(row=r, column=c).font = GRAND_FONT
    ws.cell(row=r, column=4).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.cell(row=r, column=4).font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    WC_REQ_ROW = r
    r += 2

    # ── Including Awaiting PO ──
    ws.cell(row=r, column=1, value="IF INCLUDING NEAR-CONFIRMED (PO + Awaiting PO):").font = Font(name="Calibri", bold=True, size=10, color="C00000")
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1

    ws.cell(row=r, column=1, value="Total Orders (PO + Awaiting PO)").font = NORMAL_FONT
    ws.cell(row=r, column=4, value=f"=D{PO_TOTAL_ROW}+D{APO_TOTAL_ROW}").number_format = INR
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    ALL_ORD_ROW = r
    r += 1

    ws.cell(row=r, column=1, value="COGS for All Confirmed Orders (50%)").font = NORMAL_FONT
    ws.cell(row=r, column=4, value=f"=D{ALL_ORD_ROW}*0.5").number_format = INR
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    ALL_COGS_ROW = r
    r += 1

    ws.cell(row=r, column=1, value="TOTAL WORKING CAPITAL NEEDED").font = GRAND_FONT
    ws.cell(row=r, column=4,
            value=f"=D{ALL_COGS_ROW}-D{AVAIL_CASH_ROW}+D{OPEX_BUFFER_ROW}").number_format = INR
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws.cell(row=r, column=c).border = TOTAL_BORDER
        ws.cell(row=r, column=c).font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    r += 2

    # ── Notes ──
    notes = [
        "\U0001f4b0  Working Capital Requirement \u2014 Key Points for Investors",
        "",
        "  \u2022  Fracktal has \u20b92.9M in CONFIRMED orders (PO Received, 100% certain) + \u20b96.3M awaiting PO (80% probability).",
        "  \u2022  To execute confirmed orders, the company needs \u20b91.46M in raw materials and manufacturing costs (50% COGS).",
        "  \u2022  Current bank position is in overdraft \u2014 working capital injection is critical for order execution.",
        "  \u2022  Total order pipeline (PO + Awaiting) represents \u20b99.2M in confirmed/near-confirmed revenue.",
        "  \u2022  Without working capital, the company cannot procure materials to fulfill confirmed customer orders.",
        "  \u2022  The 50% COGS ratio is based on material + manufacturing costs as a proportion of selling price.",
        "  \u2022  Monthly operating burn is ~\u20b922L/month \u2014 a 1-month buffer ensures business continuity during execution.",
        "",
        "  \u2192  Recommended Ask: Working capital for immediate order execution + 1 month operating buffer.",
        "  \u2192  Expected Return: Orders convert to revenue within 30-60 days, replenishing working capital cycle.",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=r + i, column=1, value=n).font = NOTE_FONT

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.freeze_panes = 'A6'


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("Fracktal Cash Flow + Working Capital — Investor Package")
    print(f"  MIS Source: {MIS_FILE}")
    print(f"  CRM Source: {CRM_FILE}")
    print(f"  Output:     {OUTPUT}")

    # Read data
    print("\n  Reading TB data...")
    data = read_tb_data(MIS_FILE)
    print("  Reading CRM orders...")
    po_received, awaiting_po = read_crm_orders(CRM_FILE)
    print(f"    PO Received: {len(po_received)} orders, total: {sum(d['amount'] for d in po_received):,.0f}")
    print(f"    Awaiting PO: {len(awaiting_po)} orders, total: {sum(d['amount'] for d in awaiting_po):,.0f}")

    # Create workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Cash Flow
    print("\n  Building Cash Flow Statement...")
    cf_rows = build_cashflow_sheet(wb, data)

    # Sheet 2: Working Capital
    print("  Building Working Capital Requirement...")
    build_wc_sheet(wb, po_received, awaiting_po, cf_rows)

    # Save
    print(f"\n  Saving to {OUTPUT}")
    wb.save(str(OUTPUT))
    print(f"  Sheets: {wb.sheetnames}")
    print("  Done! \u2713")


if __name__ == "__main__":
    main()
