#!/usr/bin/env python3
"""
Fracktal Works — MIS Master Sheet Generator (Formula-Linked to Trial Balance)

Creates a professional 8-sheet Excel workbook for Fracktal Works Private Limited:
  1. TB            — Trial Balance input (paste monthly Tally data here)
  2. P&L           — Profit & Loss Statement (formula-linked to TB)
  3. Cash Flow     — Monthly cash inflows / outflows / closing balance
  4. OPEX Schedule — Detailed breakup of indirect expenses
  5. Balance Sheet — Monthly balance sheet from TB
  6. Performance Summary — Quarterly aggregation with FY totals
  7. KPIs          — Key financial ratios with benchmarks
  8. Dashboard     — Charts and visual summary

All numbers flow from Excel formulas. Zero hardcoded values.
Every sheet includes plain-English explanations at the bottom.

Usage:
    python execution/fracktal_mis_generator.py --source "path/to/TrialBalance.xls"
    python execution/fracktal_mis_generator.py --source "path/to/TrialBalance.xls" --output "MIS.xlsx"
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
NAVY       = "1F3864"
DARK_BLUE  = "2F5496"
MID_BLUE   = "4472C4"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN= "E2EFDA"
LIGHT_YELLOW="FFF2CC"
WHITE      = "FFFFFF"
RED        = "C00000"
DARK_GREEN = "548235"

TITLE_FONT       = Font(name="Calibri", bold=True, size=14, color=NAVY)
SUBTITLE_FONT    = Font(name="Calibri", bold=True, size=12, color=MID_BLUE)
HEADER_FONT      = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL      = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SECTION_FONT     = Font(name="Calibri", bold=True, size=11, color=NAVY)
SECTION_FILL     = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
SUBSEC_FONT      = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
TOTAL_FONT       = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL       = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
GRAND_TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
GRAND_TOTAL_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
INPUT_FILL       = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
GROUP_FILL       = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
NORMAL_FONT      = Font(name="Calibri", size=11)
SMALL_FONT       = Font(name="Calibri", size=10, color="808080", italic=True)
PCT_FONT         = Font(name="Calibri", italic=True, size=10, color=MID_BLUE)
NOTE_FONT        = Font(name="Calibri", size=10, color="808080")

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
TOTAL_BORDER = Border(
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='double', color=NAVY),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)

INR  = '#,##0'
INR2 = '#,##0.00'
PCT  = '0.0%'
NUM2 = '0.00'

MONTHS = ['APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP',
          'OCT', 'NOV', 'DEC', 'JAN', 'FEB', 'MAR']
MONTH_FULL = ['April', 'May', 'June', 'July', 'August', 'September',
              'October', 'November', 'December', 'January', 'February', 'March']

# ═══════════════════════════════════════════════════════════════════════════════
#  TB SHEET LAYOUT CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
# TB sheet:
#   Row 1: Title
#   Row 2: Subtitle
#   Row 3: Headers row (Particulars, APR..MAR, FY Total)
#   Row 4+: Data rows (month cols B-M, FY=N)

TB_HEADER_ROW = 3           # Row with column headers
TB_DATA_START = 4            # First data row
TB_MONTH_START_COL = 2       # Col B = APR
TB_FY_COL = 14               # Col N = FY Total
TB_LAST_COL = 14             # Col N

# Report sheets use same column layout: B=APR..M=MAR, N=FY Total
RPT_COL_B = 2
RPT_COL_M = 13
RPT_COL_N = 14
RPT_LAST = 14


# ═══════════════════════════════════════════════════════════════════════════════
#  ACCOUNT DEFINITIONS — Mapping each Tally TB row to our TB input sheet
# ═══════════════════════════════════════════════════════════════════════════════
# Each entry: (tally_row_in_source, display_name, category, nature)
# nature: 'Dr' = debit-side, 'Cr' = credit-side, 'Both' = net
# category: grouping for the TB sheet sections

# The Tally TB file has rows 12-78 as account data. Source row indices below
# map to the pandas DataFrame row index (0-based source, but Tally rows are
# 12..78 in the original file).

# We'll build the TB input sheet in a logical P&L-friendly order
# rather than mirroring Tally's BS-first layout.

# --- P&L accounts (what goes into TB Input) ---

TB_ACCOUNTS = [
    # --- REVENUE ---
    # (section_header, None, None, None) for headers
    ("REVENUE", None, None, None),
    ("Sale of Products",           45, "Revenue",  "Cr"),
    ("Sale of Service",            46, "Revenue",  "Cr"),
    ("Export Sales",               43, "Revenue",  "Cr"),
    ("Printsticks",                44, "Revenue",  "Cr"),
    ("Discount Received",          60, "Revenue",  "Cr"),
    ("Interest Income",            61, "Revenue",  "Cr"),
    ("Other Income",               62, "Revenue",  "Cr"),

    # --- COGS ---
    ("COST OF GOODS SOLD (COGS)", None, None, None),
    ("Opening Stock", -1, "COGS_OPENING", "Dr"),
    ("PURCHASE of Raw Materials",  49, "COGS_PURCHASE", "Dr"),
    ("IMPORT of Raw Materials",    48, "COGS_PURCHASE", "Dr"),
    ("OTHER PURCHASES",            51, "COGS_PURCHASE", "Dr"),
    ("Closing Stock", -1, "COGS_CLOSING", "Dr"),

    # --- DIRECT EXPENSES ---
    ("DIRECT / MANUFACTURING EXPENSES", None, None, None),
    ("Salaries (Production)",      58, "Direct", "Dr"),
    ("Overtime Pay",               57, "Direct", "Dr"),
    ("Electricity - Factory",      53, "Direct", "Dr"),
    ("Electricity - Manufacturing",54, "Direct", "Dr"),
    ("Freight Inward",             55, "Direct", "Dr"),
    ("Loading & Unloading",        56, "Direct", "Dr"),
    ("Discount Allowed",           52, "Direct", "Dr"),

    # --- INDIRECT EXPENSES / OPEX ---
    ("INDIRECT EXPENSES (OPEX)", None, None, None),

    ("Office & Admin Overheads", None, None, None),
    ("Office & Administrative Overheads", 66, "Admin", "Dr"),
    ("Travelling Expense",        70, "Admin", "Dr"),
    ("Rates & Taxes",             68, "Admin", "Both"),
    ("Foreign Exchange Gain/Loss",71, "Admin", "Dr"),
    ("Razorpay Charges",          74, "Admin", "Dr"),
    ("Tender Fee",                76, "Admin", "Dr"),

    ("Finance Cost", None, None, None),
    ("Finance Cost (Interest/Bank)", 65, "Finance", "Dr"),

    ("HR / Payroll Expenses", None, None, None),
    ("Payroll Expenses",          67, "Payroll", "Dr"),

    ("Marketing & Ads", None, None, None),
    ("Advertisement/Marketing",   64, "Marketing", "Dr"),
    ("Freight Outward",           72, "Marketing", "Dr"),

    ("R&D Expenses", None, None, None),
    ("R&D",                       69, "R&D", "Dr"),

    ("Professional Service Charges", 73, "Professional", "Dr"),
    ("Round Off",                 75, "Roundoff", "Both"),
    ("Write Off / Write Back",   77, "WriteOff", "Both"),

    # --- BALANCE SHEET ITEMS ---
    ("BALANCE SHEET ITEMS", None, None, None),

    ("Capital Account", None, None, None),
    ("Reserves & Surplus",        13, "Capital", "Cr"),
    ("Issued, Subscribed & Paid Up", 14, "Capital", "Cr"),

    ("Loans (Liability)", None, None, None),
    ("Unsecured Loans",           16, "Loans", "Cr"),

    ("Current Liabilities", None, None, None),
    ("Duties & Taxes",            18, "CL", "Both"),
    ("Sundry Creditors",          19, "CL", "Both"),
    ("Reimbursements Payable",    20, "CL", "Both"),
    ("Salaries Payable",          21, "CL", "Cr"),

    ("Fixed Assets", None, None, None),
    ("Intangible Assets",         23, "FA", "Dr"),
    ("Machinery & Equipment",     24, "FA", "Dr"),
    ("Tangible Assets",           25, "FA", "Dr"),
    ("Fracktal Studio",           26, "FA", "Dr"),
    ("SLS PAMM Project",          27, "FA", "Dr"),
    ("WIP",                       28, "FA", "Dr"),

    ("Current Assets", None, None, None),
    ("Inventories",               30, "CA", "Dr"),
    ("Deposits (Asset)",          31, "CA", "Dr"),
    ("Loans & Advances (Asset)",  32, "CA", "Dr"),
    ("Sundry Debtors",            33, "CA", "Both"),
    ("Cash-in-Hand",              34, "CA", "Dr"),
    ("Bank Accounts",             35, "CA", "Both"),
    ("Deferred Expenses",         36, "CA", "Dr"),
    ("Deferred Tax Asset",        37, "CA", "Cr"),
    ("TDS - GST Receivable",      38, "CA", "Dr"),
    ("Tds Receivable",            39, "CA", "Dr"),

    ("Suspense", None, None, None),
    ("Suspense Account",          41, "Suspense", "Dr"),

    ("Profit & Loss A/c",        78, "PnL", "Both"),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_float(val, default=0.0):
    if val is None or (isinstance(val, float) and val != val):
        return default
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace(',', '').replace('\u20b9', '').replace('\xa0', '')
        if s in ('', '-'):
            return default
        try:
            return float(s)
        except ValueError:
            return default
    return default


def col_letter(idx):
    return get_column_letter(idx)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_section(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER


def style_total(ws, row, max_col, font=None, fill=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or TOTAL_FONT
        cell.fill = fill or TOTAL_FILL
        cell.border = TOTAL_BORDER


def style_grand_total(ws, row, max_col):
    style_total(ws, row, max_col, font=GRAND_TOTAL_FONT, fill=GRAND_TOTAL_FILL)


def fmt_number(ws, row, scol, ecol, fmt=INR):
    for c in range(scol, ecol + 1):
        cell = ws.cell(row=row, column=c)
        cell.number_format = fmt
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')


def fmt_pct(ws, row, scol, ecol):
    for c in range(scol, ecol + 1):
        cell = ws.cell(row=row, column=c)
        cell.number_format = PCT
        cell.font = PCT_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')


def write_notes(ws, start_row, notes):
    """Write plain-English explanation notes at the bottom of a sheet."""
    for i, note in enumerate(notes):
        cell = ws.cell(row=start_row + i, column=1, value=note)
        cell.font = NOTE_FONT


# ═══════════════════════════════════════════════════════════════════════════════
#  READ TALLY TB
# ═══════════════════════════════════════════════════════════════════════════════

def read_tally_tb(source_path):
    """
    Read Tally-exported TB. Returns dict: tally_row -> {month_idx: (dr, cr)}.
    month_idx: 0=Apr..8=Dec (only months present in source).
    Also returns num_months found.
    """
    ext = Path(source_path).suffix.lower()
    try:
        if ext == '.xls':
            try:
                df = pd.read_excel(source_path, header=None, engine='openpyxl')
            except Exception:
                df = pd.read_excel(source_path, header=None, engine='xlrd')
        else:
            df = pd.read_excel(source_path, header=None, engine='openpyxl')
    except Exception as e:
        print(f"ERROR reading file: {e}")
        sys.exit(1)

    num_total_cols = df.shape[1]
    # Cols: 0=Particulars, 1-2=Cumulative Dr/Cr, then pairs per month
    num_months = (num_total_cols - 3) // 2
    print(f"  Found {num_months} monthly periods in TB file")

    tb_data = {}
    for tally_row in range(12, min(79, df.shape[0])):
        row_data = {}
        for m in range(num_months):
            dr_col = 3 + m * 2
            cr_col = 4 + m * 2
            dr = safe_float(df.iloc[tally_row, dr_col]) if dr_col < num_total_cols else 0.0
            cr = safe_float(df.iloc[tally_row, cr_col]) if cr_col < num_total_cols else 0.0
            row_data[m] = (dr, cr)
        tb_data[tally_row] = row_data

    return tb_data, num_months


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1: TB (Trial Balance Input)
# ═══════════════════════════════════════════════════════════════════════════════

def build_tb_sheet(wb, tb_data, num_months):
    """
    Build TB sheet layout:
      Row 1: Title
      Row 2: Instructions
      Row 3: Headers (Particulars, APR..MAR, FY Total)
      Row 4+: Data (sections, input rows, totals with SUM formulas)
    
    Returns: dict mapping account_name -> TB row number (for formula refs)
    """
    ws = wb.create_sheet("TB")
    ws.sheet_properties.tabColor = "FFC000"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED \u2014 Trial Balance Input").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Paste monthly Trial Balance numbers below. FY Total auto-calculates.").font = SMALL_FONT

    # Header row 3
    headers = ['Particulars'] + MONTHS + ['FY Total']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=TB_HEADER_ROW, column=ci, value=h)
    style_header(ws, TB_HEADER_ROW, TB_LAST_COL)

    # Build data rows
    row = TB_DATA_START
    acct_rows = {}       # name -> row number
    section_stacks = []  # track section start rows for sub-totals

    # For grouping: we need to track where sections start for sub-totals
    # Revenue items
    rev_items = []
    rev_income_items = []  # other income subset
    cogs_opening_row = None
    cogs_purchase_items = []
    cogs_closing_row = None
    direct_items = []
    admin_items = []
    finance_items = []
    payroll_items = []
    marketing_items = []
    rnd_items = []
    other_opex_items = []

    # Current category tracker
    current_section = None
    current_subsection = None

    for name, tally_row, category, nature in TB_ACCOUNTS:
        if tally_row is None:
            # Section header
            ws.cell(row=row, column=1, value=name).font = SECTION_FONT
            style_section(ws, row, TB_LAST_COL)
            current_section = name
            if name in ("Office & Admin Overheads", "Finance Cost", "HR / Payroll Expenses",
                        "Marketing & Ads", "R&D Expenses",
                        "Capital Account", "Loans (Liability)", "Current Liabilities",
                        "Fixed Assets", "Current Assets", "Suspense"):
                current_subsection = name
            row += 1
            continue

        # Manual input row (e.g., Opening Stock, Closing Stock for COGS)
        if tally_row == -1:
            ws.cell(row=row, column=1, value=name).font = NORMAL_FONT
            acct_rows[name] = row
            for m in range(12):
                col = TB_MONTH_START_COL + m
                cell = ws.cell(row=row, column=col)
                cell.fill = INPUT_FILL
                cell.number_format = INR2
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='right')
            # FY Total = SUM(B:M)
            fy_cell = ws.cell(row=row, column=TB_FY_COL)
            fy_cell.value = f"=SUM(B{row}:M{row})"
            fy_cell.number_format = INR2
            fy_cell.border = THIN_BORDER
            if category == "COGS_OPENING":
                cogs_opening_row = row
            elif category == "COGS_CLOSING":
                cogs_closing_row = row
            row += 1
            continue

        # Data row
        ws.cell(row=row, column=1, value=name).font = NORMAL_FONT
        acct_rows[name] = row

        # Populate month columns with data from source + FY formula
        for m in range(12):
            col = TB_MONTH_START_COL + m
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.number_format = INR2
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

            # Fill data if available
            if tally_row in tb_data and m in tb_data[tally_row]:
                dr, cr = tb_data[tally_row][m]
                # Store NET value based on nature
                if nature == 'Dr':
                    net = dr - cr if (dr or cr) else 0
                elif nature == 'Cr':
                    net = cr - dr if (dr or cr) else 0
                else:  # Both — store Dr-Cr (can be negative)
                    net = dr - cr if (dr or cr) else 0

                if net != 0:
                    cell.value = round(net, 2)

        # FY Total = SUM(B:M)
        fy_cell = ws.cell(row=row, column=TB_FY_COL)
        fy_cell.value = f"=SUM(B{row}:M{row})"
        fy_cell.number_format = INR2
        fy_cell.border = THIN_BORDER

        # Track which category this belongs to for totals
        if category == "Revenue" and name not in ("Discount Received", "Interest Income", "Other Income"):
            rev_items.append(row)
        elif category == "Revenue":
            rev_income_items.append(row)
        elif category == "COGS_PURCHASE":
            cogs_purchase_items.append(row)
        elif category == "Direct":
            direct_items.append(row)
        elif category == "Admin":
            admin_items.append(row)
        elif category == "Finance":
            finance_items.append(row)
        elif category == "Payroll":
            payroll_items.append(row)
        elif category == "Marketing":
            marketing_items.append(row)
        elif category == "R&D":
            rnd_items.append(row)
        elif category in ("Professional", "Roundoff", "WriteOff"):
            other_opex_items.append(row)

        row += 1

    # ── Calculated total rows ──

    # Total Revenue from Operations
    def write_total_row(label, items, key_name=None):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = TOTAL_FONT
        for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
            cl = col_letter(c)
            if len(items) > 0:
                refs = '+'.join(f"{cl}{r}" for r in items)
                ws.cell(row=row, column=c, value=f"={refs}")
            else:
                ws.cell(row=row, column=c, value=0)
            ws.cell(row=row, column=c).number_format = INR2
        style_total(ws, row, TB_LAST_COL)
        if key_name:
            acct_rows[key_name] = row
        rr = row
        row += 1
        return rr

    # Insert totals after the data
    row += 1  # blank line

    # --- Calculated P&L summary section ---
    ws.cell(row=row, column=1, value="CALCULATED TOTALS (auto-formulas \u2014 do not edit)").font = SECTION_FONT
    style_section(ws, row, TB_LAST_COL)
    row += 1

    total_rev_ops_row = write_total_row("Total Revenue from Operations", rev_items, "Total Revenue Ops")
    total_other_income_row = write_total_row("Total Other Income", rev_income_items, "Total Other Income")

    # Total Revenue
    total_rev_row = row
    ws.cell(row=row, column=1, value="Total Revenue").font = TOTAL_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"={cl}{total_rev_ops_row}+{cl}{total_other_income_row}")
        ws.cell(row=row, column=c).number_format = INR2
    style_total(ws, row, TB_LAST_COL)
    acct_rows["Total Revenue"] = row
    row += 1

    total_cogs_purchases_row = write_total_row("Total Purchases", cogs_purchase_items, "Total Purchases")

    # Total COGS = Opening Stock + Total Purchases - Closing Stock
    total_cogs_row = row
    ws.cell(row=row, column=1, value="Total COGS (Opening + Purchases \u2212 Closing Stock)").font = TOTAL_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"={cl}{cogs_opening_row}+{cl}{total_cogs_purchases_row}-{cl}{cogs_closing_row}")
        ws.cell(row=row, column=c).number_format = INR2
    style_total(ws, row, TB_LAST_COL)
    acct_rows["Total COGS"] = row
    row += 1

    total_direct_row = write_total_row("Total Direct Expenses", direct_items, "Total Direct Expenses")

    # Gross Profit
    gp_row = row
    ws.cell(row=row, column=1, value="Gross Profit").font = TOTAL_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"={cl}{total_rev_row}-{cl}{total_cogs_row}-{cl}{total_direct_row}")
        ws.cell(row=row, column=c).number_format = INR2
    style_total(ws, row, TB_LAST_COL)
    acct_rows["Gross Profit"] = row
    row += 1

    # GP %
    gp_pct_row = row
    ws.cell(row=row, column=1, value="Gross Profit %").font = PCT_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"=IFERROR({cl}{gp_row}/{cl}{total_rev_row},0)")
        ws.cell(row=row, column=c).number_format = PCT
        ws.cell(row=row, column=c).font = PCT_FONT
    acct_rows["Gross Profit %"] = row
    row += 1

    # OPEX sub-totals
    total_admin_row = write_total_row("Admin Overheads Total", admin_items, "Admin Overheads Total")
    total_finance_row = write_total_row("Finance Cost Total", finance_items, "Finance Cost Total")
    total_payroll_row = write_total_row("HR Expenses Total", payroll_items, "HR Expenses Total")
    total_marketing_row = write_total_row("Marketing Total", marketing_items, "Marketing Total")
    total_rnd_row = write_total_row("R&D Total", rnd_items, "R&D Total")
    total_other_opex_row = write_total_row("Other OPEX (Professional, Round Off, Write Off)", other_opex_items, "Other OPEX Total")

    # Total OPEX
    all_opex = [total_admin_row, total_finance_row, total_payroll_row,
                total_marketing_row, total_rnd_row, total_other_opex_row]
    total_opex_row = row
    ws.cell(row=row, column=1, value="Total Indirect Expenses (OPEX)").font = TOTAL_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        refs = '+'.join(f"{cl}{r}" for r in all_opex)
        ws.cell(row=row, column=c, value=f"={refs}")
        ws.cell(row=row, column=c).number_format = INR2
    style_total(ws, row, TB_LAST_COL)
    acct_rows["Total OPEX"] = row
    row += 1

    # EBITDA
    ebitda_row = row
    ws.cell(row=row, column=1, value="EBITDA").font = TOTAL_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"={cl}{gp_row}-{cl}{total_opex_row}")
        ws.cell(row=row, column=c).number_format = INR2
    style_grand_total(ws, row, TB_LAST_COL)
    acct_rows["EBITDA"] = row
    row += 1

    # EBITDA %
    ebitda_pct_row = row
    ws.cell(row=row, column=1, value="EBITDA %").font = PCT_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"=IFERROR({cl}{ebitda_row}/{cl}{total_rev_row},0)")
        ws.cell(row=row, column=c).number_format = PCT
        ws.cell(row=row, column=c).font = PCT_FONT
    acct_rows["EBITDA %"] = row
    row += 1

    # Depreciation (manual input — placeholder)
    dep_row = row
    ws.cell(row=row, column=1, value="Depreciation").font = NORMAL_FONT
    for c in range(TB_MONTH_START_COL, TB_LAST_COL + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = INPUT_FILL
        cell.number_format = INR2
        cell.border = THIN_BORDER
    ws.cell(row=row, column=TB_FY_COL, value=f"=SUM(B{row}:M{row})")
    acct_rows["Depreciation"] = row
    row += 1

    # PBT
    pbt_row = row
    ws.cell(row=row, column=1, value="Profit Before Tax (PBT)").font = TOTAL_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"={cl}{ebitda_row}-{cl}{dep_row}")
        ws.cell(row=row, column=c).number_format = INR2
    style_grand_total(ws, row, TB_LAST_COL)
    acct_rows["PBT"] = row
    row += 1

    # PBT %
    ws.cell(row=row, column=1, value="PBT %").font = PCT_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"=IFERROR({cl}{pbt_row}/{cl}{total_rev_row},0)")
        ws.cell(row=row, column=c).number_format = PCT
        ws.cell(row=row, column=c).font = PCT_FONT
    acct_rows["PBT %"] = row
    row += 1

    # Tax Provision (manual)
    tax_row = row
    ws.cell(row=row, column=1, value="Tax Provision").font = NORMAL_FONT
    for c in range(TB_MONTH_START_COL, TB_LAST_COL + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = INPUT_FILL
        cell.number_format = INR2
        cell.border = THIN_BORDER
    ws.cell(row=row, column=TB_FY_COL, value=f"=SUM(B{row}:M{row})")
    acct_rows["Tax Provision"] = row
    row += 1

    # PAT
    pat_row = row
    ws.cell(row=row, column=1, value="Profit After Tax (PAT)").font = TOTAL_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"={cl}{pbt_row}-{cl}{tax_row}")
        ws.cell(row=row, column=c).number_format = INR2
    style_grand_total(ws, row, TB_LAST_COL)
    acct_rows["PAT"] = row
    row += 1

    # PAT %
    ws.cell(row=row, column=1, value="PAT %").font = PCT_FONT
    for c in range(TB_MONTH_START_COL, TB_FY_COL + 1):
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"=IFERROR({cl}{pat_row}/{cl}{total_rev_row},0)")
        ws.cell(row=row, column=c).number_format = PCT
        ws.cell(row=row, column=c).font = PCT_FONT
    acct_rows["PAT %"] = row
    row += 2

    # ── Explanatory notes ──
    notes = [
        "\U0001f4cb  How to Use This Sheet (for non-finance readers)",
        "  \u2022  Trial Balance is simply a list of all money that came IN (Revenue) and went OUT (Costs) during each month.",
        "  \u2022  Yellow cells = input cells. Paste the monthly Tally TB numbers here. Everything else auto-calculates.",
        "  \u2022  Revenue section: All income streams \u2014 Product Sales, Services, Exports, Printsticks, and Other Income.",
        "  \u2022  COGS (Cost of Goods Sold): Opening Stock + Purchases \u2212 Closing Stock. Enter stock values manually each month.",
        "  \u2022  Direct Expenses: Manufacturing costs \u2014 production salaries, electricity, freight inward, overtime.",
        "  \u2022  Gross Profit = Revenue minus COGS minus Direct Expenses. Shows profitability of core operations.",
        "  \u2022  Indirect Expenses (OPEX): Ongoing costs to run the business \u2014 admin, payroll, marketing, R&D, etc.",
        "  \u2022  EBITDA = Gross Profit minus OPEX. Core operating profit before depreciation and tax.",
        "  \u2022  PAT (Profit After Tax) = The final bottom line after all expenses, depreciation, and taxes.",
        "  \u2022  FY Total (last column) auto-sums all 12 months. Percentage rows show each metric as a share of Revenue.",
        "  \u2022  Balance Sheet items at the bottom are for the BS report \u2014 paste the net balances from Tally.",
    ]
    write_notes(ws, row, notes)

    # Column widths
    ws.column_dimensions['A'].width = 38
    for c in range(TB_MONTH_START_COL, TB_LAST_COL + 1):
        ws.column_dimensions[col_letter(c)].width = 14
    ws.freeze_panes = 'B4'

    return ws, acct_rows


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2: P&L (Profit & Loss Statement)
# ═══════════════════════════════════════════════════════════════════════════════

def build_pnl_sheet(wb, ar):
    """P&L sheet — all cells are =TB!{cell} references."""
    ws = wb.create_sheet("P&L")
    ws.sheet_properties.tabColor = "548235"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Profit & Loss Statement  |  FY 2025-26").font = SUBTITLE_FONT

    headers = ['Particulars'] + MONTHS + ['FY Total']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header(ws, 3, RPT_LAST)

    def ref_tb(tb_row, rpt_row, label, bold=False, is_section=False, is_pct=False):
        ws.cell(row=rpt_row, column=1, value=label)
        if is_section:
            ws.cell(row=rpt_row, column=1).font = SECTION_FONT
            style_section(ws, rpt_row, RPT_LAST)
            return
        if bold:
            ws.cell(row=rpt_row, column=1).font = TOTAL_FONT
        elif is_pct:
            ws.cell(row=rpt_row, column=1).font = PCT_FONT
        else:
            ws.cell(row=rpt_row, column=1).font = NORMAL_FONT

        for c in range(RPT_COL_B, RPT_LAST + 1):
            cl = col_letter(c)
            ws.cell(row=rpt_row, column=c, value=f"=TB!{cl}{tb_row}")
        if is_pct:
            fmt_pct(ws, rpt_row, RPT_COL_B, RPT_LAST)
        else:
            fmt_number(ws, rpt_row, RPT_COL_B, RPT_LAST)
        if bold:
            style_total(ws, rpt_row, RPT_LAST)

    # Map: label, tb_row_key, bold, section, pct
    pnl_lines = [
        ("Revenue from Operations", None, False, True, False),
        ("Sale of Products", "Sale of Products", False, False, False),
        ("Sale of Service", "Sale of Service", False, False, False),
        ("Export Sales", "Export Sales", False, False, False),
        ("Printsticks", "Printsticks", False, False, False),
        ("Total Revenue from Operations", "Total Revenue Ops", True, False, False),
        (None, None, False, False, False),  # blank
        ("Other Income", None, False, True, False),
        ("Discount Received", "Discount Received", False, False, False),
        ("Interest Income", "Interest Income", False, False, False),
        ("Other Income (Misc)", "Other Income", False, False, False),
        ("Total Other Income", "Total Other Income", True, False, False),
        (None, None, False, False, False),
        ("Total Revenue", "Total Revenue", True, False, False),
        (None, None, False, False, False),
        ("Cost of Goods Sold", None, False, True, False),
        ("Opening Stock", "Opening Stock", False, False, False),
        ("Purchase of Raw Materials (Domestic)", "PURCHASE of Raw Materials", False, False, False),
        ("Import of Raw Materials", "IMPORT of Raw Materials", False, False, False),
        ("Other Purchases", "OTHER PURCHASES", False, False, False),
        ("Total Purchases", "Total Purchases", True, False, False),
        ("Less: Closing Stock", "Closing Stock", False, False, False),
        ("Total COGS", "Total COGS", True, False, False),
        (None, None, False, False, False),
        ("Direct / Manufacturing Expenses", None, False, True, False),
        ("Salaries (Production)", "Salaries (Production)", False, False, False),
        ("Overtime Pay", "Overtime Pay", False, False, False),
        ("Electricity \u2014 Factory", "Electricity - Factory", False, False, False),
        ("Electricity \u2014 Manufacturing", "Electricity - Manufacturing", False, False, False),
        ("Freight Inward", "Freight Inward", False, False, False),
        ("Loading & Unloading", "Loading & Unloading", False, False, False),
        ("Discount Allowed", "Discount Allowed", False, False, False),
        ("Total Direct Expenses", "Total Direct Expenses", True, False, False),
        (None, None, False, False, False),
        ("Gross Profit", "Gross Profit", True, False, False),
        ("Gross Profit %", "Gross Profit %", False, False, True),
        (None, None, False, False, False),
        ("Indirect Expenses (OPEX)", None, False, True, False),
        ("Admin Overheads", "Admin Overheads Total", False, False, False),
        ("Finance Cost", "Finance Cost Total", False, False, False),
        ("HR / Payroll Expenses", "HR Expenses Total", False, False, False),
        ("Marketing & Ads", "Marketing Total", False, False, False),
        ("R&D", "R&D Total", False, False, False),
        ("Professional Services", "Professional Service Charges", False, False, False),
        ("Round Off", "Round Off", False, False, False),
        ("Write Off / Write Back", "Write Off / Write Back", False, False, False),
        ("Total OPEX", "Total OPEX", True, False, False),
        (None, None, False, False, False),
        ("EBITDA", "EBITDA", True, False, False),
        ("EBITDA %", "EBITDA %", False, False, True),
        (None, None, False, False, False),
        ("Depreciation", "Depreciation", False, False, False),
        (None, None, False, False, False),
        ("Profit Before Tax (PBT)", "PBT", True, False, False),
        ("PBT %", "PBT %", False, False, True),
        (None, None, False, False, False),
        ("Tax Provision", "Tax Provision", False, False, False),
        (None, None, False, False, False),
        ("Profit After Tax (PAT)", "PAT", True, False, False),
        ("PAT %", "PAT %", False, False, True),
    ]

    r = 4
    for label, key, bold, is_sec, is_pct in pnl_lines:
        if label is None:
            r += 1
            continue
        if key is None:
            ref_tb(None, r, label, bold, is_section=True)
        else:
            tb_row = ar.get(key)
            if tb_row:
                ref_tb(tb_row, r, label, bold, is_pct=is_pct)
            else:
                ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        r += 1

    # Notes
    r += 2
    notes = [
        "\U0001f4ca  Understanding the Profit & Loss Statement",
        "  \u2022  The P&L answers one question: 'Did the company make money or lose money this period?'",
        "  \u2022  Revenue (top): Total income from all product lines and services.",
        "  \u2022  COGS (middle): Opening Stock + Purchases \u2212 Closing Stock. Reflects actual material consumed, not just purchased.",
        "  \u2022  Direct Expenses: Manufacturing costs (production salaries, electricity, freight). Directly tied to output.",
        "  \u2022  Gross Profit = Revenue \u2212 COGS \u2212 Direct Expenses. A healthy GP % means the business keeps most of what it earns.",
        "  \u2022  OPEX: All indirect running costs \u2014 admin, payroll, marketing, R&D, etc.",
        "  \u2022  EBITDA: The core operating profit. Positive EBITDA = the business generates cash from operations.",
        "  \u2022  PBT: EBITDA minus depreciation. This is the taxable profit.",
        "  \u2022  PAT: The final profit after taxes. This is what flows to the owners\u2019 equity.",
        "  \u2022  All numbers are pulled automatically from the TB sheet \u2014 no manual entry needed here.",
    ]
    write_notes(ws, r, notes)

    ws.column_dimensions['A'].width = 38
    for c in range(RPT_COL_B, RPT_LAST + 1):
        ws.column_dimensions[col_letter(c)].width = 14
    ws.freeze_panes = 'B4'
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 3: CASH FLOW
# ═══════════════════════════════════════════════════════════════════════════════

def build_cashflow_sheet(wb, ar):
    """Cash Flow statement — monthly inflows, outflows, and closing balance."""
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_properties.tabColor = "4472C4"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Cash Flow Statement  |  FY 2025-26  (Cash Basis = P&L)").font = SUBTITLE_FONT

    headers = ['Particulars'] + MONTHS + ['FY Total']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header(ws, 3, RPT_LAST)

    r = 4

    # Opening Cash / Bank Balance
    OPEN_ROW = r
    ws.cell(row=r, column=1, value="Opening Cash / Bank Balance").font = TOTAL_FONT
    # APR opening = manual input (yellow)
    ws.cell(row=r, column=RPT_COL_B).fill = INPUT_FILL
    ws.cell(row=r, column=RPT_COL_B).number_format = INR
    ws.cell(row=r, column=RPT_COL_B).border = THIN_BORDER
    # Remaining months: = previous month's closing
    # We'll set the closing row reference after we build it
    # FY Total = APR opening
    ws.cell(row=r, column=RPT_LAST, value=f"=B{r}")
    fmt_number(ws, r, RPT_COL_B, RPT_LAST)
    style_total(ws, r, RPT_LAST)
    r += 1
    r += 1  # blank

    # CASH INFLOWS
    ws.cell(row=r, column=1, value="CASH INFLOWS").font = SECTION_FONT
    style_section(ws, r, RPT_LAST)
    r += 1

    inflow_start = r
    inflow_items = ["Sale of Products", "Sale of Service", "Export Sales",
                    "Printsticks", "Discount Received", "Interest Income", "Other Income"]
    inflow_labels = ["Sale of Products", "Sale of Service", "Export Sales",
                     "Printsticks", "Discount Received", "Interest Income", "Other Income"]
    for label, key in zip(inflow_labels, inflow_items):
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        tb_row = ar.get(key)
        if tb_row:
            for c in range(RPT_COL_B, RPT_LAST + 1):
                cl = col_letter(c)
                ws.cell(row=r, column=c, value=f"=TB!{cl}{tb_row}")
        fmt_number(ws, r, RPT_COL_B, RPT_LAST)
        r += 1

    # Total Cash Inflows = Total Revenue from TB
    TOTAL_INFLOW_ROW = r
    ws.cell(row=r, column=1, value="Total Cash Inflows").font = TOTAL_FONT
    tb_rev = ar.get("Total Revenue")
    if tb_rev:
        for c in range(RPT_COL_B, RPT_LAST + 1):
            cl = col_letter(c)
            ws.cell(row=r, column=c, value=f"=TB!{cl}{tb_rev}")
    fmt_number(ws, r, RPT_COL_B, RPT_LAST)
    style_total(ws, r, RPT_LAST)
    r += 1
    r += 1  # blank

    # CASH OUTFLOWS
    ws.cell(row=r, column=1, value="CASH OUTFLOWS").font = SECTION_FONT
    style_section(ws, r, RPT_LAST)
    r += 1

    outflow_keys = [
        ("Purchase of Raw Materials", "PURCHASE of Raw Materials"),
        ("Import of Raw Materials", "IMPORT of Raw Materials"),
        ("Other Purchases", "OTHER PURCHASES"),
        ("Salaries (Production)", "Salaries (Production)"),
        ("Overtime Pay", "Overtime Pay"),
        ("Electricity \u2014 Factory", "Electricity - Factory"),
        ("Electricity \u2014 Manufacturing", "Electricity - Manufacturing"),
        ("Freight Inward", "Freight Inward"),
        ("Loading & Unloading", "Loading & Unloading"),
        ("Discount Allowed", "Discount Allowed"),
        ("Admin Overheads", "Admin Overheads Total"),
        ("Finance Cost", "Finance Cost Total"),
        ("HR / Payroll Expenses", "HR Expenses Total"),
        ("Marketing & Ads", "Marketing Total"),
        ("R&D", "R&D Total"),
        ("Professional Services", "Professional Service Charges"),
        ("Round Off", "Round Off"),
        ("Write Off / Write Back", "Write Off / Write Back"),
        ("Depreciation", "Depreciation"),
        ("Tax Paid", "Tax Provision"),
    ]
    outflow_start = r
    for label, key in outflow_keys:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        tb_row = ar.get(key)
        if tb_row:
            for c in range(RPT_COL_B, RPT_LAST + 1):
                cl = col_letter(c)
                ws.cell(row=r, column=c, value=f"=TB!{cl}{tb_row}")
        fmt_number(ws, r, RPT_COL_B, RPT_LAST)
        r += 1
    outflow_end = r - 1

    # Total Cash Outflows
    TOTAL_OUTFLOW_ROW = r
    ws.cell(row=r, column=1, value="Total Cash Outflows").font = TOTAL_FONT
    for c in range(RPT_COL_B, RPT_LAST + 1):
        cl = col_letter(c)
        ws.cell(row=r, column=c, value=f"=SUM({cl}{outflow_start}:{cl}{outflow_end})")
    fmt_number(ws, r, RPT_COL_B, RPT_LAST)
    style_total(ws, r, RPT_LAST)
    r += 1
    r += 1  # blank

    # Net Cash Flow
    NET_CF_ROW = r
    ws.cell(row=r, column=1, value="Net Cash Flow (Inflow \u2212 Outflow)").font = TOTAL_FONT
    for c in range(RPT_COL_B, RPT_LAST + 1):
        cl = col_letter(c)
        ws.cell(row=r, column=c, value=f"={cl}{TOTAL_INFLOW_ROW}-{cl}{TOTAL_OUTFLOW_ROW}")
    fmt_number(ws, r, RPT_COL_B, RPT_LAST)
    style_grand_total(ws, r, RPT_LAST)
    r += 1
    r += 1  # blank

    # Closing Cash / Bank Balance
    CLOSE_ROW = r
    ws.cell(row=r, column=1, value="Closing Cash / Bank Balance").font = TOTAL_FONT
    for c in range(RPT_COL_B, RPT_LAST + 1):
        cl = col_letter(c)
        ws.cell(row=r, column=c, value=f"={cl}{OPEN_ROW}+{cl}{NET_CF_ROW}")
    fmt_number(ws, r, RPT_COL_B, RPT_LAST)
    style_grand_total(ws, r, RPT_LAST)

    # Now wire up Opening balance for months after APR
    for m in range(1, 12):  # MAY onward
        c = RPT_COL_B + m
        prev_cl = col_letter(c - 1)
        ws.cell(row=OPEN_ROW, column=c, value=f"={prev_cl}{CLOSE_ROW}")

    # FY closing
    ws.cell(row=r, column=RPT_LAST, value=f"={col_letter(RPT_LAST)}{OPEN_ROW}+{col_letter(RPT_LAST)}{NET_CF_ROW}")

    r += 2
    ws.cell(row=r, column=1, value="Note: Opening balance for MAY onwards auto-links from previous month's closing. Only enter APR opening.").font = SMALL_FONT
    r += 2

    notes = [
        "\U0001f4b0  Understanding Cash Flow (for non-finance readers)",
        "  \u2022  Cash Flow tracks the actual money moving IN and OUT of the bank account each month.",
        "  \u2022  Unlike P&L (which includes accruals and non-cash items), Cash Flow = what you can actually spend.",
        "  \u2022  Opening Balance: How much cash was in the bank at the start of the month.",
        "  \u2022  Cash Inflows: Money received from customers across all revenue streams.",
        "  \u2022  Cash Outflows: All payments made \u2014 purchases, salaries, admin, marketing, taxes, depreciation, etc.",
        "  \u2022  Net Cash Flow = Inflows \u2212 Outflows. Positive = cash is accumulating. Negative = cash is depleting.",
        "  \u2022  Closing Balance = Opening + Net Cash Flow. This is how much cash the company has at month-end.",
        "  \u2022  Each month's closing balance auto-links as the next month's opening \u2014 only APR opening needs manual entry.",
        "  \u2022  Key insight: A company can be profitable on P&L but still run out of cash if collections are slow.",
    ]
    write_notes(ws, r, notes)

    ws.column_dimensions['A'].width = 38
    for c in range(RPT_COL_B, RPT_LAST + 1):
        ws.column_dimensions[col_letter(c)].width = 14
    ws.freeze_panes = 'B4'


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 4: OPEX SCHEDULE
# ═══════════════════════════════════════════════════════════════════════════════

def build_opex_schedule(wb, ar):
    """Detailed OPEX breakup — every indirect expense item."""
    ws = wb.create_sheet("OPEX Schedule")
    ws.sheet_properties.tabColor = "C00000"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="OPEX Schedule (Indirect Expenses)  |  FY 2025-26").font = SUBTITLE_FONT

    headers = ['Particulars'] + MONTHS + ['FY Total']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header(ws, 3, RPT_LAST)

    def tb_ref_row(r, label, key, is_subtotal=False, is_section=False):
        ws.cell(row=r, column=1, value=label)
        if is_section:
            ws.cell(row=r, column=1).font = SECTION_FONT
            style_section(ws, r, RPT_LAST)
            return
        if is_subtotal:
            ws.cell(row=r, column=1).font = TOTAL_FONT
        else:
            ws.cell(row=r, column=1).font = NORMAL_FONT

        tb_row = ar.get(key)
        if tb_row:
            for c in range(RPT_COL_B, RPT_LAST + 1):
                cl = col_letter(c)
                ws.cell(row=r, column=c, value=f"=TB!{cl}{tb_row}")
        fmt_number(ws, r, RPT_COL_B, RPT_LAST)
        if is_subtotal:
            style_total(ws, r, RPT_LAST)

    r = 4

    # Admin
    tb_ref_row(r, "Office & Admin Overheads", None, is_section=True); r += 1
    tb_ref_row(r, "Office & Administrative Overheads", "Office & Administrative Overheads"); r += 1
    tb_ref_row(r, "Travelling Expense", "Travelling Expense"); r += 1
    tb_ref_row(r, "Rates & Taxes", "Rates & Taxes"); r += 1
    tb_ref_row(r, "Foreign Exchange Gain/Loss", "Foreign Exchange Gain/Loss"); r += 1
    tb_ref_row(r, "Razorpay Charges", "Razorpay Charges"); r += 1
    tb_ref_row(r, "Tender Fee", "Tender Fee"); r += 1
    tb_ref_row(r, "Sub-total", "Admin Overheads Total", is_subtotal=True); r += 1
    r += 1

    # Finance
    tb_ref_row(r, "Finance Cost", None, is_section=True); r += 1
    tb_ref_row(r, "Finance Cost (Interest/Bank)", "Finance Cost (Interest/Bank)"); r += 1
    tb_ref_row(r, "Sub-total", "Finance Cost Total", is_subtotal=True); r += 1
    r += 1

    # HR
    tb_ref_row(r, "HR / Payroll Expenses", None, is_section=True); r += 1
    tb_ref_row(r, "Payroll Expenses", "Payroll Expenses"); r += 1
    tb_ref_row(r, "Sub-total", "HR Expenses Total", is_subtotal=True); r += 1
    r += 1

    # Marketing
    tb_ref_row(r, "Marketing & Ads", None, is_section=True); r += 1
    tb_ref_row(r, "Advertisement/Marketing", "Advertisement/Marketing"); r += 1
    tb_ref_row(r, "Freight Outward", "Freight Outward"); r += 1
    tb_ref_row(r, "Sub-total", "Marketing Total", is_subtotal=True); r += 1
    r += 1

    # R&D
    tb_ref_row(r, "R&D Expenses", None, is_section=True); r += 1
    tb_ref_row(r, "R&D", "R&D"); r += 1
    tb_ref_row(r, "Sub-total", "R&D Total", is_subtotal=True); r += 1
    r += 1

    # Others
    tb_ref_row(r, "Professional Service Charges", "Professional Service Charges"); r += 1
    tb_ref_row(r, "Round Off", "Round Off"); r += 1
    tb_ref_row(r, "Write Off / Write Back", "Write Off / Write Back"); r += 1
    r += 1

    # Grand Total OPEX
    tb_ref_row(r, "GRAND TOTAL OPEX", "Total OPEX", is_subtotal=True)
    style_grand_total(ws, r, RPT_LAST)
    r += 2

    notes = [
        "\U0001f3e2  Understanding OPEX (Operating Expenses)",
        "  \u2022  OPEX = all costs to keep the business running that are NOT directly tied to production.",
        "  \u2022  Office & Admin: Day-to-day office costs \u2014 travel, taxes, forex charges, payment processing.",
        "  \u2022  Finance Cost: Interest paid on loans and bank charges. Lower is better.",
        "  \u2022  HR / Payroll: Usually the biggest OPEX category. Includes all employee-related costs.",
        "  \u2022  Marketing & Ads: Advertising, exhibitions, freight outward. Track as % of Revenue.",
        "  \u2022  R&D: Research & development expenses for new products/services.",
        "  \u2022  Professional Services, Round Off, Write Off: Periodic or one-off items.",
        "  \u2022  Grand Total OPEX: The total monthly overhead. If this exceeds Gross Profit, the company is losing money.",
        "  \u2022  All figures auto-pull from TB \u2014 this is a detailed breakup of OPEX in the P&L.",
    ]
    write_notes(ws, r, notes)

    ws.column_dimensions['A'].width = 38
    for c in range(RPT_COL_B, RPT_LAST + 1):
        ws.column_dimensions[col_letter(c)].width = 14
    ws.freeze_panes = 'B4'


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 5: BALANCE SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_balance_sheet(wb, ar):
    """Monthly Balance Sheet from TB data."""
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_properties.tabColor = "4472C4"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Balance Sheet  |  FY 2025-26").font = SUBTITLE_FONT

    headers = ['Particulars'] + MONTHS
    bs_last = 1 + 12  # Col M = 13
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header(ws, 3, bs_last)

    def bs_ref(r, label, key, is_section=False, is_subtotal=False, is_grand=False):
        ws.cell(row=r, column=1, value=label)
        if is_section:
            ws.cell(row=r, column=1).font = SECTION_FONT
            style_section(ws, r, bs_last)
            return
        if is_grand:
            ws.cell(row=r, column=1).font = GRAND_TOTAL_FONT
        elif is_subtotal:
            ws.cell(row=r, column=1).font = TOTAL_FONT
        else:
            ws.cell(row=r, column=1).font = NORMAL_FONT

        tb_row = ar.get(key) if key else None
        if tb_row:
            for c in range(RPT_COL_B, bs_last + 1):
                cl = col_letter(c)
                ws.cell(row=r, column=c, value=f"=TB!{cl}{tb_row}")
        fmt_number(ws, r, RPT_COL_B, bs_last)
        if is_grand:
            style_grand_total(ws, r, bs_last)
        elif is_subtotal:
            style_total(ws, r, bs_last)

    def bs_sum(r, label, rows, is_grand=False):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = GRAND_TOTAL_FONT if is_grand else TOTAL_FONT
        for c in range(RPT_COL_B, bs_last + 1):
            cl = col_letter(c)
            refs = '+'.join(f"{cl}{rr}" for rr in rows)
            ws.cell(row=r, column=c, value=f"={refs}")
        fmt_number(ws, r, RPT_COL_B, bs_last)
        if is_grand:
            style_grand_total(ws, r, bs_last)
        else:
            style_total(ws, r, bs_last)

    r = 4

    # LIABILITIES
    bs_ref(r, "LIABILITIES", None, is_section=True); r += 1

    bs_ref(r, "Capital Account", None, is_section=True); r += 1
    cap1 = r; bs_ref(r, "Reserves & Surplus", "Reserves & Surplus"); r += 1
    cap2 = r; bs_ref(r, "Share Capital (Issued & Paid Up)", "Issued, Subscribed & Paid Up"); r += 1
    cap_total = r; bs_sum(r, "Capital Account Total", [cap1, cap2]); r += 1
    r += 1

    bs_ref(r, "Loans (Liability)", None, is_section=True); r += 1
    loan1 = r; bs_ref(r, "Unsecured Loans", "Unsecured Loans"); r += 1
    loan_total = r; bs_sum(r, "Loans Total", [loan1]); r += 1
    r += 1

    bs_ref(r, "Current Liabilities", None, is_section=True); r += 1
    cl1 = r; bs_ref(r, "Duties & Taxes", "Duties & Taxes"); r += 1
    cl2 = r; bs_ref(r, "Sundry Creditors", "Sundry Creditors"); r += 1
    cl3 = r; bs_ref(r, "Reimbursements Payable", "Reimbursements Payable"); r += 1
    cl4 = r; bs_ref(r, "Salaries Payable", "Salaries Payable"); r += 1
    cl_total = r; bs_sum(r, "Current Liabilities Total", [cl1, cl2, cl3, cl4]); r += 1
    r += 1

    # Current Period P&L (cumulative PAT)
    pnl_row = r
    ws.cell(row=r, column=1, value="Current Period P&L (Cumulative PAT)").font = NORMAL_FONT
    pat_tb = ar.get("PAT")
    if pat_tb:
        for c in range(RPT_COL_B, bs_last + 1):
            # Cumulative PAT = SUM of TB!B{pat}:{current_col}{pat}
            start_cl = col_letter(RPT_COL_B)
            end_cl = col_letter(c)
            ws.cell(row=r, column=c, value=f"=SUM(TB!{start_cl}{pat_tb}:{end_cl}{pat_tb})")
    fmt_number(ws, r, RPT_COL_B, bs_last)
    r += 1
    pnl_total = r
    bs_sum(r, "P&L Total", [pnl_row]); r += 1
    r += 1

    # TOTAL LIABILITIES
    TOTAL_LIAB_ROW = r
    bs_sum(r, "TOTAL LIABILITIES", [cap_total, loan_total, cl_total, pnl_total], is_grand=True)
    r += 2

    # ASSETS
    bs_ref(r, "ASSETS", None, is_section=True); r += 1

    bs_ref(r, "Fixed Assets", None, is_section=True); r += 1
    fa1 = r; bs_ref(r, "Intangible Assets", "Intangible Assets"); r += 1
    fa2 = r; bs_ref(r, "Machinery & Equipment", "Machinery & Equipment"); r += 1
    fa3 = r; bs_ref(r, "Tangible Assets", "Tangible Assets"); r += 1
    fa4 = r; bs_ref(r, "Fracktal Studio", "Fracktal Studio"); r += 1
    fa5 = r; bs_ref(r, "SLS PAMM Project", "SLS PAMM Project"); r += 1
    fa6 = r; bs_ref(r, "WIP", "WIP"); r += 1
    fa_total = r; bs_sum(r, "Fixed Assets Total", [fa1, fa2, fa3, fa4, fa5, fa6]); r += 1
    r += 1

    bs_ref(r, "Current Assets", None, is_section=True); r += 1
    ca1 = r; bs_ref(r, "Inventories", "Inventories"); r += 1
    ca2 = r; bs_ref(r, "Deposits (Asset)", "Deposits (Asset)"); r += 1
    ca3 = r; bs_ref(r, "Loans & Advances (Asset)", "Loans & Advances (Asset)"); r += 1
    ca4 = r; bs_ref(r, "Sundry Debtors", "Sundry Debtors"); r += 1
    ca5 = r; bs_ref(r, "Cash-in-Hand", "Cash-in-Hand"); r += 1
    ca6 = r; bs_ref(r, "Bank Accounts", "Bank Accounts"); r += 1
    ca7 = r; bs_ref(r, "Deferred Expenses", "Deferred Expenses"); r += 1
    ca8 = r; bs_ref(r, "Deferred Tax Asset", "Deferred Tax Asset"); r += 1
    ca9 = r; bs_ref(r, "TDS - GST Receivable", "TDS - GST Receivable"); r += 1
    ca10 = r; bs_ref(r, "TDS Receivable", "Tds Receivable"); r += 1
    ca_total = r; bs_sum(r, "Current Assets Total", [ca1,ca2,ca3,ca4,ca5,ca6,ca7,ca8,ca9,ca10]); r += 1
    r += 1

    # Suspense
    susp = r; bs_ref(r, "Suspense Account", "Suspense Account"); r += 1
    r += 1

    # TOTAL ASSETS
    TOTAL_ASSETS_ROW = r
    bs_sum(r, "TOTAL ASSETS", [fa_total, ca_total, susp], is_grand=True)
    r += 2

    # DIFFERENCE
    DIFF_ROW = r
    ws.cell(row=r, column=1, value="Difference (Assets \u2212 Liabilities)").font = Font(
        name="Calibri", bold=True, size=11, color=RED)
    for c in range(RPT_COL_B, bs_last + 1):
        cl = col_letter(c)
        ws.cell(row=r, column=c, value=f"={cl}{TOTAL_ASSETS_ROW}-{cl}{TOTAL_LIAB_ROW}")
        ws.cell(row=r, column=c).number_format = INR2
        ws.cell(row=r, column=c).font = Font(name="Calibri", bold=True, size=11, color=RED)
        ws.cell(row=r, column=c).border = TOTAL_BORDER

    # Conditional formatting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for c in range(RPT_COL_B, bs_last + 1):
        cell_ref = f"{col_letter(c)}{DIFF_ROW}"
        ws.conditional_formatting.add(cell_ref, CellIsRule(operator='notEqual', formula=['0'], fill=red_fill))

    r += 2
    notes = [
        "\U0001f3e6  Understanding the Balance Sheet",
        "  \u2022  The Balance Sheet is a \u2018net worth statement\u2019 \u2014 what the company OWNS (Assets) vs what it OWES (Liabilities).",
        "  \u2022  LIABILITIES: Capital (owner\u2019s investment), Loans (borrowed funds), Current Liabilities (short-term dues).",
        "  \u2022  ASSETS: Fixed Assets (equipment, IP, WIP), Current Assets (cash, bank, stock, debtors, deposits).",
        "  \u2022  Golden rule: Total Assets MUST equal Total Liabilities. The Difference row should be \u20b90.",
        "  \u2022  If Difference \u2260 0, some number is missing or wrong \u2014 check the TB input.",
        "  \u2022  Current Period P&L links cumulative PAT from the TB sheet into the balance sheet.",
        "  \u2022  All data auto-pulls from TB \u2014 this is the formatted view for presentation.",
    ]
    write_notes(ws, r, notes)

    ws.column_dimensions['A'].width = 38
    for c in range(RPT_COL_B, bs_last + 1):
        ws.column_dimensions[col_letter(c)].width = 16
    ws.freeze_panes = 'B4'

    return {
        'total_liab': TOTAL_LIAB_ROW,
        'total_assets': TOTAL_ASSETS_ROW,
        'cap_total': cap_total,
        'loan_total': loan_total,
        'cl_total': cl_total,
        'fa_total': fa_total,
        'ca_total': ca_total,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 6: PERFORMANCE SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def build_performance_summary(wb, ar):
    """Quarterly performance summary with FY totals."""
    ws = wb.create_sheet("Performance Summary")
    ws.sheet_properties.tabColor = "548235"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Performance Summary  |  FY 2025-26").font = SUBTITLE_FONT

    headers = ['Particulars', 'Q1 (Apr-Jun)', 'Q2 (Jul-Sep)', 'Q3 (Oct-Dec)', 'Q4 (Jan-Mar)', 'FY Total', 'FY %']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header(ws, 3, 7)

    def qsum(tb_key, rpt_row, label, is_pct=False, fy_denom_key=None):
        """Write quarterly sum row referencing TB."""
        ws.cell(row=rpt_row, column=1, value=label).font = NORMAL_FONT
        tb_row = ar.get(tb_key)
        if tb_row:
            # Q1 = SUM(TB!B:D), Q2 = SUM(TB!E:G), Q3 = SUM(TB!H:J), Q4 = SUM(TB!K:M)
            q_ranges = [('B','D'), ('E','G'), ('H','J'), ('K','M')]
            for qi, (s, e) in enumerate(q_ranges):
                ws.cell(row=rpt_row, column=2+qi, value=f"=SUM(TB!{s}{tb_row}:{e}{tb_row})")

            # FY Total
            ws.cell(row=rpt_row, column=6, value=f"=SUM(B{rpt_row}:E{rpt_row})")

            # FY % (if applicable)
            if fy_denom_key:
                denom_tb = ar.get(fy_denom_key)
                if denom_tb:
                    ws.cell(row=rpt_row, column=7, value=f"=IFERROR(TB!N{tb_row}/TB!N{denom_tb},0)")

        if is_pct:
            fmt_pct(ws, rpt_row, 2, 7)
        else:
            fmt_number(ws, rpt_row, 2, 6)

    def qpct(num_key, denom_key, rpt_row, label):
        """Percentage row: ratio of two TB rows."""
        ws.cell(row=rpt_row, column=1, value=label).font = PCT_FONT
        num_tb = ar.get(num_key)
        denom_tb = ar.get(denom_key)
        if num_tb and denom_tb:
            q_ranges = [('B','D'), ('E','G'), ('H','J'), ('K','M')]
            for qi, (s, e) in enumerate(q_ranges):
                ws.cell(row=rpt_row, column=2+qi,
                    value=f"=IFERROR(SUM(TB!{s}{num_tb}:{e}{num_tb})/SUM(TB!{s}{denom_tb}:{e}{denom_tb}),0)")
            ws.cell(row=rpt_row, column=6, value=f"=IFERROR(TB!N{num_tb}/TB!N{denom_tb},0)")
        fmt_pct(ws, rpt_row, 2, 6)

    r = 4
    qsum("Total Revenue", r, "Revenue"); r += 1
    qsum("Total Revenue Ops", r, "  Revenue from Operations"); r += 1
    qsum("Total Other Income", r, "  Other Income"); r += 1
    r += 1

    qsum("Total COGS", r, "COGS"); r += 1
    r += 1

    qsum("Gross Profit", r, "Gross Profit"); r += 1
    qpct("Gross Profit", "Total Revenue", r, "GP %"); r += 1
    r += 1

    qsum("Total OPEX", r, "Total OPEX"); r += 1
    r += 1

    qsum("EBITDA", r, "EBITDA"); r += 1
    qpct("EBITDA", "Total Revenue", r, "EBITDA %"); r += 1
    r += 1

    qsum("Depreciation", r, "Depreciation"); r += 1
    qsum("PBT", r, "PBT"); r += 1
    qpct("PBT", "Total Revenue", r, "PBT %"); r += 1
    r += 1

    qsum("PAT", r, "PAT"); r += 1
    qpct("PAT", "Total Revenue", r, "PAT %"); r += 1
    r += 2

    notes = [
        "\U0001f4c8  How to Read the Performance Summary",
        "  \u2022  This sheet aggregates monthly P&L data into quarters so you can spot trends at a glance.",
        "  \u2022  Revenue trend: Is quarterly revenue growing, flat, or declining? Consistent growth = healthy demand.",
        "  \u2022  Gross Profit %: Should stay stable. If it drops, COGS is rising faster than revenue.",
        "  \u2022  OPEX trend: Rising OPEX is normal during growth, but as a % of Revenue should stay flat or decline.",
        "  \u2022  EBITDA %: The most watched metric. Improving EBITDA % means efficiency is rising.",
        "  \u2022  PAT: The ultimate measure. Compare Q-over-Q to see if profitability is sustainable.",
        "  \u2022  FY % column: Full-year ratio for industry benchmark comparison.",
    ]
    write_notes(ws, r, notes)

    ws.column_dimensions['A'].width = 35
    for c in range(2, 8):
        ws.column_dimensions[col_letter(c)].width = 16
    ws.freeze_panes = 'B4'


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 7: KPIs
# ═══════════════════════════════════════════════════════════════════════════════

def build_kpis_sheet(wb, ar, bs_rows):
    """Key financial ratios with benchmarks and explanations."""
    ws = wb.create_sheet("KPIs")
    ws.sheet_properties.tabColor = "7030A0"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Key Financial Ratios  |  FY 2025-26").font = SUBTITLE_FONT

    headers = ['KPI / Ratio', 'Q1', 'Q2', 'Q3', 'Q4', 'FY', 'Benchmark']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    style_header(ws, 3, 7)

    def kpi_pct(r, label, num_key, denom_key, benchmark):
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=7, value=benchmark).font = SMALL_FONT
        num_tb = ar.get(num_key)
        denom_tb = ar.get(denom_key)
        if num_tb and denom_tb:
            q_ranges = [('B','D'), ('E','G'), ('H','J'), ('K','M')]
            for qi, (s, e) in enumerate(q_ranges):
                ws.cell(row=r, column=2+qi,
                    value=f"=IFERROR(SUM(TB!{s}{num_tb}:{e}{num_tb})/SUM(TB!{s}{denom_tb}:{e}{denom_tb}),0)")
            ws.cell(row=r, column=6, value=f"=IFERROR(TB!N{num_tb}/TB!N{denom_tb},0)")
        fmt_pct(ws, r, 2, 6)

    def kpi_val(r, label, formula_func, benchmark):
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=7, value=benchmark).font = SMALL_FONT
        formula_func(ws, r)
        fmt_number(ws, r, 2, 6)

    r = 4
    ws.cell(row=r, column=1, value="PROFITABILITY RATIOS").font = SECTION_FONT
    style_section(ws, r, 7); r += 1

    kpi_pct(r, "Gross Profit Margin", "Gross Profit", "Total Revenue", "60-70%"); r += 1
    kpi_pct(r, "EBITDA Margin", "EBITDA", "Total Revenue", "15-25%"); r += 1
    kpi_pct(r, "PAT Margin", "PAT", "Total Revenue", "10-15%"); r += 1
    kpi_pct(r, "COGS / Revenue", "Total COGS", "Total Revenue", "25-35%"); r += 1
    kpi_pct(r, "OPEX / Revenue", "Total OPEX", "Total Revenue", "<40%"); r += 1
    kpi_pct(r, "HR Cost / Revenue", "HR Expenses Total", "Total Revenue", "25-35%"); r += 1
    r += 1

    ws.cell(row=r, column=1, value="OPERATIONAL METRICS").font = SECTION_FONT
    style_section(ws, r, 7); r += 1

    # Avg Monthly Revenue
    rev_tb = ar.get("Total Revenue")
    if rev_tb:
        ws.cell(row=r, column=1, value="Avg Monthly Revenue").font = NORMAL_FONT
        ws.cell(row=r, column=7, value="Growth trend").font = SMALL_FONT
        q_ranges = [('B','D'), ('E','G'), ('H','J'), ('K','M')]
        for qi, (s, e) in enumerate(q_ranges):
            ws.cell(row=r, column=2+qi, value=f"=IFERROR(SUM(TB!{s}{rev_tb}:{e}{rev_tb})/3,0)")
        ws.cell(row=r, column=6, value=f"=IFERROR(TB!N{rev_tb}/12,0)")
        fmt_number(ws, r, 2, 6)
    r += 1

    # Monthly Breakeven Revenue
    opex_tb = ar.get("Total OPEX")
    cogs_tb = ar.get("Total COGS")
    if rev_tb and opex_tb and cogs_tb:
        ws.cell(row=r, column=1, value="Monthly Breakeven Revenue").font = NORMAL_FONT
        ws.cell(row=r, column=7, value="Rev needed to cover OPEX").font = SMALL_FONT
        q_ranges = [('B','D'), ('E','G'), ('H','J'), ('K','M')]
        for qi, (s, e) in enumerate(q_ranges):
            ws.cell(row=r, column=2+qi,
                value=f"=IFERROR(SUM(TB!{s}{opex_tb}:{e}{opex_tb})/3/(1-SUM(TB!{s}{cogs_tb}:{e}{cogs_tb})/SUM(TB!{s}{rev_tb}:{e}{rev_tb})),0)")
        fmt_number(ws, r, 2, 6)
    r += 1
    r += 1

    # Liquidity ratios from Balance Sheet
    ws.cell(row=r, column=1, value="LIQUIDITY RATIOS (from Balance Sheet)").font = SECTION_FONT
    style_section(ws, r, 7); r += 1

    # Current Ratio = CA / CL (monthly, show last month of each quarter)
    ca_total = bs_rows.get('ca_total')
    cl_total = bs_rows.get('cl_total')
    if ca_total and cl_total:
        ws.cell(row=r, column=1, value="Current Ratio").font = NORMAL_FONT
        ws.cell(row=r, column=7, value="> 1.5").font = SMALL_FONT
        # Q1=Jun(D), Q2=Sep(G), Q3=Dec(J), Q4=Mar(M)
        q_end_cols = ['D', 'G', 'J', 'M']
        for qi, ec in enumerate(q_end_cols):
            ws.cell(row=r, column=2+qi,
                value=f"=IFERROR('Balance Sheet'!{ec}{ca_total}/'Balance Sheet'!{ec}{cl_total},0)")
        for c in range(2, 6):
            ws.cell(row=r, column=c).number_format = NUM2
            ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1

    # Cash Position
    ca_cash = ar.get("Cash-in-Hand")
    ca_bank = ar.get("Bank Accounts")
    if ca_cash and ca_bank:
        ws.cell(row=r, column=1, value="Cash Position (Cash + Bank)").font = NORMAL_FONT
        ws.cell(row=r, column=7, value="Positive & growing").font = SMALL_FONT
        q_end_cols = ['D', 'G', 'J', 'M']
        for qi, ec in enumerate(q_end_cols):
            ws.cell(row=r, column=2+qi,
                value=f"=TB!{ec}{ca_cash}+TB!{ec}{ca_bank}")
        fmt_number(ws, r, 2, 5)
    r += 1

    # Cash Runway
    if ca_cash and ca_bank and opex_tb:
        ws.cell(row=r, column=1, value="Cash Runway (months)").font = NORMAL_FONT
        ws.cell(row=r, column=7, value="> 3 months").font = SMALL_FONT
        q_ranges = [('B','D'), ('E','G'), ('H','J'), ('K','M')]
        q_end_cols = ['D', 'G', 'J', 'M']
        for qi, ((s, e), ec) in enumerate(zip(q_ranges, q_end_cols)):
            ws.cell(row=r, column=2+qi,
                value=f"=IFERROR((TB!{ec}{ca_cash}+TB!{ec}{ca_bank})/(SUM(TB!{s}{opex_tb}:{e}{opex_tb})/3),0)")
        for c in range(2, 6):
            ws.cell(row=r, column=c).number_format = NUM2
            ws.cell(row=r, column=c).border = THIN_BORDER
    r += 2

    # Detailed explanations
    notes = [
        "\U0001f4d0  What Each Ratio Means (Plain English)",
        "  \u2022  GROSS PROFIT MARGIN (60-70%): For every \u20b9100 of revenue, how much is left after direct costs.",
        "  \u2022  EBITDA MARGIN (15-25%): Operating profitability before accounting adjustments. Key metric for investors.",
        "  \u2022  PAT MARGIN (10-15%): Final profit percentage after ALL expenses. This grows owners' equity.",
        "  \u2022  COGS / REVENUE (25-35%): How much of every rupee goes to direct material costs. Lower is better.",
        "  \u2022  OPEX / REVENUE (<40%): How much is spent on overheads. Should decline as the business scales.",
        "  \u2022  HR COST / REVENUE (25-35%): Staff cost as a share of revenue. Manufacturing typically 25-35%.",
        "  \u2022  AVG MONTHLY REVENUE: Revenue \u00f7 months. Compare Q-over-Q for growth trends.",
        "  \u2022  MONTHLY BREAKEVEN: Minimum revenue needed to cover OPEX. Above this = profitable.",
        "  \u2022  CURRENT RATIO (>1.5): Can the company pay short-term bills? >1.5 = comfortable.",
        "  \u2022  CASH POSITION: Total cash + bank. Should be positive and growing.",
        "  \u2022  CASH RUNWAY (>3 months): How many months the company can survive on current cash if revenue stopped.",
        "",
        "\U0001f50d  How to Assess Financial Position",
        "  \u2022  REVENUE CHECK: Is revenue growing Q-over-Q? Growth = demand is healthy.",
        "  \u2022  COST CHECK: Are COGS% and OPEX% stable? If yes while revenue grows, the company is scaling well.",
        "  \u2022  PROFIT CHECK: Is EBITDA positive and improving? = business model works + efficiency rising.",
        "  \u2022  CASH CHECK: Is cash growing? If profit is positive but cash is dropping, check collections.",
        "  \u2022  RISK CHECK: Cash Runway > 3 months and Current Ratio > 1.5? Yes = safe buffer.",
        "  \u2022  OVERALL: Green flags = rising revenue + stable margins + growing cash. Red flags = flat revenue + shrinking margins + cash burn.",
    ]
    write_notes(ws, r, notes)

    ws.column_dimensions['A'].width = 35
    for c in range(2, 8):
        ws.column_dimensions[col_letter(c)].width = 18
    ws.freeze_panes = 'B4'


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 8: DASHBOARD (Charts)
# ═══════════════════════════════════════════════════════════════════════════════

def build_dashboard(wb, ar):
    """Dashboard with charts — Revenue vs EBITDA, Expense breakdown."""
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "ED7D31"

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Executive Dashboard \u2014 FY 2025-26").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value="All values auto-calculated from TB").font = SMALL_FONT

    # Data table for charts (hidden in a small area)
    headers = ['Metric'] + MONTHS
    for ci, h in enumerate(headers, 1):
        ws.cell(row=5, column=ci, value=h)
    style_header(ws, 5, 13)

    # Row 6: Revenue
    ws.cell(row=6, column=1, value="Revenue").font = NORMAL_FONT
    rev_tb = ar.get("Total Revenue")
    if rev_tb:
        for m in range(12):
            cl = col_letter(RPT_COL_B + m)
            ws.cell(row=6, column=2 + m, value=f"=TB!{cl}{rev_tb}")
    fmt_number(ws, 6, 2, 13)

    # Row 7: EBITDA
    ws.cell(row=7, column=1, value="EBITDA").font = NORMAL_FONT
    ebitda_tb = ar.get("EBITDA")
    if ebitda_tb:
        for m in range(12):
            cl = col_letter(RPT_COL_B + m)
            ws.cell(row=7, column=2 + m, value=f"=TB!{cl}{ebitda_tb}")
    fmt_number(ws, 7, 2, 13)

    # Row 8: PAT
    ws.cell(row=8, column=1, value="PAT").font = NORMAL_FONT
    pat_tb = ar.get("PAT")
    if pat_tb:
        for m in range(12):
            cl = col_letter(RPT_COL_B + m)
            ws.cell(row=8, column=2 + m, value=f"=TB!{cl}{pat_tb}")
    fmt_number(ws, 8, 2, 13)

    # Row 9: GP
    ws.cell(row=9, column=1, value="Gross Profit").font = NORMAL_FONT
    gp_tb = ar.get("Gross Profit")
    if gp_tb:
        for m in range(12):
            cl = col_letter(RPT_COL_B + m)
            ws.cell(row=9, column=2 + m, value=f"=TB!{cl}{gp_tb}")
    fmt_number(ws, 9, 2, 13)

    # Expense breakdown data (Row 11-16)
    ws.cell(row=11, column=1, value="Expense Category").font = HEADER_FONT
    ws.cell(row=11, column=2, value="FY Amount").font = HEADER_FONT
    style_header(ws, 11, 2)

    exp_items = [
        ("Material Cost", "Total COGS"),
        ("Direct Expenses", "Total Direct Expenses"),
        ("Employee Expenses", "HR Expenses Total"),
        ("Marketing & Selling", "Marketing Total"),
        ("Admin & General", "Admin Overheads Total"),
        ("Finance Cost", "Finance Cost Total"),
    ]
    for i, (label, key) in enumerate(exp_items):
        rr = 12 + i
        ws.cell(row=rr, column=1, value=label).font = NORMAL_FONT
        tb_row = ar.get(key)
        if tb_row:
            ws.cell(row=rr, column=2, value=f"=TB!N{tb_row}")
        ws.cell(row=rr, column=2).number_format = INR
        ws.cell(row=rr, column=2).border = THIN_BORDER

    # ── CHART 1: Revenue vs EBITDA bar chart ──
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Monthly Revenue vs EBITDA vs PAT"
    chart1.y_axis.title = "Amount (INR)"
    chart1.x_axis.title = "Month"
    chart1.width = 28
    chart1.height = 14

    cats = Reference(ws, min_col=2, max_col=13, min_row=5)
    rev_data = Reference(ws, min_col=2, max_col=13, min_row=6)
    ebitda_data = Reference(ws, min_col=2, max_col=13, min_row=7)
    pat_data = Reference(ws, min_col=2, max_col=13, min_row=8)

    chart1.add_data(rev_data, from_rows=True, titles_from_data=False)
    chart1.add_data(ebitda_data, from_rows=True, titles_from_data=False)
    chart1.add_data(pat_data, from_rows=True, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.series[0].title = SeriesLabel(v="Revenue")
    chart1.series[1].title = SeriesLabel(v="EBITDA")
    chart1.series[2].title = SeriesLabel(v="PAT")

    ws.add_chart(chart1, "A20")

    # ── CHART 2: Expense breakdown pie ──
    pie = PieChart()
    pie.title = "FY Expense Breakdown"
    pie.width = 16
    pie.height = 14
    cats_pie = Reference(ws, min_col=1, min_row=12, max_row=17)
    data_pie = Reference(ws, min_col=2, min_row=12, max_row=17)
    pie.add_data(data_pie, titles_from_data=False)
    pie.set_categories(cats_pie)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "I20")

    # ── CHART 3: GP trend line ──
    line = LineChart()
    line.title = "Monthly Gross Profit Trend"
    line.y_axis.title = "Amount (INR)"
    line.width = 28
    line.height = 14
    line.style = 10
    gp_data = Reference(ws, min_col=2, max_col=13, min_row=9)
    line.add_data(gp_data, from_rows=True, titles_from_data=False)
    line.set_categories(cats)
    line.series[0].title = SeriesLabel(v="Gross Profit")
    ws.add_chart(line, "A38")

    ws.column_dimensions['A'].width = 22
    for c in range(2, 14):
        ws.column_dimensions[col_letter(c)].width = 12


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Generate Fracktal Works MIS Master Workbook")
    parser.add_argument('--source', required=True, help='Path to Tally Trial Balance file (.xls/.xlsx)')
    parser.add_argument('--output', default=None, help='Output path (default: .tmp/Fracktal_MIS_Master_<ts>.xlsx)')
    args = parser.parse_args()

    source_path = Path(args.source)
    if not source_path.exists():
        print(f"ERROR: Source file not found: {source_path}")
        sys.exit(1)

    if args.output:
        output_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = TMP_DIR / f"Fracktal_MIS_Master_{ts}.xlsx"

    print(f"Fracktal MIS Generator")
    print(f"  Source: {source_path}")
    print(f"  Output: {output_path}")

    # Read TB
    print("  Reading Trial Balance...")
    tb_data, num_months = read_tally_tb(source_path)
    print(f"  Parsed {len(tb_data)} account rows")

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Build all 8 sheets
    print("  Building TB sheet...")
    _, acct_rows = build_tb_sheet(wb, tb_data, num_months)

    print("  Building P&L sheet...")
    build_pnl_sheet(wb, acct_rows)

    print("  Building Cash Flow sheet...")
    build_cashflow_sheet(wb, acct_rows)

    print("  Building OPEX Schedule sheet...")
    build_opex_schedule(wb, acct_rows)

    print("  Building Balance Sheet...")
    bs_rows = build_balance_sheet(wb, acct_rows)

    print("  Building Performance Summary...")
    build_performance_summary(wb, acct_rows)

    print("  Building KPIs sheet...")
    build_kpis_sheet(wb, acct_rows, bs_rows)

    print("  Building Dashboard...")
    build_dashboard(wb, acct_rows)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\n  MIS Master saved to: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print("  Done!")

    return str(output_path)


if __name__ == "__main__":
    main()
