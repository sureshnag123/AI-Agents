#!/usr/bin/env python3
"""
master_financial_model.py — Generates a formula-driven Master Excel for THOTA HOSPITALITY LLP.

Architecture
────────────
Sheet 1 : TB_Input       — User pastes Trial Balance numbers month-on-month
Sheet 2 : P&L            — Formulae pull from TB_Input
Sheet 3 : Cash Flow      — Same format as P&L (cash basis, no accruals)
Sheet 4 : OPEX Schedule  — Detailed indirect-expense breakup from TB_Input
Sheet 5 : BS_Input       — User pastes Balance Sheet trial balance quarterly
Sheet 6 : Balance Sheet  — Formulae pull from BS_Input
Sheet 7 : Performance    — Quarterly aggregation from P&L
Sheet 8 : KPIs           — Ratios computed from P&L and Balance Sheet

Usage
─────
    python execution/master_financial_model.py                         # empty template
    python execution/master_financial_model.py --prefill <source.xlsx>  # pre-fill with actuals
"""

import os
import sys
import argparse
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle,
)
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(exist_ok=True)

# ── Style constants ──────────────────────────────────────────────────────────

FONT_NAME = "Calibri"
HEADER_FONT    = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
HEADER_FILL    = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT   = Font(name=FONT_NAME, bold=True, size=11, color="2F5496")
SECTION_FILL   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT     = Font(name=FONT_NAME, bold=True, size=11)
TOTAL_FILL     = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TITLE_FONT     = Font(name=FONT_NAME, bold=True, size=16, color="2F5496")
SUBTITLE_FONT  = Font(name=FONT_NAME, bold=True, size=13, color="4472C4")
PCT_FONT       = Font(name=FONT_NAME, italic=True, size=10, color="808080")
INPUT_FILL     = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # light yellow for input cells
NORMAL_FONT    = Font(name=FONT_NAME, size=11)
BOLD_FONT      = Font(name=FONT_NAME, bold=True, size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="2F5496"))
DOUBLE_BOTTOM  = Border(bottom=Side(style="double", color="2F5496"))

# Explanation block styles
EXPLAIN_TITLE_FONT  = Font(name=FONT_NAME, bold=True, size=12, color="2F5496")
EXPLAIN_TITLE_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
EXPLAIN_BULLET_FONT = Font(name=FONT_NAME, size=10, color="333333")
EXPLAIN_BULLET_BOLD = Font(name=FONT_NAME, bold=True, size=10, color="333333")

INR_FMT  = '₹#,##0'
INR_DEC  = '₹#,##0.00'
PCT_FMT  = '0.0%'
NUM_FMT  = '#,##0'

MONTHS = ["APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC", "JAN", "FEB", "MAR"]
# Column B=APR(2), C=MAY(3) … M=MAR(13), N=FY Total(14)
MONTH_START_COL = 2
FY_TOTAL_COL    = 14   # column N

# Quarter column ranges (1-indexed col within months B-M)
Q1_COLS = (2, 3, 4)     # APR MAY JUN  → cols B C D
Q2_COLS = (5, 6, 7)     # JUL AUG SEP  → cols E F G
Q3_COLS = (8, 9, 10)    # OCT NOV DEC  → cols H I J
Q4_COLS = (11, 12, 13)  # JAN FEB MAR  → cols K L M


# ═══════════════════════════════════════════════════════════════════════════════
#  Helper: styling
# ═══════════════════════════════════════════════════════════════════════════════

def _style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_section_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER


def _style_total_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = DOUBLE_BOTTOM


def _style_data_row(ws, row, max_col, bold=False, indent=False, pct=False):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        if pct:
            cell.font = PCT_FONT
        elif bold:
            cell.font = BOLD_FONT
        else:
            cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if indent and c == 1:
            cell.alignment = Alignment(indent=2)


def _set_col_widths(ws, widths: dict):
    """widths = {col_num: width}"""
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def _apply_number_format(ws, row, start_col, end_col, fmt):
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).number_format = fmt


def _input_highlight(ws, row, start_col, end_col):
    """Mark cells as user-input cells (light yellow bg)."""
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = INPUT_FILL


def _fy_total_formula(row, start_col=2, end_col=13):
    """=SUM(B{row}:M{row})"""
    sc = get_column_letter(start_col)
    ec = get_column_letter(end_col)
    return f"=SUM({sc}{row}:{ec}{row})"


def _col(c):
    return get_column_letter(c)


def _add_explanation_block(ws, start_row, max_col, title, bullets):
    """
    Add an explanation section below the data table.
    title  – Section heading (e.g., "How to Read This Sheet")
    bullets – list of strings; lines starting with "**" are bold-prefix lines.
    Returns the next free row after the block.
    """
    from openpyxl.utils import get_column_letter
    # Blank spacer row
    r = start_row + 2

    # Merged title row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1, value=title)
    cell.font = EXPLAIN_TITLE_FONT
    cell.fill = EXPLAIN_TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    r += 1

    # Bullet lines
    for text in bullets:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
        cell = ws.cell(row=r, column=1, value=f"  •  {text}")
        cell.font = EXPLAIN_BULLET_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Make the merged area tall enough for wrapped text
        ws.row_dimensions[r].height = max(30, 15 * (1 + len(text) // 80))
        r += 1

    return r


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet 1: TB_Input — Trial Balance Input
# ═══════════════════════════════════════════════════════════════════════════════

# Master list of account heads.  Each entry: (label, type, indent)
# type: "section" | "input" | "formula" | "pct" | "blank"
# For "formula" type, a formula-builder func is provided later.

TB_ACCOUNTS = [
    # --- REVENUE ---
    ("REVENUE",                       "section", False),
    ("Morning Glory",                 "input",   True),
    ("Al Fresco",                     "input",   True),
    ("Sunset Soiree",                 "input",   True),
    ("Studio",                        "input",   True),
    ("Other Income",                  "input",   True),
    ("Credit Note / Returns",         "input",   True),
    ("Total Revenue",                 "formula", False),   # SUM of above
    ("",                              "blank",   False),
    # --- COGS ---
    ("COST OF GOODS SOLD (COGS)",     "section", False),
    ("Purchase - Thota Kitchen",      "input",   True),
    ("Purchase - Thota Decor",        "input",   True),
    ("Purchase - Studio Decor",       "input",   True),
    ("Fuel Expenses",                 "input",   True),
    ("Transportation Expense",        "input",   True),
    ("Total COGS",                    "formula", False),   # SUM
    ("",                              "blank",   False),
    ("Gross Profit",                  "formula", False),   # Revenue − COGS
    ("Gross Profit %",                "pct",     False),   # GP / Revenue
    ("",                              "blank",   False),
    # --- INDIRECT EXPENSES ---
    ("INDIRECT EXPENSES (OPEX)",      "section", False),
    ("Office & Admin Overheads",      "section_sub", False),
    ("Local Conveyance",              "input",   True),
    ("Office Expenses",               "input",   True),
    ("Printing & Stationery",         "input",   True),
    ("Registration & Subscription",   "input",   True),
    ("Staff Welfare",                 "input",   True),
    ("Telephone & Internet Charges",  "input",   True),
    ("Admin Overheads Total",         "formula", False),
    ("",                              "blank",   False),
    ("Finance Cost",                  "section_sub", False),
    ("Bank Charges",                  "input",   True),
    ("Interest on Bank Loan",         "input",   True),
    ("Finance Cost Total",            "formula", False),
    ("",                              "blank",   False),
    ("HR / Payroll Expenses",         "section_sub", False),
    ("Employee Salaries",             "input",   True),
    ("Incentives",                    "input",   True),
    ("Partner Remuneration",          "input",   True),
    ("PF Admin Charges",              "input",   True),
    ("Stipend",                       "input",   True),
    ("HR Expenses Total",             "formula", False),
    ("",                              "blank",   False),
    ("Marketing & Ads",               "section_sub", False),
    ("Meta / Instagram",              "input",   True),
    ("Events / Exhibitions",          "input",   True),
    ("Marketing Total",               "formula", False),
    ("",                              "blank",   False),
    ("Professional Service Charges",  "input",   False),
    ("Rates & Taxes",                 "input",   False),
    ("Repairs & Maintenance",         "input",   False),
    ("Misc. Expenses",                "input",   False),
    ("Write Off / (-)Write Back",     "input",   False),
    ("",                              "blank",   False),
    ("Total Indirect Expenses",       "formula", False),
    ("",                              "blank",   False),
    # --- EBITDA ---
    ("EBITDA",                        "formula", False),
    ("EBITDA %",                      "pct",     False),
    ("",                              "blank",   False),
    ("Depreciation",                  "input",   False),
    ("",                              "blank",   False),
    ("Profit Before Tax (PBT)",       "formula", False),
    ("PBT %",                         "pct",     False),
    ("",                              "blank",   False),
    ("Tax Provision",                 "input",   False),
    ("",                              "blank",   False),
    ("Profit After Tax (PAT)",        "formula", False),
    ("PAT %",                         "pct",     False),
]

# We need to know the row numbers for formula references.
# TB_Input data starts at row 4 (rows 1-3 are title/header).
TB_DATA_START = 4


def _tb_row(label: str) -> int:
    """Return the actual Excel row for a TB account label."""
    for i, (lbl, *_) in enumerate(TB_ACCOUNTS):
        if lbl == label:
            return TB_DATA_START + i
    raise KeyError(f"TB label not found: {label}")


# Formula builders for TB_Input.  They receive a column letter.
def _tb_formulas():
    """Return {label: formula_template} for TB_Input."""
    r = _tb_row
    return {
        "Total Revenue": lambda c: f"=SUM({c}{r('Morning Glory')}:{c}{r('Credit Note / Returns')})",
        "Total COGS": lambda c: f"=SUM({c}{r('Purchase - Thota Kitchen')}:{c}{r('Transportation Expense')})",
        "Gross Profit": lambda c: f"={c}{r('Total Revenue')}-{c}{r('Total COGS')}",
        "Gross Profit %": lambda c: f"=IFERROR({c}{r('Gross Profit')}/{c}{r('Total Revenue')},0)",
        "Admin Overheads Total": lambda c: f"=SUM({c}{r('Local Conveyance')}:{c}{r('Telephone & Internet Charges')})",
        "Finance Cost Total": lambda c: f"=SUM({c}{r('Bank Charges')}:{c}{r('Interest on Bank Loan')})",
        "HR Expenses Total": lambda c: f"=SUM({c}{r('Employee Salaries')}:{c}{r('Stipend')})",
        "Marketing Total": lambda c: f"=SUM({c}{r('Meta / Instagram')}:{c}{r('Events / Exhibitions')})",
        "Total Indirect Expenses": lambda c: (
            f"={c}{r('Admin Overheads Total')}+{c}{r('Finance Cost Total')}"
            f"+{c}{r('HR Expenses Total')}+{c}{r('Marketing Total')}"
            f"+{c}{r('Professional Service Charges')}+{c}{r('Rates & Taxes')}"
            f"+{c}{r('Repairs & Maintenance')}+{c}{r('Misc. Expenses')}"
            f"+{c}{r('Write Off / (-)Write Back')}"
        ),
        "EBITDA": lambda c: f"={c}{r('Gross Profit')}-{c}{r('Total Indirect Expenses')}",
        "EBITDA %": lambda c: f"=IFERROR({c}{r('EBITDA')}/{c}{r('Total Revenue')},0)",
        "Profit Before Tax (PBT)": lambda c: f"={c}{r('EBITDA')}-{c}{r('Depreciation')}",
        "PBT %": lambda c: f"=IFERROR({c}{r('Profit Before Tax (PBT)')}/{c}{r('Total Revenue')},0)",
        "Profit After Tax (PAT)": lambda c: f"={c}{r('Profit Before Tax (PBT)')}-{c}{r('Tax Provision')}",
        "PAT %": lambda c: f"=IFERROR({c}{r('Profit After Tax (PAT)')}/{c}{r('Total Revenue')},0)",
    }


def build_tb_input(wb):
    """Build the TB_Input sheet."""
    ws = wb.create_sheet("TB_Input", 0)
    max_col = FY_TOTAL_COL  # 14

    # Title rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP — Trial Balance Input").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value="Paste monthly Trial Balance numbers below. FY Total auto-calculates.").font = Font(
        name=FONT_NAME, italic=True, size=10, color="808080",
    )
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # Header row 3
    headers = ["Particulars"] + MONTHS + ["FY Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, max_col)

    # Column widths
    _set_col_widths(ws, {1: 35})
    for c in range(2, max_col + 1):
        _set_col_widths(ws, {c: 14})

    # Account rows
    formulas = _tb_formulas()

    for i, (label, typ, indent) in enumerate(TB_ACCOUNTS):
        row = TB_DATA_START + i
        ws.cell(row=row, column=1, value=label)

        if typ == "section":
            _style_section_row(ws, row, max_col)
        elif typ == "section_sub":
            ws.cell(row=row, column=1).font = BOLD_FONT
            ws.cell(row=row, column=1).fill = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid")
            for c in range(2, max_col + 1):
                ws.cell(row=row, column=c).fill = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid")
                ws.cell(row=row, column=c).border = THIN_BORDER
        elif typ == "formula":
            _style_total_row(ws, row, max_col)
            if label in formulas:
                for c in range(2, 14):  # B to M
                    col_letter = _col(c)
                    ws.cell(row=row, column=c, value=formulas[label](col_letter))
                # FY Total
                ws.cell(row=row, column=FY_TOTAL_COL, value=_fy_total_formula(row))
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif typ == "pct":
            if label in formulas:
                for c in range(2, 14):
                    ws.cell(row=row, column=c, value=formulas[label](_col(c)))
                # FY pct = FY GP / FY Revenue (not SUM of monthly pcts)
                # Special handling for each % row
                if "Gross Profit %" in label:
                    ws.cell(row=row, column=FY_TOTAL_COL,
                            value=f"=IFERROR({_col(FY_TOTAL_COL)}{_tb_row('Gross Profit')}/{_col(FY_TOTAL_COL)}{_tb_row('Total Revenue')},0)")
                elif "EBITDA %" in label:
                    ws.cell(row=row, column=FY_TOTAL_COL,
                            value=f"=IFERROR({_col(FY_TOTAL_COL)}{_tb_row('EBITDA')}/{_col(FY_TOTAL_COL)}{_tb_row('Total Revenue')},0)")
                elif "PBT %" in label:
                    ws.cell(row=row, column=FY_TOTAL_COL,
                            value=f"=IFERROR({_col(FY_TOTAL_COL)}{_tb_row('Profit Before Tax (PBT)')}/{_col(FY_TOTAL_COL)}{_tb_row('Total Revenue')},0)")
                elif "PAT %" in label:
                    ws.cell(row=row, column=FY_TOTAL_COL,
                            value=f"=IFERROR({_col(FY_TOTAL_COL)}{_tb_row('Profit After Tax (PAT)')}/{_col(FY_TOTAL_COL)}{_tb_row('Total Revenue')},0)")
            _style_data_row(ws, row, max_col, pct=True)
            _apply_number_format(ws, row, 2, max_col, PCT_FMT)
        elif typ == "input":
            _style_data_row(ws, row, max_col, indent=indent)
            # Highlight input cells
            _input_highlight(ws, row, 2, 13)
            # FY Total formula
            ws.cell(row=row, column=FY_TOTAL_COL, value=_fy_total_formula(row))
            ws.cell(row=row, column=FY_TOTAL_COL).font = BOLD_FONT
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif typ == "blank":
            pass  # empty spacer row

    # Freeze panes: freeze row 3 and column A
    ws.freeze_panes = "B4"

    # ── Explanation block ──
    last_data_row = TB_DATA_START + len(TB_ACCOUNTS) - 1
    _add_explanation_block(ws, last_data_row, max_col,
        "📋  How to Use This Sheet (for non-finance readers)",
        [
            "Trial Balance is simply a list of all money that came IN (Revenue) and went OUT (Costs) during each month.",
            "Yellow cells = input cells. Paste the monthly accounting numbers here. Everything else auto-calculates.",
            "Revenue section: All income streams of the business — Morning Glory, Al Fresco, Sunset Soiree, Studio, and Other Income.",
            "COGS (Cost of Goods Sold): The direct costs to deliver services — raw materials (Kitchen, Decor purchases), fuel, and transport.",
            "Gross Profit = Revenue minus COGS. This tells you how much money is left after paying for direct delivery costs.",
            "Indirect Expenses (OPEX): Ongoing costs to run the business — salaries, rent, admin, marketing, etc. These are NOT tied to a specific order.",
            "EBITDA = Gross Profit minus OPEX. Think of this as 'operating cash earned' before accounting adjustments like depreciation and tax.",
            "PAT (Profit After Tax) = The final bottom line. This is what the company actually earned after all expenses and taxes.",
            "FY Total (last column) auto-sums all 12 months. Percentage rows show each metric as a share of Total Revenue.",
        ]
    )

    # Print area
    ws.print_area = f"A1:{_col(max_col)}{last_data_row}"

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet 2: P&L — Profit & Loss (formula-linked to TB_Input)
# ═══════════════════════════════════════════════════════════════════════════════

def build_pl_sheet(wb):
    """Create P&L sheet with all cells referencing TB_Input."""
    ws = wb.create_sheet("P&L")
    max_col = FY_TOTAL_COL
    TB = "TB_Input"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value="Profit & Loss Statement  |  FY 2025-26").font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # Header row
    headers = ["Particulars"] + MONTHS + ["FY Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, max_col)

    # Column widths
    _set_col_widths(ws, {1: 35})
    for c in range(2, max_col + 1):
        _set_col_widths(ws, {c: 14})

    # P&L line items — each references the corresponding TB_Input row
    # format: (label, tb_label, style_type)
    # style_type: "section" | "item" | "subtotal" | "total" | "pct" | "blank"
    PL_LINES = [
        ("Revenue from Operations",   None,                       "section"),
        ("Morning Glory",             "Morning Glory",            "item"),
        ("Al Fresco",                 "Al Fresco",                "item"),
        ("Sunset Soiree",             "Sunset Soiree",            "item"),
        ("Studio",                    "Studio",                   "item"),
        ("Other Income",              "Other Income",             "item"),
        ("Credit Note / Returns",     "Credit Note / Returns",    "item"),
        ("Total Revenue",             "Total Revenue",            "subtotal"),
        ("",                          None,                       "blank"),
        ("Cost of Goods Sold",        None,                       "section"),
        ("Purchase - Thota Kitchen",  "Purchase - Thota Kitchen", "item"),
        ("Purchase - Thota Decor",    "Purchase - Thota Decor",   "item"),
        ("Purchase - Studio Decor",   "Purchase - Studio Decor",  "item"),
        ("Fuel Expenses",             "Fuel Expenses",            "item"),
        ("Transportation Expense",    "Transportation Expense",   "item"),
        ("Total COGS",                "Total COGS",               "subtotal"),
        ("",                          None,                       "blank"),
        ("Gross Profit",              "Gross Profit",             "total"),
        ("Gross Profit %",            "Gross Profit %",           "pct"),
        ("",                          None,                       "blank"),
        ("Indirect Expenses (OPEX)",  None,                       "section"),
        ("Admin Overheads",           "Admin Overheads Total",    "item"),
        ("Finance Cost",              "Finance Cost Total",       "item"),
        ("HR / Payroll Expenses",     "HR Expenses Total",        "item"),
        ("Marketing & Ads",           "Marketing Total",          "item"),
        ("Professional Services",     "Professional Service Charges", "item"),
        ("Rates & Taxes",             "Rates & Taxes",            "item"),
        ("Repairs & Maintenance",     "Repairs & Maintenance",    "item"),
        ("Misc. Expenses",            "Misc. Expenses",           "item"),
        ("Write Off / (-)Write Back", "Write Off / (-)Write Back","item"),
        ("Total OPEX",                "Total Indirect Expenses",  "subtotal"),
        ("",                          None,                       "blank"),
        ("EBITDA",                    "EBITDA",                   "total"),
        ("EBITDA %",                  "EBITDA %",                 "pct"),
        ("",                          None,                       "blank"),
        ("Depreciation",              "Depreciation",             "item"),
        ("",                          None,                       "blank"),
        ("Profit Before Tax (PBT)",   "Profit Before Tax (PBT)", "total"),
        ("PBT %",                     "PBT %",                    "pct"),
        ("",                          None,                       "blank"),
        ("Tax Provision",             "Tax Provision",            "item"),
        ("",                          None,                       "blank"),
        ("Profit After Tax (PAT)",    "Profit After Tax (PAT)",   "total"),
        ("PAT %",                     "PAT %",                    "pct"),
    ]

    data_start = 4
    for i, (label, tb_label, stype) in enumerate(PL_LINES):
        row = data_start + i
        ws.cell(row=row, column=1, value=label)

        if stype == "blank":
            continue

        # Put formula references to TB_Input
        if tb_label:
            tb_r = _tb_row(tb_label)
            for c in range(2, max_col + 1):
                cl = _col(c)
                ws.cell(row=row, column=c, value=f"={TB}!{cl}{tb_r}")

        # Styling
        if stype == "section":
            _style_section_row(ws, row, max_col)
        elif stype == "item":
            _style_data_row(ws, row, max_col, indent=True)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "subtotal":
            _style_total_row(ws, row, max_col)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "total":
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = Font(name=FONT_NAME, bold=True, size=12, color="2F5496")
                cell.border = DOUBLE_BOTTOM
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "pct":
            _style_data_row(ws, row, max_col, pct=True)
            _apply_number_format(ws, row, 2, max_col, PCT_FMT)

    ws.freeze_panes = "B4"

    # ── Explanation block ──
    last_data_row = data_start + len(PL_LINES) - 1
    _add_explanation_block(ws, last_data_row, max_col,
        "📊  Understanding the Profit & Loss Statement",
        [
            "The P&L answers one question: 'Did the company make money or lose money this period?'",
            "Revenue (top): Total income from all business units. Higher revenue = more customers or higher pricing.",
            "COGS (middle): Direct costs to serve those customers. If COGS grows faster than Revenue, pricing or sourcing needs attention.",
            "Gross Profit = Revenue − COGS. A healthy Gross Profit % (60−70%) means the business keeps most of what it earns before overheads.",
            "OPEX: All indirect running costs — salaries, admin, marketing, interest, etc. Keeping OPEX below 40% of Revenue is a good benchmark.",
            "EBITDA: The core operating profit. If EBITDA is positive, the business generates cash from operations. If negative, it's burning money.",
            "PBT (Profit Before Tax): EBITDA minus depreciation (accounting charge for asset wear). This is the taxable profit.",
            "PAT (Profit After Tax): The final profit after the government's share. This is what flows to the owners' equity.",
            "All numbers are pulled automatically from the TB_Input sheet — no manual entry needed here.",
        ]
    )

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet 3: Cash Flow — Same format as P&L (cash basis)
# ═══════════════════════════════════════════════════════════════════════════════

def build_cashflow_sheet(wb):
    """
    Cash Flow mirrors P&L because:
    - Cash burn is real-time (no accruals for purchases)
    - Revenue = Cash received, Expenses = Cash paid

    Additionally adds Opening / Closing bank balance rows.
    """
    ws = wb.create_sheet("Cash Flow")
    max_col = FY_TOTAL_COL
    TB = "TB_Input"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value="Cash Flow Statement  |  FY 2025-26  (Cash Basis = P&L)").font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["Particulars"] + MONTHS + ["FY Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, max_col)

    _set_col_widths(ws, {1: 38})
    for c in range(2, max_col + 1):
        _set_col_widths(ws, {c: 14})

    # Row layout
    CF_LINES = [
        ("Opening Cash / Bank Balance",  None,  "input_row"),   # user inputs opening
        ("",                              None,  "blank"),
        ("CASH INFLOWS",                  None,  "section"),
        ("Morning Glory",                "Morning Glory",          "item"),
        ("Al Fresco",                    "Al Fresco",              "item"),
        ("Sunset Soiree",                "Sunset Soiree",          "item"),
        ("Studio",                       "Studio",                 "item"),
        ("Other Income",                 "Other Income",           "item"),
        ("Credit Note / Returns",        "Credit Note / Returns",  "item"),
        ("Total Cash Inflows",           "Total Revenue",          "subtotal"),
        ("",                              None,  "blank"),
        ("CASH OUTFLOWS",                 None,  "section"),
        ("Purchase - Thota Kitchen",     "Purchase - Thota Kitchen",  "item"),
        ("Purchase - Thota Decor",       "Purchase - Thota Decor",    "item"),
        ("Purchase - Studio Decor",      "Purchase - Studio Decor",   "item"),
        ("Fuel Expenses",                "Fuel Expenses",             "item"),
        ("Transportation Expense",       "Transportation Expense",    "item"),
        ("Admin Overheads",              "Admin Overheads Total",     "item"),
        ("Finance Cost",                 "Finance Cost Total",        "item"),
        ("HR / Payroll Expenses",        "HR Expenses Total",         "item"),
        ("Marketing & Ads",             "Marketing Total",            "item"),
        ("Professional Services",        "Professional Service Charges", "item"),
        ("Rates & Taxes",               "Rates & Taxes",              "item"),
        ("Repairs & Maintenance",        "Repairs & Maintenance",     "item"),
        ("Misc. Expenses",              "Misc. Expenses",             "item"),
        ("Depreciation",                "Depreciation",               "item"),
        ("Tax Paid",                    "Tax Provision",              "item"),
        ("Total Cash Outflows",          None,  "outflow_total"),
        ("",                              None,  "blank"),
        ("Net Cash Flow (Inflow − Outflow)", None, "net_cf"),
        ("",                              None,  "blank"),
        ("Closing Cash / Bank Balance",   None,  "closing"),
    ]

    data_start = 4
    # Track rows for formulas
    row_map = {}
    for i, (label, tb_label, stype) in enumerate(CF_LINES):
        row = data_start + i
        row_map[label] = row
        ws.cell(row=row, column=1, value=label)

        if stype == "blank":
            continue
        elif stype == "section":
            _style_section_row(ws, row, max_col)
        elif stype == "input_row":
            _style_data_row(ws, row, max_col, bold=True)
            _input_highlight(ws, row, 2, 13)
            # FY Total for opening is just first month opening (not SUM)
            # Actually for opening balance, FY total = APR opening
            ws.cell(row=row, column=FY_TOTAL_COL, value=f"=B{row}")
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "item":
            tb_r = _tb_row(tb_label)
            for c in range(2, max_col + 1):
                ws.cell(row=row, column=c, value=f"=TB_Input!{_col(c)}{tb_r}")
            _style_data_row(ws, row, max_col, indent=True)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "subtotal":
            tb_r = _tb_row(tb_label)
            for c in range(2, max_col + 1):
                ws.cell(row=row, column=c, value=f"=TB_Input!{_col(c)}{tb_r}")
            _style_total_row(ws, row, max_col)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "outflow_total":
            # Sum outflow items (from "Purchase - Thota Kitchen" to "Tax Paid")
            first_out = row_map["Purchase - Thota Kitchen"]
            last_out = row - 1  # row above this total
            for c in range(2, max_col + 1):
                cl = _col(c)
                ws.cell(row=row, column=c, value=f"=SUM({cl}{first_out}:{cl}{last_out})")
            _style_total_row(ws, row, max_col)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "net_cf":
            inflow_row = row_map["Total Cash Inflows"]
            outflow_row = row_map["Total Cash Outflows"]
            for c in range(2, max_col + 1):
                cl = _col(c)
                ws.cell(row=row, column=c, value=f"={cl}{inflow_row}-{cl}{outflow_row}")
            # Total for net CF = sum of monthly
            ws.cell(row=row, column=FY_TOTAL_COL, value=_fy_total_formula(row, 2, 13))
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = Font(name=FONT_NAME, bold=True, size=12, color="2F5496")
                cell.border = DOUBLE_BOTTOM
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "closing":
            opening_row = row_map["Opening Cash / Bank Balance"]
            net_row = row_map["Net Cash Flow (Inflow − Outflow)"]
            for c in range(2, max_col + 1):
                cl = _col(c)
                ws.cell(row=row, column=c, value=f"={cl}{opening_row}+{cl}{net_row}")
            # Next month's opening = this month's closing
            # We'll add a note about this
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = Font(name=FONT_NAME, bold=True, size=12, color="006100")
                cell.border = DOUBLE_BOTTOM
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            _apply_number_format(ws, row, 2, max_col, INR_FMT)

    # Link Opening Balance: MAY opening = APR closing, JUN = MAY closing, etc.
    opening_row = row_map["Opening Cash / Bank Balance"]
    closing_row = row_map["Closing Cash / Bank Balance"]
    for c in range(3, 14):  # C(MAY) to M(MAR)
        prev_col = _col(c - 1)
        ws.cell(row=opening_row, column=c, value=f"={prev_col}{closing_row}")
        ws.cell(row=opening_row, column=c).fill = PatternFill()  # clear input highlight for linked cells
        ws.cell(row=opening_row, column=c).font = BOLD_FONT
    # Only APR opening remains yellow (user input)
    ws.cell(row=opening_row, column=2).fill = INPUT_FILL

    # Add note
    note_row = data_start + len(CF_LINES) + 1
    ws.cell(row=note_row, column=1,
            value="Note: Opening balance for MAY onwards auto-links from previous month's closing. Only enter APR opening.").font = Font(
        name=FONT_NAME, italic=True, size=9, color="808080")

    ws.freeze_panes = "B4"

    # ── Explanation block ──
    _add_explanation_block(ws, note_row, max_col,
        "💰  Understanding Cash Flow (for non-finance readers)",
        [
            "Cash Flow tracks the actual money moving IN and OUT of the bank account each month — not promises, just real transactions.",
            "Unlike P&L (which includes accruals and non-cash items), Cash Flow = what you can actually spend.",
            "Opening Balance: How much cash was in the bank at the start of the month.",
            "Cash Inflows: Money received from customers across all business units. Same as Revenue in P&L.",
            "Cash Outflows: All payments made — purchases, salaries, rent, marketing, taxes, etc.",
            "Net Cash Flow = Inflows − Outflows. Positive = cash is accumulating. Negative = cash is depleting.",
            "Closing Balance = Opening + Net Cash Flow. This is how much cash the company has at month-end.",
            "Each month's closing balance auto-links as the next month's opening — only APR opening needs manual entry.",
            "Key insight: A company can be profitable on P&L but still run out of cash if collections are slow or expenses spike.",
        ]
    )

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet 4: OPEX Schedule — Detailed indirect expenses from TB_Input
# ═══════════════════════════════════════════════════════════════════════════════

def build_opex_sheet(wb):
    ws = wb.create_sheet("OPEX Schedule")
    max_col = FY_TOTAL_COL
    TB = "TB_Input"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value="OPEX Schedule (Indirect Expenses)  |  FY 2025-26").font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["Particulars"] + MONTHS + ["FY Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, max_col)
    _set_col_widths(ws, {1: 35})
    for c in range(2, max_col + 1):
        _set_col_widths(ws, {c: 14})

    OPEX_ITEMS = [
        ("Office & Admin Overheads",    None,                          "section"),
        ("Local Conveyance",            "Local Conveyance",            "item"),
        ("Office Expenses",             "Office Expenses",             "item"),
        ("Printing & Stationery",       "Printing & Stationery",       "item"),
        ("Registration & Subscription", "Registration & Subscription", "item"),
        ("Staff Welfare",               "Staff Welfare",               "item"),
        ("Telephone & Internet",        "Telephone & Internet Charges","item"),
        ("Sub-total",                   "Admin Overheads Total",       "subtotal"),
        ("",                            None,                          "blank"),
        ("Finance Cost",                None,                          "section"),
        ("Bank Charges",                "Bank Charges",                "item"),
        ("Interest on Bank Loan",       "Interest on Bank Loan",       "item"),
        ("Sub-total",                   "Finance Cost Total",          "subtotal"),
        ("",                            None,                          "blank"),
        ("HR / Payroll Expenses",       None,                          "section"),
        ("Employee Salaries",           "Employee Salaries",           "item"),
        ("Incentives",                  "Incentives",                  "item"),
        ("Partner Remuneration",        "Partner Remuneration",        "item"),
        ("PF Admin Charges",            "PF Admin Charges",            "item"),
        ("Stipend",                     "Stipend",                     "item"),
        ("Sub-total",                   "HR Expenses Total",           "subtotal"),
        ("",                            None,                          "blank"),
        ("Marketing & Ads",             None,                          "section"),
        ("Meta / Instagram",            "Meta / Instagram",            "item"),
        ("Events / Exhibitions",        "Events / Exhibitions",        "item"),
        ("Sub-total",                   "Marketing Total",             "subtotal"),
        ("",                            None,                          "blank"),
        ("Professional Service Charges","Professional Service Charges","item_bold"),
        ("Rates & Taxes",               "Rates & Taxes",               "item_bold"),
        ("Repairs & Maintenance",       "Repairs & Maintenance",       "item_bold"),
        ("Misc. Expenses",              "Misc. Expenses",              "item_bold"),
        ("Write Off / (-)Write Back",   "Write Off / (-)Write Back",   "item_bold"),
        ("",                            None,                          "blank"),
        ("GRAND TOTAL OPEX",            "Total Indirect Expenses",     "grand"),
    ]

    data_start = 4
    for i, (label, tb_label, stype) in enumerate(OPEX_ITEMS):
        row = data_start + i
        ws.cell(row=row, column=1, value=label)

        if stype == "blank":
            continue
        if stype == "section":
            _style_section_row(ws, row, max_col)
            continue

        # Link to TB_Input
        if tb_label:
            tb_r = _tb_row(tb_label)
            for c in range(2, max_col + 1):
                ws.cell(row=row, column=c, value=f"={TB}!{_col(c)}{tb_r}")

        if stype == "item":
            _style_data_row(ws, row, max_col, indent=True)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "item_bold":
            _style_data_row(ws, row, max_col, bold=True)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "subtotal":
            _style_total_row(ws, row, max_col)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif stype == "grand":
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                cell.border = DOUBLE_BOTTOM
            _apply_number_format(ws, row, 2, max_col, INR_FMT)

    ws.freeze_panes = "B4"

    # ── Explanation block ──
    last_data_row = data_start + len(OPEX_ITEMS) - 1
    _add_explanation_block(ws, last_data_row, max_col,
        "🏢  Understanding OPEX (Operating Expenses)",
        [
            "OPEX = all costs to keep the business running that are NOT directly tied to a specific customer order.",
            "Office & Admin: Day-to-day office costs — conveyance, stationery, internet, staff welfare. Typically small but can creep up.",
            "Finance Cost: Interest paid on loans and bank charges. Lower is better — it means less dependency on debt.",
            "HR / Payroll: The biggest OPEX category usually. Includes salaries, incentives, partner remuneration, PF charges, and stipends.",
            "Marketing & Ads: Spend on Meta/Instagram and events. Track this as a % of Revenue — ideally under 5-8% for hospitality.",
            "Professional Services, Rates & Taxes, Repairs, Misc: One-off or periodic costs. Watch for sudden spikes.",
            "Grand Total OPEX: The total monthly overhead. If this exceeds Gross Profit, the company is losing money operationally.",
            "All figures auto-pull from TB_Input — this sheet is a detailed breakup of the OPEX line you see in the P&L.",
        ]
    )

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet 5: BS_Input — Balance Sheet Trial Balance Input
# ═══════════════════════════════════════════════════════════════════════════════

BS_ACCOUNTS = [
    ("LIABILITIES",                   "section",  False),
    ("Capital Account",               "section_sub", False),
    ("Reserves & Surplus",            "input",    True),
    ("Partners Capital",              "input",    True),
    ("Capital Account Total",         "formula",  False),
    ("",                              "blank",    False),
    ("Loans (Liability)",             "section_sub", False),
    ("Partner's Loan - Akshatha",     "input",    True),
    ("Partner's Loan - Tejas",        "input",    True),
    ("TVS Credit Services Ltd",       "input",    True),
    ("Other Loans",                   "input",    True),
    ("Loans Total",                   "formula",  False),
    ("",                              "blank",    False),
    ("Current Liabilities",           "section_sub", False),
    ("Duties & Taxes",                "input",    True),
    ("Provisions",                    "input",    True),
    ("Sundry Creditors",              "input",    True),
    ("Reimbursements Payable",        "input",    True),
    ("Salaries Payable",              "input",    True),
    ("Other Current Liabilities",     "input",    True),
    ("Current Liabilities Total",     "formula",  False),
    ("",                              "blank",    False),
    ("Profit & Loss A/c",             "section_sub", False),
    ("Current Period P&L",            "input",    True),
    ("P&L Total",                     "formula",  False),
    ("",                              "blank",    False),
    ("TOTAL LIABILITIES",             "formula",  False),
    ("",                              "blank",    False),
    ("",                              "blank",    False),
    ("ASSETS",                        "section",  False),
    ("Fixed Assets",                  "section_sub", False),
    ("Intangible Assets",             "input",    True),
    ("Tangible Assets",               "input",    True),
    ("Fixed Assets Total",            "formula",  False),
    ("",                              "blank",    False),
    ("Current Assets",                "section_sub", False),
    ("Closing Stock",                 "input",    True),
    ("Deposits (Asset)",              "input",    True),
    ("Loans & Advances (Asset)",      "input",    True),
    ("Sundry Debtors",                "input",    True),
    ("Cash-in-Hand",                  "input",    True),
    ("Bank Accounts",                 "input",    True),
    ("Deferred Expenses",             "input",    True),
    ("TDS Receivable",                "input",    True),
    ("Other Current Assets",          "input",    True),
    ("Current Assets Total",          "formula",  False),
    ("",                              "blank",    False),
    ("TOTAL ASSETS",                  "formula",  False),
    ("",                              "blank",    False),
    ("Difference (Assets − Liabilities)", "formula", False),
]

BS_DATA_START = 4
BS_QUARTERS = ["Q1 (Apr-Jun)", "Q2 (Jul-Sep)", "Q3 (Oct-Dec)", "Q4 (Jan-Mar)"]
BS_QTR_COLS = {1: 2, 2: 3, 3: 4, 4: 5}  # quarter num → column
BS_MAX_COL = 5  # A=Particulars, B=Q1, C=Q2, D=Q3, E=Q4


def _bs_row(label: str) -> int:
    for i, (lbl, *_) in enumerate(BS_ACCOUNTS):
        if lbl == label:
            return BS_DATA_START + i
    raise KeyError(f"BS label not found: {label}")


def _bs_formulas():
    r = _bs_row
    # Pre-compute rows for names with apostrophes
    loan_a_r = _bs_row("Partner's Loan - Akshatha")
    loan_t_r = _bs_row("Partner's Loan - Tejas")
    tvs_r = _bs_row("TVS Credit Services Ltd")
    other_l_r = _bs_row("Other Loans")
    return {
        "Capital Account Total": lambda c: f"=SUM({c}{r('Reserves & Surplus')}:{c}{r('Partners Capital')})",
        "Loans Total": lambda c: f"={c}{loan_a_r}+{c}{loan_t_r}+{c}{tvs_r}+{c}{other_l_r}",
        "Current Liabilities Total": lambda c: f"=SUM({c}{r('Duties & Taxes')}:{c}{r('Other Current Liabilities')})",
        "P&L Total": lambda c: f"={c}{r('Current Period P&L')}",
        "TOTAL LIABILITIES": lambda c: f"={c}{r('Capital Account Total')}+{c}{r('Loans Total')}+{c}{r('Current Liabilities Total')}+{c}{r('P&L Total')}",
        "Fixed Assets Total": lambda c: f"=SUM({c}{r('Intangible Assets')}:{c}{r('Tangible Assets')})",
        "Current Assets Total": lambda c: f"=SUM({c}{r('Closing Stock')}:{c}{r('Other Current Assets')})",
        "TOTAL ASSETS": lambda c: f"={c}{r('Fixed Assets Total')}+{c}{r('Current Assets Total')}",
        "Difference (Assets − Liabilities)": lambda c: f"={c}{r('TOTAL ASSETS')}-{c}{r('TOTAL LIABILITIES')}",
    }


def build_bs_input(wb):
    ws = wb.create_sheet("BS_Input")
    max_col = BS_MAX_COL

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP — Balance Sheet Input").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value="Paste quarterly Balance Sheet numbers below").font = Font(
        name=FONT_NAME, italic=True, size=10, color="808080")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["Particulars"] + BS_QUARTERS
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, max_col)

    _set_col_widths(ws, {1: 35, 2: 18, 3: 18, 4: 18, 5: 18})

    formulas = _bs_formulas()

    for i, (label, typ, indent) in enumerate(BS_ACCOUNTS):
        row = BS_DATA_START + i
        ws.cell(row=row, column=1, value=label)

        if typ == "section":
            _style_section_row(ws, row, max_col)
        elif typ == "section_sub":
            ws.cell(row=row, column=1).font = BOLD_FONT
            ws.cell(row=row, column=1).fill = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid")
            for c in range(2, max_col + 1):
                ws.cell(row=row, column=c).fill = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid")
                ws.cell(row=row, column=c).border = THIN_BORDER
        elif typ == "input":
            _style_data_row(ws, row, max_col, indent=indent)
            _input_highlight(ws, row, 2, max_col)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif typ == "formula":
            if label in formulas:
                for c in range(2, max_col + 1):
                    ws.cell(row=row, column=c, value=formulas[label](_col(c)))
            if label in ("TOTAL LIABILITIES", "TOTAL ASSETS"):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
                    cell.fill = HEADER_FILL
                    cell.border = DOUBLE_BOTTOM
            elif label == "Difference (Assets − Liabilities)":
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.font = Font(name=FONT_NAME, bold=True, size=11, color="C00000")
                    cell.border = DOUBLE_BOTTOM
            else:
                _style_total_row(ws, row, max_col)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)

    ws.freeze_panes = "B4"

    # ── Explanation block ──
    last_data_row = BS_DATA_START + len(BS_ACCOUNTS) - 1
    _add_explanation_block(ws, last_data_row, max_col,
        "📋  How to Use This Sheet (Balance Sheet Input)",
        [
            "The Balance Sheet is a snapshot of what the company OWNS (Assets) vs what it OWES (Liabilities) at a point in time.",
            "Yellow cells = input cells. Paste quarterly Balance Sheet trial balance numbers here. Totals auto-calculate.",
            "Liabilities side: Capital (owner's investment), Loans (borrowed money), Current Liabilities (short-term dues like creditors, salaries payable).",
            "Assets side: Fixed Assets (property, equipment), Current Assets (cash, bank, stock, debtors, deposits).",
            "Golden rule: Total Assets MUST equal Total Liabilities. The 'Difference' row at the bottom should be ₹0 if data is correct.",
            "If Difference ≠ 0, some number is missing or wrong — go back to the source and check.",
        ]
    )

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet 6: Balance Sheet — Formulae from BS_Input
# ═══════════════════════════════════════════════════════════════════════════════

def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    max_col = BS_MAX_COL
    BSI = "BS_Input"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value="Balance Sheet  |  FY 2025-26").font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["Particulars"] + BS_QUARTERS
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, max_col)
    _set_col_widths(ws, {1: 35, 2: 18, 3: 18, 4: 18, 5: 18})

    # Mirror BS_ACCOUNTS structure but link to BS_Input
    data_start = 4
    for i, (label, typ, indent) in enumerate(BS_ACCOUNTS):
        row = data_start + i
        bs_r = BS_DATA_START + i
        ws.cell(row=row, column=1, value=label)

        if typ == "blank":
            continue
        if typ == "section":
            _style_section_row(ws, row, max_col)
            continue
        if typ == "section_sub":
            ws.cell(row=row, column=1).font = BOLD_FONT
            ws.cell(row=row, column=1).fill = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid")
            for c in range(2, max_col + 1):
                ws.cell(row=row, column=c).fill = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid")
                ws.cell(row=row, column=c).border = THIN_BORDER
            continue

        # Link every data/formula cell to BS_Input
        for c in range(2, max_col + 1):
            ws.cell(row=row, column=c, value=f"={BSI}!{_col(c)}{bs_r}")

        if typ == "input":
            _style_data_row(ws, row, max_col, indent=indent)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)
        elif typ == "formula":
            if label in ("TOTAL LIABILITIES", "TOTAL ASSETS"):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
                    cell.fill = HEADER_FILL
                    cell.border = DOUBLE_BOTTOM
            elif label == "Difference (Assets − Liabilities)":
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.font = Font(name=FONT_NAME, bold=True, size=11, color="C00000")
                    cell.border = DOUBLE_BOTTOM
            else:
                _style_total_row(ws, row, max_col)
            _apply_number_format(ws, row, 2, max_col, INR_FMT)

    ws.freeze_panes = "B4"

    # ── Explanation block ──
    last_data_row = data_start + len(BS_ACCOUNTS) - 1
    _add_explanation_block(ws, last_data_row, max_col,
        "🏦  Understanding the Balance Sheet",
        [
            "Think of the Balance Sheet as a 'net worth statement' for the company — what it owns minus what it owes.",
            "LIABILITIES (what the company owes): Capital Account = money invested by partners. Loans = borrowed funds. Current Liabilities = short-term debts (creditors, unpaid salaries, taxes due).",
            "ASSETS (what the company owns): Fixed Assets = long-term things like equipment, furniture, IP. Current Assets = things that can be converted to cash quickly — stock, bank balance, deposits, receivables.",
            "Total Assets must equal Total Liabilities (the accounting equation). If they don't match, there's a data entry error.",
            "Current Assets vs Current Liabilities: If Current Assets > Current Liabilities, the company can comfortably pay its short-term bills (good sign).",
            "Cash + Bank balance: The most liquid assets. If this number is growing quarter-over-quarter, the business is accumulating cash.",
            "P&L Account links here: The profit (or loss) from the P&L statement flows into this Balance Sheet as retained earnings.",
            "All data auto-pulls from BS_Input — this is the formatted view for presentation.",
        ]
    )

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet 7: Performance Summary — Quarterly from P&L
# ═══════════════════════════════════════════════════════════════════════════════

def build_performance_summary(wb):
    ws = wb.create_sheet("Performance Summary")
    TB = "TB_Input"
    max_col = 7  # A=Particulars, B=Q1, C=Q2, D=Q3, E=Q4, F=FY Total, G=FY%

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value="Performance Summary  |  FY 2025-26").font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["Particulars", "Q1 (Apr-Jun)", "Q2 (Jul-Sep)", "Q3 (Oct-Dec)", "Q4 (Jan-Mar)", "FY Total", "FY %"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, max_col)
    _set_col_widths(ws, {1: 35, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16, 7: 10})

    # Quarterly sum from TB_Input: Q1=SUM(TB_Input B:D), Q2=SUM(TB_Input E:G), etc.
    PERF_LINES = [
        ("Revenue",              "Total Revenue",               "subtotal"),
        ("Morning Glory",        "Morning Glory",               "item"),
        ("Al Fresco",            "Al Fresco",                   "item"),
        ("Sunset Soiree",        "Sunset Soiree",               "item"),
        ("Studio",               "Studio",                      "item"),
        ("Other Income",         "Other Income",                "item"),
        ("",                     None,                          "blank"),
        ("COGS",                 "Total COGS",                  "subtotal"),
        ("",                     None,                          "blank"),
        ("Gross Profit",         "Gross Profit",                "total"),
        ("GP %",                 "Gross Profit",                "pct_gp"),
        ("",                     None,                          "blank"),
        ("Total OPEX",           "Total Indirect Expenses",     "subtotal"),
        ("",                     None,                          "blank"),
        ("EBITDA",               "EBITDA",                      "total"),
        ("EBITDA %",             "EBITDA",                      "pct_ebitda"),
        ("",                     None,                          "blank"),
        ("Depreciation",         "Depreciation",                "item"),
        ("PBT",                  "Profit Before Tax (PBT)",     "total"),
        ("PBT %",                "Profit Before Tax (PBT)",     "pct_pbt"),
        ("",                     None,                          "blank"),
        ("PAT",                  "Profit After Tax (PAT)",      "total"),
        ("PAT %",                "Profit After Tax (PAT)",      "pct_pat"),
    ]

    # Quarter col mappings to TB_Input columns
    # Q1 = B+C+D in TB (cols 2,3,4) → TB cols B,C,D
    # Q2 = E+F+G in TB (cols 5,6,7) → TB cols E,F,G
    # etc.
    qtrs = [
        ("B", "D", 2),   # Q1: TB B:D → Perf col 2
        ("E", "G", 3),   # Q2: TB E:G → Perf col 3
        ("H", "J", 4),   # Q3: TB H:J → Perf col 4
        ("K", "M", 5),   # Q4: TB K:M → Perf col 5
    ]

    rev_tb_row = _tb_row("Total Revenue")

    data_start = 4
    for i, (label, tb_label, stype) in enumerate(PERF_LINES):
        row = data_start + i
        ws.cell(row=row, column=1, value=label)

        if stype == "blank":
            continue
        if stype in ("item", "subtotal", "total"):
            tb_r = _tb_row(tb_label)
            for (sc, ec, pc) in qtrs:
                ws.cell(row=row, column=pc, value=f"=SUM({TB}!{sc}{tb_r}:{ec}{tb_r})")
            # FY = sum of quarters
            ws.cell(row=row, column=6, value=f"=SUM(B{row}:E{row})")
            # FY% = this / revenue
            if stype == "item":
                _style_data_row(ws, row, max_col, indent=True)
            elif stype == "subtotal":
                _style_total_row(ws, row, max_col)
            elif stype == "total":
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.font = Font(name=FONT_NAME, bold=True, size=12, color="2F5496")
                    cell.border = DOUBLE_BOTTOM
            _apply_number_format(ws, row, 2, 6, INR_FMT)

        elif stype.startswith("pct_"):
            # Calculate percentage for each quarter
            tb_r = _tb_row(tb_label)
            for (sc, ec, pc) in qtrs:
                ws.cell(row=row, column=pc,
                        value=f"=IFERROR(SUM({TB}!{sc}{tb_r}:{ec}{tb_r})/SUM({TB}!{sc}{rev_tb_row}:{ec}{rev_tb_row}),0)")
            # FY %
            ws.cell(row=row, column=6, value=f"=IFERROR({TB}!N{tb_r}/{TB}!N{rev_tb_row},0)")
            _style_data_row(ws, row, max_col, pct=True)
            _apply_number_format(ws, row, 2, 7, PCT_FMT)

    ws.freeze_panes = "B4"

    # ── Explanation block ──
    last_data_row = data_start + len(PERF_LINES) - 1
    _add_explanation_block(ws, last_data_row, max_col,
        "📈  How to Read the Performance Summary",
        [
            "This sheet aggregates monthly P&L data into quarters (Q1-Q4) so you can spot trends at a glance.",
            "Revenue trend: Is quarterly revenue growing, flat, or declining? Consistent growth = healthy demand.",
            "Gross Profit %: Should stay stable (60-70%). If it drops quarter-over-quarter, COGS is rising faster than revenue.",
            "OPEX trend: Rising OPEX is normal during growth, but OPEX as a % of Revenue should stay flat or decline over time (operating leverage).",
            "EBITDA %: The most watched metric. Improving EBITDA % means the business is becoming more efficient.",
            "PAT: The ultimate measure. Compare Q-over-Q to see if profitability is sustainable.",
            "FY % column: Shows the full-year ratio — useful for comparing against industry benchmarks.",
            "Look for seasonality: Hospitality businesses often have strong Q2/Q3 (wedding/event season) and softer Q1/Q4.",
        ]
    )

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet 8: KPIs — Key Financial Ratios
# ═══════════════════════════════════════════════════════════════════════════════

def build_kpis_sheet(wb):
    ws = wb.create_sheet("KPIs")
    TB = "TB_Input"
    BSI = "BS_Input"
    max_col = 7  # same layout as Performance

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(row=1, column=1, value="THOTA HOSPITALITY LLP").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(row=2, column=1, value="Key Financial Ratios  |  FY 2025-26").font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["KPI / Ratio", "Q1", "Q2", "Q3", "Q4", "FY", "Benchmark"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, max_col)
    _set_col_widths(ws, {1: 35, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14, 7: 20})

    # TB row references
    rev_r    = _tb_row("Total Revenue")
    gp_r     = _tb_row("Gross Profit")
    ebitda_r = _tb_row("EBITDA")
    pat_r    = _tb_row("Profit After Tax (PAT)")
    pbt_r    = _tb_row("Profit Before Tax (PBT)")
    opex_r   = _tb_row("Total Indirect Expenses")
    cogs_r   = _tb_row("Total COGS")
    hr_r     = _tb_row("HR Expenses Total")
    dep_r    = _tb_row("Depreciation")

    # BS row references
    total_liab_r = _bs_row("TOTAL LIABILITIES")
    total_asset_r = _bs_row("TOTAL ASSETS")
    curr_asset_r = _bs_row("Current Assets Total")
    curr_liab_r = _bs_row("Current Liabilities Total")
    cash_r = _bs_row("Cash-in-Hand")
    bank_r = _bs_row("Bank Accounts")
    capital_r = _bs_row("Capital Account Total")

    qtrs = [
        ("B", "D", 2),
        ("E", "G", 3),
        ("H", "J", 4),
        ("K", "M", 5),
    ]

    data_start = 4

    # Profitability section
    ws.cell(row=data_start, column=1, value="PROFITABILITY RATIOS")
    _style_section_row(ws, data_start, max_col)

    kpi_rows = [
        (data_start + 1, "Gross Profit Margin",
         lambda sc, ec: f"=IFERROR(SUM({TB}!{sc}{gp_r}:{ec}{gp_r})/SUM({TB}!{sc}{rev_r}:{ec}{rev_r}),0)",
         f"=IFERROR({TB}!N{gp_r}/{TB}!N{rev_r},0)", "60-70%", PCT_FMT),
        (data_start + 2, "EBITDA Margin",
         lambda sc, ec: f"=IFERROR(SUM({TB}!{sc}{ebitda_r}:{ec}{ebitda_r})/SUM({TB}!{sc}{rev_r}:{ec}{rev_r}),0)",
         f"=IFERROR({TB}!N{ebitda_r}/{TB}!N{rev_r},0)", "15-25%", PCT_FMT),
        (data_start + 3, "PAT Margin",
         lambda sc, ec: f"=IFERROR(SUM({TB}!{sc}{pat_r}:{ec}{pat_r})/SUM({TB}!{sc}{rev_r}:{ec}{rev_r}),0)",
         f"=IFERROR({TB}!N{pat_r}/{TB}!N{rev_r},0)", "10-15%", PCT_FMT),
        (data_start + 4, "COGS / Revenue",
         lambda sc, ec: f"=IFERROR(SUM({TB}!{sc}{cogs_r}:{ec}{cogs_r})/SUM({TB}!{sc}{rev_r}:{ec}{rev_r}),0)",
         f"=IFERROR({TB}!N{cogs_r}/{TB}!N{rev_r},0)", "25-35%", PCT_FMT),
        (data_start + 5, "OPEX / Revenue",
         lambda sc, ec: f"=IFERROR(SUM({TB}!{sc}{opex_r}:{ec}{opex_r})/SUM({TB}!{sc}{rev_r}:{ec}{rev_r}),0)",
         f"=IFERROR({TB}!N{opex_r}/{TB}!N{rev_r},0)", "<40%", PCT_FMT),
        (data_start + 6, "HR Cost / Revenue",
         lambda sc, ec: f"=IFERROR(SUM({TB}!{sc}{hr_r}:{ec}{hr_r})/SUM({TB}!{sc}{rev_r}:{ec}{rev_r}),0)",
         f"=IFERROR({TB}!N{hr_r}/{TB}!N{rev_r},0)", "25-35%", PCT_FMT),
    ]

    for row, label, qtr_formula, fy_formula, benchmark, fmt in kpi_rows:
        ws.cell(row=row, column=1, value=label)
        for (sc, ec, pc) in qtrs:
            ws.cell(row=row, column=pc, value=qtr_formula(sc, ec))
        ws.cell(row=row, column=6, value=fy_formula)
        ws.cell(row=row, column=7, value=benchmark)
        _style_data_row(ws, row, max_col, indent=True)
        _apply_number_format(ws, row, 2, 6, fmt)
        ws.cell(row=row, column=7).font = Font(name=FONT_NAME, italic=True, size=10, color="808080")

    # Operational section
    op_start = data_start + 8
    ws.cell(row=op_start, column=1, value="OPERATIONAL METRICS")
    _style_section_row(ws, op_start, max_col)

    # Monthly Revenue (avg per quarter)
    r = op_start + 1
    ws.cell(row=r, column=1, value="Avg Monthly Revenue")
    for (sc, ec, pc) in qtrs:
        ws.cell(row=r, column=pc, value=f"=IFERROR(SUM({TB}!{sc}{rev_r}:{ec}{rev_r})/3,0)")
    ws.cell(row=r, column=6, value=f"=IFERROR({TB}!N{rev_r}/12,0)")
    ws.cell(row=r, column=7, value="Growth trend")
    _style_data_row(ws, r, max_col, indent=True)
    _apply_number_format(ws, r, 2, 6, INR_FMT)

    # Monthly Breakeven
    r = op_start + 2
    ws.cell(row=r, column=1, value="Monthly Breakeven Revenue")
    for (sc, ec, pc) in qtrs:
        ws.cell(row=r, column=pc,
                value=f"=IFERROR(SUM({TB}!{sc}{opex_r}:{ec}{opex_r})/3/(1-SUM({TB}!{sc}{cogs_r}:{ec}{cogs_r})/SUM({TB}!{sc}{rev_r}:{ec}{rev_r})),0)")
    ws.cell(row=r, column=7, value="Rev needed to cover OPEX")
    _style_data_row(ws, r, max_col, indent=True)
    _apply_number_format(ws, r, 2, 6, INR_FMT)

    # Liquidity section
    liq_start = op_start + 4
    ws.cell(row=liq_start, column=1, value="LIQUIDITY RATIOS (from Balance Sheet)")
    _style_section_row(ws, liq_start, max_col)

    r = liq_start + 1
    ws.cell(row=r, column=1, value="Current Ratio")
    for q in range(1, 5):
        qc = _col(q + 1)
        ws.cell(row=r, column=q + 1,
                value=f"=IFERROR({BSI}!{qc}{curr_asset_r}/{BSI}!{qc}{curr_liab_r},0)")
    ws.cell(row=r, column=7, value="> 1.5")
    _style_data_row(ws, r, max_col, indent=True)
    _apply_number_format(ws, r, 2, 6, "0.00")

    r = liq_start + 2
    ws.cell(row=r, column=1, value="Cash Position (Cash + Bank)")
    for q in range(1, 5):
        qc = _col(q + 1)
        ws.cell(row=r, column=q + 1,
                value=f"={BSI}!{qc}{cash_r}+{BSI}!{qc}{bank_r}")
    ws.cell(row=r, column=7, value="Positive & growing")
    _style_data_row(ws, r, max_col, indent=True)
    _apply_number_format(ws, r, 2, 6, INR_FMT)

    r = liq_start + 3
    ws.cell(row=r, column=1, value="Cash Runway (months)")
    for q in range(1, 5):
        qc = _col(q + 1)
        sc, ec, pc = qtrs[q - 1]
        ws.cell(row=r, column=q + 1,
                value=f"=IFERROR(({BSI}!{qc}{cash_r}+{BSI}!{qc}{bank_r})/(SUM({TB}!{sc}{opex_r}:{ec}{opex_r})/3),0)")
    ws.cell(row=r, column=7, value="> 3 months")
    _style_data_row(ws, r, max_col, indent=True)
    _apply_number_format(ws, r, 2, 6, "0.0")

    ws.freeze_panes = "B4"

    # ── Explanation block: Ratio explanations ──
    last_kpi_row = liq_start + 3
    explain_row = _add_explanation_block(ws, last_kpi_row, max_col,
        "📐  What Each Ratio Means (Plain English)",
        [
            "GROSS PROFIT MARGIN (Benchmark: 60-70%) — For every ₹100 of revenue, how much is left after paying for raw materials and direct delivery costs. Example: 65% means ₹65 stays with the company after COGS. If this drops, either prices are too low or input costs are rising.",
            "EBITDA MARGIN (Benchmark: 15-25%) — The operating profitability before accounting adjustments. If EBITDA is 20%, the business earns ₹20 from operations for every ₹100 of revenue. This is the key metric investors look at for a company's health.",
            "PAT MARGIN (Benchmark: 10-15%) — The final profit percentage after ALL expenses including depreciation and taxes. A 12% PAT means ₹12 net profit per ₹100 revenue. This is what actually grows the owners' wealth.",
            "COGS / REVENUE (Benchmark: 25-35%) — How much of every rupee goes to direct costs. Lower is better. If this exceeds 40%, the business needs to renegotiate supplier rates or raise prices.",
            "OPEX / REVENUE (Benchmark: <40%) — How much is spent on overheads per rupee of revenue. As the business grows, this should decline (called operating leverage). If it stays high, costs aren't scaling efficiently.",
            "HR COST / REVENUE (Benchmark: 25-35%) — Staff cost as a share of revenue. Hospitality is labor-intensive, so 25-35% is normal. Above 35% may mean overstaffing or under-revenue.",
            "AVG MONTHLY REVENUE (Trend: should grow) — Revenue divided by months. Compare Q-over-Q to see if the topline is growing. Flat or declining = demand issue.",
            "MONTHLY BREAKEVEN REVENUE — The minimum monthly revenue needed to cover all OPEX. If actual revenue is above this number, the company is profitable. Below = losses.",
            "CURRENT RATIO (Benchmark: >1.5) — Current Assets ÷ Current Liabilities. Measures if the company can pay its short-term bills. >1.5 = comfortable. <1.0 = the company may struggle to pay creditors on time.",
            "CASH POSITION (Benchmark: positive & growing) — Total cash + bank balance. This is the company's liquidity cushion. Growing cash = healthy. Declining cash = potential trouble even if P&L shows profit.",
            "CASH RUNWAY (Benchmark: >3 months) — How many months the company can survive on its current cash if revenue stopped. Cash ÷ Monthly OPEX. <3 months = urgent need for funding or cost cuts.",
        ]
    )

    # ── Company position summary ──
    _add_explanation_block(ws, explain_row, max_col,
        "🔍  How to Assess the Company's Financial Position",
        [
            "REVENUE CHECK: Is revenue growing quarter-over-quarter? Compare Q1 → Q2 → Q3 → Q4. Growth = demand is healthy.",
            "COST CHECK: Are COGS% and OPEX% stable or declining? If both are flat while revenue grows, the company is scaling well.",
            "PROFIT CHECK: Is EBITDA positive and improving? Positive EBITDA = the business model works. Improving EBITDA % = efficiency is rising.",
            "CASH CHECK: Is cash position growing? A profitable company should accumulate cash. If profit is positive but cash is dropping, look for payment collection delays.",
            "RISK CHECK: Is Cash Runway > 3 months and Current Ratio > 1.5? If yes, the company has a safety buffer. If no, short-term survival is at risk.",
            "OVERALL: Green flags = rising revenue + stable margins + growing cash. Red flags = flat revenue + shrinking margins + cash burn.",
        ]
    )

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  Pre-fill TB_Input with actuals from source Excel
# ═══════════════════════════════════════════════════════════════════════════════

def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, str):
        val = val.strip().replace(",", "").replace("₹", "").replace("\xa0", "")
        if val in ("", "-", "#REF!"):
            return default
        try:
            return float(val)
        except ValueError:
            return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def prefill_from_source(wb_master, source_path: str):
    """Read source Excel and fill TB_Input + BS_Input with actuals."""
    src = openpyxl.load_workbook(source_path, data_only=True)
    ws_tb = wb_master["TB_Input"]

    # ── P&L Detailed ──
    if "P&L Detailed" in src.sheetnames:
        ws_pl = src["P&L Detailed"]
        # Row 3 = headers (APR-25 … ), Row 4+ = data
        # Map source row labels → TB_Input labels
        pl_map = {
            "Morning Glory":              "Morning Glory",
            "Al Fresco":                  "Al Fresco",
            "Sunset Soiree":              "Sunset Soiree",
            "Studio":                     "Studio",
            "Other Income":               "Other Income",
            "Purchase Cost":              None,  # skip, use sub-items
            "Thota Kitchen":              "Purchase - Thota Kitchen",
            "Thota Decor":                "Purchase - Thota Decor",
            "Studio Decor":               "Purchase - Studio Decor",
            "Fuel Expenses":              "Fuel Expenses",
            "Transportation Expense":     "Transportation Expense",
            "Depreciation":               "Depreciation",
        }
        for r in range(4, min(60, ws_pl.max_row + 1)):
            src_label = str(ws_pl.cell(row=r, column=1).value or "").strip()
            if src_label in pl_map and pl_map[src_label]:
                tb_label = pl_map[src_label]
                try:
                    tb_r = _tb_row(tb_label)
                except KeyError:
                    continue
                for c in range(2, min(14, ws_pl.max_column + 1)):
                    val = safe_float(ws_pl.cell(row=r, column=c).value)
                    if val != 0:
                        ws_tb.cell(row=tb_r, column=c, value=val)

    # ── Schedule_OPEX ──
    if "Schedule_OPEX" in src.sheetnames:
        ws_opex = src["Schedule_OPEX"]
        opex_map = {
            "Local Conveyance":             "Local Conveyance",
            "Office Expenses":              "Office Expenses",
            "Printing & Stationery":        "Printing & Stationery",
            "Registration & Subscription":  "Registration & Subscription",
            "Staff Welfare":                "Staff Welfare",
            "Telephone & Internet Charges": "Telephone & Internet Charges",
            "Bank Charges":                 "Bank Charges",
            "Interest on Bank loan":        "Interest on Bank Loan",
            "Interest on Bank Loan":        "Interest on Bank Loan",
            "Employee Salaries":            "Employee Salaries",
            "Incentives":                   "Incentives",
            "Partner Remuneration":         "Partner Remuneration",
            "PF Admin Charges":             "PF Admin Charges",
            "Stipend":                      "Stipend",
            "Meta / Instagram":             "Meta / Instagram",
            "Events / Exhibitions":         "Events / Exhibitions",
            "Professional Service Charges": "Professional Service Charges",
            "Rates & Taxes":                "Rates & Taxes",
            "Repairs & Maintenance":        "Repairs & Maintenance",
            "Misc. Expenses":               "Misc. Expenses",
        }
        for r in range(2, min(35, ws_opex.max_row + 1)):
            src_label = str(ws_opex.cell(row=r, column=1).value or "").strip()
            if src_label in opex_map:
                tb_label = opex_map[src_label]
                try:
                    tb_r = _tb_row(tb_label)
                except KeyError:
                    continue
                for c in range(2, min(14, ws_opex.max_column + 1)):
                    val = safe_float(ws_opex.cell(row=r, column=c).value)
                    if val != 0:
                        ws_tb.cell(row=tb_r, column=c, value=val)

    # ── Balance Sheet ──
    if "BalanceSheet Summary" in src.sheetnames:
        ws_bs_src = src["BalanceSheet Summary"]
        ws_bsi = wb_master["BS_Input"]
        bs_map = {
            "Reserves & Surplus":       "Reserves & Surplus",
            "Partners Capital":         "Partners Capital",
            "Partner's Loan-Akshatha":  "Partner's Loan - Akshatha",
            "Partner's Loan-Tejas":     "Partner's Loan - Tejas",
            "TVS Credit Services Ltd":  "TVS Credit Services Ltd",
            "Duties & Taxes":           "Duties & Taxes",
            "Provisions":               "Provisions",
            "Sundry Creditors":         "Sundry Creditors",
            "Reimbursments Payable":    "Reimbursements Payable",
            "Salaries Payable":         "Salaries Payable",
            "Current Period (As per th Books)": "Current Period P&L",
            "Intangible Assets":        "Intangible Assets",
            "Tangible Assets":          "Tangible Assets",
            "Closing Stock":            "Closing Stock",
            "Deposits (Asset)":         "Deposits (Asset)",
            "Loans & Advances (Asset)": "Loans & Advances (Asset)",
            "Sundry Debtors":           "Sundry Debtors",
            "Cash-in-Hand":             "Cash-in-Hand",
            "Bank Accounts":            "Bank Accounts",
            "Deferred Expenses":        "Deferred Expenses",
            "TDS - Receivable":         "TDS Receivable",
        }
        # BS has data in col C (Particulars), col D (Subtotal), col E (Total)
        # Quarterly data might be in cols F, G, H (Q1, Q2, Q3)
        for r in range(4, ws_bs_src.max_row + 1):
            src_label = str(ws_bs_src.cell(row=r, column=3).value or "").strip()
            if src_label in bs_map:
                bs_label = bs_map[src_label]
                try:
                    bs_r = _bs_row(bs_label)
                except KeyError:
                    continue
                # Try sub-total column (D=4) for overall, and Q columns (F=6, G=7, H=8)
                sub_val = safe_float(ws_bs_src.cell(row=r, column=4).value)
                # If quarterly data available
                for qi, qcol in enumerate([6, 7, 8], 1):  # Q1=col6, Q2=col7, Q3=col8
                    val = safe_float(ws_bs_src.cell(row=r, column=qcol).value)
                    if val != 0:
                        ws_bsi.cell(row=bs_r, column=qi + 1, value=val)  # BS_Input Q1=col2, etc.
                # If no quarterly data, use subtotal for Q3 (latest)
                if sub_val != 0 and all(safe_float(ws_bs_src.cell(row=r, column=qc).value) == 0 for qc in [6, 7, 8]):
                    ws_bsi.cell(row=bs_r, column=4, value=sub_val)  # Put in Q3

    # ── Cashflow Opening Balance ──
    if "Cashflow" in src.sheetnames:
        ws_cf_src = src["Cashflow"]
        ws_cf = wb_master["Cash Flow"]
        # Row 4 has "Bank/Cash Opening Balance", col 2+ has monthly values
        opening_val = safe_float(ws_cf_src.cell(row=4, column=2).value)
        if opening_val != 0:
            ws_cf.cell(row=4, column=2, value=opening_val)  # APR opening

    src.close()
    print(f"  ✓ Pre-filled TB_Input and BS_Input from source: {source_path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  Main: generate the master file
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Generate formula-driven Master Financial Model Excel for THOTA HOSPITALITY LLP."
    )
    parser.add_argument(
        "--prefill", default=None,
        help="Path to source Excel to pre-fill actuals (e.g., FW_P&L_2025-26 - V2.xlsx)",
    )
    parser.add_argument(
        "--output", default=None,
        help="Custom output path (default: .tmp/Thota_Master_Financial_Model.xlsx)",
    )
    args = parser.parse_args()

    output_path = args.output or str(TMP_DIR / "Thota_Master_Financial_Model.xlsx")

    print(f"╔{'═'*58}╗")
    print(f"║  Master Financial Model Generator                        ║")
    print(f"╚{'═'*58}╝")
    print(f"\nOutput: {output_path}")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("  Building TB_Input …")
    build_tb_input(wb)
    print("  Building P&L …")
    build_pl_sheet(wb)
    print("  Building Cash Flow …")
    build_cashflow_sheet(wb)
    print("  Building OPEX Schedule …")
    build_opex_sheet(wb)
    print("  Building BS_Input …")
    build_bs_input(wb)
    print("  Building Balance Sheet …")
    build_balance_sheet(wb)
    print("  Building Performance Summary …")
    build_performance_summary(wb)
    print("  Building KPIs …")
    build_kpis_sheet(wb)

    if args.prefill and os.path.isfile(args.prefill):
        print(f"\n  Pre-filling with actuals from source …")
        prefill_from_source(wb, args.prefill)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()

    print(f"\n{'─'*60}")
    print(f"  Master file saved: {output_path}")
    print(f"{'─'*60}")
    print(f"\n  How to use:")
    print(f"  1. Open TB_Input sheet")
    print(f"  2. Paste Trial Balance numbers into yellow cells (month by month)")
    print(f"  3. All other sheets auto-update via formulas")
    print(f"  4. For Balance Sheet, paste quarterly numbers in BS_Input")
    print(f"  5. Cash Flow Opening Balance: enter APR value; rest auto-links")
    print()

    return output_path


if __name__ == "__main__":
    main()
