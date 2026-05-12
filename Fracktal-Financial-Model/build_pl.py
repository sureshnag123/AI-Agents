import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.load_workbook('C:/Users/Lenovo/Downloads/Fracktal_Financials_FY2526.xlsx')
if 'P&L FY25-26 Provisional' in wb.sheetnames:
    del wb['P&L FY25-26 Provisional']
ws = wb.create_sheet('P&L FY25-26 Provisional', 0)

DARK_BLUE   = '1F3864'
MID_BLUE    = '2E5F9A'
LIGHT_BLUE  = 'D9E1F2'
HEADER_BLUE = 'BDD7EE'
GREEN_BG    = 'E2EFDA'
RED_BG      = 'FCE4D6'
WHITE       = 'FFFFFF'
GRAY_BG     = 'F2F2F2'
YELLOW_BG   = 'FFF2CC'

def S(s='thin'): return Side(style=s)
def med_top(): return Border(top=S('medium'), bottom=S('double'))
fmt_inr = '_(\u20b9* #,##0_);_(\u20b9* (#,##0);_(\u20b9* "-"_);_(@_)'
fmt_pct = '0.0%'

for col, w in [('A',3),('B',36),('C',18),('D',18),('E',18),('F',18)]:
    ws.column_dimensions[col].width = w

# ── Title ──────────────────────────────────────────────────────────────────
for r, txt, sz in [(1,'Fracktal Works Private Limited',14),
                   (2,'Provisional Profit & Loss Statement  |  FY 2025-26',12),
                   (3,'CIN: U30009KA2013PTC070124  |  MSME: UDYAM-KR-03-0093853  |  (Amounts in INR)',9)]:
    ws.merge_cells(f'B{r}:F{r}')
    c = ws[f'B{r}']
    c.value = txt
    c.font = Font(name='Calibri', bold=(r<3), size=sz, color=WHITE, italic=(r==3))
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 22 if r==1 else 18 if r==2 else 14
ws.row_dimensions[4].height = 6

# ── Column headers ─────────────────────────────────────────────────────────
for col, txt in [('B','Particulars'),('C','FY 2023-24\n(Audited)'),
                  ('D','FY 2024-25\n(Audited)'),('E','FY 2025-26\n(Actuals)'),
                  ('F','FY 2025-26\n(Provisional)')]:
    c = ws[f'{col}5']
    c.value = txt
    c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
    c.fill = PatternFill('solid', fgColor=MID_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = Border(left=S(),right=S(),top=S(),bottom=S())
ws.row_dimensions[5].height = 32

# ── Data rows ──────────────────────────────────────────────────────────────
# Row numbers for key items — used in formulas below
# fmt: (row, label, fy24, fy25, actuals, provisional, bold, bg, indent, is_pct)
# For pct rows: values are formula strings

rows = [
    # REVENUE
    (6,  'REVENUE FROM OPERATIONS', None,None,None,None, True,LIGHT_BLUE,False,False),
    (7,  'Sale of Goods',    21543478,12499238,15963651,25187841, False,WHITE,True,False),
    (8,  'Sale of Services',  3110583, 1286955, 2055584, 2151205, False,WHITE,True,False),
    (9,  'Other Income',        47836, 1651824,        0,       0, False,WHITE,True,False),
    (10, 'TOTAL REVENUE',        None,    None,     None,    None, True,HEADER_BLUE,False,False),
    (11, '',None,None,None,None,False,WHITE,False,False),
    # COGS
    (12, 'COST OF REVENUE',    None,None,None,None, True,LIGHT_BLUE,False,False),
    (13, 'Cost of Goods Sold (COGS)', 8546242,6970774,5910668,9638592, False,WHITE,True,False),
    (14, 'Direct Expenses',    1394144,1463225,1600540,2311323, False,WHITE,True,False),
    (15, 'TOTAL COST OF REVENUE', None,None,None,None, True,HEADER_BLUE,False,False),
    (16, '',None,None,None,None,False,WHITE,False,False),
    (17, 'GROSS PROFIT',        None,None,None,None, True,GREEN_BG,False,False),
    (18, 'Gross Profit %',      None,None,None,None, False,GREEN_BG,False,True),
    (19, '',None,None,None,None,False,WHITE,False,False),
    # OPEX
    (20, 'OPERATING EXPENSES (OPEX)', None,None,None,None, True,LIGHT_BLUE,False,False),
    (21, 'Payroll & HR',       None,None,10022991,10022991, False,WHITE,True,False),
    (22, 'Advertisement & Marketing', None,None,1486550,1486550, False,WHITE,True,False),
    (23, 'Office & Admin Overheads',  None,None,2258166,2258166, False,WHITE,True,False),
    (24, 'Professional Services',     None,None,2667779,2667779, False,WHITE,True,False),
    (25, 'Travel Expenses',           None,None, 466069, 466069, False,WHITE,True,False),
    (26, 'R&D Expenses',              None,None, 428068, 428068, False,WHITE,True,False),
    (27, 'Freight & Courier',         None,None, 327469, 327469, False,WHITE,True,False),
    (28, 'Rates & Taxes',             None,None,  68049,  68049, False,WHITE,True,False),
    (29, 'Finance Cost (Interest)',   846520,436905, 372923, 437923, False,WHITE,True,False),
    (30, 'TOTAL OPEX',         None,None,None,None, True,HEADER_BLUE,False,False),
    (31, '',None,None,None,None,False,WHITE,False,False),
    (32, 'EBITDA',             None,None,None,None, True,LIGHT_BLUE,False,False),
    (33, 'EBITDA %',           None,None,None,None, False,LIGHT_BLUE,False,True),
    (34, '',None,None,None,None,False,WHITE,False,False),
    (35, 'Depreciation',       362205,281891, 256667, 291667, False,WHITE,True,False),
    (36, '',None,None,None,None,False,WHITE,False,False),
    (37, 'NET PROFIT / (LOSS) BEFORE TAX', None,None,None,None, True,RED_BG,False,False),
    (38, 'PAT %',              None,None,None,None, False,RED_BG,False,True),
    (39, 'Income Tax',         0,0,0,0, False,WHITE,True,False),
    (40, 'NET PROFIT / (LOSS) AFTER TAX',  None,None,None,None, True,RED_BG,False,False),
]

# Write input rows first (hardcoded values)
for (row, lbl, v_fy24, v_fy25, v_act, v_prov, bold, bg, indent, is_pct) in rows:
    ws.row_dimensions[row].height = 16 if lbl else 5
    c = ws[f'B{row}']
    c.value = ('    ' if indent else '') + lbl
    c.font = Font(name='Calibri', bold=bold, size=10)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(vertical='center')
    is_total = bold and lbl not in ('REVENUE FROM OPERATIONS','COST OF REVENUE','OPERATING EXPENSES (OPEX)')
    for col, val in [('C',v_fy24),('D',v_fy25),('E',v_act),('F',v_prov)]:
        cell = ws[f'{col}{row}']
        cell.value = val
        cell.font = Font(name='Calibri', bold=bold, size=10)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = fmt_pct if is_pct else fmt_inr
        if is_total and val is not None:
            cell.border = med_top()

# ── Now write FORMULA rows ─────────────────────────────────────────────────
# Row 10: TOTAL REVENUE = SUM of goods+services+other
for col in ['C','D','E','F']:
    ws[f'{col}10'].value = f'={col}7+{col}8+{col}9'
    ws[f'{col}10'].number_format = fmt_inr
    ws[f'{col}10'].border = med_top()
    ws[f'{col}10'].font = Font(name='Calibri', bold=True, size=10)
    ws[f'{col}10'].fill = PatternFill('solid', fgColor=HEADER_BLUE)
    ws[f'{col}10'].alignment = Alignment(horizontal='right', vertical='center')

# Row 15: TOTAL COST OF REVENUE
for col in ['C','D','E','F']:
    ws[f'{col}15'].value = f'={col}13+{col}14'
    ws[f'{col}15'].number_format = fmt_inr
    ws[f'{col}15'].border = med_top()
    ws[f'{col}15'].font = Font(name='Calibri', bold=True, size=10)
    ws[f'{col}15'].fill = PatternFill('solid', fgColor=HEADER_BLUE)
    ws[f'{col}15'].alignment = Alignment(horizontal='right', vertical='center')

# Row 17: GROSS PROFIT = Revenue - Cost
for col in ['C','D','E','F']:
    ws[f'{col}17'].value = f'={col}10-{col}15'
    ws[f'{col}17'].number_format = fmt_inr
    ws[f'{col}17'].border = med_top()
    ws[f'{col}17'].font = Font(name='Calibri', bold=True, size=10)
    ws[f'{col}17'].fill = PatternFill('solid', fgColor=GREEN_BG)
    ws[f'{col}17'].alignment = Alignment(horizontal='right', vertical='center')

# Row 18: Gross Profit %
for col in ['C','D','E','F']:
    ws[f'{col}18'].value = f'=IF({col}10=0,0,{col}17/{col}10)'
    ws[f'{col}18'].number_format = fmt_pct
    ws[f'{col}18'].font = Font(name='Calibri', size=10)
    ws[f'{col}18'].fill = PatternFill('solid', fgColor=GREEN_BG)
    ws[f'{col}18'].alignment = Alignment(horizontal='right', vertical='center')

# Row 30: TOTAL OPEX = SUM rows 21:29
for col in ['C','D','E','F']:
    ws[f'{col}30'].value = f'=SUM({col}21:{col}29)'
    ws[f'{col}30'].number_format = fmt_inr
    ws[f'{col}30'].border = med_top()
    ws[f'{col}30'].font = Font(name='Calibri', bold=True, size=10)
    ws[f'{col}30'].fill = PatternFill('solid', fgColor=HEADER_BLUE)
    ws[f'{col}30'].alignment = Alignment(horizontal='right', vertical='center')

# Row 32: EBITDA = Gross Profit - OPEX (finance cost already in opex row 29)
for col in ['C','D','E','F']:
    ws[f'{col}32'].value = f'={col}17-{col}30'
    ws[f'{col}32'].number_format = fmt_inr
    ws[f'{col}32'].border = med_top()
    ws[f'{col}32'].font = Font(name='Calibri', bold=True, size=10)
    ws[f'{col}32'].fill = PatternFill('solid', fgColor=LIGHT_BLUE)
    ws[f'{col}32'].alignment = Alignment(horizontal='right', vertical='center')

# Row 33: EBITDA %
for col in ['C','D','E','F']:
    ws[f'{col}33'].value = f'=IF({col}10=0,0,{col}32/{col}10)'
    ws[f'{col}33'].number_format = fmt_pct
    ws[f'{col}33'].font = Font(name='Calibri', size=10)
    ws[f'{col}33'].fill = PatternFill('solid', fgColor=LIGHT_BLUE)
    ws[f'{col}33'].alignment = Alignment(horizontal='right', vertical='center')

# Row 37: PBT = EBITDA - Depreciation
for col in ['C','D','E','F']:
    ws[f'{col}37'].value = f'={col}32-{col}35'
    ws[f'{col}37'].number_format = fmt_inr
    ws[f'{col}37'].border = med_top()
    ws[f'{col}37'].font = Font(name='Calibri', bold=True, size=10)
    ws[f'{col}37'].fill = PatternFill('solid', fgColor=RED_BG)
    ws[f'{col}37'].alignment = Alignment(horizontal='right', vertical='center')

# Row 38: PAT %
for col in ['C','D','E','F']:
    ws[f'{col}38'].value = f'=IF({col}10=0,0,{col}37/{col}10)'
    ws[f'{col}38'].number_format = fmt_pct
    ws[f'{col}38'].font = Font(name='Calibri', size=10)
    ws[f'{col}38'].fill = PatternFill('solid', fgColor=RED_BG)
    ws[f'{col}38'].alignment = Alignment(horizontal='right', vertical='center')

# Row 40: PAT = PBT - Tax
for col in ['C','D','E','F']:
    ws[f'{col}40'].value = f'={col}37-{col}39'
    ws[f'{col}40'].number_format = fmt_inr
    ws[f'{col}40'].border = med_top()
    ws[f'{col}40'].font = Font(name='Calibri', bold=True, size=10)
    ws[f'{col}40'].fill = PatternFill('solid', fgColor=RED_BG)
    ws[f'{col}40'].alignment = Alignment(horizontal='right', vertical='center')

# ── Note ───────────────────────────────────────────────────────────────────
ws.merge_cells('B42:F42')
c = ws['B42']
c.value = ('Note: FY25-26 Provisional includes INR 93.19L of confirmed Orders in Hand (PO received INR 80.2L + Awaiting PO INR 13.0L) '
           'expected to be billed by 31-Mar-26. All totals and ratios are formula-linked.')
c.font = Font(name='Calibri', italic=True, size=9, color='444444')
c.alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[42].height = 30

wb.save('C:/Users/Lenovo/Downloads/Fracktal_Financials_FY2526.xlsx')
print('P&L sheet rebuilt with formula linkage.')
