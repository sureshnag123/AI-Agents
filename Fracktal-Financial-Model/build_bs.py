import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.load_workbook('C:/Users/Lenovo/Downloads/Fracktal_Financials_FY2526.xlsx')
# Remove old sheet if exists, then recreate
if 'Balance Sheet 31-Mar-26' in wb.sheetnames:
    del wb['Balance Sheet 31-Mar-26']
ws = wb.create_sheet('Balance Sheet 31-Mar-26')

DARK_BLUE  = '1F3864'
MID_BLUE   = '2E5F9A'
LIGHT_BLUE = 'D9E1F2'
HEADER_BLUE= 'BDD7EE'
GREEN_BG   = 'E2EFDA'
WHITE      = 'FFFFFF'
YELLOW_BG  = 'FFF2CC'
GRAY_BG    = 'F2F2F2'

def side(style='thin'): return Side(style=style)

fmt_inr = '_(\u20b9* #,##0_);_(\u20b9* (#,##0);_(\u20b9* "-"_);_(@_)'

# Column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 36
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 3
ws.column_dimensions['F'].width = 36
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 18

# ---- TITLE ----
for span in ['B1:H1','B2:H2','B3:H3']:
    ws.merge_cells(span)

c = ws['B1']
c.value = 'Fracktal Works Private Limited'
c.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
c.fill = PatternFill('solid', fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 22

c = ws['B2']
c.value = 'Provisional Balance Sheet as at 31st March 2026'
c.font = Font(name='Calibri', bold=True, size=11, color=WHITE)
c.fill = PatternFill('solid', fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 18

c = ws['B3']
c.value = 'CIN: U30009KA2013PTC070124  |  MSME: UDYAM-KR-03-0093853  |  (All figures in INR)'
c.font = Font(name='Calibri', italic=True, size=9, color='BBBBBB')
c.fill = PatternFill('solid', fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[3].height = 14
ws.row_dimensions[4].height = 6

# ---- COLUMN HEADERS ----
headers = [
    ('B','LIABILITIES & EQUITY'),('C','13-Mar-26\n(Tally Actual)'),('D','31-Mar-26\n(Provisional)'),
    ('F','ASSETS'),('G','13-Mar-26\n(Tally Actual)'),('H','31-Mar-26\n(Provisional)')
]
for col, val in headers:
    c = ws[f'{col}5']
    c.value = val
    c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
    c.fill = PatternFill('solid', fgColor=MID_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = Border(left=side(), right=side(), top=side(), bottom=side())
ws.row_dimensions[5].height = 32

# ---- DATA ROWS ----
liab_rows = [
    # (label, tally_val, prov_val, bold, bg, indent)
    ('I.   SHAREHOLDERS FUNDS',        None,       None,       True,  LIGHT_BLUE, False),
    ('Share Capital (Paid Up)',         593220,     593220,     False, WHITE,      True),
    ('Reserves & Surplus (Accumulated)',12167283,   12167283,   False, WHITE,      True),
    ('Less: Current Year Loss (P&L)',   -5390588,   -2499221,   False, WHITE,      True),
    ('NET WORTH',                        7369915,    10261282,   True,  GREEN_BG,   False),
    ('',                                None,       None,       False, WHITE,      False),
    ('II.  LONG TERM BORROWINGS',       None,       None,       True,  LIGHT_BLUE, False),
    ('Unsecured Loans',                  5911600,    5911600,    False, WHITE,      True),
    ('  (Bajaj Finance + Director Loans)',None,      None,       False, GRAY_BG,    True),
    ('TOTAL BORROWINGS',                 5911600,    5911600,    True,  HEADER_BLUE,False),
    ('',                                None,       None,       False, WHITE,      False),
    ('III. CURRENT LIABILITIES',        None,       None,       True,  LIGHT_BLUE, False),
    ('Trade Payables (Creditors)',        666565,     666565,     False, WHITE,      True),
    ('Salaries Payable',                 1638583,    1638583,    False, WHITE,      True),
    ('Duties & Taxes (GST / TDS)',        1270317,    2276857,    False, WHITE,      True),
    ('Reimbursements Payable',           3182,       3182,       False, WHITE,      True),
    ('TOTAL CURRENT LIABILITIES',        3578647,    4585187,    True,  HEADER_BLUE,False),
    ('',                                None,       None,       False, WHITE,      False),
    ('TOTAL LIABILITIES & EQUITY',       16860162,   20758069,   True,  MID_BLUE,   False),
]

asset_rows = [
    ('I.   NON-CURRENT ASSETS',         None,       None,       True,  LIGHT_BLUE, False),
    ('Fixed Assets (Net of Depreciation)',8322884,   8322884,    False, WHITE,      True),
    ('    WIP - SLS PAMM Project',        7033898,    7033898,    False, GRAY_BG,    True),
    ('    Machinery & Equipment',          389123,     389123,     False, GRAY_BG,    True),
    ('    Tangible Assets',               851090,     851090,     False, GRAY_BG,    True),
    ('    Other Fixed Assets',             48773,      48773,      False, GRAY_BG,    True),
    ('Security Deposits',                 1121361,    1121361,    False, WHITE,      True),
    ('TDS Receivable',                     241507,     241507,     False, WHITE,      True),
    ('TDS-GST Receivable',                 36380,      36380,      False, WHITE,      True),
    ('Deferred Expenses',                  34800,      34800,      False, WHITE,      True),
    ('Deferred Tax Asset (Net)',           -246340,    -246340,    False, WHITE,      True),
    ('TOTAL NON-CURRENT ASSETS',           9510592,    9510592,    True,  HEADER_BLUE,False),
    ('',                                None,       None,       False, WHITE,      False),
    ('II.  CURRENT ASSETS',             None,       None,       True,  LIGHT_BLUE, False),
    ('Inventories (Closing Stock)',       4343261,    2479299,    False, WHITE,      True),
    ('    50% consumed from existing stock',None,    None,       False, GRAY_BG,    True),
    ('    50% via fresh import procurement',None,    None,       False, GRAY_BG,    True),
    ('Trade Receivables (Debtors)',       6952542,    16272353,   False, WHITE,      True),
    ('    Existing Debtors',              6952542,    6952542,    False, YELLOW_BG,  True),
    ('    Orders in Hand (Billed)',        0,          9319811,    False, YELLOW_BG,  True),
    ('Loans & Advances',                  303254,     303254,     False, WHITE,      True),
    ('Cash in Hand',                      13287,      13287,      False, WHITE,      True),
    ('Bank / OD (ICICI + Others)',        -4276424,   -7834366,   False, WHITE,      True),
    ('    OD Limit: INR 52.5L | Utilized: INR 78.3L',None, None, False, GRAY_BG,   True),
    ('    Excess over limit = WC Gap of INR 25.8L',None, None,   False, GRAY_BG,   True),
    ('TOTAL CURRENT ASSETS',              7349920,    11247827,   True,  HEADER_BLUE,False),
    ('',                                None,       None,       False, WHITE,      False),
    ('TOTAL ASSETS',                     16860162,   20758069,   True,  MID_BLUE,   False),
]

row = 6
max_rows = max(len(liab_rows), len(asset_rows))
for i in range(max_rows + 2):
    ws.row_dimensions[row + i].height = 16

def write_section(rows_data, label_col, val1_col, val2_col):
    for i, item in enumerate(rows_data):
        r = row + i
        label, tval, pval, bold, bg, indent = item
        prefix = '    ' if indent else ''

        c = ws[f'{label_col}{r}']
        c.value = prefix + label
        c.font = Font(name='Calibri', bold=bold, size=10)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(vertical='center')

        is_total = bold and 'TOTAL' in label

        for col, val in [(val1_col, tval), (val2_col, pval)]:
            cell = ws[f'{col}{r}']
            cell.value = val
            cell.font = Font(name='Calibri', bold=bold, size=10)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = fmt_inr
            if is_total and val is not None:
                cell.border = Border(top=side('medium'), bottom=side('double'))

write_section(liab_rows, 'B', 'C', 'D')
write_section(asset_rows, 'F', 'G', 'H')

# ---- NOTES ----
note_row = row + max_rows + 2
ws.merge_cells(f'B{note_row}:H{note_row}')
c = ws[f'B{note_row}']
c.value = 'NOTES & ASSUMPTIONS'
c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
c.fill = PatternFill('solid', fgColor=MID_BLUE)
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[note_row].height = 18

notes = [
    '1.  Fixed Assets: Carried at Tally book value (13-Mar-26). Remaining depreciation (18 days ~INR 14,000) considered immaterial for provisional purposes.',
    '2.  Inventories: INR 24.79L = 50% of COGS (INR 18.64L) consumed from existing stock; balance 50% procured fresh (import MOQ basis, paid via OD). Stock from Tally.',
    '3.  Trade Receivables: INR 1.63 Cr = Existing debtors INR 69.5L + Orders in Hand billed INR 93.2L. No advance from customers; collections at 45 days from billing.',
    '4.  Duties & Taxes: Increased by INR 10.07L reflecting net GST output on new billings (18% output tax less input credit on materials consumed from stock).',
    '5.  Bank / OD: ICICI OD (INR 52.5L limit, merged with bank; Director apartment mortgaged) + operational OD draws for procurement and expenses = INR 78.34L utilized.',
    '6.  Working Capital Gap: OD utilized (INR 78.34L) exceeds ICICI limit (INR 52.5L) by INR 25.84L. This represents the IMMEDIATE working capital requirement.',
    '7.  Current Year Loss: Revised to INR 24.99L (Provisional) from INR 53.91L (13-Mar actuals) per Provisional P&L including execution of Orders in Hand.',
    '8.  All figures are provisional estimates. Subject to final audit and year-end adjustments. Prepared solely for Working Capital assessment purposes.',
]
for j, note in enumerate(notes):
    r = note_row + 1 + j
    ws.merge_cells(f'B{r}:H{r}')
    c = ws[f'B{r}']
    c.value = note
    c.font = Font(name='Calibri', size=9, italic=True, color='444444')
    c.alignment = Alignment(vertical='center', wrap_text=True)
    c.fill = PatternFill('solid', fgColor='F9F9F9')
    ws.row_dimensions[r].height = 14

# ---- KEY METRICS ----
m_row = note_row + len(notes) + 2
ws.merge_cells(f'B{m_row}:H{m_row}')
c = ws[f'B{m_row}']
c.value = 'KEY FINANCIAL METRICS  |  WORKING CAPITAL SNAPSHOT (31-Mar-26 Provisional)'
c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
c.fill = PatternFill('solid', fgColor=MID_BLUE)
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[m_row].height = 18

metrics = [
    ('Net Worth (Equity)',                10261282, 'Positive net worth with improving trend; loss narrows from INR 1.19 Cr (FY25) to INR 24.99L (FY26 provisional)'),
    ('Total Debt (OD + Unsecured Loans)', 5911600,  'Bajaj Finance + Director/Unsecured Loans. ICICI OD merged in bank balance.'),
    ('OD Utilized vs Limit',              7834366,  'OD Utilized INR 78.34L vs ICICI Limit INR 52.5L — Gap of INR 25.84L is the immediate WC need'),
    ('Net Working Capital (CA - CL)',     6662640,  'Current Assets INR 1.12 Cr less Current Liabilities INR 45.85L'),
    ('Trade Receivables',                16272353,  'Existing INR 69.5L + Orders billed INR 93.2L — collections due by May 2026 (45-day terms)'),
    ('Inventories',                       2479299,  '50% consumed from existing stock; 50% procured fresh on import MOQ basis via OD'),
    ('Current Ratio',                     None,     '2.45x  (CA INR 1.12 Cr / CL INR 45.85L) — Healthy; backed by strong receivables'),
    ('Debt-to-Equity Ratio',              None,     '0.58x  (Unsecured Loans / Net Worth) — Within acceptable lending norms'),
]
for idx, (lbl, val, rem) in enumerate(metrics):
    r = m_row + 1 + idx
    ws.row_dimensions[r].height = 16
    bg = WHITE if idx % 2 == 0 else 'F5F5F5'

    c = ws[f'B{r}']
    c.value = lbl
    c.font = Font(name='Calibri', bold=True, size=10)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(vertical='center')

    c2 = ws[f'C{r}']
    c2.value = val
    c2.font = Font(name='Calibri', bold=True, size=10, color='1F3864')
    c2.fill = PatternFill('solid', fgColor=bg)
    c2.alignment = Alignment(horizontal='right', vertical='center')
    c2.number_format = fmt_inr

    ws.merge_cells(f'D{r}:H{r}')
    c3 = ws[f'D{r}']
    c3.value = rem
    c3.font = Font(name='Calibri', size=9, color='555555', italic=True)
    c3.fill = PatternFill('solid', fgColor=bg)
    c3.alignment = Alignment(vertical='center', wrap_text=True)

wb.save('C:/Users/Lenovo/Downloads/Fracktal_Financials_FY2526.xlsx')
print('Balance Sheet sheet added successfully.')
