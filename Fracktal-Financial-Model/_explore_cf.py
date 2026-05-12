"""Explore Cash Flow sheet structure and TB data for cash flow preparation."""
import openpyxl

f = r"D:\Suresh_AGENTS\Fracktal-Financial-Model\.tmp\Fracktal_MIS_Master_20260310_122834.xlsx"
wb = openpyxl.load_workbook(f)

# ── Cash Flow sheet ──
ws = wb['Cash Flow']
print("=" * 70)
print("CASH FLOW SHEET — Full dump")
print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
for r in range(1, ws.max_row + 1):
    vals = []
    for c in range(1, min(ws.max_column + 1, 16)):
        v = ws.cell(r, c).value
        if v is not None:
            vals.append(f"c{c}={v!r}")
    if vals:
        print(f"  Row {r:3d}: {' | '.join(vals)}")

# ── TB Sheet: all rows with data (to build the cash flow mapping) ──
ws2 = wb['TB']
print("\n" + "=" * 70)
print("TB SHEET — All labeled rows with APR-FEB data")
for r in range(1, ws2.max_row + 1):
    label = ws2.cell(r, 1).value
    if label:
        # Print label and month values (B=APR ... M=MAR)
        vals = {}
        for c in range(2, 14):
            v = ws2.cell(r, c).value
            if v is not None:
                vals[c] = v
        if vals:
            sample = f"APR(B)={vals.get(2,'-')} ... FEB(L)={vals.get(12,'-')}"
        else:
            sample = "(no data)"
        print(f"  Row {r:3d}: {str(label):55s} {sample}")

# ── Check formula rows 79-100 ──
print("\n" + "=" * 70)
print("TB Calculated Totals (rows 79-100)")
for r in range(79, min(ws2.max_row + 1, 115)):
    label = ws2.cell(r, 1).value
    val_b = ws2.cell(r, 2).value
    if label or val_b:
        print(f"  Row {r:3d}: {str(label or ''):55s} B={val_b!r}")

# ── CRM Data — orders in hand (PO Received) ──
ws3 = wb['CRM Data']
print("\n" + "=" * 70)
print("CRM DATA — PO Received deals (orders in hand)")
for r in range(5, ws3.max_row + 1):
    stage = ws3.cell(r, 2).value
    if stage and 'PO Received' in str(stage):
        name = ws3.cell(r, 1).value
        amt = ws3.cell(r, 4).value
        weighted = ws3.cell(r, 7).value
        print(f"  {name}: Amount={amt}, Weighted={weighted}")

# Also check Awaiting PO
print("\nCRM DATA — Awaiting PO deals")
for r in range(5, ws3.max_row + 1):
    stage = ws3.cell(r, 2).value
    if stage and 'Awaiting PO' in str(stage):
        name = ws3.cell(r, 1).value
        amt = ws3.cell(r, 4).value
        print(f"  {name}: Amount={amt}")

wb.close()
