import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.load_workbook('C:/Users/Lenovo/Downloads/Fracktal_Financials_FY2526.xlsx')
if 'Cash Deficit — Apr-May 26' in wb.sheetnames:
    del wb['Cash Deficit — Apr-May 26']
ws = wb.create_sheet('Cash Deficit — Apr-May 26')

# ── Colours ─────────────────────────────────────────────────────────────────
DARK_BLUE   = '1F3864'
MID_BLUE    = '2E5F9A'
LIGHT_BLUE  = 'D9E1F2'
HEADER_BLUE = 'BDD7EE'
GREEN_BG    = 'E2EFDA'
DARK_GREEN  = '375623'
RED_BG      = 'FCE4D6'
DARK_RED    = 'C00000'
YELLOW_BG   = 'FFF2CC'
WHITE       = 'FFFFFF'
GRAY_BG     = 'F2F2F2'
ORANGE_BG   = 'FDEBD0'

def S(s='thin'): return Side(style=s)
def bdr():       return Border(left=S(), right=S(), top=S(), bottom=S())
def med_top():   return Border(top=S('medium'), bottom=S('double'))

fmt_inr = '_(\u20b9* #,##0_);_(\u20b9* (#,##0);_(\u20b9* "-"_);_(@_)'

for col, w in [('A',3),('B',50),('C',18),('D',32)]:
    ws.column_dimensions[col].width = w

# ── Key numbers ──────────────────────────────────────────────────────────────
# Derived from actual order and OPEX data (consistent with other sheets)
ORDERS_VAL   = 9319811   # Confirmed orders in hand
COGS_ORDERS  = 3727924   # COGS on those orders
FRESH_COGS   = 1863962   # 50% fresh procurement (import MOQ)
IMPORT_ADV   = 1304773   # 70% advance to import supplier
CRED_AMT     =  559189   # 30% supplier credit (30-day deferral)
DIRECT_EXP   =  710782   # Direct/manufacturing costs
PAYROLL_PM   =  835249   # Monthly payroll (₹1.00Cr annual ÷ 12)
OTHER_PM     =  572542   # Other monthly overheads (admin, professional, etc.)
FINANCE_PM   =   31077   # Monthly finance charges / interest
MONTHLY_OPEX = 1438868   # Total monthly fixed costs (= payroll + other + finance)
ICICI_LIMIT  = 5250000   # ICICI OD sanctioned limit (100% utilized)
BAJAJ_LIMIT  = 1500000   # Bajaj Finance (100% utilized)
EXISTING_DBT = 6952542   # Existing debtors collectible in April–May

# Derived
APR_OPEX  = MONTHLY_OPEX                      # April fixed costs
MAY_OPEX  = MONTHLY_OPEX                      # May fixed costs
FREE_CASH = EXISTING_DBT - ICICI_LIMIT        # Cash left after OD repayment from collections

TOTAL_OUTFLOW = IMPORT_ADV + DIRECT_EXP + APR_OPEX + MAY_OPEX
TOTAL_INFLOW  = FREE_CASH + CRED_AMT
DEFICIT       = TOTAL_OUTFLOW - TOTAL_INFLOW

# ── Helpers ──────────────────────────────────────────────────────────────────
def title_row(r, txt, sz=12, bold=True, italic=False, h=18):
    ws.merge_cells(f'B{r}:D{r}')
    c = ws[f'B{r}']
    c.value = txt
    c.font = Font(name='Calibri', bold=bold, size=sz, color=WHITE, italic=italic)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = h

def sec(r, title, bg=MID_BLUE, h=20):
    ws.merge_cells(f'B{r}:D{r}')
    c = ws[f'B{r}']
    c.value = title
    c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = h

def row(r, label, amount=None, note='', bold=False, bg=WHITE, indent=1,
        lcolor='000000', acolor='000000', h=16, italic=False):
    ws.row_dimensions[r].height = h
    c = ws[f'B{r}']
    c.value = ('    ' * indent) + label
    c.font = Font(name='Calibri', bold=bold, size=10, color=lcolor, italic=italic)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(vertical='center')

    ca = ws[f'C{r}']
    ca.value = amount
    ca.font = Font(name='Calibri', bold=bold, size=10, color=acolor)
    ca.fill = PatternFill('solid', fgColor=bg)
    ca.alignment = Alignment(horizontal='right', vertical='center')
    ca.number_format = fmt_inr

    if note:
        nd = ws[f'D{r}']
        nd.value = note
        nd.font = Font(name='Calibri', size=9, italic=True, color='555555')
        nd.fill = PatternFill('solid', fgColor=bg)
        nd.alignment = Alignment(vertical='center', wrap_text=True)

def tot(r, label, amount, bg=HEADER_BLUE, lcolor='000000', acolor='000000', sz=10, h=18):
    ws.row_dimensions[r].height = h
    c = ws[f'B{r}']
    c.value = label
    c.font = Font(name='Calibri', bold=True, size=sz, color=lcolor)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(vertical='center')

    ca = ws[f'C{r}']
    ca.value = amount
    ca.font = Font(name='Calibri', bold=True, size=sz, color=acolor)
    ca.fill = PatternFill('solid', fgColor=bg)
    ca.alignment = Alignment(horizontal='right', vertical='center')
    ca.number_format = fmt_inr
    ca.border = med_top()

def note_box(r, text, bg='DEEAF1', h=30, color='1F3864'):
    ws.merge_cells(f'B{r}:D{r}')
    c = ws[f'B{r}']
    c.value = text
    c.font = Font(name='Calibri', size=9, italic=True, color=color)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[r].height = h

def sp(r, h=7): ws.row_dimensions[r].height = h

# ════════════════════════════════════════════════════════════════════════════
# TITLE
# ════════════════════════════════════════════════════════════════════════════
title_row(1, 'Fracktal Works Private Limited', sz=14, h=22)
title_row(2, 'Cash Deficit Analysis  |  April & May 2026  |  To Fulfil Confirmed Orders & Cover Fixed Costs', sz=11)
title_row(3, 'CIN: U30009KA2013PTC070124  |  A simple month-by-month cash picture  |  (All amounts in INR)', sz=9, bold=False, italic=True, h=14)
sp(4, 10)

# ════════════════════════════════════════════════════════════════════════════
# STEP 1 — THE OPPORTUNITY
# ════════════════════════════════════════════════════════════════════════════
sec(5, '  STEP 1 :  THE CONFIRMED ORDERS WE HAVE IN HAND  (Ready to Execute)')
row(6,  'Total confirmed orders — goods to be made and billed', ORDERS_VAL,
    note='PO received ₹80.2L + Awaiting PO ₹13.0L', bold=True, bg=GREEN_BG,
    lcolor=DARK_GREEN, acolor=DARK_GREEN, indent=1)
note_box(7, '✔  These orders are CONFIRMED.  Once we procure materials, assemble, and ship — customers will pay ₹93.19L within 45 days of billing.',
         bg=GREEN_BG, color=DARK_GREEN, h=20)
sp(8, 8)

# ════════════════════════════════════════════════════════════════════════════
# STEP 2 — MONEY WE NEED TO SPEND (OUTFLOWS)
# ════════════════════════════════════════════════════════════════════════════
sec(9, '  STEP 2 :  MONEY WE NEED TO SPEND — APRIL & MAY 2026')

# Part A — Order execution
row(10, 'A.  TO MAKE AND DELIVER THE MACHINES  (Order Execution Costs)', bg=LIGHT_BLUE,
    bold=True, indent=0, lcolor=DARK_BLUE)

row(11, 'Raw materials — imported components  (advance payment to supplier)',
    IMPORT_ADV, note='50% of COGS procured fresh via import. Must pay 70% in advance when placing PO.',
    bg=WHITE, indent=1)

row(12, 'Materials from existing stock  (already in our warehouse)',
    None, note='50% of COGS consumed from stock we already hold — NO cash payment needed',
    bg=GRAY_BG, italic=True, indent=1, lcolor='666666')

row(13, 'Manufacturing costs  (assembly, testing, quality check, packing)',
    DIRECT_EXP, note='Direct production costs to convert components into finished machines',
    bg=WHITE, indent=1)

tot(14, '    Sub-Total A — Cost to Execute Orders', f'=C{11}+C{13}', bg=HEADER_BLUE)
sp(15, 5)

# Part B — Fixed costs April
row(16, 'B.  FIXED COSTS — APRIL 2026  (Business must run while machines are being made)', bg=LIGHT_BLUE,
    bold=True, indent=0, lcolor=DARK_BLUE)

row(17, 'Staff Salaries & HR (April)',     PAYROLL_PM,  note='~₹8.35L/month — due 30th April', bg=WHITE, indent=1)
row(18, 'Office & Admin Overheads (April)', OTHER_PM,   note='Rent, utilities, professional fees, travel, R&D, etc.', bg=WHITE, indent=1)
row(19, 'Finance Charges / Interest (April)', FINANCE_PM, note='Interest on ICICI OD + Bajaj EMI', bg=WHITE, indent=1)
tot(20, '    Sub-Total B — April Fixed Costs', f'=C{17}+C{18}+C{19}', bg=HEADER_BLUE)
sp(21, 5)

# Part C — Fixed costs May
row(22, 'C.  FIXED COSTS — MAY 2026  (Business continues; collections expected by end of May)', bg=LIGHT_BLUE,
    bold=True, indent=0, lcolor=DARK_BLUE)

row(23, 'Staff Salaries & HR (May)',      PAYROLL_PM,  note='~₹8.35L/month — due 31st May', bg=WHITE, indent=1)
row(24, 'Office & Admin Overheads (May)', OTHER_PM,    note='Same monthly overhead continues', bg=WHITE, indent=1)
row(25, 'Finance Charges / Interest (May)', FINANCE_PM, note='Interest on OD + Bajaj EMI', bg=WHITE, indent=1)
tot(26, '    Sub-Total C — May Fixed Costs', f'=C{23}+C{24}+C{25}', bg=HEADER_BLUE)
sp(27, 8)

# GRAND TOTAL OUTFLOW
tot(28, 'TOTAL MONEY WE NEED TO SPEND  (A + B + C)',
    f'=C{14}+C{20}+C{26}',
    bg='C00000', lcolor=WHITE, acolor=WHITE, sz=11, h=22)
sp(29, 10)

# ════════════════════════════════════════════════════════════════════════════
# STEP 3 — CASH WE HAVE TODAY
# ════════════════════════════════════════════════════════════════════════════
sec(30, '  STEP 3 :  CASH WE HAVE TODAY  (Right Now — as at 14 March 2026)', bg='7B3F00')

row(31, 'Cash in hand / Current account balance', 0,
    note='Bank accounts are running on OD — no free cash balance',
    bg=RED_BG, bold=True, lcolor=DARK_RED, acolor=DARK_RED, indent=1)

row(32, 'ICICI Bank OD — Sanctioned limit ₹52.5L   →   Amount drawn: ₹52.5L   →   Available: ₹0',
    0, note='100% utilized. Cannot draw further without repaying first.',
    bg=RED_BG, lcolor=DARK_RED, indent=1)

row(33, 'Bajaj Finance — Sanctioned limit ₹15.0L   →   Amount drawn: ₹15.0L   →   Available: ₹0',
    0, note='100% utilized.',
    bg=RED_BG, lcolor=DARK_RED, indent=1)

tot(34, 'TOTAL CASH AVAILABLE TODAY', 0, bg=DARK_RED, lcolor=WHITE, acolor=WHITE, sz=11, h=20)
sp(35, 10)

# ════════════════════════════════════════════════════════════════════════════
# STEP 4 — MONEY COMING IN (INFLOWS)
# ════════════════════════════════════════════════════════════════════════════
sec(36, '  STEP 4 :  MONEY COMING IN — APRIL & MAY 2026')

row(37, 'Existing customers who owe us money  (as at today)',
    EXISTING_DBT,
    note='These debtors are expected to pay within 45 days of their invoice date — April / May 2026',
    bg=YELLOW_BG, bold=True, indent=1)

row(38, 'Less: Goes directly to repay ICICI OD  (bank automatically applies)',
    -ICICI_LIMIT,
    note='When customers deposit money, bank applies it to reduce the OD balance first',
    bg=GRAY_BG, italic=True, lcolor='555555', indent=2)

tot(39, '    Free Cash After OD Repayment  (₹69.5L collections − ₹52.5L OD repaid)',
    FREE_CASH, bg=GREEN_BG, lcolor=DARK_GREEN, acolor=DARK_GREEN)

row(40, 'Supplier credit  (30% of import materials — 30-day credit terms)',
    CRED_AMT,
    note='30% of fresh material cost can be deferred — supplier gives 30-day credit',
    bg=YELLOW_BG, indent=1)

tot(41, 'TOTAL USABLE FUNDS  (Free cash + Supplier credit)',
    f'=C{39}+C{40}',
    bg=HEADER_BLUE, sz=11, h=20)
sp(42, 10)

# ════════════════════════════════════════════════════════════════════════════
# STEP 5 — THE GAP (DEFICIT)
# ════════════════════════════════════════════════════════════════════════════
sec(43, '  STEP 5 :  THE CASH GAP  (What We Are Short Of)', bg=DARK_RED)

row(44, 'Total Money We Need to Spend  (Step 2)',
    f'=C{28}', bg=RED_BG, bold=True, lcolor=DARK_RED, acolor=DARK_RED, indent=1)

row(45, 'Less: Total Usable Funds  (Step 4)',
    f'=-C{41}', bg=RED_BG, italic=True, lcolor='555555', indent=2)

ws.row_dimensions[46].height = 26
ws.merge_cells('B46:B46')
c = ws['B46']
c.value = 'CASH SHORTFALL  —  This is the amount we need from outside to keep things moving'
c.font = Font(name='Calibri', bold=True, size=12, color=WHITE)
c.fill = PatternFill('solid', fgColor=DARK_RED)
c.alignment = Alignment(vertical='center')

ca = ws['C46']
ca.value = f'=C{44}+C{45}'
ca.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
ca.fill = PatternFill('solid', fgColor=DARK_RED)
ca.alignment = Alignment(horizontal='right', vertical='center')
ca.number_format = fmt_inr
ca.border = med_top()

nd = ws['D46']
nd.value = 'This is the minimum external funding needed to execute orders AND pay staff / overheads for April & May'
nd.font = Font(name='Calibri', size=9, italic=True, color=WHITE)
nd.fill = PatternFill('solid', fgColor=DARK_RED)
nd.alignment = Alignment(vertical='center', wrap_text=True)

sp(47, 10)

# ════════════════════════════════════════════════════════════════════════════
# STEP 6 — HOW WILL IT BE REPAID?
# ════════════════════════════════════════════════════════════════════════════
sec(48, '  STEP 6 :  HOW AND WHEN WILL THE MONEY COME BACK?', bg=DARK_GREEN)

row(49, 'Once machines are made and shipped, we raise invoices on customers',
    None, note='Billing happens April–May 2026 as machines are dispatched',
    bg=GREEN_BG, lcolor=DARK_GREEN, indent=1)

row(50, 'New customers pay within 45 days of invoice  →  Collections: May–June 2026',
    ORDERS_VAL, note='₹93.19L new order collections — primary repayment source',
    bg=GREEN_BG, bold=True, lcolor=DARK_GREEN, acolor=DARK_GREEN, indent=1)

row(51, 'Remaining balance after repaying the loan  (surplus to business)',
    f'=C{50}-C{46}', note='After repaying the working capital loan, this balance is available to the business',
    bg=GREEN_BG, lcolor=DARK_GREEN, acolor=DARK_GREEN, indent=2)

note_box(52,
    '✔  This is a SELF-LIQUIDATING loan — the money lent for procurement comes back directly from customer payments within 90 days.  '
    'No separate assets need to be sold to repay.  The confirmed orders themselves are the repayment source.',
    bg=GREEN_BG, color=DARK_GREEN, h=28)

sp(53, 10)

# ════════════════════════════════════════════════════════════════════════════
# SUMMARY BOX
# ════════════════════════════════════════════════════════════════════════════
sec(54, '  SUMMARY AT A GLANCE', bg=DARK_BLUE, h=20)

summary_rows = [
    (55, 'Orders confirmed — ready to execute',      ORDERS_VAL, '₹93.19L',      GREEN_BG, DARK_GREEN),
    (56, 'Total cash needed (Apr + May)',            f'=C{28}',   '~₹48.9L',      RED_BG,   DARK_RED),
    (57, 'Cash available from our own sources',      f'=C{41}',   '~₹17.6L',      YELLOW_BG,'000000'),
    (58, 'CASH SHORTFALL  (funding gap)',            f'=C{46}',   '~₹26–30L',     RED_BG,   DARK_RED),
    (59, 'Collections from new orders (May–Jun 26)', ORDERS_VAL,  '₹93.19L',      GREEN_BG, DARK_GREEN),
    (60, 'Loan repaid from order collections',       None,        'Within 90 days',GREEN_BG, DARK_GREEN),
]

for r, lbl, val, approx, bg, color in summary_rows:
    ws.row_dimensions[r].height = 17
    bold = 'SHORTFALL' in lbl or 'Orders confirmed' in lbl
    sz = 11 if bold else 10

    c = ws[f'B{r}']
    c.value = ('    ' if not bold else '') + lbl
    c.font = Font(name='Calibri', bold=bold, size=sz, color=color)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(vertical='center')

    ca = ws[f'C{r}']
    ca.value = val
    ca.font = Font(name='Calibri', bold=bold, size=sz, color=color)
    ca.fill = PatternFill('solid', fgColor=bg)
    ca.alignment = Alignment(horizontal='right', vertical='center')
    ca.number_format = fmt_inr
    if bold: ca.border = med_top()

    nd = ws[f'D{r}']
    nd.value = approx
    nd.font = Font(name='Calibri', bold=bold, size=sz, color=color, italic=(not bold))
    nd.fill = PatternFill('solid', fgColor=bg)
    nd.alignment = Alignment(horizontal='center', vertical='center')

sp(61, 12)

# Final note
note_box(62,
    'Note: 50% of raw material cost is sourced from existing warehouse stock (no cash outflow). '
    'Bajaj Finance ₹15L (term loan) is covered by regular EMIs included in monthly Finance Charges above. '
    'All figures are based on confirmed order values and actual monthly cost run-rates.',
    bg=GRAY_BG, color='555555', h=24)

wb.save('C:/Users/Lenovo/Downloads/Fracktal_Financials_FY2526.xlsx')
print('Cash Deficit sheet built successfully.')
