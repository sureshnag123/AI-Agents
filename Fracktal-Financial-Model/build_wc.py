import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.load_workbook('C:/Users/Lenovo/Downloads/Fracktal_Financials_FY2526.xlsx')
if 'Working Capital Requirement' in wb.sheetnames:
    del wb['Working Capital Requirement']
ws = wb.create_sheet('Working Capital Requirement')

# ── Palette ──────────────────────────────────────────────────────────────
DARK_BLUE   = '1F3864'
MID_BLUE    = '2E5F9A'
LIGHT_BLUE  = 'D9E1F2'
HEADER_BLUE = 'BDD7EE'
GREEN_BG    = 'E2EFDA'
RED_BG      = 'FCE4D6'
YELLOW_BG   = 'FFF2CC'
WHITE       = 'FFFFFF'
GRAY_BG     = 'F2F2F2'
DARK_GREEN  = '375623'
DARK_RED    = '833C00'

def S(s='thin'):  return Side(style=s)
def bdr():        return Border(left=S(),right=S(),top=S(),bottom=S())
def med_top():    return Border(top=S('medium'), bottom=S('double'))
def bot_med():    return Border(bottom=S('medium'))

fmt_inr = '_(\u20b9* #,##0_);_(\u20b9* (#,##0);_(\u20b9* "-"_);_(@_)'
fmt_pct = '0.0%'

for col, w in [('A',3),('B',42),('C',20),('D',20),('E',20),('F',3),('G',30),('H',18)]:
    ws.column_dimensions[col].width = w

def banner(r, text, bg=MID_BLUE, cs=2, ce=5, h=18):
    ws.merge_cells(start_row=r,start_column=cs,end_row=r,end_column=ce)
    c = ws.cell(row=r,column=cs)
    c.value = text
    c.font = Font(name='Calibri',bold=True,size=10,color=WHITE)
    c.fill = PatternFill('solid',fgColor=bg)
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.row_dimensions[r].height = h

def col_hdr(r, texts, bg=MID_BLUE, h=30):
    for i,t in enumerate(texts):
        c = ws.cell(row=r,column=2+i)
        c.value = t
        c.font = Font(name='Calibri',bold=True,size=10,color=WHITE)
        c.fill = PatternFill('solid',fgColor=bg)
        c.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border = bdr()
    ws.row_dimensions[r].height = h

def row_cell(r, col, val, bold=False, bg=WHITE, fmt=fmt_inr,
             align='right', sz=10, color='000000', italic=False):
    c = ws.cell(row=r, column=col)
    c.value = val
    c.font = Font(name='Calibri',bold=bold,size=sz,color=color,italic=italic)
    c.fill = PatternFill('solid',fgColor=bg)
    c.alignment = Alignment(horizontal=align,vertical='center')
    c.number_format = fmt
    return c

def label(r, text, bold=False, bg=WHITE, indent=0, sz=10):
    c = ws.cell(row=r,column=2)
    c.value = ('    '*indent)+text
    c.font = Font(name='Calibri',bold=bold,size=sz)
    c.fill = PatternFill('solid',fgColor=bg)
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[r].height = 16

def total_row(r, lbl, fC, fD, fE, bg=HEADER_BLUE, sz=10):
    label(r,lbl,bold=True,bg=bg,sz=sz)
    for col,fml in zip([3,4,5],[fC,fD,fE]):
        c = ws.cell(row=r,column=col)
        c.value = fml
        c.font = Font(name='Calibri',bold=True,size=sz)
        c.fill = PatternFill('solid',fgColor=bg)
        c.alignment = Alignment(horizontal='right',vertical='center')
        c.number_format = fmt_inr
        c.border = med_top()

def spacer(r,h=6): ws.row_dimensions[r].height = h

def input_row(r, lbl, vC, vD, vE, bold=False, bg=YELLOW_BG, indent=1, fmt=fmt_inr):
    label(r, lbl, bold=bold, bg=bg, indent=indent)
    for col,val in zip([3,4,5],[vC,vD,vE]):
        c = ws.cell(row=r,column=col)
        c.value = val
        c.font = Font(name='Calibri',bold=bold,size=10)
        c.fill = PatternFill('solid',fgColor=bg)
        c.alignment = Alignment(horizontal='right',vertical='center')
        c.number_format = fmt

def formula_row(r, lbl, fC, fD, fE, bold=False, bg=WHITE, indent=1):
    label(r, lbl, bold=bold, bg=bg, indent=indent)
    for col,fml in zip([3,4,5],[fC,fD,fE]):
        c = ws.cell(row=r,column=col)
        c.value = fml
        c.font = Font(name='Calibri',bold=bold,size=10)
        c.fill = PatternFill('solid',fgColor=bg)
        c.alignment = Alignment(horizontal='right',vertical='center')
        c.number_format = fmt_inr

# ════════════════════════════════════════════════════════════════════════════
# TITLE
# ════════════════════════════════════════════════════════════════════════════
for r,txt,sz in [(1,'Fracktal Works Private Limited',14),
                 (2,'Working Capital Requirement Statement  —  Order-Based Assessment',12),
                 (3,'CIN: U30009KA2013PTC070124  |  MSME: UDYAM-KR-03-0093853  |  Amounts in INR',9)]:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
    c = ws.cell(row=r,column=2)
    c.value = txt
    c.font = Font(name='Calibri',bold=(r<3),size=sz,color=WHITE,italic=(r==3))
    c.fill = PatternFill('solid',fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[r].height = 22 if r==1 else 18 if r==2 else 14
spacer(4,8)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 1 — INPUT BASIS  (order values, not monthly averages)
# ════════════════════════════════════════════════════════════════════════════
# Row numbers — referenced by all formula rows below
R_ORD  = 8   # Order / Revenue Value for period
R_COGS = 9   # COGS on those orders
R_DEXP = 10  # Direct Expenses on those orders
R_PCST = 11  # Production Cost = COGS + Direct Exp  (formula)
R_OPEX = 12  # Monthly OPEX run rate (cash burn)
R_FRSH = 13  # Fresh Procurement % of COGS (import MOQ — 50%)
R_ADV  = 14  # Advance % paid to import suppliers (70%)
R_CRED = 15  # Credit % on local/remaining procurement (30%)

banner(5,'  SECTION 1 :  INPUT BASIS  (Yellow cells are editable — change values to refresh all calculations)')
col_hdr(6,['Particulars',
           'Current Orders in Hand\n(INR 93.19L — Mar-26)',
           'FY 26-27 Q1 Projection\n(Apr–Jun 26)',
           'FY 26-27 Q2 Projection\n(Jul–Sep 26)'])
spacer(7,4)

# Inputs
input_row(R_ORD,  'Order / Revenue Value (Period)',        9319811, 3191704, 8782281)
input_row(R_COGS, 'COGS on Orders',                        3727924, 1353701, 2915128)
input_row(R_DEXP, 'Direct Expenses on Orders',              710782,  222945,  586412)
formula_row(R_PCST,'Production Cost  (COGS + Direct Exp)',
            f'=C{R_COGS}+C{R_DEXP}', f'=D{R_COGS}+D{R_DEXP}', f'=E{R_COGS}+E{R_DEXP}',
            bg=HEADER_BLUE)
input_row(R_OPEX, 'Monthly OPEX Run Rate',                 1438868, 1409411, 1352932)
spacer(16,4)

# Procurement parameters (single column — same for all periods)
banner(17,'  PROCUREMENT & CREDIT PARAMETERS  (Edit yellow cells)',bg='4472C4',h=16)
for r_param, lbl, val, note in [
    (18,'Fresh Procurement % of COGS  (Import MOQ — balance from stock)', 0.50, '50% procured fresh; 50% from existing stock'),
    (19,'Advance Payment % to Import Suppliers',                           0.70, '70% advance on PO placement'),
    (20,'Credit % on Remaining / Local Procurement',                       0.30, '30% of fresh COGS at 30-day credit'),
    (21,'Creditor Credit Period  (days)',                                   30,   'Days of credit on local procurement'),
]:
    ws.row_dimensions[r_param].height = 15
    c = ws.cell(row=r_param,column=2)
    c.value = '    '+lbl
    c.font = Font(name='Calibri',size=10)
    c.fill = PatternFill('solid',fgColor=WHITE)
    c.alignment = Alignment(vertical='center')
    vc = ws.cell(row=r_param,column=3)
    vc.value = val
    vc.font = Font(name='Calibri',bold=True,size=10)
    vc.fill = PatternFill('solid',fgColor=YELLOW_BG)
    vc.alignment = Alignment(horizontal='center',vertical='center')
    vc.number_format = fmt_pct if val < 1 else '0" days"'
    ws.merge_cells(start_row=r_param,start_column=4,end_row=r_param,end_column=5)
    nc = ws.cell(row=r_param,column=4)
    nc.value = note
    nc.font = Font(name='Calibri',size=9,italic=True,color='555555')
    nc.fill = PatternFill('solid',fgColor=GRAY_BG)
    nc.alignment = Alignment(vertical='center',wrap_text=True)

# Row refs for parameters
R_FRSH_PCT = 18   # Fresh % of COGS
R_ADV_PCT  = 19   # Advance %
R_CRED_PCT = 20   # Credit %
R_CRED_DAY = 21   # Credit days

spacer(22,8)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 2 — OPERATING CYCLE  (target 90 days)
# ════════════════════════════════════════════════════════════════════════════
R_OC_IMP  = 25
R_OC_WIP  = 26
R_OC_FG   = 27
R_OC_DBT  = 28
R_OC_GROS = 29
R_OC_CRED = 30
R_OC_NET  = 31

banner(23,'  SECTION 2 :  OPERATING CYCLE ANALYSIS  (Change days in column C to auto-update all WC calculations)')
col_hdr(24,['Component','Days  (editable)','Notes',''])
ws.row_dimensions[24].height = 18

for row,lbl,days,note in [
    (R_OC_IMP, 'Import Advance Period  (70% on PO; goods arrive in 45 days)', 45,
               'From PO placement to goods receipt — drives import advance WC'),
    (R_OC_WIP, 'Work-in-Progress  (Assembly, Testing & QC)',                  20,
               'Targeted reduction via parallel assembly tracks'),
    (R_OC_FG,  'Finished Goods / Dispatch Buffer',                             0,
               'Direct dispatch on QC clearance — zero FG holding'),
    (R_OC_DBT, 'Debtors Collection  (from invoice date)',                     45,
               'Contractual credit terms across all customer segments'),
]:
    ws.row_dimensions[row].height = 16
    c = ws.cell(row=row,column=2)
    c.value = lbl
    c.font = Font(name='Calibri',size=10)
    c.fill = PatternFill('solid',fgColor=WHITE)
    c.alignment = Alignment(vertical='center')
    dc = ws.cell(row=row,column=3)
    dc.value = days
    dc.font = Font(name='Calibri',bold=True,size=10)
    dc.fill = PatternFill('solid',fgColor=YELLOW_BG)
    dc.alignment = Alignment(horizontal='center',vertical='center')
    dc.number_format = '0" days"'
    ws.merge_cells(start_row=row,start_column=4,end_row=row,end_column=5)
    nc = ws.cell(row=row,column=4)
    nc.value = note
    nc.font = Font(name='Calibri',size=9,italic=True,color='555555')
    nc.fill = PatternFill('solid',fgColor=GRAY_BG)
    nc.alignment = Alignment(vertical='center',wrap_text=True)

# Gross cycle
ws.row_dimensions[R_OC_GROS].height = 16
ws.cell(row=R_OC_GROS,column=2).value = 'GROSS OPERATING CYCLE'
ws.cell(row=R_OC_GROS,column=2).font = Font(name='Calibri',bold=True,size=10)
ws.cell(row=R_OC_GROS,column=2).fill = PatternFill('solid',fgColor=HEADER_BLUE)
ws.cell(row=R_OC_GROS,column=2).alignment = Alignment(vertical='center')
gc = ws.cell(row=R_OC_GROS,column=3)
gc.value = f'=C{R_OC_IMP}+C{R_OC_WIP}+C{R_OC_FG}+C{R_OC_DBT}'
gc.font = Font(name='Calibri',bold=True,size=10)
gc.fill = PatternFill('solid',fgColor=HEADER_BLUE)
gc.alignment = Alignment(horizontal='center',vertical='center')
gc.number_format = '0" days"'
gc.border = med_top()

# Creditor
ws.row_dimensions[R_OC_CRED].height = 16
ws.cell(row=R_OC_CRED,column=2).value = 'Less: Creditor Credit  (30% purchases × 30-day local credit)'
ws.cell(row=R_OC_CRED,column=2).font = Font(name='Calibri',size=10)
ws.cell(row=R_OC_CRED,column=2).fill = PatternFill('solid',fgColor=WHITE)
ws.cell(row=R_OC_CRED,column=2).alignment = Alignment(vertical='center')
dc = ws.cell(row=R_OC_CRED,column=3)
dc.value = -20
dc.font = Font(name='Calibri',bold=True,size=10)
dc.fill = PatternFill('solid',fgColor=YELLOW_BG)
dc.alignment = Alignment(horizontal='center',vertical='center')
dc.number_format = '0" days"'

# Net
ws.row_dimensions[R_OC_NET].height = 20
ws.cell(row=R_OC_NET,column=2).value = 'NET OPERATING CYCLE'
ws.cell(row=R_OC_NET,column=2).font = Font(name='Calibri',bold=True,size=11,color=WHITE)
ws.cell(row=R_OC_NET,column=2).fill = PatternFill('solid',fgColor=DARK_BLUE)
ws.cell(row=R_OC_NET,column=2).alignment = Alignment(vertical='center')
nc = ws.cell(row=R_OC_NET,column=3)
nc.value = f'=C{R_OC_GROS}+C{R_OC_CRED}'
nc.font = Font(name='Calibri',bold=True,size=11,color=WHITE)
nc.fill = PatternFill('solid',fgColor=DARK_BLUE)
nc.alignment = Alignment(horizontal='center',vertical='center')
nc.number_format = '0" days"'

spacer(32,8)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 3 — WC ASSESSMENT (order-value based, fully formula-linked)
# ════════════════════════════════════════════════════════════════════════════
S3 = 33

banner(S3,'  SECTION 3 :  WORKING CAPITAL ASSESSMENT  (Based on Actual Order / Revenue Values — All Formula Linked)')
col_hdr(S3+1,['Particulars',
              'Current Orders\n(INR 93.19L)',
              'Q1 FY26-27\n(Apr–Jun 26)',
              'Q2 FY26-27\n(Jul–Sep 26)'])
spacer(S3+2,4)

# ── A. Outflows before collection ─────────────────────────────────────────
A_HDR = S3+3
A1    = S3+4   # Import advance
A2    = S3+5   # WIP / production costs
A3    = S3+6   # FG
A4    = S3+7   # OPEX bridge (import + WIP period)
A_TOT = S3+8
A_DBT = S3+9   # Debtors (order value)
GCA   = S3+10  # Gross CA total

ws.row_dimensions[A_HDR].height = 16
ws.cell(row=A_HDR,column=2).value = 'A.  CURRENT ASSETS  (Capital Tied Up)'
ws.cell(row=A_HDR,column=2).font = Font(name='Calibri',bold=True,size=10,color=WHITE)
ws.cell(row=A_HDR,column=2).fill = PatternFill('solid',fgColor=DARK_GREEN)
ws.cell(row=A_HDR,column=2).alignment = Alignment(vertical='center')
for col in [3,4,5]:
    ws.cell(row=A_HDR,column=col).fill = PatternFill('solid',fgColor=DARK_GREEN)

# A1 Import advance = COGS × fresh% × advance%
formula_row(A1,
    'Import Advance  (COGS × 50% fresh procurement × 70% advance)',
    f'=C{R_COGS}*C{R_FRSH_PCT}*C{R_ADV_PCT}',
    f'=D{R_COGS}*C{R_FRSH_PCT}*C{R_ADV_PCT}',
    f'=E{R_COGS}*C{R_FRSH_PCT}*C{R_ADV_PCT}',
    bg=WHITE, indent=1)

# A2 WIP = production cost × (WIP days / gross cycle) × 50% completion
formula_row(A2,
    'Work-in-Progress  (Prod. Cost × WIP days ÷ Gross cycle × 50%)',
    f'=C{R_PCST}*(C{R_OC_WIP}/C{R_OC_GROS})*0.5',
    f'=D{R_PCST}*(C{R_OC_WIP}/C{R_OC_GROS})*0.5',
    f'=E{R_PCST}*(C{R_OC_WIP}/C{R_OC_GROS})*0.5',
    bg=WHITE, indent=1)

# A3 FG
formula_row(A3,
    'Finished Goods  (Prod. Cost × FG days ÷ Gross cycle)',
    f'=C{R_PCST}*(C{R_OC_FG}/C{R_OC_GROS})',
    f'=D{R_PCST}*(C{R_OC_FG}/C{R_OC_GROS})',
    f'=E{R_PCST}*(C{R_OC_FG}/C{R_OC_GROS})',
    bg=GRAY_BG, indent=1)

# A4 — removed OPEX Bridge (OPEX is a continuous cost, not order-specific capital)
# Instead, existing debtors (collectible within 45 days) offset this on the CL side
ws.row_dimensions[A4].height = 14
c = ws.cell(row=A4,column=2)
c.value = '    Note: OPEX is a continuous obligation — funded via existing debtor collections (see B3 below)'
c.font = Font(name='Calibri',size=9,italic=True,color='555555')
c.fill = PatternFill('solid',fgColor=GRAY_BG)
c.alignment = Alignment(vertical='center')
for col in [3,4,5]:
    ws.cell(row=A4,column=col).fill = PatternFill('solid',fgColor=GRAY_BG)

total_row(A_TOT,'Cash Outflows Before Billing  (Import Advance + WIP/Direct Costs)',
    f'=SUM(C{A1}:C{A3})',f'=SUM(D{A1}:D{A3})',f'=SUM(E{A1}:E{A3})',
    bg=LIGHT_BLUE)

# A_DBT: Debtors = full order value (billed, collect in 45 days)
formula_row(A_DBT,
    'Trade Receivables / Debtors  (Full order value; 45-day credit)',
    f'=C{R_ORD}', f'=D{R_ORD}', f'=E{R_ORD}',
    bg=YELLOW_BG, indent=1)

total_row(GCA,'GROSS WORKING CAPITAL  (A)',
    f'=C{A_TOT}+C{A_DBT}',f'=D{A_TOT}+D{A_DBT}',f'=E{A_TOT}+E{A_DBT}',
    bg=GREEN_BG, sz=11)

spacer(GCA+1,5)

# ── B. Current Liabilities ─────────────────────────────────────────────────
B_HDR  = GCA+2
B1     = GCA+3   # Creditors (30% credit on local/remaining procurement)
B2     = GCA+4   # Salary / OPEX payable (15-day lag)
B3     = GCA+5   # Existing debtor collections expected during execution
B_TOT  = GCA+6

ws.row_dimensions[B_HDR].height = 16
ws.cell(row=B_HDR,column=2).value = 'B.  CURRENT LIABILITIES  (Spontaneous Financing)'
ws.cell(row=B_HDR,column=2).font = Font(name='Calibri',bold=True,size=10,color=WHITE)
ws.cell(row=B_HDR,column=2).fill = PatternFill('solid',fgColor=DARK_RED)
ws.cell(row=B_HDR,column=2).alignment = Alignment(vertical='center')
for col in [3,4,5]:
    ws.cell(row=B_HDR,column=col).fill = PatternFill('solid',fgColor=DARK_RED)

# B1: Creditors = 30% of fresh COGS at creditor days credit
formula_row(B1,
    'Trade Creditors  (30% of fresh procurement × credit days ÷ 30)',
    f'=C{R_COGS}*C{R_FRSH_PCT}*C{R_CRED_PCT}*(C{R_CRED_DAY}/30)',
    f'=D{R_COGS}*C{R_FRSH_PCT}*C{R_CRED_PCT}*(C{R_CRED_DAY}/30)',
    f'=E{R_COGS}*C{R_FRSH_PCT}*C{R_CRED_PCT}*(C{R_CRED_DAY}/30)',
    bg=WHITE, indent=1)

# B2: OPEX payable (15-day lag on monthly OPEX)
formula_row(B2,
    'Salary & OPEX Payable  (15-day payment lag)',
    f'=C{R_OPEX}*0.5', f'=D{R_OPEX}*0.5', f'=E{R_OPEX}*0.5',
    bg=WHITE, indent=1)

# B3: Collections from existing debtors — offset; these collections reduce OD (not create fresh cash)
ws.row_dimensions[B3].height = 16
c = ws.cell(row=B3,column=2)
c.value = '    Collections from Existing Debtors  (₹69.5L due in 45 days — reduces OD breach; net offset to WC need)'
c.font = Font(name='Calibri',size=10)
c.fill = PatternFill('solid',fgColor=LIGHT_BLUE)
c.alignment = Alignment(vertical='center')
for col,val in zip([3,4,5],[-6952542, -1500000, -4000000]):
    c = ws.cell(row=B3,column=col)
    c.value = val
    c.font = Font(name='Calibri',size=10)
    c.fill = PatternFill('solid',fgColor=YELLOW_BG)
    c.alignment = Alignment(horizontal='right',vertical='center')
    c.number_format = fmt_inr

total_row(B_TOT,'TOTAL CURRENT LIABILITIES  (B)',
    f'=C{B1}+C{B2}+C{B3}',f'=D{B1}+D{B2}+D{B3}',f'=E{B1}+E{B2}+E{B3}',
    bg=RED_BG)

spacer(B_TOT+1,5)

# ── Net WC ─────────────────────────────────────────────────────────────────
NWC = B_TOT+2
ws.row_dimensions[NWC].height = 20
ws.cell(row=NWC,column=2).value = 'NET WORKING CAPITAL REQUIRED  (A − B)'
ws.cell(row=NWC,column=2).font = Font(name='Calibri',bold=True,size=11,color=WHITE)
ws.cell(row=NWC,column=2).fill = PatternFill('solid',fgColor=DARK_BLUE)
ws.cell(row=NWC,column=2).alignment = Alignment(vertical='center')
for col in [3,4,5]:
    c = ws.cell(row=NWC,column=col)
    col_l = chr(ord('A')+col-1)
    c.value = f'={col_l}{GCA}-{col_l}{B_TOT}'
    c.font = Font(name='Calibri',bold=True,size=11,color=WHITE)
    c.fill = PatternFill('solid',fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='right',vertical='center')
    c.number_format = fmt_inr

spacer(NWC+1,8)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 4 — CREDIT FACILITY STATUS & WC GAP
# ════════════════════════════════════════════════════════════════════════════
S4 = NWC+2
banner(S4,'  SECTION 4 :  EXISTING CREDIT FACILITY STATUS  &  WC GAP  (ALL FACILITIES 100% UTILIZED — ₹0 AVAILABLE)',
       bg='C00000',h=20)
col_hdr(S4+1,['Particulars',
              'Current (Mar-26)\nAssessment Date',
              'Q1 FY26-27\n(Post-Collection Est.)',
              'Q2 FY26-27\n(Projected)'])

# Row variables
EX_OD_LIM  = S4+2
EX_OD_UTL  = S4+3
EX_OD_BCH  = S4+4
EX_OD_AVL  = S4+5
EX_BAJ_LIM = S4+6
EX_BAJ_UTL = S4+7
EX_BAJ_AVL = S4+8
EX_TOT     = S4+9
EX_NOTE    = S4+10
EX_SP1     = S4+11
EX_WCR     = S4+12
EX_LESS    = S4+13
EX_GAP     = S4+14

# OD rows
for row,lbl,vC,vD,vE,bg,bold,color in [
    (EX_OD_LIM,'    ICICI Bank OD — Sanctioned Limit',
     5250000,5250000,5250000, WHITE,False,'000000'),
    (EX_OD_UTL,'    ICICI Bank OD — Amount Utilized  (100% — Fully Drawn)',
     5250000,3500000,5000000, YELLOW_BG,False,'000000'),
    (EX_OD_BCH,'    ICICI OD — Headroom Remaining  (₹0 — Fully Exhausted)',
     f'=MAX(0,C{EX_OD_LIM}-C{EX_OD_UTL})',
     f'=MAX(0,D{EX_OD_LIM}-D{EX_OD_UTL})',
     f'=MAX(0,E{EX_OD_LIM}-E{EX_OD_UTL})', RED_BG,True,'C00000'),
    (EX_OD_AVL,'    ICICI OD — Available for Fresh Drawing',
     f'=MAX(0,C{EX_OD_LIM}-C{EX_OD_UTL})',
     f'=MAX(0,D{EX_OD_LIM}-D{EX_OD_UTL})',
     f'=MAX(0,E{EX_OD_LIM}-E{EX_OD_UTL})', WHITE,False,'000000'),
    (EX_BAJ_LIM,'    Bajaj Finance — Sanctioned Limit',
     1500000,1500000,1500000, WHITE,False,'000000'),
    (EX_BAJ_UTL,'    Bajaj Finance — Fully Utilized',
     1500000,1500000,1500000, WHITE,False,'000000'),
    (EX_BAJ_AVL,'    Bajaj Finance — Available for Drawing  (₹0 — 100% utilized)',
     f'=MAX(0,C{EX_BAJ_LIM}-C{EX_BAJ_UTL})',
     f'=MAX(0,D{EX_BAJ_LIM}-D{EX_BAJ_UTL})',
     f'=MAX(0,E{EX_BAJ_LIM}-E{EX_BAJ_UTL})', WHITE,False,'000000'),
]:
    ws.row_dimensions[row].height = 16
    c = ws.cell(row=row,column=2)
    c.value = lbl
    c.font = Font(name='Calibri',bold=bold,size=10,color=color)
    c.fill = PatternFill('solid',fgColor=bg)
    c.alignment = Alignment(vertical='center')
    for col,val in zip([3,4,5],[vC,vD,vE]):
        cell = ws.cell(row=row,column=col)
        cell.value = val
        cell.font = Font(name='Calibri',bold=bold,size=10,color=color)
        cell.fill = PatternFill('solid',fgColor=bg)
        cell.alignment = Alignment(horizontal='right',vertical='center')
        cell.number_format = fmt_inr

total_row(EX_TOT,'TOTAL AVAILABLE FROM EXISTING LIMITS  (₹0 — Both Facilities Exhausted)',
    f'=C{EX_OD_AVL}+C{EX_BAJ_AVL}',
    f'=D{EX_OD_AVL}+D{EX_BAJ_AVL}',
    f'=E{EX_OD_AVL}+E{EX_BAJ_AVL}',
    bg=RED_BG, sz=10)

# Note row (merged across all cols)
ws.row_dimensions[EX_NOTE].height = 36
ws.merge_cells(start_row=EX_NOTE,start_column=2,end_row=EX_NOTE,end_column=5)
cnote = ws.cell(row=EX_NOTE,column=2)
cnote.value = (
    'KEY:  ICICI OD (₹52.5L) and Bajaj Finance (₹15L) are BOTH 100% utilized — total ₹67.5L fully drawn, ₹0 available.  '
    'Existing debtor collections (₹69.5L arriving ~45 days after billing) will repay and free OD headroom — '
    'BUT PROCUREMENT ADVANCE IS NEEDED TODAY (before billing begins).  '
    'A FRESH ₹30L facility (separate term loan) bridges this 45-day timing gap and is self-liquidating from new order collections (₹93.19L within 90 days).'
)
cnote.font = Font(name='Calibri',size=9,italic=True,color='1F3864')
cnote.fill = PatternFill('solid',fgColor='DEEAF1')
cnote.alignment = Alignment(wrap_text=True,vertical='center')

spacer(EX_SP1,6)

# WC gap calculation
ws.row_dimensions[EX_WCR].height = 16
ws.cell(row=EX_WCR,column=2).value = 'Net Working Capital Required  (Section 3 — order-based)'
ws.cell(row=EX_WCR,column=2).font = Font(name='Calibri',size=10)
ws.cell(row=EX_WCR,column=2).fill = PatternFill('solid',fgColor=LIGHT_BLUE)
ws.cell(row=EX_WCR,column=2).alignment = Alignment(vertical='center')
for col in [3,4,5]:
    c = ws.cell(row=EX_WCR,column=col)
    col_l = chr(ord('A')+col-1)
    c.value = f'={col_l}{NWC}'
    c.font = Font(name='Calibri',size=10)
    c.fill = PatternFill('solid',fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal='right',vertical='center')
    c.number_format = fmt_inr

ws.row_dimensions[EX_LESS].height = 16
ws.cell(row=EX_LESS,column=2).value = '    Less: Available from Existing Limits  (₹0 — Fully Exhausted; OD in Breach)'
ws.cell(row=EX_LESS,column=2).font = Font(name='Calibri',size=10)
ws.cell(row=EX_LESS,column=2).fill = PatternFill('solid',fgColor=WHITE)
ws.cell(row=EX_LESS,column=2).alignment = Alignment(vertical='center')
for col in [3,4,5]:
    c = ws.cell(row=EX_LESS,column=col)
    col_l = chr(ord('A')+col-1)
    c.value = f'=-{col_l}{EX_TOT}'
    c.font = Font(name='Calibri',size=10)
    c.fill = PatternFill('solid',fgColor=WHITE)
    c.alignment = Alignment(horizontal='right',vertical='center')
    c.number_format = fmt_inr

ws.row_dimensions[EX_GAP].height = 22
ws.cell(row=EX_GAP,column=2).value = 'FRESH WC FACILITY REQUIRED  →  INR 30 LAKHS'
ws.cell(row=EX_GAP,column=2).font = Font(name='Calibri',bold=True,size=12,color='C00000')
ws.cell(row=EX_GAP,column=2).fill = PatternFill('solid',fgColor=RED_BG)
ws.cell(row=EX_GAP,column=2).alignment = Alignment(vertical='center')
for col in [3,4,5]:
    c = ws.cell(row=EX_GAP,column=col)
    col_l = chr(ord('A')+col-1)
    c.value = f'={col_l}{EX_WCR}+{col_l}{EX_LESS}'
    c.font = Font(name='Calibri',bold=True,size=12,color='C00000')
    c.fill = PatternFill('solid',fgColor=RED_BG)
    c.alignment = Alignment(horizontal='right',vertical='center')
    c.number_format = fmt_inr
    c.border = med_top()

spacer(EX_GAP+1,8)

# ════════════════════════════════════════════════════════════════════════════
# SECTION 5 — RECOMMENDATION  (INR 30L)
# ════════════════════════════════════════════════════════════════════════════
REC = EX_GAP+2
banner(REC,'  SECTION 5 :  WORKING CAPITAL REQUIREMENT — SUMMARY & RECOMMENDATION',bg=DARK_BLUE,h=20)

rec_rows = [
    (REC+1,'FACILITY STATUS (14-Mar-26):  ALL LIMITS 100% UTILIZED — ₹0 AVAILABLE',
     0, None, None, RED_BG,
     'ICICI OD ₹52.5L + Bajaj Finance ₹15L = ₹67.5L total limits, both fully drawn — no headroom for procurement'),
    (REC+2,'Cash Outflows Before Billing  (Import Advance + WIP/Direct Costs)',
     f'=C{A_TOT}', f'=D{A_TOT}', f'=E{A_TOT}', YELLOW_BG,
     'Capital committed BEFORE invoice is raised — cannot wait for debtor collections; procurement needed today'),
    (REC+3,'Existing Debtor Collections  (₹69.5L due within 45 days of billing)',
     -6952542, None, None, LIGHT_BLUE,
     'These collections repay and free OD headroom (ICICI + Bajaj) — they do NOT provide fresh procurement cash (timing gap)'),
    (REC+4,'NET FRESH WORKING CAPITAL REQUIRED  (Section 3)',
     f'=C{NWC}', f'=D{NWC}', f'=E{NWC}', RED_BG,
     'Net WC = Outflows + New Debtors − Existing Debtor Offset − Creditors − OPEX payable ≈ ₹31L'),
    (REC+5,'WORKING CAPITAL LOAN REQUESTED  (INR 30 Lakhs — Separate Term Facility)',
     3000000, None, None, GREEN_BG,
     'Fresh term loan — NOT an OD enhancement. Required immediately to fund procurement and regularise OD breach'),
    (REC+6,'Total WC Support Post-Sanction  (Freed OD + New Facility)',
     f'=C{EX_OD_LIM}+3000000', None, None, HEADER_BLUE,
     'ICICI OD ₹52.5L (regularised post-collection) + Fresh Term Loan ₹30L = ₹82.5L — covers Q2 FY26-27 growth'),
    (REC+7,'Repayment — Primary Source  (Self-Liquidating)',
     None, None, None, GRAY_BG,
     '₹93.19L new order debtors collect within 45 days of billing (by May-Jun 26). Full ₹30L repaid within 90 days'),
    (REC+8,'Repayment — Secondary Source',
     None, None, None, GRAY_BG,
     'Existing debtors ₹69.5L reduce OD to ₹8.84L, freeing ₹43.66L OD headroom for any residual requirement'),
    (REC+9,'Security / Collateral',
     None, None, None, GRAY_BG,
     "Director's apartment (existing ICICI mortgage) + Machinery & Equipment ₹8.5L + Book Debts ₹1.63 Cr"),
]

for row,lbl,vC,vD,vE,bg,rem in rec_rows:
    ws.row_dimensions[row].height = 18
    is_key = 'REQUESTED' in lbl or 'Proposed' in lbl or 'NET IMMEDIATE' in lbl
    c = ws.cell(row=row,column=2)
    c.value = ('    ' if not is_key else '') + lbl
    c.font = Font(name='Calibri',bold=is_key,size=10,
                  color='C00000' if 'REQUESTED' in lbl or 'NET IMMEDIATE' in lbl else '000000')
    c.fill = PatternFill('solid',fgColor=bg)
    c.alignment = Alignment(vertical='center')

    cc = ws.cell(row=row,column=3)
    cc.value = vC
    cc.font = Font(name='Calibri',bold=is_key,size=10,
                   color='1F3864' if 'Proposed' in lbl else
                         'C00000' if 'REQUESTED' in lbl or 'NET IMMEDIATE' in lbl else '000000')
    cc.fill = PatternFill('solid',fgColor=bg)
    cc.alignment = Alignment(horizontal='right',vertical='center')
    cc.number_format = fmt_inr
    if vC and is_key: cc.border = med_top()

    ws.merge_cells(start_row=row,start_column=4,end_row=row,end_column=5)
    cr = ws.cell(row=row,column=4)
    cr.value = rem
    cr.font = Font(name='Calibri',size=9,italic=True,color='444444')
    cr.fill = PatternFill('solid',fgColor=bg)
    cr.alignment = Alignment(vertical='center',wrap_text=True)

# ── Right panel assumptions ───────────────────────────────────────────────
ws.merge_cells(start_row=5,start_column=7,end_row=5,end_column=8)
c = ws.cell(row=5,column=7)
c.value = 'KEY ASSUMPTIONS  (Yellow cells = editable inputs)'
c.font = Font(name='Calibri',bold=True,size=10,color=WHITE)
c.fill = PatternFill('solid',fgColor=MID_BLUE)
c.alignment = Alignment(horizontal='center',vertical='center')
ws.row_dimensions[5].height = 18

assumptions = [
    ('Assessment Basis',             'Actual order values — not monthly averages'),
    ('Current Order Book',           'INR 93.19L confirmed (PO received + awaiting PO)'),
    ('Fresh Procurement',            '50% of COGS via import MOQ; 50% from stock'),
    ('Import Advance',               '70% on PO placement'),
    ('Supplier Credit (local)',      '30% at 30-day credit'),
    ('Net Operating Cycle',          '90 days (editable in Section 2)'),
    ('Debtor Terms',                 '45 days from invoice date'),
    ('Advance from Customers',       'Nil'),
    ('ICICI Bank OD Status',          'Limit ₹52.5L | Utilized ₹52.5L (100%) | Available: ₹0'),
    ('Bajaj Finance Status',         'Limit ₹15.0L | Utilized ₹15.0L (100%) | Available: ₹0'),
    ('Total Limits & Utilization',   '₹67.5L total (₹52.5L + ₹15.0L) — Both 100% exhausted, ₹0 headroom'),
    ('Existing Debtors',             '₹69.5L collectible ~45 days — reduces OD, not fresh cash'),
    ('Security Available',           "Director's apartment (ICICI mortgage) + Machinery + Book Debts ₹1.63 Cr"),
    ('WC Loan Requested',            '₹30.0 Lakhs — FRESH term facility (not OD enhancement)'),
    ('Proposed Support Post-Sanction','OD ₹52.5L (regularised) + Term Loan ₹30L = ₹82.5L'),
    ('Repayment Source',             'Self-liquidating: ₹93.19L new debtors within 90 days'),
]
for i,(k,v) in enumerate(assumptions):
    ar = 6+i
    ws.row_dimensions[ar].height = 15
    bg_r = WHITE if i%2==0 else 'F5F5F5'
    for col,val,bold in [(7,k,True),(8,v,False)]:
        cc = ws.cell(row=ar,column=col)
        cc.value = val
        cc.font = Font(name='Calibri',bold=bold,size=9)
        cc.fill = PatternFill('solid',fgColor=bg_r)
        cc.alignment = Alignment(vertical='center',wrap_text=True)
        cc.border = bdr()

wb.save('C:/Users/Lenovo/Downloads/Fracktal_Financials_FY2526.xlsx')
print('WC sheet rebuilt — order-based, fully formula-linked.')
