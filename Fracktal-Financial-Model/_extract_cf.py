"""Extract actual TB values (Apr-Feb) for cash flow construction."""
import openpyxl

f = r"D:\Suresh_AGENTS\Fracktal-Financial-Model\.tmp\Fracktal_MIS_Master_20260310_122834.xlsx"
wb = openpyxl.load_workbook(f)
ws = wb['TB']

# Month columns: B(2)=APR ... L(12)=FEB, M(13)=MAR
# We need APR to FEB = cols 2 to 12 (11 months)

# Revenue rows
rev_rows = [
    (5, "Sale of Products"),
    (6, "Sale of Service"),
    (7, "Export Sales"),
    (8, "Printsticks"),
]
other_income_rows = [
    (9, "Discount Received"),
    (10, "Interest Income"),
    (11, "Other Income"),
]

# COGS Purchases (cash items — exclude Opening/Closing Stock as non-cash)
purchase_rows = [
    (14, "PURCHASE of Raw Materials"),
    (15, "IMPORT of Raw Materials"),
    (16, "OTHER PURCHASES"),
]

# Direct expenses
direct_rows = [
    (19, "Salaries (Production)"),
    (20, "Overtime Pay"),
    (21, "Electricity - Factory"),
    (22, "Electricity - Manufacturing"),
    (23, "Freight Inward"),
    (24, "Loading & Unloading"),
    (25, "Discount Allowed"),
]

# OPEX line items (individual, not subtotals)
opex_rows = [
    (28, "Office & Admin Overheads"),
    (29, "Travelling Expense"),
    (30, "Rates & Taxes"),
    (31, "FX Gain/Loss"),
    (32, "Razorpay Charges"),
    (33, "Tender Fee"),
    (35, "Finance Cost"),
    (37, "Payroll Expenses"),
    (39, "Advertisement/Marketing"),
    (40, "Freight Outward"),
    (42, "R&D"),
    (43, "Professional Services"),
    (44, "Round Off"),
    (45, "Write Off / Write Back"),
]

# Bank + Cash balances (for opening/closing cash)
bank_rows = [
    (69, "Cash-in-Hand"),
    (70, "Bank Accounts"),
]

print("=" * 80)
print("REVENUE (Apr-Feb)")
total_rev = [0]*11
for rn, label in rev_rows + other_income_rows:
    vals = []
    for c in range(2, 13):
        v = ws.cell(rn, c).value
        if isinstance(v, str) and v.startswith('='):
            v = 0
        vals.append(float(v) if v and v != '-' else 0)
    total_rev = [total_rev[i] + vals[i] for i in range(11)]
    total = sum(vals)
    if total != 0:
        print(f"  Row {rn:3d} {label:35s} FY={total:>12,.0f}")
print(f"  {'TOTAL REVENUE':39s} FY={sum(total_rev):>12,.0f}")

print("\nPURCHASES (Apr-Feb)")
total_purch = [0]*11
for rn, label in purchase_rows:
    vals = []
    for c in range(2, 13):
        v = ws.cell(rn, c).value
        if isinstance(v, str) and v.startswith('='):
            v = 0
        vals.append(float(v) if v and v != '-' else 0)
    total_purch = [total_purch[i] + vals[i] for i in range(11)]
    total = sum(vals)
    if total != 0:
        print(f"  Row {rn:3d} {label:35s} FY={total:>12,.0f}")
print(f"  {'TOTAL PURCHASES':39s} FY={sum(total_purch):>12,.0f}")

print("\nDIRECT EXPENSES (Apr-Feb)")
total_direct = [0]*11
for rn, label in direct_rows:
    vals = []
    for c in range(2, 13):
        v = ws.cell(rn, c).value
        if isinstance(v, str) and v.startswith('='):
            v = 0
        vals.append(float(v) if v and v != '-' else 0)
    total_direct = [total_direct[i] + vals[i] for i in range(11)]
    total = sum(vals)
    if total != 0:
        print(f"  Row {rn:3d} {label:35s} FY={total:>12,.0f}")
print(f"  {'TOTAL DIRECT':39s} FY={sum(total_direct):>12,.0f}")

print("\nOPEX (Apr-Feb)")
total_opex = [0]*11
for rn, label in opex_rows:
    vals = []
    for c in range(2, 13):
        v = ws.cell(rn, c).value
        if isinstance(v, str) and v.startswith('='):
            v = 0
        vals.append(float(v) if v and v != '-' else 0)
    total_opex = [total_opex[i] + vals[i] for i in range(11)]
    total = sum(vals)
    if total != 0:
        print(f"  Row {rn:3d} {label:35s} FY={total:>12,.0f}")
print(f"  {'TOTAL OPEX':39s} FY={sum(total_opex):>12,.0f}")

print("\nCASH / BANK BALANCES (Apr-Feb)")
for rn, label in bank_rows:
    vals = []
    for c in range(2, 13):
        v = ws.cell(rn, c).value
        if isinstance(v, str) and v.startswith('='):
            v = 0
        vals.append(float(v) if v and v != '-' else 0)
    print(f"  Row {rn:3d} {label:15s}  APR={vals[0]:>12,.0f}  FEB={vals[10]:>12,.0f}")
    print(f"    Monthly: {['%.0f' % v for v in vals]}")

print("\nSUMMARY")
total_out = [total_purch[i] + total_direct[i] + total_opex[i] for i in range(11)]
net = [total_rev[i] - total_out[i] for i in range(11)]
print(f"  Total Inflows (11m):  {sum(total_rev):>12,.0f}")
print(f"  Total Outflows (11m): {sum(total_out):>12,.0f}")
print(f"  Net Cash Flow (11m):  {sum(net):>12,.0f}")

# CRM orders in hand
print("\nORDERS IN HAND (PO Received)")
ws3 = wb['CRM Data']
total_orders = 0
for r in range(5, ws3.max_row + 1):
    stage = ws3.cell(r, 2).value
    if stage and 'PO Received' in str(stage):
        name = ws3.cell(r, 1).value
        amt = ws3.cell(r, 4).value or 0
        total_orders += float(amt)
        print(f"  {name}: {float(amt):>12,.0f}")
print(f"  TOTAL ORDERS IN HAND: {total_orders:>12,.0f}")
print(f"  COGS at 50%:          {total_orders*0.5:>12,.0f}")

# Also get Awaiting PO
total_await = 0
for r in range(5, ws3.max_row + 1):
    stage = ws3.cell(r, 2).value
    if stage and 'Awaiting PO' in str(stage):
        amt = ws3.cell(r, 4).value or 0
        total_await += float(amt)
print(f"  AWAITING PO ORDERS:   {total_await:>12,.0f}")
print(f"  ALL CONFIRMED:        {total_orders+total_await:>12,.0f}")

wb.close()
