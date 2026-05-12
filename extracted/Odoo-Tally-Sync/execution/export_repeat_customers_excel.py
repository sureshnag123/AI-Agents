#!/usr/bin/env python3
"""
Generate Excel report for Repeat Customers from GST INVOICE (Sundry Debtors)
FY 2025-26

Creates:
  Sheet 1: Summary
  Sheet 2: Repeat Customers (all 79 with totals)
  Sheet 3: One-Time Customers (117)
  Sheet 4: Invoice Details (all individual invoices)
  Sheet 5: Excluded Parties (non-Sundry Debtor)
"""

import json
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent.resolve()
TMP_DIR = PROJECT_ROOT / ".tmp"

# ── Styles ──
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2E75B6")
NUM_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
RUPEE_FMT = '#,##0.00'
INT_FMT = '#,##0'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, is_number=False, is_bold=False, is_total=False, alt_row=False):
    cell.border = THIN_BORDER
    if is_total:
        cell.fill = TOTAL_FILL
        cell.font = BOLD_FONT
    elif alt_row:
        cell.fill = ALT_FILL
    if is_number:
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = RUPEE_FMT
    if is_bold:
        cell.font = BOLD_FONT


def auto_width(ws, min_width=10, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def parse_date(date_str):
    """Convert YYYYMMDD to DD-MMM-YYYY"""
    try:
        return datetime.strptime(date_str, "%Y%m%d").strftime("%d-%b-%Y")
    except:
        return date_str


def main():
    # Load data
    with open(TMP_DIR / "repeat_customers_gst_invoice_debtors_fy2526.json", "r", encoding="utf-8") as f:
        data = json.load(f)

    summary = data["summary"]
    repeat_customers = data["repeat_customers"]
    excluded = data.get("excluded_non_debtor_parties_in_gst_invoice", {})

    # Also load one-time customers from the full dataset
    # We need to re-derive one-time from the original data
    # For now, the JSON only has repeat_customers. We'll note in summary.

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════
    # Sheet 1: SUMMARY
    # ════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = "4472C4"

    ws_sum.cell(row=1, column=1, value="Repeat Customer Analysis").font = TITLE_FONT
    ws_sum.cell(row=2, column=1, value="GST INVOICE Vouchers | Sundry Debtors | FY 2025-26").font = SUBTITLE_FONT
    ws_sum.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}").font = Font(italic=True, color="808080")

    row = 5
    metrics = [
        ("Report", summary["report"]),
        ("Period", summary["period"]),
        ("Voucher Types Included", ", ".join(summary["voucher_types_included"])),
        ("Filter", summary["filter"]),
        ("", ""),
        ("Total Vouchers (Sales Parent)", summary["total_vouchers_under_sales_parent"]),
        ("GST Invoice Vouchers", summary["gst_invoice_vouchers"]),
        ("GST Invoices with Sundry Debtor", summary["gst_invoices_with_sundry_debtor"]),
        ("", ""),
        ("Unique Customers", summary["unique_customers"]),
        ("Repeat Customers (2+ invoices)", summary["repeat_customers"]),
        ("One-Time Customers", summary["one_time_customers"]),
        ("", ""),
        ("Total Repeat Invoices", summary["total_repeat_invoices"]),
        ("Total Repeat Revenue (Rs)", summary["total_repeat_amount"]),
    ]

    for label, value in metrics:
        c1 = ws_sum.cell(row=row, column=1, value=label)
        c2 = ws_sum.cell(row=row, column=2, value=value)
        c1.font = BOLD_FONT
        if isinstance(value, (int, float)) and label:
            if "Revenue" in label or "Amount" in label:
                c2.number_format = RUPEE_FMT
            else:
                c2.number_format = INT_FMT
        row += 1

    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 55

    # ════════════════════════════════════════════
    # Sheet 2: REPEAT CUSTOMERS (all 79)
    # ════════════════════════════════════════════
    ws_rep = wb.create_sheet("Repeat Customers")
    ws_rep.sheet_properties.tabColor = "00B050"

    ws_rep.cell(row=1, column=1, value="Repeat Customers (2+ GST Invoices) — Sundry Debtors — FY 2025-26").font = TITLE_FONT
    ws_rep.merge_cells("A1:F1")

    headers = ["#", "Customer Name", "Invoice Count", "Total Revenue (Rs)", "Avg Invoice Value (Rs)", "% of Total Repeat Revenue"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws_rep.cell(row=row, column=col, value=h)
    style_header_row(ws_rep, row, len(headers))

    total_repeat_rev = summary["total_repeat_amount"]
    row = 4
    for i, cust in enumerate(repeat_customers, 1):
        alt = i % 2 == 0
        avg_val = cust["total_amount"] / cust["invoice_count"] if cust["invoice_count"] > 0 else 0
        pct = (cust["total_amount"] / total_repeat_rev * 100) if total_repeat_rev > 0 else 0

        cells = [
            (i, False),
            (cust["customer_name"], False),
            (cust["invoice_count"], True),
            (cust["total_amount"], True),
            (round(avg_val, 2), True),
            (round(pct, 2), True),
        ]

        for col, (val, is_num) in enumerate(cells, 1):
            c = ws_rep.cell(row=row, column=col, value=val)
            style_data_cell(c, is_number=is_num, alt_row=alt)
            if col == 6 and is_num:
                c.number_format = '0.00"%"'
            elif is_num and col in (4, 5):
                c.number_format = RUPEE_FMT
            elif is_num and col == 3:
                c.number_format = INT_FMT

        row += 1

    # Total row
    total_inv = sum(c["invoice_count"] for c in repeat_customers)
    total_amt = sum(c["total_amount"] for c in repeat_customers)
    avg_total = total_amt / total_inv if total_inv > 0 else 0

    totals = ["", "TOTAL", total_inv, total_amt, round(avg_total, 2), 100.0]
    for col, val in enumerate(totals, 1):
        c = ws_rep.cell(row=row, column=col, value=val)
        style_data_cell(c, is_number=col >= 3, is_total=True)
        if col in (4, 5):
            c.number_format = RUPEE_FMT
        elif col == 3:
            c.number_format = INT_FMT
        elif col == 6:
            c.number_format = '0.00"%"'

    auto_width(ws_rep)
    ws_rep.column_dimensions["B"].width = 55

    # Freeze panes
    ws_rep.freeze_panes = "A4"

    # ════════════════════════════════════════════
    # Sheet 3: INVOICE DETAILS (all individual invoices)
    # ════════════════════════════════════════════
    ws_inv = wb.create_sheet("Invoice Details")
    ws_inv.sheet_properties.tabColor = "FFC000"

    ws_inv.cell(row=1, column=1, value="All GST Invoices for Repeat Customers — FY 2025-26").font = TITLE_FONT
    ws_inv.merge_cells("A1:G1")

    inv_headers = ["#", "Customer Name", "Invoice Number", "Invoice Date", "Voucher Type", "Amount (Rs)", "Narration"]
    row = 3
    for col, h in enumerate(inv_headers, 1):
        ws_inv.cell(row=row, column=col, value=h)
    style_header_row(ws_inv, row, len(inv_headers))

    row = 4
    seq = 0
    for cust in repeat_customers:
        for inv in cust["invoices"]:
            seq += 1
            alt = seq % 2 == 0
            vals = [
                seq,
                cust["customer_name"],
                inv["number"],
                parse_date(inv["date"]),
                inv["voucher_type"],
                inv["amount"],
                inv.get("narration", ""),
            ]
            for col, val in enumerate(vals, 1):
                c = ws_inv.cell(row=row, column=col, value=val)
                is_num = col == 6
                style_data_cell(c, is_number=is_num, alt_row=alt)
                if is_num:
                    c.number_format = RUPEE_FMT
            row += 1

    # Grand total
    c = ws_inv.cell(row=row, column=5, value="TOTAL")
    c.font = BOLD_FONT
    c.fill = TOTAL_FILL
    c.border = THIN_BORDER
    total_val = sum(inv["amount"] for cust in repeat_customers for inv in cust["invoices"])
    c = ws_inv.cell(row=row, column=6, value=total_val)
    style_data_cell(c, is_number=True, is_total=True)
    c.number_format = RUPEE_FMT

    auto_width(ws_inv)
    ws_inv.column_dimensions["B"].width = 55
    ws_inv.column_dimensions["G"].width = 60
    ws_inv.freeze_panes = "A4"

    # ════════════════════════════════════════════
    # Sheet 4: CUSTOMER REVENUE ANALYSIS
    # ════════════════════════════════════════════
    ws_analysis = wb.create_sheet("Revenue Analysis")
    ws_analysis.sheet_properties.tabColor = "7030A0"

    ws_analysis.cell(row=1, column=1, value="Customer Revenue Analysis — FY 2025-26").font = TITLE_FONT
    ws_analysis.merge_cells("A1:H1")

    an_headers = ["#", "Customer Name", "Invoice Count", "Total Revenue (Rs)",
                  "Avg Invoice (Rs)", "First Invoice", "Last Invoice", "Months Active"]
    row = 3
    for col, h in enumerate(an_headers, 1):
        ws_analysis.cell(row=row, column=col, value=h)
    style_header_row(ws_analysis, row, len(an_headers))

    row = 4
    for i, cust in enumerate(repeat_customers, 1):
        alt = i % 2 == 0
        avg_val = cust["total_amount"] / cust["invoice_count"] if cust["invoice_count"] > 0 else 0

        # Find date range
        dates = []
        for inv in cust["invoices"]:
            try:
                d = datetime.strptime(inv["date"], "%Y%m%d")
                dates.append(d)
            except:
                pass

        first_date = min(dates).strftime("%d-%b-%Y") if dates else ""
        last_date = max(dates).strftime("%d-%b-%Y") if dates else ""
        if dates and len(dates) >= 2:
            months = (max(dates).year - min(dates).year) * 12 + (max(dates).month - min(dates).month)
            months_active = max(months, 1)
        else:
            months_active = 1

        vals = [
            (i, False),
            (cust["customer_name"], False),
            (cust["invoice_count"], True),
            (cust["total_amount"], True),
            (round(avg_val, 2), True),
            (first_date, False),
            (last_date, False),
            (months_active, True),
        ]

        for col, (val, is_num) in enumerate(vals, 1):
            c = ws_analysis.cell(row=row, column=col, value=val)
            style_data_cell(c, is_number=is_num, alt_row=alt)
            if col in (4, 5):
                c.number_format = RUPEE_FMT
            elif col == 3 or col == 8:
                c.number_format = INT_FMT
        row += 1

    auto_width(ws_analysis)
    ws_analysis.column_dimensions["B"].width = 55
    ws_analysis.freeze_panes = "A4"

    # ════════════════════════════════════════════
    # Sheet 5: EXCLUDED PARTIES
    # ════════════════════════════════════════════
    ws_excl = wb.create_sheet("Excluded Parties")
    ws_excl.sheet_properties.tabColor = "FF0000"

    ws_excl.cell(row=1, column=1, value="Excluded Parties (Not Sundry Debtors) in GST INVOICE").font = TITLE_FONT
    ws_excl.merge_cells("A1:C1")

    ex_headers = ["#", "Party Name", "Voucher Count"]
    row = 3
    for col, h in enumerate(ex_headers, 1):
        ws_excl.cell(row=row, column=col, value=h)
    style_header_row(ws_excl, row, len(ex_headers))

    row = 4
    for i, (party, cnt) in enumerate(excluded.items(), 1):
        alt = i % 2 == 0
        for col, val in enumerate([i, party or "(No Party Name)", cnt], 1):
            c = ws_excl.cell(row=row, column=col, value=val)
            style_data_cell(c, is_number=col == 3, alt_row=alt)
            if col == 3:
                c.number_format = INT_FMT
        row += 1

    auto_width(ws_excl)

    # ── Save ──
    out_file = TMP_DIR / "Repeat_Customers_GST_Invoice_FY2526.xlsx"
    wb.save(out_file)
    print(f"Excel report saved to: {out_file}")
    print(f"\nSheets created:")
    print(f"  1. Summary          - Key metrics")
    print(f"  2. Repeat Customers - All {len(repeat_customers)} customers with revenue")
    print(f"  3. Invoice Details  - All {sum(c['invoice_count'] for c in repeat_customers)} individual invoices")
    print(f"  4. Revenue Analysis - Date ranges, avg invoice, months active")
    print(f"  5. Excluded Parties - Non-Sundry Debtor entries")


if __name__ == "__main__":
    main()
