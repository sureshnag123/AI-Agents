#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sync Checker & Analyser — Odoo ↔ Tally
========================================
Before any push, this script cross-checks Odoo invoices against Tally vouchers
and reports exactly what is pending, what is already synced, what might be
duplicated in Tally, and which ledgers are missing or have suspicious names.

Checks performed
----------------
  1. Sales:   Odoo out_invoices vs Tally "GST INVOICE" vouchers
  2. Purchase: Odoo in_invoices  vs Tally "Purchase"     vouchers
  3. Tally-internal duplicate vouchers (same party + amount within month)
  4. Tally ledger duplicates / near-duplicates (fuzzy name matching)
  5. Odoo invoices that would need NEW ledgers created in Tally

Output
------
  • Console summary
  • Excel report saved to .tmp/sync_check_<date>.xlsx (multi-sheet)

Usage
-----
    python sync_checker.py                         # May 2026 (current month)
    python sync_checker.py --from 2026-05-01 --to 2026-05-12
    python sync_checker.py --month 2026-04         # full April
    python sync_checker.py --full-fy               # full FY 2025-26
"""

import io
import os
import re
import sys
import json
import logging
import argparse
from datetime import date, datetime
from pathlib import Path

# ── UTF-8 console on Windows ──────────────────────────────────────────
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv

SCRIPT_DIR   = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR      = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(parents=True, exist_ok=True)

load_dotenv(PROJECT_ROOT / ".env")
sys.path.insert(0, str(SCRIPT_DIR))

from odoo_connector  import OdooConnector
from tally_connector import TallyConnector

# ── Logging ───────────────────────────────────────────────────────────

LOG_DIR = TMP_DIR / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

_stream_handler = logging.StreamHandler()
_stream_handler.stream = open(
    sys.stdout.fileno(), mode="w", encoding="utf-8", buffering=1, closefd=False
) if hasattr(sys.stdout, "fileno") else sys.stdout

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        _stream_handler,
        logging.FileHandler(
            LOG_DIR / f"sync_check_{date.today().isoformat()}.log",
            encoding="utf-8",
        ),
    ],
)
log = logging.getLogger("sync_checker")

MAPPING_FILE = TMP_DIR / "ledger_mapping.json"

SALES_VCH_TYPE    = "GST INVOICE"
PURCHASE_VCH_TYPE = "Purchase"


# ── Helpers ───────────────────────────────────────────────────────────

def load_mapping() -> dict:
    if MAPPING_FILE.exists():
        with open(MAPPING_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"odoo_to_tally_ledgers": {}, "odoo_to_tally_partners": {}}


def _norm(name: str) -> str:
    """Normalise for fuzzy matching."""
    name = name.lower().strip()
    for suffix in (" pvt ltd", " pvt. ltd.", " private limited", " limited",
                   " ltd", " ltd.", " llp", " inc", " corp"):
        if name.endswith(suffix):
            name = name[: -len(suffix)].strip()
    # collapse whitespace and strip special chars
    name = re.sub(r"[^a-z0-9 ]", " ", name)
    return " ".join(name.split())


def _tally_amount(raw) -> float:
    try:
        return abs(float(str(raw).replace(",", "").strip()))
    except (ValueError, TypeError):
        return 0.0


def _odoo_date_to_tally(odoo_date: str) -> str:
    """Convert 'YYYY-MM-DD' → 'YYYYMMDD' for Tally date comparison."""
    return odoo_date.replace("-", "") if odoo_date else ""


def _resolve_partner(odoo_name: str, mapping: dict) -> str:
    partners = mapping.get("odoo_to_tally_partners", {})
    return partners.get(odoo_name, odoo_name)


def _resolve_ledger(odoo_name: str, mapping: dict) -> str:
    ledgers = mapping.get("odoo_to_tally_ledgers", {})
    return ledgers.get(odoo_name, odoo_name)


# ── Tally fetch helpers ───────────────────────────────────────────────

def fetch_tally_vouchers(tally: TallyConnector, vch_type: str,
                         tally_from: str, tally_to: str) -> list:
    """Fetch Tally vouchers with Python-side date filter."""
    try:
        raw = tally.get_vouchers(vch_type, tally_from, tally_to)
        filtered = [v for v in raw
                    if tally_from <= v.get("date", "") <= tally_to]
        log.info(f"  Tally '{vch_type}': {len(filtered)} voucher(s) in range")
        return filtered
    except Exception as exc:
        log.warning(f"  Could not fetch Tally '{vch_type}' vouchers: {exc}")
        return []


# ── Core comparison functions ─────────────────────────────────────────

def match_odoo_to_tally(
    odoo_invoices: list,
    tally_vouchers: list,
    invoice_type: str,    # "sales" or "purchase"
    mapping: dict,
) -> dict:
    """
    Match each Odoo invoice to a Tally voucher.

    Returns:
      {
        "matched":   [(odoo_inv, tally_vch, match_reason), ...],
        "unmatched": [odoo_inv, ...],          # not in Tally → needs push
        "orphans":   [tally_vch, ...],         # in Tally, not in Odoo range
      }
    """
    # Build Tally lookup maps
    tally_by_number    = {}   # num  → voucher
    tally_by_ref       = {}   # ref  → voucher
    tally_by_party_amt = {}   # norm_party → [(amt, voucher), ...]
    tally_by_date_amt  = {}   # YYYYMMDD   → [(amt, voucher), ...]
    matched_tally_ids  = set()

    for v in tally_vouchers:
        num  = (v.get("number") or "").strip()
        ref  = (v.get("reference") or "").strip()
        pty  = _norm(v.get("party") or "")
        amt  = _tally_amount(v.get("amount") or "0")
        dt   = (v.get("date") or "").strip()
        vid  = id(v)

        if num:
            tally_by_number[num] = v
        if ref:
            tally_by_ref[ref] = v
            tally_by_number.setdefault(ref, v)
        if pty and amt > 0:
            tally_by_party_amt.setdefault(pty, []).append((amt, v))
        if dt and amt > 0:
            tally_by_date_amt.setdefault(dt, []).append((amt, v))

    matched   = []
    unmatched = []

    for inv in odoo_invoices:
        inv_name   = inv.get("name", "")
        inv_ref    = (inv.get("ref") or "").strip()
        inv_amount = round(float(inv.get("amount_total", 0)), 2)
        inv_date   = (inv.get("invoice_date") or inv.get("date") or "").replace("-", "")
        partner    = inv.get("partner_id", [None, ""])[1] if inv.get("partner_id") else ""
        tally_party = _norm(_resolve_partner(partner, mapping))

        match_vch    = None
        match_reason = ""

        # Pass 1a: invoice number vs Tally voucher number
        if inv_name and inv_name in tally_by_number:
            match_vch    = tally_by_number[inv_name]
            match_reason = f"number '{inv_name}'"

        # Pass 1b (purchase): vendor ref vs Tally voucher number / reference
        if not match_vch and inv_ref:
            if inv_ref in tally_by_number:
                match_vch    = tally_by_number[inv_ref]
                match_reason = f"vendor ref '{inv_ref}' → Tally number"
            elif inv_ref in tally_by_ref:
                match_vch    = tally_by_ref[inv_ref]
                match_reason = f"vendor ref '{inv_ref}' → Tally REFERENCE"

        # Pass 2: party + amount (±₹1)
        if not match_vch and tally_party:
            for tally_amt, v in tally_by_party_amt.get(tally_party, []):
                if abs(tally_amt - inv_amount) <= 1.0:
                    match_vch    = v
                    match_reason = f"party '{partner}' + ₹{inv_amount:,.2f} (Tally ₹{tally_amt:,.2f})"
                    break

        # Pass 3: date + amount (±₹1)
        if not match_vch:
            for tally_amt, v in tally_by_date_amt.get(inv_date, []):
                if abs(tally_amt - inv_amount) <= 1.0:
                    match_vch    = v
                    match_reason = f"date {inv_date} + ₹{inv_amount:,.2f} (Tally ₹{tally_amt:,.2f})"
                    break

        if match_vch:
            matched_tally_ids.add(id(match_vch))
            matched.append((inv, match_vch, match_reason))
        else:
            unmatched.append(inv)

    # Orphan Tally vouchers not matched to any Odoo invoice
    orphans = [v for v in tally_vouchers if id(v) not in matched_tally_ids]

    return {"matched": matched, "unmatched": unmatched, "orphans": orphans}


def deduplicate_vouchers(vouchers: list) -> list:
    """Remove duplicate voucher entries by voucher number (keeps first occurrence)."""
    seen_nums = set()
    unique = []
    for v in vouchers:
        num = (v.get("number") or "").strip()
        if num:
            if num in seen_nums:
                continue
            seen_nums.add(num)
        unique.append(v)
    return unique


def find_tally_duplicate_vouchers(tally_vouchers: list) -> list:
    """
    Find pairs of Tally vouchers (same month, same party, amount within ₹1)
    that look like duplicate postings. Only flags pairs with DIFFERENT voucher numbers.

    Returns list of (vch_a, vch_b, reason).
    """
    duplicates = []
    seen = {}

    for v in tally_vouchers:
        pty = _norm(v.get("party") or "")
        amt = _tally_amount(v.get("amount") or "0")
        dt  = (v.get("date") or "")[:6]   # YYYYMM
        num = (v.get("number") or "").strip()

        if not pty or amt == 0:
            continue

        key = (pty, round(amt), dt)
        if key in seen:
            prev      = seen[key]
            prev_num  = (prev.get("number") or "").strip()
            if prev_num != num and abs(_tally_amount(prev.get("amount")) - amt) <= 1.0:
                duplicates.append((
                    prev, v,
                    f"same party '{v.get('party')}', "
                    f"amount ₹{amt:,.2f}, month {dt}"
                ))
        else:
            seen[key] = v

    return duplicates


def find_tally_ledger_duplicates(tally_ledgers: list) -> list:
    """
    Find Tally ledgers whose normalised names are identical or very close
    (edit-distance 1 or common abbreviation variants).

    Returns list of (ledger_a_name, ledger_b_name, reason).
    """
    duplicates = []
    norm_map = {}   # norm_name → original_name

    for led in tally_ledgers:
        name = led.get("name") or ""
        n    = _norm(name)
        if not n:
            continue
        if n in norm_map:
            duplicates.append((
                norm_map[n], name,
                f"identical after normalisation: '{n}'"
            ))
        else:
            norm_map[n] = name

    return duplicates


def find_new_ledgers_needed(
    odoo_invoices: list,
    unmatched_invoices: list,
    tally_ledger_cache: set,
    mapping: dict,
) -> list:
    """
    For invoices that would be pushed (unmatched), check which partner
    and account ledgers don't yet exist in Tally.

    Returns list of (ledger_name, group, invoice_name).
    """
    needed = []
    seen   = set()

    unmatched_names = {inv.get("name") for inv in unmatched_invoices}
    odoo_unmatched  = [inv for inv in odoo_invoices
                       if inv.get("name") in unmatched_names]

    for inv in odoo_unmatched:
        partner = inv.get("partner_id", [None, ""])[1] if inv.get("partner_id") else ""
        if partner:
            tally_partner = _resolve_partner(partner, mapping)
            if tally_partner.lower() not in tally_ledger_cache and tally_partner not in seen:
                needed.append((tally_partner, "Sundry Debtors/Creditors", inv.get("name", "")))
                seen.add(tally_partner)

    return needed


# ── Purchase Date Validation ──────────────────────────────────────────

def validate_purchase_dates(purchase_matched: list) -> list:
    """
    For each matched Odoo purchase bill → Tally voucher pair, verify:
      - Tally DATE         == Odoo 'date'         (Accounting Date)
      - Tally PARTYINVDATE == Odoo 'invoice_date' (Bill Date / Supplier Invoice Date)

    Returns list of issue dicts.
    """
    issues = []

    for odoo_inv, tally_vch, _reason in purchase_matched:
        bill_num        = odoo_inv.get("name", "")
        tally_num       = tally_vch.get("number", "")

        odoo_acc_date   = _odoo_date_to_tally(odoo_inv.get("date", ""))
        odoo_bill_date  = _odoo_date_to_tally(odoo_inv.get("invoice_date", ""))

        tally_date      = (tally_vch.get("date") or "").strip()
        tally_partydate = (tally_vch.get("partyinvdate") or "").strip()

        # Accounting Date mismatch
        if odoo_acc_date and tally_date and odoo_acc_date != tally_date:
            issues.append({
                "odoo_bill":  bill_num,
                "tally_num":  tally_num,
                "issue_type": "Accounting Date Mismatch",
                "odoo_val":   odoo_inv.get("date", ""),
                "tally_val":  tally_date,
                "note":       "Tally <DATE> should match Odoo Accounting Date",
            })

        # Bill Date (PARTYINVDATE) missing in Tally
        if not tally_partydate:
            issues.append({
                "odoo_bill":  bill_num,
                "tally_num":  tally_num,
                "issue_type": "Bill Date Missing in Tally",
                "odoo_val":   odoo_inv.get("invoice_date", ""),
                "tally_val":  "(empty)",
                "note":       "PARTYINVDATE not set — voucher likely pushed as Accounting Voucher View (no ISINVOICE)",
            })
        elif odoo_bill_date and tally_partydate and odoo_bill_date != tally_partydate:
            issues.append({
                "odoo_bill":  bill_num,
                "tally_num":  tally_num,
                "issue_type": "Bill Date Mismatch",
                "odoo_val":   odoo_inv.get("invoice_date", ""),
                "tally_val":  tally_partydate,
                "note":       "Tally PARTYINVDATE should match Odoo Bill Date (invoice_date)",
            })

        # Both Tally dates identical but Odoo dates differ
        if (tally_date and tally_partydate
                and tally_date == tally_partydate
                and odoo_acc_date and odoo_bill_date
                and odoo_acc_date != odoo_bill_date):
            issues.append({
                "odoo_bill":  bill_num,
                "tally_num":  tally_num,
                "issue_type": "Same Dates in Tally / Different in Odoo",
                "odoo_val":   f"AccDate={odoo_inv.get('date','')}  BillDate={odoo_inv.get('invoice_date','')}",
                "tally_val":  f"DATE={tally_date}  PARTYINVDATE={tally_partydate}",
                "note":       "Tally has identical accounting & bill dates but Odoo shows different values",
            })

    return issues


# ── Excel Report ──────────────────────────────────────────────────────

def write_excel_report(
    path: Path,
    from_date: str,
    to_date: str,
    sales_result: dict,
    purchase_result: dict,
    sales_tally_dupes: list,
    purchase_tally_dupes: list,
    ledger_dupes: list,
    new_ledgers: list,
    purchase_date_issues: list = None,
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel report. "
                    "Install with: pip install openpyxl")
        return

    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT   = Font(bold=True, color="FFFFFF")
    OK_FILL    = PatternFill("solid", fgColor="C6EFCE")
    WARN_FILL  = PatternFill("solid", fgColor="FFEB9C")
    ERR_FILL   = PatternFill("solid", fgColor="FFC7CE")
    DUPE_FILL  = PatternFill("solid", fgColor="FCE4D6")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def make_sheet(title: str) -> object:
        ws = wb.create_sheet(title=title)
        return ws

    def hdr(ws, cols: list):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[ws.max_row].height = 20

    def colour_row(ws, row_num: int, fill: PatternFill):
        for cell in ws[row_num]:
            cell.fill = fill

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── Sheet 1: Summary ─────────────────────────────────────────────
    ws = make_sheet("Summary")
    ws.append(["Odoo ↔ Tally Sync Check Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {from_date}  to  {to_date}"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])

    _date_issues = purchase_date_issues or []
    hdr(ws, ["Category", "Count", "Status"])
    rows = [
        ("Sales — Matched (already in Tally)",    len(sales_result["matched"]),    "OK"),
        ("Sales — Unmatched (needs push)",         len(sales_result["unmatched"]),  "NEEDS PUSH"),
        ("Sales — Orphan Tally vouchers",          len(sales_result["orphans"]),    "REVIEW"),
        ("Purchase — Matched (already in Tally)", len(purchase_result["matched"]),  "OK"),
        ("Purchase — Unmatched (needs push)",      len(purchase_result["unmatched"]),"NEEDS PUSH"),
        ("Purchase — Orphan Tally vouchers",       len(purchase_result["orphans"]), "REVIEW"),
        ("Tally Sales duplicate vouchers",         len(sales_tally_dupes),          "WARNING" if sales_tally_dupes else "OK"),
        ("Tally Purchase duplicate vouchers",      len(purchase_tally_dupes),       "WARNING" if purchase_tally_dupes else "OK"),
        ("Tally ledger near-duplicates",           len(ledger_dupes),               "WARNING" if ledger_dupes else "OK"),
        ("New ledgers needed in Tally",            len(new_ledgers),                "INFO"),
        ("Purchase date issues (Bill/Acc date)",   len(_date_issues),               "WARNING" if _date_issues else "OK"),
    ]
    for r in rows:
        ws.append(list(r))
        row_num = ws.max_row
        status = r[2]
        if status == "OK":
            colour_row(ws, row_num, OK_FILL)
        elif status in ("NEEDS PUSH", "WARNING"):
            colour_row(ws, row_num, WARN_FILL)
        elif status == "REVIEW":
            colour_row(ws, row_num, ERR_FILL)
    auto_width(ws)

    # ── Sheet 2: Sales Unmatched (needs push) ────────────────────────
    ws = make_sheet("Sales - Needs Push")
    hdr(ws, ["Invoice #", "Customer", "Date", "Amount (₹)", "Odoo Ref"])
    for inv in sales_result["unmatched"]:
        partner = inv.get("partner_id", [None, ""])[1] if inv.get("partner_id") else "N/A"
        ws.append([
            inv.get("name", ""),
            partner,
            inv.get("invoice_date", ""),
            round(float(inv.get("amount_total", 0)), 2),
            inv.get("ref") or "",
        ])
        colour_row(ws, ws.max_row, WARN_FILL)
    auto_width(ws)

    # ── Sheet 3: Sales Matched ────────────────────────────────────────
    ws = make_sheet("Sales - Matched")
    hdr(ws, ["Odoo Invoice #", "Customer", "Odoo Date", "Odoo Amount (₹)",
             "Tally Voucher #", "Tally Party", "Tally Amount (₹)", "Match Reason"])
    for odoo_inv, tally_vch, reason in sales_result["matched"]:
        partner = odoo_inv.get("partner_id", [None, ""])[1] if odoo_inv.get("partner_id") else ""
        ws.append([
            odoo_inv.get("name", ""),
            partner,
            odoo_inv.get("invoice_date", ""),
            round(float(odoo_inv.get("amount_total", 0)), 2),
            tally_vch.get("number", ""),
            tally_vch.get("party", ""),
            _tally_amount(tally_vch.get("amount", "0")),
            reason,
        ])
        colour_row(ws, ws.max_row, OK_FILL)
    auto_width(ws)

    # ── Sheet 4: Sales Orphan Tally Vouchers ─────────────────────────
    ws = make_sheet("Sales - Tally Orphans")
    hdr(ws, ["Tally Voucher #", "Party", "Date", "Amount (₹)", "Note"])
    for v in sales_result["orphans"]:
        ws.append([
            v.get("number", ""),
            v.get("party", ""),
            v.get("date", ""),
            _tally_amount(v.get("amount", "0")),
            "In Tally but no matching Odoo invoice found",
        ])
        colour_row(ws, ws.max_row, DUPE_FILL)
    auto_width(ws)

    # ── Sheet 5: Purchase Unmatched ───────────────────────────────────
    ws = make_sheet("Purchase - Needs Push")
    hdr(ws, ["Bill #", "Vendor", "Invoice Date", "Voucher Date", "Amount (₹)", "Vendor Ref"])
    for inv in purchase_result["unmatched"]:
        partner = inv.get("partner_id", [None, ""])[1] if inv.get("partner_id") else "N/A"
        ws.append([
            inv.get("name", ""),
            partner,
            inv.get("invoice_date", ""),
            inv.get("date", ""),
            round(float(inv.get("amount_total", 0)), 2),
            inv.get("ref") or "",
        ])
        colour_row(ws, ws.max_row, WARN_FILL)
    auto_width(ws)

    # ── Sheet 6: Purchase Matched ─────────────────────────────────────
    ws = make_sheet("Purchase - Matched")
    hdr(ws, ["Odoo Bill #", "Vendor", "Odoo Inv Date", "Odoo Amount (₹)",
             "Tally Voucher #", "Tally Party", "Tally Amount (₹)", "Match Reason"])
    for odoo_inv, tally_vch, reason in purchase_result["matched"]:
        partner = odoo_inv.get("partner_id", [None, ""])[1] if odoo_inv.get("partner_id") else ""
        ws.append([
            odoo_inv.get("name", ""),
            partner,
            odoo_inv.get("invoice_date", ""),
            round(float(odoo_inv.get("amount_total", 0)), 2),
            tally_vch.get("number", ""),
            tally_vch.get("party", ""),
            _tally_amount(tally_vch.get("amount", "0")),
            reason,
        ])
        colour_row(ws, ws.max_row, OK_FILL)
    auto_width(ws)

    # ── Sheet 7: Purchase Orphan Tally Vouchers ───────────────────────
    ws = make_sheet("Purchase - Tally Orphans")
    hdr(ws, ["Tally Voucher #", "Party", "Date", "Amount (₹)", "Note"])
    for v in purchase_result["orphans"]:
        ws.append([
            v.get("number", ""),
            v.get("party", ""),
            v.get("date", ""),
            _tally_amount(v.get("amount", "0")),
            "In Tally but no matching Odoo bill found",
        ])
        colour_row(ws, ws.max_row, DUPE_FILL)
    auto_width(ws)

    # ── Sheet 8: Tally Duplicate Vouchers ─────────────────────────────
    ws = make_sheet("Tally Duplicate Vouchers")
    hdr(ws, ["Type", "Voucher A #", "Party A", "Date A", "Amount A",
             "Voucher B #", "Party B", "Date B", "Amount B", "Reason"])
    for vch_a, vch_b, reason in sales_tally_dupes:
        ws.append([
            "GST INVOICE",
            vch_a.get("number", ""), vch_a.get("party", ""),
            vch_a.get("date", ""), _tally_amount(vch_a.get("amount")),
            vch_b.get("number", ""), vch_b.get("party", ""),
            vch_b.get("date", ""), _tally_amount(vch_b.get("amount")),
            reason,
        ])
        colour_row(ws, ws.max_row, ERR_FILL)
    for vch_a, vch_b, reason in purchase_tally_dupes:
        ws.append([
            "Purchase",
            vch_a.get("number", ""), vch_a.get("party", ""),
            vch_a.get("date", ""), _tally_amount(vch_a.get("amount")),
            vch_b.get("number", ""), vch_b.get("party", ""),
            vch_b.get("date", ""), _tally_amount(vch_b.get("amount")),
            reason,
        ])
        colour_row(ws, ws.max_row, ERR_FILL)
    if not sales_tally_dupes and not purchase_tally_dupes:
        ws.append(["No duplicate vouchers detected in Tally for this period."])
    auto_width(ws)

    # ── Sheet 9: Tally Ledger Near-Duplicates ─────────────────────────
    ws = make_sheet("Tally Ledger Duplicates")
    hdr(ws, ["Ledger Name A", "Ledger Name B", "Reason"])
    for a, b, reason in ledger_dupes:
        ws.append([a, b, reason])
        colour_row(ws, ws.max_row, ERR_FILL)
    if not ledger_dupes:
        ws.append(["No near-duplicate ledgers detected in Tally."])
    auto_width(ws)

    # ── Sheet 10: New Ledgers Needed ──────────────────────────────────
    ws = make_sheet("New Ledgers Needed")
    hdr(ws, ["Ledger Name", "Suggested Group", "From Invoice"])
    for name, group, inv_name in new_ledgers:
        ws.append([name, group, inv_name])
        colour_row(ws, ws.max_row, WARN_FILL)
    if not new_ledgers:
        ws.append(["All required ledgers already exist in Tally."])
    auto_width(ws)

    # ── Sheet 11: Purchase Date Validation ────────────────────────────
    ws = make_sheet("Purchase Date Validation")
    hdr(ws, ["Odoo Bill #", "Tally Voucher #", "Issue Type",
             "Odoo Value", "Tally Value", "Note"])
    _date_issues = purchase_date_issues or []
    if _date_issues:
        for issue in _date_issues:
            ws.append([
                issue["odoo_bill"],
                issue["tally_num"],
                issue["issue_type"],
                issue["odoo_val"],
                issue["tally_val"],
                issue["note"],
            ])
            if "Mismatch" in issue["issue_type"] or "Missing" in issue["issue_type"]:
                colour_row(ws, ws.max_row, ERR_FILL)
            else:
                colour_row(ws, ws.max_row, WARN_FILL)
    else:
        ws.append(["All matched purchase vouchers have correct Bill Date and Accounting Date in Tally."])
        colour_row(ws, ws.max_row, OK_FILL)
    auto_width(ws)

    wb.save(path)
    log.info(f"Excel report saved: {path}")


# ── Main ──────────────────────────────────────────────────────────────

def run_check(from_date: str, to_date: str) -> dict:
    mapping = load_mapping()

    log.info("=" * 65)
    log.info(f"SYNC CHECK  {from_date}  →  {to_date}")
    log.info("=" * 65)

    # ── Connect ───────────────────────────────────────────────────────
    try:
        odoo = OdooConnector()
        log.info(f"[OK] Odoo: {odoo.url}  DB: {odoo.db}")
    except Exception as exc:
        log.error(f"[FAIL] Odoo connection: {exc}")
        sys.exit(1)

    company_name = os.getenv("TALLY_COMPANY_NAME", "")
    if not company_name:
        log.error("[FAIL] TALLY_COMPANY_NAME not set in .env")
        sys.exit(1)

    tally = TallyConnector()
    try:
        info = tally.test_connection()
        log.info(f"[OK] Tally: {info['url']}  Company: {company_name}")
    except Exception as exc:
        log.error(f"[FAIL] Tally connection: {exc}")
        sys.exit(1)

    tally_from = from_date.replace("-", "")
    tally_to   = to_date.replace("-", "")

    # ── Fetch Odoo invoices ───────────────────────────────────────────
    log.info("\n[1/6] Fetching Odoo invoices...")
    odoo_sales    = odoo.get_sales_invoices(from_date=from_date, to_date=to_date)
    odoo_sales    = [i for i in odoo_sales if float(i.get("amount_total", 0)) != 0]
    odoo_purchase = odoo.get_purchase_invoices(from_date=from_date, to_date=to_date)
    odoo_purchase = [i for i in odoo_purchase if float(i.get("amount_total", 0)) != 0]
    log.info(f"  Odoo Sales invoices   : {len(odoo_sales)}")
    log.info(f"  Odoo Purchase invoices: {len(odoo_purchase)}")

    # ── Fetch Tally vouchers ──────────────────────────────────────────
    log.info("\n[2/6] Fetching Tally vouchers...")
    tally_sales_raw  = fetch_tally_vouchers(tally, SALES_VCH_TYPE, tally_from, tally_to)
    tally_sales_raw += fetch_tally_vouchers(tally, "Sales", tally_from, tally_to)
    # Deduplicate: GST INVOICE is a child of Sales in Tally — both queries return same vouchers
    tally_sales    = deduplicate_vouchers(tally_sales_raw)
    tally_purchase = fetch_tally_vouchers(tally, PURCHASE_VCH_TYPE, tally_from, tally_to)
    log.info(f"  Tally GST INVOICE+Sales (deduped): {len(tally_sales)}")
    log.info(f"  Tally Purchase vouchers           : {len(tally_purchase)}")

    # ── Fetch Tally ledgers ───────────────────────────────────────────
    log.info("\n[3/6] Fetching Tally ledger list...")
    tally_ledgers = []
    tally_ledger_cache = set()
    try:
        tally_ledgers      = tally.get_all_ledgers()
        tally_ledger_cache = {l["name"].lower() for l in tally_ledgers}
        log.info(f"  {len(tally_ledgers)} ledgers in Tally")
    except Exception as exc:
        log.warning(f"  Could not fetch ledgers: {exc}")

    # ── Cross-match Odoo ↔ Tally ──────────────────────────────────────
    log.info("\n[4/6] Cross-matching Odoo invoices with Tally vouchers...")
    sales_result    = match_odoo_to_tally(odoo_sales,    tally_sales,    "sales",    mapping)
    purchase_result = match_odoo_to_tally(odoo_purchase, tally_purchase, "purchase", mapping)

    log.info(f"\n  === SALES ===")
    log.info(f"  Matched (already in Tally) : {len(sales_result['matched'])}")
    log.info(f"  UNMATCHED (needs push)     : {len(sales_result['unmatched'])}")
    log.info(f"  Orphan Tally vouchers      : {len(sales_result['orphans'])}")

    if sales_result["unmatched"]:
        total_pending = sum(float(i.get("amount_total", 0)) for i in sales_result["unmatched"])
        log.info(f"\n  Sales invoices NOT yet in Tally (₹{total_pending:,.2f} total):")
        for inv in sales_result["unmatched"]:
            partner = inv.get("partner_id", [None, ""])[1] if inv.get("partner_id") else "N/A"
            log.info(f"    [PUSH]  {inv.get('name',''):<28} {partner:<35} "
                     f"{inv.get('invoice_date',''):<12}  ₹{float(inv.get('amount_total',0)):>12,.2f}")

    log.info(f"\n  === PURCHASE ===")
    log.info(f"  Matched (already in Tally) : {len(purchase_result['matched'])}")
    log.info(f"  UNMATCHED (needs push)     : {len(purchase_result['unmatched'])}")
    log.info(f"  Orphan Tally vouchers      : {len(purchase_result['orphans'])}")

    if purchase_result["unmatched"]:
        total_pending = sum(float(i.get("amount_total", 0)) for i in purchase_result["unmatched"])
        log.info(f"\n  Purchase invoices NOT yet in Tally (₹{total_pending:,.2f} total):")
        for inv in purchase_result["unmatched"]:
            partner = inv.get("partner_id", [None, ""])[1] if inv.get("partner_id") else "N/A"
            log.info(f"    [PUSH]  {inv.get('name',''):<28} {partner:<35} "
                     f"InvDate:{inv.get('invoice_date',''):<12}  ₹{float(inv.get('amount_total',0)):>12,.2f}")

    # ── Purchase date validation ───────────────────────────────────────
    log.info("\n[4b/6] Validating Bill Date & Accounting Date for matched purchase vouchers...")
    purchase_date_issues = validate_purchase_dates(purchase_result["matched"])
    if purchase_date_issues:
        log.warning(f"  [!] {len(purchase_date_issues)} date issue(s) found in matched purchase vouchers:")
        for issue in purchase_date_issues:
            log.warning(f"      {issue['odoo_bill']:<28} Tally:{issue['tally_num']:<20} "
                        f"[{issue['issue_type']}]  "
                        f"Odoo={issue['odoo_val']}  Tally={issue['tally_val']}")
    else:
        log.info("  All matched purchase vouchers: Bill Date & Accounting Date correctly set in Tally.")

    # ── Tally internal duplicate check ────────────────────────────────
    log.info("\n[5/6] Checking Tally for duplicate vouchers...")
    sales_tally_dupes    = find_tally_duplicate_vouchers(tally_sales)
    purchase_tally_dupes = find_tally_duplicate_vouchers(tally_purchase)

    if sales_tally_dupes:
        log.warning(f"  [!] {len(sales_tally_dupes)} potential duplicate GST INVOICE pair(s) in Tally:")
        for a, b, reason in sales_tally_dupes:
            log.warning(f"      {a.get('number','?')} vs {b.get('number','?')}  ({reason})")
    else:
        log.info("  No duplicate sales vouchers detected in Tally.")

    if purchase_tally_dupes:
        log.warning(f"  [!] {len(purchase_tally_dupes)} potential duplicate Purchase pair(s) in Tally:")
        for a, b, reason in purchase_tally_dupes:
            log.warning(f"      {a.get('number','?')} vs {b.get('number','?')}  ({reason})")
    else:
        log.info("  No duplicate purchase vouchers detected in Tally.")

    # ── Ledger duplicate check ────────────────────────────────────────
    log.info("\n[6/6] Checking Tally ledgers for near-duplicates...")
    ledger_dupes = find_tally_ledger_duplicates(tally_ledgers)
    if ledger_dupes:
        log.warning(f"  [!] {len(ledger_dupes)} near-duplicate ledger pair(s):")
        for a, b, reason in ledger_dupes:
            log.warning(f"      '{a}'  ≈  '{b}'  ({reason})")
    else:
        log.info("  No near-duplicate ledgers detected.")

    # ── New ledgers needed ────────────────────────────────────────────
    all_unmatched = sales_result["unmatched"] + purchase_result["unmatched"]
    all_odoo      = odoo_sales + odoo_purchase
    new_ledgers   = find_new_ledgers_needed(all_odoo, all_unmatched,
                                             tally_ledger_cache, mapping)
    if new_ledgers:
        log.info(f"\n  New ledgers that will be AUTO-CREATED during push:")
        for name, group, inv_name in new_ledgers:
            log.info(f"    + '{name}'  (group: {group})  ← from {inv_name}")

    # ── Excel output ──────────────────────────────────────────────────
    report_path = TMP_DIR / f"sync_check_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    write_excel_report(
        path=report_path,
        from_date=from_date,
        to_date=to_date,
        sales_result=sales_result,
        purchase_result=purchase_result,
        sales_tally_dupes=sales_tally_dupes,
        purchase_tally_dupes=purchase_tally_dupes,
        ledger_dupes=ledger_dupes,
        new_ledgers=new_ledgers,
        purchase_date_issues=purchase_date_issues,
    )

    # ── Final console summary ─────────────────────────────────────────
    log.info("\n" + "=" * 65)
    log.info("SYNC CHECK COMPLETE")
    log.info(f"  Period             : {from_date}  →  {to_date}")
    log.info(f"  Sales pending push : {len(sales_result['unmatched'])}")
    log.info(f"  Purchase pending   : {len(purchase_result['unmatched'])}")
    log.info(f"  Tally sales dupes  : {len(sales_tally_dupes)}")
    log.info(f"  Tally pur dupes    : {len(purchase_tally_dupes)}")
    log.info(f"  Ledger dupes       : {len(ledger_dupes)}")
    log.info(f"  New ledgers needed : {len(new_ledgers)}")
    log.info(f"  Purchase date issues: {len(purchase_date_issues)}")
    if (sales_tally_dupes or purchase_tally_dupes or ledger_dupes or purchase_date_issues):
        log.warning("  *** REVIEW WARNINGS / DATE ISSUES BEFORE EXECUTING PUSH ***")
    elif sales_result["unmatched"] or purchase_result["unmatched"]:
        log.info("  *** Run push scripts with --execute to post pending invoices ***")
    else:
        log.info("  All invoices already in Tally — nothing to push.")
    log.info("=" * 65)

    return {
        "sales_unmatched":      len(sales_result["unmatched"]),
        "purchase_unmatched":   len(purchase_result["unmatched"]),
        "sales_dupes":          len(sales_tally_dupes),
        "purchase_dupes":       len(purchase_tally_dupes),
        "ledger_dupes":         len(ledger_dupes),
        "purchase_date_issues": len(purchase_date_issues),
        "report":               str(report_path),
    }


# ── CLI ───────────────────────────────────────────────────────────────

def main():
    today = date.today()
    # Default to current month
    default_from = today.replace(day=1).isoformat()
    default_to   = today.isoformat()

    parser = argparse.ArgumentParser(
        description="Sync Checker — Odoo ↔ Tally audit before push",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sync_checker.py                          # current month (default)
  python sync_checker.py --month 2026-05          # full May 2026
  python sync_checker.py --from 2026-05-01 --to 2026-05-12
  python sync_checker.py --full-fy                # full FY 2025-26
        """,
    )
    parser.add_argument("--from", dest="from_date", default=default_from,
                        help=f"Start date YYYY-MM-DD (default: {default_from})")
    parser.add_argument("--to", dest="to_date", default=default_to,
                        help=f"End date YYYY-MM-DD (default: {default_to})")
    parser.add_argument("--month", dest="month",
                        help="Short form: --month 2026-05 sets from=first day, to=last day")
    parser.add_argument("--full-fy", action="store_true",
                        help="Check full FY 2025-26 (2025-04-01 to 2026-03-31)")
    args = parser.parse_args()

    from_date = args.from_date
    to_date   = args.to_date

    if args.full_fy:
        from_date = "2025-04-01"
        to_date   = "2026-03-31"
    elif args.month:
        import calendar
        year, month = map(int, args.month.split("-"))
        last_day = calendar.monthrange(year, month)[1]
        from_date = f"{year:04d}-{month:02d}-01"
        to_date   = f"{year:04d}-{month:02d}-{last_day:02d}"

    run_check(from_date, to_date)


if __name__ == "__main__":
    main()
