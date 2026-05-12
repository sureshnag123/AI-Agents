#!/usr/bin/env python3
"""
Cash Flow Statement (Indirect Method) — Fracktal Works Pvt. Ltd.

Pulls Trial Balance (P&L movement + closing BS) and opening ledger balances
from Tally Prime, then builds an Indirect Method Cash Flow Statement.

Usage:
    python execution/cashflow_report.py
    python execution/cashflow_report.py --from 2025-04-01 --to 2025-09-30
    python execution/cashflow_report.py --dry-run
    python execution/cashflow_report.py --test
"""

import os
import sys
import re
import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
import xml.etree.ElementTree as ET
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

load_dotenv(PROJECT_ROOT / ".env")
sys.path.insert(0, str(SCRIPT_DIR))

from tally_connector import TallyConnector

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Defaults ───────────────────────────────────────────────────────────────────

DEFAULT_FROM = date(2025, 4, 1)
DEFAULT_TO   = date(2025, 9, 30)


# ── BS Group → CF Category Mapping ────────────────────────────────────────────
# Tally group display name → (cf_section, display_label)
#
# Sections:
#   "cash"         — Cash & Bank (used for opening/closing reconciliation)
#   "operating_wc" — Working Capital changes in Operating section
#   "investing_fa" — Fixed Assets in Investing section
#   "financing"    — Loans & Capital in Financing section
#
# Sign convention (TB NET: DR negative, CR positive):
#   Assets (DR balance): NET < 0. Change = closing_NET - opening_NET.
#     If asset grows → more negative → negative CF (cash outflow) ✓
#   Liabilities (CR balance): NET > 0. Change = closing_NET - opening_NET.
#     If liability grows → more positive → positive CF (cash inflow) ✓

BS_GROUP_MAP: Dict[str, Tuple[str, str]] = {
    # Cash & Equivalents
    "Bank Accounts":                  ("cash",         "Bank Accounts"),
    "Cash-in-Hand":                   ("cash",         "Cash-in-Hand"),
    # Working Capital — Current Assets
    "Sundry Debtors":                 ("operating_wc", "(Increase)/Decrease in Sundry Debtors"),
    "Loans & Advances (Asset)":       ("operating_wc", "(Increase)/Decrease in Loans & Advances"),
    "TDS - GST Receivable":           ("operating_wc", "(Increase)/Decrease in TDS/GST Receivable"),
    "Tds Receivable":                 ("operating_wc", "(Increase)/Decrease in TDS Receivable"),
    "Deposits (Asset)":               ("operating_wc", "(Increase)/Decrease in Deposits"),
    "Deferred Expenses":              ("operating_wc", "(Increase)/Decrease in Deferred Expenses"),
    # Working Capital — Current Liabilities
    "Sundry Creditors":               ("operating_wc", "Increase/(Decrease) in Sundry Creditors"),
    "Duties & Taxes":                 ("operating_wc", "Increase/(Decrease) in Duties & Taxes Payable"),
    "Salaries Payable":               ("operating_wc", "Increase/(Decrease) in Salaries Payable"),
    "Reimbursments Payable":          ("operating_wc", "Increase/(Decrease) in Reimbursements Payable"),
    # Fixed Assets → Investing
    "Machinery & Equipment":          ("investing_fa", "Machinery & Equipment"),
    "Tangible Assets":                ("investing_fa", "Tangible Assets"),
    "Intangible Assets":              ("investing_fa", "Intangible Assets"),
    "Fracktal Studio":                ("investing_fa", "Fracktal Studio"),
    "SLS PAMM Project":               ("investing_fa", "SLS PAMM Project"),
    "WIP":                            ("investing_fa", "Capital Work-in-Progress"),
    # Financing
    "Unsecured Loans":                ("financing",    "Proceeds/(Repayment) of Unsecured Loans"),
    "Issued, Subscribed and Paid Up": ("financing",    "Change in Share Capital"),
    # Note: Reserves & Surplus excluded — its change flows through Net Profit (P&L A/c)
}

# Ledger names containing these keywords are treated as Depreciation (case-insensitive)
DEPRECIATION_KEYWORDS = ("depreciat", "amortis", "amortiz")


# ── XML / Tally Helpers ────────────────────────────────────────────────────────

def sanitize_xml(text: str) -> str:
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    text = re.sub(r'&#x([0-8BbCcEeFf]);', '', text)
    text = re.sub(r'&#x1[0-9A-Fa-f];', '', text)
    text = re.sub(r'&#([0-8]|1[0-1]|1[4-9]|2[0-9]|3[01]);', '', text)
    return text


def parse_tb_amount(text: str) -> float:
    """Parse Tally TB amount string. DR is negative, CR is positive."""
    if not text or not text.strip():
        return 0.0
    text = text.strip().replace(",", "").replace("Rs.", "").replace("Rs", "")
    for suffix in (" Cr.", " Cr", " Dr.", " Dr"):
        if text.endswith(suffix):
            text = text[:-len(suffix)].strip()
            break
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_opening_balance(text: str) -> float:
    """
    Parse Tally OPENINGBALANCE field.
      '1,00,000 Dr' → -100000.0  (debit balance = asset, negative in TB convention)
      '50,000 Cr'   → +50000.0   (credit balance = liability/equity, positive)
      '0'           → 0.0
    """
    if not text or not text.strip() or text.strip() == "0":
        return 0.0
    raw = text.strip().replace(",", "")
    is_dr = raw.endswith(" Dr") or raw.endswith(" Dr.")
    for sfx in (" Dr.", " Dr", " Cr.", " Cr"):
        if raw.endswith(sfx):
            raw = raw[:-len(sfx)].strip()
            break
    try:
        val = float(raw)
    except ValueError:
        return 0.0
    return -val if is_dr else val


def fmt_date(d: date) -> str:
    return f"{d.year:04d}{d.month:02d}{d.day:02d}"


def fetch_tb_explode(url: str, company: str, from_date: str, to_date: str,
                     timeout: int = 120) -> Dict[str, float]:
    """
    Fetch Trial Balance with EXPLODEFLAG=Yes from Tally.
    Returns {display_name: net_value} where NET = DR + CR (DR neg, CR pos).
    P&L items: net movement for the period.
    BS items:  cumulative closing balance as of to_date.
    """
    xml_req = (
        f"<ENVELOPE><HEADER><VERSION>1</VERSION>"
        f"<TALLYREQUEST>Export</TALLYREQUEST><TYPE>Data</TYPE>"
        f"<ID>Trial Balance</ID></HEADER><BODY><DESC><STATICVARIABLES>"
        f"<SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
        f"<SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>"
        f"<SVFROMDATE>{from_date}</SVFROMDATE>"
        f"<SVTODATE>{to_date}</SVTODATE>"
        f"<EXPLODEFLAG>Yes</EXPLODEFLAG>"
        f"</STATICVARIABLES></DESC></BODY></ENVELOPE>"
    )
    try:
        resp = requests.post(
            url, data=xml_req.encode("utf-8"),
            headers={"Content-Type": "text/xml; charset=utf-8"},
            timeout=timeout,
        )
        resp.raise_for_status()
    except Exception as e:
        print(f"  [ERROR] TB fetch failed: {e}")
        return {}

    root = ET.fromstring(sanitize_xml(resp.text))
    result: Dict[str, float] = {}
    children = list(root)
    i = 0
    while i < len(children):
        if children[i].tag == "DSPACCNAME":
            name = (children[i].findtext("DSPDISPNAME") or "").strip()
            if i + 1 < len(children) and children[i + 1].tag == "DSPACCINFO":
                info = children[i + 1]
                dr = parse_tb_amount(info.findtext(".//DSPCLDRAMTA", ""))
                cr = parse_tb_amount(info.findtext(".//DSPCLCRAMTA", ""))
                if name:
                    result[name] = dr + cr
                i += 2
                continue
        i += 1
    return result


# ── Opening Balance Aggregation ────────────────────────────────────────────────

def build_group_root_map(groups: list) -> Dict[str, Optional[str]]:
    """
    Walk the Tally group parent chain and return {group_name: root_cf_group}.
    Only maps groups whose chain leads to a key in BS_GROUP_MAP.
    """
    known = set(BS_GROUP_MAP.keys())
    parent_map = {g["name"]: g["parent"] for g in groups}
    cache: Dict[str, Optional[str]] = {}

    def find_root(name: str, depth: int = 0) -> Optional[str]:
        if depth > 12 or not name:
            return None
        if name in cache:
            return cache[name]
        if name in known:
            cache[name] = name
            return name
        parent = parent_map.get(name, "")
        root = find_root(parent, depth + 1) if parent and parent != name else None
        cache[name] = root
        return root

    return {g["name"]: find_root(g["name"]) for g in groups}


def get_opening_bs(tally: TallyConnector) -> Dict[str, float]:
    """
    Fetch all ledger opening balances from Tally and aggregate by CF group.
    Returns {cf_group_name: total_opening_net} in TB sign convention
    (assets: negative, liabilities/equity: positive).

    Also stores '_reserves_opening' — the opening Reserves & Surplus balance
    used to compute Net Profit for the period (closing R&S − opening R&S).
    """
    print("  Fetching ledger opening balances...")
    ledgers = tally.get_all_ledgers()
    print("  Fetching group hierarchy...")
    groups  = tally.get_all_groups()

    group_root_map = build_group_root_map(groups)

    # Build a quick group→parent map for special lookups
    parent_of = {g["name"]: g["parent"] for g in groups}

    result: Dict[str, float] = {}
    unmapped_groups: set = set()

    for ldg in ledgers:
        lgroup = ldg.get("group", "")
        bal = parse_opening_balance(ldg.get("opening_balance", "0"))

        # Special: capture Reserves & Surplus opening for Net Profit calculation.
        # Walk up to find any group whose name contains "reserve" (case-insensitive).
        grp = lgroup
        for _ in range(8):
            if "reserve" in grp.lower() or "reserves" in grp.lower():
                result["_reserves_opening"] = result.get("_reserves_opening", 0.0) + bal
                break
            grp = parent_of.get(grp, "")
            if not grp:
                break

        # Normal CF group mapping
        root = group_root_map.get(lgroup) or (lgroup if lgroup in BS_GROUP_MAP else None)
        if root is None:
            unmapped_groups.add(lgroup)
            continue
        result[root] = result.get(root, 0.0) + bal

    if unmapped_groups:
        sample = ", ".join(sorted(unmapped_groups)[:5])
        print(f"  Note: {len(unmapped_groups)} group(s) not mapped to CF categories "
              f"(e.g. {sample})")
    return result


# ── Cash Flow Builder ──────────────────────────────────────────────────────────

def build_cashflow(
    tb: Dict[str, float],
    opening_bs: Dict[str, float],
    from_date: date,
    to_date: date,
) -> dict:
    """
    Compute the Indirect Method Cash Flow Statement.

    All output values use ECONOMIC sign convention:
      Positive = cash inflow / profit
      Negative = cash outflow / loss

    TB sign convention (DR negative, CR positive) is preserved for the
    change calculations — assets decrease cash when they grow (more negative
    closing NET), liabilities increase cash when they grow (more positive
    closing NET). Both handled uniformly by: change = closing_NET - opening_NET.
    """

    # ── Net Profit ─────────────────────────────────────────────────────
    # EXPLODEFLAG TB shows both parent groups AND their children.  The 6
    # root P&L groups (direct children of Primary in Tally's hierarchy)
    # represent the complete, non-duplicated income/expense picture:
    #   Income  (CR, positive NET): Sales Accounts, Direct Incomes, Indirect Incomes
    #   Expense (DR, negative NET): Purchase Accounts, Direct Expenses, Indirect Expenses
    # Net Profit = algebraic sum of these 6 groups.
    PL_ROOT_INCOME   = ("Sales Accounts", "Direct Incomes", "Indirect Incomes")
    PL_ROOT_EXPENSE  = ("Purchase Accounts", "Direct Expenses", "Indirect Expenses")
    net_profit = sum(tb.get(g, 0.0) for g in PL_ROOT_INCOME + PL_ROOT_EXPENSE)

    # ── Depreciation Add-back ──────────────────────────────────────────
    # Depreciation is a non-cash expense (TB NET negative).
    # Add-back = -TB_NET → positive number added back to Net Profit.
    dep_items: Dict[str, float] = {
        name: net for name, net in tb.items()
        if any(kw in name.lower() for kw in DEPRECIATION_KEYWORDS)
    }
    dep_addback = -sum(dep_items.values())   # positive

    # ── Working Capital Changes ────────────────────────────────────────
    # change = closing_TB_NET - opening_TB_NET for each group.
    # Assets grow → more negative NET → negative change → cash outflow ✓
    # Liabilities grow → more positive NET → positive change → cash inflow ✓
    wc_items: List[Tuple[str, float]] = []
    for group, (section, label) in BS_GROUP_MAP.items():
        if section != "operating_wc":
            continue
        change = tb.get(group, 0.0) - opening_bs.get(group, 0.0)
        wc_items.append((label, change))

    net_wc = sum(v for _, v in wc_items)
    net_operating = net_profit + dep_addback + net_wc

    # ── Investing: Fixed Assets ────────────────────────────────────────
    # Net FA change (TB) = Depreciation - Gross Capex
    # Investing CF = Net FA change - Dep add-back = -Gross Capex (outflow)
    #
    # Example: Net FA change = -1L (grew by 1L net), Dep = 50k
    #   Investing CF = -1L - 50k = -1.5L (gross capex = 1.5L cash out) ✓
    fa_items: List[Tuple[str, float]] = []
    fa_change_total = 0.0
    for group, (section, label) in BS_GROUP_MAP.items():
        if section != "investing_fa":
            continue
        change = tb.get(group, 0.0) - opening_bs.get(group, 0.0)
        fa_items.append((label, change))
        fa_change_total += change

    net_investing = fa_change_total - dep_addback   # = -Gross Capex

    # ── Financing ──────────────────────────────────────────────────────
    fin_items: List[Tuple[str, float]] = []
    for group, (section, label) in BS_GROUP_MAP.items():
        if section != "financing":
            continue
        change = tb.get(group, 0.0) - opening_bs.get(group, 0.0)
        fin_items.append((label, change))

    net_financing = sum(v for _, v in fin_items)

    # ── Cash Reconciliation ────────────────────────────────────────────
    # Assets have negative TB NET; flip sign for economic (positive) display.
    opening_cash = -sum(opening_bs.get(g, 0.0) for g in ("Bank Accounts", "Cash-in-Hand"))
    closing_cash_tally = -sum(tb.get(g, 0.0) for g in ("Bank Accounts", "Cash-in-Hand"))
    net_change = net_operating + net_investing + net_financing
    closing_cash_computed = opening_cash + net_change
    discrepancy = closing_cash_computed - closing_cash_tally

    return {
        "period_from":            from_date,
        "period_to":              to_date,
        "net_profit":             net_profit,
        "dep_items":              dep_items,
        "dep_addback":            dep_addback,
        "wc_items":               wc_items,
        "net_wc":                 net_wc,
        "net_operating":          net_operating,
        "fa_items":               fa_items,
        "fa_change_total":        fa_change_total,
        "net_investing":          net_investing,
        "fin_items":              fin_items,
        "net_financing":          net_financing,
        "net_change":             net_change,
        "opening_cash":           opening_cash,
        "closing_cash_computed":  closing_cash_computed,
        "closing_cash_tally":     closing_cash_tally,
        "discrepancy":            discrepancy,
    }


# ── Console Output ─────────────────────────────────────────────────────────────

def _fmt(val: float, w: int = 16) -> str:
    if abs(val) < 0.01:
        return f"{'0.00':>{w}}"
    return f"{val:>{w},.2f}"


def print_cashflow(cf: dict) -> None:
    period = (f"{cf['period_from'].strftime('%d-%b-%Y')} to "
              f"{cf['period_to'].strftime('%d-%b-%Y')}")
    W = 68

    def ruler(ch="-"):
        print("  " + ch * W)

    def row(label, val=None, indent=0):
        prefix = "  " + ("    " * indent)
        if val is None:
            print(f"{prefix}{label}")
        else:
            print(f"{prefix}{label:<52} {_fmt(val)}")

    def subtotal(label, val):
        ruler()
        print(f"  {label:<54} {_fmt(val)}")
        ruler()

    print(f"\n{'='*72}")
    print(f"  CASH FLOW STATEMENT (Indirect Method)  —  Fracktal Works Pvt. Ltd.")
    print(f"  Period: {period}")
    print(f"{'='*72}")

    # A. Operating
    print(f"\n  A. CASH FLOW FROM OPERATING ACTIVITIES")
    ruler()
    row("Net Profit / (Loss) for the period", cf["net_profit"])
    row("Add: Depreciation & Amortisation", cf["dep_addback"])
    if cf["dep_items"]:
        for name, net in cf["dep_items"].items():
            row(f"  [{name}]", -net, indent=1)
    row("Working Capital Changes:", indent=0)
    for label, val in cf["wc_items"]:
        if abs(val) > 0.01:
            row(label, val, indent=1)
    subtotal("NET CASH FROM OPERATING ACTIVITIES  (A)", cf["net_operating"])

    # B. Investing
    print(f"\n  B. CASH FLOW FROM INVESTING ACTIVITIES")
    ruler()
    any_fa = any(abs(v) > 0.01 for _, v in cf["fa_items"])
    if any_fa:
        for label, val in cf["fa_items"]:
            if abs(val) > 0.01:
                row(f"Change in {label} (net)", val)
        if cf["dep_addback"] > 0.01:
            row("Less: Depreciation included in above (non-cash)", -cf["dep_addback"])
    else:
        row("No investing activities in the period", 0.0)
    subtotal("NET CASH FROM INVESTING ACTIVITIES  (B)", cf["net_investing"])

    # C. Financing
    print(f"\n  C. CASH FLOW FROM FINANCING ACTIVITIES")
    ruler()
    any_fin = any(abs(v) > 0.01 for _, v in cf["fin_items"])
    if any_fin:
        for label, val in cf["fin_items"]:
            if abs(val) > 0.01:
                row(label, val)
    else:
        row("No financing activities in the period", 0.0)
    subtotal("NET CASH FROM FINANCING ACTIVITIES  (C)", cf["net_financing"])

    # Summary
    print(f"\n{'='*72}")
    print(f"  NET CHANGE IN CASH (A+B+C){'':30} {_fmt(cf['net_change'])}")
    print(f"  Opening Cash & Bank Balance (01-Apr-2025){'':15} {_fmt(cf['opening_cash'])}")
    print(f"  Closing Cash & Bank (Computed){'':26} {_fmt(cf['closing_cash_computed'])}")
    print(f"  Closing Cash & Bank (Per Tally TB){'':22} {_fmt(cf['closing_cash_tally'])}")
    disc = cf["discrepancy"]
    if abs(disc) > 1.0:
        print(f"\n  ** DISCREPANCY: Rs.{disc:,.2f} — check unmapped accounts **")
    else:
        print(f"\n  Cash reconciliation: OK")
    print(f"{'='*72}\n")


# ── Excel Output ───────────────────────────────────────────────────────────────

def write_excel(cf: dict, filepath: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 20

    # Colour palette
    C_DARK    = "1F3864"
    C_MID     = "2E5090"
    C_LIGHT   = "D6E4F0"
    C_SUBTOT  = "BDD7EE"
    C_WHITE   = "FFFFFF"

    def _fill(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def _font(size=10, bold=False, color="000000"):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    _thin = Border(bottom=Side(style="thin", color="BBBBBB"))
    _thick = Border(
        top=Side(style="medium", color=C_DARK),
        bottom=Side(style="medium", color=C_DARK),
    )

    r = [0]

    def nr():
        r[0] += 1
        return r[0]

    def add(label="", amount=None, style="normal", indent=0):
        row_n = nr()
        ws.cell(row_n, 1, ("    " * indent) + label)
        if amount is not None:
            ws.cell(row_n, 2, round(amount, 2))
            ws.cell(row_n, 2).number_format = '#,##0.00;[Red]-#,##0.00'
            ws.cell(row_n, 2).alignment = Alignment(horizontal="right")

        if style == "title":
            for c in (1, 2):
                ws.cell(row_n, c).font  = _font(12, True, C_WHITE)
                ws.cell(row_n, c).fill  = _fill(C_DARK)
            ws.merge_cells(f"A{row_n}:B{row_n}")
            ws.cell(row_n, 1).alignment = Alignment(horizontal="center")
        elif style == "section":
            for c in (1, 2):
                ws.cell(row_n, c).font  = _font(10, True, C_WHITE)
                ws.cell(row_n, c).fill  = _fill(C_MID)
        elif style == "group":
            for c in (1, 2):
                ws.cell(row_n, c).font  = _font(10, True)
                ws.cell(row_n, c).fill  = _fill(C_LIGHT)
        elif style == "subtotal":
            for c in (1, 2):
                ws.cell(row_n, c).font   = _font(10, True)
                ws.cell(row_n, c).fill   = _fill(C_SUBTOT)
                ws.cell(row_n, c).border = _thick
        elif style == "total":
            for c in (1, 2):
                ws.cell(row_n, c).font   = _font(11, True, C_WHITE)
                ws.cell(row_n, c).fill   = _fill(C_DARK)
                ws.cell(row_n, c).border = _thick
        else:
            ws.cell(row_n, 1).font   = _font(10)
            ws.cell(row_n, 2).font   = _font(10)
            ws.cell(row_n, 1).border = _thin
            ws.cell(row_n, 2).border = _thin

    period = (f"{cf['period_from'].strftime('%d-%b-%Y')} to "
              f"{cf['period_to'].strftime('%d-%b-%Y')}")

    add("CASH FLOW STATEMENT (Indirect Method)", style="title")
    add(f"Fracktal Works Pvt. Ltd.   |   Period: {period}", style="title")
    add()

    # A. Operating
    add("A.  CASH FLOW FROM OPERATING ACTIVITIES", style="section")
    add("Net Profit / (Loss) for the period", cf["net_profit"])
    add("Add: Depreciation & Amortisation", cf["dep_addback"])
    if cf["dep_items"]:
        for name, net in cf["dep_items"].items():
            add(f"[{name}]", -net, indent=1)
    add("Working Capital Changes:", style="group")
    for label, val in cf["wc_items"]:
        if abs(val) > 0.01:
            add(label, val, indent=1)
    add("NET CASH FROM OPERATING ACTIVITIES  (A)", cf["net_operating"], style="subtotal")
    add()

    # B. Investing
    add("B.  CASH FLOW FROM INVESTING ACTIVITIES", style="section")
    any_fa = any(abs(v) > 0.01 for _, v in cf["fa_items"])
    if any_fa:
        for label, val in cf["fa_items"]:
            if abs(val) > 0.01:
                add(f"Change in {label} (net of depreciation)", val)
        if cf["dep_addback"] > 0.01:
            add("Less: Depreciation included above (non-cash)", -cf["dep_addback"])
    else:
        add("No investing activities in the period", 0.0)
    add("NET CASH FROM INVESTING ACTIVITIES  (B)", cf["net_investing"], style="subtotal")
    add()

    # C. Financing
    add("C.  CASH FLOW FROM FINANCING ACTIVITIES", style="section")
    any_fin = any(abs(v) > 0.01 for _, v in cf["fin_items"])
    if any_fin:
        for label, val in cf["fin_items"]:
            if abs(val) > 0.01:
                add(label, val)
    else:
        add("No financing activities in the period", 0.0)
    add("NET CASH FROM FINANCING ACTIVITIES  (C)", cf["net_financing"], style="subtotal")
    add()

    # Summary
    add("NET CHANGE IN CASH AND CASH EQUIVALENTS  (A+B+C)",
        cf["net_change"], style="total")
    add()
    add(f"Opening Cash & Bank Balance  (01-Apr-2025)", cf["opening_cash"])
    add("Closing Cash & Bank Balance  (Computed)", cf["closing_cash_computed"])
    add("Closing Cash & Bank Balance  (Per Tally TB)", cf["closing_cash_tally"])
    disc = cf["discrepancy"]
    add(f"Reconciliation: {'OK' if abs(disc) < 1.0 else f'DISCREPANCY Rs.{disc:,.2f}'}")
    add()
    add(f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}")

    ws.freeze_panes = "A4"
    wb.save(filepath)
    print(f"  Excel saved: {filepath}")


# ── Entry Point ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Cash Flow Statement (Indirect Method) from Tally Prime"
    )
    parser.add_argument("--from", dest="from_date", default=None,
                        help="Period start YYYY-MM-DD (default: 2025-04-01)")
    parser.add_argument("--to",   dest="to_date",   default=None,
                        help="Period end   YYYY-MM-DD (default: 2025-09-30)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print to console only, skip Excel export")
    parser.add_argument("--test", action="store_true",
                        help="Test Tally connection and exit")
    args = parser.parse_args()

    from_date = date.fromisoformat(args.from_date) if args.from_date else DEFAULT_FROM
    to_date   = date.fromisoformat(args.to_date)   if args.to_date   else DEFAULT_TO

    company    = os.getenv("TALLY_COMPANY_NAME", "")
    tally_host = os.getenv("TALLY_HOST", "localhost")
    tally_port = int(os.getenv("TALLY_PORT", "9000"))
    url        = f"http://{tally_host}:{tally_port}"

    tally = TallyConnector()
    try:
        tally.test_connection()
        print(f"[OK] Tally connected: {url}")
    except ConnectionError as e:
        print(f"[FAIL] {e}")
        sys.exit(1)

    if args.test:
        return

    print(f"\nCash Flow Report — {company}")
    print(f"Period : {from_date.strftime('%d-%b-%Y')} to {to_date.strftime('%d-%b-%Y')}")

    # Step 1: Period TB (P&L movements + closing BS)
    print(f"\n[1/3] Fetching Trial Balance ({fmt_date(from_date)} to {fmt_date(to_date)})...")
    tb = fetch_tb_explode(url, company, fmt_date(from_date), fmt_date(to_date))
    if not tb:
        print("[FAIL] No data returned from Tally.")
        sys.exit(1)
    print(f"      {len(tb)} line items received.")

    # Step 2: Opening BS (April 1 ledger opening balances)
    print(f"\n[2/3] Fetching opening balance sheet (1-Apr-2025)...")
    opening_bs = get_opening_bs(tally)
    print(f"      Aggregated {len(opening_bs)} CF group(s).")

    # Step 3: Build & output
    print(f"\n[3/3] Building cash flow statement...")
    cf = build_cashflow(tb, opening_bs, from_date, to_date)
    print_cashflow(cf)

    if not args.dry_run:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = TMP_DIR / f"cashflow_{from_date.strftime('%b%Y')}_{to_date.strftime('%b%Y')}_{ts}.xlsx"
        TMP_DIR.mkdir(parents=True, exist_ok=True)
        write_excel(cf, out)


if __name__ == "__main__":
    main()
