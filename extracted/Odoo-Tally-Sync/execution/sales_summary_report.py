#!/usr/bin/env python3
"""
Sales Summary Report - Odoo -> Excel

Generates a month-on-month sales analysis for the last 3 financial years
(Indian FY: April-March) with quarterly and annual totals.

Output: two-sheet Excel workbook
  Sheet 1 - Total Sales Summary  (Gross Sales | Credit Notes | Net Sales)
  Sheet 2 - Service & Goods Breakdown (same 3-col pattern per category)

Usage:
    python sales_summary_report.py
    python sales_summary_report.py --output my_report.xlsx
"""

import os
import sys
import argparse
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict

from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
load_dotenv(PROJECT_ROOT / ".env")
sys.path.insert(0, str(SCRIPT_DIR))

from odoo_connector import OdooConnector


# ── Financial Year Helpers ──────────────────────────────────────────────────

def fy_label(yr: int) -> str:
    return f"FY {yr}-{str(yr + 1)[2:]}"


def date_to_fy_and_month(invoice_date: str):
    d = datetime.strptime(invoice_date[:10], "%Y-%m-%d")
    fy_start = d.year if d.month >= 4 else d.year - 1
    return fy_start, d.month


FY_MONTHS = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
MONTH_NAMES = {
    1: "January",  2: "February",  3: "March",
    4: "April",    5: "May",       6: "June",
    7: "July",     8: "August",    9: "September",
    10: "October", 11: "November", 12: "December",
}
QUARTERS = {
    "Q1 (Apr-Jun)": [4, 5, 6],
    "Q2 (Jul-Sep)": [7, 8, 9],
    "Q3 (Oct-Dec)": [10, 11, 12],
    "Q4 (Jan-Mar)": [1, 2, 3],
}


# ── Data Fetching ───────────────────────────────────────────────────────────

def _empty() -> dict:
    return defaultdict(lambda: defaultdict(float))


def fetch_sales_data(odoo: OdooConnector, fy_years: list):
    """
    Returns six dicts, each keyed [fy_start][month] -> float (always positive):
      gross_sales  - out_invoice amount_untaxed
      credit_notes - out_refund  amount_untaxed  (positive; shown separately)
      svc_gross    - service lines from invoices
      svc_cn       - service lines from credit notes
      gds_gross    - goods lines from invoices
      gds_cn       - goods lines from credit notes
    """
    from_date = f"{min(fy_years)}-04-01"
    to_date   = f"{max(fy_years) + 1}-03-31"

    print(f"Fetching sales invoices {from_date} to {to_date} ...")
    invoices = odoo.search_read(
        "account.move",
        [
            ("move_type", "in", ["out_invoice", "out_refund"]),
            ("state", "=", "posted"),
            ("invoice_date", ">=", from_date),
            ("invoice_date", "<=", to_date),
        ],
        ["id", "name", "move_type", "invoice_date", "amount_untaxed"],
        order="invoice_date asc",
    )
    print(f"  -> {len(invoices)} records found")

    # inv_meta: id -> (fy, month, is_cn, amount)
    inv_meta = {}
    for inv in invoices:
        if not inv["invoice_date"]:
            continue
        fy, mo = date_to_fy_and_month(inv["invoice_date"])
        is_cn  = inv["move_type"] == "out_refund"
        inv_meta[inv["id"]] = (fy, mo, is_cn, inv["amount_untaxed"])

    gross_sales  = _empty()
    credit_notes = _empty()

    for inv_id, (fy, mo, is_cn, amt) in inv_meta.items():
        if fy not in fy_years:
            continue
        if is_cn:
            credit_notes[fy][mo] += amt
        else:
            gross_sales[fy][mo]  += amt

    # ── Service / Goods breakdown via invoice lines ──────────────────────
    svc_gross = _empty()
    svc_cn    = _empty()
    gds_gross = _empty()
    gds_cn    = _empty()

    inv_ids = list(inv_meta.keys())
    if inv_ids:
        print(f"Fetching invoice lines for {len(inv_ids)} invoices ...")
        CHUNK = 200
        all_lines = []
        for i in range(0, len(inv_ids), CHUNK):
            chunk = inv_ids[i : i + CHUNK]
            batch = odoo.search_read(
                "account.move.line",
                [
                    ("move_id", "in", chunk),
                    ("product_id", "!=", False),
                    ("display_type", "=", "product"),
                ],
                ["move_id", "product_id", "price_subtotal"],
            )
            all_lines.extend(batch)
        print(f"  -> {len(all_lines)} line items fetched")

        product_ids = list({l["product_id"][0] for l in all_lines if l["product_id"]})
        print(f"Fetching product types for {len(product_ids)} products ...")
        PCHUNK = 500
        prod_type_map = {}
        for i in range(0, len(product_ids), PCHUNK):
            chunk = product_ids[i : i + PCHUNK]
            prods = odoo.read("product.product", chunk, ["id", "type"])
            for p in prods:
                prod_type_map[p["id"]] = p.get("type", "")

        for line in all_lines:
            if not line["product_id"] or not line["move_id"]:
                continue
            pid = line["product_id"][0]
            mid = line["move_id"][0] if isinstance(line["move_id"], list) else line["move_id"]
            if mid not in inv_meta:
                continue
            fy, mo, is_cn, _ = inv_meta[mid]
            if fy not in fy_years:
                continue
            amt   = abs(line["price_subtotal"] or 0)
            ptype = prod_type_map.get(pid, "")
            if ptype == "service":
                if is_cn: svc_cn[fy][mo]    += amt
                else:     svc_gross[fy][mo] += amt
            else:
                if is_cn: gds_cn[fy][mo]    += amt
                else:     gds_gross[fy][mo] += amt

    return gross_sales, credit_notes, svc_gross, svc_cn, gds_gross, gds_cn


# ── Excel Styling ────────────────────────────────────────────────────────────

BLUE_DARK  = "1F3864"
BLUE_MID   = "2F5496"
BLUE_LIGHT = "BDD7EE"
BLUE_PALE  = "DEEAF1"
ORANGE_HDR = "C55A11"   # dark orange for CN header
ORANGE_MID = "F4B942"
ORANGE_PAL = "FCE4D6"
GREEN_DARK = "375623"
GREEN_MID  = "70AD47"
GREEN_PAL  = "E2EFDA"
GREY_LIGHT = "F2F2F2"
WHITE      = "FFFFFF"
BLACK      = "000000"
RED_FONT   = "C00000"
GREEN_FONT = "375623"

# Column style definitions
STYLE = {
    # (sub-header fill, sub-header font, qtotal fill, qtotal font, annual fill, annual font, value font)
    "normal": (BLUE_LIGHT,  BLUE_DARK,  BLUE_PALE,  BLUE_DARK,  BLUE_DARK,  WHITE,      BLACK),
    "cn":     (ORANGE_PAL,  ORANGE_HDR, ORANGE_PAL, ORANGE_HDR, ORANGE_HDR, WHITE,      RED_FONT),
    "net":    (GREEN_PAL,   GREEN_DARK, GREEN_PAL,  GREEN_DARK, GREEN_DARK, WHITE,      GREEN_FONT),
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


FMT_AMOUNT = '#,##0.00'


def set_cell(ws, row, col, value, bold=False, fill_hex=None,
             font_color=BLACK, size=11, number_format=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=size, name="Calibri")
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if number_format:
        cell.number_format = number_format
    cell.border = _border()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    return cell


def _style_formula_cell(ws, row, col, formula, style_key, bold=False, size=11):
    s = STYLE[style_key]
    fill_hex   = s[2]   # quarter total fill
    font_color = s[3]   # quarter total font
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font   = Font(bold=bold, color=font_color, size=size, name="Calibri")
    cell.fill   = _fill(fill_hex)
    cell.border = _border()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = FMT_AMOUNT
    return cell


def _style_annual_cell(ws, row, col, formula, style_key, size=12):
    s = STYLE[style_key]
    fill_hex   = s[4]
    font_color = s[5]
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font   = Font(bold=True, color=font_color, size=size, name="Calibri")
    cell.fill   = _fill(fill_hex)
    cell.border = _border()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = FMT_AMOUNT
    return cell


# ── Sheet Builder ────────────────────────────────────────────────────────────

def build_summary_sheet(wb, sheet_name: str, fy_years: list,
                        data_cols: list):
    """
    data_cols: list of (header_label, data_dict, style_key)
      style_key: "normal" | "cn" | "net"
      data_dict[fy_start][month] = float (always positive)
    """
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "B4"

    fy_labels     = [fy_label(y) for y in fy_years]
    num_fy        = len(fy_years)
    num_dc        = len(data_cols)
    total_cols    = num_fy * num_dc

    ws.column_dimensions["A"].width = 22
    for col in range(2, 2 + total_cols):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # ── Row 1: Sheet title ─────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value=sheet_name)
    title.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
    title.fill      = _fill(BLUE_DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + total_cols)

    # ── Row 2: FY group headers (merged over sub-columns) ─────────────
    ws.row_dimensions[2].height = 22
    set_cell(ws, 2, 1, "Month / Quarter", bold=True, fill_hex=BLUE_DARK, font_color=WHITE)
    for fi, fy_lbl in enumerate(fy_labels):
        sc = 2 + fi * num_dc
        ec = sc + num_dc - 1
        cell = ws.cell(row=2, column=sc, value=fy_lbl)
        cell.font      = Font(bold=True, color=WHITE, size=11, name="Calibri")
        cell.fill      = _fill(BLUE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
        if num_dc > 1:
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
            for c in range(sc + 1, ec + 1):
                ws.cell(row=2, column=c).fill   = _fill(BLUE_MID)
                ws.cell(row=2, column=c).border = _border()

    # ── Row 3: Sub-column headers ──────────────────────────────────────
    ws.row_dimensions[3].height = 22
    set_cell(ws, 3, 1, "", fill_hex=BLUE_LIGHT, font_color=BLACK)
    for fi in range(num_fy):
        for di, (lbl, _, sk) in enumerate(data_cols):
            c  = 2 + fi * num_dc + di
            sh = STYLE[sk][0]  # sub-header fill
            sf = STYLE[sk][1]  # sub-header font color
            set_cell(ws, 3, c, lbl, bold=True, fill_hex=sh, font_color=sf)

    current_row = 4
    annual_refs = {fy: defaultdict(list) for fy in fy_years}

    # ── Quarter groups ─────────────────────────────────────────────────
    for q_label, q_months in QUARTERS.items():
        ws.row_dimensions[current_row].height = 18
        set_cell(ws, current_row, 1, q_label, bold=True,
                 fill_hex=BLUE_MID, font_color=WHITE)
        for fi in range(num_fy):
            for di in range(num_dc):
                set_cell(ws, current_row, 2 + fi * num_dc + di, "", fill_hex=BLUE_MID)
        current_row += 1

        month_rows = []
        for idx, mo in enumerate(q_months):
            ws.row_dimensions[current_row].height = 16
            row_fill = WHITE if idx % 2 == 0 else GREY_LIGHT
            set_cell(ws, current_row, 1, MONTH_NAMES[mo],
                     fill_hex=row_fill, font_color=BLACK, align="left")

            for fi, fy in enumerate(fy_years):
                for di, (_, data_dict, sk) in enumerate(data_cols):
                    c   = 2 + fi * num_dc + di
                    val = data_dict[fy].get(mo, 0.0)
                    fnt = STYLE[sk][6]  # value font color
                    cell = ws.cell(row=current_row, column=c,
                                   value=val if val else None)
                    cell.font   = Font(color=fnt, name="Calibri")
                    cell.fill   = _fill(row_fill)
                    cell.border = _border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = FMT_AMOUNT
                    annual_refs[fy][di].append(current_row)
            month_rows.append(current_row)
            current_row += 1

        # Quarter Total row
        ws.row_dimensions[current_row].height = 18
        set_cell(ws, current_row, 1, f"{q_label} Total", bold=True,
                 fill_hex=BLUE_PALE, font_color=BLUE_DARK)
        for fi in range(num_fy):
            for di, (_, _, sk) in enumerate(data_cols):
                c       = 2 + fi * num_dc + di
                cl      = get_column_letter(c)
                formula = "=" + "+".join(f"{cl}{r}" for r in month_rows)
                _style_formula_cell(ws, current_row, c, formula, sk, bold=True)
        current_row += 1

    # ── Annual Total row ───────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 22
    set_cell(ws, current_row, 1, "Annual Total", bold=True, size=12,
             fill_hex=BLUE_DARK, font_color=WHITE)
    for fi in range(num_fy):
        for di, (_, _, sk) in enumerate(data_cols):
            c       = 2 + fi * num_dc + di
            cl      = get_column_letter(c)
            rows    = annual_refs[fy_years[fi]][di]
            formula = "=" + "+".join(f"{cl}{r}" for r in rows)
            _style_annual_cell(ws, current_row, c, formula, sk, size=12)

    return ws


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Sales Summary Report")
    parser.add_argument("--output", default=None,
                        help="Output Excel file path (default: auto-named in .tmp/reports/)")
    args = parser.parse_args()

    today = date.today()
    current_fy_start = today.year if today.month >= 4 else today.year - 1
    fy_years = [current_fy_start - 2, current_fy_start - 1, current_fy_start]
    print(f"Report period: {', '.join(fy_label(y) for y in fy_years)}")

    odoo = OdooConnector()
    print(f"Connected to Odoo: {odoo.url} / {odoo.db}")

    gross_sales, credit_notes, svc_gross, svc_cn, gds_gross, gds_cn = \
        fetch_sales_data(odoo, fy_years)

    # Pre-compute net dicts
    def net_dict(inv_d, cn_d):
        nd = _empty()
        for fy in fy_years:
            for mo in FY_MONTHS:
                nd[fy][mo] = inv_d[fy].get(mo, 0.0) - cn_d[fy].get(mo, 0.0)
        return nd

    net_sales = net_dict(gross_sales, credit_notes)
    svc_net   = net_dict(svc_gross,   svc_cn)
    gds_net   = net_dict(gds_gross,   gds_cn)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Total Sales Summary — 3 cols per FY
    build_summary_sheet(
        wb, "Sales Summary", fy_years,
        [
            ("Gross Sales (INR)", gross_sales,  "normal"),
            ("Credit Notes (INR)", credit_notes, "cn"),
            ("Net Sales (INR)",    net_sales,    "net"),
        ],
    )

    # Sheet 2: Service & Goods — 6 cols per FY
    build_summary_sheet(
        wb, "Service & Goods Breakdown", fy_years,
        [
            ("Service Gross (INR)",  svc_gross, "normal"),
            ("Service CN (INR)",     svc_cn,    "cn"),
            ("Service Net (INR)",    svc_net,   "net"),
            ("Goods Gross (INR)",    gds_gross, "normal"),
            ("Goods CN (INR)",       gds_cn,    "cn"),
            ("Goods Net (INR)",      gds_net,   "net"),
        ],
    )

    if args.output:
        out_path = Path(args.output)
    else:
        report_dir = PROJECT_ROOT / ".tmp" / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)
        out_path = report_dir / f"Sales_Summary_{date.today().isoformat()}.xlsx"

    wb.save(out_path)
    print(f"\nReport saved: {out_path}")
    return str(out_path)


if __name__ == "__main__":
    main()
