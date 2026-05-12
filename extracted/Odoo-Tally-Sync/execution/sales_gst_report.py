#!/usr/bin/env python3
"""
Sales Report with GST Breakup (CGST / SGST / IGST) — Excel Export

Connects to Odoo via XML-RPC, fetches posted sales invoices for a given
month, breaks down each invoice's tax lines into CGST, SGST, and IGST,
and writes the result to a formatted Excel workbook.

Usage:
    python execution/sales_gst_report.py                           # Current month
    python execution/sales_gst_report.py --month 1 --year 2026     # January 2026
    python execution/sales_gst_report.py --from 2026-01-01 --to 2026-01-31
"""

import os
import sys
import re
import argparse
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

from dotenv import load_dotenv

# ── Paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

# Add execution dir so we can import the connector
sys.path.insert(0, str(SCRIPT_DIR))

load_dotenv(PROJECT_ROOT / ".env")

from odoo_connector import OdooConnector  # noqa: E402

# ── Tax classification helpers ──────────────────────────────────────────

_CGST_RE = re.compile(r"\bCGST\b", re.IGNORECASE)

# ── HSN summary helpers ────────────────────────────────────────────────
def get_product_hsn(odoo, product_id):
    if not product_id:
        return ""
    pid = product_id[0] if isinstance(product_id, (list, tuple)) else product_id
    prod = odoo.read("product.product", [pid], ["id", "name", "default_code", "hsn_code", "l10n_in_hsn_code"])
    if prod:
        return prod[0].get("hsn_code") or prod[0].get("l10n_in_hsn_code") or ""
    return ""

def summarize_hsn_lines(odoo, invoices):
    hsn_summary = defaultdict(lambda: {
        "hsn_code": "",
        "description": "",
        "qty": 0,
        "taxable_value": 0,
        "cgst": 0,
        "sgst": 0,
        "igst": 0,
    })
    for inv in invoices:
        inv_id = inv["id"]
        lines = odoo.get_invoice_lines(inv_id)
        for line in lines:
            product_id = line.get("product_id")
            hsn_code = get_product_hsn(odoo, product_id)
            qty = line.get("quantity", 0)
            taxable = line.get("price_subtotal", 0)
            # Tax calculation
            cgst = sgst = igst = 0
            for tid in line.get("tax_ids", []):
                tax = odoo.get_tax_details([tid])[0]
                name = tax.get("name", "")
                rate = tax.get("amount", 0)
                if "CGST" in name:
                    cgst += round(taxable * rate / 100, 2)
                elif "SGST" in name:
                    sgst += round(taxable * rate / 100, 2)
                elif "IGST" in name:
                    igst += round(taxable * rate / 100, 2)
            key = hsn_code or "Unknown"
            hsn_summary[key]["hsn_code"] = key
            hsn_summary[key]["description"] = line.get("name", "")
            hsn_summary[key]["qty"] += qty
            hsn_summary[key]["taxable_value"] += taxable
            hsn_summary[key]["cgst"] += cgst
            hsn_summary[key]["sgst"] += sgst
            hsn_summary[key]["igst"] += igst
    return list(hsn_summary.values())

# ── Excel writer for HSN summary ───────────────────────────────────────
def write_hsn_excel(rows: list, output_path: Path, from_date: str, to_date: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HSN Summary"
    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    subtitle_font = Font(name="Calibri", size=11, color="666666")
    currency_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    # Title rows
    ws.merge_cells("A1:G1")
    ws["A1"] = "GSTR-1 HSN Summary Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Period: {from_date} to {to_date}  |  Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")
    # Column headers
    columns = [
        ("HSN/SAC", 14),
        ("Description", 30),
        ("Total Qty", 12),
        ("Taxable Value (₹)", 18),
        ("CGST Amount (₹)", 16),
        ("SGST Amount (₹)", 16),
        ("IGST Amount (₹)", 16),
    ]
    header_row = 4
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width
    # Data rows
    for i, row in enumerate(rows, 1):
        r = header_row + i
        ws.cell(row=r, column=1, value=row["hsn_code"]).border = thin_border
        ws.cell(row=r, column=2, value=row["description"]).border = thin_border
        ws.cell(row=r, column=3, value=row["qty"]).border = thin_border
        ws.cell(row=r, column=4, value=row["taxable_value"]).number_format = currency_fmt
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=5, value=row["cgst"]).number_format = currency_fmt
        ws.cell(row=r, column=5).border = thin_border
        ws.cell(row=r, column=6, value=row["sgst"]).number_format = currency_fmt
        ws.cell(row=r, column=6).border = thin_border
        ws.cell(row=r, column=7, value=row["igst"]).number_format = currency_fmt
        ws.cell(row=r, column=7).border = thin_border
    wb.save(output_path)

_SGST_RE = re.compile(r"\bSGST\b", re.IGNORECASE)
_IGST_RE = re.compile(r"\bIGST\b", re.IGNORECASE)


def classify_tax(tax_name: str) -> str:
    """Return 'CGST', 'SGST', 'IGST', or 'OTHER' based on the tax name."""
    if _CGST_RE.search(tax_name):
        return "CGST"
    if _SGST_RE.search(tax_name):
        return "SGST"
    if _IGST_RE.search(tax_name):
        return "IGST"
    return "OTHER"


# ── Core logic ──────────────────────────────────────────────────────

def fetch_sales_report(odoo: OdooConnector, from_date: str, to_date: str):
    """
    Fetch sales invoices and build report rows with GST breakup.

    Returns a list of dicts, one per invoice, with keys:
        invoice_number, date, customer, gstin, place_of_supply,
        taxable_value, cgst_rate, cgst_amount, sgst_rate, sgst_amount,
        igst_rate, igst_amount, total_tax, invoice_total
    """
    print(f"\nFetching posted sales invoices from {from_date} to {to_date} ...")
    invoices = odoo.get_sales_invoices(from_date=from_date, to_date=to_date)
    print(f"  Found {len(invoices)} posted sales invoice(s)")

    if not invoices:
        return []

    # (Removed broken partner_ids block)
    def main():
        parser = argparse.ArgumentParser(description="GSTR-1 HSN Summary Report")
        parser.add_argument("--month", type=int, default=None, help="Month (1-12)")
        parser.add_argument("--year", type=int, default=None, help="Year (YYYY)")
        parser.add_argument("--from", dest="from_date", type=str, default=None, help="Start date (YYYY-MM-DD)")
        parser.add_argument("--to", dest="to_date", type=str, default=None, help="End date (YYYY-MM-DD)")
        parser.add_argument("--output", type=str, default=None, help="Output Excel file path")
        args = parser.parse_args()

        # Date range logic
        if args.month and args.year:
            from_date = date(args.year, args.month, 1)
            to_date = date(args.year, args.month, 1).replace(day=28) + timedelta(days=4)
            to_date = to_date - timedelta(days=to_date.day)
        else:
            from_date = args.from_date or date.today().replace(day=1)
            to_date = args.to_date or date.today()
        from_date_str = str(from_date)
        to_date_str = str(to_date)

        odoo = OdooConnector()
        invoices = odoo.get_sales_invoices(from_date=from_date_str, to_date=to_date_str)

        print(f"Fetched {len(invoices)} sales invoices from Odoo.")

        hsn_rows = summarize_hsn_lines(odoo, invoices)
        output_path = args.output or TMP_DIR / f"hsn_summary_{from_date_str}_to_{to_date_str}.xlsx"
        write_hsn_excel(hsn_rows, output_path, from_date_str, to_date_str)
        print(f"HSN summary report written to {output_path}")
        tax_rates: dict = {}                     # {tax_category: rate %}

        for line in lines:
            tax_line_ref = line.get("tax_line_id")
            if tax_line_ref:
                tax_id = tax_line_ref[0]
                tax_name = tax_line_ref[1] if len(tax_line_ref) > 1 else ""

                # Get full tax detail if not cached
                if tax_id not in tax_cache:
                    details = odoo.get_tax_details([tax_id])
                    tax_cache[tax_id] = details[0] if details else {"name": tax_name, "amount": 0}

                tax_detail = tax_cache[tax_id]
                category = classify_tax(tax_detail.get("name", tax_name))
                amount = abs(line.get("balance", 0))
                tax_amounts[category] += amount
                if category not in tax_rates:
                    tax_rates[category] = tax_detail.get("amount", 0)

        # If no tax lines found via tax_line_id, fall back to computing from product lines
        if not tax_amounts:
            for line in lines:
                if line.get("tax_ids") and not line.get("tax_line_id"):
                    subtotal = abs(line.get("price_subtotal", 0))
                    # Fetch tax details
                    for tid in line["tax_ids"]:
                        if tid not in tax_cache:
                            details = odoo.get_tax_details([tid])
                            tax_cache[tid] = details[0] if details else {"name": "", "amount": 0}
                        td = tax_cache[tid]
                        category = classify_tax(td.get("name", ""))
                        rate = td.get("amount", 0)
                        computed = round(subtotal * rate / 100, 2)
                        tax_amounts[category] += computed
                        if category not in tax_rates:
                            tax_rates[category] = rate

        cgst_amt = round(tax_amounts.get("CGST", 0), 2)
        sgst_amt = round(tax_amounts.get("SGST", 0), 2)
        igst_amt = round(tax_amounts.get("IGST", 0), 2)
        other_tax = round(tax_amounts.get("OTHER", 0), 2)

        total_tax = round(cgst_amt + sgst_amt + igst_amt + other_tax, 2)

        rows.append({
            "invoice_number": inv_name,
            "date": inv_date,
            "customer": partner_name,
            "gstin": gstin,
            "place_of_supply": place_of_supply,
            "hsn_sac": "",  # Can be enriched per-line if needed
            "taxable_value": inv.get("amount_untaxed", 0),
            "cgst_rate": tax_rates.get("CGST", 0),
            "cgst_amount": cgst_amt,
            "sgst_rate": tax_rates.get("SGST", 0),
            "sgst_amount": sgst_amt,
            "igst_rate": tax_rates.get("IGST", 0),
            "igst_amount": igst_amt,
            "other_tax": other_tax,
            "total_tax": total_tax,
            "invoice_total": inv.get("amount_total", 0),
            "payment_status": inv.get("payment_state", ""),
        })

        if idx % 10 == 0:
            print(f"  Processed {idx}/{len(invoices)} invoices …")

    print(f"  Done. {len(rows)} invoice rows ready for export.\n")
    return rows


# ── Excel writer ────────────────────────────────────────────────────

def write_excel(rows: list, output_path: Path, from_date: str, to_date: str):
    """Write report rows to a formatted Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales GST Report"

    # ── Styles ──────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    subtitle_font = Font(name="Calibri", size=11, color="666666")
    currency_fmt = '#,##0.00'
    pct_fmt = '0.00"%"'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, size=11)

    # ── Title rows ──────────────────────────────────────────────────
    ws.merge_cells("A1:Q1")
    ws["A1"] = "Sales Report with GST Breakup"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:Q2")
    ws["A2"] = f"Period: {from_date} to {to_date}  |  Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Column headers ──────────────────────────────────────────────
    columns = [
        ("S.No", 6),
        ("Invoice No.", 18),
        ("Date", 12),
        ("Customer Name", 30),
        ("GSTIN/UIN", 18),
        ("Place of Supply", 20),
        ("HSN/SAC", 12),
        ("Taxable Value (₹)", 18),
        ("CGST Rate %", 12),
        ("CGST Amount (₹)", 16),
        ("SGST Rate %", 12),
        ("SGST Amount (₹)", 16),
        ("IGST Rate %", 12),
        ("IGST Amount (₹)", 16),
        ("Total Tax (₹)", 16),
        ("Invoice Total (₹)", 18),
        ("Payment Status", 15),
    ]

    header_row = 4
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width

    # ── Data rows ───────────────────────────────────────────────────
    for i, row in enumerate(rows, 1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i).border = thin_border
        ws.cell(row=r, column=2, value=row["invoice_number"]).border = thin_border
        ws.cell(row=r, column=3, value=row["date"]).border = thin_border
        ws.cell(row=r, column=4, value=row["customer"]).border = thin_border
        ws.cell(row=r, column=5, value=row["gstin"]).border = thin_border
        ws.cell(row=r, column=6, value=row["place_of_supply"]).border = thin_border
        ws.cell(row=r, column=7, value=row["hsn_sac"]).border = thin_border

        for col, key, fmt in [
            (8, "taxable_value", currency_fmt),
            (9, "cgst_rate", pct_fmt),
            (10, "cgst_amount", currency_fmt),
            (11, "sgst_rate", pct_fmt),
            (12, "sgst_amount", currency_fmt),
            (13, "igst_rate", pct_fmt),
            (14, "igst_amount", currency_fmt),
            (15, "total_tax", currency_fmt),
            (16, "invoice_total", currency_fmt),
        ]:
            cell = ws.cell(row=r, column=col, value=row[key])
            cell.number_format = fmt
            cell.border = thin_border

        ws.cell(row=r, column=17, value=row["payment_status"]).border = thin_border

    # ── Totals row ──────────────────────────────────────────────────
    if rows:
        total_r = header_row + len(rows) + 1
        ws.cell(row=total_r, column=1, value="").border = thin_border
        ws.merge_cells(start_row=total_r, start_column=2, end_row=total_r, end_column=7)
        total_label = ws.cell(row=total_r, column=2, value="TOTAL")
        total_label.font = total_font
        total_label.fill = total_fill
        total_label.alignment = Alignment(horizontal="right")
        total_label.border = thin_border
        # Apply border to merged cells
        for c in range(3, 8):
            ws.cell(row=total_r, column=c).border = thin_border
            ws.cell(row=total_r, column=c).fill = total_fill

        sum_fields = [
            (8, "taxable_value"),
            (10, "cgst_amount"),
            (12, "sgst_amount"),
            (14, "igst_amount"),
            (15, "total_tax"),
            (16, "invoice_total"),
        ]
        for col, key in sum_fields:
            total_val = round(sum(r[key] for r in rows), 2)
            cell = ws.cell(row=total_r, column=col, value=total_val)
            cell.number_format = currency_fmt
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border

        # Empty styled cells for rate columns and others in total row
        for c in [1, 9, 11, 13, 17]:
            cell = ws.cell(row=total_r, column=c)
            cell.fill = total_fill
            cell.border = thin_border

    # ── Freeze & filter ─────────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:Q{header_row + len(rows)}"

    # ── Save ────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"✓ Excel report saved: {output_path}")
    print(f"  Invoices: {len(rows)}")
    if rows:
        print(f"  Total taxable: ₹{sum(r['taxable_value'] for r in rows):,.2f}")
        print(f"  Total CGST:    ₹{sum(r['cgst_amount'] for r in rows):,.2f}")
        print(f"  Total SGST:    ₹{sum(r['sgst_amount'] for r in rows):,.2f}")
        print(f"  Total IGST:    ₹{sum(r['igst_amount'] for r in rows):,.2f}")
        print(f"  Grand Total:   ₹{sum(r['invoice_total'] for r in rows):,.2f}")


# ── CLI ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export Odoo Sales Report with GST breakup to Excel"
    )
    parser.add_argument("--month", type=int, default=None,
                        help="Month number (1-12). Defaults to current month.")
    parser.add_argument("--year", type=int, default=None,
                        help="Year (e.g. 2026). Defaults to current year.")
    parser.add_argument("--from", dest="date_from", type=str,
                        help="Start date YYYY-MM-DD (overrides --month/--year)")
    parser.add_argument("--to", dest="date_to", type=str,
                        help="End date YYYY-MM-DD (overrides --month/--year)")
    parser.add_argument("--output", type=str, default=None,
                        help="Output Excel file path")
    args = parser.parse_args()

    # Determine date range
    if args.date_from and args.date_to:
        from_date = args.date_from
        to_date = args.date_to
    else:
        year = args.year or date.today().year
        month = args.month or date.today().month
        from_date = f"{year}-{month:02d}-01"
        # Last day of month
        if month == 12:
            to_date = f"{year}-12-31"
        else:
            next_month_first = date(year, month + 1, 1)
            last_day = next_month_first.replace(day=1).toordinal() - 1
            to_date = date.fromordinal(last_day).isoformat()

    # Output path
    if args.output:
        output_path = Path(args.output)
    else:
        period_label = f"{from_date}_to_{to_date}"
        output_path = TMP_DIR / f"Sales_GST_Report_{period_label}.xlsx"

    # Connect
    try:
        odoo = OdooConnector()
        ver = odoo.version()
        print(f"✓ Connected to Odoo {ver.get('server_version', '?')}")
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    # Fetch & export
    # (Removed old fetch_sales_report logic; only HSN summary logic runs)


if __name__ == "__main__":
    main()
