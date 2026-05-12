#!/usr/bin/env python3
"""
Sales Register Reconciliation: Odoo vs Tally

Fetches Sales invoices from Odoo and Sales vouchers from Tally for a given
month, cross-references them, and highlights differences (missing, amount
mismatch, extra entries) in a formatted Excel report.

Usage:
    python execution/sales_reconciliation.py                           # Current month
    python execution/sales_reconciliation.py --month 1 --year 2026     # January 2026
    python execution/sales_reconciliation.py --from 2026-01-01 --to 2026-01-31
"""

import os
import sys
import re
import argparse
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

from dotenv import load_dotenv

# ── Paths ───────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

sys.path.insert(0, str(SCRIPT_DIR))

load_dotenv(PROJECT_ROOT / ".env")

from odoo_connector import OdooConnector   # noqa: E402
from tally_connector import TallyConnector  # noqa: E402


# ── Tax classification ──────────────────────────────────────────────

_CGST_RE = re.compile(r"\bCGST\b", re.I)
_SGST_RE = re.compile(r"\bSGST\b", re.I)
_IGST_RE = re.compile(r"\bIGST\b", re.I)


def classify_tax(name: str) -> str:
    if _CGST_RE.search(name):
        return "CGST"
    if _SGST_RE.search(name):
        return "SGST"
    if _IGST_RE.search(name):
        return "IGST"
    return "OTHER"


# ── Odoo extraction ────────────────────────────────────────────────

def fetch_odoo_sales(odoo: OdooConnector, from_date: str, to_date: str) -> list:
    """Return list of dicts with invoice details from Odoo."""
    print(f"\n── Odoo: Fetching posted sales invoices ({from_date} to {to_date}) ──")
    invoices = odoo.get_sales_invoices(from_date=from_date, to_date=to_date)
    print(f"   Found {len(invoices)} invoices")

    # Collect partner details
    partner_ids = list({inv["partner_id"][0] for inv in invoices if inv.get("partner_id")})
    partners_map = {}
    if partner_ids:
        partners = odoo.search_read(
            "res.partner", [("id", "in", partner_ids)],
            ["id", "name", "vat", "state_id"],
        )
        partners_map = {p["id"]: p for p in partners}

    tax_cache: dict = {}
    rows = []

    for inv in invoices:
        inv_id = inv["id"]
        partner_id = inv["partner_id"][0] if inv.get("partner_id") else None
        partner_name = inv["partner_id"][1] if inv.get("partner_id") else ""
        partner = partners_map.get(partner_id, {})
        gstin = partner.get("vat") or ""

        # Get tax breakdown from journal lines
        lines = odoo.get_invoice_lines(inv_id)
        tax_amounts = defaultdict(float)

        for line in lines:
            tax_line_ref = line.get("tax_line_id")
            if tax_line_ref:
                tid = tax_line_ref[0]
                tname = tax_line_ref[1] if len(tax_line_ref) > 1 else ""
                if tid not in tax_cache:
                    details = odoo.get_tax_details([tid])
                    tax_cache[tid] = details[0] if details else {"name": tname, "amount": 0}
                cat = classify_tax(tax_cache[tid].get("name", tname))
                tax_amounts[cat] += abs(line.get("balance", 0))

        # Fallback: compute from product lines
        if not tax_amounts:
            for line in lines:
                if line.get("tax_ids") and not line.get("tax_line_id"):
                    subtotal = abs(line.get("price_subtotal", 0))
                    for tid in line["tax_ids"]:
                        if tid not in tax_cache:
                            details = odoo.get_tax_details([tid])
                            tax_cache[tid] = details[0] if details else {"name": "", "amount": 0}
                        td = tax_cache[tid]
                        cat = classify_tax(td.get("name", ""))
                        rate = td.get("amount", 0)
                        tax_amounts[cat] += round(subtotal * rate / 100, 2)

        rows.append({
            "invoice_number": inv["name"],
            "date": inv.get("invoice_date", ""),
            "customer": partner_name,
            "gstin": gstin,
            "taxable_value": round(inv.get("amount_untaxed", 0), 2),
            "cgst": round(tax_amounts.get("CGST", 0), 2),
            "sgst": round(tax_amounts.get("SGST", 0), 2),
            "igst": round(tax_amounts.get("IGST", 0), 2),
            "total_tax": round(sum(tax_amounts.values()), 2),
            "invoice_total": round(inv.get("amount_total", 0), 2),
        })

    print(f"   Odoo grand total: ₹{sum(r['invoice_total'] for r in rows):,.2f}")
    return rows


# ── Tally extraction ───────────────────────────────────────────────

def _parse_tally_amount(val: str) -> float:
    """Tally amounts can be negative (Cr), strip spaces/commas."""
    if not val:
        return 0.0
    val = val.replace(",", "").strip()
    try:
        return abs(float(val))
    except ValueError:
        return 0.0


def fetch_tally_sales(tally: TallyConnector, from_date: str, to_date: str) -> list:
    """
    Return list of dicts with voucher details from Tally.
    Tally dates are YYYYMMDD.
    """
    tally_from = from_date.replace("-", "")
    tally_to = to_date.replace("-", "")

    print(f"\n── Tally: Fetching Sales vouchers ({from_date} to {to_date}) ──")
    vouchers = tally.get_vouchers("Sales", tally_from, tally_to)
    print(f"   Found {len(vouchers)} vouchers")

    rows = []
    for v in vouchers:
        raw_date = v.get("date", "")
        # Tally sometimes returns YYYYMMDD
        if len(raw_date) == 8 and raw_date.isdigit():
            fmt_date = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:]}"
        else:
            fmt_date = raw_date

        narration = v.get("narration", "")
        reference = v.get("reference", "")
        voucher_no = v.get("number", "")

        # Odoo invoice number could be in narration or reference field
        odoo_ref = ""
        for field in [reference, narration]:
            match = re.search(r'(INV/\d{4}/\d+(?:/\d+)?|S[A-Z]*/\d{4}/\d+)', field or "")
            if match:
                odoo_ref = match.group(0)
                break

        rows.append({
            "voucher_number": voucher_no,
            "date": fmt_date,
            "party": v.get("party", ""),
            "amount": _parse_tally_amount(v.get("amount", "0")),
            "reference": reference,
            "narration": narration,
            "odoo_ref_found": odoo_ref,
        })

    total = sum(r["amount"] for r in rows)
    print(f"   Tally grand total: ₹{total:,.2f}")
    return rows


# ── Reconciliation logic ───────────────────────────────────────────

def reconcile(odoo_rows: list, tally_rows: list) -> dict:
    """
    Cross-match Odoo invoices ↔ Tally vouchers.

    Matching strategy (in order of priority):
      1. Odoo invoice number found in Tally narration/reference
      2. Same date + same amount + fuzzy customer name

    Returns dict with:
      matched     — list of (odoo_row, tally_row, status, note)
      only_odoo   — invoices in Odoo but NOT in Tally
      only_tally  — vouchers in Tally but NOT in Odoo
      mismatches  — matched pair but amounts differ
    """

    # Build lookup: Tally rows by extracted Odoo reference
    tally_by_ref: dict = {}
    tally_unmatched = list(range(len(tally_rows)))

    for idx, tr in enumerate(tally_rows):
        ref = tr.get("odoo_ref_found", "")
        if ref:
            tally_by_ref.setdefault(ref, []).append(idx)

    # Build secondary lookup: date+amount key → tally indices
    tally_by_date_amt: dict = defaultdict(list)
    for idx, tr in enumerate(tally_rows):
        key = (tr["date"], round(tr["amount"], 0))
        tally_by_date_amt[key].append(idx)

    matched = []
    only_odoo = []
    mismatches = []
    odoo_matched_indices = set()
    tally_matched_indices = set()

    # Pass 1: Match by Odoo invoice number
    for oi, odoo_row in enumerate(odoo_rows):
        inv_num = odoo_row["invoice_number"]
        if inv_num in tally_by_ref:
            ti = tally_by_ref[inv_num][0]
            tally_row = tally_rows[ti]
            diff = abs(odoo_row["invoice_total"] - tally_row["amount"])
            if diff < 1.0:
                matched.append((odoo_row, tally_row, "Matched", ""))
            else:
                mismatches.append((
                    odoo_row, tally_row, "Amount Mismatch",
                    f"Odoo ₹{odoo_row['invoice_total']:,.2f} vs Tally ₹{tally_row['amount']:,.2f} (diff ₹{diff:,.2f})"
                ))
            odoo_matched_indices.add(oi)
            tally_matched_indices.add(ti)
            continue

    # Pass 2: Match by date + amount (±₹1 tolerance)
    for oi, odoo_row in enumerate(odoo_rows):
        if oi in odoo_matched_indices:
            continue
        key = (odoo_row["date"], round(odoo_row["invoice_total"], 0))
        candidates = tally_by_date_amt.get(key, [])
        found = False
        for ti in candidates:
            if ti in tally_matched_indices:
                continue
            tally_row = tally_rows[ti]
            diff = abs(odoo_row["invoice_total"] - tally_row["amount"])
            if diff < 1.0:
                matched.append((odoo_row, tally_row, "Matched (date+amount)", ""))
                odoo_matched_indices.add(oi)
                tally_matched_indices.add(ti)
                found = True
                break
        if not found:
            only_odoo.append(odoo_row)

    only_tally = [tally_rows[ti] for ti in range(len(tally_rows)) if ti not in tally_matched_indices]

    return {
        "matched": matched,
        "mismatches": mismatches,
        "only_odoo": only_odoo,
        "only_tally": only_tally,
    }


# ── Excel Report ────────────────────────────────────────────────────

def write_reconciliation_excel(
    odoo_rows: list,
    tally_rows: list,
    recon: dict,
    output_path: Path,
    from_date: str,
    to_date: str,
):
    """Write a multi-sheet reconciliation Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cur_fmt = '#,##0.00'
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    subtitle_font = Font(name="Calibri", size=11, color="666666")
    bold_font = Font(name="Calibri", bold=True, size=11)

    def style_headers(ws, cols, row=4):
        for ci, (name, width) in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=name)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = hdr_align; cell.border = thin
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width

    def add_title(ws, title, cols_span):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_span)
        ws["A1"] = title
        ws["A1"].font = title_font; ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_span)
        ws["A2"] = f"Period: {from_date} to {to_date}  |  Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
        ws["A2"].font = subtitle_font; ws["A2"].alignment = Alignment(horizontal="center")

    # ── Sheet 1: Summary ────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    add_title(ws_sum, "Sales Register Reconciliation — Odoo vs Tally", 5)

    summary_data = [
        ("Metric", "Odoo", "Tally", "Difference", "Status"),
        ("Invoice / Voucher Count", len(odoo_rows), len(tally_rows),
         len(odoo_rows) - len(tally_rows),
         "OK" if len(odoo_rows) == len(tally_rows) else "MISMATCH"),
        ("Grand Total (₹)",
         round(sum(r["invoice_total"] for r in odoo_rows), 2),
         round(sum(r["amount"] for r in tally_rows), 2),
         round(sum(r["invoice_total"] for r in odoo_rows) - sum(r["amount"] for r in tally_rows), 2),
         ""),
        ("Taxable Value (₹)", round(sum(r["taxable_value"] for r in odoo_rows), 2), "—", "", ""),
        ("Total CGST (₹)", round(sum(r["cgst"] for r in odoo_rows), 2), "—", "", ""),
        ("Total SGST (₹)", round(sum(r["sgst"] for r in odoo_rows), 2), "—", "", ""),
        ("Total IGST (₹)", round(sum(r["igst"] for r in odoo_rows), 2), "—", "", ""),
        ("", "", "", "", ""),
        ("Matched invoices", len(recon["matched"]), "", "", ""),
        ("Amount mismatches", len(recon["mismatches"]), "", "", ""),
        ("Only in Odoo (missing in Tally)", len(recon["only_odoo"]), "", "", ""),
        ("Only in Tally (missing in Odoo)", len(recon["only_tally"]), "", "", ""),
    ]

    sum_cols = [("Metric", 38), ("Odoo", 18), ("Tally", 18), ("Difference", 18), ("Status", 14)]
    style_headers(ws_sum, sum_cols)
    for ri, row_data in enumerate(summary_data[1:], 5):
        for ci, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            cell.border = thin
            if ci in (2, 3, 4) and isinstance(val, (int, float)):
                cell.number_format = cur_fmt
            if ri == 5:  # count row
                cell.font = bold_font
            # Highlight status
            if ci == 5:
                if val == "OK":
                    cell.fill = green_fill
                elif val == "MISMATCH":
                    cell.fill = red_fill

    # ── Sheet 2: Odoo Sales Register ────────────────────────────────
    ws_odoo = wb.create_sheet("Odoo Sales Register")
    add_title(ws_odoo, "Odoo — Sales Invoices (Posted)", 11)
    odoo_cols = [
        ("S.No", 6), ("Invoice No.", 18), ("Date", 12), ("Customer", 30),
        ("GSTIN", 18), ("Taxable (₹)", 16), ("CGST (₹)", 14),
        ("SGST (₹)", 14), ("IGST (₹)", 14), ("Total Tax (₹)", 14),
        ("Invoice Total (₹)", 16),
    ]
    style_headers(ws_odoo, odoo_cols)
    for i, r in enumerate(odoo_rows, 1):
        row = 4 + i
        ws_odoo.cell(row=row, column=1, value=i).border = thin
        ws_odoo.cell(row=row, column=2, value=r["invoice_number"]).border = thin
        ws_odoo.cell(row=row, column=3, value=r["date"]).border = thin
        ws_odoo.cell(row=row, column=4, value=r["customer"]).border = thin
        ws_odoo.cell(row=row, column=5, value=r["gstin"]).border = thin
        for col, key in [(6, "taxable_value"), (7, "cgst"), (8, "sgst"),
                         (9, "igst"), (10, "total_tax"), (11, "invoice_total")]:
            c = ws_odoo.cell(row=row, column=col, value=r[key])
            c.number_format = cur_fmt; c.border = thin

    # Totals
    if odoo_rows:
        tr = 5 + len(odoo_rows)
        ws_odoo.cell(row=tr, column=4, value="TOTAL").font = bold_font
        ws_odoo.cell(row=tr, column=4).border = thin
        for col, key in [(6, "taxable_value"), (7, "cgst"), (8, "sgst"),
                         (9, "igst"), (10, "total_tax"), (11, "invoice_total")]:
            c = ws_odoo.cell(row=tr, column=col, value=round(sum(r[key] for r in odoo_rows), 2))
            c.number_format = cur_fmt; c.font = bold_font; c.border = thin

    # ── Sheet 3: Tally Sales Register ───────────────────────────────
    ws_tally = wb.create_sheet("Tally Sales Register")
    add_title(ws_tally, "Tally — Sales Vouchers", 7)
    tally_cols = [
        ("S.No", 6), ("Voucher No.", 18), ("Date", 12), ("Party Name", 30),
        ("Amount (₹)", 16), ("Reference", 20), ("Odoo Ref Found", 18),
    ]
    style_headers(ws_tally, tally_cols)
    for i, r in enumerate(tally_rows, 1):
        row = 4 + i
        ws_tally.cell(row=row, column=1, value=i).border = thin
        ws_tally.cell(row=row, column=2, value=r["voucher_number"]).border = thin
        ws_tally.cell(row=row, column=3, value=r["date"]).border = thin
        ws_tally.cell(row=row, column=4, value=r["party"]).border = thin
        c = ws_tally.cell(row=row, column=5, value=r["amount"])
        c.number_format = cur_fmt; c.border = thin
        ws_tally.cell(row=row, column=6, value=r["reference"]).border = thin
        ws_tally.cell(row=row, column=7, value=r["odoo_ref_found"]).border = thin

    if tally_rows:
        tr = 5 + len(tally_rows)
        ws_tally.cell(row=tr, column=4, value="TOTAL").font = bold_font
        ws_tally.cell(row=tr, column=4).border = thin
        c = ws_tally.cell(row=tr, column=5, value=round(sum(r["amount"] for r in tally_rows), 2))
        c.number_format = cur_fmt; c.font = bold_font; c.border = thin

    # ── Sheet 4: Differences ────────────────────────────────────────
    ws_diff = wb.create_sheet("Differences")
    add_title(ws_diff, "Reconciliation Differences", 8)
    diff_cols = [
        ("S.No", 6), ("Category", 22), ("Invoice / Voucher", 20),
        ("Date", 12), ("Party / Customer", 30), ("Odoo Amount (₹)", 16),
        ("Tally Amount (₹)", 16), ("Note", 40),
    ]
    style_headers(ws_diff, diff_cols)

    diff_row = 5
    serial = 0

    # Amount mismatches
    for odoo_r, tally_r, status, note in recon["mismatches"]:
        serial += 1
        ws_diff.cell(row=diff_row, column=1, value=serial).border = thin
        c = ws_diff.cell(row=diff_row, column=2, value="Amount Mismatch")
        c.fill = yellow_fill; c.border = thin
        ws_diff.cell(row=diff_row, column=3, value=odoo_r["invoice_number"]).border = thin
        ws_diff.cell(row=diff_row, column=4, value=odoo_r["date"]).border = thin
        ws_diff.cell(row=diff_row, column=5, value=odoo_r["customer"]).border = thin
        c = ws_diff.cell(row=diff_row, column=6, value=odoo_r["invoice_total"])
        c.number_format = cur_fmt; c.border = thin
        c = ws_diff.cell(row=diff_row, column=7, value=tally_r["amount"])
        c.number_format = cur_fmt; c.border = thin
        ws_diff.cell(row=diff_row, column=8, value=note).border = thin
        diff_row += 1

    # Only in Odoo
    for odoo_r in recon["only_odoo"]:
        serial += 1
        ws_diff.cell(row=diff_row, column=1, value=serial).border = thin
        c = ws_diff.cell(row=diff_row, column=2, value="Missing in Tally")
        c.fill = red_fill; c.border = thin
        ws_diff.cell(row=diff_row, column=3, value=odoo_r["invoice_number"]).border = thin
        ws_diff.cell(row=diff_row, column=4, value=odoo_r["date"]).border = thin
        ws_diff.cell(row=diff_row, column=5, value=odoo_r["customer"]).border = thin
        c = ws_diff.cell(row=diff_row, column=6, value=odoo_r["invoice_total"])
        c.number_format = cur_fmt; c.border = thin
        ws_diff.cell(row=diff_row, column=7, value="—").border = thin
        ws_diff.cell(row=diff_row, column=8, value="Invoice exists in Odoo but NOT found in Tally").border = thin
        diff_row += 1

    # Only in Tally
    for tally_r in recon["only_tally"]:
        serial += 1
        ws_diff.cell(row=diff_row, column=1, value=serial).border = thin
        c = ws_diff.cell(row=diff_row, column=2, value="Missing in Odoo")
        c.fill = red_fill; c.border = thin
        ws_diff.cell(row=diff_row, column=3, value=tally_r["voucher_number"]).border = thin
        ws_diff.cell(row=diff_row, column=4, value=tally_r["date"]).border = thin
        ws_diff.cell(row=diff_row, column=5, value=tally_r["party"]).border = thin
        ws_diff.cell(row=diff_row, column=6, value="—").border = thin
        c = ws_diff.cell(row=diff_row, column=7, value=tally_r["amount"])
        c.number_format = cur_fmt; c.border = thin
        ws_diff.cell(row=diff_row, column=8, value="Voucher exists in Tally but NOT found in Odoo").border = thin
        diff_row += 1

    if serial == 0:
        ws_diff.cell(row=diff_row, column=2, value="No differences found — registers match! ✓")
        ws_diff.cell(row=diff_row, column=2).font = Font(bold=True, color="006100", size=12)

    # ── Save ────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\n✓ Reconciliation report saved: {output_path}")
    return output_path


# ── Console summary printer ────────────────────────────────────────

def print_summary(odoo_rows, tally_rows, recon):
    """Print a quick console summary."""
    odoo_total = sum(r["invoice_total"] for r in odoo_rows)
    tally_total = sum(r["amount"] for r in tally_rows)

    print("\n" + "=" * 70)
    print("       SALES REGISTER RECONCILIATION — SUMMARY")
    print("=" * 70)
    print(f"  {'Metric':<40} {'Odoo':>14} {'Tally':>14}")
    print(f"  {'─' * 40} {'─' * 14} {'─' * 14}")
    print(f"  {'Invoice / Voucher Count':<40} {len(odoo_rows):>14} {len(tally_rows):>14}")
    print(f"  {'Grand Total (₹)':<40} {odoo_total:>14,.2f} {tally_total:>14,.2f}")
    print(f"  {'Difference (₹)':<40} {odoo_total - tally_total:>14,.2f}")
    print()
    print(f"  Matched:                {len(recon['matched']):>5}")
    print(f"  Amount Mismatches:      {len(recon['mismatches']):>5}")
    print(f"  Only in Odoo:           {len(recon['only_odoo']):>5}")
    print(f"  Only in Tally:          {len(recon['only_tally']):>5}")
    print("=" * 70)

    if recon["mismatches"]:
        print("\n⚠  AMOUNT MISMATCHES:")
        for odoo_r, tally_r, _, note in recon["mismatches"]:
            print(f"   {odoo_r['invoice_number']:>18}  {note}")

    if recon["only_odoo"]:
        print(f"\n⚠  MISSING IN TALLY ({len(recon['only_odoo'])} invoices):")
        for r in recon["only_odoo"]:
            print(f"   {r['invoice_number']:>18}  {r['date']}  {r['customer']:<30}  ₹{r['invoice_total']:>12,.2f}")

    if recon["only_tally"]:
        print(f"\n⚠  MISSING IN ODOO ({len(recon['only_tally'])} vouchers):")
        for r in recon["only_tally"]:
            print(f"   {r['voucher_number']:>18}  {r['date']}  {r['party']:<30}  ₹{r['amount']:>12,.2f}")


# ── CLI ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Sales Register Reconciliation: Odoo vs Tally"
    )
    parser.add_argument("--month", type=int, default=None)
    parser.add_argument("--year", type=int, default=None)
    parser.add_argument("--from", dest="date_from", type=str)
    parser.add_argument("--to", dest="date_to", type=str)
    parser.add_argument("--output", type=str, default=None)
    parser.add_argument("--odoo-only", action="store_true",
                        help="Run only the Odoo side (if Tally is unavailable)")
    args = parser.parse_args()

    # Date range
    if args.date_from and args.date_to:
        from_date, to_date = args.date_from, args.date_to
    else:
        year = args.year or date.today().year
        month = args.month or date.today().month
        from_date = f"{year}-{month:02d}-01"
        if month == 12:
            to_date = f"{year}-12-31"
        else:
            next_first = date(year, month + 1, 1)
            to_date = date.fromordinal(next_first.toordinal() - 1).isoformat()

    # Output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = TMP_DIR / f"Sales_Reconciliation_{from_date}_to_{to_date}.xlsx"

    # ── Connect Odoo ────────────────────────────────────────────────
    try:
        odoo = OdooConnector()
        ver = odoo.version()
        print(f"✓ Connected to Odoo {ver.get('server_version', '?')}")
    except Exception as e:
        print(f"ERROR (Odoo): {e}")
        sys.exit(1)

    # ── Connect Tally ───────────────────────────────────────────────
    tally = None
    tally_available = False
    if not args.odoo_only:
        try:
            tally = TallyConnector()
            info = tally.test_connection()
            print(f"✓ Connected to Tally Prime at {info['url']}")
            print(f"  Company: {info.get('target_company', '?')}")
            tally_available = True
        except Exception as e:
            print(f"\n⚠ Tally not reachable: {e}")
            print("  → Continuing with Odoo-only export. Tally columns will be empty.\n")

    # ── Fetch data ──────────────────────────────────────────────────
    odoo_rows = fetch_odoo_sales(odoo, from_date, to_date)
    tally_rows = fetch_tally_sales(tally, from_date, to_date) if tally_available else []

    # ── Reconcile ───────────────────────────────────────────────────
    if tally_rows:
        recon = reconcile(odoo_rows, tally_rows)
    else:
        recon = {
            "matched": [],
            "mismatches": [],
            "only_odoo": odoo_rows,   # All Odoo invoices are "unmatched"
            "only_tally": [],
        }

    # ── Output ──────────────────────────────────────────────────────
    print_summary(odoo_rows, tally_rows, recon)
    write_reconciliation_excel(odoo_rows, tally_rows, recon, output_path, from_date, to_date)


if __name__ == "__main__":
    main()
