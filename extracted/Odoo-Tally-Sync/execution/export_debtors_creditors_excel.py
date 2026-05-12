#!/usr/bin/env python3
"""
Export Sundry Debtors & Sundry Creditors from Tally to Excel

Fetches all ledgers under the Sundry Debtors and Sundry Creditors groups
from Tally Prime, including their closing balance for a given date range,
and writes them to an Excel file with two separate sheets.

Usage:
    python execution/export_debtors_creditors_excel.py
    python execution/export_debtors_creditors_excel.py --as-of 20260331
    python execution/export_debtors_creditors_excel.py --from-date 20250401 --to-date 20260331
    python execution/export_debtors_creditors_excel.py --output MyReport.xlsx
    python execution/export_debtors_creditors_excel.py --test
"""

import os
import re
import sys
import argparse
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

sys.path.insert(0, str(SCRIPT_DIR))
from tally_connector import TallyConnector

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── XML helpers ──────────────────────────────────────────────────────

def sanitize_xml(xml_text: str) -> str:
    xml_text = re.sub(r'&#x([0-8BbCcEeFf]);', '', xml_text)
    xml_text = re.sub(r'&#x1[0-9A-Fa-f];', '', xml_text)
    xml_text = re.sub(r'&#([0-8]|1[0-1]|1[4-9]|2[0-9]|3[01]);', '', xml_text)
    xml_text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', xml_text)
    xml_text = re.sub(r'<(/?)(\w+):', r'<\1\2_', xml_text)
    xml_text = re.sub(r'\s+xmlns:\w+="[^"]*"', '', xml_text)
    return xml_text


def parse_tally_amount(text: str) -> float:
    """
    Parse a Tally amount string to a signed float.

    Tally returns amounts with a sign prefix or ' Dr' / ' Cr' suffix:
      '-12345.67'  to debit (asset)  to negative in Tally convention
      '12345.67'   to credit (liability) to positive

    For Sundry Debtors (assets): closing balance is Dr to stored negative in Tally.
    For Sundry Creditors (liabilities): closing balance is Cr to stored positive.

    We return the raw signed value and let the caller interpret it.
    """
    if not text:
        return 0.0
    text = text.strip().replace(",", "")
    text = text.replace("₹", "").replace("Rs.", "").replace("Rs", "")
    # Strip Dr/Cr suffix — it is already encoded in the sign
    for suffix in (" Cr.", " Cr", " Dr.", " Dr"):
        if text.endswith(suffix):
            text = text[: -len(suffix)].strip()
            break
    try:
        return float(text)
    except ValueError:
        return 0.0


# ── Tally Queries ─────────────────────────────────────────────────────

def fetch_group_ledgers_with_balance(
    tally: TallyConnector,
    group_name: str,
    from_date: str,
    to_date: str,
    timeout: int = 60,
) -> list[dict]:
    """
    Fetch all ledgers under *group_name* with their closing balance
    as of *to_date* (opening balance + movements from *from_date* to *to_date*).

    Returns list of dicts:
      { name, parent, closing_balance, opening_balance, debit_total, credit_total }
    """
    xml = f"""<ENVELOPE>
  <HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Collection</TYPE>
    <ID>LedgerBalances_{group_name.replace(' ', '')}</ID>
  </HEADER>
  <BODY>
    <DESC>
      <STATICVARIABLES>
        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
        <SVCURRENTCOMPANY>{tally.company}</SVCURRENTCOMPANY>
        <SVFROMDATE>{from_date}</SVFROMDATE>
        <SVTODATE>{to_date}</SVTODATE>
      </STATICVARIABLES>
      <TDL>
        <TDLMESSAGE>
          <COLLECTION NAME="LedgerBalances_{group_name.replace(' ', '')}" ISMODIFY="No">
            <TYPE>Ledger</TYPE>
            <BELONGSTO>Yes</BELONGSTO>
            <CHILDOF>{group_name}</CHILDOF>
            <FETCH>NAME, PARENT, OPENINGBALANCE, CLOSINGBALANCE,
                   BILLALLOCATIONS.LIST, LEDSTATENAME, LEDGERMOBILE,
                   LEDGEREMAIL, PARTYGSTIN, LEDADDRESS</FETCH>
          </COLLECTION>
        </TDLMESSAGE>
      </TDL>
    </DESC>
  </BODY>
</ENVELOPE>"""

    resp = tally._send_xml(xml, timeout=timeout)
    root = ET.fromstring(sanitize_xml(resp))

    ledgers = []
    for ldg in root.iter("LEDGER"):
        name = (ldg.get("NAME") or ldg.findtext(".//NAME", "")).strip()
        if not name:
            continue
        parent = ldg.findtext("PARENT", "").strip()
        opening_raw = ldg.findtext("OPENINGBALANCE", "0")
        closing_raw = ldg.findtext("CLOSINGBALANCE", "0")
        state = ldg.findtext("LEDSTATENAME", "").strip()
        mobile = ldg.findtext("LEDGERMOBILE", "").strip()
        email = ldg.findtext("LEDGEREMAIL", "").strip()
        gstin = ldg.findtext("PARTYGSTIN", "").strip()
        address = " ".join(
            (a.text or "").strip()
            for a in ldg.findall(".//LEDADDRESS")
            if a.text
        )

        opening = parse_tally_amount(opening_raw)
        closing = parse_tally_amount(closing_raw)

        ledgers.append({
            "name": name,
            "parent": parent,
            "opening_balance": opening,
            "closing_balance": closing,
            "gstin": gstin,
            "state": state,
            "mobile": mobile,
            "email": email,
            "address": address,
        })

    return ledgers


# ── Excel Styles ──────────────────────────────────────────────────────

def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(ws, row, cols, fill_hex: str, font_color: str = "FFFFFF"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_color, size=11)
    border = _thin_border()
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _data_row_style(ws, row, cols, amount_cols: list[int]):
    border = _thin_border()
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if col in amount_cols:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")


def _total_row_style(ws, row, cols, amount_cols: list[int]):
    fill = PatternFill("solid", fgColor="F2F2F2")
    font = Font(bold=True, size=11)
    border = _thin_border()
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        if col in amount_cols:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")


# ── Sheet Writer ──────────────────────────────────────────────────────

def write_sheet(
    wb: openpyxl.Workbook,
    sheet_title: str,
    group_name: str,
    ledgers: list[dict],
    from_date_str: str,
    to_date_str: str,
    fill_hex: str,
    balance_label: str,  # e.g. "Receivable (Dr)" or "Payable (Cr)"
    balance_sign: int,   # +1 to show as positive for assets, -1 for liabilities
):
    ws = wb.create_sheet(title=sheet_title)

    # ── Title ──
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{group_name} — Closing Balances as of {to_date_str[:4]}-{to_date_str[4:6]}-{to_date_str[6:]}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = f"From: {from_date_str[:4]}-{from_date_str[4:6]}-{from_date_str[6:]}  |  Tally Company: (as configured in .env)"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Column Headers (row 4) ──
    headers = [
        "#",
        "Ledger Name",
        "GSTIN",
        "State",
        f"Opening Balance",
        f"{balance_label}",   # closing balance column
        "Mobile",
        "Email",
        "Address",
        "Parent Group",
    ]
    AMOUNT_COLS = [5, 6]
    NUM_COLS = len(headers)

    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    _header_style(ws, 4, NUM_COLS, fill_hex)
    ws.row_dimensions[4].height = 30

    # ── Data rows ──
    # Sort by closing balance descending (largest first)
    sorted_ledgers = sorted(
        ledgers,
        key=lambda x: abs(x["closing_balance"]),
        reverse=True,
    )

    data_start = 5
    total_opening = 0.0
    total_closing = 0.0

    for i, ldg in enumerate(sorted_ledgers, 1):
        row = data_start + i - 1
        opening_display = balance_sign * ldg["opening_balance"]
        closing_display = balance_sign * ldg["closing_balance"]

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=ldg["name"])
        ws.cell(row=row, column=3, value=ldg["gstin"])
        ws.cell(row=row, column=4, value=ldg["state"])
        ws.cell(row=row, column=5, value=opening_display if opening_display != 0 else None)
        ws.cell(row=row, column=6, value=closing_display if closing_display != 0 else None)
        ws.cell(row=row, column=7, value=ldg["mobile"])
        ws.cell(row=row, column=8, value=ldg["email"])
        ws.cell(row=row, column=9, value=ldg["address"])
        ws.cell(row=row, column=10, value=ldg["parent"])

        _data_row_style(ws, row, NUM_COLS, AMOUNT_COLS)
        # Alternate row shading
        if i % 2 == 0:
            light_fill = PatternFill("solid", fgColor="F9F9F9")
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=row, column=col).fill = light_fill

        total_opening += opening_display
        total_closing += closing_display

    # ── Total row ──
    total_row = data_start + len(sorted_ledgers)
    ws.cell(row=total_row, column=1, value="")
    ws.cell(row=total_row, column=2, value=f"TOTAL  ({len(sorted_ledgers)} accounts)")
    ws.cell(row=total_row, column=5, value=total_opening)
    ws.cell(row=total_row, column=6, value=total_closing)
    _total_row_style(ws, total_row, NUM_COLS, AMOUNT_COLS)

    # ── Column widths ──
    col_widths = [5, 52, 18, 18, 18, 18, 16, 28, 40, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze top rows ──
    ws.freeze_panes = "A5"

    # ── Auto-filter on headers ──
    ws.auto_filter.ref = f"A4:{get_column_letter(NUM_COLS)}{total_row - 1}"

    return total_closing


# ── Summary Sheet ─────────────────────────────────────────────────────

def write_summary_sheet(
    wb: openpyxl.Workbook,
    to_date_str: str,
    debtors_total: float,
    creditors_total: float,
    debtor_count: int,
    creditor_count: int,
):
    ws = wb.create_sheet(title="Summary", index=0)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = "Tally — Debtors & Creditors Summary"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Balances as of {to_date_str[:4]}-{to_date_str[4:6]}-{to_date_str[6:]}"
    ws["A2"].font = Font(italic=True, size=11, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    rows = [
        (4, "Particulars", "Amount (₹)", True, "1F4E79", "FFFFFF"),
        (5, f"Sundry Debtors  ({debtor_count} accounts)", debtors_total, False, None, None),
        (6, f"Sundry Creditors  ({creditor_count} accounts)", creditors_total, False, None, None),
        (8, "Net Position (Debtors − Creditors)", debtors_total - creditors_total, False, "FFF2CC", "000000"),
    ]

    border = _thin_border()
    for r in rows:
        row_num, label, value, is_header, fill_hex, font_color = r
        cell_a = ws.cell(row=row_num, column=1, value=label)
        cell_b = ws.cell(row=row_num, column=2, value=value if not is_header else value)

        cell_a.border = border
        cell_b.border = border
        cell_a.alignment = Alignment(vertical="center")
        cell_b.alignment = Alignment(horizontal="right", vertical="center")

        if is_header:
            fill = PatternFill("solid", fgColor=fill_hex)
            font = Font(bold=True, color=font_color, size=11)
            cell_a.fill = fill
            cell_b.fill = fill
            cell_a.font = font
            cell_b.font = font
        else:
            cell_b.number_format = '#,##0.00'
            if fill_hex:
                fill = PatternFill("solid", fgColor=fill_hex)
                cell_a.fill = fill
                cell_b.fill = fill
                cell_a.font = Font(bold=True)
                cell_b.font = Font(bold=True)
            ws.row_dimensions[row_num].height = 22

    ws.row_dimensions[4].height = 28


# ── Main ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export Sundry Debtors & Creditors from Tally to Excel"
    )
    parser.add_argument(
        "--from-date", default="20250401",
        help="Start date YYYYMMDD (default: 20250401 — FY start)"
    )
    parser.add_argument(
        "--as-of", "--to-date",
        default=date.today().strftime("%Y%m%d"),
        dest="to_date",
        help="End/as-of date YYYYMMDD (default: today)"
    )
    parser.add_argument(
        "--output", default=None,
        help="Output Excel filename (default: auto-named in .tmp/)"
    )
    parser.add_argument(
        "--test", action="store_true",
        help="Test Tally connection only"
    )
    parser.add_argument(
        "--timeout", type=int, default=60,
        help="Per-query timeout in seconds (default: 60)"
    )
    args = parser.parse_args()

    TMP_DIR.mkdir(parents=True, exist_ok=True)

    # ── Connect ──────────────────────────────────────────────────────
    tally = TallyConnector()
    print(f"Connecting to Tally at {tally.base_url}...")
    try:
        info = tally.test_connection()
        print(f"[OK] Connected: {info['target_company']}")
    except Exception as e:
        print(f"[FAIL] Connection failed: {e}")
        sys.exit(1)

    if args.test:
        return

    from_date = args.from_date
    to_date   = args.to_date
    print(f"  Period: {from_date} to {to_date}\n")

    # ── Fetch Sundry Debtors ─────────────────────────────────────────
    print("Fetching Sundry Debtors...")
    debtors = fetch_group_ledgers_with_balance(
        tally, "Sundry Debtors", from_date, to_date, timeout=args.timeout
    )
    # Filter out zero-balance (optional — comment out to include all)
    debtors_nonzero = [d for d in debtors if d["closing_balance"] != 0]
    debtors_zero    = [d for d in debtors if d["closing_balance"] == 0]
    print(f"  Found {len(debtors)} ledgers  ({len(debtors_nonzero)} with balance, "
          f"{len(debtors_zero)} zero-balance)")

    # ── Fetch Sundry Creditors ────────────────────────────────────────
    print("Fetching Sundry Creditors...")
    creditors = fetch_group_ledgers_with_balance(
        tally, "Sundry Creditors", from_date, to_date, timeout=args.timeout
    )
    creditors_nonzero = [c for c in creditors if c["closing_balance"] != 0]
    creditors_zero    = [c for c in creditors if c["closing_balance"] == 0]
    print(f"  Found {len(creditors)} ledgers  ({len(creditors_nonzero)} with balance, "
          f"{len(creditors_zero)} zero-balance)")

    # ── Build Excel ───────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    # Remove default sheet — we create named sheets ourselves
    wb.remove(wb.active)

    #  Sundry Debtors:  Tally stores Dr balance as negative to multiply by -1 for display
    debtors_total = write_sheet(
        wb,
        sheet_title="Sundry Debtors",
        group_name="Sundry Debtors",
        ledgers=debtors,          # include ALL (even zero) in sheet
        from_date_str=from_date,
        to_date_str=to_date,
        fill_hex="1F4E79",        # dark blue
        balance_label="Closing Balance (Receivable ₹)",
        balance_sign=-1,          # Dr balance to negate to show as positive receivable
    )

    # Sundry Creditors: Tally stores Cr balance as positive to multiply by +1 for display as payable
    creditors_total = write_sheet(
        wb,
        sheet_title="Sundry Creditors",
        group_name="Sundry Creditors",
        ledgers=creditors,
        from_date_str=from_date,
        to_date_str=to_date,
        fill_hex="833C00",        # dark orange
        balance_label="Closing Balance (Payable ₹)",
        balance_sign=1,           # Cr balance to positive payable
    )

    write_summary_sheet(
        wb,
        to_date_str=to_date,
        debtors_total=debtors_total,
        creditors_total=creditors_total,
        debtor_count=len(debtors),
        creditor_count=len(creditors),
    )

    # Summary sheet should be first
    # (already created with index=0 above)

    # ── Save ──────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.output:
        out_path = Path(args.output)
    else:
        out_path = TMP_DIR / f"Debtors_Creditors_{to_date}_{timestamp}.xlsx"

    wb.save(str(out_path))

    print(f"\n{'='*65}")
    print("EXPORT COMPLETE")
    print(f"{'='*65}")
    print(f"  Sundry Debtors  : {len(debtors):>5} accounts   "
          f"Total Receivable : Rs.{debtors_total:>15,.2f}")
    print(f"  Sundry Creditors: {len(creditors):>5} accounts   "
          f"Total Payable    : Rs.{creditors_total:>15,.2f}")
    print(f"  Net (Dr - Cr)   :                   "
          f"                  Rs.{debtors_total - creditors_total:>15,.2f}")
    print(f"\n  Saved: {out_path}")


if __name__ == "__main__":
    main()
