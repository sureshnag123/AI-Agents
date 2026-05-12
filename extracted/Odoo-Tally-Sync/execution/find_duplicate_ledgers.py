"""
Parse Tally Master.xml and find duplicate/similar ledger names
for Sundry Debtors and Sundry Creditors.
"""

import re
import unicodedata
from difflib import SequenceMatcher
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MASTER_XML = r"D:\TallyGST62SU\TallyExportFiles\Master.xml"
OUTPUT_EXCEL = r"D:\TallyGST62SU\TallyExportFiles\Duplicate_Ledgers.xlsx"

TARGET_GROUPS = {"Sundry Debtors", "Sundry Creditors"}
SIMILARITY_THRESHOLD = 0.82  # 82%+ match = flagged as similar


# ── Parse ────────────────────────────────────────────────────────────────────

def parse_master_xml(path):
    with open(path, encoding="utf-16", errors="replace") as f:
        content = f.read()

    ledgers = []
    # Split into records by CALEDGERSLNO
    blocks = re.split(r"<\s*CALEDGERSLNO\s*>", content)[1:]
    for block in blocks:
        name_m = re.search(r"<\s*CAACCTYPENAME\s*>(.*?)<\s*/\s*CAACCTYPENAME\s*>", block, re.DOTALL)
        parent_m = re.search(r"<\s*CALEDGERPARENT\s*>(.*?)<\s*/\s*CALEDGERPARENT\s*>", block, re.DOTALL)
        slno_m = re.search(r"^(.*?)<", block)

        if name_m and parent_m:
            name = name_m.group(1).strip()
            parent = parent_m.group(1).strip()
            if parent in TARGET_GROUPS:
                ledgers.append({"name": name, "group": parent})

    return ledgers


# ── Normalise for comparison ─────────────────────────────────────────────────

def normalise(name):
    name = name.upper()
    name = unicodedata.normalize("NFKD", name)
    # remove common suffixes/noise
    noise = [
        r"\bPVT\b", r"\bPRIVATE\b", r"\bLIMITED\b", r"\bLTD\b",
        r"\bPLC\b", r"\bLLP\b", r"\bINC\b", r"\bCORPORATION\b",
        r"\bCORP\b", r"\bENTERPRISES\b", r"\bENTERPRISE\b",
        r"\bTRADERS\b", r"\bTRADER\b", r"\bINDUSTRIES\b",
        r"\bINDUSTRY\b", r"\bAND\b", r"&", r"\.", r",", r"-", r"\(",
        r"\)", r"M/S", r"\bCO\b",
    ]
    for n in noise:
        name = re.sub(n, " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def similarity(a, b):
    return SequenceMatcher(None, normalise(a), normalise(b)).ratio()


# ── Find duplicates ──────────────────────────────────────────────────────────

def find_duplicates(ledgers):
    by_group = defaultdict(list)
    for l in ledgers:
        by_group[l["group"]].append(l["name"])

    results = {}
    for group, names in by_group.items():
        seen = []       # list of (canonical_name, [all_similar_names])
        assigned = set()

        for i, name in enumerate(names):
            if name in assigned:
                continue
            cluster = [name]
            for j, other in enumerate(names):
                if i == j or other in assigned:
                    continue
                if similarity(name, other) >= SIMILARITY_THRESHOLD:
                    cluster.append(other)
                    assigned.add(other)
            assigned.add(name)
            if len(cluster) > 1:
                seen.append(cluster)

        results[group] = seen

    return results


# ── Excel output ─────────────────────────────────────────────────────────────

COLORS = [
    "FFF2CC", "DAEEF3", "E2EFDA", "FCE4D6", "EAD1DC",
    "D9D2E9", "CFE2F3", "D9EAD3", "FFF2CC", "FCE4D6",
]

HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
GRP_FILL  = PatternFill("solid", fgColor="2E75B6")
GRP_FONT  = Font(bold=True, color="FFFFFF", size=11)
THIN      = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_excel(duplicates, all_ledgers, path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Duplicates ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Duplicates"

    headers = ["#", "Group", "Duplicate Set", "Ledger Name", "Action Suggested"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    row = 2
    set_no = 0
    for group in ["Sundry Creditors", "Sundry Debtors"]:
        clusters = duplicates.get(group, [])
        if not clusters:
            continue

        # Group header row
        ws.merge_cells(f"A{row}:E{row}")
        gc = ws.cell(row, 1, f"  {group}  ({len(clusters)} duplicate sets found)")
        gc.fill = GRP_FILL
        gc.font = GRP_FONT
        gc.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        for cluster in clusters:
            set_no += 1
            color = COLORS[(set_no - 1) % len(COLORS)]
            fill = PatternFill("solid", fgColor=color)

            for idx, name in enumerate(cluster):
                action = "KEEP (primary)" if idx == 0 else "REVIEW / DELETE"
                font = Font(bold=(idx == 0))
                for col, val in enumerate([set_no if idx == 0 else "", group if idx == 0 else "", set_no if idx == 0 else "", name, action], 1):
                    c = ws.cell(row, col, val)
                    c.fill = fill
                    c.font = font
                    c.border = BORDER
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                row += 1

            # blank separator
            row += 1

    # column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 22
    ws.freeze_panes = "A2"

    # ── Sheet 2: All Creditors ───────────────────────────────────────────────
    for group in ["Sundry Creditors", "Sundry Debtors"]:
        ws2 = wb.create_sheet(group.replace("Sundry ", "All "))
        ws2.append(["#", "Ledger Name", "Group"])
        for col in range(1, 4):
            c = ws2.cell(1, col)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER

        dup_names = set()
        for cluster in duplicates.get(group, []):
            for n in cluster:
                dup_names.add(n)

        names = sorted([l["name"] for l in all_ledgers if l["group"] == group])
        for i, name in enumerate(names, 1):
            is_dup = name in dup_names
            fill = PatternFill("solid", fgColor="FFE0E0") if is_dup else PatternFill("solid", fgColor="FFFFFF")
            font = Font(bold=is_dup, color="C00000" if is_dup else "000000")
            for col, val in enumerate([i, name, group], 1):
                c = ws2.cell(i + 1, col)
                c.value = val
                c.fill = fill
                c.font = font
                c.border = BORDER
                c.alignment = Alignment(vertical="center")

        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 60
        ws2.column_dimensions["C"].width = 22
        ws2.freeze_panes = "A2"

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    total_cred = len([l for l in all_ledgers if l["group"] == "Sundry Creditors"])
    total_debt = len([l for l in all_ledgers if l["group"] == "Sundry Debtors"])
    dup_cred   = sum(len(c) for c in duplicates.get("Sundry Creditors", []))
    dup_debt   = sum(len(c) for c in duplicates.get("Sundry Debtors", []))
    set_cred   = len(duplicates.get("Sundry Creditors", []))
    set_debt   = len(duplicates.get("Sundry Debtors", []))

    rows = [
        ("Metric", "Value"),
        ("── Sundry Creditors ──", ""),
        ("Total Creditor Ledgers", total_cred),
        ("Duplicate Sets Found", set_cred),
        ("Ledgers in Duplicate Sets", dup_cred),
        ("Suggested Deletions (extras)", dup_cred - set_cred),
        ("", ""),
        ("── Sundry Debtors ──", ""),
        ("Total Debtor Ledgers", total_debt),
        ("Duplicate Sets Found", set_debt),
        ("Ledgers in Duplicate Sets", dup_debt),
        ("Suggested Deletions (extras)", dup_debt - set_debt),
        ("", ""),
        ("TOTAL SUGGESTED DELETIONS", (dup_cred - set_cred) + (dup_debt - set_debt)),
    ]

    for r, (label, val) in enumerate(rows, 1):
        cl = ws3.cell(r, 1, label)
        cv = ws3.cell(r, 2, val)
        if r == 1:
            cl.fill = HDR_FILL; cl.font = HDR_FONT
            cv.fill = HDR_FILL; cv.font = HDR_FONT
        elif "──" in str(label):
            cl.fill = GRP_FILL; cl.font = GRP_FONT
            cv.fill = GRP_FILL
        elif label == "TOTAL SUGGESTED DELETIONS":
            cl.font = Font(bold=True, color="C00000")
            cv.font = Font(bold=True, color="C00000")
        for c in [cl, cv]:
            c.border = BORDER
            c.alignment = Alignment(vertical="center")

    wb.save(path)
    print(f"\nExcel saved to: {path}")
    return total_cred, total_debt, set_cred, set_debt, dup_cred, dup_debt


# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Reading Master.xml ...")
    ledgers = parse_master_xml(MASTER_XML)
    print(f"  Found {len(ledgers)} Debtor/Creditor ledgers")

    cred = [l for l in ledgers if l["group"] == "Sundry Creditors"]
    debt = [l for l in ledgers if l["group"] == "Sundry Debtors"]
    print(f"  Creditors: {len(cred)}  |  Debtors: {len(debt)}")

    print("Finding duplicates ...")
    duplicates = find_duplicates(ledgers)

    tc, td, sc, sd, dc, dd = write_excel(duplicates, ledgers, OUTPUT_EXCEL)

    print(f"\n{'='*55}")
    print(f"  Sundry Creditors : {tc} total | {sc} dup sets | {dc} in sets | {dc-sc} to delete")
    print(f"  Sundry Debtors   : {td} total | {sd} dup sets | {dd} in sets | {dd-sd} to delete")
    print(f"  Total suggested deletions : {(dc-sc)+(dd-sd)}")
    print(f"{'='*55}")
    print(f"\nOpen: {OUTPUT_EXCEL}")
