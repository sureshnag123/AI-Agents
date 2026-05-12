#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FY 2025-26 Sales & Purchase Reconciliation — Odoo vs Tally
===========================================================
Audit-grade month-on-month reconciliation for April 2025 – March 2026.

Rules:
  • Tally voucher types used:
      Sales    → "GST INVOICE"  (NOT "Sales", NOT "Credit Note")
      Purchase → "Purchase"     (NOT "Debit Note", NOT "Journal")
  • ALL other Tally voucher types are IGNORED (Contra, Payment, Receipt,
    Journal, Credit Note, Debit Note, etc.)
  • Only POSTED Odoo records (out_invoice / in_invoice) are included.
  • Credit Notes (out_refund) and Debit Notes (in_refund) are EXCLUDED.

Matching strategy:
  PURCHASE (3 passes):
    1. Odoo bill number  ==  Tally voucher number  (exact, e.g. BILL/2025/04/0001)
    2. Odoo vendor-ref   in  Tally Reference field
    3. Same date + same amount  (±₹1 tolerance)

  SALES (3 passes):
    1. Normalised party/customer name  +  amount  (±₹1)
       [Tally GST INVOICE has Tally-own numbers, no Odoo ref in narration]
    2. Date + amount  (±₹1)
    3. Amount only within same month  (if unique)

Output:
  Two separate Excel files — one for Sales, one for Purchase.
  Both saved to .tmp/

Usage:
    python execution/fy2526_reconciliation.py
    python execution/fy2526_reconciliation.py --sales-only
    python execution/fy2526_reconciliation.py --purchase-only
    python execution/fy2526_reconciliation.py --odoo-only
    python execution/fy2526_reconciliation.py --output-dir D:/Audit/
"""

import os
import io
import re
import sys
import argparse
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict
from calendar import monthrange

# ── UTF-8 console output on Windows ──────────────────────────────────
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if sys.stderr.encoding and sys.stderr.encoding.lower() != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv

SCRIPT_DIR   = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR      = PROJECT_ROOT / ".tmp"

sys.path.insert(0, str(SCRIPT_DIR))
load_dotenv(PROJECT_ROOT / ".env")

from odoo_connector  import OdooConnector
from tally_connector import TallyConnector


# ── FY 2025-26 months ─────────────────────────────────────────────────

FY_MONTHS = [
    (2025,  4), (2025,  5), (2025,  6),
    (2025,  7), (2025,  8), (2025,  9),
    (2025, 10), (2025, 11), (2025, 12),
    (2026,  1), (2026,  2), (2026,  3),
]
MONTH_NAMES = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
               7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

def month_label(y, m): return f"{MONTH_NAMES[m]}-{str(y)[2:]}"
def month_range(y, m):
    last = monthrange(y, m)[1]
    return f"{y}-{m:02d}-01", f"{y}-{m:02d}-{last:02d}"


# ── Tax helpers ───────────────────────────────────────────────────────

_CGST_RE = re.compile(r"\bCGST\b", re.I)
_SGST_RE = re.compile(r"\bSGST\b", re.I)
_IGST_RE = re.compile(r"\bIGST\b", re.I)

def classify_tax(name):
    if _CGST_RE.search(name): return "CGST"
    if _SGST_RE.search(name): return "SGST"
    if _IGST_RE.search(name): return "IGST"
    return "OTHER"


# ── Party name normalisation (for fuzzy sales matching) ───────────────

_STRIP_WORDS = re.compile(
    r"\b(private|pvt|limited|ltd|llp|llc|inc|corp|co|india|"
    r"industries|enterprises|solutions|services|technologies|"
    r"manufacturing|works|group|and|&)\b\.?",
    re.I
)
_STRIP_PUNCT = re.compile(r"[^a-z0-9\s]")
_MULTI_SPACE = re.compile(r"\s{2,}")

def _norm_name(name: str) -> str:
    """Normalise a company name for fuzzy matching."""
    n = (name or "").lower()
    n = _STRIP_WORDS.sub(" ", n)
    n = _STRIP_PUNCT.sub(" ", n)
    n = _MULTI_SPACE.sub(" ", n)
    return n.strip()

def _names_match(a: str, b: str) -> bool:
    """True if normalised names are equal, or one contains the other (min 6 chars)."""
    na, nb = _norm_name(a), _norm_name(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    # Containment check (longer must be at least 60% of shorter to avoid false matches)
    shorter, longer = (na, nb) if len(na) <= len(nb) else (nb, na)
    if len(shorter) >= 6 and shorter in longer:
        # Guard: shorter must be at least 55% the length of longer
        return len(shorter) / len(longer) >= 0.55
    return False


# ── Odoo fetch ────────────────────────────────────────────────────────

def _build_partner_map(odoo, invoices):
    ids = list({inv["partner_id"][0] for inv in invoices if inv.get("partner_id")})
    if not ids:
        return {}
    partners = odoo.search_read("res.partner", [("id", "in", ids)],
                                ["id", "name", "vat", "state_id"])
    return {p["id"]: p for p in partners}


def _extract_taxes(odoo, inv_id, tax_cache):
    lines = odoo.get_invoice_lines(inv_id)
    amounts = defaultdict(float)
    for line in lines:
        tlr = line.get("tax_line_id")
        if tlr:
            tid = tlr[0]; tname = tlr[1] if len(tlr) > 1 else ""
            if tid not in tax_cache:
                d = odoo.get_tax_details([tid])
                tax_cache[tid] = d[0] if d else {"name": tname, "amount": 0}
            amounts[classify_tax(tax_cache[tid].get("name", tname))] += \
                abs(line.get("balance", 0))
    if not any(amounts.values()):
        for line in lines:
            if line.get("tax_ids") and not line.get("tax_line_id"):
                sub = abs(line.get("price_subtotal", 0))
                for tid in line["tax_ids"]:
                    if tid not in tax_cache:
                        d = odoo.get_tax_details([tid])
                        tax_cache[tid] = d[0] if d else {"name": "", "amount": 0}
                    td = tax_cache[tid]
                    amounts[classify_tax(td.get("name", ""))] += \
                        round(sub * td.get("amount", 0) / 100, 2)
    return dict(amounts)


def fetch_odoo_invoices(odoo, move_type, from_date, to_date, tax_cache):
    """Fetch posted Odoo invoices/bills (no refunds)."""
    raw = (odoo.get_sales_invoices if move_type == "out_invoice"
           else odoo.get_purchase_invoices)(from_date=from_date, to_date=to_date)
    if not raw:
        return []
    pmap = _build_partner_map(odoo, raw)
    rows = []
    for inv in raw:
        pid   = inv["partner_id"][0] if inv.get("partner_id") else None
        pname = inv["partner_id"][1] if inv.get("partner_id") else ""
        gstin = (pmap.get(pid) or {}).get("vat") or ""
        taxes = _extract_taxes(odoo, inv["id"], tax_cache)
        rows.append({
            "invoice_number": inv["name"],
            "vendor_ref"    : (inv.get("ref") or "").strip(),
            "date"          : inv.get("invoice_date", ""),
            "partner"       : pname,
            "gstin"         : gstin,
            "taxable_value" : round(inv.get("amount_untaxed", 0), 2),
            "cgst"          : round(taxes.get("CGST", 0), 2),
            "sgst"          : round(taxes.get("SGST", 0), 2),
            "igst"          : round(taxes.get("IGST", 0), 2),
            "total_tax"     : round(sum(taxes.values()), 2),
            "invoice_total" : round(inv.get("amount_total", 0), 2),
            "payment_state" : inv.get("payment_state", ""),
        })
    return rows


# ── Tally fetch (correct voucher type filtering) ───────────────────────

def _sanitize(xml_text: str) -> str:
    xml_text = re.sub(r'&#x([0-8BbCcEeFf]);', '', xml_text)
    xml_text = re.sub(r'&#x1[0-9A-Fa-f];', '', xml_text)
    xml_text = re.sub(r'&#([0-8]|1[0-1]|1[4-9]|2[0-9]|3[01]);', '', xml_text)
    xml_text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', xml_text)
    xml_text = re.sub(r'<(/?)(\w+):', r'<\1\2_', xml_text)
    xml_text = re.sub(r'\s+xmlns:\w+="[^"]*"', '', xml_text)
    return xml_text


def _parse_amount(val: str) -> float:
    if not val: return 0.0
    try: return abs(float(val.replace(",", "").strip()))
    except ValueError: return 0.0


_CGST_LDG = re.compile(r"\bCGST\b", re.I)
_SGST_LDG = re.compile(r"\bSGST\b", re.I)
_IGST_LDG = re.compile(r"\bIGST\b", re.I)


def _classify_ledger(name: str) -> str:
    """Classify a Tally ledger entry as CGST / SGST / IGST / TAXABLE / PARTY."""
    if _CGST_LDG.search(name): return "CGST"
    if _SGST_LDG.search(name): return "SGST"
    if _IGST_LDG.search(name): return "IGST"
    return "TAXABLE"


def _parse_ledger_entries(voucher_elem, party_name: str) -> dict:
    """
    Walk ALLLEDGERENTRIES.LIST entries and split into:
      taxable_value, cgst, sgst, igst, total_tax

    Approach:
      - Skip the AR/AP (party) entry — identified by matching party_name
        OR by ISDEEMEDPOSITIVE = Yes (the receivable / payable side)
      - Classify remaining entries by ledger name
      - Use abs() on all amounts (Tally sign convention varies by voucher type)
    """
    taxable = cgst = sgst = igst = 0.0

    for entry in voucher_elem.findall("ALLLEDGERENTRIES.LIST"):
        ldg_name = (entry.findtext("LEDGERNAME", "") or "").strip()
        ldg_amt  = _parse_amount(entry.findtext("AMOUNT", "0") or "0")

        if not ldg_name or ldg_amt == 0:
            continue

        # Skip the party / AR / AP ledger entry
        if ldg_name == party_name:
            continue
        # Also skip if ISDEEMEDPOSITIVE = Yes  (that's the party-side debit/credit)
        deemed = (entry.findtext("ISDEEMEDPOSITIVE", "") or "").strip().lower()
        if deemed == "yes":
            continue

        cat = _classify_ledger(ldg_name)
        if   cat == "CGST":    cgst    += ldg_amt
        elif cat == "SGST":    sgst    += ldg_amt
        elif cat == "IGST":    igst    += ldg_amt
        else:                  taxable += ldg_amt   # sales/purchase/income/expense ledger

    total_tax = round(cgst + sgst + igst, 2)
    return {
        "taxable_value": round(taxable, 2),
        "cgst"         : round(cgst, 2),
        "sgst"         : round(sgst, 2),
        "igst"         : round(igst, 2),
        "total_tax"    : total_tax,
    }


def fetch_tally_vouchers(tally: TallyConnector,
                         voucher_type_name: str,   # "GST INVOICE" | "Purchase"
                         from_date: str,            # YYYY-MM-DD
                         to_date: str) -> list:
    """
    Fetch Tally vouchers of an EXACT voucher type, including full GST breakdown.

    Strategy:
      1. Query Tally with ALLLEDGERENTRIES.LIST to get per-voucher ledger entries.
      2. Filter Python-side by exact VOUCHERTYPENAME.
      3. Apply Python-side date guard (Tally TDL date filter is unreliable).
      4. Parse ledger entries to extract Taxable, CGST, SGST, IGST amounts.
    """
    tally_from = from_date.replace("-", "")
    tally_to   = to_date.replace("-", "")

    xml = f"""<ENVELOPE>
  <HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Collection</TYPE>
    <ID>ReconVouchers</ID>
  </HEADER>
  <BODY>
    <DESC>
      <STATICVARIABLES>
        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
        <SVCURRENTCOMPANY>{tally.company}</SVCURRENTCOMPANY>
        <SVFROMDATE>{tally_from}</SVFROMDATE>
        <SVTODATE>{tally_to}</SVTODATE>
      </STATICVARIABLES>
      <TDL>
        <TDLMESSAGE>
          <COLLECTION NAME="ReconVouchers" ISMODIFY="No">
            <TYPE>Voucher</TYPE>
            <FETCH>DATE,VOUCHERNUMBER,VOUCHERTYPENAME,PARTYLEDGERNAME,
                   BASICBUYERNAME,NARRATION,AMOUNT,REFERENCE,
                   ALLLEDGERENTRIES.LIST</FETCH>
          </COLLECTION>
        </TDLMESSAGE>
      </TDL>
    </DESC>
  </BODY>
</ENVELOPE>"""

    resp = tally._send_xml(xml, timeout=180)
    resp = _sanitize(resp)
    root = ET.fromstring(resp)

    rows = []
    for v in root.iter("VOUCHER"):
        vtype = (v.findtext("VOUCHERTYPENAME", "") or "").strip()
        # ── Filter 1: exact voucher type only ────────────────────────
        if vtype != voucher_type_name:
            continue

        raw_date = (v.findtext("DATE", "") or "").strip()
        compact  = raw_date.replace("-", "")
        # ── Filter 2: Python-side date guard ─────────────────────────
        if len(compact) == 8 and compact.isdigit():
            if not (tally_from <= compact <= tally_to):
                continue
        # Format date
        if len(raw_date) == 8 and raw_date.isdigit():
            fmt_date = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:]}"
        else:
            fmt_date = raw_date

        party  = (v.findtext("PARTYLEDGERNAME", "") or "").strip()
        buyer  = (v.findtext("BASICBUYERNAME",  "") or "").strip()
        narr   = (v.findtext("NARRATION",       "") or "").strip()
        ref    = (v.findtext("REFERENCE",       "") or "").strip()
        vno    = (v.findtext("VOUCHERNUMBER",   "") or "").strip()
        amount = _parse_amount(v.findtext("AMOUNT", "0") or "0")

        # ── GST breakdown from ledger entries ─────────────────────────
        gst = _parse_ledger_entries(v, party)

        rows.append({
            "voucher_number" : vno,
            "date"           : fmt_date,
            "party"          : party,
            "buyer_name"     : buyer,
            "amount"         : amount,
            "taxable_value"  : gst["taxable_value"],
            "cgst"           : gst["cgst"],
            "sgst"           : gst["sgst"],
            "igst"           : gst["igst"],
            "total_tax"      : gst["total_tax"],
            "reference"      : ref,
            "narration"      : narr,
        })

    return rows


# ── Reconciliation engine ─────────────────────────────────────────────

def reconcile_purchase(odoo_rows: list, tally_rows: list) -> dict:
    """
    Purchase matching — 3 passes:
      Pass 1: Odoo bill number  ==  Tally voucher number  (exact)
      Pass 2: Odoo vendor_ref   in  Tally reference field  (or vice-versa)
      Pass 3: Same date + same amount  (±₹1)
    """
    tally_matched = set()
    odoo_matched  = set()
    matched = []; mismatches = []

    # Lookup structures
    tally_by_vno: dict  = {}            # voucher_number → index
    tally_by_ref: dict  = defaultdict(list)  # reference → indices
    tally_by_da:  dict  = defaultdict(list)  # (date, round_amt) → indices

    for ti, tr in enumerate(tally_rows):
        vno = (tr.get("voucher_number") or "").strip().upper()
        if vno:
            tally_by_vno[vno] = ti
        ref = (tr.get("reference") or "").strip().upper()
        if ref:
            tally_by_ref[ref].append(ti)
        key = (tr["date"], round(tr["amount"], 0))
        tally_by_da[key].append(ti)

    def _try(oi, odoo_row, ti, method):
        if ti in tally_matched: return False
        tr   = tally_rows[ti]
        diff = abs(odoo_row["invoice_total"] - tr["amount"])
        tally_matched.add(ti); odoo_matched.add(oi)
        if diff < 1.0:
            matched.append((odoo_row, tr, method))
        else:
            mismatches.append((odoo_row, tr, method,
                f"Odoo Rs.{odoo_row['invoice_total']:,.2f}  vs  "
                f"Tally Rs.{tr['amount']:,.2f}  (diff Rs.{diff:,.2f})"))
        return True

    # Pass 1 — bill number match
    for oi, or_ in enumerate(odoo_rows):
        inv_no = (or_["invoice_number"] or "").strip().upper()
        if inv_no in tally_by_vno:
            _try(oi, or_, tally_by_vno[inv_no], "Bill# match")

    # Pass 2 — vendor ref match
    for oi, or_ in enumerate(odoo_rows):
        if oi in odoo_matched: continue
        vref = (or_.get("vendor_ref") or "").strip().upper()
        if not vref: continue
        # Tally reference contains vendor ref?
        for ti in tally_by_ref.get(vref, []):
            if _try(oi, or_, ti, "Vendor-ref match"): break
        if oi in odoo_matched: continue
        # Odoo vendor_ref inside tally reference (partial)
        for ti, tr in enumerate(tally_rows):
            if ti in tally_matched: continue
            tref = (tr.get("reference") or "").strip().upper()
            if vref and vref in tref:
                if _try(oi, or_, ti, "Vendor-ref partial match"): break

    # Pass 3 — date + amount
    for oi, or_ in enumerate(odoo_rows):
        if oi in odoo_matched: continue
        key = (or_["date"], round(or_["invoice_total"], 0))
        for ti in tally_by_da.get(key, []):
            if _try(oi, or_, ti, "Date+Amount match"): break

    return {
        "matched"   : matched,
        "mismatches": mismatches,
        "only_odoo" : [r for i, r in enumerate(odoo_rows)  if i not in odoo_matched],
        "only_tally": [r for i, r in enumerate(tally_rows) if i not in tally_matched],
    }


def reconcile_sales(odoo_rows: list, tally_rows: list) -> dict:
    """
    Sales matching — 3 passes:
      Pass 1: Normalised party/customer name  +  amount  (±₹1)
              [Tally uses its own INV/2526/FW1 sequence — no Odoo ref embedded]
      Pass 2: Date  +  amount  (±₹1)
      Pass 3: Amount-only within month if unique (rare, safety net)
    """
    tally_matched = set()
    odoo_matched  = set()
    matched = []; mismatches = []

    # Lookup: tally amount → indices
    tally_by_amt: dict = defaultdict(list)
    for ti, tr in enumerate(tally_rows):
        tally_by_amt[round(tr["amount"], 0)].append(ti)

    # Lookup: tally (date, round_amt) → indices
    tally_by_da: dict = defaultdict(list)
    for ti, tr in enumerate(tally_rows):
        tally_by_da[(tr["date"], round(tr["amount"], 0))].append(ti)

    def _try(oi, odoo_row, ti, method):
        if ti in tally_matched: return False
        tr   = tally_rows[ti]
        diff = abs(odoo_row["invoice_total"] - tr["amount"])
        tally_matched.add(ti); odoo_matched.add(oi)
        if diff < 1.0:
            matched.append((odoo_row, tr, method))
        else:
            mismatches.append((odoo_row, tr, method,
                f"Odoo Rs.{odoo_row['invoice_total']:,.2f}  vs  "
                f"Tally Rs.{tr['amount']:,.2f}  (diff Rs.{diff:,.2f})"))
        return True

    # Pass 1 — party name + amount
    for oi, or_ in enumerate(odoo_rows):
        odoo_name = or_["partner"]
        target_amt = round(or_["invoice_total"], 0)
        for ti in tally_by_amt.get(target_amt, []):
            if ti in tally_matched: continue
            tr = tally_rows[ti]
            # Try both PARTYLEDGERNAME and BASICBUYERNAME
            if (_names_match(odoo_name, tr["party"]) or
                    _names_match(odoo_name, tr.get("buyer_name", ""))):
                diff = abs(or_["invoice_total"] - tr["amount"])
                if diff < 1.0:
                    if _try(oi, or_, ti, "Party+Amount match"): break

    # Pass 2 — date + amount
    for oi, or_ in enumerate(odoo_rows):
        if oi in odoo_matched: continue
        key = (or_["date"], round(or_["invoice_total"], 0))
        for ti in tally_by_da.get(key, []):
            if _try(oi, or_, ti, "Date+Amount match"): break

    # Pass 3 — unique amount in full dataset (safety net)
    for oi, or_ in enumerate(odoo_rows):
        if oi in odoo_matched: continue
        target_amt = round(or_["invoice_total"], 0)
        candidates = [ti for ti in tally_by_amt.get(target_amt, [])
                      if ti not in tally_matched]
        if len(candidates) == 1:
            _try(oi, or_, candidates[0], "Unique-Amount match")

    return {
        "matched"   : matched,
        "mismatches": mismatches,
        "only_odoo" : [r for i, r in enumerate(odoo_rows)  if i not in odoo_matched],
        "only_tally": [r for i, r in enumerate(tally_rows) if i not in tally_matched],
    }


def month_stats(odoo, tally, recon):
    return {
        "odoo_count"      : len(odoo),
        "tally_count"     : len(tally),
        "odoo_total"      : round(sum(r["invoice_total"] for r in odoo), 2),
        "tally_total"     : round(sum(r["amount"]        for r in tally), 2),
        "matched"         : len(recon["matched"]),
        "mismatches"      : len(recon["mismatches"]),
        "missing_tally"   : len(recon["only_odoo"]),
        "missing_odoo"    : len(recon["only_tally"]),
    }


# ── Excel styling helpers ─────────────────────────────────────────────

def _styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
    med  = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("medium"))
    return {
        "thin"       : thin,
        "med_bot"    : med,
        "hdr_font"   : Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "hdr_blue"   : PatternFill("solid", fgColor="2F5496"),
        "hdr_green"  : PatternFill("solid", fgColor="375623"),
        "title_font" : Font(name="Calibri", bold=True, size=13, color="1F3864"),
        "sub_font"   : Font(name="Calibri", size=10, color="595959"),
        "bold"       : Font(name="Calibri", bold=True, size=10),
        "bold_w"     : Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "hdr_align"  : Alignment(horizontal="center", vertical="center", wrap_text=True),
        "ctr"        : Alignment(horizontal="center", vertical="center"),
        "lft"        : Alignment(horizontal="left",   vertical="center"),
        "rgt"        : Alignment(horizontal="right",  vertical="center"),
        "green_f"    : PatternFill("solid", fgColor="C6EFCE"),
        "red_f"      : PatternFill("solid", fgColor="FFC7CE"),
        "yellow_f"   : PatternFill("solid", fgColor="FFEB9C"),
        "orange_f"   : PatternFill("solid", fgColor="FCE4D6"),
        "grey_f"     : PatternFill("solid", fgColor="D9D9D9"),
        "cur"        : '#,##0.00',
        "int"        : '#,##0',
    }


def _title(ws, title, sub, span, st):
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    ws["A1"] = title
    ws["A1"].font      = st["title_font"]
    ws["A1"].alignment = st["ctr"]
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=span)
    ws["A2"] = sub
    ws["A2"].font      = st["sub_font"]
    ws["A2"].alignment = st["ctr"]


def _hdr(ws, cols, row, st, fill_key="hdr_blue"):
    import openpyxl.utils
    for ci, (label, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font = st["hdr_font"]; c.fill = st[fill_key]
        c.alignment = st["hdr_align"]; c.border = st["thin"]
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width
    ws.row_dimensions[row].height = 30


def _cell(ws, row, col, val, st, fmt=None, fill=None, bold=False, align="lft"):
    c = ws.cell(row=row, column=col, value=val)
    c.border = st["thin"]; c.alignment = st[align]
    if fmt:  c.number_format = fmt
    if fill: c.fill = st[fill]
    if bold: c.font = st["bold"]
    return c


# ── Sheet builders ────────────────────────────────────────────────────

def _sheet_summary(wb, st, monthly, doc_type, voucher_type, fill_key):
    ws = wb.active if doc_type == "Sales" else wb.create_sheet("Summary")
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    sub = (f"FY 2025-26  |  Tally: '{voucher_type}'  |  "
           f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}  |  "
           f"Credit/Debit Notes excluded")
    _title(ws, f"{doc_type} Reconciliation — Odoo vs Tally  |  FY 2025-26", sub, 11, st)

    cols = [
        ("Month",11),("Odoo Count",11),("Odoo Total (Rs.)",18),
        ("Tally Count",11),("Tally Total (Rs.)",18),("Diff (Rs.)",16),
        ("Matched",10),("Mismatches",12),("Missing in Tally",16),("Missing in Odoo",16),
        ("Status",12),
    ]
    _hdr(ws, cols, 4, st, fill_key)
    ws.freeze_panes = "A5"

    tc = tt = oc = ot = m = mm = mit = mio = 0
    for row_idx, (label, ms) in enumerate(monthly.items(), 5):
        diff   = round(ms["odoo_total"] - ms["tally_total"], 2)
        issues = ms["mismatches"] + ms["missing_tally"] + ms["missing_odoo"]
        status = "OK" if issues == 0 else ("REVIEW" if issues <= 3 else "ACTION NEEDED")
        sf     = "green_f" if status=="OK" else ("yellow_f" if status=="REVIEW" else "red_f")

        _cell(ws, row_idx, 1,  label,               st, bold=True,  align="ctr")
        _cell(ws, row_idx, 2,  ms["odoo_count"],     st, fmt=st["int"],  align="ctr")
        _cell(ws, row_idx, 3,  ms["odoo_total"],     st, fmt=st["cur"],  align="rgt")
        _cell(ws, row_idx, 4,  ms["tally_count"],    st, fmt=st["int"],  align="ctr")
        _cell(ws, row_idx, 5,  ms["tally_total"],    st, fmt=st["cur"],  align="rgt")
        cd = _cell(ws, row_idx, 6, diff,             st, fmt=st["cur"],  align="rgt")
        if abs(diff) > 1: cd.fill = st["yellow_f"]
        _cell(ws, row_idx, 7,  ms["matched"],        st, fmt=st["int"],  align="ctr")
        cmm = _cell(ws, row_idx, 8, ms["mismatches"],st, fmt=st["int"],  align="ctr")
        if ms["mismatches"]: cmm.fill = st["yellow_f"]
        cmt = _cell(ws, row_idx, 9,  ms["missing_tally"], st, fmt=st["int"], align="ctr")
        if ms["missing_tally"]: cmt.fill = st["red_f"]
        cmo = _cell(ws, row_idx, 10, ms["missing_odoo"],  st, fmt=st["int"], align="ctr")
        if ms["missing_odoo"]: cmo.fill = st["orange_f"]
        _cell(ws, row_idx, 11, status, st, fill=sf, align="ctr", bold=True)

        oc += ms["odoo_count"];  ot += ms["odoo_total"]
        tc += ms["tally_count"]; tt += ms["tally_total"]
        m  += ms["matched"];     mm += ms["mismatches"]
        mit += ms["missing_tally"]; mio += ms["missing_odoo"]

    # Totals row
    tr = 5 + len(monthly)
    diff_total = round(ot - tt, 2)
    for ci, val in enumerate(["TOTAL", oc, ot, tc, tt, diff_total, m, mm, mit, mio, ""], 1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.border = st["med_bot"]; c.font = st["bold"]; c.fill = st["grey_f"]
        if ci in (3,5,6) and isinstance(val, float): c.number_format = st["cur"]
        if ci in (2,4,7,8,9,10) and isinstance(val, int): c.number_format = st["int"]

    # Legend
    from openpyxl.styles import PatternFill as PF
    lr = tr + 2
    ws.cell(row=lr,   column=1, value="LEGEND").font = st["bold"]
    for i, (lbl, col, desc) in enumerate([
        ("Green",  "C6EFCE", "No issues — fully matched"),
        ("Red",    "FFC7CE", "Missing invoice/voucher needs action"),
        ("Yellow", "FFEB9C", "Amount mismatch — review and correct"),
        ("Orange", "FCE4D6", "Extra Tally voucher with no Odoo match — investigate"),
    ], lr+1):
        ws.cell(row=i, column=1, value=lbl).fill = PF("solid", fgColor=col)
        ws.cell(row=i, column=2, value=desc)


def _sheet_odoo_register(wb, sheet_name, rows, st, fill_key, is_purchase):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    doc = "Purchase Bills (Vendor Invoices)" if is_purchase else "Sales Invoices (Customer)"
    _title(ws, f"Odoo — {doc}  |  FY 2025-26",
           "Posted only. Credit/Debit Notes excluded.", 13, st)
    cols = [
        ("S.No",6),("Month",10),("Invoice No.",22),("Vendor Ref",22),
        ("Date",12),("Partner",32),("GSTIN",20),
        ("Taxable (Rs.)",16),("CGST (Rs.)",14),("SGST (Rs.)",14),
        ("IGST (Rs.)",14),("Total Tax (Rs.)",16),("Invoice Total (Rs.)",18),
    ]
    _hdr(ws, cols, 4, st, fill_key)
    ws.freeze_panes = "A5"
    for i, r in enumerate(rows, 1):
        row = 4 + i
        _cell(ws, row, 1,  i,                     st, fmt=st["int"],  align="ctr")
        _cell(ws, row, 2,  r.get("_month",""),    st, align="ctr")
        _cell(ws, row, 3,  r["invoice_number"],   st)
        _cell(ws, row, 4,  r.get("vendor_ref",""),st)
        _cell(ws, row, 5,  r["date"],             st, align="ctr")
        _cell(ws, row, 6,  r["partner"],          st)
        _cell(ws, row, 7,  r["gstin"],            st, align="ctr")
        _cell(ws, row, 8,  r["taxable_value"],    st, fmt=st["cur"],  align="rgt")
        _cell(ws, row, 9,  r["cgst"],             st, fmt=st["cur"],  align="rgt")
        _cell(ws, row, 10, r["sgst"],             st, fmt=st["cur"],  align="rgt")
        _cell(ws, row, 11, r["igst"],             st, fmt=st["cur"],  align="rgt")
        _cell(ws, row, 12, r["total_tax"],        st, fmt=st["cur"],  align="rgt")
        _cell(ws, row, 13, r["invoice_total"],    st, fmt=st["cur"],  align="rgt")
    if rows:
        tr = 5 + len(rows)
        ws.cell(row=tr, column=6, value="GRAND TOTAL").font = st["bold"]
        for ci, key in [(8,"taxable_value"),(9,"cgst"),(10,"sgst"),
                        (11,"igst"),(12,"total_tax"),(13,"invoice_total")]:
            c = ws.cell(row=tr, column=ci, value=round(sum(r[key] for r in rows), 2))
            c.number_format = st["cur"]; c.font = st["bold"]
            c.border = st["thin"]; c.fill = st["grey_f"]


def _sheet_tally_register(wb, sheet_name, rows, st, fill_key, voucher_type):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    _title(ws, f"Tally — '{voucher_type}' Vouchers  |  FY 2025-26",
           "Only this voucher type included. GST breakdown from ledger entries.", 14, st)
    cols = [
        ("S.No",6),("Month",10),("Voucher No.",22),
        ("Date",12),("Party / Ledger",32),("Buyer Name",28),
        ("Taxable (Rs.)",16),("CGST (Rs.)",14),("SGST (Rs.)",14),
        ("IGST (Rs.)",14),("Total Tax (Rs.)",16),("Invoice Total (Rs.)",18),
        ("Reference",24),("Narration",35),
    ]
    _hdr(ws, cols, 4, st, fill_key)
    ws.freeze_panes = "A5"
    for i, r in enumerate(rows, 1):
        row = 4 + i
        _cell(ws, row,  1, i,                       st, fmt=st["int"], align="ctr")
        _cell(ws, row,  2, r.get("_month",""),      st, align="ctr")
        _cell(ws, row,  3, r["voucher_number"],     st)
        _cell(ws, row,  4, r["date"],               st, align="ctr")
        _cell(ws, row,  5, r["party"],              st)
        _cell(ws, row,  6, r.get("buyer_name",""),  st)
        _cell(ws, row,  7, r.get("taxable_value",0),st, fmt=st["cur"], align="rgt")
        _cell(ws, row,  8, r.get("cgst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row,  9, r.get("sgst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 10, r.get("igst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 11, r.get("total_tax",0),    st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 12, r["amount"],             st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 13, r["reference"],          st)
        _cell(ws, row, 14, r["narration"],          st)
    if rows:
        tr = 5 + len(rows)
        ws.cell(row=tr, column=5, value="GRAND TOTAL").font = st["bold"]
        for ci, key in [(7,"taxable_value"),(8,"cgst"),(9,"sgst"),
                        (10,"igst"),(11,"total_tax"),(12,"amount")]:
            c = ws.cell(row=tr, column=ci,
                        value=round(sum(r.get(key,0) for r in rows), 2))
            c.number_format = st["cur"]; c.font = st["bold"]
            c.border = st["thin"]; c.fill = st["grey_f"]


def _sheet_missing_in_tally(wb, sheet_name, rows, st, fill_key, is_purchase):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    doc = "Bills" if is_purchase else "Invoices"
    _title(ws,
           f"Action Required — {doc} in Odoo NOT FOUND in Tally  |  FY 2025-26",
           "These must be posted in Tally. Verify each before posting.", 13, st)
    cols = [
        ("S.No",6),("Month",10),("Invoice No.",22),("Vendor Ref",22),
        ("Date",12),("Partner",32),("GSTIN",20),
        ("Taxable (Rs.)",16),("CGST (Rs.)",14),("SGST (Rs.)",14),
        ("IGST (Rs.)",14),("Total Tax (Rs.)",16),("Invoice Total (Rs.)",18),
    ]
    _hdr(ws, cols, 4, st, fill_key)
    ws.freeze_panes = "A5"
    for i, r in enumerate(rows, 1):
        row = 4 + i
        _cell(ws, row, 1,  i,                   st, fmt=st["int"], align="ctr")
        _cell(ws, row, 2,  r.get("_month",""),  st, align="ctr",  fill="red_f", bold=True)
        _cell(ws, row, 3,  r["invoice_number"], st, fill="red_f")
        _cell(ws, row, 4,  r.get("vendor_ref",""), st)
        _cell(ws, row, 5,  r["date"],           st, align="ctr")
        _cell(ws, row, 6,  r["partner"],        st)
        _cell(ws, row, 7,  r["gstin"],          st, align="ctr")
        _cell(ws, row, 8,  r["taxable_value"],  st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 9,  r["cgst"],           st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 10, r["sgst"],           st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 11, r["igst"],           st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 12, r["total_tax"],      st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 13, r["invoice_total"],  st, fmt=st["cur"], align="rgt")
    if not rows:
        ws.cell(row=5, column=2,
                value=f"Nothing missing — all Odoo {doc.lower()} found in Tally").font = st["bold"]
    else:
        tr = 5 + len(rows)
        ws.cell(row=tr, column=6,
                value=f"TOTAL MISSING  ({len(rows)} {doc})").font = st["bold"]
        c = ws.cell(row=tr, column=13,
                    value=round(sum(r["invoice_total"] for r in rows), 2))
        c.number_format = st["cur"]; c.font = st["bold"]
        c.border = st["thin"]; c.fill = st["red_f"]


def _sheet_missing_in_odoo(wb, sheet_name, rows, st, fill_key, is_purchase, voucher_type):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    doc = "Purchase Vouchers" if is_purchase else "Sales Vouchers (GST INVOICE)"
    _title(ws,
           f"Investigate — '{voucher_type}' in Tally NOT FOUND in Odoo  |  FY 2025-26",
           "Entries posted directly in Tally without an Odoo record. Verify each.", 14, st)
    cols = [
        ("S.No",6),("Month",10),("Voucher No.",22),
        ("Date",12),("Party / Ledger",32),("Buyer Name",28),
        ("Taxable (Rs.)",16),("CGST (Rs.)",14),("SGST (Rs.)",14),
        ("IGST (Rs.)",14),("Total Tax (Rs.)",16),("Invoice Total (Rs.)",18),
        ("Reference",24),("Narration",40),
    ]
    _hdr(ws, cols, 4, st, fill_key)
    ws.freeze_panes = "A5"
    for i, r in enumerate(rows, 1):
        row = 4 + i
        _cell(ws, row,  1, i,                       st, fmt=st["int"], align="ctr")
        _cell(ws, row,  2, r.get("_month",""),      st, align="ctr",  fill="orange_f", bold=True)
        _cell(ws, row,  3, r["voucher_number"],     st, fill="orange_f")
        _cell(ws, row,  4, r["date"],               st, align="ctr")
        _cell(ws, row,  5, r["party"],              st)
        _cell(ws, row,  6, r.get("buyer_name",""),  st)
        _cell(ws, row,  7, r.get("taxable_value",0),st, fmt=st["cur"], align="rgt")
        _cell(ws, row,  8, r.get("cgst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row,  9, r.get("sgst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 10, r.get("igst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 11, r.get("total_tax",0),    st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 12, r["amount"],             st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 13, r["reference"],          st)
        _cell(ws, row, 14, r["narration"],          st)
    if not rows:
        ws.cell(row=5, column=2,
                value=f"No extra {doc} found in Tally — Tally matches Odoo").font = st["bold"]
    else:
        tr = 5 + len(rows)
        ws.cell(row=tr, column=5,
                value=f"TOTAL  ({len(rows)} vouchers)").font = st["bold"]
        for ci, key in [(7,"taxable_value"),(8,"cgst"),(9,"sgst"),
                        (10,"igst"),(11,"total_tax"),(12,"amount")]:
            c = ws.cell(row=tr, column=ci,
                        value=round(sum(r.get(key,0) for r in rows), 2))
            c.number_format = st["cur"]; c.font = st["bold"]
            c.border = st["thin"]; c.fill = st["orange_f"]


def _sheet_mismatches(wb, sheet_name, rows, st, fill_key, is_purchase):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    doc = "Bills" if is_purchase else "Invoices"
    _title(ws,
           f"Amount Mismatches — {doc}  |  FY 2025-26",
           "Matched by reference/party/date but amounts differ. Correct before finalising.", 19, st)
    cols = [
        ("S.No",6),("Month",10),("Match Method",20),
        # Odoo side
        ("Odoo Invoice No.",22),("Odoo Date",12),("Odoo Partner",28),
        ("Odoo Taxable",16),("Odoo CGST",13),("Odoo SGST",13),
        ("Odoo IGST",13),("Odoo Total Tax",14),("Odoo Total (Rs.)",18),
        # Tally side
        ("Tally Voucher No.",22),("Tally Date",12),("Tally Party",28),
        ("Tally Taxable",16),("Tally CGST",13),("Tally SGST",13),
        ("Tally IGST",13),
        # Diff
        ("Diff (Rs.)",16),
    ]
    _hdr(ws, cols, 4, st, fill_key)
    ws.freeze_panes = "A5"
    for i, row_tuple in enumerate(rows, 1):
        or_, tr, method = row_tuple[0], row_tuple[1], row_tuple[2]
        row  = 4 + i
        diff = round(or_["invoice_total"] - tr["amount"], 2)
        _cell(ws, row,  1, i,                         st, fmt=st["int"], align="ctr")
        _cell(ws, row,  2, or_.get("_month",""),      st, align="ctr",  fill="yellow_f")
        _cell(ws, row,  3, method,                    st, align="ctr")
        # Odoo
        _cell(ws, row,  4, or_["invoice_number"],     st, fill="yellow_f")
        _cell(ws, row,  5, or_["date"],               st, align="ctr")
        _cell(ws, row,  6, or_["partner"],            st)
        _cell(ws, row,  7, or_.get("taxable_value",0),st, fmt=st["cur"], align="rgt")
        _cell(ws, row,  8, or_.get("cgst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row,  9, or_.get("sgst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 10, or_.get("igst",0),         st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 11, or_.get("total_tax",0),    st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 12, or_["invoice_total"],      st, fmt=st["cur"], align="rgt")
        # Tally
        _cell(ws, row, 13, tr["voucher_number"],      st)
        _cell(ws, row, 14, tr["date"],                st, align="ctr")
        _cell(ws, row, 15, tr["party"],               st)
        _cell(ws, row, 16, tr.get("taxable_value",0), st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 17, tr.get("cgst",0),          st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 18, tr.get("sgst",0),          st, fmt=st["cur"], align="rgt")
        _cell(ws, row, 19, tr.get("igst",0),          st, fmt=st["cur"], align="rgt")
        # Diff
        cd = _cell(ws, row, 20, diff,                 st, fmt=st["cur"], align="rgt")
        cd.fill = st["yellow_f"]
    if not rows:
        ws.cell(row=5, column=2,
                value="No amount mismatches found").font = st["bold"]


def _sheet_matched(wb, sheet_name, rows, st, fill_key, is_purchase):
    """Full matched-pairs sheet for audit cross-reference."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    doc = "Bills" if is_purchase else "Invoices"
    _title(ws, f"Matched Pairs — {doc}  |  FY 2025-26",
           "Odoo and Tally records confirmed as the same transaction.", 14, st)
    cols = [
        ("S.No",6),("Month",10),("Match Method",20),
        ("Odoo Invoice No.",22),("Odoo Date",12),("Odoo Partner",30),
        ("Odoo Amount (Rs.)",18),
        ("Tally Voucher No.",22),("Tally Date",12),("Tally Party",30),
        ("Tally Amount (Rs.)",18),
    ]
    _hdr(ws, cols, 4, st, fill_key)
    ws.freeze_panes = "A5"
    for i, (or_, tr, method) in enumerate(rows, 1):
        row = 4 + i
        _cell(ws, row,  1, i,                    st, fmt=st["int"], align="ctr")
        _cell(ws, row,  2, or_.get("_month",""), st, align="ctr")
        _cell(ws, row,  3, method,               st, align="ctr")
        _cell(ws, row,  4, or_["invoice_number"],st)
        _cell(ws, row,  5, or_["date"],          st, align="ctr")
        _cell(ws, row,  6, or_["partner"],       st)
        _cell(ws, row,  7, or_["invoice_total"], st, fmt=st["cur"], align="rgt")
        _cell(ws, row,  8, tr["voucher_number"], st)
        _cell(ws, row,  9, tr["date"],           st, align="ctr")
        _cell(ws, row, 10, tr["party"],          st)
        _cell(ws, row, 11, tr["amount"],         st, fmt=st["cur"], align="rgt")
    if rows:
        tr_r = 5 + len(rows)
        ws.cell(row=tr_r, column=6,
                value=f"TOTAL MATCHED  ({len(rows)})").font = st["bold"]
        c = ws.cell(row=tr_r, column=7,
                    value=round(sum(o["invoice_total"] for o,_,_ in rows), 2))
        c.number_format = st["cur"]; c.font = st["bold"]
        c.border = st["thin"]; c.fill = st["green_f"]


# ── Excel writers (one per register) ─────────────────────────────────

def write_sales_excel(data: dict, output_path: Path):
    try: import openpyxl
    except ImportError: print("ERROR: pip install openpyxl"); sys.exit(1)
    st = _styles()
    wb = openpyxl.Workbook()

    _sheet_summary(wb, st, data["monthly"], "Sales", "GST INVOICE", "hdr_blue")
    _sheet_odoo_register(wb, "Odoo Sales Register", data["odoo_all"],   st, "hdr_blue", False)
    _sheet_tally_register(wb, "Tally GST Invoice Register", data["tally_all"], st, "hdr_blue", "GST INVOICE")
    _sheet_missing_in_tally(wb, "Missing in Tally", data["miss_tally"], st, "hdr_blue", False)
    _sheet_missing_in_odoo(wb, "Extra in Tally (Not in Odoo)", data["miss_odoo"],  st, "hdr_blue", False, "GST INVOICE")
    _sheet_mismatches(wb, "Amount Mismatches",   data["mismatches"],    st, "hdr_blue", False)
    _sheet_matched(wb, "Matched Pairs",          data["matched_pairs"], st, "hdr_blue", False)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Sales Excel    : {output_path}")


def write_purchase_excel(data: dict, output_path: Path):
    try: import openpyxl
    except ImportError: print("ERROR: pip install openpyxl"); sys.exit(1)
    st = _styles()
    wb = openpyxl.Workbook()

    _sheet_summary(wb, st, data["monthly"], "Purchase", "Purchase", "hdr_green")
    _sheet_odoo_register(wb, "Odoo Purchase Register", data["odoo_all"],  st, "hdr_green", True)
    _sheet_tally_register(wb, "Tally Purchase Register", data["tally_all"], st, "hdr_green", "Purchase")
    _sheet_missing_in_tally(wb, "Missing in Tally", data["miss_tally"], st, "hdr_green", True)
    _sheet_missing_in_odoo(wb, "Extra in Tally (Not in Odoo)", data["miss_odoo"], st, "hdr_green", True, "Purchase")
    _sheet_mismatches(wb, "Amount Mismatches",    data["mismatches"],    st, "hdr_green", True)
    _sheet_matched(wb, "Matched Pairs",           data["matched_pairs"], st, "hdr_green", True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Purchase Excel : {output_path}")


# ── Main orchestrator ─────────────────────────────────────────────────

def run(odoo, tally, run_sales, run_purchase):
    tally_ok = tally is not None
    sales_data = {
        "odoo_all":[], "tally_all":[], "miss_tally":[], "miss_odoo":[],
        "mismatches":[], "matched_pairs":[], "monthly":{}
    }
    purch_data = {
        "odoo_all":[], "tally_all":[], "miss_tally":[], "miss_odoo":[],
        "mismatches":[], "matched_pairs":[], "monthly":{}
    }
    tax_cache = {}

    for year, month in FY_MONTHS:
        label = month_label(year, month)
        fd, td = month_range(year, month)

        print(f"\n{'─'*60}")
        print(f"  {label}  ({fd}  to  {td})")
        print(f"{'─'*60}")

        # ── SALES ─────────────────────────────────────────────────────
        if run_sales:
            print("  [Sales] Fetching Odoo out_invoice ...")
            os_ = fetch_odoo_invoices(odoo, "out_invoice", fd, td, tax_cache)
            for r in os_: r["_month"] = label
            print(f"  [Sales] Odoo: {len(os_)} invoices  Rs.{sum(r['invoice_total'] for r in os_):,.2f}")

            ts_ = []
            if tally_ok:
                print("  [Sales] Fetching Tally 'GST INVOICE' ...")
                ts_ = fetch_tally_vouchers(tally, "GST INVOICE", fd, td)
                for r in ts_: r["_month"] = label
                print(f"  [Sales] Tally: {len(ts_)} vouchers  Rs.{sum(r['amount'] for r in ts_):,.2f}")

            recon = reconcile_sales(os_, ts_)
            ms    = month_stats(os_, ts_, recon)
            print(f"  [Sales] Matched:{ms['matched']}  Mismatch:{ms['mismatches']}  "
                  f"MissingTally:{ms['missing_tally']}  ExtraTally:{ms['missing_odoo']}")

            sales_data["monthly"][label]   = ms
            sales_data["odoo_all"]        += os_
            sales_data["tally_all"]       += ts_
            sales_data["miss_tally"]      += recon["only_odoo"]
            sales_data["miss_odoo"]       += recon["only_tally"]
            sales_data["mismatches"]      += [(o,t,m,n) for o,t,m,n in recon["mismatches"]]
            sales_data["matched_pairs"]   += [(o,t,m) for o,t,m in recon["matched"]]

        # ── PURCHASE ──────────────────────────────────────────────────
        if run_purchase:
            print("  [Purchase] Fetching Odoo in_invoice ...")
            op_ = fetch_odoo_invoices(odoo, "in_invoice", fd, td, tax_cache)
            for r in op_: r["_month"] = label
            print(f"  [Purchase] Odoo: {len(op_)} bills  Rs.{sum(r['invoice_total'] for r in op_):,.2f}")

            tp_ = []
            if tally_ok:
                print("  [Purchase] Fetching Tally 'Purchase' ...")
                tp_ = fetch_tally_vouchers(tally, "Purchase", fd, td)
                for r in tp_: r["_month"] = label
                print(f"  [Purchase] Tally: {len(tp_)} vouchers  Rs.{sum(r['amount'] for r in tp_):,.2f}")

            recon = reconcile_purchase(op_, tp_)
            mp    = month_stats(op_, tp_, recon)
            print(f"  [Purchase] Matched:{mp['matched']}  Mismatch:{mp['mismatches']}  "
                  f"MissingTally:{mp['missing_tally']}  ExtraTally:{mp['missing_odoo']}")

            purch_data["monthly"][label]   = mp
            purch_data["odoo_all"]        += op_
            purch_data["tally_all"]       += tp_
            purch_data["miss_tally"]      += recon["only_odoo"]
            purch_data["miss_odoo"]       += recon["only_tally"]
            purch_data["mismatches"]      += [(o,t,m,n) for o,t,m,n in recon["mismatches"]]
            purch_data["matched_pairs"]   += [(o,t,m) for o,t,m in recon["matched"]]

    return sales_data, purch_data


# ── Console summary ───────────────────────────────────────────────────

def _print_section(title, monthly):
    print(f"\n  {'─'*62}")
    print(f"  {title}")
    print(f"  {'─'*62}")
    print(f"  {'Month':<10} {'Odoo':>5} {'Tally':>5}  {'Odoo Rs.':>16} {'Tally Rs.':>16}"
          f"  {'Diff Rs.':>12} {'MissT':>6} {'ExtraT':>7} {'Mism':>5}")
    print(f"  {'─'*10} {'─'*5} {'─'*5}  {'─'*16} {'─'*16}  {'─'*12} {'─'*6} {'─'*7} {'─'*5}")
    oc=ot=tc=tt=mt=mo=mm=0
    for label, ms in monthly.items():
        diff = ms["odoo_total"] - ms["tally_total"]
        flag = " !" if (ms["missing_tally"] or ms["mismatches"]) else ""
        print(f"  {label:<10} {ms['odoo_count']:>5} {ms['tally_count']:>5}  "
              f"{ms['odoo_total']:>16,.2f} {ms['tally_total']:>16,.2f}  "
              f"{diff:>12,.2f} {ms['missing_tally']:>6} {ms['missing_odoo']:>7}"
              f" {ms['mismatches']:>5}{flag}")
        oc+=ms['odoo_count']; ot+=ms['odoo_total']
        tc+=ms['tally_count']; tt+=ms['tally_total']
        mt+=ms['missing_tally']; mo+=ms['missing_odoo']; mm+=ms['mismatches']
    print(f"  {'─'*10} {'─'*5} {'─'*5}  {'─'*16} {'─'*16}  {'─'*12} {'─'*6} {'─'*7} {'─'*5}")
    print(f"  {'TOTAL':<10} {oc:>5} {tc:>5}  {ot:>16,.2f} {tt:>16,.2f}  "
          f"{(ot-tt):>12,.2f} {mt:>6} {mo:>7} {mm:>5}")
    if mt: print(f"\n  *** {mt} invoice(s) missing in Tally — must be posted ***")
    if mm: print(f"  *** {mm} amount mismatch(es) — correct before audit sign-off ***")


def print_summary(sales_data, purch_data, run_sales, run_purchase):
    print(f"\n{'='*72}")
    print("   FY 2025-26  RECONCILIATION SUMMARY  (Apr-2025 to Mar-2026)")
    print(f"{'='*72}")
    if run_sales:
        _print_section("SALES  (Odoo out_invoice  vs  Tally 'GST INVOICE')",
                       sales_data["monthly"])
    if run_purchase:
        _print_section("PURCHASE  (Odoo in_invoice  vs  Tally 'Purchase')",
                       purch_data["monthly"])
    print(f"\n  NOTE: Credit Notes (out_refund) and Debit Notes (in_refund) are")
    print(f"        excluded. Reconcile separately using Tally Credit Note / Debit Note reports.")
    print(f"{'='*72}\n")


# ── CLI ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="FY 2025-26 Sales & Purchase Reconciliation — Odoo vs Tally",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--sales-only",    action="store_true")
    parser.add_argument("--purchase-only", action="store_true")
    parser.add_argument("--odoo-only",     action="store_true",
                        help="Odoo extract only (Tally offline)")
    parser.add_argument("--output-dir",    type=str, default=None,
                        help="Folder for Excel files (default: .tmp/)")
    args = parser.parse_args()

    run_sales    = not args.purchase_only
    run_purchase = not args.sales_only

    out_dir = Path(args.output_dir) if args.output_dir else TMP_DIR
    ts      = datetime.now().strftime("%Y%m%d_%H%M")

    sales_path = out_dir / f"FY2526_Sales_Reconciliation_{ts}.xlsx"
    purch_path = out_dir / f"FY2526_Purchase_Reconciliation_{ts}.xlsx"

    print("=" * 65)
    print("  FY 2025-26  SALES & PURCHASE RECONCILIATION")
    print("  Odoo vs Tally  |  Audit Grade  |  Apr-2025 to Mar-2026")
    print("  Sales   : Tally 'GST INVOICE' voucher type only")
    print("  Purchase: Tally 'Purchase'    voucher type only")
    print("  Excluded: Credit Notes, Debit Notes, Payments, Receipts,")
    print("            Journals, Contra, and ALL other voucher types")
    print("=" * 65)

    # ── Connect Odoo ─────────────────────────────────────────────────
    try:
        odoo = OdooConnector()
        ver  = odoo.version()
        print(f"\nOdoo  : {ver.get('server_version','?')}  |  {odoo.url}  |  DB:{odoo.db}")
    except Exception as e:
        print(f"Odoo connection failed: {e}"); sys.exit(1)

    # ── Connect Tally ─────────────────────────────────────────────────
    tally = None
    if not args.odoo_only:
        try:
            tally = TallyConnector()
            info  = tally.test_connection()
            print(f"Tally : {info['url']}  |  {info.get('target_company','?')}")
        except Exception as e:
            print(f"Tally not reachable: {e}")
            print("  Continuing in Odoo-only mode.")

    # ── Run ───────────────────────────────────────────────────────────
    sales_data, purch_data = run(odoo, tally, run_sales, run_purchase)

    # ── Summary ───────────────────────────────────────────────────────
    print_summary(sales_data, purch_data, run_sales, run_purchase)

    # ── Excel ─────────────────────────────────────────────────────────
    print("Saving Excel files ...")
    if run_sales:
        write_sales_excel(sales_data, sales_path)
    if run_purchase:
        write_purchase_excel(purch_data, purch_path)

    print("\nWorkbook sheets (each file):")
    for s in ["Summary", "Odoo Register", "Tally Register",
              "Missing in Tally  <- post these",
              "Extra in Tally (Not in Odoo)  <- investigate",
              "Amount Mismatches  <- correct before audit",
              "Matched Pairs  <- audit cross-reference"]:
        print(f"  * {s}")

    print("\nDone.\n")


if __name__ == "__main__":
    main()
