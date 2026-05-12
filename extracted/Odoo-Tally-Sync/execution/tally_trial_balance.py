#!/usr/bin/env python3
"""
Tally Trial Balance → Excel Filler (v3 — EXPLODEFLAG Approach)

Fetches monthly trial balance data from Tally Prime XML Server using
TB Report with EXPLODEFLAG=Yes. This returns ALL sub-groups and ledgers
with DR/CR/NET amounts, including P&L items.

The EXPLODEFLAG approach solves the fundamental bug where TDL Collection
CLOSINGBALANCE returned 0 for ALL P&L items.

One query per month → parse → map to Excel rows → write.

Usage:
    python execution/tally_trial_balance.py --test          # test connection
    python execution/tally_trial_balance.py --dry-run       # compute, don't write
    python execution/tally_trial_balance.py                 # full run
    python execution/tally_trial_balance.py --months 3      # last 3 months
    python execution/tally_trial_balance.py --overwrite     # overwrite existing cells
"""

import os
import sys
import json
import time
import argparse
import calendar
import re
from pathlib import Path
from datetime import date
from collections import OrderedDict
from typing import Optional, Dict, Tuple, List

import requests
import xml.etree.ElementTree as ET

# ── Paths ───────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

sys.path.insert(0, str(SCRIPT_DIR))
from tally_connector import TallyConnector

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# ── Configuration ───────────────────────────────────────────────────

DEFAULT_EXCEL = (
    r"D:\Suresh_AGENTS\Fracktal-Financial-Model\.tmp"
    r"\Fracktal_MIS_Master_20260224_181404.xlsx"
)
SHEET_NAME = "TB"

FY_START = date(2025, 4, 1)
FY_END = date(2026, 3, 31)

FY_MONTHS = [
    ("APR", 2025, 4),  ("MAY", 2025, 5),  ("JUN", 2025, 6),
    ("JUL", 2025, 7),  ("AUG", 2025, 8),  ("SEP", 2025, 9),
    ("OCT", 2025, 10), ("NOV", 2025, 11), ("DEC", 2025, 12),
    ("JAN", 2026, 1),  ("FEB", 2026, 2),  ("MAR", 2026, 3),
]

# Column B=2 (APR) through M=13 (MAR), N=14 (FY Total, formula)
MONTH_COL = {
    "APR": 2, "MAY": 3, "JUN": 4, "JUL": 5, "AUG": 6, "SEP": 7,
    "OCT": 8, "NOV": 9, "DEC": 10, "JAN": 11, "FEB": 12, "MAR": 13,
}

QUERY_DELAY = 3  # seconds between Tally queries to avoid overload

# ── Tally Display Name → Excel Mapping ──────────────────────────────
# TB Report with EXPLODEFLAG returns display names which may differ from
# Tally object names. This maps display_name → (excel_row_name, sign, section).
#
# Sign convention:
#   TB NET = DR + CR, where DR is negative, CR is positive.
#   Revenue/Income (credit balance): NET positive → Excel positive → sign = +1
#   Expenses (debit balance):        NET negative → Excel positive → sign = -1
#   BS items (all):                  Excel = -NET always        → sign = -1
#     Assets (debit, NET<0): -NET = positive ✓
#     Liabilities (credit, NET>0): -NET = negative (shows as liability)
#     Debit-balance liability (NET<0): -NET = positive (unusual but correct)

# tally_display_name → (excel_row_name, sign_multiplier, "pnl"|"bs")
TALLY_TO_EXCEL = {
    # ── Revenue ─────────────────────────────────────────────────
    "Sale of Products":                        ("Sale of Products",                +1, "pnl"),
    "Sale of Service":                         ("Sale of Service",                 +1, "pnl"),
    "Export Sales":                            ("Export Sales",                    +1, "pnl"),
    "Printsticks":                             ("Printsticks",                     +1, "pnl"),
    "Discount Received":                       ("Discount Received",               +1, "pnl"),
    "Interest Income":                         ("Interest Income",                 +1, "pnl"),
    "Other Income":                            ("Other Income",                    +1, "pnl"),
    # ── COGS / Purchase ────────────────────────────────────────
    "PURCHASE of Raw Materials":               ("PURCHASE of Raw Materials",       -1, "pnl"),
    "IMPORT of Raw Materials":                 ("IMPORT of Raw Materials",         -1, "pnl"),
    "OTHER PURCHASES":                         ("OTHER PURCHASES",                -1, "pnl"),
    # ── Direct Expenses ─────────────────────────────────────────
    "Salaries (Production)":                   ("Salaries (Production)",           -1, "pnl"),
    "Overtime Pay":                            ("Overtime Pay",                    -1, "pnl"),
    "Electricity Charges - Fractory Services": ("Electricity - Factory",           -1, "pnl"),
    "Electricity Charges-Manufacturing":       ("Electricity - Manufacturing",     -1, "pnl"),
    "Freight Inward":                          ("Freight Inward",                  -1, "pnl"),
    "Loading & Unloading":                     ("Loading & Unloading",             -1, "pnl"),
    "Discount Allowed":                        ("Discount Allowed",                -1, "pnl"),
    # ── Indirect Expenses ───────────────────────────────────────
    "Office & Administrative Overheads":       ("Office & Administrative Overheads", -1, "pnl"),
    "Travelling Expense":                      ("Travelling Expense",              -1, "pnl"),
    "Rates & Taxes":                           ("Rates & Taxes",                   -1, "pnl"),
    "Foreign Exchange Gain/Loss":              ("Foreign Exchange Gain/Loss",      -1, "pnl"),
    "Razorpay Charges":                        ("Razorpay Charges",                -1, "pnl"),
    "Tender Fee":                              ("Tender Fee",                      -1, "pnl"),
    "Finance Cost":                            ("Finance Cost (Interest/Bank)",    -1, "pnl"),
    "Payroll Expenses":                        ("Payroll Expenses",                -1, "pnl"),
    "Advertisement/Marketing":                 ("Advertisement/Marketing",         -1, "pnl"),
    "Freight Outward":                         ("Freight Outward",                 -1, "pnl"),
    "R&D":                                     ("R&D",                             -1, "pnl"),
    "Professional Service Charges":            ("Professional Service Charges",    -1, "pnl"),
    "Round Off":                               ("Round Off",                       -1, "pnl"),
    "Write Off / Write Back":                  ("Write Off / Write Back",          -1, "pnl"),
    # ── BS: Capital Account (credit balance → positive in Excel) ──
    "Reserves & Surplus":                      ("Reserves & Surplus",              +1, "bs"),
    "Issued, Subscribed and Paid Up":          ("Issued, Subscribed & Paid Up",    +1, "bs"),
    # ── BS: Loans (credit balance → positive in Excel) ──────────
    "Unsecured Loans":                         ("Unsecured Loans",                 +1, "bs"),
    # ── BS: Current Liabilities ─────────────────────────────────
    "Duties & Taxes":                          ("Duties & Taxes",                  -1, "bs"),
    "Sundry Creditors":                        ("Sundry Creditors",                -1, "bs"),
    "Reimbursments Payable":                   ("Reimbursements Payable",          -1, "bs"),
    "Salaries Payable":                        ("Salaries Payable",                +1, "bs"),
    # ── BS: Fixed Assets ────────────────────────────────────────
    "Intangible Assets":                       ("Intangible Assets",               -1, "bs"),
    "Machinery & Equipment":                   ("Machinery & Equipment",           -1, "bs"),
    "Tangible Assets":                         ("Tangible Assets",                 -1, "bs"),
    "Fracktal Studio":                         ("Fracktal Studio",                 -1, "bs"),
    "SLS PAMM Project":                        ("SLS PAMM Project",                -1, "bs"),
    "WIP":                                     ("WIP",                             -1, "bs"),
    # ── BS: Current Assets ──────────────────────────────────────
    "Opening Stock":                           ("Opening Stock",                   -1, "bs"),
    "Deposits (Asset)":                        ("Deposits (Asset)",                -1, "bs"),
    "Loans & Advances (Asset)":                ("Loans & Advances (Asset)",        -1, "bs"),
    "Sundry Debtors":                          ("Sundry Debtors",                  -1, "bs"),
    "Cash-in-Hand":                            ("Cash-in-Hand",                    -1, "bs"),
    "Bank Accounts":                           ("Bank Accounts",                   -1, "bs"),
    "Deferred Expenses":                       ("Deferred Expenses",               -1, "bs"),
    "Deferred Tax Asset":                      ("Deferred Tax Asset",              +1, "bs"),
    "TDS - GST Receivable":                    ("TDS - GST Receivable",            -1, "bs"),
    "Tds Receivable":                          ("Tds Receivable",                  -1, "bs"),
    # ── BS: Suspense / P&L ──────────────────────────────────────
    "Suspense A/c":                            ("Suspense Account",                -1, "bs"),
    "Profit & Loss A/c":                       ("Profit & Loss A/c",               -1, "bs"),
}

# Derived sets for summary counting
PNL_EXCEL_NAMES = {v[0] for v in TALLY_TO_EXCEL.values() if v[2] == "pnl"}
BS_EXCEL_NAMES  = {v[0] for v in TALLY_TO_EXCEL.values() if v[2] == "bs"}

# Section headers (non-data rows) to skip when reading Excel
SECTION_HEADERS = {
    "REVENUE", "COST OF GOODS SOLD (COGS)",
    "DIRECT / MANUFACTURING EXPENSES",
    "INDIRECT EXPENSES (OPEX)", "Office & Admin Overheads",
    "Finance Cost", "HR / Payroll Expenses",
    "Marketing & Ads", "R&D Expenses",
    "BALANCE SHEET ITEMS", "Capital Account",
    "Loans (Liability)", "Current Liabilities",
    "Fixed Assets", "Current Assets", "Suspense",
}


# ═══════════════════════════════════════════════════════════════════
# Tally XML Helpers
# ═══════════════════════════════════════════════════════════════════

def sanitize_xml(text: str) -> str:
    """Clean Tally XML for safe parsing."""
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    text = re.sub(r'&#x([0-8BbCcEeFf]);', '', text)
    text = re.sub(r'&#x1[0-9A-Fa-f];', '', text)
    text = re.sub(r'&#([0-8]|1[0-1]|1[4-9]|2[0-9]|3[01]);', '', text)
    return text


def parse_amount(text: str) -> float:
    """Parse Tally amount string → float."""
    if not text or not text.strip():
        return 0.0
    text = text.strip().replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
    for suffix in (" Cr.", " Cr", " Dr.", " Dr"):
        if text.endswith(suffix):
            text = text[:-len(suffix)].strip()
            break
    try:
        return float(text)
    except ValueError:
        return 0.0


def send_xml(url: str, xml: str, timeout: int = 120, retries: int = 1,
             label: str = "") -> Optional[str]:
    """Send XML to Tally with retry logic."""
    for attempt in range(retries + 1):
        try:
            if attempt > 0:
                wait = 10 * attempt
                print(f"    ↻ Retry {attempt}/{retries} for {label} (wait {wait}s)")
                time.sleep(wait)
            resp = requests.post(
                url, data=xml.encode("utf-8"),
                headers={"Content-Type": "text/xml; charset=utf-8"},
                timeout=timeout,
            )
            resp.raise_for_status()
            return resp.text
        except (requests.Timeout, requests.ConnectionError) as e:
            print(f"    ✗ {label}: {type(e).__name__} (attempt {attempt+1})")
            if attempt == retries:
                return None
    return None


# ═══════════════════════════════════════════════════════════════════
# Core: TB with EXPLODEFLAG
# ═══════════════════════════════════════════════════════════════════

def fmt_date(d: date) -> str:
    """Format date as Tally YYYYMMDD."""
    return f"{d.year:04d}{d.month:02d}{d.day:02d}"


def month_range(year: int, month: int) -> Tuple[str, str]:
    """Return (from_date, to_date) strings for a month."""
    last = calendar.monthrange(year, month)[1]
    return fmt_date(date(year, month, 1)), fmt_date(date(year, month, last))


def fetch_tb_explode(url: str, company: str, from_date: str, to_date: str,
                     timeout: int = 120) -> Dict[str, float]:
    """
    Fetch Trial Balance with EXPLODEFLAG=Yes.
    Returns ALL sub-groups AND ledgers: {display_name: net_value}
    where NET = DR + CR (DR is negative, CR is positive).

    P&L items: monthly movement for the date range.
    BS items: cumulative balance as of to_date.
    """
    xml = f"""<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Data</TYPE><ID>Trial Balance</ID></HEADER><BODY><DESC>
    <STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
    <SVFROMDATE>{from_date}</SVFROMDATE><SVTODATE>{to_date}</SVTODATE>
    <EXPLODEFLAG>Yes</EXPLODEFLAG>
    </STATICVARIABLES></DESC></BODY></ENVELOPE>"""

    resp = send_xml(url, xml, timeout=timeout, retries=1,
                    label=f"TB-Explode({from_date}→{to_date})")
    if not resp:
        return {}

    root = ET.fromstring(sanitize_xml(resp))
    result = {}
    children = list(root)
    i = 0
    while i < len(children):
        if children[i].tag == "DSPACCNAME":
            name = (children[i].findtext("DSPDISPNAME") or "").strip()
            if i + 1 < len(children) and children[i + 1].tag == "DSPACCINFO":
                info = children[i + 1]
                dr = parse_amount(info.findtext(".//DSPCLDRAMTA", ""))
                cr = parse_amount(info.findtext(".//DSPCLCRAMTA", ""))
                net = dr + cr
                if name:
                    result[name] = net
                i += 2
                continue
        i += 1
    return result


def map_tally_to_excel(tb_data: Dict[str, float]) -> Dict[str, float]:
    """
    Convert Tally TB EXPLODEFLAG data → Excel-ready values.
    Applies name mapping and sign conversion from TALLY_TO_EXCEL.
    Returns: {excel_row_name: value}
    """
    result = {}
    for tally_name, net in tb_data.items():
        if tally_name in TALLY_TO_EXCEL:
            excel_name, sign, _section = TALLY_TO_EXCEL[tally_name]
            val = round(sign * net, 2)
            if val != 0:
                result[excel_name] = val
    return result


# ═══════════════════════════════════════════════════════════════════
# Excel Functions
# ═══════════════════════════════════════════════════════════════════

def find_excel_row_map(ws) -> Dict[str, int]:
    """Build mapping: account name → Excel row number."""
    row_map = {}
    for row in ws.iter_rows(min_row=4, max_row=75, min_col=1, max_col=1):
        cell = row[0]
        if cell.value and isinstance(cell.value, str):
            name = cell.value.strip()
            if name and name not in SECTION_HEADERS:
                row_map[name] = cell.row
    return row_map


def read_existing_values(ws, row_map: Dict[str, int]) -> Dict[str, Dict[str, float]]:
    """Read existing Excel values for validation. Returns {month: {name: value}}."""
    existing = {}
    for label, col in MONTH_COL.items():
        month_data = {}
        for name, row in row_map.items():
            val = ws.cell(row=row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                month_data[name] = float(val)
        if month_data:
            existing[label] = month_data
    return existing


# ═══════════════════════════════════════════════════════════════════
# Main Pipeline
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Tally Trial Balance → Excel (EXPLODEFLAG approach)")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Excel file path")
    parser.add_argument("--dry-run", action="store_true", help="Print data, don't write")
    parser.add_argument("--months", type=int, default=0, help="Last N months (0=all)")
    parser.add_argument("--timeout", type=int, default=120, help="Per-query timeout (s)")
    parser.add_argument("--save-json", action="store_true", help="Save raw data to .tmp/")
    parser.add_argument("--test", action="store_true", help="Test Tally connection only")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing cells")
    args = parser.parse_args()

    TMP_DIR.mkdir(parents=True, exist_ok=True)

    # ── Connect ─────────────────────────────────────────────────
    tally = TallyConnector()
    url = tally.base_url
    company = tally.company

    print(f"Connecting to Tally at {url}...")
    try:
        info = tally.test_connection()
        print(f"✓ Connected: {info['target_company']}")
    except Exception as e:
        print(f"✗ Connection failed: {e}")
        sys.exit(1)

    if args.test:
        return

    # ── Determine months to process ─────────────────────────────
    today = date.today()
    available_months = []
    for label, year, month in FY_MONTHS:
        if date(year, month, 1) <= today:
            available_months.append((label, year, month))
    if args.months > 0:
        available_months = available_months[-args.months:]

    month_labels = [m[0] for m in available_months]
    print(f"\nMonths: {', '.join(month_labels)} ({len(available_months)} months)")

    # ══════════════════════════════════════════════════════════════
    # Fetch each month using TB + EXPLODEFLAG
    # ══════════════════════════════════════════════════════════════
    all_data = {}       # {label: {excel_name: value}}
    all_raw = {}        # {label: {tally_display_name: net}} for debugging
    all_unmapped = {}   # {label: [(name, net)]} — items not in our mapping

    for idx, (label, year, month) in enumerate(available_months):
        m_from, m_to = month_range(year, month)
        print(f"\n[{idx+1}/{len(available_months)}] {label} {year} ({m_from}→{m_to})...")

        tb_raw = fetch_tb_explode(url, company, m_from, m_to, args.timeout)
        if not tb_raw:
            print(f"  ✗ No data for {label}")
            time.sleep(QUERY_DELAY)
            continue

        all_raw[label] = tb_raw

        # Map to Excel values
        month_values = map_tally_to_excel(tb_raw)
        all_data[label] = month_values

        # Track unmapped items (for discovering new accounts)
        mapped_tally = set(TALLY_TO_EXCEL.keys())
        unmapped = [(n, v) for n, v in tb_raw.items()
                    if n not in mapped_tally and v != 0]
        if unmapped:
            all_unmapped[label] = unmapped

        # Summary
        pnl_count = sum(1 for k in month_values if k in PNL_EXCEL_NAMES)
        bs_count = sum(1 for k in month_values if k in BS_EXCEL_NAMES)
        print(f"  → {len(tb_raw)} TB entries, mapped: {pnl_count} P&L + {bs_count} BS")

        time.sleep(QUERY_DELAY)

    # ══════════════════════════════════════════════════════════════
    # Save raw data
    # ══════════════════════════════════════════════════════════════
    if args.save_json:
        json_path = TMP_DIR / "tally_tb_data_v3.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"mapped": all_data, "raw": all_raw}, f, indent=2, default=str)
        print(f"\n✓ Data saved: {json_path}")

    # ══════════════════════════════════════════════════════════════
    # Summary table
    # ══════════════════════════════════════════════════════════════
    print(f"\n{'='*100}")
    print("TRIAL BALANCE SUMMARY")
    print(f"{'='*100}")

    header = f"{'Account':<45}" + "".join(f"{m:>12}" for m in month_labels)
    print(header)
    print("─" * len(header))

    print("── P&L ──")
    for tally_name, (excel_name, sign, section) in TALLY_TO_EXCEL.items():
        if section != "pnl":
            continue
        vals = [all_data.get(l, {}).get(excel_name, 0) for l in month_labels]
        if any(v != 0 for v in vals):
            fmt_vals = "".join(f"{v:>12,.0f}" if v else f"{'':>12}" for v in vals)
            print(f"  {excel_name:<43}{fmt_vals}")

    print("\n── Balance Sheet ──")
    for tally_name, (excel_name, sign, section) in TALLY_TO_EXCEL.items():
        if section != "bs":
            continue
        vals = [all_data.get(l, {}).get(excel_name, 0) for l in month_labels]
        if any(v != 0 for v in vals):
            fmt_vals = "".join(f"{v:>12,.0f}" if v else f"{'':>12}" for v in vals)
            print(f"  {excel_name:<43}{fmt_vals}")

    # Show unmapped items (potential missing mappings)
    if all_unmapped:
        print(f"\n── Unmapped Tally Items (not in TALLY_TO_EXCEL) ──")
        # Collect all unmapped across months
        seen = set()
        for label, items in all_unmapped.items():
            for name, net in items:
                if name not in seen:
                    print(f"  {name}: {net:>14,.2f} ({label})")
                    seen.add(name)

    # ══════════════════════════════════════════════════════════════
    # Validate against existing Excel data
    # ══════════════════════════════════════════════════════════════
    excel_path = Path(args.excel)
    if excel_path.exists():
        print(f"\n{'='*65}")
        print("VALIDATION vs existing Excel data")
        print(f"{'='*65}")
        wb_check = openpyxl.load_workbook(str(excel_path), data_only=True)
        ws_check = wb_check[SHEET_NAME]
        row_map = find_excel_row_map(ws_check)
        existing = read_existing_values(ws_check, row_map)
        wb_check.close()

        matches = 0
        mismatches = 0
        for label in month_labels:
            if label not in existing:
                continue
            for excel_name, computed_val in all_data.get(label, {}).items():
                excel_val = existing[label].get(excel_name)
                if excel_val is not None and computed_val != 0:
                    diff = abs(computed_val - excel_val)
                    pct = (diff / abs(excel_val) * 100) if excel_val != 0 else 100
                    if diff > 1.0 and pct > 1:
                        print(f"  {label} {excel_name:40s}: "
                              f"Excel={excel_val:>14,.2f}  "
                              f"Tally={computed_val:>14,.2f}  "
                              f"Δ={diff:>10,.2f} ({pct:.1f}%)")
                        mismatches += 1
                    else:
                        matches += 1

        if mismatches == 0:
            print(f"  ✓ All {matches} compared values match within 1% tolerance")
        else:
            print(f"\n  ✓ {matches} values match")
            print(f"  ⚠ {mismatches} differences found (see above)")

        # Show unmapped Excel rows (rows in Excel that we don't fill)
        all_excel_names = PNL_EXCEL_NAMES | BS_EXCEL_NAMES
        unmapped_rows = [name for name in row_map if name not in all_excel_names]
        if unmapped_rows:
            print(f"\n  Excel rows not mapped to Tally: {unmapped_rows}")

    # ══════════════════════════════════════════════════════════════
    # Write to Excel
    # ══════════════════════════════════════════════════════════════
    if args.dry_run:
        print("\n[DRY RUN] Not writing to Excel.")
        return

    if not excel_path.exists():
        print(f"\n✗ Excel not found: {excel_path}")
        sys.exit(1)

    print(f"\nWriting to {excel_path}...")
    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb[SHEET_NAME]
    row_map = find_excel_row_map(ws)

    cells_written = 0
    cells_skipped = 0

    for label in month_labels:
        col = MONTH_COL.get(label)
        if not col:
            continue
        month_values = all_data.get(label, {})
        for excel_name, value in month_values.items():
            row = row_map.get(excel_name)
            if row is None:
                print(f"  ⚠ No row for '{excel_name}'")
                continue

            existing_val = ws.cell(row=row, column=col).value
            if existing_val is not None and not args.overwrite:
                cells_skipped += 1
                continue

            ws.cell(row=row, column=col, value=value if value != 0 else None)
            cells_written += 1

    wb.save(str(excel_path))
    print(f"✓ Written {cells_written} cells ({cells_skipped} skipped, had data)")
    print(f"✓ Saved {excel_path}")


if __name__ == "__main__":
    main()
