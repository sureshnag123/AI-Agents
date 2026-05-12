"""
GST Compliance Report Builder — Excel Output
=============================================
Creates a professional, multi-sheet Excel workbook with:
  - Executive Dashboard
  - GSTR-1 tables (B2B, B2CL, B2CS, CDN, HSN)
  - Reconciliation results
  - ITC analysis & vendor reports
  - Alerts & compliance actions
  - Monthly automation schedule
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime

# ── THEME ──────────────────────────────────────────────────────

NAVY = "0F1B2D"
TEAL = "00897B"
TEAL_DARK = "00695C"
ORANGE = "E65100"
RED = "C62828"
GREEN = "2E7D32"
AMBER = "F57F17"
BLUE = "1565C0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "E0E0E0"
LIGHT_TEAL = "E0F2F1"
LIGHT_RED = "FFEBEE"
LIGHT_AMBER = "FFF8E1"
LIGHT_GREEN = "E8F5E9"
LIGHT_BLUE = "E3F2FD"

BORDER = Border(
    left=Side("thin", color="BDBDBD"), right=Side("thin", color="BDBDBD"),
    top=Side("thin", color="BDBDBD"), bottom=Side("thin", color="BDBDBD"),
)

SEVERITY_COLORS = {
    "CRITICAL": (RED, LIGHT_RED), "HIGH": (ORANGE, LIGHT_RED),
    "MEDIUM": (AMBER, LIGHT_AMBER), "LOW": (GREEN, LIGHT_GREEN),
    "INFO": (BLUE, LIGHT_BLUE), "ACTION": (TEAL, LIGHT_TEAL),
    "WARNING": (AMBER, LIGHT_AMBER),
}

def _hdr(cell, bg=NAVY, fg=WHITE, sz=10):
    cell.font = Font("Arial", bold=True, size=sz, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def _cell(cell, bold=False, color="000000", bg=None, fmt=None):
    cell.font = Font("Arial", size=10, bold=bold, color=color)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center")
    if fmt: cell.number_format = fmt

def _title(ws, text, row, col, span, bg=NAVY):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font("Arial", bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    for i in range(col, col+span):
        ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=row, column=i).border = BORDER
    ws.row_dimensions[row].height = 32

def _write_df(ws, df, sr=1, sc=1, hdr_bg=NAVY):
    if len(df) == 0:
        ws.cell(row=sr, column=sc, value="No data available")
        return sr + 1
    for ci, col in enumerate(df.columns, sc):
        _hdr(ws.cell(row=sr, column=ci, value=col), bg=hdr_bg)
    for ri, (_, row) in enumerate(df.iterrows(), sr+1):
        bg = LIGHT_GRAY if ri % 2 == 0 else None
        for ci, col in enumerate(df.columns, sc):
            val = row[col]
            if isinstance(val, (np.integer,)): val = int(val)
            elif isinstance(val, (np.floating,)): val = float(val)
            c = ws.cell(row=ri, column=ci, value=val)
            fmt = '₹#,##0.00' if isinstance(val, float) else None
            _cell(c, bg=bg, fmt=fmt)
    for ci, col in enumerate(df.columns, sc):
        ml = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) > 0 else 5)
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 4, 32)
    return sr + 1 + len(df)

def _summary_block(ws, data, sr, sc=1, title="Summary", bg_title=NAVY):
    _title(ws, title, sr, sc, 2, bg=bg_title)
    r = sr + 1
    for k, v in data.items():
        kc = ws.cell(row=r, column=sc, value=k)
        _cell(kc, bold=True, bg=LIGHT_TEAL)
        vc = ws.cell(row=r, column=sc+1, value=v)
        fmt = '₹#,##0.00' if isinstance(v, float) else ('#,##0' if isinstance(v, int) else None)
        color = RED if isinstance(v, (int,float)) and v < 0 else "000000"
        _cell(vc, fmt=fmt, color=color)
        r += 1
    ws.column_dimensions[get_column_letter(sc)].width = 32
    ws.column_dimensions[get_column_letter(sc+1)].width = 22
    return r


def build_workbook(gstr1, recon, itc, alerts_df, sales_df, purchase_df, gstr2b_df):
    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # SHEET 1: EXECUTIVE DASHBOARD
    # ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = NAVY

    _title(ws, "GST COMPLIANCE AUTOMATION REPORT — Q1 FY 2025-26", 1, 1, 10, bg=NAVY)
    ts = ws.cell(row=2, column=1, value=f"Report Generated: {datetime.now().strftime('%d %B %Y, %H:%M IST')}  |  Engine: Odoo ERP → GST Portal  |  Period: Jan–Mar 2025")
    ts.font = Font("Arial", size=9, italic=True, color="757575")

    r = _summary_block(ws, gstr1["summary"], 4, 1, "GSTR-1 SUMMARY", TEAL_DARK)
    r = _summary_block(ws, recon["summary"], r+1, 1, "GSTR-2B RECONCILIATION", NAVY)
    r = _summary_block(ws, itc["summary"], r+1, 1, "ITC ANALYSIS", TEAL_DARK)

    # Alerts panel (right side)
    _title(ws, "COMPLIANCE ALERTS & REQUIRED ACTIONS", 4, 5, 5, bg=RED)
    ar = 5
    for _, alert in alerts_df.iterrows():
        sev = alert["Severity"]
        fg, bg = SEVERITY_COLORS.get(sev, (BLUE, LIGHT_BLUE))
        sc = ws.cell(row=ar, column=5, value=sev)
        _cell(sc, bold=True, color=fg, bg=bg)
        cat = ws.cell(row=ar, column=6, value=alert["Category"])
        _cell(cat, bg=bg)
        msg = ws.cell(row=ar, column=7, value=alert["Alert"])
        _cell(msg, bg=bg)
        act = ws.cell(row=ar, column=8, value=alert["Action"])
        _cell(act, bg=bg)
        imp = ws.cell(row=ar, column=9, value=alert.get("Impact", ""))
        _cell(imp, bg=bg, color=RED if sev in ("CRITICAL","HIGH") else "000000")
        ar += 1

    for c in [5]: ws.column_dimensions[get_column_letter(c)].width = 12
    for c in [6]: ws.column_dimensions[get_column_letter(c)].width = 16
    for c in [7]: ws.column_dimensions[get_column_letter(c)].width = 48
    for c in [8]: ws.column_dimensions[get_column_letter(c)].width = 42
    for c in [9]: ws.column_dimensions[get_column_letter(c)].width = 28

    # ════════════════════════════════════════════════════════════
    # SHEET 2: GSTR-1 B2B
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("GSTR1 — B2B (Table 4)")
    ws2.sheet_properties.tabColor = TEAL
    cols = ["Invoice Number","Invoice Date","Customer GSTIN","Customer Name",
            "Place of Supply","Taxable Value","IGST","CGST","SGST","Invoice Value","HSN Code","Tax Rate"]
    _title(ws2, "GSTR-1 Table 4A/4B — B2B Supply to Registered Persons", 1, 1, len(cols))
    _write_df(ws2, gstr1["b2b"][cols], sr=3)

    # ════════════════════════════════════════════════════════════
    # SHEET 3: GSTR-1 B2C LARGE
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("GSTR1 — B2CL (Table 5)")
    ws3.sheet_properties.tabColor = TEAL
    cols3 = ["Invoice Number","Invoice Date","Place of Supply","Taxable Value","IGST","Invoice Value"]
    _title(ws3, "GSTR-1 Table 5 — B2C Large (Interstate > ₹2.5 Lakh)", 1, 1, len(cols3))
    if len(gstr1["b2c_large"]) > 0:
        _write_df(ws3, gstr1["b2c_large"][cols3], sr=3)
    else:
        ws3.cell(row=3, column=1, value="No B2C Large invoices in this period")

    # ════════════════════════════════════════════════════════════
    # SHEET 4: GSTR-1 B2C SMALL
    # ════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("GSTR1 — B2CS (Table 7)")
    ws4.sheet_properties.tabColor = TEAL
    _title(ws4, "GSTR-1 Table 7 — B2C Others (State-wise Aggregate)", 1, 1, 5)
    if len(gstr1["b2c_small"]) > 0:
        agg = gstr1["b2c_small"].groupby("Place of Supply").agg(
            Taxable_Value=("Taxable Value","sum"),
            CGST=("CGST","sum"), SGST=("SGST","sum"), IGST=("IGST","sum")
        ).reset_index().round(2)
        _write_df(ws4, agg, sr=3)

    # ════════════════════════════════════════════════════════════
    # SHEET 5: CREDIT/DEBIT NOTES
    # ════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("GSTR1 — CDN (Table 9)")
    ws5.sheet_properties.tabColor = ORANGE
    cols5 = ["Invoice Number","Invoice Date","Customer GSTIN","Customer Name",
             "Taxable Value","IGST","CGST","SGST","Invoice Value"]
    _title(ws5, "GSTR-1 Table 9 — Credit/Debit Notes", 1, 1, len(cols5))
    _write_df(ws5, gstr1["cdn"][cols5], sr=3)

    # ════════════════════════════════════════════════════════════
    # SHEET 6: EXPORTS
    # ════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("GSTR1 — Exports")
    ws6.sheet_properties.tabColor = TEAL
    _title(ws6, "GSTR-1 Table 6A — Export Invoices", 1, 1, 9)
    if len(gstr1["exports"]) > 0:
        _write_df(ws6, gstr1["exports"][cols5], sr=3)
    else:
        ws6.cell(row=3, column=1, value="No export invoices in this period")

    # ════════════════════════════════════════════════════════════
    # SHEET 7: HSN SUMMARY
    # ════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("HSN Summary")
    ws7.sheet_properties.tabColor = TEAL
    _title(ws7, "GSTR-1 HSN-wise Summary (Table 12)", 1, 1, len(gstr1["hsn_summary"].columns))
    end_r = _write_df(ws7, gstr1["hsn_summary"], sr=3)

    # Add totals row
    n_data = len(gstr1["hsn_summary"])
    if n_data > 0:
        tr = end_r + n_data
        ws7.cell(row=tr, column=1, value="TOTAL").font = Font("Arial", bold=True, size=10)
        for ci, col in enumerate(gstr1["hsn_summary"].columns):
            if col in ("Taxable_Value","IGST","CGST","SGST","Total_Value","Count"):
                c = ws7.cell(row=tr, column=ci+1)
                col_letter = get_column_letter(ci+1)
                c.value = f"=SUM({col_letter}4:{col_letter}{tr-1})"
                _cell(c, bold=True, fmt='₹#,##0.00')

    # ════════════════════════════════════════════════════════════
    # SHEET 8: MONTHLY SUMMARY
    # ════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("Monthly Breakdown")
    ws8.sheet_properties.tabColor = NAVY
    _title(ws8, "Month-wise GST Liability Summary", 1, 1, len(gstr1["monthly"].columns))
    _write_df(ws8, gstr1["monthly"], sr=3)

    # ════════════════════════════════════════════════════════════
    # SHEET 9: RECONCILIATION - FULL
    # ════════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("2B Reconciliation")
    ws9.sheet_properties.tabColor = RED
    _title(ws9, "GSTR-2B Reconciliation — Invoice Match Results", 1, 1, 13)
    if len(recon["matched"]) > 0:
        _write_df(ws9, recon["matched"], sr=3)
        for ri in range(4, 4 + len(recon["matched"])):
            st = ws9.cell(row=ri, column=11).value
            if st and "Mismatch" in str(st):
                for ci in range(1, 14):
                    ws9.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=LIGHT_RED)

    # ════════════════════════════════════════════════════════════
    # SHEET 10: MISSING IN 2B
    # ════════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("Missing in 2B")
    ws10.sheet_properties.tabColor = RED
    _title(ws10, "Invoices in Books but MISSING from GSTR-2B — ITC AT RISK", 1, 1, 10)
    _write_df(ws10, recon["missing_in_2b"], sr=3)

    # ════════════════════════════════════════════════════════════
    # SHEET 11: MISSING IN BOOKS
    # ════════════════════════════════════════════════════════════
    ws11 = wb.create_sheet("Missing in Books")
    ws11.sheet_properties.tabColor = ORANGE
    _title(ws11, "Invoices in GSTR-2B but NOT in Books — Action Required", 1, 1, 10)
    _write_df(ws11, recon["missing_in_books"], sr=3)

    # ════════════════════════════════════════════════════════════
    # SHEET 12: VENDOR ITC REPORT
    # ════════════════════════════════════════════════════════════
    ws12 = wb.create_sheet("Vendor ITC Analysis")
    ws12.sheet_properties.tabColor = ORANGE
    vcols = ["Vendor GSTIN","Vendor_Name","Invoice_Count","Books_Taxable","Books_ITC","Portal_ITC","ITC_Difference","Risk"]
    vitc = itc["vendor_itc"]
    if "Filing_Status" in vitc.columns:
        vcols.append("Filing_Status")
    avail_cols = [c for c in vcols if c in vitc.columns]
    _title(ws12, "Vendor-wise ITC Difference Report", 1, 1, len(avail_cols))
    if len(vitc) > 0:
        end_r = _write_df(ws12, vitc[avail_cols], sr=3)
        for ri in range(4, 4 + len(vitc)):
            risk_col = avail_cols.index("Risk") + 1 if "Risk" in avail_cols else None
            if risk_col:
                risk = ws12.cell(row=ri, column=risk_col).value
                _, bg = SEVERITY_COLORS.get(risk, (BLUE, LIGHT_BLUE))
                for ci in range(1, len(avail_cols)+1):
                    ws12.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=bg)

    # ════════════════════════════════════════════════════════════
    # SHEET 13: BLOCKED ITC DETAIL
    # ════════════════════════════════════════════════════════════
    ws13 = wb.create_sheet("Blocked ITC Detail")
    ws13.sheet_properties.tabColor = AMBER
    _title(ws13, "Blocked ITC Breakdown — Section 17(5)", 1, 1, 3)
    if len(itc.get("blocked_detail", pd.DataFrame())) > 0:
        _write_df(ws13, itc["blocked_detail"], sr=3)
    else:
        ws13.cell(row=3, column=1, value="No blocked ITC breakdown available")

    # ════════════════════════════════════════════════════════════
    # SHEET 14: VALIDATION ERRORS
    # ════════════════════════════════════════════════════════════
    ws14 = wb.create_sheet("Validation Errors")
    ws14.sheet_properties.tabColor = RED
    _title(ws14, "Data Validation Errors — Fix Before Filing", 1, 1, 4)
    if len(gstr1["errors"]) > 0:
        _write_df(ws14, gstr1["errors"], sr=3)
        for ri in range(4, 4 + len(gstr1["errors"])):
            sev = ws14.cell(row=ri, column=4).value
            _, bg = SEVERITY_COLORS.get(sev, (BLUE, LIGHT_BLUE))
            for ci in range(1, 5):
                ws14.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=bg)

    # ════════════════════════════════════════════════════════════
    # SHEET 15: ALERTS
    # ════════════════════════════════════════════════════════════
    ws15 = wb.create_sheet("Compliance Alerts")
    ws15.sheet_properties.tabColor = RED
    _title(ws15, "Compliance Alerts & Required Actions", 1, 1, 5)
    if len(alerts_df) > 0:
        _write_df(ws15, alerts_df, sr=3)
        for ri in range(4, 4 + len(alerts_df)):
            sev = ws15.cell(row=ri, column=1).value
            fg, bg = SEVERITY_COLORS.get(sev, (BLUE, LIGHT_BLUE))
            for ci in range(1, 6):
                ws15.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=bg)

    # ════════════════════════════════════════════════════════════
    # SHEET 16: AUTOMATION SCHEDULE
    # ════════════════════════════════════════════════════════════
    from gst_engine import get_monthly_schedule
    ws16 = wb.create_sheet("Monthly Schedule")
    ws16.sheet_properties.tabColor = TEAL
    _title(ws16, "Monthly GST Compliance Automation Schedule", 1, 1, 2)
    schedule = get_monthly_schedule(2025, 4)
    r = 3
    for k, v in schedule.items():
        _cell(ws16.cell(row=r, column=1, value=k), bold=True, bg=LIGHT_TEAL)
        _cell(ws16.cell(row=r, column=2, value=v))
        r += 1
    ws16.column_dimensions["A"].width = 18
    ws16.column_dimensions["B"].width = 55

    # ════════════════════════════════════════════════════════════
    # SHEET 17: RAW DATA — SALES
    # ════════════════════════════════════════════════════════════
    ws17 = wb.create_sheet("Raw — Sales Register")
    ws17.sheet_properties.tabColor = "757575"
    raw_cols = ["Invoice Number","Invoice Date","Customer GSTIN","Customer Name",
                "Invoice Value","Taxable Value","IGST","CGST","SGST",
                "Place of Supply","HSN Code","Tax Rate","Document Type","Category","GSTIN_Status","Is_Duplicate"]
    _write_df(ws17, gstr1["full_data"][[c for c in raw_cols if c in gstr1["full_data"].columns]], sr=1)

    # ════════════════════════════════════════════════════════════
    # SHEET 18: RAW DATA — PURCHASES
    # ════════════════════════════════════════════════════════════
    ws18 = wb.create_sheet("Raw — Purchase Register")
    ws18.sheet_properties.tabColor = "757575"
    _write_df(ws18, purchase_df, sr=1)

    # Freeze panes on all sheets
    for sheet in wb.worksheets:
        if sheet.max_row > 4:
            sheet.freeze_panes = "A4" if sheet.title != "Dashboard" else "A3"

    return wb
