#!/usr/bin/env python3
"""
Payslip Generator — Fracktal Works Private Limited

Generates statutory-compliant employee payslips from the HR/Payroll Master
Workbook (see generate_hr_master_workbook.py). Two modes:

  --mode sample   Builds ONE demo payslip with clearly-labelled placeholder
                  data, for the user to review and approve the template/layout.
                  No real employee data is read or required.

  --mode full     Reads the real payroll Excel (a "Salary Sheet_<Month> <Year>"
                  sheet + an "Employee_Details" sheet) and generates one
                  payslip per employee row for that month. Optionally embeds
                  a company logo image and an Earned Leave balance table.

Usage:
    # Review the template
    python generate_payslip.py --mode sample

    # Generate real payslips
    python generate_payslip.py --mode full \
        --workbook ".tmp/Payroll26-27_FracktalWorks_JUNE.xlsx" \
        --salary-sheet "Salary Sheet_June 2026" \
        --logo "Logo FW1.png"
"""

import sys
import re
import difflib
import argparse
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

MONTH_NAMES = {
    "JAN": "January", "FEB": "February", "MAR": "March", "APR": "April",
    "MAY": "May", "JUN": "June", "JUL": "July", "AUG": "August",
    "SEP": "September", "OCT": "October", "NOV": "November", "DEC": "December",
}

# ---------------------------------------------------------------- styling ---
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
COMPANY_FONT = Font(name="Calibri", bold=True, size=13, color="1F4E79")
SUB_FONT = Font(name="Calibri", size=10, color="444444")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LABEL_FONT = Font(name="Calibri", bold=True, size=10, color="333333")
VALUE_FONT = Font(name="Calibri", size=10, color="000000")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
NET_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
NOTE_FONT = Font(italic=True, size=8, color="777777")
SAMPLE_FONT = Font(italic=True, bold=True, size=10, color="C00000")
LOGO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _num_to_words_indian(n):
    """Convert an integer rupee amount to words, Indian numbering (Lakh/Crore)."""
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def two_digit(num):
        if num < 20:
            return ones[num]
        return tens[num // 10] + (f" {ones[num % 10]}" if num % 10 else "")

    def three_digit(num):
        if num >= 100:
            return f"{ones[num // 100]} Hundred" + (f" {two_digit(num % 100)}" if num % 100 else "")
        return two_digit(num)

    if n == 0:
        return "Zero"
    parts = []
    crore, n = divmod(n, 10000000)
    lakh, n = divmod(n, 100000)
    thousand, n = divmod(n, 1000)
    hundred = n
    if crore:
        parts.append(f"{three_digit(crore)} Crore")
    if lakh:
        parts.append(f"{three_digit(lakh)} Lakh")
    if thousand:
        parts.append(f"{three_digit(thousand)} Thousand")
    if hundred:
        parts.append(three_digit(hundred))
    return " ".join(parts)


def earnings_deduction_lines(data):
    """Build the (label, value) line items for the Earnings and Deductions
    columns from a payslip data dict. Shared by the Excel renderer
    (build_payslip_sheet) and the webapp's PDF renderer so both always show
    the identical set of line items for the identical numbers."""
    earnings = [
        ("Basic", data["basic"]),
        ("HRA", data["hra"]),
        ("Conveyance / Transport Allowance", data["conveyance"]),
        ("Medical Allowance", data["medical_allowance"]),
        ("Special Allowance", data["special_allowance"]),
        ("Overtime Payment", data["overtime"]),
        ("Incentive", data["incentive"]),
        ("Arrears", data["arrears"]),
    ]
    # Shown only when applicable (interns/contract staff paid a flat fee;
    # attendance bonus only when actually awarded) — kept off the standard
    # payslip otherwise, per the approved template.
    if data["attendance_bonus"]:
        earnings.append(("Attendance Bonus", data["attendance_bonus"]))
    if data["consolidated_pay"]:
        earnings.append(("Consolidated Pay (Interns/Contract Staff)", data["consolidated_pay"]))
    deductions = [
        ("Employee PF (12%)", data["employee_pf"]),
        ("Employee ESI (0.75%)", data["employee_esi"]),
        ("Professional Tax", data["pt"]),
        ("TDS (Income Tax)", data["tds"]),
        ("Salary Advance Recovery", data["salary_advance"]),
        ("Employer PF (as per CTC structure)", data["employer_pf"]),
        ("Employer ESI (as per CTC structure)", data["employer_esi"]),
    ]
    return earnings, deductions


def _set(ws, cell, value, font=None, align=None, fill=None, border=None, number_format=None):
    c = ws[cell]
    c.value = value
    if font:
        c.font = font
    if align:
        c.alignment = align
    if fill:
        c.fill = fill
    if border:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c


def build_payslip_sheet(wb, sheet_name, data, logo_path=None, is_sample=False):
    """
    Build one payslip on a fresh sheet.
    `data` is a flat dict of every field the template needs (see SAMPLE_DATA
    below for the full field list, or `_row_to_payslip_data` for real data).
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    col_widths = {"A": 16, "B": 20, "C": 16, "D": 20, "E": 8, "F": 18}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    if is_sample:
        ws.merge_cells("A1:F1")
        _set(ws, "A1", "SAMPLE DATA — FOR TEMPLATE REVIEW ONLY (no real employee data)",
             font=SAMPLE_FONT, align=CENTER)
        top = 3
    else:
        top = 1

    # --- Header: logo + company ------------------------------------------
    logo_row_span = 4
    ws.merge_cells(f"A{top}:B{top + logo_row_span - 1}")
    logo_cell = ws[f"A{top}"]
    logo_cell.fill = LOGO_FILL
    logo_cell.border = BORDER
    logo_cell.alignment = CENTER
    for r in range(top, top + logo_row_span):
        ws.row_dimensions[r].height = 18

    if logo_path and Path(logo_path).exists():
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            img.height = 70
            img.width = 140
            img.anchor = f"A{top}"
            ws.add_image(img)
        except Exception as e:
            logo_cell.value = f"[LOGO — could not embed: {e}]"
            logo_cell.font = NOTE_FONT
    else:
        logo_cell.value = "[ COMPANY LOGO ]\n(upload logo to embed here)"
        logo_cell.font = NOTE_FONT

    ws.merge_cells(f"C{top}:F{top}")
    _set(ws, f"C{top}", data["company_name"], font=COMPANY_FONT, align=CENTER)
    ws.merge_cells(f"C{top+1}:F{top+1}")
    _set(ws, f"C{top+1}", data["company_address"], font=SUB_FONT, align=CENTER)
    ws.merge_cells(f"C{top+2}:F{top+2}")
    _set(ws, f"C{top+2}", f"CIN: {data['cin']}", font=SUB_FONT, align=CENTER)
    ws.merge_cells(f"C{top+3}:F{top+3}")
    _set(ws, f"C{top+3}", f"PAYSLIP FOR {data['month_label'].upper()}", font=TITLE_FONT, align=CENTER)

    r = top + logo_row_span + 1

    # --- Employee details ---------------------------------------------------
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws, f"A{r}", "EMPLOYEE DETAILS", font=SECTION_FONT, fill=SECTION_FILL, align=CENTER)
    r += 1

    emp_fields = [
        ("Employee Name", data["name"], "Employee ID", data["employee_id"]),
        ("Designation", data["designation"], "Department", data["department"]),
        ("Date of Joining", data["doj"], "Pay Mode", "Bank Transfer"),
        ("PAN", data["pan"], "UAN", data["uan"]),
        ("ESI No.", data["esi_no"], "Bank Name", data["bank_name"]),
        ("Bank A/c No.", data["bank_account"], "", ""),
    ]
    for label1, val1, label2, val2 in emp_fields:
        _set(ws, f"A{r}", label1, font=LABEL_FONT, border=BORDER)
        ws.merge_cells(f"B{r}:B{r}")
        _set(ws, f"B{r}", val1, font=VALUE_FONT, border=BORDER, align=LEFT)
        _set(ws, f"C{r}", label2, font=LABEL_FONT, border=BORDER)
        ws.merge_cells(f"D{r}:F{r}")
        _set(ws, f"D{r}", val2, font=VALUE_FONT, border=BORDER, align=LEFT)
        r += 1

    r += 1
    # --- Attendance summary ---------------------------------------------------
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws, f"A{r}", "ATTENDANCE SUMMARY", font=SECTION_FONT, fill=SECTION_FILL, align=CENTER)
    r += 1
    att_fields = [
        ("Days in Month", data["calendar_days"], "Present Days", data["present_days"]),
        ("LOP Days", data["lop_days"], "Paid Days", data["paid_days"]),
        ("Earned Leave Balance", data["earned_leave_balance"], "", ""),
    ]
    for label1, val1, label2, val2 in att_fields:
        _set(ws, f"A{r}", label1, font=LABEL_FONT, border=BORDER)
        _set(ws, f"B{r}", val1, font=VALUE_FONT, border=BORDER, align=CENTER)
        _set(ws, f"C{r}", label2, font=LABEL_FONT, border=BORDER)
        ws.merge_cells(f"D{r}:F{r}")
        _set(ws, f"D{r}", val2, font=VALUE_FONT, border=BORDER, align=LEFT)
        r += 1

    r += 1
    # --- Earnings / Deductions side-by-side ---------------------------------
    earn_start = r
    ws.merge_cells(f"A{r}:B{r}")
    _set(ws, f"A{r}", "EARNINGS", font=SECTION_FONT, fill=SECTION_FILL, align=CENTER)
    ws.merge_cells(f"C{r}:D{r}")
    _set(ws, f"C{r}", "DEDUCTIONS", font=SECTION_FONT, fill=SECTION_FILL, align=CENTER)
    r += 1

    earnings, deductions = earnings_deduction_lines(data)

    row_count = max(len(earnings), len(deductions))
    for i in range(row_count):
        row = earn_start + 1 + i
        if i < len(earnings):
            _set(ws, f"A{row}", earnings[i][0], font=VALUE_FONT, border=BORDER, align=LEFT)
            _set(ws, f"B{row}", earnings[i][1], font=VALUE_FONT, border=BORDER, align=RIGHT, number_format="#,##0")
        else:
            ws[f"A{row}"].border = BORDER
            ws[f"B{row}"].border = BORDER
        if i < len(deductions):
            _set(ws, f"C{row}", deductions[i][0], font=VALUE_FONT, border=BORDER, align=LEFT)
            _set(ws, f"D{row}", deductions[i][1], font=VALUE_FONT, border=BORDER, align=RIGHT, number_format="#,##0")
        else:
            ws[f"C{row}"].border = BORDER
            ws[f"D{row}"].border = BORDER

    total_row = earn_start + 1 + row_count
    _set(ws, f"A{total_row}", "GROSS EARNINGS", font=TOTAL_FONT, border=BORDER, align=LEFT)
    _set(ws, f"B{total_row}", data["gross_earnings"], font=TOTAL_FONT, border=BORDER, align=RIGHT, number_format="#,##0")
    _set(ws, f"C{total_row}", "TOTAL DEDUCTIONS", font=TOTAL_FONT, border=BORDER, align=LEFT)
    _set(ws, f"D{total_row}", data["total_deductions"], font=TOTAL_FONT, border=BORDER, align=RIGHT, number_format="#,##0")

    r = total_row + 2
    # --- Net pay --------------------------------------------------------------
    ws.merge_cells(f"A{r}:C{r}")
    _set(ws, f"A{r}", "NET SALARY PAYABLE", font=TOTAL_FONT, fill=NET_FILL, border=BORDER, align=LEFT)
    ws.merge_cells(f"D{r}:F{r}")
    _set(ws, f"D{r}", data["net_pay"], font=TOTAL_FONT, fill=NET_FILL, border=BORDER, align=RIGHT, number_format="#,##0")
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws, f"A{r}", f"Net Salary in Words: Rupees {data['net_pay_words']} Only", font=SUB_FONT, align=LEFT)
    r += 2

    # --- Footer / compliance -------------------------------------------------
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws, f"A{r}", (f"EPFO Establishment Code: {data['pf_establishment_code']}   |   "
                        f"ESIC Employer Code: {data['esi_employer_code']}   |   "
                        f"PT Registration No.: {data['pt_reg_no']}"), font=NOTE_FONT, align=LEFT)
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws, f"A{r}", f"TAN: {data['tan']}", font=NOTE_FONT, align=LEFT)
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws, f"A{r}", "This is a system-generated payslip and does not require a physical signature.",
         font=NOTE_FONT, align=LEFT)

    return ws


SAMPLE_DATA = {
    "company_name": "Fracktal Works Private Limited",
    "company_address": "<Registered Office Address — to be filled>",
    "cin": "<TO BE FILLED>",
    "month_label": "April 2026",
    "name": "Sample Employee (placeholder)",
    "employee_id": "SAMPLE-001",
    "designation": "Sample Designation",
    "department": "Sample Department",
    "doj": "01-Apr-2020",
    "location": "Bengaluru",
    "pan": "ABCDE1234F",
    "uan": "100200300400",
    "esi_no": "1234567890",
    "bank_name": "Sample Bank",
    "bank_account": "XXXXXXXX1234",
    "email": "sample.employee@example.com",
    "calendar_days": 30,
    "present_days": 30,
    "lop_days": 0,
    "paid_days": 30,
    "earned_leave_balance": 12,
    "basic": 25000,
    "hra": 10000,
    "conveyance": 1600,
    "medical_allowance": 1250,
    "special_allowance": 5550,
    "overtime": 1500,
    "incentive": 0,
    "arrears": 0,
    "attendance_bonus": 0,
    "consolidated_pay": 0,
    "employee_pf": 3000,
    "employee_esi": 0,
    "pt": 200,
    "tds": 0,
    "salary_advance": 0,
    "employer_pf": 3250,
    "employer_esi": 0,
    "gross_earnings": 44900,
    "total_deductions": 6450,
    "net_pay": 38450,
    "net_pay_words": _num_to_words_indian(38450),
    "tan": "<TO BE FILLED>",
    "pf_establishment_code": "<TO BE FILLED>",
    "esi_employer_code": "<TO BE FILLED>",
    "pt_reg_no": "<TO BE FILLED>",
}


def generate_sample(company_name, output_path=None):
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(TMP_DIR / f"Payslip_TEMPLATE_SAMPLE_{timestamp}.xlsx")

    data = dict(SAMPLE_DATA)
    data["company_name"] = company_name

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_payslip_sheet(wb, "Payslip Template", data, logo_path=None, is_sample=True)
    wb.save(output_path)

    print(f"\n{'='*60}")
    print("PAYSLIP TEMPLATE (SAMPLE) GENERATED")
    print(f"  {output_path}")
    print("  All data on this sheet is a PLACEHOLDER for layout review only.")
    print(f"{'='*60}\n")
    return output_path


def _norm_digits(v):
    if v is None:
        return ""
    return "".join(ch for ch in str(v).strip() if ch.isdigit())


def _norm_pan(v):
    return str(v).strip().upper() if v else ""


def _norm_name(v):
    return re.sub(r"[^a-z]", "", str(v).lower())


def _build_employee_details_index(ed_ws):
    """Index Employee_Details by UAN / Bank Account / PAN (normalized) ->
    {employee_id, bank_name, designation, pan, account_no, office_email,
    personal_email} — pan/account_no are a fallback when the Salary Sheet's
    own PAN/Bank Account is blank; the emails have no Salary Sheet equivalent
    at all, so they always come from here."""
    index = {}
    for r in range(3, ed_ws.max_row + 1):
        name = ed_ws.cell(row=r, column=3).value  # C: Name
        if not name:
            continue
        if str(name).strip() == "Employee Name":
            break  # reached the appended Earned Leave balance sub-table — stop
        raw_pan = ed_ws.cell(row=r, column=18).value       # R: PAN Card No
        raw_account = ed_ws.cell(row=r, column=20).value   # T: Account No
        info = {
            "employee_id": ed_ws.cell(row=r, column=2).value,    # B: Employee ID
            "bank_name": ed_ws.cell(row=r, column=19).value,     # S: Bank Name
            "designation": ed_ws.cell(row=r, column=10).value,   # J: Designation
            "pan": raw_pan,
            "account_no": raw_account,
            "office_email": ed_ws.cell(row=r, column=12).value,    # L: Office Email ID
            "personal_email": ed_ws.cell(row=r, column=13).value,  # M: Personal Email ID
        }
        uan = _norm_digits(ed_ws.cell(row=r, column=22).value)      # V: UAN
        account_no = _norm_digits(raw_account)
        pan = _norm_pan(raw_pan)
        if uan:
            index.setdefault(("uan", uan), info)
        if account_no:
            index.setdefault(("bank", account_no), info)
        if pan:
            index.setdefault(("pan", pan), info)
    return index


def _build_leave_balance_index(ed_ws):
    """Read the Earned Leave closing-balance sub-table appended below the main
    Employee_Details list (marker row: col C = 'Employee Name', col D = the
    'EL closing balance as on ...' label). Only employees with a recorded
    balance appear here. Returns normalized-name -> (raw_name, balance)."""
    index = {}
    marker_row = None
    for r in range(1, ed_ws.max_row + 1):
        if str(ed_ws.cell(row=r, column=3).value or "").strip() == "Employee Name":
            marker_row = r
            break
    if marker_row is None:
        return index
    for r in range(marker_row + 2, ed_ws.max_row + 1):
        name = ed_ws.cell(row=r, column=3).value
        balance = ed_ws.cell(row=r, column=4).value
        if name is None:
            continue
        index[_norm_name(name)] = (name, balance)
    return index


def _lookup_leave_balance(leave_index, employee_name, min_ratio=0.75):
    if not leave_index:
        return "N.A"
    target = _norm_name(employee_name)
    best_key, best_ratio = None, 0.0
    for key in leave_index:
        ratio = difflib.SequenceMatcher(None, target, key).ratio()
        if ratio > best_ratio:
            best_key, best_ratio = key, ratio
    if best_key and (best_ratio >= min_ratio or target.startswith(best_key[:6]) or best_key.startswith(target[:6])):
        raw_name, balance = leave_index[best_key]
        return f"{balance} days" if balance is not None else "0 days"
    return "N.A"


def _row_to_payslip_data(company, month_label, salary_ws, ed_index, leave_index, r, warnings=None):
    def val(col):
        return salary_ws[f"{col}{r}"].value

    def num(col):
        v = val(col)
        return v if isinstance(v, (int, float)) else 0

    uan = _norm_digits(val("E"))
    pan = _norm_pan(val("F"))
    bank_account = _norm_digits(val("H"))
    match = (ed_index.get(("uan", uan)) or ed_index.get(("bank", bank_account))
             or ed_index.get(("pan", pan)))
    employee_id = (match["employee_id"] if match else None) or "<NOT FOUND — check Employee_Details>"
    bank_name = (match["bank_name"] if match else None) or "<NOT FOUND — check Employee_Details>"
    designation = (match["designation"] if match else None) or val("D") or ""
    # PAN / Bank Account: prefer the Salary Sheet's own value, fall back to
    # Employee_Details when the Salary Sheet leaves it blank.
    pan_value = val("F") or (match["pan"] if match else None) or ""
    bank_account_value = val("H") or (match["account_no"] if match else None) or ""
    # Email has no Salary Sheet equivalent — Personal Email preferred (payslips
    # are confidential; a personal inbox is safer than a shared/IT-visible
    # office mailbox), falling back to Office Email only if no personal one is on file.
    email_value = (match.get("personal_email") or match.get("office_email") or "") if match else ""

    # Interns/contract staff/professional-service consultants have no Basic/HRA/etc
    # breakdown — they're paid a flat consolidated amount (Eff. Gross) instead.
    has_breakdown = any(val(c) is not None for c in ("U", "V", "W", "X", "Y"))
    if has_breakdown:
        basic, hra, conv, medical, special = num("U"), num("V"), num("W"), num("X"), num("Y")
        consolidated_pay = 0
    else:
        basic = hra = conv = medical = special = 0
        consolidated_pay = num("T") or num("AB")

    incentive, overtime = num("Z"), num("AA")
    employee_pf, employee_esi = num("AC"), num("AD")
    employer_pf, employer_esi = num("AE"), num("AF")
    pt, tds, salary_advance = num("AG"), num("AH"), num("AI")
    arrears, attendance_bonus = num("AK"), num("AL")
    net_sal_source = val("AM")

    gross_earnings = (basic + hra + conv + medical + special + consolidated_pay
                       + overtime + incentive + arrears + attendance_bonus)
    total_deductions = employee_pf + employee_esi + pt + tds + salary_advance + employer_pf + employer_esi
    net_pay = gross_earnings - total_deductions

    if isinstance(net_sal_source, (int, float)) and round(net_pay) != round(net_sal_source):
        msg = (f"Row {r} ({val('C')}): computed Net Pay {net_pay} differs from "
               f"source 'Net Sal.' {net_sal_source} — review this row in the source sheet.")
        if warnings is not None:
            warnings.append(msg)
        else:
            print(f"  WARNING: {msg}")

    doj = val("I")
    doj_str = doj.strftime("%d-%b-%Y") if isinstance(doj, datetime) else (doj or "")

    return {
        "company_name": company["name"],
        "company_address": company["address"],
        "cin": company["cin"],
        "month_label": month_label,
        "name": val("C") or "",
        "employee_id": employee_id,
        "designation": designation,
        "department": val("B") or "",
        "doj": doj_str,
        "pan": pan_value,
        "uan": val("E") or "",
        "esi_no": val("G") or "-",
        "bank_name": bank_name,
        "bank_account": bank_account_value,
        "email": email_value,
        "calendar_days": val("J") or "",
        "present_days": val("K") or "",
        "lop_days": val("Q") or 0,
        "paid_days": val("P") or "",
        "earned_leave_balance": _lookup_leave_balance(leave_index, val("C") or ""),
        "basic": basic, "hra": hra, "conveyance": conv, "medical_allowance": medical,
        "special_allowance": special, "consolidated_pay": consolidated_pay,
        "overtime": overtime, "incentive": incentive,
        "arrears": arrears, "attendance_bonus": attendance_bonus,
        "employee_pf": employee_pf, "employee_esi": employee_esi, "pt": pt, "tds": tds,
        "salary_advance": salary_advance, "employer_pf": employer_pf, "employer_esi": employer_esi,
        "gross_earnings": gross_earnings, "total_deductions": total_deductions,
        "net_pay": net_pay, "net_pay_words": _num_to_words_indian(max(int(round(net_pay)), 0)),
        "tan": company["tan"],
        "pf_establishment_code": company["pf_code"],
        "esi_employer_code": company["esi_code"],
        "pt_reg_no": company["pt_reg"],
    }


def load_payslip_rows(workbook_path, salary_sheet_name, employee_sheet_name, company):
    """Load and compute per-employee payslip data from the real payroll Excel.

    `company` is a dict with address/cin/tan/pf_code/esi_code/pt_reg (name is
    read from the sheet itself and filled in here). Returns
    (month_label, rows, unmatched_names, warnings) where `rows` is a list of
    (unique_id, data_dict) — unique_id is a filesystem/sheet-name-safe string
    ("<EmployeeID>-<Name>") suitable as an Excel sheet name or PDF filename.

    Shared by the CLI (generate_from_excel, below) and webapp/payslip_data.py
    so both consume the exact same extraction/reconciliation logic.
    """
    wb_in = openpyxl.load_workbook(workbook_path, data_only=True)
    if salary_sheet_name not in wb_in.sheetnames:
        raise ValueError(f"sheet '{salary_sheet_name}' not found. Available sheets: {wb_in.sheetnames}")
    if employee_sheet_name not in wb_in.sheetnames:
        raise ValueError(f"sheet '{employee_sheet_name}' not found. Available sheets: {wb_in.sheetnames}")

    salary_ws = wb_in[salary_sheet_name]
    ed_ws = wb_in[employee_sheet_name]
    ed_index = _build_employee_details_index(ed_ws)
    leave_index = _build_leave_balance_index(ed_ws)

    company = dict(company)
    company["name"] = salary_ws["A1"].value or company.get("name") or "Company Name"
    month_line = str(salary_ws["A2"].value or "")
    month_label = (month_line.upper().split("MONTH OF")[-1].strip().title()
                   if "MONTH OF" in month_line.upper() else salary_sheet_name)

    rows = []
    unmatched = []
    warnings = []
    used_ids = set()
    for r in range(5, salary_ws.max_row + 1):
        sl_no = salary_ws.cell(row=r, column=1).value
        name = salary_ws.cell(row=r, column=3).value
        if not (isinstance(sl_no, (int, float)) and name):
            continue  # skips blank rows, section headers (e.g. "PROFESSIONAL SERVICE"), TOTAL/GRAND TOTAL

        data = _row_to_payslip_data(company, month_label, salary_ws, ed_index, leave_index, r, warnings=warnings)
        if "NOT FOUND" in str(data["employee_id"]):
            unmatched.append(str(name))

        raw_id = str(name) if "NOT FOUND" in str(data["employee_id"]) else f"{data['employee_id']}-{name}"
        safe_id = raw_id[:31]
        for bad_char in '/\\[]*?:':
            safe_id = safe_id.replace(bad_char, "-")
        final_id, n = safe_id, 2
        while final_id in used_ids:
            final_id = f"{safe_id[:28]}_{n}"
            n += 1
        used_ids.add(final_id)
        rows.append((final_id, data))

    return month_label, rows, unmatched, warnings


def generate_from_excel(workbook_path, salary_sheet_name, employee_sheet_name, company_address, cin,
                         tan, pf_code, esi_code, pt_reg, logo_path=None, output_path=None):
    company = {
        "address": company_address, "cin": cin, "tan": tan,
        "pf_code": pf_code, "esi_code": esi_code, "pt_reg": pt_reg,
    }
    try:
        month_label, rows, unmatched, warnings = load_payslip_rows(
            workbook_path, salary_sheet_name, employee_sheet_name, company)
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    for msg in warnings:
        print(f"  WARNING: {msg}")

    TMP_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_month = salary_sheet_name.replace(" ", "_").replace("/", "-")
        output_path = str(TMP_DIR / f"Payslips_{safe_month}_{timestamp}.xlsx")

    if not rows:
        print(f"ERROR: No employee rows found in '{salary_sheet_name}'.")
        sys.exit(1)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    for sheet_name, data in rows:
        build_payslip_sheet(wb_out, sheet_name, data, logo_path=logo_path, is_sample=False)
    wb_out.save(output_path)

    print(f"\n{'='*60}")
    print(f"GENERATED {len(rows)} PAYSLIP(S) from '{salary_sheet_name}'")
    print(f"  {output_path}")
    if unmatched:
        print(f"  NOTE: {len(unmatched)} employee(s) had no match in '{employee_sheet_name}' "
              f"(Employee ID / Bank Name flagged for manual fill-in): {', '.join(unmatched)}")
    print(f"{'='*60}\n")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate employee payslips")
    parser.add_argument("--mode", choices=["sample", "full"], default="sample")
    parser.add_argument("--company", default="Fracktal Works Private Limited")
    parser.add_argument("--company-address",
                         default="No. 3, 50ft Laggere Main Road, Chowdeshwari Nagar, "
                                 "Bengaluru 560058, Karnataka, India")
    parser.add_argument("--cin", default="U30009KA2013PTC070124")
    parser.add_argument("--tan", default="BLRF03155F")
    parser.add_argument("--pf-code", default="PYKRP1426103000", help="EPFO Establishment Code")
    parser.add_argument("--esi-code", default="49000552030001099", help="ESIC Employer Code")
    parser.add_argument("--pt-reg", default="356662288", help="Professional Tax Registration No.")
    parser.add_argument("--workbook", default=None, help="Path to the real payroll Excel (--mode full)")
    parser.add_argument("--salary-sheet", default=None, help="e.g. 'Salary Sheet_June 2026' (--mode full)")
    parser.add_argument("--employee-sheet", default="Employee_Details",
                         help="Sheet with Employee ID / Bank Name (--mode full)")
    parser.add_argument("--logo", default=None, help="Path to company logo image (png/jpg)")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    if args.mode == "sample":
        generate_sample(args.company, args.output)
    else:
        if not args.workbook or not args.salary_sheet:
            print("ERROR: --workbook and --salary-sheet are required for --mode full")
            sys.exit(1)
        generate_from_excel(
            args.workbook, args.salary_sheet, args.employee_sheet,
            args.company_address, args.cin, args.tan, args.pf_code, args.esi_code, args.pt_reg,
            logo_path=args.logo, output_path=args.output,
        )


if __name__ == "__main__":
    main()
