#!/usr/bin/env python3
"""
HR Workflow Chart Generator — Fracktal Works Private Limited

Generates comprehensive workflow charts for Human Resources processes
as an Excel workbook with swim-lane flowcharts, RACI matrices, leave
policy matrix, and KPI dashboards.

Usage:
    python hr_workflow_generator.py --company "Fracktal Works Private Limited"
    python hr_workflow_generator.py --workflows "Recruitment,Onboarding,Exit"
    python hr_workflow_generator.py --workflows "Attendance" --leave-policy '{"casual_leave":{"days":10,"approver":"Reporting Manager"}}'
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

# ─── Styles ───────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DECISION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CONTROL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ESCALATION_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
INDEX_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
RACI_R_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
RACI_A_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
RACI_C_FILL = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
RACI_I_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ─── Default Leave Policy ─────────────────────────────────────────────────────

DEFAULT_LEAVE_POLICY = {
    "casual_leave":    {"days": 12,  "approver": "Reporting Manager",          "description": "Short-notice personal leave"},
    "sick_leave":      {"days": 12,  "approver": "Reporting Manager",          "description": "Medical / illness leave"},
    "earned_leave":    {"days": 15,  "approver": "Reporting Manager + HR",     "description": "Planned vacation, accrual-based"},
    "maternity_leave": {"days": 182, "approver": "HR Head",                   "description": "Per Maternity Benefit Act, 1961"},
    "paternity_leave": {"days": 7,   "approver": "Reporting Manager",          "description": "Company policy benefit"},
    "loss_of_pay":     {"days": None, "approver": "HR Head",                  "description": "Unpaid leave beyond entitlement"},
}

# ─── Workflow Definitions ─────────────────────────────────────────────────────

def get_all_workflows():
    """Return all HR workflow definitions."""
    return {
        "Recruitment": recruitment_workflow(),
        "Onboarding": onboarding_workflow(),
        "Attendance": attendance_leave_workflow(),
        "Payroll Input": payroll_input_workflow(),
        "Statutory Compliance": statutory_compliance_workflow(),
        "Performance Management": performance_workflow(),
        "Exit": exit_workflow(),
        "Grievance & POSH": grievance_posh_workflow(),
        "Training & Development": training_workflow(),
        "HR Policy": hr_policy_workflow(),
    }


def recruitment_workflow():
    """Recruitment & Selection workflow steps."""
    return {
        "name": "Recruitment & Selection",
        "steps": [
            {"step": 1, "phase": "Requisition", "role": "Department Head", "action": "Raise manpower requisition with role, budget, headcount justification", "decision": False, "doc": "Manpower Requisition Form", "control": "Budgeted headcount check", "sla": "Day 0", "escalation": "HR Manager"},
            {"step": 2, "phase": "Requisition", "role": "HR Manager", "action": "Approve requisition and open the position", "decision": True, "yes": "Proceed to sourcing", "no": "Return with comments", "doc": "Approved Requisition", "control": "Budget vs headcount plan", "sla": "Day 1-2", "escalation": "HR Head"},
            {"step": 3, "phase": "Sourcing", "role": "HR Executive", "action": "Source candidates via job portals, referrals, consultants", "decision": False, "doc": "Job Posting, Applicant Tracker", "control": "JD alignment check", "sla": "Ongoing", "escalation": "HR Manager"},
            {"step": 4, "phase": "Sourcing", "role": "HR Executive", "action": "Screen resumes, shortlist candidates against JD", "decision": True, "yes": "Schedule interviews", "no": "Reject with feedback", "doc": "Applicant Tracker", "control": "Minimum qualification screen", "sla": "T+3 days", "escalation": "HR Manager"},
            {"step": 5, "phase": "Interview", "role": "Department Head", "action": "Conduct technical/functional interview rounds", "decision": True, "yes": "Recommend for HR round", "no": "Reject with feedback", "doc": "Interview Evaluation Form", "control": "Standard scorecard used", "sla": "T+7 days", "escalation": "HR Manager"},
            {"step": 6, "phase": "Interview", "role": "HR Manager", "action": "Conduct HR round — culture fit, compensation discussion", "decision": True, "yes": "Proceed to offer", "no": "Reject / hold for future role", "doc": "HR Evaluation Form", "control": "Compensation band adherence", "sla": "T+8 days", "escalation": "HR Head"},
            {"step": 7, "phase": "Offer", "role": "HR Head", "action": "Approve compensation and release offer letter", "decision": True, "yes": "Offer released", "no": "Revise and re-approve", "doc": "Offer Letter", "control": "CTC band approval", "sla": "T+10 days", "escalation": "CEO"},
            {"step": 8, "phase": "Verification", "role": "HR Executive", "action": "Initiate background verification (education, employment, criminal)", "decision": True, "yes": "Cleared — confirm joining", "no": "Escalate discrepancy to HR Head", "doc": "BGV Report", "control": "Third-party BGV vendor check", "sla": "T+20 days", "escalation": "HR Head"},
            {"step": 9, "phase": "Confirmation", "role": "HR Executive", "action": "Confirm joining date, send pre-onboarding documents", "decision": False, "doc": "Joining Confirmation Email", "control": "Document checklist sent", "sla": "T-3 days to joining", "escalation": "HR Manager"},
            {"step": 10, "phase": "Completion", "role": "HR Executive", "action": "Close requisition, update recruitment tracker and metrics", "decision": False, "doc": "Recruitment Tracker", "control": "Time-to-fill logged", "sla": "Day of joining", "escalation": "HR Manager"},
        ],
        "raci_roles": ["Candidate", "Reporting Mgr", "HR Exec", "HR Manager", "HR Head", "Finance/CFO", "Auditor"],
        "raci": [
            ["Raise Requisition",       "", "R", "", "A", "", "", ""],
            ["Approve Requisition",     "", "I", "", "R", "A", "", ""],
            ["Source Candidates",       "I", "", "R", "A", "", "", ""],
            ["Screen & Shortlist",      "I", "C", "R", "A", "", "", ""],
            ["Technical Interview",     "I", "R", "", "A", "", "", ""],
            ["HR Interview",            "I", "", "C", "R", "A", "", ""],
            ["Approve Offer",           "I", "", "", "C", "R", "A", ""],
            ["Background Verification","I", "", "R", "A", "", "", ""],
            ["Confirm Joining",         "I", "I", "R", "A", "", "", ""],
            ["Close Requisition",       "", "I", "R", "A", "", "", "I"],
        ],
    }


def onboarding_workflow():
    """Onboarding & Induction workflow steps."""
    return {
        "name": "Onboarding & Induction",
        "steps": [
            {"step": 1, "phase": "Pre-boarding", "role": "HR Executive", "action": "Collect joining documents (ID proof, education certificates, prior UAN/bank details)", "decision": True, "yes": "Documents complete — proceed", "no": "Follow up before joining date", "doc": "Document Checklist", "control": "Mandatory document list per policy", "sla": "T-3 days", "escalation": "HR Manager"},
            {"step": 2, "phase": "Pre-boarding", "role": "HR Executive", "action": "Create employee master record, assign Employee ID", "decision": False, "doc": "Master_Employees Sheet", "control": "Unique Employee ID generation", "sla": "T-1 day", "escalation": "HR Manager"},
            {"step": 3, "phase": "Day 1", "role": "HR Executive", "action": "Complete induction — company policies, code of conduct, POSH policy", "decision": False, "doc": "Induction Kit, Policy Acknowledgement", "control": "Signed acknowledgement on file", "sla": "Day 1", "escalation": "HR Manager"},
            {"step": 4, "phase": "Day 1", "role": "HR Executive", "action": "Register employee for PF (UAN generation), ESI, and PT as applicable", "decision": True, "yes": "Registered — proceed", "no": "Not applicable — note in master", "doc": "PF/ESI Portal, Master_Employees", "control": "Wage-ceiling applicability check", "sla": "Day 1-2", "escalation": "HR Manager"},
            {"step": 5, "phase": "Day 1", "role": "IT/Admin", "action": "Provision system access, email ID, workstation", "decision": False, "doc": "IT Provisioning Ticket", "control": "Access limited to role requirements", "sla": "Day 1", "escalation": "HR Manager"},
            {"step": 6, "phase": "Week 1", "role": "Reporting Manager", "action": "Assign buddy, set initial 30-60-90 day goals", "decision": False, "doc": "Goal Sheet", "control": "Goals aligned to role JD", "sla": "Week 1", "escalation": "HR Manager"},
            {"step": 7, "phase": "Probation", "role": "HR Executive", "action": "Track probation period, schedule check-ins at 30/60/90 days", "decision": False, "doc": "Probation Tracker", "control": "Check-in reminders automated", "sla": "Throughout probation", "escalation": "HR Manager"},
            {"step": 8, "phase": "Confirmation", "role": "Reporting Manager", "action": "Submit probation confirmation recommendation", "decision": True, "yes": "Recommend confirmation", "no": "Extend probation / performance improvement plan", "doc": "Confirmation Recommendation Form", "control": "Performance evidence required", "sla": "End of probation", "escalation": "HR Manager"},
            {"step": 9, "phase": "Confirmation", "role": "HR Manager", "action": "Issue confirmation letter or extension letter", "decision": True, "yes": "Confirmed — update master", "no": "Extended — set new review date", "doc": "Confirmation Letter", "control": "Letter signed and filed", "sla": "T+5 days after review", "escalation": "HR Head"},
            {"step": 10, "phase": "Completion", "role": "HR Executive", "action": "Update employee master status, archive onboarding documents", "decision": False, "doc": "Master_Employees, Document Archive", "control": "Complete personnel file check", "sla": "Post-confirmation", "escalation": "HR Manager"},
        ],
        "raci_roles": ["New Employee", "Reporting Mgr", "HR Exec", "HR Manager", "IT/Admin", "HR Head", "Auditor"],
        "raci": [
            ["Collect Documents",       "R", "", "R", "A", "", "", ""],
            ["Create Master Record",    "", "", "R", "A", "", "", ""],
            ["Induction & Policy Ack",  "R", "", "R", "A", "", "", "I"],
            ["PF/ESI/PT Registration",  "C", "", "R", "A", "", "", "I"],
            ["System Provisioning",     "C", "", "I", "A", "R", "", ""],
            ["Set 30-60-90 Goals",      "C", "R", "", "A", "", "", ""],
            ["Probation Tracking",      "I", "C", "R", "A", "", "", ""],
            ["Confirmation Recommend",  "I", "R", "", "A", "", "", ""],
            ["Issue Confirmation",      "I", "C", "", "R", "", "A", ""],
            ["Update & Archive",        "", "", "R", "A", "", "", "I"],
        ],
    }


def attendance_leave_workflow():
    """Attendance & Leave Management workflow steps."""
    return {
        "name": "Attendance & Leave Management",
        "steps": [
            {"step": 1, "phase": "Capture", "role": "Employee", "action": "Mark daily attendance via biometric/portal", "decision": False, "doc": "Attendance System", "control": "Daily capture reconciliation", "sla": "Daily", "escalation": "HR Executive"},
            {"step": 2, "phase": "Capture", "role": "HR Executive", "action": "Monitor exceptions — missed punches, late marks", "decision": True, "yes": "Regularization requested — go to Step 3", "no": "No action needed", "doc": "Attendance Exception Report", "control": "Exception threshold review", "sla": "Daily/Weekly", "escalation": "HR Manager"},
            {"step": 3, "phase": "Regularization", "role": "Reporting Manager", "action": "Approve or reject attendance regularization request", "decision": True, "yes": "Approved — update attendance", "no": "Rejected — marked as LOP/absent", "doc": "Regularization Request Form", "control": "Justification required", "sla": "T+2 days", "escalation": "HR Manager"},
            {"step": 4, "phase": "Leave Application", "role": "Employee", "action": "Apply for leave (casual/sick/earned) via portal", "decision": False, "doc": "Leave Application", "control": "Leave balance auto-check", "sla": "Per policy notice period", "escalation": "Reporting Manager"},
            {"step": 5, "phase": "Leave Approval", "role": "Reporting Manager", "action": "Review and approve/reject leave request", "decision": True, "yes": "Approved — update leave tracker", "no": "Rejected — notify employee", "doc": "Leave Tracker", "control": "Team coverage check", "sla": "T+1-2 days", "escalation": "HR Manager"},
            {"step": 6, "phase": "Leave Approval", "role": "HR Executive", "action": "Verify leave balance and statutory leave types (maternity/paternity)", "decision": True, "yes": "Balance sufficient — proceed", "no": "Convert excess to LOP", "doc": "Leave Tracker, Master_Employees", "control": "Statutory entitlement compliance", "sla": "T+2 days", "escalation": "HR Manager"},
            {"step": 7, "phase": "Monitoring", "role": "HR Executive", "action": "Track leave-without-pay (LOP) and absenteeism trends", "decision": True, "yes": "Excessive absence — flag for review", "no": "Within norms — continue monitoring", "doc": "Absenteeism Report", "control": "Monthly absenteeism threshold (>10%)", "sla": "Monthly", "escalation": "HR Manager"},
            {"step": 8, "phase": "Compilation", "role": "HR Executive", "action": "Compile monthly attendance, leave, and LOP summary per employee", "decision": False, "doc": "Monthly Attendance Summary", "control": "Cross-check with biometric export", "sla": "1st of following month", "escalation": "HR Manager"},
            {"step": 9, "phase": "Review", "role": "HR Manager", "action": "Review summary for anomalies before payroll handover", "decision": True, "yes": "Approved — hand off to payroll", "no": "Correct discrepancies", "doc": "Reviewed Attendance Summary", "control": "Variance vs prior month check", "sla": "2nd of month", "escalation": "HR Head"},
            {"step": 10, "phase": "Handover", "role": "HR Manager", "action": "Hand off finalized attendance/LOP data to Finance for payroll processing", "decision": False, "doc": "Payroll Input Sheet", "control": "Sign-off before handover", "sla": "3rd of month", "escalation": "HR Head"},
        ],
        "raci_roles": ["Employee", "Reporting Mgr", "HR Exec", "HR Manager", "HR Head", "Finance", "Auditor"],
        "raci": [
            ["Mark Attendance",         "R", "", "A", "", "", "", ""],
            ["Monitor Exceptions",      "I", "", "R", "A", "", "", ""],
            ["Regularization Approval", "I", "R", "", "A", "", "", ""],
            ["Apply for Leave",         "R", "I", "", "A", "", "", ""],
            ["Leave Approval",          "I", "R", "", "A", "", "", ""],
            ["Verify Leave Balance",    "", "", "R", "A", "", "", ""],
            ["Monitor Absenteeism",     "", "C", "R", "A", "", "", ""],
            ["Compile Monthly Summary", "", "", "R", "A", "", "", ""],
            ["Review Summary",          "", "", "I", "R", "A", "", ""],
            ["Handover to Payroll",     "", "", "I", "R", "A", "I", ""],
        ],
    }


def payroll_input_workflow():
    """Payroll Input Processing workflow steps (HR side of payroll)."""
    return {
        "name": "Payroll Input Processing",
        "steps": [
            {"step": 1, "phase": "Data Collection", "role": "HR Executive", "action": "Compile finalized attendance, leave, and LOP data", "decision": False, "doc": "Monthly Attendance Summary", "control": "Sign-off from Attendance workflow", "sla": "1st of month", "escalation": "HR Manager"},
            {"step": 2, "phase": "Data Collection", "role": "HR Executive", "action": "Update new joiners, exits, and confirmations in employee master", "decision": False, "doc": "Master_Employees", "control": "Approval letters on file for each change", "sla": "1st of month", "escalation": "HR Manager"},
            {"step": 3, "phase": "Data Collection", "role": "HR Executive", "action": "Update salary revisions, promotions, and increments", "decision": True, "yes": "Approved revision — update master", "no": "Pending approval — hold at old salary", "doc": "Increment Letter, Master_Employees", "control": "Approval letter required before update", "sla": "1st of month", "escalation": "HR Manager"},
            {"step": 4, "phase": "Verification", "role": "HR Manager", "action": "Cross-check master data against attendance and HR letters issued", "decision": True, "yes": "Data verified — proceed", "no": "Correct discrepancies", "doc": "Master_Employees, Letter Register", "control": "100% cross-check before handover", "sla": "2nd of month", "escalation": "HR Head"},
            {"step": 5, "phase": "Handover", "role": "HR Manager", "action": "Hand over verified payroll input file to Finance team", "decision": False, "doc": "Payroll Input Sheet", "control": "Formal sign-off / email confirmation", "sla": "3rd of month", "escalation": "HR Head"},
            {"step": 6, "phase": "Query Resolution", "role": "HR Executive", "action": "Respond to Finance queries on attendance/master data during payroll run", "decision": True, "yes": "Query resolved — Finance proceeds", "no": "Escalate for urgent resolution", "doc": "Query Log", "control": "Same-day turnaround target", "sla": "Ongoing during payroll cycle", "escalation": "HR Manager"},
            {"step": 7, "phase": "Payslip Support", "role": "HR Executive", "action": "Assist Finance in distributing payslips, address employee pay queries", "decision": False, "doc": "Payslips", "control": "Employee acknowledgement tracked", "sla": "7th of month", "escalation": "HR Manager"},
            {"step": 8, "phase": "Reconciliation", "role": "HR Manager", "action": "Reconcile headcount and cost-to-company reports with Finance post-payroll", "decision": True, "yes": "Matched — close cycle", "no": "Investigate variance", "doc": "Headcount Reconciliation Report", "control": "Monthly headcount vs budget check", "sla": "10th of month", "escalation": "HR Head"},
        ],
        "raci_roles": ["Employee", "HR Exec", "HR Manager", "HR Head", "Finance Exec", "Finance Manager", "CFO"],
        "raci": [
            ["Compile Attendance Data",   "", "R", "A", "", "I", "", ""],
            ["Update Joiners/Exits",      "", "R", "A", "", "", "", ""],
            ["Update Salary Revisions",   "", "R", "A", "", "", "", ""],
            ["Verify Master Data",        "", "C", "R", "A", "", "", ""],
            ["Handover to Finance",       "", "", "R", "A", "R", "I", ""],
            ["Resolve Payroll Queries",   "I", "R", "A", "", "C", "", ""],
            ["Support Payslip Distribution","I", "R", "A", "", "C", "", ""],
            ["Reconcile Headcount/Cost",  "", "", "R", "A", "C", "I", "I"],
        ],
    }


def statutory_compliance_workflow():
    """HR Statutory Compliance (PF/ESI/PT) workflow steps."""
    return {
        "name": "Statutory Compliance (PF/ESI/PT)",
        "steps": [
            {"step": 1, "phase": "Registration", "role": "HR Manager", "action": "Register establishment for PF (EPFO), ESI (ESIC), and Professional Tax as applicable", "decision": True, "yes": "Registered — obtain codes", "no": "Not yet applicable — monitor headcount/wage thresholds", "doc": "EPFO/ESIC Portal, PT Registration Certificate", "control": "Threshold monitoring (20 employees for PF, wage ceiling for ESI)", "sla": "Within statutory window of applicability", "escalation": "HR Head"},
            {"step": 2, "phase": "Employee Registration", "role": "HR Executive", "action": "Generate UAN for new PF-eligible employees; register ESI-eligible employees", "decision": True, "yes": "Registered — record in master", "no": "Not eligible — note applicability flag as No", "doc": "UAN Portal, ESIC Portal, Master_Employees", "control": "Wage ceiling check (PF ₹15,000 wage definition; ESI ₹21,000 gross)", "sla": "Within 1 day of joining", "escalation": "HR Manager"},
            {"step": 3, "phase": "Monthly Calculation", "role": "HR Executive", "action": "Compile PF/ESI/PT-applicable wages per employee from finalized payroll", "decision": False, "doc": "Salary Statement", "control": "EPF wage capped at ₹15,000; ESI at gross ≤ ₹21,000", "sla": "After payroll finalization", "escalation": "HR Manager"},
            {"step": 4, "phase": "Monthly Calculation", "role": "HR Executive", "action": "Generate PF ECR file (UAN, wages, contributions, NCP days) in ECR 2.0 format", "decision": True, "yes": "Cross-checks pass — proceed to filing", "no": "Correct wage/contribution mismatch", "doc": "ECR Text File", "control": "EPS+DIFF=EPF contribution cross-check; EPS capped at ₹1,250", "sla": "Before 15th of month", "escalation": "HR Manager"},
            {"step": 5, "phase": "Filing", "role": "HR Manager", "action": "Upload ECR on EPFO Unified Portal, generate challan, remit payment", "decision": True, "yes": "Remitted — file TRRN receipt", "no": "Correct rejection reason and re-upload", "doc": "EPFO Portal, Challan Receipt", "control": "Due date compliance (15th of month)", "sla": "By 15th of month", "escalation": "HR Head"},
            {"step": 6, "phase": "Filing", "role": "HR Manager", "action": "File monthly ESI contribution and remit payment via ESIC portal", "decision": False, "doc": "ESIC Portal, Challan Receipt", "control": "Due date compliance (15th of month)", "sla": "By 15th of month", "escalation": "HR Head"},
            {"step": 7, "phase": "Filing", "role": "HR Executive", "action": "Deduct and remit Professional Tax (PT) as per state slab", "decision": False, "doc": "PT Challan", "control": "State-specific slab accuracy", "sla": "Per state due date", "escalation": "HR Manager"},
            {"step": 8, "phase": "Reconciliation", "role": "HR Manager", "action": "Reconcile filed contributions against payroll deductions", "decision": True, "yes": "Matched — archive", "no": "Investigate and correct in next cycle", "doc": "Reconciliation Statement", "control": "Zero variance target", "sla": "Month-end", "escalation": "HR Head"},
            {"step": 9, "phase": "Annual Compliance", "role": "HR Manager", "action": "Complete annual returns (PF Form 3A/6A equivalent via ECR, ESI half-yearly return, Gratuity/Bonus Act compliance)", "decision": False, "doc": "Annual Statutory Returns", "control": "Statutory calendar tracking", "sla": "Per statutory calendar", "escalation": "HR Head"},
            {"step": 10, "phase": "Audit", "role": "HR Head", "action": "Coordinate statutory/labour law audit, provide records to auditors", "decision": False, "doc": "Compliance Register, Audit Report", "control": "Complete documentation trail", "sla": "As scheduled", "escalation": "CEO"},
        ],
        "raci_roles": ["HR Exec", "HR Manager", "HR Head", "Finance", "CEO", "Auditor"],
        "raci": [
            ["Establishment Registration", "", "R", "A", "C", "", ""],
            ["Employee PF/ESI Registration","R", "A", "", "", "", ""],
            ["Compile Applicable Wages",    "R", "A", "C", "", "", ""],
            ["Generate ECR File",           "R", "A", "", "", "", "I"],
            ["File PF & Remit",             "", "R", "A", "C", "", "I"],
            ["File ESI & Remit",            "", "R", "A", "C", "", "I"],
            ["Remit Professional Tax",      "R", "A", "", "", "", ""],
            ["Reconcile Contributions",     "C", "R", "A", "C", "", "I"],
            ["Annual Returns",              "C", "R", "A", "", "", "I"],
            ["Statutory Audit Coordination","", "C", "R", "", "A", "R"],
        ],
    }


def performance_workflow():
    """Performance Management workflow steps."""
    return {
        "name": "Performance Management",
        "steps": [
            {"step": 1, "phase": "Goal Setting", "role": "Reporting Manager", "action": "Set annual KRAs/goals with employee at start of appraisal cycle", "decision": False, "doc": "Goal Sheet / KRA Form", "control": "SMART goal criteria check", "sla": "Start of cycle", "escalation": "HR Manager"},
            {"step": 2, "phase": "Goal Setting", "role": "HR Executive", "action": "Consolidate goal sheets across departments, verify completeness", "decision": True, "yes": "Complete — proceed", "no": "Follow up with missing managers", "doc": "Goal Sheet Tracker", "control": "100% coverage target", "sla": "T+2 weeks", "escalation": "HR Manager"},
            {"step": 3, "phase": "Mid-Year Review", "role": "Reporting Manager", "action": "Conduct mid-year check-in, document progress and feedback", "decision": False, "doc": "Mid-Year Review Form", "control": "Documented feedback required", "sla": "Mid-cycle", "escalation": "HR Manager"},
            {"step": 4, "phase": "Self-Assessment", "role": "Employee", "action": "Complete self-assessment against goals at year-end", "decision": False, "doc": "Self-Assessment Form", "control": "Submission deadline enforced", "sla": "End of cycle", "escalation": "Reporting Manager"},
            {"step": 5, "phase": "Manager Review", "role": "Reporting Manager", "action": "Rate performance against goals, provide written feedback", "decision": True, "yes": "Rating finalized — submit for calibration", "no": "Revise rating with justification", "doc": "Performance Appraisal Form", "control": "Rating scale consistency check", "sla": "T+1 week after self-assessment", "escalation": "HR Manager"},
            {"step": 6, "phase": "Calibration", "role": "HR Manager", "action": "Facilitate calibration session across teams to normalize ratings", "decision": True, "yes": "Ratings calibrated — finalize", "no": "Return to manager for re-rating", "doc": "Calibration Matrix", "control": "Forced distribution guideline (bell curve reference)", "sla": "T+2 weeks", "escalation": "HR Head"},
            {"step": 7, "phase": "Approval", "role": "HR Head", "action": "Approve final ratings and increment/bonus recommendations", "decision": True, "yes": "Approved — proceed to communication", "no": "Return for revision", "doc": "Increment Recommendation Sheet", "control": "Budget vs increment pool check", "sla": "T+3 weeks", "escalation": "CEO"},
            {"step": 8, "phase": "Communication", "role": "Reporting Manager", "action": "Communicate final rating, increment, and development plan to employee", "decision": False, "doc": "Appraisal Letter", "control": "1:1 discussion documented", "sla": "T+4 weeks", "escalation": "HR Manager"},
            {"step": 9, "phase": "Documentation", "role": "HR Executive", "action": "Update employee master with new rating, increment, and promotion status", "decision": False, "doc": "Master_Employees", "control": "Signed appraisal letter on file", "sla": "T+4 weeks", "escalation": "HR Manager"},
            {"step": 10, "phase": "Handover", "role": "HR Manager", "action": "Notify Finance of approved salary revisions for next payroll cycle", "decision": False, "doc": "Increment Notification", "control": "Effective-date accuracy", "sla": "Before next payroll cut-off", "escalation": "HR Head"},
        ],
        "raci_roles": ["Employee", "Reporting Mgr", "HR Exec", "HR Manager", "HR Head", "CEO", "Finance"],
        "raci": [
            ["Set Goals",                  "R", "R", "", "A", "", "", ""],
            ["Consolidate Goal Sheets",     "", "C", "R", "A", "", "", ""],
            ["Mid-Year Review",            "R", "R", "", "A", "", "", ""],
            ["Self-Assessment",            "R", "I", "", "A", "", "", ""],
            ["Manager Rating",             "I", "R", "", "A", "", "", ""],
            ["Calibration",                "", "C", "", "R", "A", "", ""],
            ["Approve Ratings",            "I", "I", "", "C", "R", "A", ""],
            ["Communicate Results",        "I", "R", "", "A", "", "", ""],
            ["Update Master Records",      "", "", "R", "A", "", "", ""],
            ["Notify Finance",             "", "", "", "R", "A", "", "I"],
        ],
    }


def exit_workflow():
    """Exit & Full-and-Final Settlement workflow steps."""
    return {
        "name": "Exit & Full-and-Final Settlement",
        "steps": [
            {"step": 1, "phase": "Initiation", "role": "Employee", "action": "Submit resignation with notice period as per contract", "decision": False, "doc": "Resignation Letter", "control": "Notice period per employment contract", "sla": "Day 0", "escalation": "Reporting Manager"},
            {"step": 2, "phase": "Initiation", "role": "Reporting Manager", "action": "Acknowledge resignation, discuss knowledge transfer plan", "decision": True, "yes": "Accepted — proceed", "no": "Negotiate retention / withdrawal", "doc": "Acceptance Email", "control": "KT plan documented", "sla": "T+2 days", "escalation": "HR Manager"},
            {"step": 3, "phase": "Notice Period", "role": "HR Executive", "action": "Update HR master with last working day (LWD), initiate exit checklist", "decision": False, "doc": "Exit Checklist, Master_Employees", "control": "LWD matches notice period rules", "sla": "T+3 days", "escalation": "HR Manager"},
            {"step": 4, "phase": "Exit Interview", "role": "HR Executive", "action": "Conduct exit interview, capture reason for leaving and feedback", "decision": False, "doc": "Exit Interview Form", "control": "Standard questionnaire used", "sla": "Before LWD", "escalation": "HR Manager"},
            {"step": 5, "phase": "Clearance", "role": "IT/Admin/Finance", "action": "Complete department clearances — IT assets, finance dues, admin property", "decision": True, "yes": "All cleared — proceed to settlement", "no": "Recover dues / assets before settlement", "doc": "No Dues Certificate", "control": "Multi-department sign-off required", "sla": "By LWD", "escalation": "HR Manager"},
            {"step": 6, "phase": "Settlement Calculation", "role": "HR Executive", "action": "Calculate full and final settlement — pending salary, leave encashment, gratuity, deductions", "decision": False, "doc": "F&F Settlement Sheet", "control": "Leave encashment per policy; Gratuity per Payment of Gratuity Act (if eligible)", "sla": "T+7 days after LWD", "escalation": "HR Manager"},
            {"step": 7, "phase": "Approval", "role": "HR Manager", "action": "Review and approve F&F settlement calculation", "decision": True, "yes": "Approved — forward to Finance for payment", "no": "Correct calculation errors", "doc": "Approved F&F Sheet", "control": "Cross-check against no-dues certificate", "sla": "T+10 days after LWD", "escalation": "HR Head"},
            {"step": 8, "phase": "Payment", "role": "Finance Manager", "action": "Process F&F settlement payment to employee", "decision": False, "doc": "Bank Transfer, Payment Voucher", "control": "Statutory timeline (within 2 working days of LWD per Karnataka S&E Act, or as applicable)", "sla": "Per state Shops & Establishment Act", "escalation": "CFO"},
            {"step": 9, "phase": "Documentation", "role": "HR Executive", "action": "Issue relieving letter, experience certificate, Form 16 (part), PF withdrawal/transfer assistance", "decision": False, "doc": "Relieving Letter, Experience Certificate", "control": "Documents issued only after clearance", "sla": "On/after LWD", "escalation": "HR Manager"},
            {"step": 10, "phase": "Completion", "role": "HR Executive", "action": "Update employee master status to 'Exited', archive personnel file", "decision": False, "doc": "Master_Employees, Personnel Archive", "control": "Retention period per policy (statutory minimum)", "sla": "Post-settlement", "escalation": "HR Manager"},
        ],
        "raci_roles": ["Employee", "Reporting Mgr", "HR Exec", "HR Manager", "HR Head", "Finance", "IT/Admin"],
        "raci": [
            ["Submit Resignation",          "R", "A", "", "", "", "", ""],
            ["Acknowledge & KT Plan",       "C", "R", "", "A", "", "", ""],
            ["Update Master & Checklist",   "", "", "R", "A", "", "", ""],
            ["Exit Interview",              "R", "", "R", "A", "", "", ""],
            ["Department Clearance",        "R", "", "C", "A", "", "C", "R"],
            ["Calculate F&F",               "", "", "R", "A", "", "C", ""],
            ["Approve F&F",                 "", "", "I", "R", "A", "", ""],
            ["Process F&F Payment",         "I", "", "", "A", "", "R", ""],
            ["Issue Relieving Documents",   "I", "", "R", "A", "", "", ""],
            ["Close Personnel File",        "", "", "R", "A", "", "", ""],
        ],
    }


def grievance_posh_workflow():
    """Employee Grievance & POSH workflow steps."""
    return {
        "name": "Grievance & POSH",
        "steps": [
            {"step": 1, "phase": "Complaint", "role": "Employee", "action": "Raise grievance or POSH complaint via designated channel", "decision": False, "doc": "Grievance/Complaint Form", "control": "Confidential submission channel", "sla": "Day 0", "escalation": "HR Manager"},
            {"step": 2, "phase": "Acknowledgement", "role": "HR Executive", "action": "Acknowledge receipt, log complaint confidentially", "decision": True, "yes": "POSH-related — route to Internal Committee (ICC)", "no": "General grievance — route to HR Manager", "doc": "Complaint Register", "control": "Confidentiality maintained; access restricted", "sla": "Within 24 hours", "escalation": "HR Manager"},
            {"step": 3, "phase": "Investigation Setup", "role": "HR Manager / ICC", "action": "Constitute investigation — ICC for POSH (per POSH Act 2013), HR panel for general grievance", "decision": False, "doc": "Investigation Committee Order", "control": "ICC composition per statutory requirement (external member included)", "sla": "Within 7 days (POSH)", "escalation": "HR Head"},
            {"step": 4, "phase": "Investigation", "role": "ICC / HR Panel", "action": "Interview complainant, respondent, and witnesses", "decision": False, "doc": "Investigation Notes", "control": "Natural justice — both parties heard", "sla": "Within 90 days (POSH statutory limit)", "escalation": "HR Head"},
            {"step": 5, "phase": "Findings", "role": "ICC / HR Panel", "action": "Document findings and recommend action", "decision": True, "yes": "Substantiated — recommend action", "no": "Not substantiated — close with documented reasoning", "doc": "Inquiry Report", "control": "Evidence-based findings required", "sla": "Within statutory timeline", "escalation": "HR Head"},
            {"step": 6, "phase": "Action", "role": "HR Head", "action": "Approve disciplinary action or corrective measure per findings", "decision": True, "yes": "Action approved — implement", "no": "Return for further review", "doc": "Disciplinary Action Letter", "control": "Proportionate to severity; legal review for termination", "sla": "T+7 days after report", "escalation": "CEO"},
            {"step": 7, "phase": "Communication", "role": "HR Manager", "action": "Communicate outcome to complainant (and respondent, as applicable) confidentially", "decision": False, "doc": "Outcome Communication", "control": "Confidentiality maintained", "sla": "T+3 days after decision", "escalation": "HR Head"},
            {"step": 8, "phase": "Closure", "role": "HR Executive", "action": "Close case in register, retain records per statutory retention period", "decision": False, "doc": "Complaint Register, Case Archive", "control": "Restricted access archive", "sla": "Post-closure", "escalation": "HR Manager"},
            {"step": 9, "phase": "Annual Reporting", "role": "HR Head", "action": "Prepare annual POSH report for inclusion in company Annual Return (as applicable)", "decision": False, "doc": "Annual POSH Report", "control": "Statutory disclosure requirement", "sla": "Annually", "escalation": "CEO"},
        ],
        "raci_roles": ["Employee", "HR Exec", "HR Manager", "ICC/Panel", "HR Head", "CEO", "Legal"],
        "raci": [
            ["Raise Complaint",           "R", "A", "", "", "", "", ""],
            ["Acknowledge & Log",         "I", "R", "A", "", "", "", ""],
            ["Constitute Committee",      "", "", "R", "A", "A", "", "C"],
            ["Conduct Investigation",     "C", "", "", "R", "A", "", "C"],
            ["Document Findings",         "I", "", "", "R", "A", "", "C"],
            ["Approve Action",            "I", "", "", "C", "R", "A", "C"],
            ["Communicate Outcome",       "I", "", "R", "", "A", "", ""],
            ["Close & Archive",           "", "R", "A", "", "", "", ""],
            ["Annual Reporting",          "", "", "C", "", "R", "A", "C"],
        ],
    }


def training_workflow():
    """Training & Development workflow steps."""
    return {
        "name": "Training & Development",
        "steps": [
            {"step": 1, "phase": "Needs Identification", "role": "Reporting Manager", "action": "Identify skill gaps during appraisal/project reviews", "decision": False, "doc": "Training Needs Form", "control": "Linked to appraisal feedback", "sla": "Ongoing / post-appraisal", "escalation": "HR Manager"},
            {"step": 2, "phase": "Needs Identification", "role": "HR Executive", "action": "Consolidate training needs across departments into annual training calendar", "decision": False, "doc": "Training Needs Analysis (TNA)", "control": "Coverage across all departments", "sla": "Annually", "escalation": "HR Manager"},
            {"step": 3, "phase": "Planning", "role": "HR Manager", "action": "Finalize training calendar, budget, and vendor/trainer selection", "decision": True, "yes": "Approved — proceed to scheduling", "no": "Revise budget/scope", "doc": "Training Calendar, Budget", "control": "Budget approval before commitment", "sla": "T+2 weeks", "escalation": "HR Head"},
            {"step": 4, "phase": "Scheduling", "role": "HR Executive", "action": "Schedule sessions, nominate/invite participants", "decision": False, "doc": "Training Invite, Nomination List", "control": "Manager sign-off on nominations", "sla": "T-1 week before session", "escalation": "HR Manager"},
            {"step": 5, "phase": "Delivery", "role": "Trainer/Vendor", "action": "Conduct training session (internal or external)", "decision": False, "doc": "Session Material, Attendance Sheet", "control": "Attendance capture mandatory", "sla": "As scheduled", "escalation": "HR Manager"},
            {"step": 6, "phase": "Evaluation", "role": "HR Executive", "action": "Collect participant feedback and assessment scores", "decision": False, "doc": "Feedback Form, Assessment Results", "control": "Minimum response rate target (>80%)", "sla": "T+1 day after session", "escalation": "HR Manager"},
            {"step": 7, "phase": "Evaluation", "role": "HR Manager", "action": "Evaluate training effectiveness against objectives", "decision": True, "yes": "Effective — continue program", "no": "Revise content/vendor for next cycle", "doc": "Effectiveness Report", "control": "Pre/post assessment comparison", "sla": "T+2 weeks", "escalation": "HR Head"},
            {"step": 8, "phase": "Documentation", "role": "HR Executive", "action": "Update employee training records and certifications in master", "decision": False, "doc": "Master_Employees, Training Log", "control": "Certification expiry tracking (statutory trainings)", "sla": "T+1 week after session", "escalation": "HR Manager"},
        ],
        "raci_roles": ["Employee", "Reporting Mgr", "HR Exec", "HR Manager", "HR Head", "Trainer/Vendor"],
        "raci": [
            ["Identify Skill Gaps",       "C", "R", "", "A", "", ""],
            ["Consolidate TNA",           "", "C", "R", "A", "", ""],
            ["Plan Calendar & Budget",    "", "", "C", "R", "A", ""],
            ["Schedule Sessions",         "I", "A", "R", "", "", "C"],
            ["Deliver Training",          "R", "I", "C", "A", "", "R"],
            ["Collect Feedback",          "R", "", "R", "A", "", ""],
            ["Evaluate Effectiveness",    "", "C", "C", "R", "A", ""],
            ["Update Training Records",   "", "", "R", "A", "", ""],
        ],
    }


def hr_policy_workflow():
    """HR Policy & Documentation workflow steps."""
    return {
        "name": "HR Policy & Documentation",
        "steps": [
            {"step": 1, "phase": "Trigger", "role": "HR Manager", "action": "Identify need for new policy or revision (legal change, incident, gap)", "decision": False, "doc": "Policy Change Request", "control": "Linked to statutory update or audit finding", "sla": "As triggered", "escalation": "HR Head"},
            {"step": 2, "phase": "Drafting", "role": "HR Executive", "action": "Draft policy document with scope, applicability, and procedure", "decision": False, "doc": "Draft Policy", "control": "Based on statutory minimum + company practice", "sla": "T+1-2 weeks", "escalation": "HR Manager"},
            {"step": 3, "phase": "Legal Review", "role": "Legal Counsel", "action": "Review draft for statutory compliance (Labour Codes, POSH, Maternity Benefit Act, etc.)", "decision": True, "yes": "Compliant — proceed", "no": "Return with required changes", "doc": "Legal Review Comments", "control": "Statutory citation check", "sla": "T+1 week", "escalation": "HR Head"},
            {"step": 4, "phase": "Stakeholder Review", "role": "Department Heads", "action": "Review policy for operational feasibility", "decision": True, "yes": "Feasible — proceed to approval", "no": "Flag concerns for revision", "doc": "Review Comments", "control": "Cross-department sign-off", "sla": "T+1 week", "escalation": "HR Head"},
            {"step": 5, "phase": "Approval", "role": "HR Head", "action": "Finalize policy and submit for CEO/leadership approval", "decision": True, "yes": "Approved — proceed to rollout", "no": "Revise and resubmit", "doc": "Final Policy Document", "control": "Version control maintained", "sla": "T+1 week", "escalation": "CEO"},
            {"step": 6, "phase": "Rollout", "role": "HR Executive", "action": "Communicate policy company-wide, publish on HR portal/handbook", "decision": False, "doc": "Policy Communication Email, Handbook", "control": "Acknowledgement tracking", "sla": "Within 1 week of approval", "escalation": "HR Manager"},
            {"step": 7, "phase": "Acknowledgement", "role": "Employee", "action": "Read and digitally acknowledge policy", "decision": True, "yes": "Acknowledged — logged", "no": "Follow up until 100% acknowledgement", "doc": "Acknowledgement Log", "control": "100% acknowledgement target", "sla": "T+2 weeks", "escalation": "HR Manager"},
            {"step": 8, "phase": "Periodic Audit", "role": "HR Manager", "action": "Review policy annually or on statutory change for continued relevance", "decision": True, "yes": "No change needed", "no": "Trigger revision cycle (Step 1)", "doc": "Policy Audit Checklist", "control": "Annual review calendar", "sla": "Annually", "escalation": "HR Head"},
        ],
        "raci_roles": ["Employee", "HR Exec", "HR Manager", "HR Head", "Legal", "Dept Heads", "CEO"],
        "raci": [
            ["Identify Need",             "", "C", "R", "A", "", "", ""],
            ["Draft Policy",              "", "R", "A", "", "C", "", ""],
            ["Legal Review",              "", "", "C", "A", "R", "", ""],
            ["Stakeholder Review",        "", "", "C", "A", "", "R", ""],
            ["Finalize & Approve",        "", "", "C", "R", "C", "", "A"],
            ["Rollout & Communicate",     "I", "R", "A", "", "", "", ""],
            ["Track Acknowledgement",     "R", "R", "A", "", "", "", ""],
            ["Periodic Audit",            "", "C", "R", "A", "C", "", ""],
        ],
    }


# ─── Excel Generation ────────────────────────────────────────────────────────

def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_cell(cell, wrap=True, border=True, center=False):
    """Apply standard styling to a cell."""
    if wrap:
        cell.alignment = WRAP_ALIGN if not center else CENTER_ALIGN
    if border:
        cell.border = THIN_BORDER


def apply_conditional_fills(ws, row, is_decision, has_control, has_escalation):
    """Apply conditional formatting fills to a row."""
    if is_decision:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = DECISION_FILL
    if has_control:
        ws.cell(row=row, column=9).fill = CONTROL_FILL  # Control column
    if has_escalation:
        ws.cell(row=row, column=11).fill = ESCALATION_FILL  # Escalation column


def create_flowchart_sheet(wb, workflow):
    """Create a flowchart sheet for a workflow."""
    name = workflow["name"]
    # Excel sheet names max 31 chars
    sheet_name = f"{name[:22]}_Flow" if len(name) > 22 else f"{name}_Flow"
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    headers = [
        "Step #", "Phase", "Responsible Role", "Action / Task",
        "Decision?", "If Yes →", "If No →",
        "Document / System", "Control Point", "SLA / Timeline", "Escalation To"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    # Data rows
    for i, step in enumerate(workflow["steps"], start=2):
        ws.cell(row=i, column=1, value=step["step"])
        ws.cell(row=i, column=2, value=step["phase"])
        ws.cell(row=i, column=3, value=step["role"])
        ws.cell(row=i, column=4, value=step["action"])
        ws.cell(row=i, column=5, value="YES" if step["decision"] else "")
        ws.cell(row=i, column=6, value=step.get("yes", ""))
        ws.cell(row=i, column=7, value=step.get("no", ""))
        ws.cell(row=i, column=8, value=step.get("doc", ""))
        ws.cell(row=i, column=9, value=step.get("control", ""))
        ws.cell(row=i, column=10, value=step.get("sla", ""))
        ws.cell(row=i, column=11, value=step.get("escalation", ""))

        # Style cells
        for col in range(1, 12):
            cell = ws.cell(row=i, column=col)
            style_cell(cell, center=(col in [1, 5]))

        apply_conditional_fills(ws, i, step["decision"], bool(step.get("control")), bool(step.get("escalation")))

    # Column widths
    col_widths = [8, 16, 20, 45, 10, 30, 30, 30, 30, 18, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    return sheet_name


def create_raci_sheet(wb, workflow):
    """Create a RACI matrix sheet for a workflow."""
    name = workflow["name"]
    sheet_name = f"{name[:22]}_RACI" if len(name) > 22 else f"{name}_RACI"
    ws = wb.create_sheet(title=sheet_name)

    roles = workflow.get("raci_roles", [])
    raci_data = workflow.get("raci", [])

    # Headers
    headers = ["Process Step"] + roles
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    # RACI fill map
    raci_fills = {"R": RACI_R_FILL, "A": RACI_A_FILL, "C": RACI_C_FILL, "I": RACI_I_FILL}
    raci_fonts = {
        "R": Font(bold=True, color="FFFFFF"),
        "A": Font(bold=True, color="FFFFFF"),
        "C": Font(bold=True, color="FFFFFF"),
        "I": Font(bold=True, color="000000"),
    }

    # Data rows
    for i, row_data in enumerate(raci_data, start=2):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            style_cell(cell, center=(col > 1))
            if col > 1 and value in raci_fills:
                cell.fill = raci_fills[value]
                cell.font = raci_fonts[value]

    # Column widths
    ws.column_dimensions["A"].width = 32
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Data validation for RACI values
    dv = DataValidation(type="list", formula1='"R,A,C,I,"', allow_blank=True)
    dv.error = "Please enter R, A, C, or I"
    dv.errorTitle = "Invalid RACI value"
    for col in range(2, len(headers) + 1):
        for row in range(2, len(raci_data) + 2):
            dv.add(ws.cell(row=row, column=col))
    ws.add_data_validation(dv)

    # RACI Legend
    legend_row = len(raci_data) + 4
    ws.cell(row=legend_row, column=1, value="RACI Legend:").font = Font(bold=True)
    legend_items = [
        ("R", "Responsible — Does the work", RACI_R_FILL),
        ("A", "Accountable — Owns the outcome", RACI_A_FILL),
        ("C", "Consulted — Provides input", RACI_C_FILL),
        ("I", "Informed — Kept in the loop", RACI_I_FILL),
    ]
    for j, (code, desc, fill) in enumerate(legend_items):
        row = legend_row + 1 + j
        cell_code = ws.cell(row=row, column=1, value=code)
        cell_code.fill = fill
        cell_code.font = Font(bold=True, color="FFFFFF" if code != "I" else "000000")
        cell_code.alignment = CENTER_ALIGN
        ws.cell(row=row, column=2, value=desc)

    ws.freeze_panes = "B2"
    return sheet_name


def create_index_sheet(wb, sheet_map, company_name):
    """Create an index/TOC sheet."""
    ws = wb.create_sheet(title="Index")
    wb.move_sheet(ws, offset=-len(wb.sheetnames) + 1)  # Move to first position

    # Title
    ws.cell(row=1, column=1, value=f"{company_name}").font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:D1")
    ws.cell(row=2, column=1, value="Human Resources — Workflow Charts").font = Font(name="Calibri", bold=True, size=13, color="2F5496")
    ws.merge_cells("A2:D2")
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y')}").font = Font(italic=True, color="666666")

    # Headers
    headers = ["#", "Workflow", "Flowchart Sheet", "RACI Sheet"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = INDEX_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Workflow entries
    for i, (wf_key, sheets) in enumerate(sheet_map.items(), start=1):
        row = 5 + i
        ws.cell(row=row, column=1, value=i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=wf_key).border = THIN_BORDER

        flow_name = sheets.get("flow", "")
        raci_name = sheets.get("raci", "")

        flow_cell = ws.cell(row=row, column=3, value=flow_name)
        flow_cell.font = Font(color="0563C1", underline="single")
        flow_cell.border = THIN_BORDER

        raci_cell = ws.cell(row=row, column=4, value=raci_name)
        raci_cell.font = Font(color="0563C1", underline="single")
        raci_cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30

    ws.freeze_panes = "A6"


def create_leave_policy_sheet(wb, leave_policy):
    """Create the leave policy matrix sheet."""
    ws = wb.create_sheet(title="Leave_Policy")

    # Title
    ws.cell(row=1, column=1, value="Leave Entitlement & Approval Policy").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:E1")

    # Headers
    headers = ["Leave Type", "Days/Year", "Approver", "Description", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    # Data
    for i, (leave_type, info) in enumerate(leave_policy.items(), start=4):
        days = info["days"]
        days_str = "Unpaid" if days is None else str(days)
        notes = ""
        if leave_type == "maternity_leave":
            notes = "Per Maternity Benefit Act, 1961"
        elif leave_type == "loss_of_pay":
            notes = "Applied when leave balance exhausted"

        ws.cell(row=i, column=1, value=leave_type.replace("_", " ").title())
        ws.cell(row=i, column=2, value=days_str)
        ws.cell(row=i, column=3, value=info["approver"])
        ws.cell(row=i, column=4, value=info.get("description", ""))
        ws.cell(row=i, column=5, value=notes)

        for col in range(1, 6):
            style_cell(ws.cell(row=i, column=col))

    # Column widths
    col_widths = [20, 12, 25, 35, 35]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


def create_controls_summary_sheet(wb, workflows):
    """Create a controls summary sheet across all workflows."""
    ws = wb.create_sheet(title="Controls_Summary")

    ws.cell(row=1, column=1, value="Internal Controls Summary").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:E1")

    headers = ["#", "Workflow", "Step", "Control Point", "Type"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    row = 4
    counter = 1
    for wf_key, wf_data in workflows.items():
        for step in wf_data["steps"]:
            if step.get("control"):
                ws.cell(row=row, column=1, value=counter)
                ws.cell(row=row, column=2, value=wf_data["name"])
                ws.cell(row=row, column=3, value=step["action"][:60])
                ws.cell(row=row, column=4, value=step["control"])

                # Determine control type
                control_text = step["control"].lower()
                if any(w in control_text for w in ["approval", "review", "sign-off", "authorize"]):
                    ctrl_type = "Authorization"
                elif any(w in control_text for w in ["reconcil", "match", "verify", "validation", "cross-check"]):
                    ctrl_type = "Verification"
                elif any(w in control_text for w in ["segregat", "dual", "independent", "confidential"]):
                    ctrl_type = "Segregation of Duties"
                elif any(w in control_text for w in ["monitor", "track", "aging", "deadline", "threshold"]):
                    ctrl_type = "Monitoring"
                elif any(w in control_text for w in ["statutory", "act", "compliance"]):
                    ctrl_type = "Statutory Compliance"
                else:
                    ctrl_type = "Preventive"

                ws.cell(row=row, column=5, value=ctrl_type)

                for col in range(1, 6):
                    style_cell(ws.cell(row=row, column=col))

                counter += 1
                row += 1

    col_widths = [6, 28, 45, 40, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


def create_kpi_sheet(wb):
    """Create KPI metrics sheet."""
    ws = wb.create_sheet(title="KPI_Metrics")

    ws.cell(row=1, column=1, value="HR Key Performance Indicators").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:E1")

    headers = ["#", "KPI", "Workflow", "Target", "Measurement Frequency"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    kpis = [
        ("Time to Fill", "Recruitment", "≤ 30 days", "Monthly"),
        ("Offer Acceptance Rate", "Recruitment", "≥ 85%", "Monthly"),
        ("Cost per Hire", "Recruitment", "Within budget", "Quarterly"),
        ("Onboarding Completion Rate", "Onboarding", "100% within Day 1", "Monthly"),
        ("Probation Confirmation Turnaround", "Onboarding", "On/before due date", "Monthly"),
        ("Attendance Regularization TAT", "Attendance", "≤ 2 working days", "Monthly"),
        ("Absenteeism Rate", "Attendance", "≤ 5%", "Monthly"),
        ("Payroll Input Accuracy", "Payroll Input", "100% (zero errors)", "Monthly"),
        ("Payroll Handover Timeliness", "Payroll Input", "By 3rd of month", "Monthly"),
        ("Statutory Filing Compliance", "Statutory Compliance", "100% on-time filing", "Monthly"),
        ("PF/ESI Reconciliation Variance", "Statutory Compliance", "Zero variance", "Monthly"),
        ("Appraisal Completion Rate", "Performance Management", "100% by cycle end", "Annually"),
        ("Employee Attrition Rate", "Exit", "≤ 12% annually", "Monthly"),
        ("F&F Settlement TAT", "Exit", "Per state S&E Act", "Per exit"),
        ("Grievance Resolution TAT", "Grievance & POSH", "Within statutory/policy limit", "Per case"),
        ("Training Hours per Employee", "Training & Development", "≥ 20 hrs/year", "Annually"),
        ("Policy Acknowledgement Rate", "HR Policy", "100%", "Per rollout"),
        ("Employee Engagement Score", "Overall HR", "≥ 75%", "Annually"),
    ]

    for i, (kpi, workflow, target, freq) in enumerate(kpis, start=4):
        ws.cell(row=i, column=1, value=i - 3)
        ws.cell(row=i, column=2, value=kpi)
        ws.cell(row=i, column=3, value=workflow)
        ws.cell(row=i, column=4, value=target)
        ws.cell(row=i, column=5, value=freq)
        for col in range(1, 6):
            style_cell(ws.cell(row=i, column=col))

    col_widths = [6, 35, 25, 25, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


# ─── Main Generator ──────────────────────────────────────────────────────────

def generate_workbook(
    company_name: str = "Fracktal Works Private Limited",
    selected_workflows: list = None,
    leave_policy: dict = None,
    output_path: str = None,
):
    """Generate the complete HR workflow charts workbook."""
    TMP_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(TMP_DIR / f"HR_Workflow_Charts_{timestamp}.xlsx")

    if leave_policy is None:
        leave_policy = DEFAULT_LEAVE_POLICY

    all_workflows = get_all_workflows()

    # Filter workflows if specified
    if selected_workflows:
        # Normalize keys for matching
        key_map = {k.lower().replace(" ", "").replace("-", "").replace("&", ""): k for k in all_workflows}
        filtered = {}
        for sel in selected_workflows:
            normalized = sel.strip().lower().replace(" ", "").replace("-", "").replace("&", "")
            if normalized in key_map:
                real_key = key_map[normalized]
                filtered[real_key] = all_workflows[real_key]
            else:
                print(f"  ⚠ Unknown workflow: '{sel}' — skipping")
        workflows = filtered
    else:
        workflows = all_workflows

    if not workflows:
        print("ERROR: No valid workflows selected.")
        sys.exit(1)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print(f"\n{'='*60}")
    print(f"GENERATING HR WORKFLOW CHARTS")
    print(f"Company: {company_name}")
    print(f"Workflows: {len(workflows)}")
    print(f"{'='*60}\n")

    # Generate flowchart and RACI sheets for each workflow
    sheet_map = {}
    for wf_key, wf_data in workflows.items():
        print(f"  → {wf_data['name']}...")
        flow_sheet = create_flowchart_sheet(wb, wf_data)
        raci_sheet = create_raci_sheet(wb, wf_data)
        sheet_map[wf_data["name"]] = {"flow": flow_sheet, "raci": raci_sheet}
        print(f"    ✓ Flowchart: {flow_sheet}")
        print(f"    ✓ RACI: {raci_sheet}")

    # Create summary sheets
    print(f"\n  → Summary sheets...")
    create_index_sheet(wb, sheet_map, company_name)
    print(f"    ✓ Index (Table of Contents)")

    create_leave_policy_sheet(wb, leave_policy)
    print(f"    ✓ Leave Policy")

    create_controls_summary_sheet(wb, workflows)
    print(f"    ✓ Controls Summary")

    create_kpi_sheet(wb)
    print(f"    ✓ KPI Metrics")

    # Save
    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"✓ WORKBOOK SAVED: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Workflows: {len(workflows)}")
    print(f"{'='*60}\n")

    return output_path


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate HR workflow charts as Excel workbook"
    )
    parser.add_argument(
        "--company",
        default="Fracktal Works Private Limited",
        help="Company name (default: Fracktal Works Private Limited)"
    )
    parser.add_argument(
        "--workflows",
        default=None,
        help="Comma-separated workflow names (default: all). Options: Recruitment, Onboarding, Attendance, Payroll Input, Statutory Compliance, Performance Management, Exit, Grievance & POSH, Training & Development, HR Policy"
    )
    parser.add_argument(
        "--leave-policy",
        default=None,
        help="JSON string with custom leave policy"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output file path (default: .tmp/HR_Workflow_Charts_<timestamp>.xlsx)"
    )
    parser.add_argument(
        "--list-workflows",
        action="store_true",
        help="List all available workflows"
    )

    args = parser.parse_args()

    if args.list_workflows:
        all_wf = get_all_workflows()
        print("\nAvailable Workflows:")
        for key, wf in all_wf.items():
            print(f"  • {key}: {wf['name']} ({len(wf['steps'])} steps)")
        return

    selected = None
    if args.workflows:
        selected = [w.strip() for w in args.workflows.split(",")]

    leave_policy = None
    if args.leave_policy:
        try:
            leave_policy = json.loads(args.leave_policy)
        except json.JSONDecodeError as e:
            print(f"ERROR: Invalid JSON for leave policy: {e}")
            sys.exit(1)

    generate_workbook(
        company_name=args.company,
        selected_workflows=selected,
        leave_policy=leave_policy,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
