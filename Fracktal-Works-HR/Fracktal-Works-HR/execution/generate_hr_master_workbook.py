#!/usr/bin/env python3
"""
HR/Payroll Master Workbook Generator — Fracktal Works Private Limited

Generates an EMPTY employee master + monthly salary statement workbook,
structured the same way as an operational HR/payroll system (comparable
to the Thota Hospitality LLP payroll workbook), ready for the user to
populate with real employee data.

No employee data is fabricated — every data area is left blank for the
user to fill in. Formulas are pre-wired so that once Master_Employees is
populated, every monthly Salary Stmt sheet calculates automatically.

Usage:
    python generate_hr_master_workbook.py --company "Fracktal Works Private Limited" --fy "2026-27"
    python generate_hr_master_workbook.py --fy "2026-27" --employee-rows 60 --output "Fracktal_HR_Master.xlsx"
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
PLACEHOLDER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NOTE_FONT = Font(italic=True, color="C00000", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

MONTHS_FY = [
    "APR", "MAY", "JUN", "JUL", "AUG", "SEP",
    "OCT", "NOV", "DEC", "JAN", "FEB", "MAR"
]

# Master_Employees column layout
MASTER_HEADERS = {
    "A": "SlNo", "B": "Employee ID", "C": "Name", "D": "Designation", "E": "Department",
    "F": "UAN", "G": "PF No", "H": "ESI No", "I": "PAN",
    "J": "Bank Acc No", "K": "Bank Name", "L": "IFSC",
    "M": "DOJ", "N": "DOL", "O": "Status",
    "P": "Monthly Gross (CTC)", "Q": "Basic %", "R": "HRA %", "S": "Conveyance",
    "T": "PF Applicable", "U": "ESI Applicable", "V": "PT Applicable",
}

# Salary Stmt column layout (formula-driven off Master_Employees)
SALARY_HEADERS = {
    "A": "SlNo", "B": "Employee ID", "C": "Name", "D": "Designation", "F": "UAN",
    "L": "Calendar Days", "M": "Present Days", "P": "LOP Days", "R": "Paid Days",
    "T": "Monthly Gross (CTC)", "V": "Eff. Gross (after LOP)",
    "W": "Basic", "X": "HRA", "Y": "Conveyance", "Z": "Special Allowance",
    "AD": "Employee PF (12%)", "AE": "Employee ESI (0.75%)", "AF": "Employer PF (13% incl. admin)",
    "AG": "Employer ESI (3.25%)", "AH": "Professional Tax", "AI": "Net Salary Payable",
}


def style_header_row(ws, row, columns):
    for col_letter in columns:
        cell = ws[f"{col_letter}{row}"]
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def create_instructions_sheet(wb, company_name, fy_label, employee_rows):
    ws = wb.create_sheet(title="INSTRUCTIONS")
    ws.column_dimensions["A"].width = 110

    lines = [
        (f"{company_name} — HR & Payroll Master Workbook (FY {fy_label})", TITLE_FONT),
        ("", None),
        ("This workbook is a TEMPLATE. No employee data has been entered.", NOTE_FONT),
        ("", None),
        ("BEFORE FIRST USE — fill in these placeholders:", Font(bold=True)),
        ("  1. Statutory_Compliance sheet — EPFO Establishment Code, ESI Employer Code, PT Registration (state-specific)", None),
        ("  2. Master_Employees sheet — add one row per employee starting at row 3", None),
        ("     Required: Name, UAN (if PF applicable), Bank details, DOJ, Monthly Gross, Basic %, HRA %, PF/ESI/PT Applicable flags", None),
        ("", None),
        ("HOW THE WORKBOOK WORKS:", Font(bold=True)),
        ("  - Master_Employees is the single source of truth for employee data.", None),
        ("  - Each 'Salary Stmt-<MON><YY>' sheet pulls Name/UAN/Gross/etc. from Master_Employees by row position", None),
        ("    (row 3 in Master_Employees = row 3 in every Salary Stmt sheet). Keep employee row order consistent.", None),
        ("  - PF, ESI, and PT are calculated automatically based on the Applicable flags and statutory rules:", None),
        ("      PF Wage = MIN(Effective Gross used for PF components, 15000), only if PF Applicable = Yes", None),
        ("      Employee PF = ROUND(12% x PF Wage), Employer PF = ROUND(13% x PF Wage) [12% + ~1% admin/EDLI]", None),
        ("      ESI = applicable only if ESI Applicable = Yes AND Effective Gross <= 21000", None),
        ("      Employee ESI = ROUND(0.75% x Gross), Employer ESI = ROUND(3.25% x Gross)", None),
        ("      Professional Tax (PT) uses the Karnataka PT slab by default — update the PT formula if registered in another state", None),
        ("  - Attendance columns (Present Days, LOP Days) are entered manually per month before payroll run.", None),
        ("", None),
        ("MONTHLY PAYROLL CYCLE (see directives/hr_workflow_charts.md → Payroll Input Processing):", Font(bold=True)),
        ("  1. HR compiles attendance/leave and updates Master_Employees for joiners/exits/revisions", None),
        ("  2. Enter Present Days / LOP Days in the current month's Salary Stmt sheet", None),
        ("  3. Verify calculated PF/ESI/PT/Net Salary, hand off to Finance for disbursement", None),
        ("  4. Once populated, run execution/generate_ecr_file.py to produce the monthly PF ECR upload file", None),
        ("", None),
        (f"Sheets: INSTRUCTIONS, Statutory_Compliance, Master_Employees, "
         f"Salary Stmt-{MONTHS_FY[0]}{fy_label[2:4]}…{MONTHS_FY[-1]}{fy_label[-2:]}, Leave_Tracker, Annual_Summary", None),
        (f"Template capacity: {employee_rows} employee rows (extend by copying the last formula row down if needed)", None),
        ("", None),
        (f"Generated: {datetime.now().strftime('%d %B %Y')}", Font(italic=True, color="666666")),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_statutory_compliance_sheet(wb, company_name):
    ws = wb.create_sheet(title="Statutory_Compliance")
    ws.cell(row=1, column=1, value=f"{company_name} — Statutory Registration Details").font = TITLE_FONT
    ws.merge_cells("A1:C1")

    ws.cell(row=3, column=1, value="Statutory registrations are company-specific and unknown until confirmed. "
                                    "Replace every <TO BE FILLED> below.").font = NOTE_FONT
    ws.merge_cells("A3:C3")
    ws.row_dimensions[3].height = 30
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    fields = [
        ("EPFO Establishment Code", "<TO BE FILLED>"),
        ("ESIC Employer Code", "<TO BE FILLED>"),
        ("Professional Tax Registration No. (PT)", "<TO BE FILLED>"),
        ("State (for PT slab)", "Karnataka"),
        ("PF Wage Ceiling (₹)", 15000),
        ("ESI Gross Wage Ceiling (₹)", 21000),
        ("Employee PF Rate", "12%"),
        ("Employer PF Rate (incl. admin/EDLI)", "13%"),
        ("Employee ESI Rate", "0.75%"),
        ("Employer ESI Rate", "3.25%"),
        ("PF Due Date", "15th of following month"),
        ("ESI Due Date", "15th of following month"),
        ("PT Due Date", "20th of following month (Karnataka)"),
    ]

    headers = ["Field", "Value"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, ["A", "B"])

    for i, (field, value) in enumerate(fields, start=6):
        ws.cell(row=i, column=1, value=field).border = THIN_BORDER
        cell = ws.cell(row=i, column=2, value=value)
        cell.border = THIN_BORDER
        if value == "<TO BE FILLED>":
            cell.fill = PLACEHOLDER_FILL

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "A6"


def create_master_employees_sheet(wb, company_name, employee_rows):
    ws = wb.create_sheet(title="Master_Employees")

    ws.cell(row=1, column=1, value=f"{company_name} — Master Employee Data").font = TITLE_FONT
    ws.merge_cells("A1:V1")

    for col_letter, header in MASTER_HEADERS.items():
        ws[f"{col_letter}2"] = header
    style_header_row(ws, 2, list(MASTER_HEADERS.keys()))

    # SlNo formulas + data validation, no employee data
    for r in range(3, 3 + employee_rows):
        ws[f"A{r}"] = f"=IF(C{r}=\"\",\"\",ROW()-2)"
        for col_letter in MASTER_HEADERS:
            ws[f"{col_letter}{r}"].border = THIN_BORDER

    # Data validations
    status_dv = DataValidation(type="list", formula1='"Active,Exited,On Notice"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"O3:O{2 + employee_rows}")

    yesno_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f"T3:T{2 + employee_rows}")
    yesno_dv.add(f"U3:U{2 + employee_rows}")
    yesno_dv.add(f"V3:V{2 + employee_rows}")

    col_widths = {
        "A": 6, "B": 14, "C": 22, "D": 20, "E": 18, "F": 15, "G": 15, "H": 15, "I": 13,
        "J": 18, "K": 16, "L": 13, "M": 12, "N": 12, "O": 12, "P": 16, "Q": 10, "R": 10,
        "S": 12, "T": 14, "U": 14, "V": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A3"


def create_salary_sheet(wb, month_label, employee_rows):
    """Create one monthly Salary Stmt sheet, formula-linked to Master_Employees."""
    sheet_name = f"Salary Stmt-{month_label}"
    ws = wb.create_sheet(title=sheet_name)

    for col_letter, header in SALARY_HEADERS.items():
        ws[f"{col_letter}2"] = header
    style_header_row(ws, 2, list(SALARY_HEADERS.keys()))

    for r in range(3, 3 + employee_rows):
        me_row = r  # same row position in Master_Employees
        ws[f"A{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!A{me_row})"
        ws[f"B{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!B{me_row})"
        ws[f"C{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!C{me_row})"
        ws[f"D{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!D{me_row})"
        ws[f"F{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!F{me_row})"

        # Attendance — entered manually by HR each month
        ws[f"L{r}"] = 30  # Calendar Days (adjust per month)
        # M = Present Days (manual entry)
        # P = LOP Days (manual entry)
        ws[f"R{r}"] = f"=IF(C{r}=\"\",\"\",L{r}-P{r})"  # Paid Days

        ws[f"T{r}"] = f"=IF(C{r}=\"\",\"\",Master_Employees!P{me_row})"  # Monthly Gross
        ws[f"V{r}"] = f"=IF(C{r}=\"\",\"\",ROUND(T{r}/L{r}*R{r},0))"  # Effective Gross after LOP

        ws[f"W{r}"] = f"=IF(C{r}=\"\",\"\",ROUND(V{r}*Master_Employees!Q{me_row},0))"  # Basic
        ws[f"X{r}"] = f"=IF(C{r}=\"\",\"\",ROUND(V{r}*Master_Employees!R{me_row},0))"  # HRA
        ws[f"Y{r}"] = f"=IF(C{r}=\"\",\"\",Master_Employees!S{me_row})"  # Conveyance
        ws[f"Z{r}"] = f"=IF(C{r}=\"\",\"\",MAX(V{r}-W{r}-X{r}-Y{r},0))"  # Special Allowance (balancing figure)

        # PF: capped at 15000 wage ceiling, only if applicable
        ws[f"AD{r}"] = (f"=IF(OR(C{r}=\"\",Master_Employees!T{me_row}<>\"Yes\"),0,"
                         f"ROUND(MIN(W{r},15000)*0.12,0))")
        ws[f"AF{r}"] = (f"=IF(OR(C{r}=\"\",Master_Employees!T{me_row}<>\"Yes\"),0,"
                         f"ROUND(MIN(W{r},15000)*0.13,0))")

        # ESI: only if applicable AND effective gross <= 21000
        ws[f"AE{r}"] = (f"=IF(OR(C{r}=\"\",Master_Employees!U{me_row}<>\"Yes\",V{r}>21000),0,"
                         f"ROUND(V{r}*0.0075,0))")
        ws[f"AG{r}"] = (f"=IF(OR(C{r}=\"\",Master_Employees!U{me_row}<>\"Yes\",V{r}>21000),0,"
                         f"ROUND(V{r}*0.0325,0))")

        # Professional Tax — Karnataka slab (update Statutory_Compliance sheet if different state)
        ws[f"AH{r}"] = (f"=IF(OR(C{r}=\"\",Master_Employees!V{me_row}<>\"Yes\"),0,"
                         f"IF(V{r}<=15000,0,200))")

        # Net Salary Payable
        ws[f"AI{r}"] = f"=IF(C{r}=\"\",\"\",V{r}-AD{r}-AE{r}-AH{r})"

        for col_letter in SALARY_HEADERS:
            ws[f"{col_letter}{r}"].border = THIN_BORDER

    col_widths = {
        "A": 6, "B": 12, "C": 20, "D": 16, "F": 14, "L": 12, "M": 12, "P": 10, "R": 10,
        "T": 16, "V": 16, "W": 12, "X": 12, "Y": 12, "Z": 14, "AD": 14, "AE": 14,
        "AF": 16, "AG": 14, "AH": 12, "AI": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "C3"
    return sheet_name


def create_leave_tracker_sheet(wb, employee_rows):
    ws = wb.create_sheet(title="Leave_Tracker")
    ws.cell(row=1, column=1, value="Employee Leave Tracker (Annual)").font = TITLE_FONT
    ws.merge_cells("A1:H1")

    headers = ["SlNo", "Employee ID", "Name", "Casual Leave Taken", "Sick Leave Taken",
               "Earned Leave Taken", "LOP Days (YTD)", "Leave Balance (Earned)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, [get_column_letter(c) for c in range(1, len(headers) + 1)])

    for r in range(4, 4 + employee_rows):
        me_row = r - 1
        ws[f"A{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!A{me_row})"
        ws[f"B{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!B{me_row})"
        ws[f"C{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!C{me_row})"
        ws[f"H{r}"] = f"=IF(C{r}=\"\",\"\",15-F{r})"
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = THIN_BORDER

    widths = [8, 14, 22, 18, 16, 18, 16, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "D4"


def create_annual_summary_sheet(wb, fy_label, salary_sheet_names, employee_rows):
    ws = wb.create_sheet(title="Annual_Summary")
    ws.cell(row=1, column=1, value=f"Annual Payroll Summary — FY {fy_label}").font = TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["SlNo", "Employee ID", "Name", "Total Gross Paid",
               "Total Employee PF", "Total Employee ESI", "Total PT", "Total Net Paid"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, [get_column_letter(c) for c in range(1, len(headers) + 1)])

    for r in range(4, 4 + employee_rows):
        me_row = r - 1
        ws[f"A{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!A{me_row})"
        ws[f"B{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!B{me_row})"
        ws[f"C{r}"] = f"=IF(Master_Employees!C{me_row}=\"\",\"\",Master_Employees!C{me_row})"

        sheet_row = r - 1  # corresponding data row in each monthly sheet (offset by header rows: master row3->sheet row3)
        v_terms = "+".join(f"'{s}'!V{sheet_row}" for s in salary_sheet_names)
        pf_terms = "+".join(f"'{s}'!AD{sheet_row}" for s in salary_sheet_names)
        esi_terms = "+".join(f"'{s}'!AE{sheet_row}" for s in salary_sheet_names)
        pt_terms = "+".join(f"'{s}'!AH{sheet_row}" for s in salary_sheet_names)
        net_terms = "+".join(f"'{s}'!AI{sheet_row}" for s in salary_sheet_names)

        ws[f"D{r}"] = f"=IF(C{r}=\"\",\"\",SUM({v_terms}))"
        ws[f"E{r}"] = f"=IF(C{r}=\"\",\"\",SUM({pf_terms}))"
        ws[f"F{r}"] = f"=IF(C{r}=\"\",\"\",SUM({esi_terms}))"
        ws[f"G{r}"] = f"=IF(C{r}=\"\",\"\",SUM({pt_terms}))"
        ws[f"H{r}"] = f"=IF(C{r}=\"\",\"\",SUM({net_terms}))"

        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = THIN_BORDER

    widths = [8, 14, 22, 16, 16, 16, 12, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "D4"


def generate_workbook(company_name, fy_label, employee_rows, output_path=None):
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(TMP_DIR / f"Fracktal_HR_Master_FY{fy_label.replace('-', '')}_{timestamp}.xlsx")

    fy_start_yy = fy_label.split("-")[0][-2:]
    fy_end_yy = fy_label.split("-")[1]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print(f"\n{'='*60}")
    print(f"GENERATING HR/PAYROLL MASTER WORKBOOK (EMPTY TEMPLATE)")
    print(f"Company: {company_name}")
    print(f"FY: {fy_label}")
    print(f"{'='*60}\n")

    create_instructions_sheet(wb, company_name, fy_label, employee_rows)
    print("  ✓ INSTRUCTIONS")

    create_statutory_compliance_sheet(wb, company_name)
    print("  ✓ Statutory_Compliance (placeholders — fill in establishment codes)")

    create_master_employees_sheet(wb, company_name, employee_rows)
    print("  ✓ Master_Employees (empty — no employee rows)")

    salary_sheet_names = []
    for i, mon in enumerate(MONTHS_FY):
        yy = fy_start_yy if i < 9 else fy_end_yy
        label = f"{mon}{yy}"
        name = create_salary_sheet(wb, label, employee_rows)
        salary_sheet_names.append(name)
    print(f"  ✓ 12 monthly Salary Stmt sheets ({salary_sheet_names[0]} … {salary_sheet_names[-1]})")

    create_leave_tracker_sheet(wb, employee_rows)
    print("  ✓ Leave_Tracker")

    create_annual_summary_sheet(wb, fy_label, salary_sheet_names, employee_rows)
    print("  ✓ Annual_Summary")

    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"✓ WORKBOOK SAVED: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)}  |  Employee row capacity: {employee_rows}")
    print(f"  NOTE: Zero employee data included — populate Master_Employees before running payroll.")
    print(f"{'='*60}\n")

    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate an empty HR/Payroll master workbook")
    parser.add_argument("--company", default="Fracktal Works Private Limited")
    parser.add_argument("--fy", default="2026-27", help="Financial year label, e.g. 2026-27")
    parser.add_argument("--employee-rows", type=int, default=50, help="Template capacity (default: 50 rows)")
    parser.add_argument("--output", default=None, help="Output .xlsx path")
    args = parser.parse_args()

    generate_workbook(
        company_name=args.company,
        fy_label=args.fy,
        employee_rows=args.employee_rows,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
