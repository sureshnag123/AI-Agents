#!/usr/bin/env python3
"""
Fracktal Works Private Limited — F&A SOP / Workflow / Flowchart Generator
=========================================================================

Customized for MSME / Private Limited with a lean 2-person Finance team:
  • Head – Finance & Accounts  (Decision-maker, compliance owner)
  • Junior Accountant           (Day-to-day execution)

External stakeholders:
  • Director / CEO              (Final approver for strategic items)
  • Department Heads            (Sales, Engineering, Operations, HR — cross-functional)
  • CA / Tax Consultant         (External — audit, tax advisory)
  • Statutory Auditor           (External — annual audit)

POLICY: *** NO VERBAL APPROVALS / NO VERBAL COMMITMENTS ***
  Every step requires documented approval (email, signed form, ERP workflow, or digital signature).

Covers: P2P (Procure to Pay), O2C (Order to Cash), R2R (Record to Report),
        Payroll, GST, TDS, Expense, Bank Recon, Budget, Payment Approval.
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

# ─── Styles ──────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DECISION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CONTROL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ESCALATION_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
POLICY_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
NO_VERBAL_FILL = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
INDEX_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
RACI_R_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
RACI_A_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
RACI_C_FILL = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
RACI_I_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ─── Fracktal Works Roles ────────────────────────────────────────────────

ROLES = {
    "HEAD_FA": "Head – F&A",
    "JR_ACCT": "Junior Accountant",
    "DIRECTOR": "Director / CEO",
    "DEPT_HEAD": "Dept Head (Requestor)",
    "SALES": "Sales / Project Lead",
    "ENGINEERING": "Engineering Head",
    "OPERATIONS": "Operations Head",
    "HR": "HR / Admin",
    "CA": "CA / Tax Consultant",
    "AUDITOR": "Statutory Auditor",
    "EMPLOYEE": "Employee",
    "VENDOR": "Vendor",
    "CUSTOMER": "Customer",
    "BOARD": "Board of Directors",
}

R = ROLES  # shorthand

# ─── Approval Thresholds (INR) — Fracktal Works ─────────────────────────

DEFAULT_THRESHOLDS = {
    "petty_cash": {
        "limit": 2000,
        "approver": R["HEAD_FA"],
        "description": "Petty cash & minor office expenses",
        "approval_mode": "Signed petty cash voucher"
    },
    "operational_low": {
        "limit": 10000,
        "approver": R["HEAD_FA"],
        "description": "Small vendor payments, office supplies, utilities",
        "approval_mode": "Email approval from Head F&A"
    },
    "operational_high": {
        "limit": 50000,
        "approver": R["HEAD_FA"],
        "description": "Vendor payments, service charges, subscriptions",
        "approval_mode": "Email approval + PO/Invoice on file"
    },
    "major": {
        "limit": 200000,
        "approver": R["DIRECTOR"],
        "description": "Capital expenditure, large vendor contracts",
        "approval_mode": "Director email/signed approval + PO + Invoice"
    },
    "strategic": {
        "limit": 500000,
        "approver": R["DIRECTOR"],
        "description": "Strategic purchases, annual contracts",
        "approval_mode": "Director signed approval + Board awareness"
    },
    "board_level": {
        "limit": None,
        "approver": R["BOARD"],
        "description": "Above ₹5,00,000 — Board resolution required",
        "approval_mode": "Board resolution (signed minutes)"
    },
}

# ───────────────────────────────────────────────────────────────────────────
# NO VERBAL APPROVAL POLICY — embedded as a control in every decision step
# ───────────────────────────────────────────────────────────────────────────
NO_VERBAL_POLICY = (
    "⛔ NO VERBAL APPROVALS. "
    "All approvals must be documented via: email, signed form, ERP workflow approval, "
    "or WhatsApp/message with screenshot saved. Verbal commitments are NOT valid."
)

APPROVAL_TRAIL = (
    "📋 Approval trail required: Approver name, date, amount, purpose. "
    "Filed with supporting documents."
)


# ═══════════════════════════════════════════════════════════════════════════
# WORKFLOW DEFINITIONS — Tailored for Fracktal Works (2-person F&A team)
# ═══════════════════════════════════════════════════════════════════════════

def get_all_workflows():
    return {
        "P2P": p2p_workflow(),
        "O2C": o2c_workflow(),
        "R2R": r2r_workflow(),
        "Expense Reimbursement": expense_workflow(),
        "Payment Approval": payment_approval_workflow(),
        "Payroll": payroll_workflow(),
        "Bank Reconciliation": bank_recon_workflow(),
        "GST Compliance": gst_workflow(),
        "TDS Compliance": tds_workflow(),
        "Budget & MIS": budget_mis_workflow(),
    }


# ─── P2P (Procure to Pay) ───────────────────────────────────────────────

def p2p_workflow():
    return {
        "name": "P2P — Procure to Pay",
        "description": "End-to-end procurement: Purchase Requisition → PO → GRN → Invoice Verification → Payment",
        "steps": [
            {"step": 1, "phase": "Requisition", "role": R["DEPT_HEAD"],
             "action": "Raise Purchase Requisition (PR) with item details, qty, estimated cost, and business justification. Submit via email/ERP to Head F&A.",
             "decision": False, "doc": "Purchase Requisition Form (signed/email)", "control": f"Written PR mandatory. {NO_VERBAL_POLICY}", "sla": "Day 0", "escalation": R["DIRECTOR"]},

            {"step": 2, "phase": "Requisition", "role": R["HEAD_FA"],
             "action": "Review PR: check budget availability, verify business need, confirm vendor is not blacklisted. APPROVE or REJECT in writing.",
             "decision": True, "yes": "Approved — proceed to vendor selection (Step 3)", "no": "Reject with written reason; return to requestor",
             "doc": "Budget Tracker, Vendor Master", "control": f"Written approval on PR. {NO_VERBAL_POLICY}", "sla": "Day 1", "escalation": R["DIRECTOR"]},

            {"step": 3, "phase": "Vendor Selection", "role": R["JR_ACCT"],
             "action": "For new vendors: obtain 3 quotations. For existing vendors: verify rate contract validity. Prepare comparative statement.",
             "decision": True, "yes": "Existing vendor with valid rate — proceed to PO", "no": "New vendor or expired rate — get 3 quotes",
             "doc": "Quotation Comparative Sheet", "control": "Minimum 3 quotes for purchases > ₹10,000. Document reason if single source.", "sla": "Day 1-3", "escalation": R["HEAD_FA"]},

            {"step": 4, "phase": "PO Creation", "role": R["JR_ACCT"],
             "action": "Create Purchase Order in system (Tally/Odoo/Excel PO Register). Include: item, qty, rate, delivery date, payment terms, GSTIN.",
             "decision": False, "doc": "Purchase Order (PO)", "control": "PO must reference approved PR number. No procurement without PO.", "sla": "Day 2-3", "escalation": R["HEAD_FA"]},

            {"step": 5, "phase": "PO Approval", "role": R["HEAD_FA"],
             "action": "Review and approve PO. For amounts > ₹50,000: escalate to Director for written approval before dispatch to vendor.",
             "decision": True, "yes": "PO approved — dispatch to vendor (email/portal)", "no": "Reject — return to Jr Accountant with comments",
             "doc": "Approved PO", "control": f"PO > ₹50K needs Director's email/signed approval. {NO_VERBAL_POLICY}", "sla": "Day 3", "escalation": R["DIRECTOR"]},

            {"step": 6, "phase": "PO Approval", "role": R["DIRECTOR"],
             "action": "[Only if PO > ₹50,000] Review and approve PO via email or signed copy.",
             "decision": True, "yes": "Approved — send PO to vendor", "no": "Reject or request modification",
             "doc": "Email / Signed PO copy", "control": f"Director approval mandatory for PO > ₹50K. {NO_VERBAL_POLICY}", "sla": "Day 3-4", "escalation": "N/A"},

            {"step": 7, "phase": "Goods Receipt", "role": R["DEPT_HEAD"],
             "action": "Receive goods/services. Inspect for quality and quantity. Create Goods Receipt Note (GRN) with sign-off.",
             "decision": True, "yes": "Goods accepted — sign GRN, forward to Finance", "no": "Reject — raise Debit Note / return goods to vendor",
             "doc": "Goods Receipt Note (GRN), Delivery Challan", "control": "GRN must be signed by receiving person. Match GRN qty vs PO qty.", "sla": "On receipt day", "escalation": R["HEAD_FA"]},

            {"step": 8, "phase": "Invoice Verification", "role": R["JR_ACCT"],
             "action": "Receive vendor invoice. Perform 3-way match: PO ↔ GRN ↔ Invoice (item, qty, rate, GST).",
             "decision": True, "yes": "3-way match successful — proceed to booking", "no": "Mismatch found — hold invoice, raise query with vendor/dept head",
             "doc": "Vendor Invoice, PO, GRN", "control": "Mandatory 3-way match. No invoice booking without GRN. Check GSTIN on invoice.", "sla": "Day 1 after invoice receipt", "escalation": R["HEAD_FA"]},

            {"step": 9, "phase": "Invoice Booking", "role": R["JR_ACCT"],
             "action": "Book vendor invoice in Tally/Odoo: correct GL code, cost centre, GST treatment (IGST/CGST/SGST), TDS applicability.",
             "decision": False, "doc": "Tally/Odoo entry, Invoice Register", "control": "GL code per Chart of Accounts. TDS deduction check (Sec 194C/194J/etc.).", "sla": "Day 1-2 after match", "escalation": R["HEAD_FA"]},

            {"step": 10, "phase": "Payment Scheduling", "role": R["HEAD_FA"],
             "action": "Review booked invoices. Add to payment batch per vendor payment terms. Prioritize per cash flow position.",
             "decision": False, "doc": "Payment Schedule / Cash Flow Tracker", "control": "Payment only for matched & booked invoices. No advance without PO.", "sla": "Per payment terms", "escalation": R["DIRECTOR"]},

            {"step": 11, "phase": "Payment Execution", "role": R["JR_ACCT"],
             "action": "Prepare payment batch (NEFT/RTGS/UPI). Enter in banking portal. Submit for authorization.",
             "decision": False, "doc": "Bank Portal, Payment Voucher", "control": "Dual signatory for amounts > ₹25,000. Verify vendor bank details.", "sla": "Per schedule", "escalation": R["HEAD_FA"]},

            {"step": 12, "phase": "Payment Authorization", "role": R["HEAD_FA"],
             "action": "Authorize payment in banking portal. For payments > ₹2,00,000: Director must co-authorize.",
             "decision": True, "yes": "Payment released", "no": "Hold — investigate and resolve",
             "doc": "Bank Authorization, UTR reference", "control": f"Online banking authorization. {NO_VERBAL_POLICY}", "sla": "Same day", "escalation": R["DIRECTOR"]},

            {"step": 13, "phase": "Post-Payment", "role": R["JR_ACCT"],
             "action": "Record payment in books (Tally/Odoo). Update vendor ledger. Note UTR/cheque number. File all documents (PR → PO → GRN → Invoice → Payment proof).",
             "decision": False, "doc": "Tally/Odoo, Bank Statement, Document File", "control": "Complete document chain: PR → PO → GRN → Invoice → Payment. Monthly vendor ledger reconciliation.", "sla": "Day of payment", "escalation": R["HEAD_FA"]},

            {"step": 14, "phase": "Reconciliation", "role": R["JR_ACCT"],
             "action": "Monthly: Reconcile vendor ledger with vendor statement of account. Flag discrepancies.",
             "decision": True, "yes": "Vendor ledger balanced — file confirmation", "no": "Discrepancy — raise with vendor, adjust if needed",
             "doc": "Vendor Reconciliation Statement", "control": "Monthly vendor reconciliation mandatory. Quarterly vendor confirmation.", "sla": "By 10th of next month", "escalation": R["HEAD_FA"]},
        ],
        "raci_roles": [R["DEPT_HEAD"], R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["CA"], R["AUDITOR"]],
        "raci": [
            ["Raise Purchase Requisition",   "R", "", "A", "I", "", ""],
            ["Approve PR (budget check)",    "I", "", "R", "I", "", ""],
            ["Vendor Selection / Quotes",    "C", "R", "A", "", "", ""],
            ["Create Purchase Order",        "", "R", "A", "", "", ""],
            ["Approve PO (≤₹50K)",           "", "I", "R", "I", "", ""],
            ["Approve PO (>₹50K)",           "", "", "C", "R", "", ""],
            ["Goods Receipt (GRN)",          "R", "I", "A", "", "", ""],
            ["3-Way Match (PO-GRN-Invoice)", "", "R", "A", "", "", "I"],
            ["Book Invoice in ERP",          "", "R", "A", "", "", ""],
            ["Schedule Payment",             "", "C", "R", "", "", ""],
            ["Prepare Payment Batch",        "", "R", "A", "", "", ""],
            ["Authorize Payment",            "", "", "R", "A", "", ""],
            ["Record & File Documents",      "", "R", "A", "", "", "I"],
            ["Monthly Vendor Reconciliation","", "R", "A", "", "", "I"],
        ],
    }


# ─── O2C (Order to Cash) ────────────────────────────────────────────────

def o2c_workflow():
    return {
        "name": "O2C — Order to Cash",
        "description": "End-to-end revenue cycle: Customer Order → Delivery → Invoicing → Collection → Receipt → Reconciliation",
        "steps": [
            {"step": 1, "phase": "Order Receipt", "role": R["SALES"],
             "action": "Receive customer order / Purchase Order. Verify terms: item, qty, rate, delivery timeline, payment terms. Share PO copy with Finance.",
             "decision": False, "doc": "Customer PO / Order Confirmation (email)", "control": f"No work to start without written customer PO/email confirmation. {NO_VERBAL_POLICY}", "sla": "Day 0", "escalation": R["HEAD_FA"]},

            {"step": 2, "phase": "Order Review", "role": R["HEAD_FA"],
             "action": "Review customer creditworthiness. For new customers: check GSTIN validity, request advance payment. For existing: check outstanding AR balance.",
             "decision": True, "yes": "Credit approved — proceed with order", "no": "Hold order — request advance or clear past dues first",
             "doc": "Customer Master, AR Aging Report", "control": "New customer: mandatory advance or LC. Existing customer with overdue > 60 days: hold supply.", "sla": "Day 1", "escalation": R["DIRECTOR"]},

            {"step": 3, "phase": "Delivery", "role": R["OPERATIONS"],
             "action": "Fulfil order — manufacture / assemble / ship. Create Delivery Challan / Dispatch Note with customer sign-off.",
             "decision": False, "doc": "Delivery Challan (signed by customer), e-Way Bill (if applicable)", "control": "Delivery challan must have customer acknowledgement (signature/stamp/email).", "sla": "As per order terms", "escalation": R["SALES"]},

            {"step": 4, "phase": "Invoice", "role": R["JR_ACCT"],
             "action": "Generate GST-compliant Sales Invoice within 24 hours of delivery. Match with PO terms (item, qty, rate, GST).",
             "decision": False, "doc": "Tax Invoice, Customer PO, Delivery Challan", "control": "Invoice must match PO terms. Correct HSN/SAC, GSTIN, e-invoice (if applicable).", "sla": "Within 24 hrs of delivery", "escalation": R["HEAD_FA"]},

            {"step": 5, "phase": "Invoice Approval", "role": R["HEAD_FA"],
             "action": "Review and approve invoice before dispatch to customer. Verify GST, TDS applicability, revenue recognition.",
             "decision": True, "yes": "Approved — dispatch to customer", "no": "Correction needed — return to Jr Accountant",
             "doc": "Approved Invoice", "control": f"No invoice dispatch without Head F&A sign-off. {NO_VERBAL_POLICY}", "sla": "Same day", "escalation": R["DIRECTOR"]},

            {"step": 6, "phase": "Invoice Dispatch", "role": R["JR_ACCT"],
             "action": "Send invoice to customer via email (PDF) and upload on customer portal if applicable. Confirm receipt.",
             "decision": False, "doc": "Email confirmation, Portal upload screenshot", "control": "Invoice delivery confirmation required (email read receipt or portal timestamp).", "sla": "Same day as approval", "escalation": R["HEAD_FA"]},

            {"step": 7, "phase": "AR Tracking", "role": R["JR_ACCT"],
             "action": "Enter invoice in AR Aging Tracker. Set payment due date reminder. Monitor aging buckets (Current, 30, 60, 90+ days).",
             "decision": False, "doc": "AR Aging Report / Tracker", "control": "Weekly AR aging review by Head F&A.", "sla": "Same day", "escalation": R["HEAD_FA"]},

            {"step": 8, "phase": "Follow-up", "role": R["JR_ACCT"],
             "action": "Send payment reminder email 7 days before due date. Follow up on due date. Send formal demand at Due+15 days.",
             "decision": True, "yes": "Customer confirms payment date — monitor", "no": "No response — escalate to Head F&A (Step 9)",
             "doc": "Reminder Emails (saved)", "control": "Written follow-up trail. No verbal follow-ups only.", "sla": "Due date", "escalation": R["HEAD_FA"]},

            {"step": 9, "phase": "Escalation", "role": R["HEAD_FA"],
             "action": "Call customer + send formal escalation email. Negotiate payment plan if needed. For > 90 days overdue: involve Director.",
             "decision": True, "yes": "Payment committed — document commitment in email", "no": "No resolution — escalate to Director for legal/credit hold",
             "doc": "Escalation Email, Collection Notes", "control": f"Any payment plan must be documented in writing. {NO_VERBAL_POLICY}", "sla": "Due +15 to +30 days", "escalation": R["DIRECTOR"]},

            {"step": 10, "phase": "Collection", "role": R["JR_ACCT"],
             "action": "Receive payment (bank credit). Match amount with invoice. Check for TDS deducted by customer — collect TDS certificate.",
             "decision": True, "yes": "Full payment received — close invoice", "no": "Partial / short payment — update balance, continue follow-up for remainder",
             "doc": "Bank Statement, Receipt Voucher", "control": "TDS certificate to be collected within 15 days of payment. Short payments investigated.", "sla": "On receipt", "escalation": R["HEAD_FA"]},

            {"step": 11, "phase": "Receipt Booking", "role": R["JR_ACCT"],
             "action": "Book receipt in Tally/Odoo. Issue Receipt/Acknowledgement to customer. Update customer ledger and AR tracker.",
             "decision": False, "doc": "Receipt Voucher, Tally/Odoo entry", "control": "Receipt must match bank credit. Customer ledger updated same day.", "sla": "Day of receipt", "escalation": R["HEAD_FA"]},

            {"step": 12, "phase": "Reconciliation", "role": R["JR_ACCT"],
             "action": "Monthly: Reconcile customer ledger with customer records. Send Statement of Account to customers quarterly.",
             "decision": True, "yes": "Balanced — file confirmation", "no": "Discrepancy — investigate and adjust",
             "doc": "Customer Reconciliation, Statement of Account", "control": "Monthly customer ledger recon. Quarterly SOA to all customers.", "sla": "By 10th of next month", "escalation": R["HEAD_FA"]},
        ],
        "raci_roles": [R["SALES"], R["OPERATIONS"], R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["AUDITOR"]],
        "raci": [
            ["Receive Customer Order",       "R", "", "I", "A", "I", ""],
            ["Credit Check / Approval",      "C", "", "C", "R", "A", ""],
            ["Delivery & Challan",            "C", "R", "", "A", "", ""],
            ["Generate Invoice",             "", "", "R", "A", "", ""],
            ["Approve Invoice",              "", "", "I", "R", "I", ""],
            ["Dispatch Invoice",             "", "", "R", "A", "", ""],
            ["AR Aging Tracking",            "", "", "R", "A", "", "I"],
            ["Payment Follow-up",            "C", "", "R", "A", "", ""],
            ["Escalate Overdue",             "", "", "I", "R", "A", ""],
            ["Receive & Match Payment",      "", "", "R", "A", "", ""],
            ["Book Receipt & Update Ledger", "", "", "R", "A", "", ""],
            ["Monthly Customer Recon",       "", "", "R", "A", "", "I"],
        ],
    }


# ─── R2R (Record to Report) / Month-End Close ───────────────────────────

def r2r_workflow():
    return {
        "name": "R2R — Record to Report (Month-End Close)",
        "description": "Trial Balance → Adjustments → Reconciliation → Financial Statements → MIS → Sign-off",
        "steps": [
            {"step": 1, "phase": "Preparation", "role": R["JR_ACCT"],
             "action": "Ensure ALL transactions for the month are entered — AP invoices, AR invoices, bank entries, JVs. Run preliminary Trial Balance.",
             "decision": False, "doc": "Trial Balance (draft), Transaction Checklist", "control": "Cut-off checklist: all invoices dated within the month must be booked.", "sla": "T+1 (1st working day)", "escalation": R["HEAD_FA"]},

            {"step": 2, "phase": "Preparation", "role": R["JR_ACCT"],
             "action": "Collect pending information from departments: Sales confirmations, expense claims, project costs, inventory counts.",
             "decision": False, "doc": "Department data (emails)", "control": f"Department heads must respond in writing. {NO_VERBAL_POLICY}", "sla": "T+1", "escalation": R["HEAD_FA"]},

            {"step": 3, "phase": "Adjustments", "role": R["JR_ACCT"],
             "action": "Pass accrual entries: rent, salaries, utilities, interest, any unbilled expenses.",
             "decision": False, "doc": "Accrual Schedule, JV entries", "control": "Head F&A to provide accrual list. Accruals > ₹10K need written basis.", "sla": "T+2", "escalation": R["HEAD_FA"]},

            {"step": 4, "phase": "Adjustments", "role": R["JR_ACCT"],
             "action": "Pass depreciation entries (fixed asset register schedule). Pass amortization entries.",
             "decision": False, "doc": "Fixed Asset Register, Depreciation Schedule", "control": "Depreciation method per Companies Act / IT Act as applicable.", "sla": "T+2", "escalation": R["HEAD_FA"]},

            {"step": 5, "phase": "Adjustments", "role": R["JR_ACCT"],
             "action": "Pass prepaid expense adjustments, advance adjustments, and any provision entries.",
             "decision": False, "doc": "Prepaid Schedule, Provision Worksheet", "control": "Each provision must have documented basis (email/contract/estimate).", "sla": "T+2", "escalation": R["HEAD_FA"]},

            {"step": 6, "phase": "Reconciliation", "role": R["JR_ACCT"],
             "action": "Complete Bank Reconciliation Statement (BRS) for ALL bank accounts.",
             "decision": True, "yes": "All items matched or explained — proceed", "no": "Unmatched items > 30 days — escalate to Head F&A",
             "doc": "Bank Reconciliation Statement", "control": "Zero unexplained items target. Items > 30 days must be resolved.", "sla": "T+3", "escalation": R["HEAD_FA"]},

            {"step": 7, "phase": "Reconciliation", "role": R["JR_ACCT"],
             "action": "Reconcile GST Input Credit ledger with books. Reconcile TDS receivable with 26AS/AIS.",
             "decision": True, "yes": "Matched — proceed", "no": "Mismatch — investigate with vendors/CA",
             "doc": "GST Recon, TDS Recon (26AS vs Books)", "control": "ITC mismatch flagged and resolved before GSTR-3B filing.", "sla": "T+3", "escalation": R["HEAD_FA"]},

            {"step": 8, "phase": "Reconciliation", "role": R["JR_ACCT"],
             "action": "Reconcile inter-company accounts (if any), employee advance ledgers, and petty cash.",
             "decision": False, "doc": "Reconciliation Statements", "control": "Physical cash count for petty cash at month-end.", "sla": "T+3", "escalation": R["HEAD_FA"]},

            {"step": 9, "phase": "Review", "role": R["HEAD_FA"],
             "action": "Review adjusted Trial Balance. Generate draft P&L, Balance Sheet, and Cash Flow Statement. Perform analytical review (compare with prior month/year).",
             "decision": True, "yes": "Financials accurate — proceed to MIS", "no": "Corrections needed — return specific items to Jr Accountant",
             "doc": "P&L, Balance Sheet, Cash Flow, TB", "control": "A = L + E check. Revenue/expense reasonableness. Prior period comparison.", "sla": "T+4", "escalation": R["DIRECTOR"]},

            {"step": 10, "phase": "Reporting", "role": R["HEAD_FA"],
             "action": "Prepare MIS report: P&L summary, Balance Sheet summary, Cash Flow, Budget vs Actuals variance, AR/AP aging, key ratios.",
             "decision": False, "doc": "MIS Report (Excel/PDF)", "control": "Variance > 10% from budget must have explanation note.", "sla": "T+5", "escalation": R["DIRECTOR"]},

            {"step": 11, "phase": "Sign-off", "role": R["DIRECTOR"],
             "action": "Review financial package and MIS. Approve or request clarifications. Sign-off on monthly close.",
             "decision": True, "yes": "Approved — month closed. Archive all papers.", "no": "Request revisions — Head F&A to address",
             "doc": "Signed Financial Package", "control": f"Director written sign-off mandatory. {NO_VERBAL_POLICY}", "sla": "T+6", "escalation": "N/A"},

            {"step": 12, "phase": "Archive", "role": R["JR_ACCT"],
             "action": "Lock the period in ERP. Archive all supporting documents (physical file + digital backup). Update checklist as COMPLETE.",
             "decision": False, "doc": "Month-End Checklist (signed)", "control": "Period locked — no backdated entries without written Head F&A approval.", "sla": "T+6", "escalation": R["HEAD_FA"]},
        ],
        "raci_roles": [R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["DEPT_HEAD"], R["CA"], R["AUDITOR"]],
        "raci": [
            ["Cut-off & Preliminary TB",    "R", "A", "", "C", "", ""],
            ["Collect Dept Information",     "R", "A", "", "C", "", ""],
            ["Pass Accrual Entries",         "R", "A", "", "", "", "I"],
            ["Depreciation & Amortization",  "R", "A", "", "", "", "I"],
            ["Prepaid & Provision Entries",  "R", "A", "", "", "", ""],
            ["Bank Reconciliation",          "R", "A", "I", "", "", "I"],
            ["GST / TDS Reconciliation",     "R", "A", "", "", "C", "I"],
            ["Other Reconciliations",        "R", "A", "", "", "", ""],
            ["Review & Generate Financials", "C", "R", "A", "", "", ""],
            ["Prepare MIS",                  "C", "R", "A", "", "", ""],
            ["Director Sign-off",            "", "C", "R", "", "", "I"],
            ["Lock Period & Archive",        "R", "A", "", "", "", ""],
        ],
    }


# ─── Expense Reimbursement ───────────────────────────────────────────────

def expense_workflow():
    return {
        "name": "Expense Reimbursement",
        "description": "Employee claim → Dept Head approval → Finance verification → Payment",
        "steps": [
            {"step": 1, "phase": "Submission", "role": R["EMPLOYEE"],
             "action": "Fill Expense Claim Form with: date, description, amount, business purpose. Attach original bills/receipts. Submit to Dept Head.",
             "decision": False, "doc": "Expense Claim Form + original receipts", "control": f"No claim without original bills. {NO_VERBAL_POLICY}", "sla": "Within 7 days of expense", "escalation": R["DEPT_HEAD"]},

            {"step": 2, "phase": "Approval", "role": R["DEPT_HEAD"],
             "action": "Review claim: verify business purpose, check amount reasonableness, confirm within department budget. APPROVE/REJECT in writing.",
             "decision": True, "yes": "Approved — sign and forward to Finance", "no": "Reject with written reason",
             "doc": "Signed Expense Claim Form", "control": f"Dept Head written approval mandatory. {NO_VERBAL_POLICY}", "sla": "Day 1-2", "escalation": R["HEAD_FA"]},

            {"step": 3, "phase": "Verification", "role": R["JR_ACCT"],
             "action": "Verify receipts: GST invoice validity, amount accuracy, policy compliance (travel limits, meal limits etc.), duplicate check.",
             "decision": True, "yes": "All receipts valid — proceed", "no": "Discrepancy — return to employee with written query",
             "doc": "Expense Policy, Receipt verification checklist", "control": "GST input credit claimed only on valid tax invoices. No self-made bills accepted.", "sla": "Day 2-3", "escalation": R["HEAD_FA"]},

            {"step": 4, "phase": "Approval", "role": R["HEAD_FA"],
             "action": "Final approval: review claim, verify GL coding, approve for payment. For claims > ₹10,000: additional scrutiny.",
             "decision": True, "yes": "Approved for payment", "no": "Hold / partial approval with written reason",
             "doc": "Approved Expense Claim", "control": f"Head F&A written approval. {NO_VERBAL_POLICY}", "sla": "Day 3", "escalation": R["DIRECTOR"]},

            {"step": 5, "phase": "Payment", "role": R["JR_ACCT"],
             "action": "Process reimbursement: book expense in Tally/Odoo, process bank transfer to employee. For petty cash claims < ₹2,000: cash disbursement.",
             "decision": False, "doc": "Tally/Odoo entry, Bank Transfer / Petty Cash Voucher", "control": "Employee signs receipt acknowledgement for cash. Bank transfer: verified account.", "sla": "Day 4-5", "escalation": R["HEAD_FA"]},

            {"step": 6, "phase": "Closure", "role": R["JR_ACCT"],
             "action": "File completed expense claim with all supporting documents. Update expense tracker.",
             "decision": False, "doc": "Filed Expense Dossier", "control": "Complete trail: Claim → Approval → Receipts → Payment proof. Monthly expense summary.", "sla": "Day 5", "escalation": R["HEAD_FA"]},
        ],
        "raci_roles": [R["EMPLOYEE"], R["DEPT_HEAD"], R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["AUDITOR"]],
        "raci": [
            ["Submit Expense Claim",    "R", "I", "", "", "", ""],
            ["Dept Head Approval",      "I", "R", "", "A", "", ""],
            ["Receipt Verification",    "", "", "R", "A", "", "I"],
            ["Head F&A Approval",       "", "", "I", "R", "I", ""],
            ["Process Payment",         "I", "", "R", "A", "", ""],
            ["File & Archive",          "", "", "R", "A", "", "I"],
        ],
    }


# ─── Payment Approval Policy ────────────────────────────────────────────

def payment_approval_workflow():
    return {
        "name": "Payment Approval Policy",
        "description": "Threshold-based approval matrix — every payment must have documented approval",
        "steps": [
            {"step": 1, "phase": "Initiation", "role": "Requestor (Any Dept)",
             "action": "Raise payment request with supporting documents: invoice, PO, contract, or written justification. Submit to Finance (email/form).",
             "decision": False, "doc": "Payment Request + Supporting Docs", "control": f"⛔ NO PAYMENT WITHOUT WRITTEN REQUEST. {NO_VERBAL_POLICY}", "sla": "Day 0", "escalation": R["HEAD_FA"]},

            {"step": 2, "phase": "Verification", "role": R["JR_ACCT"],
             "action": "Verify supporting documents are complete. Check: invoice matches PO, budget available, no duplicate payment. Classify amount tier.",
             "decision": True, "yes": "Documents complete — route per approval matrix", "no": "Incomplete — return with checklist of missing items",
             "doc": "Payment Checklist, Budget Tracker", "control": "Document completeness checklist mandatory before routing.", "sla": "Day 1", "escalation": R["HEAD_FA"]},

            {"step": 3, "phase": "Tier 1: ≤ ₹2,000", "role": R["HEAD_FA"],
             "action": "Approve petty cash expenses up to ₹2,000. Sign petty cash voucher.",
             "decision": True, "yes": "Approved — disburse from petty cash", "no": "Reject with reason",
             "doc": "Signed Petty Cash Voucher", "control": "Petty cash replenishment when balance < ₹5,000. Daily petty cash register.", "sla": "Same day", "escalation": R["DIRECTOR"]},

            {"step": 4, "phase": "Tier 2: ₹2,001–₹50,000", "role": R["HEAD_FA"],
             "action": "Approve operational payments. Verify invoice, PO match, and GL coding. Approve via email or signed form.",
             "decision": True, "yes": "Approved — process bank payment", "no": "Reject / hold with written reason",
             "doc": "Approved Payment Request (email/signed)", "control": f"Head F&A email/signed approval. {NO_VERBAL_POLICY}", "sla": "Day 1", "escalation": R["DIRECTOR"]},

            {"step": 5, "phase": "Tier 3: ₹50,001–₹2,00,000", "role": R["DIRECTOR"],
             "action": "Review and approve payment. Head F&A presents supporting documents. Director approves via email or signed form.",
             "decision": True, "yes": "Approved — process payment", "no": "Reject / defer",
             "doc": "Director's email/signed approval + supporting docs", "control": f"Director written approval mandatory. {NO_VERBAL_POLICY}", "sla": "Day 1-2", "escalation": "N/A"},

            {"step": 6, "phase": "Tier 4: ₹2,00,001–₹5,00,000", "role": R["DIRECTOR"],
             "action": "Review strategic payment with detailed business case. Approve with Board awareness (email to Board members).",
             "decision": True, "yes": "Approved — Director signs + Board informed", "no": "Defer to Board discussion",
             "doc": "Director signed approval + Board notification email", "control": "Board awareness email for transparency. Minutes if discussed.", "sla": "Day 2-3", "escalation": R["BOARD"]},

            {"step": 7, "phase": "Tier 5: > ₹5,00,000", "role": R["BOARD"],
             "action": "Board of Directors approves via Board Resolution. Head F&A presents financial impact assessment.",
             "decision": True, "yes": "Board Resolution passed — proceed", "no": "Rejected — documented in Board minutes",
             "doc": "Board Resolution (signed by all directors)", "control": "Board quorum required. Resolution filed with company records.", "sla": "Next Board meeting / circular resolution", "escalation": "N/A"},

            {"step": 8, "phase": "Execution", "role": R["JR_ACCT"],
             "action": "Prepare payment in banking portal. For > ₹25,000: dual signatory (Jr Accountant prepares, Head F&A authorizes).",
             "decision": False, "doc": "Bank Portal, Payment Voucher", "control": "Maker-Checker: Jr Acct = Maker, Head F&A = Checker/Authorizer.", "sla": "Per terms", "escalation": R["HEAD_FA"]},

            {"step": 9, "phase": "Post-Payment", "role": R["JR_ACCT"],
             "action": "Record payment. File complete approval trail: Request → Approval (email/signed) → Invoice → Payment proof (UTR/cheque copy).",
             "decision": False, "doc": "Complete Payment Dossier", "control": f"Approval trail MUST be on file. {APPROVAL_TRAIL}", "sla": "Same day", "escalation": R["HEAD_FA"]},
        ],
        "raci_roles": ["Requestor", R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["BOARD"], R["AUDITOR"]],
        "raci": [
            ["Raise Payment Request",     "R", "I", "A", "", "", ""],
            ["Verify & Classify",          "", "R", "A", "", "", ""],
            ["Tier 1 (≤₹2K) Petty Cash",  "", "I", "R", "", "", ""],
            ["Tier 2 (₹2K–₹50K)",         "", "I", "R", "I", "", ""],
            ["Tier 3 (₹50K–₹2L)",         "", "I", "C", "R", "", ""],
            ["Tier 4 (₹2L–₹5L)",          "", "", "C", "R", "I", ""],
            ["Tier 5 (>₹5L) Board",        "", "", "C", "C", "R", ""],
            ["Execute Payment",            "", "R", "A", "", "", ""],
            ["Record & File Trail",        "", "R", "A", "", "", "I"],
        ],
    }


# ─── Payroll Processing ─────────────────────────────────────────────────

def payroll_workflow():
    return {
        "name": "Payroll Processing",
        "description": "Attendance → Salary Calculation → Approval → Disbursement → PF/ESI/PT/TDS Filing",
        "steps": [
            {"step": 1, "phase": "Data Collection", "role": R["HR"],
             "action": "Compile attendance, leave records, overtime, and any LOP (Loss of Pay) data. Share with Finance in writing (email/sheet).",
             "decision": False, "doc": "Attendance Sheet (signed by Dept Heads), Leave Register", "control": f"Attendance signed by Dept Heads — not verbal. {NO_VERBAL_POLICY}", "sla": "1st working day", "escalation": R["HEAD_FA"]},

            {"step": 2, "phase": "Data Collection", "role": R["HR"],
             "action": "Update changes: new joiners, exits, salary revisions, bonus/incentives. Provide supporting documents (letters).",
             "decision": False, "doc": "Appointment Letters, Revision Letters, Exit Letters", "control": "Every change backed by signed letter. No payroll changes on verbal instruction.", "sla": "1st working day", "escalation": R["HEAD_FA"]},

            {"step": 3, "phase": "Calculation", "role": R["JR_ACCT"],
             "action": "Calculate gross salary, deductions: PF (12% employee + 12% employer), ESI (if applicable), PT, TDS (as per Sec 192). Calculate net salary.",
             "decision": False, "doc": "Payroll Register / Salary Sheet", "control": "Statutory rates verified. Cross-check with previous month for anomalies.", "sla": "2nd–3rd of month", "escalation": R["HEAD_FA"]},

            {"step": 4, "phase": "Review", "role": R["HEAD_FA"],
             "action": "Review payroll register: verify net salary, check unusual variances, verify statutory deductions, confirm total payout.",
             "decision": True, "yes": "Payroll approved — proceed to payment", "no": "Corrections needed — return to Jr Accountant",
             "doc": "Approved Payroll Register (signed)", "control": f"Head F&A signed approval on payroll register. {NO_VERBAL_POLICY}", "sla": "3rd–4th of month", "escalation": R["DIRECTOR"]},

            {"step": 5, "phase": "Approval", "role": R["DIRECTOR"],
             "action": "Final payroll approval: review total payout, approve bank transfer. Sign or email approval.",
             "decision": True, "yes": "Authorized — process salary transfer", "no": "Hold — investigate flagged items",
             "doc": "Director's email/signed approval", "control": f"Director written approval for salary disbursement. {NO_VERBAL_POLICY}", "sla": "4th–5th of month", "escalation": "N/A"},

            {"step": 6, "phase": "Disbursement", "role": R["JR_ACCT"],
             "action": "Upload salary file to banking portal. Process NEFT/RTGS to employee accounts. Head F&A authorizes the batch.",
             "decision": False, "doc": "Bank Salary File, Bank Authorization", "control": "Maker-Checker: Jr Acct uploads, Head F&A authorizes. Verify account numbers.", "sla": "5th–7th of month", "escalation": R["HEAD_FA"]},

            {"step": 7, "phase": "Post-Disbursement", "role": R["JR_ACCT"],
             "action": "Generate payslips (PDF). Distribute to employees via email. Book salary entry in Tally/Odoo.",
             "decision": False, "doc": "Payslips (PDF), Tally/Odoo JV", "control": "Payslip matches actual credit to employee. Email delivery confirmation.", "sla": "7th of month", "escalation": R["HEAD_FA"]},

            {"step": 8, "phase": "Statutory Filing", "role": R["JR_ACCT"],
             "action": "Deposit PF contribution (employee + employer) via EPFO portal. Generate ECR and pay challan.",
             "decision": False, "doc": "ECR, PF Challan (TRRN)", "control": "PF deposit by 15th of month. ECR verified by Head F&A before submission.", "sla": "By 15th of month", "escalation": R["HEAD_FA"]},

            {"step": 9, "phase": "Statutory Filing", "role": R["JR_ACCT"],
             "action": "Deposit ESI contribution (if applicable). File online via ESIC portal.",
             "decision": False, "doc": "ESI Challan", "control": "ESI deposit by 15th. Applicable if employee gross ≤ ₹21,000/month.", "sla": "By 15th of month", "escalation": R["HEAD_FA"]},

            {"step": 10, "phase": "Statutory Filing", "role": R["JR_ACCT"],
             "action": "Deposit Professional Tax. File monthly/quarterly return on Karnataka PT portal.",
             "decision": False, "doc": "PT Challan, PT Return", "control": "PT rates as per Karnataka schedule. Monthly remittance.", "sla": "By 20th of month", "escalation": R["HEAD_FA"]},

            {"step": 11, "phase": "Statutory Filing", "role": R["JR_ACCT"],
             "action": "Deposit TDS on salaries (Sec 192). Prepare quarterly TDS return (Form 24Q) with CA support.",
             "decision": False, "doc": "TDS Challan (Form 26QB), Form 24Q", "control": "TDS deposit by 7th of next month. Quarterly return with CA review.", "sla": "7th of next month / quarterly", "escalation": R["HEAD_FA"]},

            {"step": 12, "phase": "Reconciliation", "role": R["HEAD_FA"],
             "action": "Monthly: reconcile payroll expenditure with GL. Quarterly: reconcile PF/ESI with portal records. Annual: Form 16 issuance.",
             "decision": False, "doc": "Payroll Recon, PF/ESI Recon, Form 16", "control": "Any variance > ₹500 investigated. Form 16 issued by June 15.", "sla": "Monthly / Quarterly / Annual", "escalation": R["CA"]},
        ],
        "raci_roles": [R["HR"], R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["CA"], R["AUDITOR"]],
        "raci": [
            ["Compile Attendance & Leave",    "R", "", "A", "", "", ""],
            ["Update Joiners/Exits/Revisions","R", "", "A", "", "", ""],
            ["Calculate Salary & Deductions", "", "R", "A", "", "", ""],
            ["Review Payroll Register",       "", "I", "R", "I", "", ""],
            ["Director Approval",             "", "", "C", "R", "", ""],
            ["Process Salary Transfer",       "", "R", "A", "", "", ""],
            ["Generate & Send Payslips",      "I", "R", "A", "", "", ""],
            ["Deposit PF (ECR)",              "", "R", "A", "", "C", "I"],
            ["Deposit ESI",                   "", "R", "A", "", "", "I"],
            ["Deposit Professional Tax",      "", "R", "A", "", "", ""],
            ["Deposit TDS / File 24Q",        "", "R", "A", "", "R", "I"],
            ["Monthly/Quarterly Recon",       "", "C", "R", "", "C", "I"],
        ],
    }


# ─── Bank Reconciliation ────────────────────────────────────────────────

def bank_recon_workflow():
    return {
        "name": "Bank Reconciliation",
        "description": "Bank statement ↔ Book ledger matching → Exception handling → Sign-off",
        "steps": [
            {"step": 1, "phase": "Preparation", "role": R["JR_ACCT"],
             "action": "Download bank statements for ALL bank accounts (Current A/c, OD, FD, any other). Export cash/bank ledger from Tally/Odoo.",
             "decision": False, "doc": "Bank Statements (PDF/CSV), Bank Ledger from ERP", "control": "All accounts must be reconciled — no account skipped.", "sla": "Daily (ideal) / Weekly (minimum)", "escalation": R["HEAD_FA"]},

            {"step": 2, "phase": "Matching", "role": R["JR_ACCT"],
             "action": "Match bank entries with book entries. Auto-match by amount/date/ref. Manual match for unmatched items.",
             "decision": True, "yes": "All matched — proceed to sign-off", "no": "Unmatched items — investigate (Step 3)",
             "doc": "BRS Working Sheet", "control": "Transaction-level matching. UTR / cheque numbers must match.", "sla": "T+1", "escalation": R["HEAD_FA"]},

            {"step": 3, "phase": "Exception Handling", "role": R["JR_ACCT"],
             "action": "Investigate unmatched items: timing differences (cheques issued not cleared), bank charges, interest, direct debits/credits not yet recorded.",
             "decision": True, "yes": "Timing difference — document and carry forward", "no": "Error/omission — pass correction entry",
             "doc": "Exception Report", "control": "Unmatched items > 15 days must be escalated to Head F&A in writing.", "sla": "T+2", "escalation": R["HEAD_FA"]},

            {"step": 4, "phase": "Adjustment", "role": R["JR_ACCT"],
             "action": "Pass adjustment entries in books: bank charges, interest credited, direct debits, bounced cheques, etc.",
             "decision": False, "doc": "Journal Vouchers", "control": "Head F&A approval for adjustment entries > ₹5,000.", "sla": "T+2", "escalation": R["HEAD_FA"]},

            {"step": 5, "phase": "Review", "role": R["HEAD_FA"],
             "action": "Review BRS: verify all outstanding items are genuine timing differences. Challenge old items. Sign-off.",
             "decision": True, "yes": "BRS approved — sign and archive", "no": "Further investigation required",
             "doc": "Reviewed & Signed BRS", "control": f"Written sign-off by Head F&A. Outstanding > 30 days require explanation. {NO_VERBAL_POLICY}", "sla": "T+3", "escalation": R["DIRECTOR"]},

            {"step": 6, "phase": "Archive", "role": R["JR_ACCT"],
             "action": "Archive signed BRS with bank statement copies. Update BRS tracker (running file).",
             "decision": False, "doc": "Signed BRS + Bank Statements (filed)", "control": "Monthly BRS file maintained. Available for audit at any time.", "sla": "T+3", "escalation": R["HEAD_FA"]},
        ],
        "raci_roles": [R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["AUDITOR"]],
        "raci": [
            ["Download Statements & Ledger", "R", "A", "", ""],
            ["Match Entries",                "R", "A", "", ""],
            ["Investigate Exceptions",       "R", "A", "I", ""],
            ["Pass Adjustments",             "R", "A", "", "I"],
            ["Review & Sign-off BRS",        "I", "R", "I", "I"],
            ["Archive BRS",                  "R", "A", "", "I"],
        ],
    }


# ─── GST Compliance ─────────────────────────────────────────────────────

def gst_workflow():
    return {
        "name": "GST Compliance",
        "description": "Invoice compliance → GSTR-1 → GSTR-3B → ITC Recon → Annual Return → Audit",
        "steps": [
            {"step": 1, "phase": "Ongoing", "role": R["JR_ACCT"],
             "action": "Ensure ALL sales invoices are GST-compliant: correct GSTIN, HSN/SAC, tax rate, e-invoice generation (if turnover > ₹5 Cr).",
             "decision": False, "doc": "Sales Invoice Register", "control": "Invoice checklist: GSTIN, HSN, tax rate, POS, e-invoice QR. Dept heads must provide HSN/SAC.", "sla": "Ongoing — same day as transaction", "escalation": R["HEAD_FA"]},

            {"step": 2, "phase": "Ongoing", "role": R["JR_ACCT"],
             "action": "Verify ALL purchase invoices have valid GSTIN, correct tax rate, and are filed by vendor in GSTR-1 (check in GSTR-2A/2B).",
             "decision": True, "yes": "Invoice appears in 2A/2B — claim ITC", "no": "Missing in 2A/2B — follow up with vendor in writing",
             "doc": "Purchase Register, GSTR-2A/2B download", "control": f"Monthly ITC reconciliation (Books vs 2A/2B). Written follow-up with vendors. {NO_VERBAL_POLICY}", "sla": "Monthly — by 15th", "escalation": R["HEAD_FA"]},

            {"step": 3, "phase": "GSTR-1 Filing", "role": R["JR_ACCT"],
             "action": "Prepare GSTR-1 data: B2B invoices, B2C summary, credit/debit notes, advances, HSN summary. Upload to GST portal.",
             "decision": False, "doc": "GSTR-1 working file, GST Portal", "control": "Cross-check GSTR-1 with Sales Register from Tally/Odoo. Zero mismatch target.", "sla": "By 11th of next month", "escalation": R["HEAD_FA"]},

            {"step": 4, "phase": "GSTR-1 Filing", "role": R["HEAD_FA"],
             "action": "Review GSTR-1 data before filing. Verify totals match books. Approve filing.",
             "decision": True, "yes": "Approved — file GSTR-1", "no": "Corrections needed — return to Jr Accountant",
             "doc": "GSTR-1 Review Checklist", "control": f"Head F&A must approve GSTR-1 before filing. {NO_VERBAL_POLICY}", "sla": "By 10th (1 day before deadline)", "escalation": R["CA"]},

            {"step": 5, "phase": "GSTR-3B Filing", "role": R["JR_ACCT"],
             "action": "Prepare GSTR-3B: output tax liability, ITC claimed (matched with 2B), tax payable. Calculate GST payment.",
             "decision": False, "doc": "GSTR-3B working sheet", "control": "ITC claimed ≤ ITC available in 2B. Excess ITC = risk.", "sla": "By 18th of next month", "escalation": R["HEAD_FA"]},

            {"step": 6, "phase": "GSTR-3B Filing", "role": R["HEAD_FA"],
             "action": "Review GSTR-3B data. Verify ITC reconciliation. Approve filing and authorize GST payment.",
             "decision": True, "yes": "Approved — file GSTR-3B and pay GST", "no": "Review ITC mismatch — consult CA if needed",
             "doc": "Approved GSTR-3B, Payment Challan", "control": f"Head F&A approval before filing. {NO_VERBAL_POLICY}", "sla": "By 19th (1 day before deadline)", "escalation": R["CA"]},

            {"step": 7, "phase": "GSTR-3B Filing", "role": R["JR_ACCT"],
             "action": "File GSTR-3B on portal. Make GST payment via net banking. Save acknowledgement.",
             "decision": False, "doc": "GSTR-3B filed (ARN), GST Challan", "control": "Payment confirmation screenshot. ARN saved.", "sla": "By 20th of next month", "escalation": R["HEAD_FA"]},

            {"step": 8, "phase": "e-Way Bill", "role": R["JR_ACCT"],
             "action": "Generate e-Way Bill for all consignments > ₹50,000 BEFORE dispatch. Ensure validity period matches transit time.",
             "decision": False, "doc": "e-Way Bill (printed copy with goods)", "control": "e-Way Bill mandatory before goods movement. Dept heads must inform Finance of dispatches.", "sla": "Before dispatch", "escalation": R["HEAD_FA"]},

            {"step": 9, "phase": "Annual", "role": R["HEAD_FA"],
             "action": "Prepare GSTR-9 (Annual Return) and GSTR-9C (Reconciliation Statement) with CA support. Reconcile monthly filings with annual books.",
             "decision": False, "doc": "GSTR-9, GSTR-9C, Reconciliation workings", "control": "Annual return reconciled with audited financials. CA review mandatory.", "sla": "By 31st December", "escalation": R["CA"]},

            {"step": 10, "phase": "Reconciliation", "role": R["HEAD_FA"],
             "action": "Quarterly: reconcile GST credit ledger (portal) with books. Reconcile output tax with sales register.",
             "decision": True, "yes": "Balanced — file reconciliation note", "no": "Mismatch — investigate, adjust, consult CA",
             "doc": "GST Reconciliation Report", "control": "Quarterly recon prevents year-end surprises. Document all adjustments.", "sla": "Quarterly", "escalation": R["CA"]},
        ],
        "raci_roles": [R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["DEPT_HEAD"], R["CA"], R["AUDITOR"]],
        "raci": [
            ["Sales Invoice Compliance",  "R", "A", "", "C", "", ""],
            ["Purchase ITC Verification", "R", "A", "", "", "C", ""],
            ["Prepare GSTR-1",            "R", "A", "", "", "", ""],
            ["Review & Approve GSTR-1",   "I", "R", "", "", "C", ""],
            ["Prepare GSTR-3B",           "R", "A", "", "", "", ""],
            ["Review & Approve GSTR-3B",  "I", "R", "I", "", "C", ""],
            ["File GSTR-3B & Pay GST",    "R", "A", "", "", "", ""],
            ["Generate e-Way Bills",      "R", "A", "", "C", "", ""],
            ["Annual Return (GSTR-9/9C)", "C", "R", "I", "", "R", "R"],
            ["Quarterly GST Recon",       "C", "R", "", "", "C", "I"],
        ],
    }


# ─── TDS Compliance ─────────────────────────────────────────────────────

def tds_workflow():
    return {
        "name": "TDS Compliance",
        "description": "TDS deduction at source → Deposit → Quarterly returns → Form 16/16A → Annual reconciliation",
        "steps": [
            {"step": 1, "phase": "Identification", "role": R["JR_ACCT"],
             "action": "At every payment: identify if TDS is applicable (Sec 194C/194J/194I/194H/192/etc.). Check payee PAN, nature of payment, threshold limits.",
             "decision": True, "yes": "TDS applicable — deduct at prescribed rate", "no": "TDS not applicable — document reason (threshold / exemption / 15CA-CB)",
             "doc": "TDS Applicability Checklist, PAN verification", "control": "TDS checklist for every vendor/employee payment. PAN verified on IT portal.", "sla": "At each payment", "escalation": R["HEAD_FA"]},

            {"step": 2, "phase": "Deduction", "role": R["JR_ACCT"],
             "action": "Deduct TDS at correct rate. Record TDS entry in Tally/Odoo. Note: section, rate, PAN, amount in TDS register.",
             "decision": False, "doc": "TDS Register, Tally/Odoo entry", "control": "TDS register maintained with: date, payee, PAN, section, rate, amount. Lower deduction certificate checked.", "sla": "At each payment", "escalation": R["HEAD_FA"]},

            {"step": 3, "phase": "Deposit", "role": R["JR_ACCT"],
             "action": "Deposit TDS to government via NSDL/TIN portal. Generate challan (Form 26QB). Separate challans for different sections.",
             "decision": False, "doc": "TDS Challan (CIN number), Bank payment receipt", "control": "TDS deposit by 7th of next month (30th April for March). Verify CIN on OLTAS.", "sla": "By 7th of next month", "escalation": R["HEAD_FA"]},

            {"step": 4, "phase": "Quarterly Return", "role": R["JR_ACCT"],
             "action": "Prepare quarterly TDS return data: Form 24Q (salary), Form 26Q (non-salary). Compile all challans and deductee details.",
             "decision": False, "doc": "TDS Return Working Sheet", "control": "Challan total = Sum of deductee-wise TDS. PAN validation for all deductees.", "sla": "Within 15 days of quarter end", "escalation": R["HEAD_FA"]},

            {"step": 5, "phase": "Quarterly Return", "role": R["HEAD_FA"],
             "action": "Review TDS return data. Verify with CA. Approve filing.",
             "decision": True, "yes": "Approved — file TDS return", "no": "Corrections needed — return to Jr Accountant",
             "doc": "Reviewed TDS Return", "control": f"Head F&A and CA review before filing. {NO_VERBAL_POLICY}", "sla": "5 days before due date", "escalation": R["CA"]},

            {"step": 6, "phase": "Quarterly Return", "role": R["CA"],
             "action": "File TDS return on TRACES portal. Download provisional receipts. Verify Form 26AS reflects deductions.",
             "decision": False, "doc": "TDS Return Filed (Token No.), Form 26AS", "control": "Verify 26AS within 15 days of filing. Corrections filed if mismatches found.", "sla": "31 Jul / 31 Oct / 31 Jan / 31 May", "escalation": R["HEAD_FA"]},

            {"step": 7, "phase": "Certificate", "role": R["JR_ACCT"],
             "action": "Generate TDS certificates: Form 16 (salary — annual), Form 16A (non-salary — quarterly). Download from TRACES.",
             "decision": False, "doc": "Form 16, Form 16A", "control": "Form 16 by June 15. Form 16A within 15 days of return filing.", "sla": "Per statutory deadline", "escalation": R["HEAD_FA"]},

            {"step": 8, "phase": "Distribution", "role": R["JR_ACCT"],
             "action": "Distribute TDS certificates to employees (Form 16) and vendors (Form 16A) via email. Obtain acknowledgement.",
             "decision": False, "doc": "Email with TDS certificates", "control": "Delivery confirmation (email read receipt). Maintain distribution register.", "sla": "Within 5 days of generation", "escalation": R["HEAD_FA"]},

            {"step": 9, "phase": "Reconciliation", "role": R["HEAD_FA"],
             "action": "Annual: Reconcile TDS deducted vs deposited vs returned. Verify Form 26AS matches for all deductees. Resolve any demand notices.",
             "decision": True, "yes": "All reconciled — close for the year", "no": "Mismatch — file correction return with CA",
             "doc": "TDS Annual Reconciliation, Form 26AS verification", "control": "Any TDS demand notices addressed within 30 days. Correction returns filed promptly.", "sla": "Annual — post return filing", "escalation": R["CA"]},
        ],
        "raci_roles": [R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["CA"], R["AUDITOR"]],
        "raci": [
            ["Identify TDS Applicability",  "R", "A", "", "C", ""],
            ["Deduct & Record TDS",         "R", "A", "", "", ""],
            ["Deposit TDS (Challan)",       "R", "A", "", "", "I"],
            ["Prepare Quarterly Return",    "R", "A", "", "C", ""],
            ["Review & Approve Return",     "I", "R", "", "C", ""],
            ["File Return on TRACES",       "I", "A", "", "R", ""],
            ["Generate Form 16/16A",        "R", "A", "", "", ""],
            ["Distribute Certificates",     "R", "A", "", "", ""],
            ["Annual TDS Reconciliation",   "C", "R", "", "C", "I"],
        ],
    }


# ─── Budget & MIS ───────────────────────────────────────────────────────

def budget_mis_workflow():
    return {
        "name": "Budget & MIS Reporting",
        "description": "Annual budget → Monthly tracking → Variance analysis → MIS to management",
        "steps": [
            {"step": 1, "phase": "Budget Preparation", "role": R["HEAD_FA"],
             "action": "Issue budget preparation guidelines to all departments: timeline, format, assumptions. Provide previous year actuals as reference.",
             "decision": False, "doc": "Budget Circular (email), Budget Template", "control": f"Written circular to all Dept Heads. {NO_VERBAL_POLICY}", "sla": "Q3 (Oct-Nov)", "escalation": R["DIRECTOR"]},

            {"step": 2, "phase": "Budget Preparation", "role": R["DEPT_HEAD"],
             "action": "Prepare department budget: headcount, capex, opex, project costs. Submit in standard format to Head F&A.",
             "decision": False, "doc": "Department Budget Sheet (submitted via email)", "control": f"Written submission — no verbal estimates. {NO_VERBAL_POLICY}", "sla": "4 weeks from circular", "escalation": R["HEAD_FA"]},

            {"step": 3, "phase": "Consolidation", "role": R["HEAD_FA"],
             "action": "Consolidate all department budgets. Add company-level items: taxes, interest, depreciation. Prepare consolidated budget with revenue forecast.",
             "decision": False, "doc": "Consolidated Company Budget", "control": "Revenue forecast from Sales team (written). Cross-check for completeness.", "sla": "1 week after collection", "escalation": R["DIRECTOR"]},

            {"step": 4, "phase": "Review", "role": R["DIRECTOR"],
             "action": "Review consolidated budget with Head F&A and Dept Heads. Approve or request revisions.",
             "decision": True, "yes": "Budget approved — finalize for FY", "no": "Revisions needed — specific feedback in writing",
             "doc": "Budget Review Meeting Minutes, Approved Budget", "control": f"Director written approval on final budget. {NO_VERBAL_POLICY}", "sla": "Before FY start", "escalation": "N/A"},

            {"step": 5, "phase": "Monitoring", "role": R["JR_ACCT"],
             "action": "Monthly: Extract actuals from Tally/Odoo. Prepare Budget vs Actuals comparison sheet.",
             "decision": False, "doc": "Budget vs Actuals Sheet", "control": "Actuals extracted after month-end close. Consistent GL mapping.", "sla": "By T+5 (after month close)", "escalation": R["HEAD_FA"]},

            {"step": 6, "phase": "Monitoring", "role": R["HEAD_FA"],
             "action": "Analyze variances. For variances > ±10%: request written explanation from Dept Head. Prepare MIS report.",
             "decision": True, "yes": "Variances within tolerance — proceed to MIS", "no": "Material variance — Dept Head must provide written justification",
             "doc": "Variance Analysis, MIS Report", "control": f"Variance > 10% = mandatory written explanation from Dept Head. {NO_VERBAL_POLICY}", "sla": "By T+6", "escalation": R["DIRECTOR"]},

            {"step": 7, "phase": "MIS Reporting", "role": R["HEAD_FA"],
             "action": "Compile monthly MIS package: P&L, Balance Sheet, Cash Flow, Budget vs Actuals, AR/AP Aging, Key Ratios, Cash Position.",
             "decision": False, "doc": "MIS Report (Excel/PDF)", "control": "Standard MIS template used every month for consistency.", "sla": "By T+7", "escalation": R["DIRECTOR"]},

            {"step": 8, "phase": "Presentation", "role": R["HEAD_FA"],
             "action": "Present MIS to Director. Discuss key variances, cash flow outlook, and action items.",
             "decision": False, "doc": "MIS Presentation, Action Items List (documented)", "control": f"Meeting minutes / action items documented. {NO_VERBAL_POLICY}", "sla": "By T+8", "escalation": "N/A"},
        ],
        "raci_roles": [R["DEPT_HEAD"], R["JR_ACCT"], R["HEAD_FA"], R["DIRECTOR"], R["AUDITOR"]],
        "raci": [
            ["Issue Budget Guidelines",    "", "", "R", "A", ""],
            ["Prepare Dept Budgets",       "R", "", "A", "", ""],
            ["Consolidate Budget",         "", "C", "R", "A", ""],
            ["Review & Approve Budget",    "C", "", "C", "R", ""],
            ["Extract Monthly Actuals",    "", "R", "A", "", ""],
            ["Variance Analysis",          "C", "", "R", "A", ""],
            ["Compile MIS Report",         "", "C", "R", "A", ""],
            ["Present MIS to Director",    "", "", "R", "A", "I"],
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL GENERATION
# ═══════════════════════════════════════════════════════════════════════════

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_cell(cell, wrap=True, border=True, center=False):
    if wrap:
        cell.alignment = WRAP_ALIGN if not center else CENTER_ALIGN
    if border:
        cell.border = THIN_BORDER


def create_policy_sheet(wb):
    """Create the 'No Verbal Approval' policy overview sheet."""
    ws = wb.create_sheet(title="⛔ Policy_NoVerbal")

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:D1")
    ws.cell(row=2, column=1, value="FINANCE & ACCOUNTS — STANDING POLICY").font = Font(bold=True, size=13, color="DC3545")
    ws.merge_cells("A2:D2")
    ws.cell(row=3, column=1, value=f"Effective: {datetime.now().strftime('%d %B %Y')}   |   Approved by: Director / CEO").font = Font(italic=True, color="666666")
    ws.merge_cells("A3:D3")

    policies = [
        ("⛔ NO VERBAL APPROVALS",
         "All financial approvals — purchase, payment, expense, salary, invoice — MUST be documented in writing.\n"
         "Acceptable approval modes: Email, Signed form, ERP workflow, WhatsApp message (screenshot saved).\n"
         "Verbal approvals are NOT valid and will not be honored by the Finance team."),

        ("⛔ NO VERBAL COMMITMENTS",
         "No team member may commit to any financial obligation (vendor payment terms, pricing, discounts, salaries, contracts) verbally.\n"
         "All commitments must be in writing (email / signed document) and approved by the appropriate authority."),

        ("📋 DOCUMENT TRAIL — MANDATORY",
         "Every financial transaction must have a complete audit trail:\n"
         "• Purchase: PR → PO → GRN → Invoice → Payment Proof\n"
         "• Sale: Customer PO → Delivery Challan → Invoice → Receipt\n"
         "• Expense: Claim Form → Approval → Receipts → Payment\n"
         "• Payroll: Attendance → Payroll Register → Approval → Bank Transfer\n"
         "Missing documents = payment on hold until complete."),

        ("🔑 MAKER-CHECKER PRINCIPLE",
         "Junior Accountant = Maker (prepares transactions)\n"
         "Head – F&A = Checker/Authorizer (reviews and approves)\n"
         "Director = Final authority for payments > ₹50,000\n"
         "No single person can initiate AND approve a payment."),

        ("📅 DEADLINES — NON-NEGOTIABLE",
         "• GST (GSTR-1): 11th of next month\n"
         "• GST (GSTR-3B + payment): 20th of next month\n"
         "• TDS deposit: 7th of next month\n"
         "• PF/ESI deposit: 15th of month\n"
         "• Professional Tax: 20th of month\n"
         "• Payroll disbursement: 7th of month\n"
         "• Month-end close: T+6 working days\n"
         "• MIS to Director: T+8 working days"),

        ("⚠️ DEPARTMENT RESPONSIBILITIES",
         "All departments linked to finance must:\n"
         "• Submit Purchase Requisitions IN WRITING before any procurement\n"
         "• Sign Goods Receipt Notes upon receiving goods/services\n"
         "• Submit expense claims with ORIGINAL BILLS within 7 days\n"
         "• Provide attendance data SIGNED by Dept Head by 1st of month\n"
         "• Respond to finance queries within 2 working days (in writing)\n"
         "• Submit budget proposals in standard format within deadline"),
    ]

    row = 5
    for title, description in policies:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color="DC3545")
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        row += 1
        ws.cell(row=row, column=1, value=description).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 90
        row += 2

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25


def create_flowchart_sheet(wb, workflow):
    name = workflow["name"]
    sheet_name = name[:28] + "_WF" if len(name) > 28 else name + "_WF"
    # Clean sheet name (remove invalid chars)
    sheet_name = sheet_name.replace("—", "-").replace("/", "-")[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Title row
    ws.cell(row=1, column=1, value=workflow["name"]).font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells("A1:K1")
    desc = workflow.get("description", "")
    ws.cell(row=2, column=1, value=desc).font = Font(italic=True, color="666666", size=10)
    ws.merge_cells("A2:K2")

    headers = [
        "Step #", "Phase", "Responsible Role", "Action / Task",
        "Decision?", "If Yes →", "If No →",
        "Document / System", "Control Point / Policy", "SLA / Timeline", "Escalation To"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))

    for i, step in enumerate(workflow["steps"], start=5):
        ws.cell(row=i, column=1, value=step["step"])
        ws.cell(row=i, column=2, value=step["phase"])
        ws.cell(row=i, column=3, value=step["role"])
        ws.cell(row=i, column=4, value=step["action"])
        ws.cell(row=i, column=5, value="YES" if step["decision"] else "")
        ws.cell(row=i, column=6, value=step.get("yes", ""))
        ws.cell(row=i, column=7, value=step.get("no", ""))
        ws.cell(row=i, column=8, value=step.get("doc", ""))
        ws.cell(row=i, column=9, value=step.get("control", ""))
        ws.cell(row=i, column=10, value=step.get("sla", ""))
        ws.cell(row=i, column=11, value=step.get("escalation", ""))

        for col in range(1, 12):
            cell = ws.cell(row=i, column=col)
            style_cell(cell, center=(col in [1, 5]))
            ws.row_dimensions[i].height = 45

        # Highlight decision rows
        if step["decision"]:
            for col in range(1, 12):
                ws.cell(row=i, column=col).fill = DECISION_FILL

        # Highlight control column
        if step.get("control"):
            ws.cell(row=i, column=9).fill = CONTROL_FILL

        # Highlight NO VERBAL in red if present
        ctrl = step.get("control", "")
        if "NO VERBAL" in ctrl:
            ws.cell(row=i, column=9).font = Font(bold=True, size=10, color="DC3545")

    col_widths = [7, 14, 22, 50, 10, 35, 35, 35, 45, 18, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"
    return sheet_name


def create_raci_sheet(wb, workflow):
    name = workflow["name"]
    sheet_name = name[:26] + "_RACI" if len(name) > 26 else name + "_RACI"
    sheet_name = sheet_name.replace("—", "-").replace("/", "-")[:31]
    ws = wb.create_sheet(title=sheet_name)

    roles = workflow.get("raci_roles", [])
    raci_data = workflow.get("raci", [])

    headers = ["Process Step"] + roles
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    raci_fills = {"R": RACI_R_FILL, "A": RACI_A_FILL, "C": RACI_C_FILL, "I": RACI_I_FILL}
    raci_fonts = {
        "R": Font(bold=True, color="FFFFFF"),
        "A": Font(bold=True, color="FFFFFF"),
        "C": Font(bold=True, color="FFFFFF"),
        "I": Font(bold=True, color="000000"),
    }

    for i, row_data in enumerate(raci_data, start=2):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            style_cell(cell, center=(col > 1))
            if col > 1 and value in raci_fills:
                cell.fill = raci_fills[value]
                cell.font = raci_fonts[value]

    ws.column_dimensions["A"].width = 35
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Legend
    lr = len(raci_data) + 4
    ws.cell(row=lr, column=1, value="RACI Legend:").font = Font(bold=True)
    for j, (code, desc, fill) in enumerate([
        ("R", "Responsible — Does the work", RACI_R_FILL),
        ("A", "Accountable — Owns the outcome (approves)", RACI_A_FILL),
        ("C", "Consulted — Provides input before action", RACI_C_FILL),
        ("I", "Informed — Kept in the loop after action", RACI_I_FILL),
    ]):
        r = lr + 1 + j
        c = ws.cell(row=r, column=1, value=code)
        c.fill = fill
        c.font = Font(bold=True, color="FFFFFF" if code != "I" else "000000")
        c.alignment = CENTER_ALIGN
        ws.cell(row=r, column=2, value=desc)

    ws.freeze_panes = "B2"
    return sheet_name


def create_index_sheet(wb, sheet_map):
    ws = wb.create_sheet(title="Index")
    wb.move_sheet(ws, offset=-len(wb.sheetnames) + 1)

    ws.cell(row=1, column=1, value="FRACKTAL WORKS PRIVATE LIMITED").font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:E1")
    ws.cell(row=2, column=1, value="Finance & Accounts — SOP / Workflow Charts").font = Font(bold=True, size=13, color="2F5496")
    ws.merge_cells("A2:E2")
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y')}   |   F&A Team: Head F&A + Junior Accountant").font = Font(italic=True, color="666666")
    ws.merge_cells("A3:E3")
    ws.cell(row=4, column=1, value="⛔ POLICY: All approvals must be in WRITING. No verbal approvals or commitments.").font = Font(bold=True, size=11, color="DC3545")
    ws.merge_cells("A4:E4")
    ws.cell(row=4, column=1).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    headers = ["#", "Process / Workflow", "Type", "Workflow Sheet", "RACI Sheet"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = INDEX_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for i, (wf_key, sheets) in enumerate(sheet_map.items(), start=1):
        row = 6 + i
        ws.cell(row=row, column=1, value=i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        # Determine type
        wtype = "P2P" if "P2P" in wf_key else "O2C" if "O2C" in wf_key else "R2R" if "R2R" in wf_key else "Compliance" if any(x in wf_key for x in ["GST", "TDS"]) else "Operational"
        ws.cell(row=row, column=2, value=wf_key).border = THIN_BORDER
        ws.cell(row=row, column=3, value=wtype).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN

        flow_name = sheets.get("flow", "")
        flow_cell = ws.cell(row=row, column=4, value=flow_name)
        flow_cell.font = Font(color="0563C1", underline="single")
        flow_cell.border = THIN_BORDER
        if flow_name:
            flow_cell.hyperlink = f"#'{flow_name}'!A1"

        raci_name = sheets.get("raci", "")
        raci_cell = ws.cell(row=row, column=5, value=raci_name)
        raci_cell.font = Font(color="0563C1", underline="single")
        raci_cell.border = THIN_BORDER
        if raci_name:
            raci_cell.hyperlink = f"#'{raci_name}'!A1"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 32
    ws.freeze_panes = "A7"


def create_approval_matrix_sheet(wb, thresholds):
    ws = wb.create_sheet(title="Approval_Matrix")

    ws.cell(row=1, column=1, value="Payment Approval Matrix — Fracktal Works").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:F1")
    ws.cell(row=2, column=1, value="⛔ Every payment MUST follow this matrix. No exceptions without Director's written approval.").font = Font(bold=True, color="DC3545")
    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    headers = ["Tier", "Amount Range (₹)", "Primary Approver", "Approval Mode", "Description", "Additional Controls"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))

    sorted_tiers = sorted(thresholds.items(), key=lambda x: x[1].get("limit") or float("inf"))
    prev_limit = 0
    for i, (tier, info) in enumerate(sorted_tiers, start=5):
        limit = info["limit"]
        if limit is None:
            range_str = f"Above ₹{prev_limit:,}"
        else:
            range_str = f"Up to ₹{limit:,}" if prev_limit == 0 else f"₹{prev_limit + 1:,} — ₹{limit:,}"

        controls = "Petty cash voucher (signed)" if limit and limit <= 2000 else ""
        if limit and limit > 2000:
            controls = "Maker-Checker (Jr Acct prepares, Head F&A authorizes)"
        if limit and limit > 50000:
            controls += " + Director email/signed approval"
        if limit is None:
            controls = "Board Resolution (signed minutes) + Dual signatory"

        ws.cell(row=i, column=1, value=tier.replace("_", " ").title())
        ws.cell(row=i, column=2, value=range_str)
        ws.cell(row=i, column=3, value=info["approver"])
        ws.cell(row=i, column=4, value=info.get("approval_mode", "Written approval"))
        ws.cell(row=i, column=5, value=info.get("description", ""))
        ws.cell(row=i, column=6, value=controls)

        for col in range(1, 7):
            style_cell(ws.cell(row=i, column=col))
        ws.row_dimensions[i].height = 30

        prev_limit = limit if limit else prev_limit

    col_widths = [18, 22, 22, 35, 35, 40]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"


def create_controls_summary_sheet(wb, workflows):
    ws = wb.create_sheet(title="Controls_Summary")

    ws.cell(row=1, column=1, value="Internal Controls & Compliance Points").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:F1")

    headers = ["#", "Workflow", "Step", "Control Point", "Type", "No-Verbal Flag"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    row = 4
    counter = 1
    for wf_key, wf_data in workflows.items():
        for step in wf_data["steps"]:
            if step.get("control"):
                ws.cell(row=row, column=1, value=counter)
                ws.cell(row=row, column=2, value=wf_data["name"])
                ws.cell(row=row, column=3, value=step["action"][:70])
                ws.cell(row=row, column=4, value=step["control"])

                ctrl_text = step["control"].lower()
                if any(w in ctrl_text for w in ["approval", "review", "sign", "authorize", "approve"]):
                    ctrl_type = "Authorization"
                elif any(w in ctrl_text for w in ["reconcil", "match", "verif", "validation", "check"]):
                    ctrl_type = "Verification"
                elif any(w in ctrl_text for w in ["segregat", "dual", "maker", "checker"]):
                    ctrl_type = "Segregation of Duties"
                elif any(w in ctrl_text for w in ["monitor", "track", "aging", "deadline"]):
                    ctrl_type = "Monitoring"
                else:
                    ctrl_type = "Preventive"

                ws.cell(row=row, column=5, value=ctrl_type)

                no_verbal = "⛔ YES" if "NO VERBAL" in step["control"] else ""
                ws.cell(row=row, column=6, value=no_verbal)
                if no_verbal:
                    ws.cell(row=row, column=6).font = Font(bold=True, color="DC3545")

                for col in range(1, 7):
                    style_cell(ws.cell(row=row, column=col))

                counter += 1
                row += 1

    col_widths = [6, 28, 50, 50, 22, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


def create_kpi_sheet(wb):
    ws = wb.create_sheet(title="KPI_Metrics")

    ws.cell(row=1, column=1, value="F&A Key Performance Indicators — Fracktal Works").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:F1")

    headers = ["#", "KPI", "Workflow / Process", "Target", "Frequency", "Owner"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    kpis = [
        ("Invoice Processing Time (PR to Payment)", "P2P", "≤ 5 working days (after GRN)", "Per invoice", R["JR_ACCT"]),
        ("3-Way Match Success Rate", "P2P", "≥ 98%", "Monthly", R["JR_ACCT"]),
        ("Vendor Payment On-Time Rate", "P2P", "≥ 95% per payment terms", "Monthly", R["HEAD_FA"]),
        ("Days Payable Outstanding (DPO)", "P2P", "≤ 30 days", "Monthly", R["HEAD_FA"]),
        ("Invoice-to-Dispatch Time", "O2C", "≤ 24 hours after delivery", "Per invoice", R["JR_ACCT"]),
        ("Days Sales Outstanding (DSO)", "O2C", "≤ 45 days", "Monthly", R["HEAD_FA"]),
        ("Collection Efficiency Rate", "O2C", "≥ 95%", "Monthly", R["HEAD_FA"]),
        ("AR Aging > 90 Days", "O2C", "< 5% of total AR", "Monthly", R["HEAD_FA"]),
        ("Month-End Close Time", "R2R", "≤ T+6 working days", "Monthly", R["HEAD_FA"]),
        ("BRS Outstanding Items > 30 Days", "Bank Recon", "Zero", "Monthly", R["JR_ACCT"]),
        ("Expense Claim Processing Time", "Expense", "≤ 5 working days", "Per claim", R["JR_ACCT"]),
        ("Payroll Accuracy (zero errors)", "Payroll", "100%", "Monthly", R["HEAD_FA"]),
        ("Payroll Disbursement Date", "Payroll", "≤ 7th of month", "Monthly", R["JR_ACCT"]),
        ("PF/ESI Deposit Timeliness", "Payroll", "By 15th of month — 100%", "Monthly", R["JR_ACCT"]),
        ("GSTR-1 Filing Timeliness", "GST", "By 11th — 100%", "Monthly", R["HEAD_FA"]),
        ("GSTR-3B Filing Timeliness", "GST", "By 20th — 100%", "Monthly", R["HEAD_FA"]),
        ("ITC Match Rate (Books vs 2B)", "GST", "≥ 98%", "Monthly", R["JR_ACCT"]),
        ("TDS Deposit Timeliness", "TDS", "By 7th — 100%", "Monthly", R["JR_ACCT"]),
        ("TDS Return Filing (24Q/26Q)", "TDS", "On or before due date", "Quarterly", R["HEAD_FA"]),
        ("Budget Variance (actuals vs budget)", "Budget & MIS", "Within ±10%", "Monthly", R["HEAD_FA"]),
        ("MIS Report Delivery", "Budget & MIS", "≤ T+8 working days", "Monthly", R["HEAD_FA"]),
        ("Verbal Approval Incidents", "ALL", "ZERO — target 0 incidents", "Monthly", R["HEAD_FA"]),
    ]

    for i, (kpi, workflow, target, freq, owner) in enumerate(kpis, start=4):
        ws.cell(row=i, column=1, value=i - 3)
        ws.cell(row=i, column=2, value=kpi)
        ws.cell(row=i, column=3, value=workflow)
        ws.cell(row=i, column=4, value=target)
        ws.cell(row=i, column=5, value=freq)
        ws.cell(row=i, column=6, value=owner)
        for col in range(1, 7):
            style_cell(ws.cell(row=i, column=col))

    col_widths = [6, 42, 18, 32, 14, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


def create_dept_responsibility_sheet(wb):
    """Cross-department responsibilities linked to Finance."""
    ws = wb.create_sheet(title="Dept_Responsibilities")

    ws.cell(row=1, column=1, value="Department Responsibilities — Finance Linkages").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:E1")
    ws.cell(row=2, column=1, value="Every department has obligations to Finance. Non-compliance = process delays.").font = Font(italic=True, color="666666")
    ws.merge_cells("A2:E2")

    headers = ["Department", "Responsibility", "Deliverable to Finance", "Deadline", "Approval Mode"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))

    responsibilities = [
        ("Sales / Projects", "Share customer PO before work begins", "Customer PO (email/scan)", "Before work starts", "Email"),
        ("Sales / Projects", "Confirm delivery for invoicing", "Signed Delivery Challan / Completion Report", "Day of delivery", "Signed document"),
        ("Sales / Projects", "Provide customer contact for collections", "Customer contact details (email)", "With PO", "Email"),
        ("Sales / Projects", "Submit revenue forecast for budgeting", "Revenue Forecast Sheet", "Per budget circular", "Email + signed"),
        ("Engineering / R&D", "Raise Purchase Requisition for materials", "Signed PR Form (email)", "Before procurement", "Email / signed form"),
        ("Engineering / R&D", "Sign GRN upon receiving goods", "Signed GRN", "Day of receipt", "Physical signature"),
        ("Engineering / R&D", "Submit project cost estimates", "Cost Estimate Sheet", "Per budget circular", "Email"),
        ("Operations / Production", "Raise PR for raw materials / consumables", "Signed PR Form", "Per production schedule", "Email / signed form"),
        ("Operations / Production", "Sign GRN for materials received", "Signed GRN", "Day of receipt", "Physical signature"),
        ("Operations / Production", "Inventory count data for month-end", "Stock Count Sheet (signed)", "Last working day of month", "Signed by Ops Head"),
        ("Operations / Production", "Generate e-Way Bill info (dispatch details)", "Dispatch details to Finance", "Before dispatch", "Email"),
        ("HR / Admin", "Submit attendance & leave data", "Signed Attendance Sheet", "1st working day of month", "Signed by Dept Heads"),
        ("HR / Admin", "Inform joiners, exits, salary revisions", "Appointment/Exit/Revision letters", "Immediately on occurrence", "Signed letters"),
        ("HR / Admin", "Submit expense claims with original bills", "Expense Claim Form + receipts", "Within 7 days of expense", "Signed by Dept Head"),
        ("All Departments", "Respond to Finance queries", "Written response (email)", "Within 2 working days", "Email"),
        ("All Departments", "Submit department budget", "Budget Sheet in standard template", "Per budget circular deadline", "Email + signed"),
        ("All Departments", "No verbal approvals or commitments", "N/A", "Always", "⛔ POLICY"),
        ("Director / CEO", "Approve payments > ₹50,000", "Email / signed approval", "Within 2 working days", "Email / signed"),
        ("Director / CEO", "Sign-off on monthly financials", "Signed financial package", "By T+6", "Physical / email signature"),
        ("Director / CEO", "Approve annual budget", "Signed budget", "Before FY start", "Meeting minutes + sign"),
    ]

    for i, (dept, resp, deliverable, deadline, mode) in enumerate(responsibilities, start=5):
        ws.cell(row=i, column=1, value=dept)
        ws.cell(row=i, column=2, value=resp)
        ws.cell(row=i, column=3, value=deliverable)
        ws.cell(row=i, column=4, value=deadline)
        ws.cell(row=i, column=5, value=mode)
        for col in range(1, 6):
            style_cell(ws.cell(row=i, column=col))
        if "POLICY" in mode:
            for col in range(1, 6):
                ws.cell(row=i, column=col).font = Font(bold=True, color="DC3545")

    col_widths = [22, 40, 40, 25, 25]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"


# ─── Main Generator ─────────────────────────────────────────────────────

def generate_workbook(output_path=None):
    TMP_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(TMP_DIR / f"Fracktal_FA_SOPs_Workflows_{timestamp}.xlsx")

    all_workflows = get_all_workflows()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print(f"\n{'='*70}")
    print(f"  FRACKTAL WORKS — F&A SOP / WORKFLOW GENERATOR")
    print(f"  Team: Head F&A + Junior Accountant")
    print(f"  Policy: ⛔ NO VERBAL APPROVALS")
    print(f"  Workflows: {len(all_workflows)}")
    print(f"{'='*70}\n")

    # Policy sheet first
    create_policy_sheet(wb)
    print(f"  ✓ Policy: No Verbal Approvals & Controls")

    # Generate workflow + RACI sheets
    sheet_map = {}
    for wf_key, wf_data in all_workflows.items():
        print(f"  → {wf_data['name']}...")
        flow_sheet = create_flowchart_sheet(wb, wf_data)
        raci_sheet = create_raci_sheet(wb, wf_data)
        sheet_map[wf_data["name"]] = {"flow": flow_sheet, "raci": raci_sheet}
        print(f"    ✓ Workflow: {flow_sheet}")
        print(f"    ✓ RACI: {raci_sheet}")

    # Summary sheets
    print(f"\n  → Summary sheets...")
    create_index_sheet(wb, sheet_map)
    print(f"    ✓ Index (Table of Contents)")

    create_approval_matrix_sheet(wb, DEFAULT_THRESHOLDS)
    print(f"    ✓ Approval Matrix")

    create_controls_summary_sheet(wb, all_workflows)
    print(f"    ✓ Controls Summary (all control points)")

    create_kpi_sheet(wb)
    print(f"    ✓ KPI Metrics")

    create_dept_responsibility_sheet(wb)
    print(f"    ✓ Department Responsibilities (cross-functional)")

    wb.save(output_path)
    print(f"\n{'='*70}")
    print(f"  ✅ WORKBOOK SAVED: {output_path}")
    print(f"  Total Sheets: {len(wb.sheetnames)}")
    print(f"{'='*70}\n")

    return output_path


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Fracktal Works F&A SOP/Workflow Generator")
    parser.add_argument("--output", default=None, help="Output file path")
    args = parser.parse_args()
    generate_workbook(output_path=args.output)
