#!/usr/bin/env python3
"""
F&A Workflow Chart Generator — Fracktal Works Private Limited

Generates comprehensive workflow charts for Finance & Accounting processes
as an Excel workbook with swim-lane flowcharts, RACI matrices, approval
matrices, and KPI dashboards.

Usage:
    python fa_workflow_generator.py --company "Fracktal Works Private Limited"
    python fa_workflow_generator.py --workflows "AP,AR,Payment Approval"
    python fa_workflow_generator.py --workflows "Payment Approval" --thresholds '{"petty_cash":{"limit":5000,"approver":"Team Lead"}}'
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"

# ─── Styles ───────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DECISION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CONTROL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ESCALATION_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
INDEX_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
RACI_R_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
RACI_A_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
RACI_C_FILL = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
RACI_I_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ─── Default Approval Thresholds (INR) ───────────────────────────────────────

DEFAULT_THRESHOLDS = {
    "petty_cash":  {"limit": 5000,   "approver": "Department Head",   "description": "Petty cash & minor expenses"},
    "operational": {"limit": 50000,  "approver": "Finance Manager",   "description": "Operational expenses & vendor payments"},
    "capital":     {"limit": 200000, "approver": "CFO",               "description": "Capital expenditure & large procurements"},
    "strategic":   {"limit": 500000, "approver": "CEO",               "description": "Strategic investments & contracts"},
    "board_level": {"limit": None,   "approver": "Board of Directors", "description": "Above ₹5,00,000 — requires Board approval"},
}

# ─── Workflow Definitions ─────────────────────────────────────────────────────

def get_all_workflows():
    """Return all F&A workflow definitions."""
    return {
        "AP": ap_workflow(),
        "AR": ar_workflow(),
        "Month-End Close": month_end_workflow(),
        "Expense Reimbursement": expense_workflow(),
        "Payment Approval": payment_approval_workflow(),
        "Financial Reporting": financial_reporting_workflow(),
        "Payroll": payroll_workflow(),
        "Bank Reconciliation": bank_recon_workflow(),
        "GST Compliance": gst_workflow(),
        "Budget Planning": budget_workflow(),
    }


def ap_workflow():
    """Accounts Payable workflow steps."""
    return {
        "name": "Accounts Payable (AP)",
        "steps": [
            {"step": 1, "phase": "Initiation", "role": "Vendor/Dept Head", "action": "Submit invoice with PO reference", "decision": False, "doc": "Invoice, Purchase Order", "control": "3-way matching (PO-GRN-Invoice)", "sla": "Day 0", "escalation": "Finance Manager"},
            {"step": 2, "phase": "Initiation", "role": "Finance Executive", "action": "Receive and log invoice in register", "decision": False, "doc": "Invoice Register", "control": "Unique invoice number check", "sla": "Day 0", "escalation": "Finance Manager"},
            {"step": 3, "phase": "Processing", "role": "Finance Executive", "action": "Verify invoice details (amount, GST, PO match)", "decision": True, "yes": "Proceed to Step 4", "no": "Return to vendor for correction", "doc": "PO, GRN, Invoice", "control": "3-way match verification", "sla": "Day 1", "escalation": "Finance Manager"},
            {"step": 4, "phase": "Processing", "role": "Finance Executive", "action": "Record invoice in accounting system (Tally/ERP)", "decision": False, "doc": "Tally/ERP", "control": "GL code validation", "sla": "Day 1", "escalation": "Finance Manager"},
            {"step": 5, "phase": "Approval", "role": "Department Head", "action": "Review and approve invoice for payment", "decision": True, "yes": "Proceed to Step 6", "no": "Return with comments for rework", "doc": "Approval Form", "control": "Budget availability check", "sla": "Day 2", "escalation": "Finance Manager"},
            {"step": 6, "phase": "Approval", "role": "Finance Manager", "action": "Verify approval and check payment terms", "decision": True, "yes": "Schedule payment per terms", "no": "Hold — escalate discrepancy", "doc": "Payment Schedule", "control": "Approval threshold check", "sla": "Day 2", "escalation": "CFO"},
            {"step": 7, "phase": "Payment", "role": "Finance Manager", "action": "Prepare payment batch (NEFT/RTGS/Cheque)", "decision": False, "doc": "Bank Portal / Cheque Register", "control": "Dual signatory for amounts > ₹50,000", "sla": "Per payment terms", "escalation": "CFO"},
            {"step": 8, "phase": "Payment", "role": "CFO/CEO", "action": "Authorize payment (per approval matrix)", "decision": True, "yes": "Release payment", "no": "Hold and investigate", "doc": "Bank Authorization", "control": "Segregation of duties", "sla": "Day 3", "escalation": "CEO/Board"},
            {"step": 9, "phase": "Completion", "role": "Finance Executive", "action": "Record payment in books, update vendor ledger", "decision": False, "doc": "Tally/ERP, Bank Statement", "control": "Payment reference tracking", "sla": "Day 3", "escalation": "Finance Manager"},
            {"step": 10, "phase": "Completion", "role": "Finance Executive", "action": "Reconcile with bank statement, file documents", "decision": False, "doc": "Bank Reconciliation Statement", "control": "Monthly reconciliation", "sla": "Month-end", "escalation": "Finance Manager"},
        ],
        "raci_roles": ["Vendor", "Dept Head", "Finance Exec", "Finance Manager", "CFO", "CEO", "Auditor"],
        "raci": [
            ["Submit Invoice",        "R", "I", "I", "I", "", "", ""],
            ["Log Invoice",           "I", "", "R", "A", "", "", ""],
            ["Verify Invoice",        "", "", "R", "A", "", "", "I"],
            ["Record in System",      "", "", "R", "A", "", "", ""],
            ["Department Approval",   "", "R", "I", "A", "", "", ""],
            ["Finance Approval",      "", "", "I", "R", "A", "", ""],
            ["Prepare Payment",       "", "", "I", "R", "A", "", ""],
            ["Authorize Payment",     "", "", "", "I", "R", "A", ""],
            ["Record Payment",        "", "", "R", "A", "", "", ""],
            ["Bank Reconciliation",   "", "", "R", "A", "I", "", "I"],
        ],
    }


def ar_workflow():
    """Accounts Receivable workflow steps."""
    return {
        "name": "Accounts Receivable (AR)",
        "steps": [
            {"step": 1, "phase": "Initiation", "role": "Sales/Project Team", "action": "Confirm deliverable completion, request billing", "decision": False, "doc": "Delivery Note / Completion Report", "control": "Milestone verification", "sla": "Upon completion", "escalation": "Finance Manager"},
            {"step": 2, "phase": "Initiation", "role": "Finance Executive", "action": "Generate invoice based on contract/PO terms", "decision": False, "doc": "Sales Invoice, Contract", "control": "Rate & quantity verification", "sla": "T+1 day", "escalation": "Finance Manager"},
            {"step": 3, "phase": "Initiation", "role": "Finance Manager", "action": "Review and approve invoice before dispatch", "decision": True, "yes": "Dispatch to customer", "no": "Return for correction", "doc": "Invoice Approval", "control": "GST, TDS applicability check", "sla": "T+1 day", "escalation": "CFO"},
            {"step": 4, "phase": "Processing", "role": "Finance Executive", "action": "Dispatch invoice to customer (email/portal)", "decision": False, "doc": "Email / Customer Portal", "control": "Delivery confirmation", "sla": "T+2 days", "escalation": "Finance Manager"},
            {"step": 5, "phase": "Processing", "role": "Finance Executive", "action": "Record invoice in AR ledger and aging tracker", "decision": False, "doc": "AR Aging Report", "control": "Due date tracking", "sla": "T+2 days", "escalation": "Finance Manager"},
            {"step": 6, "phase": "Follow-up", "role": "Finance Executive", "action": "Send payment reminder at due date", "decision": True, "yes": "Payment received — go to Step 8", "no": "Escalate — go to Step 7", "doc": "Reminder Email", "control": "Aging bucket monitoring (30/60/90)", "sla": "Due date", "escalation": "Finance Manager"},
            {"step": 7, "phase": "Follow-up", "role": "Finance Manager", "action": "Escalate overdue — call customer, negotiate payment plan", "decision": True, "yes": "Payment committed — monitor", "no": "Escalate to CFO for legal action", "doc": "Collection Notes", "control": "Write-off threshold review", "sla": "Due +15 days", "escalation": "CFO"},
            {"step": 8, "phase": "Collection", "role": "Finance Executive", "action": "Receive payment, match against invoice", "decision": True, "yes": "Full payment — close invoice", "no": "Partial — update balance, follow up remainder", "doc": "Bank Statement, Receipt", "control": "TDS certificate collection", "sla": "Upon receipt", "escalation": "Finance Manager"},
            {"step": 9, "phase": "Completion", "role": "Finance Executive", "action": "Issue receipt, update customer ledger", "decision": False, "doc": "Receipt Voucher", "control": "Ledger reconciliation", "sla": "T+1 day", "escalation": "Finance Manager"},
            {"step": 10, "phase": "Completion", "role": "Finance Executive", "action": "Reconcile AR aging report, archive documents", "decision": False, "doc": "AR Aging Report", "control": "Monthly aging review", "sla": "Month-end", "escalation": "Finance Manager"},
        ],
        "raci_roles": ["Sales Team", "Dept Head", "Finance Exec", "Finance Manager", "CFO", "CEO", "Auditor"],
        "raci": [
            ["Request Billing",       "R", "A", "I", "I", "", "", ""],
            ["Generate Invoice",      "C", "", "R", "A", "", "", ""],
            ["Approve Invoice",       "", "", "I", "R", "A", "", ""],
            ["Dispatch Invoice",      "", "", "R", "A", "", "", ""],
            ["Record in AR Ledger",   "", "", "R", "A", "", "", ""],
            ["Payment Follow-up",     "C", "", "R", "A", "", "", ""],
            ["Escalate Overdue",      "", "", "I", "R", "A", "", ""],
            ["Receive Payment",       "", "", "R", "A", "", "", ""],
            ["Issue Receipt",         "", "", "R", "A", "", "", ""],
            ["Reconcile AR",          "", "", "R", "A", "I", "", "I"],
        ],
    }


def month_end_workflow():
    """Month-End Close workflow steps."""
    return {
        "name": "Month-End Close",
        "steps": [
            {"step": 1, "phase": "Preparation", "role": "Finance Executive", "action": "Export Trial Balance from Tally/ERP", "decision": False, "doc": "Trial Balance", "control": "TB date validation", "sla": "T+1", "escalation": "Finance Manager"},
            {"step": 2, "phase": "Preparation", "role": "Finance Executive", "action": "Complete all pending invoice entries (AP & AR)", "decision": False, "doc": "Invoice Register", "control": "Cut-off check", "sla": "T+1", "escalation": "Finance Manager"},
            {"step": 3, "phase": "Adjustments", "role": "Finance Executive", "action": "Pass accrual and provision entries", "decision": False, "doc": "Accrual Schedule", "control": "Accrual completeness checklist", "sla": "T+2", "escalation": "Finance Manager"},
            {"step": 4, "phase": "Adjustments", "role": "Finance Executive", "action": "Pass depreciation and amortization entries", "decision": False, "doc": "Fixed Asset Register", "control": "Depreciation schedule review", "sla": "T+2", "escalation": "Finance Manager"},
            {"step": 5, "phase": "Adjustments", "role": "Finance Executive", "action": "Record prepaid expenses and advances adjustment", "decision": False, "doc": "Prepaid Schedule", "control": "Amortization schedule", "sla": "T+2", "escalation": "Finance Manager"},
            {"step": 6, "phase": "Reconciliation", "role": "Finance Executive", "action": "Complete bank reconciliation for all accounts", "decision": True, "yes": "All matched — proceed", "no": "Investigate unmatched items", "doc": "BRS", "control": "Zero unmatched target", "sla": "T+3", "escalation": "Finance Manager"},
            {"step": 7, "phase": "Reconciliation", "role": "Finance Executive", "action": "Reconcile inter-company, vendor, customer ledgers", "decision": True, "yes": "Balanced — proceed", "no": "Adjust and recheck", "doc": "Reconciliation Statements", "control": "Confirmation from counterparties", "sla": "T+3", "escalation": "Finance Manager"},
            {"step": 8, "phase": "Reporting", "role": "Finance Manager", "action": "Review adjusted TB, generate P&L and Balance Sheet", "decision": True, "yes": "Reports accurate — proceed", "no": "Return for corrections", "doc": "P&L, Balance Sheet", "control": "BS must balance (A = L + E)", "sla": "T+4", "escalation": "CFO"},
            {"step": 9, "phase": "Reporting", "role": "Finance Manager", "action": "Prepare MIS report with variance analysis", "decision": False, "doc": "MIS Report", "control": "Budget vs Actuals comparison", "sla": "T+4", "escalation": "CFO"},
            {"step": 10, "phase": "Sign-off", "role": "CFO", "action": "Review and sign-off on monthly financials", "decision": True, "yes": "Approved — distribute reports", "no": "Request revisions", "doc": "Signed Financial Package", "control": "Management review control", "sla": "T+5", "escalation": "CEO"},
        ],
        "raci_roles": ["Finance Exec", "Dept Heads", "Finance Manager", "CFO", "CEO", "Auditor"],
        "raci": [
            ["Export Trial Balance",   "R", "", "A", "I", "", ""],
            ["Complete Pending Entries","R", "C", "A", "", "", ""],
            ["Pass Accruals",          "R", "", "A", "", "", "I"],
            ["Depreciation Entries",   "R", "", "A", "", "", "I"],
            ["Prepaid Adjustments",    "R", "", "A", "", "", ""],
            ["Bank Reconciliation",    "R", "", "A", "I", "", "I"],
            ["Ledger Reconciliation",  "R", "", "A", "", "", "I"],
            ["Generate Reports",       "C", "", "R", "A", "", ""],
            ["Prepare MIS",            "C", "", "R", "A", "I", ""],
            ["Sign-off Financials",    "", "", "I", "R", "A", "I"],
        ],
    }


def expense_workflow():
    """Expense Reimbursement workflow steps."""
    return {
        "name": "Expense Reimbursement",
        "steps": [
            {"step": 1, "phase": "Submission", "role": "Employee", "action": "Fill expense claim form with receipts/bills", "decision": False, "doc": "Expense Claim Form, Receipts", "control": "Policy compliance (per diem limits)", "sla": "Within 7 days of expense", "escalation": "Department Head"},
            {"step": 2, "phase": "Submission", "role": "Employee", "action": "Submit claim to Department Head for approval", "decision": False, "doc": "Expense Claim Form", "control": "Expense within budget allocation", "sla": "Day 0", "escalation": "Department Head"},
            {"step": 3, "phase": "Approval", "role": "Department Head", "action": "Review claim — verify business purpose and amounts", "decision": True, "yes": "Approve and forward to Finance", "no": "Reject with reason / request modification", "doc": "Approved Claim Form", "control": "Business purpose validation", "sla": "Day 1-2", "escalation": "Finance Manager"},
            {"step": 4, "phase": "Approval", "role": "Finance Executive", "action": "Verify receipts, check policy limits, validate GST input", "decision": True, "yes": "Accept for processing", "no": "Return with discrepancy note", "doc": "Expense Policy, GST Invoices", "control": "Receipt authenticity, GST input check", "sla": "Day 2-3", "escalation": "Finance Manager"},
            {"step": 5, "phase": "Approval", "role": "Finance Manager", "action": "Final approval for claims above ₹10,000", "decision": True, "yes": "Approve and schedule payment", "no": "Reject or request additional info", "doc": "Approval Register", "control": "Threshold-based approval", "sla": "Day 3", "escalation": "CFO"},
            {"step": 6, "phase": "Processing", "role": "Finance Executive", "action": "Record expense in books, process reimbursement", "decision": False, "doc": "Tally/ERP, Payment Voucher", "control": "GL code accuracy", "sla": "Day 4-5", "escalation": "Finance Manager"},
            {"step": 7, "phase": "Payment", "role": "Finance Manager", "action": "Release payment via bank transfer / petty cash", "decision": False, "doc": "Bank Transfer / Petty Cash Register", "control": "Payment method per policy", "sla": "Day 5-7", "escalation": "CFO"},
            {"step": 8, "phase": "Completion", "role": "Finance Executive", "action": "Notify employee, update records, file documents", "decision": False, "doc": "Payment Confirmation", "control": "Employee acknowledgement", "sla": "Day 7", "escalation": "Finance Manager"},
        ],
        "raci_roles": ["Employee", "Dept Head", "Finance Exec", "Finance Manager", "CFO", "Auditor"],
        "raci": [
            ["Fill Expense Claim",     "R", "I", "", "", "", ""],
            ["Submit to Dept Head",    "R", "A", "", "", "", ""],
            ["Dept Head Approval",     "I", "R", "I", "A", "", ""],
            ["Finance Verification",   "", "", "R", "A", "", "I"],
            ["Finance Mgr Approval",   "", "", "I", "R", "A", ""],
            ["Record & Process",       "", "", "R", "A", "", ""],
            ["Release Payment",        "", "", "I", "R", "A", ""],
            ["Notify & Archive",       "I", "", "R", "A", "", ""],
        ],
    }


def payment_approval_workflow():
    """Payment Approval Policy workflow."""
    return {
        "name": "Payment Approval Policy",
        "steps": [
            {"step": 1, "phase": "Initiation", "role": "Requestor", "action": "Raise payment request with supporting documents", "decision": False, "doc": "Payment Request Form, Invoice/PO", "control": "Complete documentation check", "sla": "Day 0", "escalation": "Department Head"},
            {"step": 2, "phase": "Verification", "role": "Finance Executive", "action": "Verify documentation, check budget, classify amount tier", "decision": True, "yes": "Documents complete — route per approval matrix", "no": "Return for missing documents", "doc": "Budget Tracker", "control": "Budget availability", "sla": "Day 1", "escalation": "Finance Manager"},
            {"step": 3, "phase": "Approval Tier 1", "role": "Department Head", "action": "Approve payments up to ₹5,000 (Petty Cash tier)", "decision": True, "yes": "Approved — proceed to payment", "no": "Reject / request modification", "doc": "Approval Form", "control": "Amount ≤ ₹5,000", "sla": "Day 1", "escalation": "Finance Manager"},
            {"step": 4, "phase": "Approval Tier 2", "role": "Finance Manager", "action": "Approve payments ₹5,001 — ₹50,000 (Operational tier)", "decision": True, "yes": "Approved — proceed to payment", "no": "Reject / escalate", "doc": "Approval Form", "control": "Amount ≤ ₹50,000", "sla": "Day 1-2", "escalation": "CFO"},
            {"step": 5, "phase": "Approval Tier 3", "role": "CFO", "action": "Approve payments ₹50,001 — ₹2,00,000 (Capital tier)", "decision": True, "yes": "Approved — proceed to payment", "no": "Reject / request Board review", "doc": "Capital Approval Form", "control": "Amount ≤ ₹2,00,000", "sla": "Day 2-3", "escalation": "CEO"},
            {"step": 6, "phase": "Approval Tier 4", "role": "CEO", "action": "Approve payments ₹2,00,001 — ₹5,00,000 (Strategic tier)", "decision": True, "yes": "Approved — proceed to payment", "no": "Reject / defer to Board", "doc": "Strategic Approval Form", "control": "Amount ≤ ₹5,00,000", "sla": "Day 3-5", "escalation": "Board"},
            {"step": 7, "phase": "Approval Tier 5", "role": "Board of Directors", "action": "Approve payments above ₹5,00,000 (Board tier)", "decision": True, "yes": "Board resolution passed — proceed", "no": "Rejected — record minutes", "doc": "Board Resolution", "control": "Board quorum required", "sla": "Next Board meeting", "escalation": "Chairman"},
            {"step": 8, "phase": "Execution", "role": "Finance Manager", "action": "Execute approved payment through banking channel", "decision": False, "doc": "Bank Portal / Cheque", "control": "Dual signatory for > ₹50,000", "sla": "Per payment terms", "escalation": "CFO"},
            {"step": 9, "phase": "Completion", "role": "Finance Executive", "action": "Record payment, update ledger, file approval trail", "decision": False, "doc": "Tally/ERP, Filing", "control": "Complete audit trail", "sla": "Day of payment", "escalation": "Finance Manager"},
        ],
        "raci_roles": ["Requestor", "Dept Head", "Finance Exec", "Finance Manager", "CFO", "CEO", "Board"],
        "raci": [
            ["Raise Request",          "R", "I", "I", "I", "", "", ""],
            ["Verify & Classify",       "", "", "R", "A", "", "", ""],
            ["Tier 1 (≤₹5K)",          "I", "R", "I", "A", "", "", ""],
            ["Tier 2 (≤₹50K)",         "I", "", "I", "R", "A", "", ""],
            ["Tier 3 (≤₹2L)",          "I", "", "", "I", "R", "A", ""],
            ["Tier 4 (≤₹5L)",          "I", "", "", "", "I", "R", "A"],
            ["Tier 5 (>₹5L)",          "I", "", "", "", "I", "I", "R"],
            ["Execute Payment",        "I", "", "I", "R", "A", "", ""],
            ["Record & File",          "", "", "R", "A", "", "", ""],
        ],
    }


def financial_reporting_workflow():
    """Financial Reporting workflow steps."""
    return {
        "name": "Financial Reporting",
        "steps": [
            {"step": 1, "phase": "Data Collection", "role": "Finance Executive", "action": "Collect data from all departments (revenue, costs, headcount)", "decision": False, "doc": "Department Reports", "control": "Deadline adherence tracker", "sla": "T+1", "escalation": "Finance Manager"},
            {"step": 2, "phase": "Data Collection", "role": "Finance Executive", "action": "Download bank statements, reconcile cash position", "decision": False, "doc": "Bank Statements", "control": "All accounts covered", "sla": "T+1", "escalation": "Finance Manager"},
            {"step": 3, "phase": "Report Generation", "role": "Finance Executive", "action": "Generate P&L, Balance Sheet, Cash Flow from TB", "decision": False, "doc": "Financial Statements", "control": "TB must be finalized", "sla": "T+2", "escalation": "Finance Manager"},
            {"step": 4, "phase": "Report Generation", "role": "Finance Manager", "action": "Prepare variance analysis (Budget vs Actuals)", "decision": False, "doc": "Variance Report", "control": "Material variance investigation (>10%)", "sla": "T+3", "escalation": "CFO"},
            {"step": 5, "phase": "Report Generation", "role": "Finance Manager", "action": "Compile MIS dashboard with KPIs", "decision": False, "doc": "MIS Dashboard", "control": "KPI trend consistency", "sla": "T+3", "escalation": "CFO"},
            {"step": 6, "phase": "Review", "role": "CFO", "action": "Review financial package, verify key metrics", "decision": True, "yes": "Approve for distribution", "no": "Return for corrections", "doc": "Review Checklist", "control": "Analytical review", "sla": "T+4", "escalation": "CEO"},
            {"step": 7, "phase": "Distribution", "role": "Finance Manager", "action": "Distribute reports to stakeholders (management, board)", "decision": False, "doc": "Email / Shared Drive", "control": "Distribution list verification", "sla": "T+5", "escalation": "CFO"},
            {"step": 8, "phase": "Distribution", "role": "CFO", "action": "Present financials in management/board meeting", "decision": False, "doc": "Presentation Deck", "control": "Minutes of meeting", "sla": "As scheduled", "escalation": "CEO"},
        ],
        "raci_roles": ["Finance Exec", "Dept Heads", "Finance Manager", "CFO", "CEO", "Board", "Auditor"],
        "raci": [
            ["Collect Dept Data",     "R", "C", "A", "", "", "", ""],
            ["Download Bank Stmts",   "R", "", "A", "", "", "", ""],
            ["Generate Statements",   "R", "", "A", "I", "", "", ""],
            ["Variance Analysis",     "C", "", "R", "A", "", "", ""],
            ["Compile MIS",           "C", "", "R", "A", "", "", ""],
            ["Review & Approve",      "", "", "I", "R", "A", "", ""],
            ["Distribute Reports",    "I", "I", "R", "A", "I", "I", ""],
            ["Present to Board",      "", "", "C", "R", "", "A", "I"],
        ],
    }


def payroll_workflow():
    """Payroll Processing workflow steps."""
    return {
        "name": "Payroll Processing",
        "steps": [
            {"step": 1, "phase": "Data Collection", "role": "HR Executive", "action": "Compile attendance, leave records, overtime data", "decision": False, "doc": "Attendance Register, Leave Tracker", "control": "Attendance vs leave reconciliation", "sla": "1st of month", "escalation": "HR Manager"},
            {"step": 2, "phase": "Data Collection", "role": "HR Executive", "action": "Update new joiners, exits, salary revisions", "decision": False, "doc": "HR Master Data", "control": "Approval letters on file", "sla": "1st of month", "escalation": "HR Manager"},
            {"step": 3, "phase": "Calculation", "role": "Finance Executive", "action": "Calculate gross salary, deductions (PF, ESI, TDS, PT)", "decision": False, "doc": "Payroll Register", "control": "Statutory rate validation", "sla": "2nd-3rd of month", "escalation": "Finance Manager"},
            {"step": 4, "phase": "Calculation", "role": "Finance Executive", "action": "Calculate net salary and prepare payroll summary", "decision": False, "doc": "Payroll Summary", "control": "Cross-check with previous month", "sla": "3rd of month", "escalation": "Finance Manager"},
            {"step": 5, "phase": "Approval", "role": "Finance Manager", "action": "Review payroll register, verify calculations", "decision": True, "yes": "Approve payroll for processing", "no": "Return for corrections", "doc": "Approved Payroll Register", "control": "Variance analysis (monthly)", "sla": "4th of month", "escalation": "CFO"},
            {"step": 6, "phase": "Approval", "role": "CFO", "action": "Final payroll approval and authorize bank transfer", "decision": True, "yes": "Authorize salary disbursement", "no": "Hold — investigate anomalies", "doc": "Bank Authorization", "control": "Total payroll budget check", "sla": "5th of month", "escalation": "CEO"},
            {"step": 7, "phase": "Disbursement", "role": "Finance Manager", "action": "Process salary transfer via bank (NEFT/RTGS)", "decision": False, "doc": "Bank Transfer File", "control": "Account number verification", "sla": "5th-7th of month", "escalation": "CFO"},
            {"step": 8, "phase": "Disbursement", "role": "Finance Executive", "action": "Generate and distribute payslips to employees", "decision": False, "doc": "Payslips (PDF)", "control": "Employee acknowledgement", "sla": "7th of month", "escalation": "Finance Manager"},
            {"step": 9, "phase": "Statutory", "role": "Finance Executive", "action": "Deposit PF, ESI contributions; file returns", "decision": False, "doc": "PF/ESI Challan", "control": "Due date compliance (15th)", "sla": "15th of month", "escalation": "Finance Manager"},
            {"step": 10, "phase": "Statutory", "role": "Finance Executive", "action": "Deposit TDS and file TDS return (quarterly)", "decision": False, "doc": "TDS Challan, Form 24Q", "control": "TDS rate accuracy", "sla": "7th of next month / quarterly", "escalation": "Finance Manager"},
        ],
        "raci_roles": ["HR Exec", "HR Manager", "Finance Exec", "Finance Manager", "CFO", "Auditor"],
        "raci": [
            ["Compile Attendance",     "R", "A", "", "", "", ""],
            ["Update HR Master",       "R", "A", "", "", "", ""],
            ["Calculate Gross/Deductions", "", "", "R", "A", "", "I"],
            ["Calculate Net Salary",   "", "", "R", "A", "", ""],
            ["Review Payroll",         "", "", "I", "R", "A", ""],
            ["Final Approval",         "", "", "", "I", "R", ""],
            ["Process Bank Transfer",  "", "", "I", "R", "A", ""],
            ["Distribute Payslips",    "", "I", "R", "A", "", ""],
            ["Deposit PF/ESI",         "", "", "R", "A", "", "I"],
            ["Deposit TDS & File",     "", "", "R", "A", "I", "I"],
        ],
    }


def bank_recon_workflow():
    """Bank Reconciliation workflow steps."""
    return {
        "name": "Bank Reconciliation",
        "steps": [
            {"step": 1, "phase": "Preparation", "role": "Finance Executive", "action": "Download bank statements for all accounts", "decision": False, "doc": "Bank Statements (PDF/CSV)", "control": "All accounts included", "sla": "Daily / Month-end", "escalation": "Finance Manager"},
            {"step": 2, "phase": "Preparation", "role": "Finance Executive", "action": "Export cash/bank ledger from Tally/ERP", "decision": False, "doc": "Cash & Bank Ledger", "control": "Date range match with bank statement", "sla": "Same day", "escalation": "Finance Manager"},
            {"step": 3, "phase": "Matching", "role": "Finance Executive", "action": "Match bank entries with book entries (auto + manual)", "decision": True, "yes": "All matched — proceed to sign-off", "no": "Identify unmatched items — go to Step 4", "doc": "BRS Working Sheet", "control": "Transaction-level matching", "sla": "T+1", "escalation": "Finance Manager"},
            {"step": 4, "phase": "Exception Handling", "role": "Finance Executive", "action": "Investigate unmatched items (timing differences, errors, omissions)", "decision": True, "yes": "Timing difference — note and carry forward", "no": "Error found — pass correction entry", "doc": "Exception Report", "control": "Aging of unmatched items", "sla": "T+2", "escalation": "Finance Manager"},
            {"step": 5, "phase": "Exception Handling", "role": "Finance Executive", "action": "Pass adjustment entries for identified errors", "decision": False, "doc": "Journal Voucher", "control": "Approval for adjustment entries", "sla": "T+2", "escalation": "Finance Manager"},
            {"step": 6, "phase": "Review", "role": "Finance Manager", "action": "Review BRS — verify outstanding items are valid", "decision": True, "yes": "BRS approved", "no": "Request further investigation", "doc": "Reviewed BRS", "control": "Outstanding item aging < 30 days target", "sla": "T+3", "escalation": "CFO"},
            {"step": 7, "phase": "Sign-off", "role": "Finance Manager", "action": "Sign-off BRS, archive with supporting documents", "decision": False, "doc": "Signed BRS", "control": "Monthly sign-off deadline", "sla": "T+3", "escalation": "CFO"},
        ],
        "raci_roles": ["Finance Exec", "Finance Manager", "CFO", "Auditor"],
        "raci": [
            ["Download Bank Statements", "R", "A", "", ""],
            ["Export Book Ledger",       "R", "A", "", ""],
            ["Match Entries",            "R", "A", "", ""],
            ["Investigate Exceptions",   "R", "A", "I", ""],
            ["Pass Adjustments",         "R", "A", "", "I"],
            ["Review BRS",               "I", "R", "A", "I"],
            ["Sign-off & Archive",       "I", "R", "A", "I"],
        ],
    }


def gst_workflow():
    """GST Compliance workflow steps."""
    return {
        "name": "GST Compliance",
        "steps": [
            {"step": 1, "phase": "Data Capture", "role": "Finance Executive", "action": "Ensure all sales and purchase invoices are GST-compliant", "decision": False, "doc": "Invoice Register", "control": "GSTIN validation, HSN/SAC codes", "sla": "Ongoing", "escalation": "Finance Manager"},
            {"step": 2, "phase": "Data Capture", "role": "Finance Executive", "action": "Reconcile purchase register with GSTR-2A/2B (ITC matching)", "decision": True, "yes": "ITC fully matched", "no": "Follow up with vendors for missing invoices", "doc": "GSTR-2A/2B, Purchase Register", "control": "ITC reconciliation report", "sla": "20th of next month", "escalation": "Finance Manager"},
            {"step": 3, "phase": "Filing", "role": "Finance Executive", "action": "Prepare and file GSTR-1 (outward supplies)", "decision": False, "doc": "GSTR-1", "control": "Sales register match", "sla": "11th of next month", "escalation": "Finance Manager"},
            {"step": 4, "phase": "Filing", "role": "Finance Manager", "action": "Review GSTR-3B data, verify tax liability calculation", "decision": True, "yes": "Approve for filing", "no": "Correct discrepancies and recompute", "doc": "GSTR-3B Working", "control": "Liability vs ITC reconciliation", "sla": "18th of next month", "escalation": "CFO"},
            {"step": 5, "phase": "Filing", "role": "Finance Executive", "action": "File GSTR-3B and make GST payment", "decision": False, "doc": "GSTR-3B, GST Challan", "control": "Payment confirmation from portal", "sla": "20th of next month", "escalation": "Finance Manager"},
            {"step": 6, "phase": "Annual", "role": "Finance Manager", "action": "Prepare GSTR-9 (Annual Return) and GSTR-9C (Reconciliation)", "decision": False, "doc": "GSTR-9, GSTR-9C", "control": "Auditor coordination", "sla": "31st December", "escalation": "CFO"},
            {"step": 7, "phase": "Reconciliation", "role": "Finance Manager", "action": "Reconcile GST credit ledger with books quarterly", "decision": True, "yes": "Balanced — no action needed", "no": "Investigate and adjust", "doc": "GST Credit Ledger, Books", "control": "Quarterly reconciliation", "sla": "Quarterly", "escalation": "CFO"},
        ],
        "raci_roles": ["Finance Exec", "Finance Manager", "CFO", "Tax Consultant", "Auditor"],
        "raci": [
            ["Invoice Compliance",     "R", "A", "", "C", ""],
            ["ITC Reconciliation",     "R", "A", "", "C", "I"],
            ["File GSTR-1",            "R", "A", "", "", ""],
            ["Review GSTR-3B",         "I", "R", "A", "C", ""],
            ["File GSTR-3B & Pay",     "R", "A", "", "", ""],
            ["Annual Return (GSTR-9)", "C", "R", "A", "C", "R"],
            ["Credit Ledger Recon",    "C", "R", "A", "", "I"],
        ],
    }


def budget_workflow():
    """Budget Planning workflow steps."""
    return {
        "name": "Budget Planning",
        "steps": [
            {"step": 1, "phase": "Planning", "role": "CFO", "action": "Issue budget preparation guidelines and timeline", "decision": False, "doc": "Budget Circular", "control": "Aligned with strategic plan", "sla": "Q3 (Oct)", "escalation": "CEO"},
            {"step": 2, "phase": "Collection", "role": "Department Heads", "action": "Prepare department-level budget proposals", "decision": False, "doc": "Department Budget Templates", "control": "Historical data comparison", "sla": "4 weeks from circular", "escalation": "CFO"},
            {"step": 3, "phase": "Collection", "role": "Finance Manager", "action": "Collect, compile, and consolidate all department budgets", "decision": False, "doc": "Consolidated Budget Draft", "control": "Completeness check (all depts)", "sla": "1 week after collection", "escalation": "CFO"},
            {"step": 4, "phase": "Analysis", "role": "Finance Manager", "action": "Analyze consolidated budget — identify gaps, overspends", "decision": False, "doc": "Budget Analysis Report", "control": "Revenue vs expense balance", "sla": "1 week", "escalation": "CFO"},
            {"step": 5, "phase": "Review", "role": "CFO", "action": "Review consolidated budget with department heads", "decision": True, "yes": "Budget acceptable — proceed to approval", "no": "Request revisions from departments", "doc": "Review Meeting Minutes", "control": "Strategic alignment check", "sla": "1-2 weeks", "escalation": "CEO"},
            {"step": 6, "phase": "Approval", "role": "CEO", "action": "Present budget to Board for approval", "decision": True, "yes": "Board approved — finalize", "no": "Revise per Board feedback", "doc": "Board Presentation", "control": "Board resolution required", "sla": "Board meeting", "escalation": "Board"},
            {"step": 7, "phase": "Finalization", "role": "Finance Manager", "action": "Finalize approved budget, distribute to departments", "decision": False, "doc": "Approved Budget", "control": "Version control & sign-off", "sla": "Before FY start", "escalation": "CFO"},
            {"step": 8, "phase": "Monitoring", "role": "Finance Manager", "action": "Track actuals vs budget monthly, report variances", "decision": True, "yes": "Within tolerance (±10%) — continue", "no": "Flag variance — require explanation from dept", "doc": "Budget vs Actuals Report", "control": "Monthly variance review", "sla": "Monthly", "escalation": "CFO"},
        ],
        "raci_roles": ["Dept Heads", "Finance Exec", "Finance Manager", "CFO", "CEO", "Board"],
        "raci": [
            ["Issue Guidelines",       "I", "", "I", "R", "A", ""],
            ["Prepare Dept Budgets",   "R", "", "C", "A", "", ""],
            ["Consolidate Budgets",    "", "C", "R", "A", "", ""],
            ["Analyze Budget",         "", "C", "R", "A", "", ""],
            ["Review with Dept Heads", "C", "", "C", "R", "A", ""],
            ["Board Approval",         "", "", "C", "C", "R", "A"],
            ["Finalize & Distribute",  "I", "", "R", "A", "", ""],
            ["Monthly Monitoring",     "C", "R", "R", "A", "I", ""],
        ],
    }


# ─── Excel Generation ────────────────────────────────────────────────────────

def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_cell(cell, wrap=True, border=True, center=False):
    """Apply standard styling to a cell."""
    if wrap:
        cell.alignment = WRAP_ALIGN if not center else CENTER_ALIGN
    if border:
        cell.border = THIN_BORDER


def apply_conditional_fills(ws, row, is_decision, has_control, has_escalation):
    """Apply conditional formatting fills to a row."""
    if is_decision:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = DECISION_FILL
    if has_control:
        ws.cell(row=row, column=9).fill = CONTROL_FILL  # Control column
    if has_escalation:
        ws.cell(row=row, column=11).fill = ESCALATION_FILL  # Escalation column


def create_flowchart_sheet(wb, workflow):
    """Create a flowchart sheet for a workflow."""
    name = workflow["name"]
    # Excel sheet names max 31 chars
    sheet_name = f"{name[:22]}_Flow" if len(name) > 22 else f"{name}_Flow"
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    headers = [
        "Step #", "Phase", "Responsible Role", "Action / Task",
        "Decision?", "If Yes →", "If No →",
        "Document / System", "Control Point", "SLA / Timeline", "Escalation To"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    # Data rows
    for i, step in enumerate(workflow["steps"], start=2):
        ws.cell(row=i, column=1, value=step["step"])
        ws.cell(row=i, column=2, value=step["phase"])
        ws.cell(row=i, column=3, value=step["role"])
        ws.cell(row=i, column=4, value=step["action"])
        ws.cell(row=i, column=5, value="YES" if step["decision"] else "")
        ws.cell(row=i, column=6, value=step.get("yes", ""))
        ws.cell(row=i, column=7, value=step.get("no", ""))
        ws.cell(row=i, column=8, value=step.get("doc", ""))
        ws.cell(row=i, column=9, value=step.get("control", ""))
        ws.cell(row=i, column=10, value=step.get("sla", ""))
        ws.cell(row=i, column=11, value=step.get("escalation", ""))

        # Style cells
        for col in range(1, 12):
            cell = ws.cell(row=i, column=col)
            style_cell(cell, center=(col in [1, 5]))

        apply_conditional_fills(ws, i, step["decision"], bool(step.get("control")), bool(step.get("escalation")))

    # Column widths
    col_widths = [8, 14, 20, 45, 10, 30, 30, 30, 30, 18, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    return sheet_name


def create_raci_sheet(wb, workflow):
    """Create a RACI matrix sheet for a workflow."""
    name = workflow["name"]
    sheet_name = f"{name[:22]}_RACI" if len(name) > 22 else f"{name}_RACI"
    ws = wb.create_sheet(title=sheet_name)

    roles = workflow.get("raci_roles", [])
    raci_data = workflow.get("raci", [])

    # Headers
    headers = ["Process Step"] + roles
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    # RACI fill map
    raci_fills = {"R": RACI_R_FILL, "A": RACI_A_FILL, "C": RACI_C_FILL, "I": RACI_I_FILL}
    raci_fonts = {
        "R": Font(bold=True, color="FFFFFF"),
        "A": Font(bold=True, color="FFFFFF"),
        "C": Font(bold=True, color="FFFFFF"),
        "I": Font(bold=True, color="000000"),
    }

    # Data rows
    for i, row_data in enumerate(raci_data, start=2):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            style_cell(cell, center=(col > 1))
            if col > 1 and value in raci_fills:
                cell.fill = raci_fills[value]
                cell.font = raci_fonts[value]

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Data validation for RACI values
    dv = DataValidation(type="list", formula1='"R,A,C,I,"', allow_blank=True)
    dv.error = "Please enter R, A, C, or I"
    dv.errorTitle = "Invalid RACI value"
    for col in range(2, len(headers) + 1):
        for row in range(2, len(raci_data) + 2):
            dv.add(ws.cell(row=row, column=col))
    ws.add_data_validation(dv)

    # RACI Legend
    legend_row = len(raci_data) + 4
    ws.cell(row=legend_row, column=1, value="RACI Legend:").font = Font(bold=True)
    legend_items = [
        ("R", "Responsible — Does the work", RACI_R_FILL),
        ("A", "Accountable — Owns the outcome", RACI_A_FILL),
        ("C", "Consulted — Provides input", RACI_C_FILL),
        ("I", "Informed — Kept in the loop", RACI_I_FILL),
    ]
    for j, (code, desc, fill) in enumerate(legend_items):
        row = legend_row + 1 + j
        cell_code = ws.cell(row=row, column=1, value=code)
        cell_code.fill = fill
        cell_code.font = Font(bold=True, color="FFFFFF" if code != "I" else "000000")
        cell_code.alignment = CENTER_ALIGN
        ws.cell(row=row, column=2, value=desc)

    ws.freeze_panes = "B2"
    return sheet_name


def create_index_sheet(wb, sheet_map, company_name):
    """Create an index/TOC sheet."""
    ws = wb.create_sheet(title="Index")
    wb.move_sheet(ws, offset=-len(wb.sheetnames) + 1)  # Move to first position

    # Title
    ws.cell(row=1, column=1, value=f"{company_name}").font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:D1")
    ws.cell(row=2, column=1, value="Finance & Accounting — Workflow Charts").font = Font(name="Calibri", bold=True, size=13, color="2F5496")
    ws.merge_cells("A2:D2")
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y')}").font = Font(italic=True, color="666666")

    # Headers
    headers = ["#", "Workflow", "Flowchart Sheet", "RACI Sheet"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = INDEX_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Workflow entries
    for i, (wf_key, sheets) in enumerate(sheet_map.items(), start=1):
        row = 5 + i
        ws.cell(row=row, column=1, value=i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=wf_key).border = THIN_BORDER

        flow_name = sheets.get("flow", "")
        raci_name = sheets.get("raci", "")

        flow_cell = ws.cell(row=row, column=3, value=flow_name)
        flow_cell.font = Font(color="0563C1", underline="single")
        flow_cell.border = THIN_BORDER

        raci_cell = ws.cell(row=row, column=4, value=raci_name)
        raci_cell.font = Font(color="0563C1", underline="single")
        raci_cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30

    ws.freeze_panes = "A6"


def create_approval_matrix_sheet(wb, thresholds):
    """Create the payment approval matrix sheet."""
    ws = wb.create_sheet(title="Approval_Matrix")

    # Title
    ws.cell(row=1, column=1, value="Payment Approval Matrix").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:E1")

    # Headers
    headers = ["Tier", "Amount Range (₹)", "Primary Approver", "Description", "Additional Controls"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    # Data
    sorted_tiers = sorted(thresholds.items(), key=lambda x: x[1].get("limit") or float("inf"))
    prev_limit = 0
    for i, (tier, info) in enumerate(sorted_tiers, start=4):
        limit = info["limit"]
        if limit is None:
            range_str = f"Above ₹{prev_limit:,}"
        else:
            range_str = f"₹{prev_limit:,} — ₹{limit:,}" if prev_limit > 0 else f"Up to ₹{limit:,}"

        controls = ""
        if limit and limit > 50000:
            controls = "Dual signatory required"
        if limit is None:
            controls = "Board resolution + Dual signatory"

        ws.cell(row=i, column=1, value=tier.replace("_", " ").title())
        ws.cell(row=i, column=2, value=range_str)
        ws.cell(row=i, column=3, value=info["approver"])
        ws.cell(row=i, column=4, value=info.get("description", ""))
        ws.cell(row=i, column=5, value=controls)

        for col in range(1, 6):
            style_cell(ws.cell(row=i, column=col))

        prev_limit = limit if limit else prev_limit

    # Column widths
    col_widths = [18, 25, 22, 35, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


def create_controls_summary_sheet(wb, workflows):
    """Create a controls summary sheet across all workflows."""
    ws = wb.create_sheet(title="Controls_Summary")

    ws.cell(row=1, column=1, value="Internal Controls Summary").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:E1")

    headers = ["#", "Workflow", "Step", "Control Point", "Type"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    row = 4
    counter = 1
    for wf_key, wf_data in workflows.items():
        for step in wf_data["steps"]:
            if step.get("control"):
                ws.cell(row=row, column=1, value=counter)
                ws.cell(row=row, column=2, value=wf_data["name"])
                ws.cell(row=row, column=3, value=step["action"][:60])
                ws.cell(row=row, column=4, value=step["control"])

                # Determine control type
                control_text = step["control"].lower()
                if any(w in control_text for w in ["approval", "review", "sign-off", "authorize"]):
                    ctrl_type = "Authorization"
                elif any(w in control_text for w in ["reconcil", "match", "verify", "validation"]):
                    ctrl_type = "Verification"
                elif any(w in control_text for w in ["segregat", "dual", "independent"]):
                    ctrl_type = "Segregation of Duties"
                elif any(w in control_text for w in ["monitor", "track", "aging", "deadline"]):
                    ctrl_type = "Monitoring"
                else:
                    ctrl_type = "Preventive"

                ws.cell(row=row, column=5, value=ctrl_type)

                for col in range(1, 6):
                    style_cell(ws.cell(row=row, column=col))

                counter += 1
                row += 1

    col_widths = [6, 25, 45, 40, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


def create_kpi_sheet(wb):
    """Create KPI metrics sheet."""
    ws = wb.create_sheet(title="KPI_Metrics")

    ws.cell(row=1, column=1, value="F&A Key Performance Indicators").font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:E1")

    headers = ["#", "KPI", "Workflow", "Target", "Measurement Frequency"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    kpis = [
        ("Days Payable Outstanding (DPO)", "AP", "≤ 30 days", "Monthly"),
        ("Invoice Processing Time", "AP", "≤ 3 working days", "Weekly"),
        ("3-Way Match Rate", "AP", "≥ 98%", "Monthly"),
        ("Days Sales Outstanding (DSO)", "AR", "≤ 45 days", "Monthly"),
        ("Collection Efficiency", "AR", "≥ 95%", "Monthly"),
        ("AR Aging > 90 Days", "AR", "< 5% of total AR", "Monthly"),
        ("Month-End Close Time", "Month-End Close", "≤ T+5 working days", "Monthly"),
        ("BRS Outstanding Items", "Bank Reconciliation", "Zero items > 30 days", "Monthly"),
        ("Expense Claim Processing Time", "Expense Reimbursement", "≤ 7 working days", "Weekly"),
        ("Payroll Accuracy", "Payroll", "100% (zero errors)", "Monthly"),
        ("Payroll Processing Time", "Payroll", "≤ 5th of month", "Monthly"),
        ("Statutory Filing Compliance", "GST / Payroll", "100% on-time filing", "Monthly"),
        ("GST ITC Match Rate", "GST Compliance", "≥ 98%", "Monthly"),
        ("Budget Variance", "Budget Planning", "Within ±10%", "Monthly"),
        ("Financial Report Delivery", "Financial Reporting", "≤ T+5 working days", "Monthly"),
    ]

    for i, (kpi, workflow, target, freq) in enumerate(kpis, start=4):
        ws.cell(row=i, column=1, value=i - 3)
        ws.cell(row=i, column=2, value=kpi)
        ws.cell(row=i, column=3, value=workflow)
        ws.cell(row=i, column=4, value=target)
        ws.cell(row=i, column=5, value=freq)
        for col in range(1, 6):
            style_cell(ws.cell(row=i, column=col))

    col_widths = [6, 35, 25, 25, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"


# ─── Main Generator ──────────────────────────────────────────────────────────

def generate_workbook(
    company_name: str = "Fracktal Works Private Limited",
    selected_workflows: list = None,
    thresholds: dict = None,
    output_path: str = None,
):
    """Generate the complete F&A workflow charts workbook."""
    TMP_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(TMP_DIR / f"FA_Workflow_Charts_{timestamp}.xlsx")

    if thresholds is None:
        thresholds = DEFAULT_THRESHOLDS

    all_workflows = get_all_workflows()

    # Filter workflows if specified
    if selected_workflows:
        # Normalize keys for matching
        key_map = {k.lower().replace(" ", "").replace("-", ""): k for k in all_workflows}
        filtered = {}
        for sel in selected_workflows:
            normalized = sel.strip().lower().replace(" ", "").replace("-", "")
            if normalized in key_map:
                real_key = key_map[normalized]
                filtered[real_key] = all_workflows[real_key]
            else:
                print(f"  ⚠ Unknown workflow: '{sel}' — skipping")
        workflows = filtered
    else:
        workflows = all_workflows

    if not workflows:
        print("ERROR: No valid workflows selected.")
        sys.exit(1)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print(f"\n{'='*60}")
    print(f"GENERATING F&A WORKFLOW CHARTS")
    print(f"Company: {company_name}")
    print(f"Workflows: {len(workflows)}")
    print(f"{'='*60}\n")

    # Generate flowchart and RACI sheets for each workflow
    sheet_map = {}
    for wf_key, wf_data in workflows.items():
        print(f"  → {wf_data['name']}...")
        flow_sheet = create_flowchart_sheet(wb, wf_data)
        raci_sheet = create_raci_sheet(wb, wf_data)
        sheet_map[wf_data["name"]] = {"flow": flow_sheet, "raci": raci_sheet}
        print(f"    ✓ Flowchart: {flow_sheet}")
        print(f"    ✓ RACI: {raci_sheet}")

    # Create summary sheets
    print(f"\n  → Summary sheets...")
    create_index_sheet(wb, sheet_map, company_name)
    print(f"    ✓ Index (Table of Contents)")

    create_approval_matrix_sheet(wb, thresholds)
    print(f"    ✓ Approval Matrix")

    create_controls_summary_sheet(wb, workflows)
    print(f"    ✓ Controls Summary")

    create_kpi_sheet(wb)
    print(f"    ✓ KPI Metrics")

    # Save
    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"✓ WORKBOOK SAVED: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Workflows: {len(workflows)}")
    print(f"{'='*60}\n")

    return output_path


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate F&A workflow charts as Excel workbook"
    )
    parser.add_argument(
        "--company",
        default="Fracktal Works Private Limited",
        help="Company name (default: Fracktal Works Private Limited)"
    )
    parser.add_argument(
        "--workflows",
        default=None,
        help="Comma-separated workflow names (default: all). Options: AP, AR, Month-End Close, Expense Reimbursement, Payment Approval, Financial Reporting, Payroll, Bank Reconciliation, GST Compliance, Budget Planning"
    )
    parser.add_argument(
        "--thresholds",
        default=None,
        help="JSON string with custom approval thresholds"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output file path (default: .tmp/FA_Workflow_Charts_<timestamp>.xlsx)"
    )
    parser.add_argument(
        "--list-workflows",
        action="store_true",
        help="List all available workflows"
    )

    args = parser.parse_args()

    if args.list_workflows:
        all_wf = get_all_workflows()
        print("\nAvailable Workflows:")
        for key, wf in all_wf.items():
            print(f"  • {key}: {wf['name']} ({len(wf['steps'])} steps)")
        return

    selected = None
    if args.workflows:
        selected = [w.strip() for w in args.workflows.split(",")]

    thresholds = None
    if args.thresholds:
        try:
            thresholds = json.loads(args.thresholds)
        except json.JSONDecodeError as e:
            print(f"ERROR: Invalid JSON for thresholds: {e}")
            sys.exit(1)

    generate_workbook(
        company_name=args.company,
        selected_workflows=selected,
        thresholds=thresholds,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
