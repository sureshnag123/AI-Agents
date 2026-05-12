"""
Generate a test Excel file with 10 realistic March 2026 entries
for THOTA HOSPITALITY LLP — based on actual expense patterns.
"""
import warnings
warnings.filterwarnings('ignore')
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = SCRIPT_DIR.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(parents=True, exist_ok=True)

# Load the template to get structure + reference sheets
template_path = TMP_DIR / "Thota_Tally_Upload_Template.xlsx"
wb = openpyxl.load_workbook(str(template_path))
ws = wb["Expense_Entries"]

# Clear sample rows (rows 2-7)
for row_idx in range(2, 8):
    for col_idx in range(1, 14):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = None
        cell.fill = PatternFill()  # clear fill

# 10 realistic test entries for March 2026
test_entries = [
    # Row, Date, Voucher_Type, Department, Description, Expense_Ledger, Vendor_Name, Amount, GST_Rate, GSTIN, Bill_Number, Paid_By, Payment_Mode, Comment
    (datetime(2026, 3, 1), "Journal", "Thota Kitchen", "Vegetables from local market",
     "Thota Kitchen", "VENDOR for Non GST Purchases", 580, 0, "", "", "Niveditha", "Cash",
     "No GST bill"),

    (datetime(2026, 3, 1), "Purchase", "Thota Kitchen", "Chicken from Meat Shop",
     "Thota Kitchen", "Meat Shop", 1650, 5, "29AABCM1234A1Z5", "MS-4521", "Niveditha", "Cash",
     "GST invoice available"),

    (datetime(2026, 3, 2), "Journal", "Thota Decor", "Flowers for hall decoration",
     "Thota Decor", "VENDOR for Non GST Purchases", 450, 0, "", "", "Account", "Cash",
     "Local flower merchant"),

    (datetime(2026, 3, 2), "Purchase", "Logistics", "Pickup vehicle fuel",
     "Fuel Expenses", "Petrol Bunk", 1500, 18, "29AABCP5678B2Z1", "PB-7890", "Account", "UPI",
     ""),

    (datetime(2026, 3, 3), "Purchase", "Thota Kitchen", "Monthly grocery from Big Market",
     "Thota Kitchen", "The Big Market", 2850, 12, "29AABCT9012C3Z2", "BM-2026-312", "Akshata", "Bank",
     "Monthly grocery stock"),

    (datetime(2026, 3, 3), "Journal", "Staff Welfare", "Staff meals - chicken",
     "Staff Welfare", "VENDOR for Non GST Purchases", 370, 0, "", "", "Niveditha", "Cash",
     "Local non-GST vendor"),

    (datetime(2026, 3, 4), "Journal", "Thota Kitchen", "Eggs and curd from provision store",
     "Thota Kitchen", "VENDOR for Non GST Purchases", 275, 0, "", "", "Niveditha", "Cash",
     "Manjunath provision store"),

    (datetime(2026, 3, 4), "Purchase", "Thota Kitchen", "Swiggy Instamart grocery order",
     "Thota Kitchen", "Swiggy Instamart", 1120, 18, "29AABCS3456D4Z3", "SWG-98765", "Account", "Bank",
     "Deducted from bank"),

    (datetime(2026, 3, 5), "Journal", "Maintenance", "Plumbing repair - kitchen sink",
     "Repairs & Maintenance", "VENDOR for Non GST Purchases", 800, 0, "", "", "Account", "Cash",
     "Local plumber - no bill"),

    (datetime(2026, 3, 5), "Purchase", "Thota Kitchen", "Monthly milk bill - Feb",
     "Thota Kitchen", "Akshay - Milk", 3915, 0, "", "MILK-FEB26", "Account", "Bank",
     "Monthly settlement"),
]

# Write entries
for row_idx, data in enumerate(test_entries, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if col_idx == 1:  # Date
            cell.number_format = "DD-MM-YYYY"
        elif col_idx == 7:  # Amount
            cell.number_format = '#,##0.00'

output_path = TMP_DIR / "Thota_TEST_March2026.xlsx"
wb.save(str(output_path))
print(f"✅ Test file created: {output_path}")
print(f"   10 entries: March 1-5, 2026")

# Quick summary
journal_count = sum(1 for e in test_entries if e[1] == "Journal")
purchase_count = sum(1 for e in test_entries if e[1] == "Purchase")
total = sum(e[6] for e in test_entries)
print(f"   Journal: {journal_count}, Purchase: {purchase_count}")
print(f"   Total: ₹{total:,.2f}")
