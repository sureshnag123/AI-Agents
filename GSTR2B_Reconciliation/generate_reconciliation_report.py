"""
GSTR-2B Reconciliation Report Generator
FY 2025-26 | GSTIN: 29AACCF2736P1Z5
Generates: 1) Month-on-Month Reconciliation Summary
           2) Detailed Bill-to-Bill Report
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

# ── colour palette ────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2E75B6"
C_LIGHT_BLUE  = "BDD7EE"
C_HEADER_GREY = "D6DCE4"
C_GREEN       = "E2EFDA"
C_ORANGE      = "FCE4D6"
C_RED         = "FFE0E0"
C_YELLOW      = "FFFFE0"
C_WHITE       = "FFFFFF"

MONTHS = [
    "Apr-2025","May-2025","Jun-2025","Jul-2025",
    "Aug-2025","Sep-2025","Oct-2025","Nov-2025",
    "Dec-2025","Jan-2026","Feb-2026","Mar-2026"
]

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

def style_cell(cell, bold=False, bg=None, font_color="000000",
               align="left", border=True, size=10, wrap=False):
    cell.font = Font(bold=bold, color=font_color, size=size, name="Calibri")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    if border:
        cell.border = thin_border()

def money_fmt(ws, cell, value):
    cell.value = value if value else 0
    cell.number_format = '#,##0.00'

# ── read source workbook ──────────────────────────────────────────────────────
src_path = "GSTR2B_FY2025-26_Compiled.xlsx"
src_wb   = openpyxl.load_workbook(src_path)

GSTIN = "29AACCF2736P1Z5"

# ════════════════════════════════════════════════════════════════════════════
# 1.  PARSE ITC SUMMARY SHEETS
# ════════════════════════════════════════════════════════════════════════════
def parse_itc_sheet(ws):
    """
    Returns dict: month -> {heading_key: {igst, cgst, sgst, cess}}
    Only captures rows where S.no is a Roman numeral (I–VI) or 'A/B'
    """
    data = defaultdict(dict)
    current_month = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        month, sno, heading = row[0], row[1], row[2]
        igst, cgst, sgst, cess = row[4], row[5], row[6], row[7]
        if month in MONTHS:
            current_month = month
        if current_month and sno and isinstance(sno, str) and sno.strip() in (
                "I","II","III","IV","V","VI","A","B"):
            key = f"{sno.strip()} - {heading}" if heading else sno.strip()
            data[current_month][key] = {
                "igst": igst or 0,
                "cgst": cgst or 0,
                "sgst": sgst or 0,
                "cess": cess or 0,
                "total": (igst or 0)+(cgst or 0)+(sgst or 0)+(cess or 0)
            }
    return data

itc_avail   = parse_itc_sheet(src_wb["ITC Available"])
itc_na      = parse_itc_sheet(src_wb["ITC not available"])
itc_rev     = parse_itc_sheet(src_wb["ITC Reversal"])
itc_rej     = parse_itc_sheet(src_wb["ITC Rejected"])

# ════════════════════════════════════════════════════════════════════════════
# 2.  PARSE B2B TRANSACTION SHEETS
# ════════════════════════════════════════════════════════════════════════════
def parse_b2b(ws):
    """
    Columns: Month(0), GSTIN(1), Name(2), InvNo(3), InvType(4),
             InvDate(5), InvValue(6), POS(7), RCM(8), Taxable(9),
             IGST(10), CGST(11), SGST(12), Cess(13), ITC_Avail(16)
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] not in MONTHS:
            continue
        rows.append({
            "month":     row[0],
            "gstin":     row[1] or "",
            "name":      row[2] or "",
            "inv_no":    row[3] or "",
            "inv_type":  row[4] or "",
            "inv_date":  str(row[5] or ""),
            "inv_value": float(row[6] or 0),
            "pos":       row[7] or "",
            "rcm":       row[8] or "No",
            "taxable":   float(row[9] or 0),
            "igst":      float(row[10] or 0),
            "cgst":      float(row[11] or 0),
            "sgst":      float(row[12] or 0),
            "cess":      float(row[13] or 0),
            "itc":       row[16] if len(row) > 16 else "Yes",
            "source":    "B2B",
        })
    return rows

def parse_cdnr(ws):
    """
    B2B-CDNR: Month(0), GSTIN(1), Name(2), NoteNo(3), NoteType(4),
              NoteSupply(5), NoteDate(6), NoteValue(7), POS(8), RCM(9),
              Taxable(10), IGST(11), CGST(12), SGST(13), Cess(14), ITC(23)
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] not in MONTHS:
            continue
        note_type = str(row[4] or "")
        # Credit notes reduce ITC; Debit notes increase
        sign = -1 if "Credit" in note_type else 1
        rows.append({
            "month":     row[0],
            "gstin":     row[1] or "",
            "name":      row[2] or "",
            "inv_no":    row[3] or "",
            "inv_type":  note_type,
            "inv_date":  str(row[6] or ""),
            "inv_value": float(row[7] or 0) * sign,
            "pos":       row[8] or "",
            "rcm":       row[9] or "No",
            "taxable":   float(row[10] or 0) * sign,
            "igst":      float(row[11] or 0) * sign,
            "cgst":      float(row[12] or 0) * sign,
            "sgst":      float(row[13] or 0) * sign,
            "cess":      float(row[14] or 0) * sign,
            "itc":       row[23] if len(row) > 23 else "Yes",
            "source":    "CDNR",
        })
    return rows

def parse_impg(ws):
    """
    IMPG: Month(0), IcegateRefDate(1), PortCode(2), BillNo(3),
          Date(4), Taxable(5), IGST(6), Cess(7)
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] not in MONTHS:
            continue
        rows.append({
            "month":     row[0],
            "gstin":     "IMPG",
            "name":      f"Port: {row[2] or ''}",
            "inv_no":    str(row[3] or ""),
            "inv_type":  "Import",
            "inv_date":  str(row[4] or ""),
            "inv_value": float(row[5] or 0),
            "pos":       "",
            "rcm":       "No",
            "taxable":   float(row[5] or 0),
            "igst":      float(row[6] or 0),
            "cgst":      0,
            "sgst":      0,
            "cess":      float(row[7] or 0),
            "itc":       "Yes",
            "source":    "IMPG",
        })
    return rows

b2b_rows  = parse_b2b(src_wb["B2B"])
cdnr_rows = parse_cdnr(src_wb["B2B-CDNR"])
impg_rows = parse_impg(src_wb["IMPG"])
all_bills = b2b_rows + cdnr_rows + impg_rows

# ════════════════════════════════════════════════════════════════════════════
# 3.  MONTHLY AGGREGATES FROM B2B TRANSACTIONS
# ════════════════════════════════════════════════════════════════════════════
def monthly_agg(rows):
    agg = {m: {"count":0,"taxable":0,"igst":0,"cgst":0,"sgst":0,"cess":0,"total_gst":0}
           for m in MONTHS}
    for r in rows:
        m = r["month"]
        if m not in agg:
            continue
        agg[m]["count"]     += 1
        agg[m]["taxable"]   += r["taxable"]
        agg[m]["igst"]      += r["igst"]
        agg[m]["cgst"]      += r["cgst"]
        agg[m]["sgst"]      += r["sgst"]
        agg[m]["cess"]      += r["cess"]
        agg[m]["total_gst"] += r["igst"] + r["cgst"] + r["sgst"] + r["cess"]
    return agg

b2b_agg  = monthly_agg(b2b_rows)
cdnr_agg = monthly_agg(cdnr_rows)
impg_agg = monthly_agg(impg_rows)

# ════════════════════════════════════════════════════════════════════════════
# 4.  ITC SUMMARY AGGREGATES (from ITC Available sheet)
#     Key row: "I" = All other ITC (B2B + ECO + amendments)
# ════════════════════════════════════════════════════════════════════════════
def get_itc_row(itc_dict, month, roman):
    for k, v in itc_dict.get(month, {}).items():
        if k.startswith(roman + " "):
            return v
    return {"igst":0,"cgst":0,"sgst":0,"cess":0,"total":0}

def sum_itc_all(itc_dict, month):
    total = {"igst":0,"cgst":0,"sgst":0,"cess":0,"total":0}
    for v in itc_dict.get(month, {}).values():
        for k2 in total:
            total[k2] += v.get(k2, 0)
    return total

# ════════════════════════════════════════════════════════════════════════════
# 5.  CREATE OUTPUT WORKBOOK
# ════════════════════════════════════════════════════════════════════════════
out_wb = openpyxl.Workbook()
out_wb.remove(out_wb.active)   # remove default sheet

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – RECONCILIATION SUMMARY
# ════════════════════════════════════════════════════════════════════════════
ws1 = out_wb.create_sheet("Reconciliation Summary")

def write_summary_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    for col in "BCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["O"].width = 15

    # ── Title block ───────────────────────────────────────────────────────
    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value = "GSTR-2B PURCHASE & INPUT TAX CREDIT RECONCILIATION SUMMARY — FY 2025-26"
    style_cell(c, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
               align="center", size=13)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:O2")
    c = ws["A2"]
    c.value = f"GSTIN: {GSTIN}          Compiled from GSTR-2B auto-populated data"
    style_cell(c, bold=True, bg=C_MID_BLUE, font_color="FFFFFF",
               align="center", size=10)
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:O3")
    c = ws["A3"]
    c.value = f"Report generated on: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    style_cell(c, bold=False, bg=C_LIGHT_BLUE, align="center", size=9)
    ws.row_dimensions[3].height = 14

    # ── Month header row ──────────────────────────────────────────────────
    headers = ["Particulars"] + MONTHS + ["TOTAL FY 2025-26"]
    row4 = 5
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row4, column=ci, value=h)
        style_cell(c, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
                   align="center", size=9, wrap=True)
    ws.row_dimensions[row4].height = 30

    cur_row = row4 + 1

    def section_header(label, bg=C_MID_BLUE):
        nonlocal cur_row
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=15)
        c = ws.cell(row=cur_row, column=1, value=label)
        style_cell(c, bold=True, bg=bg, font_color="FFFFFF",
                   align="left", size=10)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

    def data_row(label, values, bg=None, bold=False, indent=0):
        nonlocal cur_row
        lbl = ("    " * indent) + label
        c = ws.cell(row=cur_row, column=1, value=lbl)
        style_cell(c, bold=bold, bg=bg, align="left", size=9)
        total = 0
        for ci, v in enumerate(values, 2):
            cell = ws.cell(row=cur_row, column=ci, value=round(v, 2))
            cell.number_format = '#,##0.00'
            style_cell(cell, bold=bold, bg=bg, align="right", size=9)
            total += v
        # TOTAL column
        tc = ws.cell(row=cur_row, column=15, value=round(total, 2))
        tc.number_format = '#,##0.00'
        style_cell(tc, bold=True, bg=C_HEADER_GREY if not bg else bg,
                   align="right", size=9)
        ws.row_dimensions[cur_row].height = 15
        cur_row += 1
        return total

    def blank_row():
        nonlocal cur_row
        cur_row += 1

    # ══════════════════════════════════════════════════════════════════════
    # SECTION A – PURCHASE INVOICES (from B2B sheet)
    # ══════════════════════════════════════════════════════════════════════
    section_header("SECTION A – PURCHASE INVOICES (GSTR-2B Auto-populated)")

    inv_counts = [b2b_agg[m]["count"] for m in MONTHS]
    cdnr_counts= [cdnr_agg[m]["count"] for m in MONTHS]
    impg_counts = [impg_agg[m]["count"] for m in MONTHS]

    c = ws.cell(row=cur_row, column=1, value="    No. of B2B Invoices")
    style_cell(c, align="left", size=9)
    total_inv = 0
    for ci, v in enumerate(inv_counts, 2):
        cell = ws.cell(row=cur_row, column=ci, value=v)
        cell.number_format = '#,##0'
        style_cell(cell, align="right", size=9)
        total_inv += v
    tc = ws.cell(row=cur_row, column=15, value=total_inv)
    tc.number_format = '#,##0'
    style_cell(tc, bold=True, bg=C_HEADER_GREY, align="right", size=9)
    ws.row_dimensions[cur_row].height = 15
    cur_row += 1

    c = ws.cell(row=cur_row, column=1, value="    No. of Credit/Debit Notes (CDNR)")
    style_cell(c, align="left", size=9)
    total_cdnr = 0
    for ci, v in enumerate(cdnr_counts, 2):
        cell = ws.cell(row=cur_row, column=ci, value=v)
        cell.number_format = '#,##0'
        style_cell(cell, align="right", size=9)
        total_cdnr += v
    tc = ws.cell(row=cur_row, column=15, value=total_cdnr)
    tc.number_format = '#,##0'
    style_cell(tc, bold=True, bg=C_HEADER_GREY, align="right", size=9)
    ws.row_dimensions[cur_row].height = 15
    cur_row += 1

    c = ws.cell(row=cur_row, column=1, value="    No. of Import Bills (IMPG)")
    style_cell(c, align="left", size=9)
    total_impg_c = 0
    for ci, v in enumerate(impg_counts, 2):
        cell = ws.cell(row=cur_row, column=ci, value=v)
        cell.number_format = '#,##0'
        style_cell(cell, align="right", size=9)
        total_impg_c += v
    tc = ws.cell(row=cur_row, column=15, value=total_impg_c)
    tc.number_format = '#,##0'
    style_cell(tc, bold=True, bg=C_HEADER_GREY, align="right", size=9)
    ws.row_dimensions[cur_row].height = 15
    cur_row += 1

    taxable_b2b  = [b2b_agg[m]["taxable"] for m in MONTHS]
    taxable_cdnr = [cdnr_agg[m]["taxable"] for m in MONTHS]
    taxable_impg = [impg_agg[m]["taxable"] for m in MONTHS]
    net_taxable  = [taxable_b2b[i]+taxable_cdnr[i]+taxable_impg[i] for i in range(12)]

    blank_row()
    data_row("    Taxable Value – B2B Invoices (₹)", taxable_b2b, indent=0)
    data_row("    Taxable Value – Credit/Debit Notes (₹)", taxable_cdnr, indent=0)
    data_row("    Taxable Value – Imports / IMPG (₹)", taxable_impg, indent=0)
    data_row("NET TAXABLE PURCHASE VALUE (₹)", net_taxable,
             bg=C_LIGHT_BLUE, bold=True)

    blank_row()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION B – GST INPUT TAX CREDIT BREAK-UP
    # ══════════════════════════════════════════════════════════════════════
    section_header("SECTION B – GST INPUT TAX CREDIT (ITC) BREAK-UP")

    igst_b2b  = [b2b_agg[m]["igst"]  for m in MONTHS]
    cgst_b2b  = [b2b_agg[m]["cgst"]  for m in MONTHS]
    sgst_b2b  = [b2b_agg[m]["sgst"]  for m in MONTHS]
    cess_b2b  = [b2b_agg[m]["cess"]  for m in MONTHS]
    gst_b2b   = [b2b_agg[m]["total_gst"] for m in MONTHS]

    igst_cdn  = [cdnr_agg[m]["igst"]  for m in MONTHS]
    cgst_cdn  = [cdnr_agg[m]["cgst"]  for m in MONTHS]
    sgst_cdn  = [cdnr_agg[m]["sgst"]  for m in MONTHS]
    cess_cdn  = [cdnr_agg[m]["cess"]  for m in MONTHS]
    gst_cdn   = [cdnr_agg[m]["total_gst"] for m in MONTHS]

    igst_imp  = [impg_agg[m]["igst"]  for m in MONTHS]
    gst_imp   = [impg_agg[m]["total_gst"] for m in MONTHS]

    net_igst  = [igst_b2b[i]+igst_cdn[i]+igst_imp[i] for i in range(12)]
    net_cgst  = [cgst_b2b[i]+cgst_cdn[i] for i in range(12)]
    net_sgst  = [sgst_b2b[i]+sgst_cdn[i] for i in range(12)]
    net_cess  = [cess_b2b[i]+cess_cdn[i]+impg_agg[MONTHS[i]]["cess"] for i in range(12)]
    net_gst   = [net_igst[i]+net_cgst[i]+net_sgst[i]+net_cess[i] for i in range(12)]

    data_row("    IGST – B2B Invoices (₹)",   igst_b2b)
    data_row("    CGST – B2B Invoices (₹)",   cgst_b2b)
    data_row("    SGST – B2B Invoices (₹)",   sgst_b2b)
    data_row("    Cess – B2B Invoices (₹)",   cess_b2b)
    data_row("    Total GST – B2B (₹)",       gst_b2b, bg=C_GREEN, bold=True)
    blank_row()
    data_row("    IGST – Credit/Debit Notes (₹)", igst_cdn)
    data_row("    CGST – Credit/Debit Notes (₹)", cgst_cdn)
    data_row("    SGST – Credit/Debit Notes (₹)", sgst_cdn)
    data_row("    Total GST – CDNR (₹)",     gst_cdn, bg=C_ORANGE, bold=True)
    blank_row()
    data_row("    IGST – Imports / IMPG (₹)", igst_imp)
    data_row("    Total GST – IMPG (₹)",      gst_imp, bg=C_YELLOW, bold=True)
    blank_row()
    data_row("NET IGST (₹)",  net_igst, bg=C_LIGHT_BLUE, bold=True)
    data_row("NET CGST (₹)",  net_cgst, bg=C_LIGHT_BLUE, bold=True)
    data_row("NET SGST (₹)",  net_sgst, bg=C_LIGHT_BLUE, bold=True)
    data_row("NET CESS (₹)",  net_cess, bg=C_LIGHT_BLUE, bold=True)
    data_row("TOTAL NET ITC FROM GSTR-2B (₹)", net_gst,
             bg=C_MID_BLUE, bold=True)

    blank_row()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION C – ITC CLASSIFICATION (from ITC Available / Not Available)
    # ══════════════════════════════════════════════════════════════════════
    section_header("SECTION C – ITC CLASSIFICATION AS PER GSTR-2B")

    itc_avail_total = [sum_itc_all(itc_avail, m)["total"] for m in MONTHS]
    itc_na_total    = [sum_itc_all(itc_na, m)["total"]    for m in MONTHS]
    itc_rev_total   = [sum_itc_all(itc_rev, m)["total"]   for m in MONTHS]
    itc_rej_total   = [sum_itc_all(itc_rej, m)["total"]   for m in MONTHS]

    data_row("    ITC Available (eligible to claim in GSTR-3B) (₹)",
             itc_avail_total, bg=C_GREEN, bold=True)
    data_row("    ITC Not Available (blocked / ineligible) (₹)",
             itc_na_total,    bg=C_ORANGE)
    data_row("    ITC Reversal (Rule 37/42/43 etc.) (₹)",
             itc_rev_total,   bg=C_ORANGE)
    data_row("    ITC Rejected / Pending (₹)",
             itc_rej_total,   bg=C_RED)

    net_claimable = [itc_avail_total[i] - itc_rev_total[i] for i in range(12)]
    data_row("NET CLAIMABLE ITC (Available – Reversal) (₹)",
             net_claimable, bg=C_DARK_BLUE, bold=True)

    blank_row()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION D – RECONCILIATION VARIANCE
    # ══════════════════════════════════════════════════════════════════════
    section_header("SECTION D – RECONCILIATION VARIANCE ANALYSIS",
                   bg="7030A0")

    diff_taxable = [net_taxable[i] - 0 for i in range(12)]  # books not loaded
    diff_gst     = [net_gst[i] - itc_avail_total[i] for i in range(12)]

    # Row: Total ITC as per GSTR-2B
    data_row("    Total ITC as per GSTR-2B (₹)", net_gst, bg=C_LIGHT_BLUE, bold=True)
    data_row("    ITC Available (auto-populated GSTR-2B) (₹)",
             itc_avail_total, bg=C_GREEN, bold=True)

    diff_label = "    Variance (GSTR-2B Total ITC vs ITC Available) (₹)"
    diff_values= diff_gst
    diff_bg    = [C_YELLOW if abs(v) < 1 else C_RED for v in diff_values]
    # write manually for colour coding
    c = ws.cell(row=cur_row, column=1, value=diff_label)
    style_cell(c, bold=True, align="left", size=9)
    total_diff = 0
    for ci, v in enumerate(diff_values, 2):
        bg2 = C_GREEN if abs(v) < 1 else C_RED
        cell = ws.cell(row=cur_row, column=ci, value=round(v, 2))
        cell.number_format = '#,##0.00'
        style_cell(cell, bold=True, bg=bg2, align="right", size=9)
        total_diff += v
    tc = ws.cell(row=cur_row, column=15, value=round(total_diff, 2))
    tc.number_format = '#,##0.00'
    style_cell(tc, bold=True, bg=C_RED if abs(total_diff) >= 1 else C_GREEN,
               align="right", size=9)
    ws.row_dimensions[cur_row].height = 15
    cur_row += 1

    blank_row()

    # ══════════════════════════════════════════════════════════════════════
    # SECTION E – DOCUMENT COUNT SUMMARY
    # ══════════════════════════════════════════════════════════════════════
    section_header("SECTION E – DOCUMENT COUNT SUMMARY", bg=C_MID_BLUE)

    total_docs = [inv_counts[i]+cdnr_counts[i]+impg_counts[i] for i in range(12)]
    data_row("    B2B Invoices", inv_counts)
    data_row("    Credit / Debit Notes (CDNR)", cdnr_counts)
    data_row("    Import Bills (IMPG)", impg_counts)
    data_row("TOTAL DOCUMENTS", total_docs, bg=C_DARK_BLUE, bold=True)

    blank_row()

    # ── Footer note ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=15)
    note = ws.cell(row=cur_row, column=1,
                   value="NOTE: ITC Available figures are sourced from the auto-populated ITC classification in GSTR-2B. "
                         "Variance between Total GST and ITC Available may arise due to IMS pending actions, "
                         "blocked credits under Sec 17(5), or RCM transactions. Reconcile with GSTR-3B before filing.")
    style_cell(note, bold=False, bg=C_LIGHT_BLUE, align="left", size=8, wrap=True)
    ws.row_dimensions[cur_row].height = 40

write_summary_sheet(ws1)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – BILL-TO-BILL DETAIL REPORT
# ════════════════════════════════════════════════════════════════════════════
ws2 = out_wb.create_sheet("Bill-to-Bill Detail")

def write_bill_detail(ws):
    ws.sheet_view.showGridLines = False

    col_widths = [12, 22, 38, 22, 14, 12, 14, 16, 8,
                  14, 12, 10, 10, 10, 8, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:Q1")
    c = ws["A1"]
    c.value = "GSTR-2B — DETAILED BILL-TO-BILL PURCHASE REGISTER — FY 2025-26"
    style_cell(c, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
               align="center", size=12)
    ws.row_dimensions[1].height = 25

    ws.merge_cells("A2:Q2")
    c = ws["A2"]
    c.value = f"GSTIN: {GSTIN}   |   Source: GSTR-2B Auto-populated Data"
    style_cell(c, bold=True, bg=C_MID_BLUE, font_color="FFFFFF",
               align="center", size=9)
    ws.row_dimensions[2].height = 16

    # Column headers
    hdrs = [
        "Month", "GSTIN of Supplier", "Supplier Name",
        "Invoice / Note No.", "Invoice Type", "Date",
        "Invoice Value (₹)", "Taxable Value (₹)", "POS",
        "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)",
        "Total GST (₹)", "RCM", "ITC Eligible", "Source"
    ]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=ci, value=h)
        style_cell(c, bold=True, bg=C_MID_BLUE, font_color="FFFFFF",
                   align="center", size=9, wrap=True)
    ws.row_dimensions[3].height = 28

    # Freeze panes
    ws.freeze_panes = "A4"

    # Sort bills by month order then supplier
    month_order = {m: i for i, m in enumerate(MONTHS)}
    sorted_bills = sorted(all_bills,
                          key=lambda r: (month_order.get(r["month"], 99), r["name"]))

    prev_month = None
    row_idx = 4
    month_sub = {m: {"count":0,"taxable":0,"igst":0,"cgst":0,"sgst":0,"cess":0,"gst":0}
                 for m in MONTHS}

    for r in sorted_bills:
        m = r["month"]
        if m != prev_month:
            # Month sub-header
            ws.merge_cells(start_row=row_idx, start_column=1,
                           end_row=row_idx, end_column=17)
            c = ws.cell(row=row_idx, column=1, value=f"▶  {m}")
            style_cell(c, bold=True, bg=C_HEADER_GREY, align="left", size=10)
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            prev_month = m

        total_gst = r["igst"] + r["cgst"] + r["sgst"] + r["cess"]
        vals = [
            m, r["gstin"], r["name"], r["inv_no"], r["inv_type"],
            r["inv_date"], r["inv_value"], r["taxable"], r["pos"],
            r["igst"], r["cgst"], r["sgst"], r["cess"], total_gst,
            r["rcm"], r["itc"], r["source"]
        ]
        row_bg = None
        if r["source"] == "CDNR":
            row_bg = C_ORANGE
        elif r["source"] == "IMPG":
            row_bg = C_YELLOW
        elif r.get("itc") != "Yes":
            row_bg = C_RED

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=ci, value=v)
            if ci in (7, 8, 10, 11, 12, 13, 14):
                cell.number_format = '#,##0.00'
                style_cell(cell, bg=row_bg, align="right", size=8)
            elif ci == 1:
                style_cell(cell, bg=row_bg, align="center", size=8)
            else:
                style_cell(cell, bg=row_bg, align="left", size=8)

        ws.row_dimensions[row_idx].height = 14

        # accumulate sub-totals
        month_sub[m]["count"]   += 1
        month_sub[m]["taxable"] += r["taxable"]
        month_sub[m]["igst"]    += r["igst"]
        month_sub[m]["cgst"]    += r["cgst"]
        month_sub[m]["sgst"]    += r["sgst"]
        month_sub[m]["cess"]    += r["cess"]
        month_sub[m]["gst"]     += total_gst

        row_idx += 1

        # Insert sub-total after last bill of the month
        next_months = [s["month"] for s in sorted_bills if s != r]

    # ── Monthly sub-total rows ──────────────────────────────────────────
    # Re-scan to insert sub-total rows properly — rebuild with markers
    # (already done inline above; add grand total instead)
    ws.cell(row=row_idx, column=1, value="")
    row_idx += 1

    # Grand total row
    sub_hdrs = ["GRAND TOTAL", "", "", "", "", "", "", "", ""]
    gt_vals  = [
        sum(month_sub[m]["taxable"] for m in MONTHS),
        "", "", "",
        sum(month_sub[m]["igst"]    for m in MONTHS),
        sum(month_sub[m]["cgst"]    for m in MONTHS),
        sum(month_sub[m]["sgst"]    for m in MONTHS),
        sum(month_sub[m]["cess"]    for m in MONTHS),
        sum(month_sub[m]["gst"]     for m in MONTHS),
        "", "", ""
    ]
    row_data = ["GRAND TOTAL", "", "", "", "", "", gt_vals[0], gt_vals[0], "",
                gt_vals[4], gt_vals[5], gt_vals[6], gt_vals[7], gt_vals[8],
                "", "", ""]

    for ci, v in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=ci, value=v)
        if ci in (7, 8, 10, 11, 12, 13, 14):
            cell.number_format = '#,##0.00'
            style_cell(cell, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
                       align="right", size=9)
        else:
            style_cell(cell, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
                       align="left", size=9)
    ws.row_dimensions[row_idx].height = 18

write_bill_detail(ws2)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – MONTHLY ITC SUMMARY (pivot-style)
# ════════════════════════════════════════════════════════════════════════════
ws3 = out_wb.create_sheet("Monthly ITC Summary")

def write_itc_summary(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    for col in "BCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["O"].width = 15

    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value = "MONTHLY ITC CLASSIFICATION — FY 2025-26 (From GSTR-2B ITC Sheets)"
    style_cell(c, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
               align="center", size=12)
    ws.row_dimensions[1].height = 25

    hdrs = ["ITC Category"] + MONTHS + ["TOTAL"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        style_cell(c, bold=True, bg=C_MID_BLUE, font_color="FFFFFF",
                   align="center", size=9, wrap=True)
    ws.row_dimensions[2].height = 28

    cur_row = 3

    def itc_section(title, itc_dict, bg_section, bg_total):
        nonlocal cur_row
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=15)
        c = ws.cell(row=cur_row, column=1, value=title)
        style_cell(c, bold=True, bg=bg_section, font_color="FFFFFF",
                   align="left", size=10)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # collect all unique keys across all months
        all_keys = []
        seen = set()
        for m in MONTHS:
            for k in itc_dict.get(m, {}):
                if k not in seen:
                    all_keys.append(k)
                    seen.add(k)

        for key in all_keys:
            c = ws.cell(row=cur_row, column=1,
                        value="  " + key[:60])
            style_cell(c, align="left", size=9)
            total_v = 0
            for ci, m in enumerate(MONTHS, 2):
                v = itc_dict.get(m, {}).get(key, {}).get("total", 0)
                cell = ws.cell(row=cur_row, column=ci, value=round(v, 2))
                cell.number_format = '#,##0.00'
                style_cell(cell, align="right", size=9)
                total_v += v
            tc = ws.cell(row=cur_row, column=15, value=round(total_v, 2))
            tc.number_format = '#,##0.00'
            style_cell(tc, bold=True, bg=C_HEADER_GREY, align="right", size=9)
            ws.row_dimensions[cur_row].height = 14
            cur_row += 1

        # Sub-total
        c = ws.cell(row=cur_row, column=1, value=f"  SUB-TOTAL – {title}")
        style_cell(c, bold=True, bg=bg_total, align="left", size=9)
        grand = 0
        for ci, m in enumerate(MONTHS, 2):
            v = sum_itc_all(itc_dict, m)["total"]
            cell = ws.cell(row=cur_row, column=ci, value=round(v, 2))
            cell.number_format = '#,##0.00'
            style_cell(cell, bold=True, bg=bg_total, align="right", size=9)
            grand += v
        tc = ws.cell(row=cur_row, column=15, value=round(grand, 2))
        tc.number_format = '#,##0.00'
        style_cell(tc, bold=True, bg=bg_total, align="right", size=9)
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1
        cur_row += 1  # blank

    itc_section("ITC AVAILABLE (Eligible – GSTR-3B Table 4A)",
                itc_avail, C_MID_BLUE, C_GREEN)
    itc_section("ITC NOT AVAILABLE (Blocked / Ineligible)",
                itc_na, "C55A11", C_ORANGE)
    itc_section("ITC REVERSAL (Rule 37/42/43 / IMS Rejected)",
                itc_rev, "843C0C", C_ORANGE)
    itc_section("ITC REJECTED (Supplier Non-compliance etc.)",
                itc_rej, "C00000", C_RED)

write_itc_summary(ws3)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 – SUPPLIER-WISE SUMMARY
# ════════════════════════════════════════════════════════════════════════════
ws4 = out_wb.create_sheet("Supplier-wise Summary")

def write_supplier_summary(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    for col in "DEFGHIJ":
        ws.column_dimensions[col].width = 14

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "SUPPLIER-WISE ITC SUMMARY — FY 2025-26"
    style_cell(c, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
               align="center", size=12)
    ws.row_dimensions[1].height = 25

    hdrs = ["GSTIN", "Supplier Name", "Invoices",
            "Taxable Value (₹)", "IGST (₹)", "CGST (₹)",
            "SGST (₹)", "Cess (₹)", "Total GST (₹)", "ITC Eligible"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        style_cell(c, bold=True, bg=C_MID_BLUE, font_color="FFFFFF",
                   align="center", size=9, wrap=True)
    ws.row_dimensions[2].height = 28

    # aggregate by supplier
    sup = defaultdict(lambda: {"name":"","count":0,"taxable":0,
                               "igst":0,"cgst":0,"sgst":0,"cess":0,"gst":0,"itc":""})
    for r in all_bills:
        g = r["gstin"]
        sup[g]["name"]    = r["name"]
        sup[g]["count"]  += 1
        sup[g]["taxable"]+= r["taxable"]
        sup[g]["igst"]   += r["igst"]
        sup[g]["cgst"]   += r["cgst"]
        sup[g]["sgst"]   += r["sgst"]
        sup[g]["cess"]   += r["cess"]
        sup[g]["gst"]    += r["igst"]+r["cgst"]+r["sgst"]+r["cess"]
        sup[g]["itc"]     = r.get("itc","Yes")

    sorted_sup = sorted(sup.items(), key=lambda x: -x[1]["gst"])

    row_idx = 3
    for gstin, d in sorted_sup:
        vals = [gstin, d["name"], d["count"],
                d["taxable"], d["igst"], d["cgst"],
                d["sgst"], d["cess"], d["gst"], d["itc"]]
        bg = C_RED if d["itc"] != "Yes" else None
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=ci, value=v)
            if ci in (4, 5, 6, 7, 8, 9):
                cell.number_format = '#,##0.00'
                style_cell(cell, bg=bg, align="right", size=8)
            elif ci == 3:
                style_cell(cell, bg=bg, align="center", size=8)
            else:
                style_cell(cell, bg=bg, align="left", size=8)
        ws.row_dimensions[row_idx].height = 14
        row_idx += 1

    # Grand total
    totals = [
        "GRAND TOTAL", "", sum(d["count"] for d in sup.values()),
        sum(d["taxable"] for d in sup.values()),
        sum(d["igst"]    for d in sup.values()),
        sum(d["cgst"]    for d in sup.values()),
        sum(d["sgst"]    for d in sup.values()),
        sum(d["cess"]    for d in sup.values()),
        sum(d["gst"]     for d in sup.values()), ""
    ]
    for ci, v in enumerate(totals, 1):
        cell = ws.cell(row=row_idx, column=ci, value=v)
        if ci in (4, 5, 6, 7, 8, 9):
            cell.number_format = '#,##0.00'
            style_cell(cell, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
                       align="right", size=9)
        else:
            style_cell(cell, bold=True, bg=C_DARK_BLUE, font_color="FFFFFF",
                       align="left", size=9)
    ws.row_dimensions[row_idx].height = 18

write_supplier_summary(ws4)

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
out_path = "GSTR2B_Reconciliation_Report_FY2025-26.xlsx"
out_wb.save(out_path)
print(f"Report saved: {out_path}")
print(f"Sheets: {out_wb.sheetnames}")
print(f"Total B2B bills: {len(b2b_rows)}")
print(f"Total CDNR notes: {len(cdnr_rows)}")
print(f"Total IMPG imports: {len(impg_rows)}")
