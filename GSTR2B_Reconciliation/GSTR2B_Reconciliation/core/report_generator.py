"""
Report Generator - Creates formatted Excel reconciliation reports.
"""
import logging
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from core.reconciler import ReconciliationResult
import config

logger = logging.getLogger("gstr2b_reco")

# Styling constants
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Arial", size=10, bold=True, color="1B3A5C")
DATA_FONT = Font(name="Arial", size=9)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1B3A5C")
SUMMARY_FONT = Font(name="Arial", size=11, bold=True)
NUMBER_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Status-based row colours
STATUS_COLORS = {
    "Matched": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "Amount Mismatch": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "In Tally, Missing in GSTR2B": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    "In GSTR2B, Missing in Tally": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    "No Bill / No Payment in Tally": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
}

ALT_ROW_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")


def _safe_dates(df: pd.DataFrame) -> pd.DataFrame:
    """Convert all datetime/date columns to dd-mm-yyyy strings, NaT → empty string."""
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].apply(
                lambda x: x.strftime("%d-%m-%Y") if x is not pd.NaT and pd.notna(x) else ""
            )
        else:
            # Also catch object columns that may hold datetime/Timestamp values
            df[col] = df[col].apply(
                lambda x: x.strftime("%d-%m-%Y")
                if isinstance(x, (pd.Timestamp, datetime)) and x is not pd.NaT
                else ("" if x is pd.NaT else x)
            )
    return df


def _write_sheet(ws, df: pd.DataFrame, title: str = ""):
    """Write a DataFrame to a worksheet with formatting."""
    if df is None or df.empty:
        ws.append(["No records found"])
        return

    df = _safe_dates(df)

    # Title row
    if title:
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws.cell(1, 1).font = TITLE_FONT
        ws.append([])  # blank row

    start_row = ws.max_row + 1

    # Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = start_row + 1 + row_idx
        status = str(row.get("Status", ""))
        row_fill = STATUS_COLORS.get(status, ALT_ROW_FILL if row_idx % 2 == 0 else None)

        for col_idx, col_name in enumerate(df.columns, 1):
            val = row[col_name]
            # Convert timestamps to strings — check NaT first (NaT is a Timestamp but crashes strftime)
            if val is pd.NaT or val is None:
                val = ""
            elif isinstance(val, (pd.Timestamp, datetime)):
                try:
                    val = val.strftime("%d-%m-%Y")
                except Exception:
                    val = ""

            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            # Number formatting for amount columns
            if any(kw in col_name.lower() for kw in ["value", "igst", "cgst", "sgst", "tax", "amount", "difference"]):
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(wrap_text=True)

            if row_fill:
                cell.fill = row_fill

    # Auto-width columns
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(df.columns[col_idx - 1])),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto-filter
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"


def _write_summary_sheet(ws, summary: dict, month_str: str):
    """Write the reconciliation summary dashboard."""
    ws.append(["GSTR2B Reconciliation Summary"])
    ws.merge_cells("A1:D1")
    ws.cell(1, 1).font = Font(name="Arial", size=16, bold=True, color="1B3A5C")
    ws.append([f"Period: {month_str}"])
    ws.cell(2, 1).font = Font(name="Arial", size=11, color="666666")
    ws.append([f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"])
    ws.cell(3, 1).font = Font(name="Arial", size=11, color="666666")
    ws.append([])

    # Key metrics
    metrics = [
        ("METRIC", "COUNT / VALUE", "", ""),
        ("Tally - Total Records", summary.get("tally_total_records", 0), "", ""),
        ("GSTR2B - Total Records", summary.get("gstr2b_total_records", 0), "", ""),
        ("", "", "", ""),
        ("Matched Entries", summary.get("matched_count", 0), "", ""),
        ("Amount Mismatches", summary.get("amount_mismatch", 0), "", ""),
        ("In Tally, NOT in GSTR2B", summary.get("in_tally_not_gstr2b", 0), "", ""),
        ("In GSTR2B, NOT in Tally", summary.get("in_gstr2b_not_tally", 0), "", ""),
        ("GSTR2B - No Bill/Payment in Tally", summary.get("no_bill_no_payment", 0), "", ""),
        ("Duplicate Entries (Tally)", summary.get("duplicates_tally", 0), "", ""),
        ("Duplicate Entries (GSTR2B)", summary.get("duplicates_gstr2b", 0), "", ""),
        ("", "", "", ""),
        ("ITC as per Tally (\u20B9)", summary.get("tally_itc_total", 0), "", ""),
        ("ITC as per GSTR2B (\u20B9)", summary.get("gstr2b_itc_total", 0), "", ""),
        ("ITC Difference (\u20B9)", round(summary.get("tally_itc_total", 0) - summary.get("gstr2b_itc_total", 0), 2), "", ""),
        ("", "", "", ""),
        ("Processing Time (seconds)", summary.get("run_time_seconds", 0), "", ""),
        ("Errors / Warnings", summary.get("errors_count", 0), "", ""),
    ]

    for i, row in enumerate(metrics):
        ws.append(list(row))
        r = ws.max_row
        if i == 0:
            ws.cell(r, 1).font = HEADER_FONT
            ws.cell(r, 1).fill = HEADER_FILL
            ws.cell(r, 2).font = HEADER_FONT
            ws.cell(r, 2).fill = HEADER_FILL
        else:
            ws.cell(r, 1).font = Font(name="Arial", size=11, bold=True)
            ws.cell(r, 2).font = Font(name="Arial", size=11)
            if "ITC" in str(row[0]):
                ws.cell(r, 2).number_format = NUMBER_FORMAT

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20


def generate_reports(
    result: ReconciliationResult,
    output_dir: str | Path,
    month_str: str = "",
) -> dict:
    """
    Generate Excel reconciliation reports.

    Returns dict with file paths.
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    if not month_str:
        month_str = datetime.now().strftime("%Y-%m")

    files = {}

    # ---- Detailed Report ----
    detail_path = output_path / "Detailed_Reco_Report.xlsx"
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = config.REPORT_SHEETS["summary"]
    _write_summary_sheet(ws_summary, result.summary, month_str)

    # Matched
    ws = wb.create_sheet(config.REPORT_SHEETS["matched"])
    _write_sheet(ws, result.matched, "Matched Entries")

    # In Tally NOT in GSTR2B
    ws = wb.create_sheet(config.REPORT_SHEETS["in_tally_not_gstr2b"])
    _write_sheet(ws, result.in_tally_not_gstr2b, "In Tally but NOT in GSTR2B")

    # In GSTR2B NOT in Tally
    ws = wb.create_sheet(config.REPORT_SHEETS["in_gstr2b_not_tally"])
    _write_sheet(ws, result.in_gstr2b_not_tally, "In GSTR2B but NOT in Tally")

    # No Bill / No Payment
    ws = wb.create_sheet(config.REPORT_SHEETS["no_bill_no_payment"])
    _write_sheet(ws, result.no_bill_no_payment, "In GSTR2B but No Bill / No Payment in Tally")

    # Amount Mismatch
    ws = wb.create_sheet(config.REPORT_SHEETS["amount_mismatch"])
    _write_sheet(ws, result.amount_mismatch, "Amount Mismatch Entries")

    # Duplicates
    ws = wb.create_sheet(config.REPORT_SHEETS["duplicates"])
    if not result.duplicates_tally.empty or not result.duplicates_gstr2b.empty:
        combined = pd.concat([result.duplicates_tally, result.duplicates_gstr2b], ignore_index=True)
        _write_sheet(ws, combined, "Duplicate Entries")
    else:
        ws.append(["No duplicate entries found"])

    wb.save(detail_path)
    files["detailed"] = str(detail_path)
    logger.info(f"Saved: {detail_path}")

    # ---- Summary Report (compact) ----
    summary_path = output_path / "Reconciliation_Summary.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Summary"
    _write_summary_sheet(ws2, result.summary, month_str)
    wb2.save(summary_path)
    files["summary"] = str(summary_path)
    logger.info(f"Saved: {summary_path}")

    return files
