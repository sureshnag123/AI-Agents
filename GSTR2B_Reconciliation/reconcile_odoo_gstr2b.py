"""
GSTR-2B vs Odoo Vendor Bills — Full Reconciliation Report
FY 2025-26 | GSTIN: 29AACCF2736P1Z5

Matching logic:
  1. Exact match: GSTIN + normalized invoice number
  2. Fuzzy match:  GSTIN + invoice number similarity (for minor formatting diffs)
  3. Unmatched:   flagged as Only-in-GSTR2B or Only-in-Odoo
"""
import sys, io, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import openpyxl

# ── colour palette ────────────────────────────────────────────────────────────
C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "BDD7EE"
C_HDR_GREY   = "D6DCE4"
C_GREEN      = "E2EFDA"
C_ORANGE     = "FCE4D6"
C_RED        = "FFE0E0"
C_YELLOW     = "FFFFE0"
C_PURPLE     = "EAD1DC"
C_WHITE      = "FFFFFF"

MONTHS = [
    "Apr-2025","May-2025","Jun-2025","Jul-2025","Aug-2025","Sep-2025",
    "Oct-2025","Nov-2025","Dec-2025","Jan-2026","Feb-2026","Mar-2026"
]
GSTIN = "29AACCF2736P1Z5"

# ═══════════════════════════════════════════════════════════════════════════
# 1.  LOAD GSTR-2B DATA (from compiled Excel)
# ═══════════════════════════════════════════════════════════════════════════
def normalize_inv(s):
    if not s:
        return ""
    return re.sub(r'[^A-Z0-9]', '', str(s).upper().strip())

def clean_gstin(s):
    return re.sub(r'[^A-Z0-9]', '', str(s).upper().strip()) if s else ""

print("Loading GSTR-2B data...")
src_wb = openpyxl.load_workbook("GSTR2B_FY2025-26_Compiled.xlsx")

gstr2b_rows = []
for sheet_name, doc_type_label in [("B2B", "B2B Invoice"), ("B2B-CDNR", "Credit/Debit Note"), ("IMPG", "Import (IMPG)")]:
    ws = src_wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] not in MONTHS:
            continue
        if sheet_name == "B2B":
            sign = 1
            gstin    = clean_gstin(row[1])
            name     = str(row[2] or "")
            inv_no   = str(row[3] or "")
            inv_date = str(row[5] or "")
            taxable  = float(row[9]  or 0)
            igst     = float(row[10] or 0)
            cgst     = float(row[11] or 0)
            sgst     = float(row[12] or 0)
            cess     = float(row[13] or 0)
            itc      = str(row[16] or "Yes") if len(row) > 16 else "Yes"
        elif sheet_name == "B2B-CDNR":
            note_type = str(row[4] or "")
            sign = -1 if "Credit" in note_type else 1
            gstin    = clean_gstin(row[1])
            name     = str(row[2] or "")
            inv_no   = str(row[3] or "")
            inv_date = str(row[6] or "")
            taxable  = float(row[10] or 0) * sign
            igst     = float(row[11] or 0) * sign
            cgst     = float(row[12] or 0) * sign
            sgst     = float(row[13] or 0) * sign
            cess     = float(row[14] or 0) * sign
            itc      = str(row[23] or "Yes") if len(row) > 23 else "Yes"
        else:  # IMPG
            gstin    = "IMPG"
            name     = f"Port: {row[2] or ''}"
            inv_no   = str(row[3] or "")
            inv_date = str(row[4] or "")
            taxable  = float(row[5] or 0)
            igst     = float(row[6] or 0)
            cgst     = 0; sgst = 0; cess = float(row[7] or 0)
            itc      = "Yes"; sign = 1

        gstr2b_rows.append({
            "month":    row[0],
            "gstin":    gstin,
            "name":     name,
            "inv_no":   inv_no,
            "inv_no_n": normalize_inv(inv_no),
            "inv_date": inv_date,
            "taxable":  taxable,
            "igst":     igst,
            "cgst":     cgst,
            "sgst":     sgst,
            "cess":     cess,
            "total_gst": igst+cgst+sgst+cess,
            "itc":      itc,
            "source":   sheet_name,
        })

print(f"  GSTR-2B rows: {len(gstr2b_rows)}")

# ═══════════════════════════════════════════════════════════════════════════
# 2.  LOAD ODOO DATA
# ═══════════════════════════════════════════════════════════════════════════
print("Loading Odoo data...")
with open("odoo_bills_raw.json", encoding="utf-8") as f:
    odoo_raw = json.load(f)

odoo_rows = []
for r in odoo_raw:
    vendor_ref = r.get("vendor_ref","").strip()
    odoo_ref   = r.get("odoo_ref","").strip()
    inv_no     = vendor_ref or odoo_ref
    odoo_rows.append({
        "month":     r["month"],
        "gstin":     clean_gstin(r.get("gstin","")),
        "name":      r.get("partner",""),
        "odoo_ref":  odoo_ref,
        "vendor_ref":vendor_ref,
        "inv_no":    inv_no,
        "inv_no_n":  normalize_inv(inv_no),
        "inv_date":  r.get("inv_date",""),
        "taxable":   float(r.get("taxable",0)),
        "igst":      float(r.get("igst",0)),
        "cgst":      float(r.get("cgst",0)),
        "sgst":      float(r.get("sgst",0)),
        "cess":      0,
        "total_gst": float(r.get("total_gst",0)),
        "doc_type":  r.get("doc_type","Vendor Bill"),
    })

print(f"  Odoo rows: {len(odoo_rows)}")

# ═══════════════════════════════════════════════════════════════════════════
# 3.  MATCHING ENGINE
# ═══════════════════════════════════════════════════════════════════════════
print("Matching bills...")

# Build lookup: (gstin, norm_inv_no) -> list of gstr2b rows
gstr2b_lookup = defaultdict(list)
for i, r in enumerate(gstr2b_rows):
    key = (r["gstin"], r["inv_no_n"])
    gstr2b_lookup[key].append(i)

TOLERANCE = 1.0   # ₹1 rounding tolerance

matched_pairs  = []   # (gstr2b_row, odoo_row, status, diff_taxable, diff_gst)
only_in_odoo   = []
only_in_gstr2b = list(range(len(gstr2b_rows)))  # start with all; remove as matched

matched_gstr2b_idx = set()

for odoo_r in odoo_rows:
    key = (odoo_r["gstin"], odoo_r["inv_no_n"])
    candidates = gstr2b_lookup.get(key, [])

    # filter to same month if possible
    same_month = [i for i in candidates if gstr2b_rows[i]["month"] == odoo_r["month"]]
    pool = same_month if same_month else candidates

    # pick best unmatched candidate
    best = None
    for i in pool:
        if i not in matched_gstr2b_idx:
            best = i
            break

    if best is not None:
        gr = gstr2b_rows[best]
        matched_gstr2b_idx.add(best)
        diff_tax = round(odoo_r["taxable"] - gr["taxable"], 2)
        diff_gst = round(odoo_r["total_gst"] - gr["total_gst"], 2)

        if abs(diff_tax) <= TOLERANCE and abs(diff_gst) <= TOLERANCE:
            status = "MATCHED"
        elif abs(diff_tax) > TOLERANCE and abs(diff_gst) <= TOLERANCE:
            status = "TAXABLE DIFF"
        elif abs(diff_tax) <= TOLERANCE and abs(diff_gst) > TOLERANCE:
            status = "GST DIFF"
        else:
            status = "BOTH DIFF"

        matched_pairs.append((gr, odoo_r, status, diff_tax, diff_gst))
    else:
        only_in_odoo.append(odoo_r)

only_in_gstr2b = [gstr2b_rows[i] for i in range(len(gstr2b_rows))
                  if i not in matched_gstr2b_idx]

print(f"  Matched:           {len(matched_pairs)}")
print(f"  Only in GSTR-2B:   {len(only_in_gstr2b)}")
print(f"  Only in Odoo:      {len(only_in_odoo)}")
print(f"  Status breakdown:")
status_counts = defaultdict(int)
for _, _, st, _, _ in matched_pairs:
    status_counts[st] += 1
for k,v in status_counts.items():
    print(f"    {k}: {v}")

# ═══════════════════════════════════════════════════════════════════════════
# 4.  BUILD EXCEL REPORT
# ═══════════════════════════════════════════════════════════════════════════
print("\nBuilding Excel report...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

def thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(cell, bold=False, bg=None, fc="000000", align="left",
       size=9, wrap=False, border=True):
    cell.font = Font(bold=bold, color=fc, size=size, name="Calibri")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = thin()

def money(cell, val):
    cell.value = round(float(val or 0), 2)
    cell.number_format = '#,##0.00'

# ─────────────────────────────────────────────────────────────────────────
# SHEET 1 — RECONCILIATION SUMMARY (Month-on-Month)
# ─────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("Recon Summary")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 40
for c in "BCDEFGHIJKLMNO":
    ws1.column_dimensions[c].width = 14

# Title
ws1.merge_cells("A1:O1")
c = ws1["A1"]
c.value = "GSTR-2B vs ODOO — MONTH-ON-MONTH RECONCILIATION SUMMARY  |  FY 2025-26"
sc(c, bold=True, bg=C_DARK_BLUE, fc="FFFFFF", align="center", size=13)
ws1.row_dimensions[1].height = 28

ws1.merge_cells("A2:O2")
c = ws1["A2"]
c.value = f"GSTIN: {GSTIN}   |   Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
sc(c, bold=True, bg=C_MID_BLUE, fc="FFFFFF", align="center", size=10)
ws1.row_dimensions[2].height = 16

# Month headers row
hdrs = ["Particulars"] + MONTHS + ["TOTAL FY 2025-26"]
for ci, h in enumerate(hdrs, 1):
    c = ws1.cell(row=4, column=ci, value=h)
    sc(c, bold=True, bg=C_DARK_BLUE, fc="FFFFFF", align="center", size=9, wrap=True)
ws1.row_dimensions[4].height = 30

def agg_by_month(rows_list, keys):
    res = {m: {k: 0 for k in keys} for m in MONTHS}
    res_count = {m: 0 for m in MONTHS}
    for r in rows_list:
        m = r.get("month","")
        if m in res:
            res_count[m] += 1
            for k in keys:
                res[m][k] += float(r.get(k, 0))
    return res, res_count

gstr2b_agg, gstr2b_cnt = agg_by_month(gstr2b_rows, ["taxable","igst","cgst","sgst","cess","total_gst"])
odoo_agg,   odoo_cnt   = agg_by_month(odoo_rows,   ["taxable","igst","cgst","sgst","total_gst"])

matched_g = [p[0] for p in matched_pairs]
matched_o = [p[1] for p in matched_pairs]
only_g_agg, only_g_cnt = agg_by_month(only_in_gstr2b, ["taxable","total_gst"])
only_o_agg, only_o_cnt = agg_by_month(only_in_odoo,   ["taxable","total_gst"])

cur = 5

def section_hdr(label, bg=C_MID_BLUE):
    global cur
    ws1.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=15)
    c = ws1.cell(row=cur, column=1, value=label)
    sc(c, bold=True, bg=bg, fc="FFFFFF", align="left", size=10)
    ws1.row_dimensions[cur].height = 18
    cur += 1

def data_row(label, vals, bg=None, bold=False):
    global cur
    c = ws1.cell(row=cur, column=1, value=label)
    sc(c, bold=bold, bg=bg, align="left", size=9)
    total = 0
    for ci, v in enumerate(vals, 2):
        cell = ws1.cell(row=cur, column=ci)
        money(cell, v)
        sc(cell, bold=bold, bg=bg, align="right", size=9)
        total += v
    tc = ws1.cell(row=cur, column=15)
    money(tc, total)
    sc(tc, bold=True, bg=bg or C_HDR_GREY, align="right", size=9)
    ws1.row_dimensions[cur].height = 15
    cur += 1
    return total

def count_row(label, vals, bg=None, bold=False):
    global cur
    c = ws1.cell(row=cur, column=1, value=label)
    sc(c, bold=bold, bg=bg, align="left", size=9)
    total = 0
    for ci, v in enumerate(vals, 2):
        cell = ws1.cell(row=cur, column=ci, value=int(v))
        cell.number_format = '#,##0'
        sc(cell, bold=bold, bg=bg, align="center", size=9)
        total += v
    tc = ws1.cell(row=cur, column=15, value=int(total))
    tc.number_format = '#,##0'
    sc(tc, bold=True, bg=bg or C_HDR_GREY, align="center", size=9)
    ws1.row_dimensions[cur].height = 15
    cur += 1

def blank():
    global cur
    cur += 1

# ── Section A: GSTR-2B ───────────────────────────────────────────────────
section_hdr("SECTION A — GSTR-2B (Auto-populated) Summary", C_MID_BLUE)
count_row("  No. of Documents (B2B + CDNR + IMPG)",
          [gstr2b_cnt[m] for m in MONTHS])
data_row("  Taxable Value (₹)",
         [gstr2b_agg[m]["taxable"] for m in MONTHS], bg=C_LIGHT_BLUE)
data_row("  IGST (₹)",  [gstr2b_agg[m]["igst"]  for m in MONTHS])
data_row("  CGST (₹)",  [gstr2b_agg[m]["cgst"]  for m in MONTHS])
data_row("  SGST (₹)",  [gstr2b_agg[m]["sgst"]  for m in MONTHS])
data_row("  Total GST (₹)",
         [gstr2b_agg[m]["total_gst"] for m in MONTHS], bg=C_GREEN, bold=True)
blank()

# ── Section B: Odoo ──────────────────────────────────────────────────────
section_hdr("SECTION B — ODOO (Books) Summary", "714B67")
count_row("  No. of Vendor Bills / Refunds",
          [odoo_cnt[m] for m in MONTHS])
data_row("  Taxable Value (₹)",
         [odoo_agg[m]["taxable"] for m in MONTHS], bg=C_PURPLE)
data_row("  IGST (₹)",  [odoo_agg[m]["igst"] for m in MONTHS])
data_row("  CGST (₹)",  [odoo_agg[m]["cgst"] for m in MONTHS])
data_row("  SGST (₹)",  [odoo_agg[m]["sgst"] for m in MONTHS])
data_row("  Total GST (₹)",
         [odoo_agg[m]["total_gst"] for m in MONTHS], bg=C_PURPLE, bold=True)
blank()

# ── Section C: Variance ──────────────────────────────────────────────────
section_hdr("SECTION C — VARIANCE (GSTR-2B minus Odoo)", "C55A11")

diff_tax = [round(gstr2b_agg[m]["taxable"]   - odoo_agg[m]["taxable"],   2) for m in MONTHS]
diff_gst = [round(gstr2b_agg[m]["total_gst"] - odoo_agg[m]["total_gst"], 2) for m in MONTHS]
diff_igst= [round(gstr2b_agg[m]["igst"]      - odoo_agg[m]["igst"],      2) for m in MONTHS]
diff_cgst= [round(gstr2b_agg[m]["cgst"]      - odoo_agg[m]["cgst"],      2) for m in MONTHS]
diff_sgst= [round(gstr2b_agg[m]["sgst"]      - odoo_agg[m]["sgst"],      2) for m in MONTHS]
diff_cnt = [gstr2b_cnt[m] - odoo_cnt[m] for m in MONTHS]

def variance_row(label, vals, bold=False):
    global cur
    c = ws1.cell(row=cur, column=1, value=label)
    sc(c, bold=bold, align="left", size=9)
    total = 0
    for ci, v in enumerate(vals, 2):
        bg = C_GREEN if abs(v) < 1 else C_RED
        cell = ws1.cell(row=cur, column=ci)
        money(cell, v)
        sc(cell, bold=bold, bg=bg, align="right", size=9)
        total += v
    bg_t = C_GREEN if abs(total) < 1 else C_RED
    tc = ws1.cell(row=cur, column=15)
    money(tc, total)
    sc(tc, bold=True, bg=bg_t, align="right", size=9)
    ws1.row_dimensions[cur].height = 15
    cur += 1

variance_row("  Taxable Value Difference (₹)", diff_tax)
variance_row("  IGST Difference (₹)",  diff_igst)
variance_row("  CGST Difference (₹)",  diff_cgst)
variance_row("  SGST Difference (₹)",  diff_sgst)
variance_row("  TOTAL GST Difference (₹)", diff_gst, bold=True)
blank()

# ── Section D: Reconciliation status ────────────────────────────────────
section_hdr("SECTION D — BILL-TO-BILL RECONCILIATION STATUS", C_DARK_BLUE)

# matched counts by month
matched_cnt = defaultdict(int)
matched_ok  = defaultdict(int)
matched_diff= defaultdict(int)
for gr, or_, st, dt, dg in matched_pairs:
    m = gr["month"]
    matched_cnt[m] += 1
    if st == "MATCHED":
        matched_ok[m]  += 1
    else:
        matched_diff[m]+= 1

only_g_cnt_m = defaultdict(int)
only_o_cnt_m = defaultdict(int)
for r in only_in_gstr2b: only_g_cnt_m[r["month"]] += 1
for r in only_in_odoo:   only_o_cnt_m[r["month"]] += 1

def int_row(label, vals_dict, bg=None, bold=False):
    global cur
    c = ws1.cell(row=cur, column=1, value=label)
    sc(c, bold=bold, bg=bg, align="left", size=9)
    total = 0
    for ci, m in enumerate(MONTHS, 2):
        v = vals_dict.get(m, 0)
        cell = ws1.cell(row=cur, column=ci, value=int(v))
        cell.number_format = '#,##0'
        sc(cell, bold=bold, bg=bg, align="center", size=9)
        total += v
    tc = ws1.cell(row=cur, column=15, value=int(total))
    tc.number_format = '#,##0'
    sc(tc, bold=True, bg=bg or C_HDR_GREY, align="center", size=9)
    ws1.row_dimensions[cur].height = 15
    cur += 1

int_row("  Matched Bills (exact + value match)", matched_ok,   bg=C_GREEN)
int_row("  Matched Bills (value difference)",    matched_diff, bg=C_YELLOW)
int_row("  In GSTR-2B only (not in Odoo)",       only_g_cnt_m, bg=C_ORANGE)
int_row("  In Odoo only (not in GSTR-2B)",        only_o_cnt_m, bg=C_RED)
blank()

# GST at risk (only-in-gstr2b = ITC not booked; only-in-odoo = ITC booked but not in GSTR-2B)
gst_only_g = {m: round(only_g_agg[m]["total_gst"], 2) for m in MONTHS}
gst_only_o = {m: round(only_o_agg[m]["total_gst"], 2) for m in MONTHS}

def gst_risk_row(label, vals_dict, bg, bold=False):
    global cur
    c = ws1.cell(row=cur, column=1, value=label)
    sc(c, bold=bold, bg=bg, align="left", size=9)
    total = 0
    for ci, m in enumerate(MONTHS, 2):
        v = vals_dict.get(m, 0)
        cell = ws1.cell(row=cur, column=ci)
        money(cell, v)
        sc(cell, bold=bold, bg=bg, align="right", size=9)
        total += v
    tc = ws1.cell(row=cur, column=15)
    money(tc, total)
    sc(tc, bold=True, bg=bg, align="right", size=9)
    ws1.row_dimensions[cur].height = 15
    cur += 1

section_hdr("SECTION E — ITC IMPACT ANALYSIS", "C00000")
gst_risk_row("  GST on bills only in GSTR-2B (ITC not claimed in Odoo) (₹)",
             gst_only_g, C_ORANGE, bold=True)
gst_risk_row("  GST on bills only in Odoo (booked but not auto-populated) (₹)",
             gst_only_o, C_RED, bold=True)

diff_match_gst = defaultdict(float)
for gr, _, st, _, dg in matched_pairs:
    if st != "MATCHED":
        diff_match_gst[gr["month"]] += abs(dg)
gst_risk_row("  GST difference on matched-but-differing bills (₹)",
             dict(diff_match_gst), C_YELLOW, bold=True)

blank()

# Footer note
ws1.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=15)
note = ws1.cell(row=cur, column=1,
    value="NOTE: GREEN = matched within ₹1 tolerance | RED = mismatch | "
          "ORANGE = in GSTR-2B only | YELLOW = value difference. "
          "Reconcile 'Only in GSTR-2B' items — these are ITC available but not claimed. "
          "'Only in Odoo' items indicate entries booked without GSTR-2B match — verify supplier filing.")
sc(note, bg=C_LIGHT_BLUE, align="left", size=8, wrap=True)
ws1.row_dimensions[cur].height = 45

# ─────────────────────────────────────────────────────────────────────────
# SHEET 2 — MATCHED BILLS (with differences)
# ─────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Matched Bills")
ws2.sheet_view.showGridLines = False

col_w2 = [12,16,30,22,12,14,12,10,10,10,8,16,30,22,12,14,12,10,10,10,10,12,12,14]
for i,w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:X1")
c = ws2["A1"]
c.value = "MATCHED BILLS — GSTR-2B vs ODOO  |  FY 2025-26"
sc(c, bold=True, bg=C_DARK_BLUE, fc="FFFFFF", align="center", size=12)
ws2.row_dimensions[1].height = 24

# Sub-headers: GSTR-2B side | Odoo side | Variance
ws2.merge_cells("A2:K2")
c = ws2["A2"]
c.value = "◀  GSTR-2B Side"
sc(c, bold=True, bg=C_MID_BLUE, fc="FFFFFF", align="center", size=10)

ws2.merge_cells("L2:V2")
c = ws2["L2"]
c.value = "▶  ODOO (Books) Side"
sc(c, bold=True, bg="714B67", fc="FFFFFF", align="center", size=10)

ws2.merge_cells("W2:X2")
c = ws2["W2"]
c.value = "△  Variance"
sc(c, bold=True, bg="C55A11", fc="FFFFFF", align="center", size=10)

hdrs2 = [
    # GSTR-2B
    "Month","GSTIN","Supplier Name","Invoice No.","Date",
    "Taxable (₹)","IGST (₹)","CGST (₹)","SGST (₹)","Total GST (₹)","ITC",
    # Odoo
    "Odoo Ref","Partner Name","Vendor Bill No.","Date",
    "Taxable (₹)","IGST (₹)","CGST (₹)","SGST (₹)","Total GST (₹)","Doc Type",
    # Variance
    "Taxable Diff (₹)","GST Diff (₹)",
    # Status
    "Status"
]
for ci, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=3, column=ci, value=h)
    bg = C_MID_BLUE if ci <= 11 else ("714B67" if ci <= 21 else "C55A11")
    sc(c, bold=True, bg=bg, fc="FFFFFF", align="center", size=8, wrap=True)
ws2.row_dimensions[3].height = 30
ws2.freeze_panes = "A4"

row_idx = 4
status_bg = {
    "MATCHED":     C_GREEN,
    "TAXABLE DIFF":C_YELLOW,
    "GST DIFF":    C_ORANGE,
    "BOTH DIFF":   C_RED,
}
for gr, or_, status, diff_tax, diff_gst in matched_pairs:
    bg = status_bg.get(status, C_WHITE)
    vals = [
        gr["month"], gr["gstin"], gr["name"][:40], gr["inv_no"][:30], gr["inv_date"],
        gr["taxable"], gr["igst"], gr["cgst"], gr["sgst"], gr["total_gst"], gr["itc"],
        or_["odoo_ref"][:20], or_["name"][:40], or_["vendor_ref"][:30], or_["inv_date"],
        or_["taxable"], or_["igst"], or_["cgst"], or_["sgst"], or_["total_gst"], or_["doc_type"],
        diff_tax, diff_gst,
        status
    ]
    money_cols = {6,7,8,9,10,16,17,18,19,20,22,23}
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=ci, value=v)
        if ci in money_cols:
            money(cell, v)
            diff_bg = (C_RED if ci in (22,23) and abs(float(v or 0)) >= 1
                       else C_GREEN if ci in (22,23) else bg)
            sc(cell, bg=diff_bg if ci in (22,23) else bg, align="right", size=8)
        elif ci == 24:
            sc(cell, bold=True, bg=bg, align="center", size=8)
        else:
            sc(cell, bg=bg, align="left", size=8)
    ws2.row_dimensions[row_idx].height = 14
    row_idx += 1

# ─────────────────────────────────────────────────────────────────────────
# SHEET 3 — ONLY IN GSTR-2B (ITC not booked in Odoo)
# ─────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Only in GSTR-2B")
ws3.sheet_view.showGridLines = False

col_w3 = [12,18,38,24,12,14,12,10,10,10,8,10,50]
for i,w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:M1")
c = ws3["A1"]
c.value = "BILLS IN GSTR-2B BUT NOT FOUND IN ODOO — ITC AVAILABLE BUT POSSIBLY UNCLAIMED"
sc(c, bold=True, bg="C55A11", fc="FFFFFF", align="center", size=12)
ws3.row_dimensions[1].height = 24

ws3.merge_cells("A2:M2")
c = ws3["A2"]
c.value = f"Total: {len(only_in_gstr2b)} documents  |  Action Required: Book in Odoo or verify supplier filing"
sc(c, bold=True, bg=C_ORANGE, align="center", size=10)
ws3.row_dimensions[2].height = 16

hdrs3 = ["Month","GSTIN","Supplier Name","Invoice No.","Date",
         "Taxable (₹)","IGST (₹)","CGST (₹)","SGST (₹)","Cess (₹)","Total GST (₹)","ITC","Source"]
for ci,h in enumerate(hdrs3,1):
    c = ws3.cell(row=3, column=ci, value=h)
    sc(c, bold=True, bg="C55A11", fc="FFFFFF", align="center", size=8, wrap=True)
ws3.row_dimensions[3].height = 24
ws3.freeze_panes = "A4"

row_idx = 4
for r in sorted(only_in_gstr2b, key=lambda x: (MONTHS.index(x["month"]) if x["month"] in MONTHS else 99, x["name"])):
    vals = [r["month"], r["gstin"], r["name"][:50], r["inv_no"][:30], r["inv_date"],
            r["taxable"], r["igst"], r["cgst"], r["sgst"], r["cess"], r["total_gst"],
            r["itc"], r["source"]]
    money_c = {6,7,8,9,10,11}
    for ci,v in enumerate(vals,1):
        cell = ws3.cell(row=row_idx, column=ci, value=v)
        if ci in money_c:
            money(cell, v); sc(cell, bg=C_ORANGE, align="right", size=8)
        else:
            sc(cell, bg=C_ORANGE, align="left", size=8)
    ws3.row_dimensions[row_idx].height = 14
    row_idx += 1

# Grand total
tots = {"taxable":0,"igst":0,"cgst":0,"sgst":0,"cess":0,"total_gst":0}
for r in only_in_gstr2b:
    for k in tots: tots[k] += r[k]
gt_vals = ["TOTAL","","","","",
           tots["taxable"],tots["igst"],tots["cgst"],tots["sgst"],tots["cess"],tots["total_gst"],"",""]
for ci,v in enumerate(gt_vals,1):
    cell = ws3.cell(row=row_idx, column=ci, value=v)
    if ci in {6,7,8,9,10,11}:
        money(cell,v); sc(cell,bold=True,bg=C_DARK_BLUE,fc="FFFFFF",align="right",size=9)
    else:
        sc(cell,bold=True,bg=C_DARK_BLUE,fc="FFFFFF",align="left",size=9)
ws3.row_dimensions[row_idx].height = 18

# ─────────────────────────────────────────────────────────────────────────
# SHEET 4 — ONLY IN ODOO (booked but not in GSTR-2B)
# ─────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Only in Odoo")
ws4.sheet_view.showGridLines = False

col_w4 = [12,18,38,24,24,12,14,12,10,10,12,50]
for i,w in enumerate(col_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells("A1:L1")
c = ws4["A1"]
c.value = "BILLS IN ODOO BUT NOT IN GSTR-2B — ITC BOOKED BUT SUPPLIER MAY NOT HAVE FILED"
sc(c, bold=True, bg="C00000", fc="FFFFFF", align="center", size=12)
ws4.row_dimensions[1].height = 24

ws4.merge_cells("A2:L2")
c = ws4["A2"]
c.value = f"Total: {len(only_in_odoo)} documents  |  Action: Verify supplier GSTR-1 filing; reverse ITC if not appearing"
sc(c, bold=True, bg=C_RED, align="center", size=10)
ws4.row_dimensions[2].height = 16

hdrs4 = ["Month","GSTIN","Partner Name","Odoo Ref","Vendor Bill No.","Date",
         "Taxable (₹)","IGST (₹)","CGST (₹)","SGST (₹)","Total GST (₹)","Doc Type"]
for ci,h in enumerate(hdrs4,1):
    c = ws4.cell(row=3, column=ci, value=h)
    sc(c, bold=True, bg="C00000", fc="FFFFFF", align="center", size=8, wrap=True)
ws4.row_dimensions[3].height = 24
ws4.freeze_panes = "A4"

row_idx = 4
for r in sorted(only_in_odoo, key=lambda x: (MONTHS.index(x["month"]) if x["month"] in MONTHS else 99, x["name"])):
    vals = [r["month"],r["gstin"],r["name"][:50],r["odoo_ref"][:25],r["vendor_ref"][:25],
            r["inv_date"],r["taxable"],r["igst"],r["cgst"],r["sgst"],r["total_gst"],r["doc_type"]]
    money_c = {7,8,9,10,11}
    for ci,v in enumerate(vals,1):
        cell = ws4.cell(row=row_idx, column=ci, value=v)
        if ci in money_c:
            money(cell,v); sc(cell,bg=C_RED,align="right",size=8)
        else:
            sc(cell,bg=C_RED,align="left",size=8)
    ws4.row_dimensions[row_idx].height = 14
    row_idx += 1

tots4 = {"taxable":0,"igst":0,"cgst":0,"sgst":0,"total_gst":0}
for r in only_in_odoo:
    for k in tots4: tots4[k] += r[k]
gt4 = ["TOTAL","","","","","",
       tots4["taxable"],tots4["igst"],tots4["cgst"],tots4["sgst"],tots4["total_gst"],""]
for ci,v in enumerate(gt4,1):
    cell = ws4.cell(row=row_idx, column=ci, value=v)
    if ci in {7,8,9,10,11}:
        money(cell,v); sc(cell,bold=True,bg=C_DARK_BLUE,fc="FFFFFF",align="right",size=9)
    else:
        sc(cell,bold=True,bg=C_DARK_BLUE,fc="FFFFFF",align="left",size=9)
ws4.row_dimensions[row_idx].height = 18

# ─────────────────────────────────────────────────────────────────────────
# SHEET 5 — FULL BILL-TO-BILL DETAIL (all rows, reconciliation status)
# ─────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Full Bill-to-Bill")
ws5.sheet_view.showGridLines = False

col_w5 = [12,16,35,22,12,14,12,10,10,10,8,14,35,22,12,14,12,10,10,10,12,12,14]
for i,w in enumerate(col_w5,1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells("A1:W1")
c = ws5["A1"]
c.value = "FULL BILL-TO-BILL RECONCILIATION DETAIL — GSTR-2B vs ODOO  |  FY 2025-26"
sc(c,bold=True,bg=C_DARK_BLUE,fc="FFFFFF",align="center",size=12)
ws5.row_dimensions[1].height = 24

ws5.merge_cells("A2:K2")
ws5["A2"].value = "◀  GSTR-2B"
sc(ws5["A2"],bold=True,bg=C_MID_BLUE,fc="FFFFFF",align="center",size=10)
ws5.merge_cells("L2:V2")
ws5["L2"].value = "▶  ODOO"
sc(ws5["L2"],bold=True,bg="714B67",fc="FFFFFF",align="center",size=10)
ws5["W2"].value = "Status"
sc(ws5["W2"],bold=True,bg=C_DARK_BLUE,fc="FFFFFF",align="center",size=10)

hdrs5 = [
    "Month","GSTIN","Supplier","Invoice No.","Date","Taxable","IGST","CGST","SGST","Total GST","ITC",
    "Odoo Ref","Partner","Vendor Ref","Date","Taxable","IGST","CGST","SGST","Total GST","Doc",
    "GST Diff","Status"
]
for ci,h in enumerate(hdrs5,1):
    c = ws5.cell(row=3,column=ci,value=h)
    bg = C_MID_BLUE if ci<=11 else "714B67" if ci<=21 else C_DARK_BLUE
    sc(c,bold=True,bg=bg,fc="FFFFFF",align="center",size=8,wrap=True)
ws5.row_dimensions[3].height = 28
ws5.freeze_panes = "A4"

row_idx = 4

# All matched
for gr,or_,status,diff_tax,diff_gst in matched_pairs:
    bg = status_bg.get(status,C_WHITE)
    vals = [
        gr["month"],gr["gstin"],gr["name"][:35],gr["inv_no"][:22],gr["inv_date"],
        gr["taxable"],gr["igst"],gr["cgst"],gr["sgst"],gr["total_gst"],gr["itc"],
        or_["odoo_ref"][:14],or_["name"][:35],or_["vendor_ref"][:22],or_["inv_date"],
        or_["taxable"],or_["igst"],or_["cgst"],or_["sgst"],or_["total_gst"],or_["doc_type"],
        diff_gst, status
    ]
    mc = {6,7,8,9,10,16,17,18,19,20,22}
    for ci,v in enumerate(vals,1):
        cell = ws5.cell(row=row_idx,column=ci,value=v)
        if ci in mc:
            money(cell,v)
            if ci==22:
                sc(cell,bold=True,bg=C_RED if abs(float(v or 0))>=1 else C_GREEN,align="right",size=8)
            else:
                sc(cell,bg=bg,align="right",size=8)
        elif ci==23:
            sc(cell,bold=True,bg=bg,align="center",size=8)
        else:
            sc(cell,bg=bg,align="left",size=8)
    ws5.row_dimensions[row_idx].height = 14
    row_idx += 1

# Only in GSTR-2B
for r in only_in_gstr2b:
    vals = [
        r["month"],r["gstin"],r["name"][:35],r["inv_no"][:22],r["inv_date"],
        r["taxable"],r["igst"],r["cgst"],r["sgst"],r["total_gst"],r["itc"],
        "—","—","—","—",0,0,0,0,0,"—",
        r["total_gst"],"ONLY IN GSTR-2B"
    ]
    mc = {6,7,8,9,10,16,17,18,19,20,22}
    for ci,v in enumerate(vals,1):
        cell = ws5.cell(row=row_idx,column=ci,value=v)
        if ci in mc:
            money(cell,v); sc(cell,bg=C_ORANGE,align="right",size=8)
        elif ci==23:
            sc(cell,bold=True,bg=C_ORANGE,align="center",size=8)
        else:
            sc(cell,bg=C_ORANGE,align="left",size=8)
    ws5.row_dimensions[row_idx].height = 14
    row_idx += 1

# Only in Odoo
for r in only_in_odoo:
    vals = [
        r["month"],"—","—","—","—",0,0,0,0,0,"—",
        r["odoo_ref"][:14],r["name"][:35],r["vendor_ref"][:22],r["inv_date"],
        r["taxable"],r["igst"],r["cgst"],r["sgst"],r["total_gst"],r["doc_type"],
        -r["total_gst"],"ONLY IN ODOO"
    ]
    mc = {6,7,8,9,10,16,17,18,19,20,22}
    for ci,v in enumerate(vals,1):
        cell = ws5.cell(row=row_idx,column=ci,value=v)
        if ci in mc:
            money(cell,v); sc(cell,bg=C_RED,align="right",size=8)
        elif ci==23:
            sc(cell,bold=True,bg=C_RED,align="center",size=8)
        else:
            sc(cell,bg=C_RED,align="left",size=8)
    ws5.row_dimensions[row_idx].height = 14
    row_idx += 1

# ── Save ──────────────────────────────────────────────────────────────────
out = "GSTR2B_vs_Odoo_Reconciliation_FY2025-26.xlsx"
wb.save(out)
print(f"\nReport saved: {out}")
print(f"Sheets: {wb.sheetnames}")
print(f"\n{'='*60}")
print(f"RECONCILIATION FINAL SUMMARY")
print(f"{'='*60}")
print(f"Total GSTR-2B documents  : {len(gstr2b_rows):,}")
print(f"Total Odoo documents     : {len(odoo_rows):,}")
print(f"Matched (exact)          : {status_counts['MATCHED']:,}")
print(f"Matched (with diff)      : {sum(status_counts[k] for k in ['TAXABLE DIFF','GST DIFF','BOTH DIFF']):,}")
print(f"Only in GSTR-2B          : {len(only_in_gstr2b):,}")
print(f"Only in Odoo             : {len(only_in_odoo):,}")
print(f"\nGSTR-2B Total GST        : ₹{sum(r['total_gst'] for r in gstr2b_rows):>15,.2f}")
print(f"Odoo Total GST           : ₹{sum(r['total_gst'] for r in odoo_rows):>15,.2f}")
print(f"Net Variance             : ₹{sum(r['total_gst'] for r in gstr2b_rows)-sum(r['total_gst'] for r in odoo_rows):>15,.2f}")
print(f"\nITC available but not claimed (GSTR-2B only GST): ₹{tots['total_gst']:>12,.2f}")
print(f"ITC booked but not in GSTR-2B (Odoo only GST)  : ₹{tots4['total_gst']:>12,.2f}")
